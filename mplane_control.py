"""
Load M-Plane Excel workbooks for the GUI Control tab.
Reference-only sheets (not loaded): LLTE LLRE, RSRP_EffBW
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl

SKIP_SHEETS = frozenset({"LLTE", "LLRE", "RSRP_EffBW"})
SEND_ORDER = ["CUplane-interface", "Processing-element", "PDSCH", "PUSCH", "PRACH", "ACTIVE"]
# Control-Sheet detail grid: valid endpoint / carrier names (skip Korean notes / blank rows).
_UPLANE_ENDPOINT_NAME_RE = re.compile(r"^(llre_|llte_|rac_)", re.IGNORECASE)


def normalize_header(s: str) -> str:
    # Normalize common Excel header variations:
    # - treat "_" and "-" as spaces
    # - collapse whitespace
    txt = str(s).strip().lower()
    txt = txt.replace("_", " ").replace("-", " ")
    return re.sub(r"\s+", " ", txt).strip()


def read_sheet_rpc_text(ws) -> str:
    """Rebuild line-oriented RPC XML from worksheet (handles merged / multi-column rows)."""
    lines: list[str] = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        parts: list[str] = []
        for c in row:
            if c is None:
                parts.append("")
            else:
                parts.append(str(c))
        raw = "".join(parts)
        # Preserve leading spaces (indentation); only trim trailing CR/LF artifacts.
        if raw.strip():
            lines.append(raw.rstrip())
    return "\n".join(lines)


def extract_excel_table(ws, required_headers: frozenset[str], max_scan_rows: int = 500) -> tuple[list[str], list[list[str]]]:
    """
    Extract a contiguous table from an Excel sheet.
    - Find the first row whose normalized headers include all `required_headers`
    - Use that row as the header row (from first non-empty to last non-empty cell)
    - Read subsequent rows until an all-empty row
    Returns (headers, rows) as strings.
    """
    rows = list(ws.iter_rows(values_only=True))
    header_i = None
    header_j0 = 0
    header_j1 = 0
    scan_lim = min(len(rows), max_scan_rows)
    for i in range(scan_lim):
        r = rows[i] or ()
        norms: list[str] = []
        for c in r:
            if c is None:
                norms.append("")
            else:
                norms.append(normalize_header(c))
        present = {n for n in norms if n}
        if required_headers.issubset(present):
            header_i = i
            # Find first/last non-empty header cell
            nonempty = [j for j, v in enumerate(r) if v not in (None, "")]
            if not nonempty:
                continue
            header_j0 = nonempty[0]
            header_j1 = nonempty[-1]
            break
    if header_i is None:
        return [], []

    header_row = rows[header_i]
    headers = ["" if header_row[j] is None else str(header_row[j]).strip() for j in range(header_j0, header_j1 + 1)]
    out_rows: list[list[str]] = []
    for i in range(header_i + 1, len(rows)):
        r = rows[i] or ()
        cells: list[str] = []
        any_val = False
        for j in range(header_j0, header_j1 + 1):
            v = r[j] if j < len(r) else None
            s = "" if v is None else str(v).strip()
            if s:
                any_val = True
            cells.append(s)
        if not any_val:
            break
        out_rows.append(cells)
    return headers, out_rows


def build_table_from_rpc_xml(xml: str, kind: str) -> tuple[list[str], list[list[str]]]:
    """
    Build an editable, row-aligned table view from the sheet's RPC XML text.

    The workbook tabs (PDSCH/PUSCH/PRACH) in this project are line-oriented XML, not a literal Excel grid.
    This function derives a "spreadsheet-like" view by extracting the Nth occurrences of key tags.

    kind: "PDSCH" | "PUSCH" | "PRACH"
    """
    text = (xml or "").strip()
    if not text:
        return [], []

    # Endpoint tag differs by direction.
    endpoint_tag = "low-level-tx-endpoint" if kind.upper() == "PDSCH" else "low-level-rx-endpoint"
    endpoints = extract_nth_tag_values(text, endpoint_tag, limit=64)
    prbs = extract_nth_tag_values(text, "number-of-prb", limit=64)
    iq = extract_nth_tag_values(text, "iq-bitwidth", limit=64)
    comp_method = extract_nth_tag_values(text, "compression-method", limit=64)
    comp_type = extract_nth_tag_values(text, "compression-type", limit=64)
    typ = extract_nth_tag_values(text, "type", limit=64)
    scs = extract_nth_tag_values(text, "scs", limit=64)
    center_hz = extract_nth_tag_values(text, "center-of-channel-bandwidth", limit=64)

    # Build a wide view; some columns can be shorter (that's ok).
    headers = [
        "end_point",
        "scs",
        "number-of-prb",
        "iq-bitwidth",
        "compression-type",
        "compression-method",
        "type",
    ]
    if kind.upper() in ("PDSCH", "PUSCH"):
        headers += ["center-freq(MHz)"]

    row_count = max(len(endpoints), len(prbs), len(iq), len(comp_method), len(comp_type), len(typ), len(scs), len(center_hz), 1)
    rows: list[list[str]] = []
    for i in range(row_count):
        row = [
            endpoints[i] if i < len(endpoints) else "",
            scs[i] if i < len(scs) else "",
            prbs[i] if i < len(prbs) else "",
            iq[i] if i < len(iq) else "",
            comp_type[i] if i < len(comp_type) else "",
            comp_method[i] if i < len(comp_method) else "",
            typ[i] if i < len(typ) else "",
        ]
        if kind.upper() in ("PDSCH", "PUSCH"):
            hz = center_hz[i] if i < len(center_hz) else ""
            row.append(hz_to_mhz_string(hz))
        rows.append(row)
    return headers, rows


_LEAF_TAG_RE = re.compile(r"<([a-zA-Z0-9_\-:]+)>([^<]*)</\1>")


def _local_tag(tag: str) -> str:
    return tag.split(":", 1)[-1] if ":" in tag else tag


def _split_endpoint_segments(text: str, endpoint_tag: str) -> tuple[str, list[str]]:
    """
    Split RPC XML text into:
      - prefix before the first endpoint leaf
      - per-endpoint segments [endpoint_i_start .. endpoint_{i+1}_start)
    """
    ep_pat = re.compile(rf"<{re.escape(endpoint_tag)}>([^<]*)</{re.escape(endpoint_tag)}>")
    starts = [m.start() for m in ep_pat.finditer(text)]
    if not starts:
        return text, []
    prefix = text[: starts[0]]
    segs: list[str] = []
    for i, st in enumerate(starts):
        ed = starts[i + 1] if i + 1 < len(starts) else len(text)
        segs.append(text[st:ed])
    return prefix, segs


def _split_block_segments(text: str, block_tag: str) -> tuple[str, list[str]]:
    """
    Split text by container block start tags, e.g. <low-level-tx-endpoints>.
    Returns prefix before first block and per-block slices.
    """
    pat = re.compile(rf"<{re.escape(block_tag)}>")
    starts = [m.start() for m in pat.finditer(text)]
    if not starts:
        return text, []
    prefix = text[: starts[0]]
    segs: list[str] = []
    for i, st in enumerate(starts):
        ed = starts[i + 1] if i + 1 < len(starts) else len(text)
        segs.append(text[st:ed])
    return prefix, segs


def _closed_block_pattern(block_tag: str) -> re.Pattern[str]:
    return re.compile(
        rf"<{re.escape(block_tag)}\b[^>]*>[\s\S]*?</{re.escape(block_tag)}>",
        re.DOTALL,
    )


def _extract_closed_blocks(text: str, block_tag: str) -> list[str]:
    """Return each fully closed ``<block_tag>…</block_tag>`` fragment in document order."""
    return [m.group(0) for m in _closed_block_pattern(block_tag).finditer(text or "")]


def _replace_closed_blocks(text: str, block_tag: str, blocks: list[str]) -> str:
    """Substitute the Nth closed ``block_tag`` block with ``blocks[N]`` (in-order)."""
    idx = 0

    def _sub(m: re.Match[str]) -> str:
        nonlocal idx
        if idx < len(blocks):
            rep = blocks[idx]
            idx += 1
            return rep
        return m.group(0)

    return _closed_block_pattern(block_tag).sub(_sub, text)


def comment_out_endpoint_rows(xml: str, kind: str, off_row_indices_1based: list[int]) -> str:
    """
    Comment-out endpoint rows by 1-based row index for the given sheet kind.
    """
    text = (xml or "").strip()
    if not text:
        return xml
    kind_u = kind.upper()
    endpoint_tag = "low-level-tx-endpoint" if kind_u == "PDSCH" else "low-level-rx-endpoint"
    block_tag = "low-level-tx-endpoints" if kind_u == "PDSCH" else "low-level-rx-endpoints"
    link_tag = "low-level-tx-links" if kind_u == "PDSCH" else "low-level-rx-links"
    # First choice: comment-out the full per-CC chunk so endpoint/carrier/link
    # are wrapped together in one XML comment block.
    cc_chunk_pat = re.compile(
        rf"<{re.escape(block_tag)}\b[^>]*>[\s\S]*?</{re.escape(link_tag)}>",
        re.DOTALL,
    )
    if cc_chunk_pat.search(text):
        return _comment_out_nth_cc_chunks(text, block_tag, link_tag, off_row_indices_1based)

    # Prefer exact CC container blocks (1 block ~= 1 CC) to avoid
    # over-capturing tail text after the last block.
    block_pat = re.compile(
        rf"<{re.escape(block_tag)}\b[^>]*>[\s\S]*?</{re.escape(block_tag)}>",
        re.DOTALL,
    )
    if block_pat.search(text):
        return _comment_out_nth_blocks(text, block_tag, off_row_indices_1based)

    # Fallback for non-standard vendor XML that may not have container blocks.
    prefix, segments = _split_endpoint_segments(text, endpoint_tag)
    if not segments:
        return xml
    off = {i for i in off_row_indices_1based if i > 0}
    out: list[str] = []
    for i, seg in enumerate(segments, start=1):
        seg_keep = seg or ""
        if i in off:
            out.append(f"<!--\n{seg_keep}\n-->")
        else:
            out.append(seg_keep)
    return prefix + "".join(out)


def uncomment_endpoint_rows(xml: str, kind: str) -> str:
    """
    Remove XML comment wrappers around endpoint segments for the given sheet kind.
    This normalizes previously OFF-commented payloads so ON state can be rebuilt cleanly.
    """
    text = (xml or "").strip()
    if not text:
        return xml
    kind_u = kind.upper()
    endpoint_tag = "low-level-tx-endpoint" if kind_u == "PDSCH" else "low-level-rx-endpoint"
    block_tag = "low-level-tx-endpoints" if kind_u == "PDSCH" else "low-level-rx-endpoints"
    # Unwrap only comments that contain the target CC block (fallback: endpoint leaf).
    pat = re.compile(
        rf"<!--\s*(.*?<\s*{re.escape(block_tag)}\s*>.*?</\s*{re.escape(block_tag)}\s*>.*?)\s*-->",
        re.DOTALL,
    )
    out = text
    # Repeat until stable in case of multiple comment blocks.
    while True:
        new_out, n = pat.subn(r"\1", out)
        out = new_out
        if n == 0:
            break
    # Backward compatibility: older files may have endpoint-leaf based comments.
    pat_leaf = re.compile(
        rf"<!--\s*(.*?<\s*{re.escape(endpoint_tag)}\s*>.*?</\s*{re.escape(endpoint_tag)}\s*>.*?)\s*-->",
        re.DOTALL,
    )
    while True:
        new_out, n = pat_leaf.subn(r"\1", out)
        out = new_out
        if n == 0:
            break
    return out


def _comment_out_nth_blocks(xml: str, block_tag: str, off_row_indices_1based: list[int]) -> str:
    text = xml or ""
    pat = re.compile(
        rf"<{re.escape(block_tag)}\b[^>]*>[\s\S]*?</{re.escape(block_tag)}>",
        re.DOTALL,
    )
    off = {i for i in off_row_indices_1based if i > 0}
    out: list[str] = []
    last = 0
    idx = 0
    for m in pat.finditer(text):
        idx += 1
        out.append(text[last : m.start()])
        blk = m.group(0)
        if idx in off:
            out.append(f"<!--\n{blk.strip(chr(10))}\n-->")
        else:
            out.append(blk)
        last = m.end()
    out.append(text[last:])
    return "".join(out)


def _comment_out_nth_cc_chunks(xml: str, start_tag: str, end_tag: str, off_row_indices_1based: list[int]) -> str:
    """
    Comment-out Nth CC chunks spanning:
      <start_tag> ... </end_tag>
    """
    text = xml or ""
    pat = re.compile(
        rf"<{re.escape(start_tag)}\b[^>]*>[\s\S]*?</{re.escape(end_tag)}>",
        re.DOTALL,
    )
    off = {i for i in off_row_indices_1based if i > 0}
    out: list[str] = []
    last = 0
    idx = 0
    for m in pat.finditer(text):
        idx += 1
        out.append(text[last : m.start()])
        blk = m.group(0)
        if idx in off:
            out.append(f"<!--\n{blk.strip(chr(10))}\n-->")
        else:
            out.append(blk)
        last = m.end()
    out.append(text[last:])
    return "".join(out)


def _uncomment_block_comments(xml: str, block_tag: str) -> str:
    text = (xml or "").strip()
    if not text:
        return xml
    pat = re.compile(
        rf"<!--\s*(<\s*{re.escape(block_tag)}\b[\s\S]*?</\s*{re.escape(block_tag)}\s*>)\s*-->",
        re.DOTALL,
    )
    out = text
    while True:
        new_out, n = pat.subn(r"\1", out)
        out = new_out
        if n == 0:
            break
    return out


def comment_out_active_rows(xml: str, off_row_indices_1based: list[int]) -> str:
    """
    Comment-out ACTIVE CC rows for both tx/rx-array-carrier(s) blocks by 1-based index.
    """
    text = (xml or "").strip()
    if not text:
        return xml
    tx_tag = "tx-array-carriers" if re.search(r"<tx-array-carriers\b", text) else "tx-array-carrier"
    rx_tag = "rx-array-carriers" if re.search(r"<rx-array-carriers\b", text) else "rx-array-carrier"
    out = _comment_out_nth_blocks(text, tx_tag, off_row_indices_1based)
    out = _comment_out_nth_blocks(out, rx_tag, off_row_indices_1based)
    return out


def uncomment_active_rows(xml: str) -> str:
    """
    Remove ACTIVE row comments created by comment_out_active_rows.
    """
    text = (xml or "").strip()
    if not text:
        return xml
    out = _uncomment_block_comments(text, "tx-array-carrier")
    out = _uncomment_block_comments(out, "rx-array-carrier")
    out = _uncomment_block_comments(out, "tx-array-carriers")
    out = _uncomment_block_comments(out, "rx-array-carriers")
    return out


_TAC_CARRIER_RE = re.compile(r"^tac_", re.IGNORECASE)
_RAC_CARRIER_RE = re.compile(r"^rac_", re.IGNORECASE)


def _array_carrier_name_from_row(
    headers: list[str], row: list[str], prefix_re: re.Pattern[str]
) -> str:
    """Pick tx/rx array-carrier name from a Control-Sheet detail row."""
    for j, _h in enumerate(headers):
        if j >= len(row):
            continue
        val = (row[j] or "").strip()
        if val and prefix_re.match(val):
            return val
    for j, h in enumerate(headers):
        nh = normalize_header(h)
        if "array-carrier" not in nh and "array carrier" not in nh:
            continue
        val = (row[j] if j < len(row) else "").strip()
        if val:
            return val
    return ""


def _replace_active_carrier_blocks(active_xml: str, carrier_inner: str) -> str:
    """Replace list entries inside ``user-plane-configuration`` for ACTIVE RPC."""
    text = (active_xml or "").strip()
    if not text:
        return active_xml
    pat = re.compile(
        r"(<user-plane-configuration\b[^>]*>)([\s\S]*?)(</user-plane-configuration>)",
        re.IGNORECASE,
    )
    m = pat.search(text)
    if not m:
        return active_xml
    inner = carrier_inner.strip()
    if inner and not inner.endswith("\n"):
        inner += "\n"
    return text[: m.start()] + m.group(1) + "\n" + inner + "      " + m.group(3) + text[m.end() :]


def apply_active_from_control_tables(
    active_xml: str,
    pdsch: tuple[list[str], list[list[str]]],
    pusch: tuple[list[str], list[list[str]]],
    off_row_indices_1based: list[int] | None = None,
) -> tuple[str, list[str]]:
    """
    Rebuild ACTIVE ``tx-array-carriers`` / ``rx-array-carriers`` lists from Control-Sheet
    PDSCH/PUSCH detail tables (``tac_*`` / ``rac_*``). The ACTIVE worksheet skeleton often
    has empty or missing ``<name>`` leaves; carrier keys must come from the grid.
    """
    warns: list[str] = []
    text = uncomment_active_rows((active_xml or "").strip())
    if not text:
        warns.append("ACTIVE: empty RPC template")
        return active_xml, warns

    off = {i for i in (off_row_indices_1based or []) if i > 0}
    ph, pr = pdsch
    uh, ur = pusch
    blocks: list[str] = []

    if ph and pr:
        for i, row in enumerate(pr):
            if (i + 1) in off:
                continue
            if not _is_control_uplane_data_row(row):
                continue
            name = _array_carrier_name_from_row(ph, row, _TAC_CARRIER_RE)
            if not name:
                warns.append(f"ACTIVE: PDSCH row {i + 1}: missing tx-array-carrier name (tac_*)")
                continue
            blocks.append(
                "        <tx-array-carriers>\n"
                f"          <name>{name}</name>\n"
                "          <active>ACTIVE</active>\n"
                "        </tx-array-carriers>"
            )
    else:
        warns.append("ACTIVE: no PDSCH Control-Sheet table — keeping ACTIVE worksheet carriers")

    if uh and ur:
        for i, row in enumerate(ur):
            if (i + 1) in off:
                continue
            if not _is_control_uplane_data_row(row):
                continue
            name = _array_carrier_name_from_row(uh, row, _RAC_CARRIER_RE)
            if not name:
                warns.append(f"ACTIVE: PUSCH row {i + 1}: missing rx-array-carrier name (rac_*)")
                continue
            blocks.append(
                "        <rx-array-carriers>\n"
                f"          <name>{name}</name>\n"
                "          <active>ACTIVE</active>\n"
                "        </rx-array-carriers>"
            )
    elif ph and pr:
        warns.append("ACTIVE: no PUSCH Control-Sheet table — rx-array-carriers not rebuilt")

    if blocks:
        text = _replace_active_carrier_blocks(text, "\n".join(blocks))
    return text, warns


def parse_control_detail_cc_on_flags(ws) -> list[bool]:
    """Read column-A ON/OFF flags for PDSCH detail rows (same order as section 3 grid)."""
    rows = list(ws.iter_rows(values_only=True))
    si = None
    for i, row in enumerate(rows):
        b = row[1] if row and len(row) > 1 else None
        if b is not None and normalize_header(str(b)) == "3. pdsch":
            si = i
            break
    if si is None:
        return []
    header_i = None
    for j in range(si + 1, min(si + 12, len(rows))):
        row = rows[j] or ()
        c = row[2] if len(row) > 2 else None
        d = row[3] if len(row) > 3 else None
        if normalize_header(str(c or "")) == "name" and normalize_header(str(d or "")) == "scs":
            header_i = j
            break
    if header_i is None:
        return []
    flags: list[bool] = []
    for j in range(header_i + 1, len(rows)):
        row = rows[j] or ()
        b = row[1] if len(row) > 1 else None
        if b is not None and str(b).strip() and re.match(r"^\d+\.", str(b).strip()):
            break
        cells = [row[c] if c < len(row) else None for c in range(2, 12)]
        if not any(c is not None and str(c).strip() for c in cells):
            break
        if len(row) > 2 and normalize_header(str(row[2] or "")) in ("NAME", "END-POINT", "END POINT"):
            continue
        if not _is_control_uplane_data_row(
            ["" if (c >= len(row) or row[c] is None) else str(row[c]).strip() for c in range(2, 12)]
        ):
            continue
        enabled = True
        if len(row) > 0 and row[0] is not None:
            s = str(row[0]).strip().upper()
            enabled = s in ("ON", "1", "TRUE", "YES", "Y", "O", "ENABLE", "ENABLED")
        flags.append(enabled)
    return flags


def derive_full_table_from_rpc(xml: str, kind: str) -> tuple[list[str], list[list[str]]]:
    """
    Derive a 'full' editable table from line-oriented RPC XML.

    We split the RPC text by endpoint markers and then collect ALL leaf tags
    (<tag>value</tag> on the same line) within each segment as columns.
    """
    text = (xml or "").strip()
    if not text:
        return [], []

    endpoint_tag = "low-level-tx-endpoint" if kind.upper() == "PDSCH" else "low-level-rx-endpoint"
    _prefix, segments = _split_endpoint_segments(text, endpoint_tag)
    if not segments:
        return [], []
    endpoint_vals = extract_nth_tag_values(text, endpoint_tag, limit=256)

    kind_u = kind.upper()
    if kind_u == "PDSCH":
        columns: list[tuple[str, str]] = [
            ("end_point", endpoint_tag),
            ("scs", "scs"),
            ("number-of-prb", "number-of-prb"),
            ("eaxc-id", "eaxc-id"),
            ("processing-element", "processing-element"),
            ("tx-array-carrier", "tx-array-carrier"),
            ("low-level-tx-endpoint", endpoint_tag),
            ("iq-bitwidth", "iq-bitwidth"),
            ("compression", "compression-method"),
            ("Center-FREQ", "center-of-channel-bandwidth"),
            ("Channel-BW", "channel-bandwidth"),
            ("type", "type"),
            ("gain", "gain"),
            ("T-da-offset", "t-da-offset"),
        ]
    elif kind_u == "PUSCH":
        columns = [
            ("end_point", endpoint_tag),
            ("scs", "scs"),
            ("number-of-prb", "number-of-prb"),
            ("ul-fft-sampling-offset", "ul-fft-sampling-offset"),
            ("eaxc-id", "eaxc-id"),
            ("processing-element", "processing-element"),
            ("rx-array-carrier", "rx-array-carrier"),
            ("low-level-rx-endpoint", endpoint_tag),
            ("iq-bitwidth", "iq-bitwidth"),
            ("compression", "compression-method"),
            ("Center-FREQ", "center-of-channel-bandwidth"),
            ("Channel-BW", "channel-bandwidth"),
            ("type", "type"),
            ("n-ta-offset", "n-ta-offset"),
            ("T-au-offset", "t-au-offset"),
            ("Gain-Correction", "gain-correction"),
        ]
    else:  # PRACH
        columns = [
            ("low-level-rx-endpoint", endpoint_tag),
            ("scs", "scs"),
            ("number-of-prb", "number-of-prb"),
            ("ul-fft-sampling-offset", "ul-fft-sampling-offset"),
            ("eaxc-id", "eaxc-id"),
            ("processing-element", "processing-element"),
            ("rx-array-carrier", "rx-array-carrier"),
            ("iq-bitwidth", "iq-bitwidth"),
            ("compression", "compression-method"),
        ]

    row_n = len(endpoint_vals)
    if row_n < 1:
        return headers, []
    tag_values: dict[str, list[str]] = {}
    unique_tags = {t for _, t in columns}
    for t in unique_tags:
        vals = extract_nth_tag_values(text, t, limit=4096)
        # Some vendors use <compression-type> instead of <compression-method>.
        if t == "compression-method" and not vals:
            vals = extract_nth_tag_values(text, "compression-type", limit=4096)
        if len(vals) > row_n:
            vals = vals[-row_n:]
        tag_values[t] = vals
    # Optional compression variants used by some vendors.
    comp_type_vals = extract_nth_tag_values(text, "compression-type", limit=4096)
    exp_vals = extract_nth_tag_values(text, "exponent", limit=4096)
    if len(comp_type_vals) > row_n:
        comp_type_vals = comp_type_vals[-row_n:]
    if len(exp_vals) > row_n:
        exp_vals = exp_vals[-row_n:]

    headers = [h for h, _ in columns]
    rows: list[list[str]] = []
    for i in range(row_n):
        row: list[str] = []
        for _h, tag in columns:
            vals = tag_values.get(tag, [])
            val = vals[i] if i < len(vals) else ""
            if tag == "center-of-channel-bandwidth":
                val = hz_to_mhz_string(val)
            if tag == "compression-method" and val.strip().upper() == "BLOCK_FLOATING_POINT":
                val = "BFP"
            if tag == "compression-method" and val.strip().upper() == "STATIC":
                val = "no comp"
            if tag == "compression-method" and not val:
                e = exp_vals[i].strip() if i < len(exp_vals) else ""
                ct = comp_type_vals[i].strip() if i < len(comp_type_vals) else ""
                if e:
                    val = "exponent"
                elif ct.upper() == "STATIC":
                    val = "no comp"
                elif ct.upper() == "BLOCK_FLOATING_POINT":
                    val = "BFP"
            row.append(val)
        rows.append(row)
    return headers, rows


def apply_full_table_to_rpc(xml: str, kind: str, headers: list[str], rows: list[list[str]]) -> tuple[str, list[str]]:
    """
    Apply edited table values back into the RPC by segment-local leaf tag replacement.

    - Splits RPC text by endpoint marker
    - For each row i, replaces <tag>...</tag> within segment i (first occurrence)
    - center-freq(MHz) -> <center-of-channel-bandwidth> in Hz
    """
    warns: list[str] = []
    text = (xml or "").strip()
    if not text or not headers or not rows:
        return xml, warns

    endpoint_tag = "low-level-tx-endpoint" if kind.upper() == "PDSCH" else "low-level-rx-endpoint"
    prefix, segments = _split_endpoint_segments(text, endpoint_tag)
    if not segments:
        warns.append(f"{kind}: no endpoint segments found for <{endpoint_tag}>")
        return xml, warns

    kind_u = kind.upper()
    if kind_u == "PDSCH":
        ordered_tags = [
            endpoint_tag, "scs", "number-of-prb", "eaxc-id", "processing-element",
            "tx-array-carrier", endpoint_tag, "iq-bitwidth", "compression-method",
            "center-of-channel-bandwidth", "channel-bandwidth", "type", "gain", "t-da-offset",
        ]
    elif kind_u == "PUSCH":
        ordered_tags = [
            endpoint_tag, "scs", "number-of-prb", "ul-fft-sampling-offset", "eaxc-id",
            "processing-element", "rx-array-carrier", endpoint_tag, "iq-bitwidth", "compression-method",
            "center-of-channel-bandwidth", "channel-bandwidth", "type", "n-ta-offset", "t-au-offset", "gain-correction",
        ]
    else:
        ordered_tags = [
            endpoint_tag, "scs", "number-of-prb", "ul-fft-sampling-offset", "eaxc-id",
            "processing-element", "rx-array-carrier", "iq-bitwidth", "compression-method",
        ]
    # Safety fallback if UI/header count diverges
    if len(ordered_tags) != len(headers):
        ordered_tags = []
        for h in headers:
            nh = normalize_header(h)
            if "center" in nh and ("freq" in nh or "bandwidth" in nh):
                ordered_tags.append("center-of-channel-bandwidth")
            elif nh in ("end point", "end_point", "low level tx endpoint", "low level rx endpoint"):
                ordered_tags.append(endpoint_tag)
            else:
                ordered_tags.append(nh.replace(" ", "-"))

    # Replace by global Nth occurrence aligned to the last row_n values for each tag.
    out = text
    row_n = len(rows)
    iq_col = None
    for ii, t in enumerate(ordered_tags):
        if t == "iq-bitwidth":
            iq_col = ii
            break

    def _nth_compression_block_span(text_in: str, index0: int) -> tuple[int, int] | None:
        """
        Return (start,end) span for the Nth <compression>...</compression> block (0-based).
        """
        pat = re.compile(r"<compression>\s*[\s\S]*?</compression>", re.DOTALL)
        n = -1
        for m in pat.finditer(text_in):
            n += 1
            if n == index0:
                return m.start(), m.end()
        return None

    def _set_tag_in_compression(text_in: str, idx0: int, tag: str, val: str) -> tuple[str, bool]:
        span = _nth_compression_block_span(text_in, idx0)
        if span is None:
            return text_in, False
        st, ed = span
        block = text_in[st:ed]
        pat = re.compile(rf"(<{re.escape(tag)}>\s*)([^<]*)(\s*</{re.escape(tag)}>)")
        if pat.search(block):
            block2 = pat.sub(lambda m: f"{m.group(1)}{val}{m.group(3)}", block, count=1)
        else:
            # Insert just before the closing </compression> within this same block.
            close_i = block.rfind("</compression>")
            if close_i < 0:
                return text_in, False
            insert = f"  <{tag}>{val}</{tag}>\n"
            block2 = block[:close_i] + insert + block[close_i:]
        return text_in[:st] + block2 + text_in[ed:], True

    def _remove_tag_in_compression(text_in: str, idx0: int, tag: str) -> tuple[str, bool]:
        span = _nth_compression_block_span(text_in, idx0)
        if span is None:
            return text_in, False
        st, ed = span
        block = text_in[st:ed]
        pat = re.compile(rf"\s*<{re.escape(tag)}>\s*[^<]*\s*</{re.escape(tag)}>\s*\n?", re.DOTALL)
        if not pat.search(block):
            return text_in, False
        block2 = pat.sub("\n", block, count=1)
        return text_in[:st] + block2 + text_in[ed:], True

    def _nth_block_span(text_in: str, block_tag: str, index0: int) -> tuple[int, int] | None:
        pat = re.compile(rf"<{re.escape(block_tag)}\b[^>]*>[\s\S]*?</{re.escape(block_tag)}>", re.DOTALL)
        n = -1
        for m in pat.finditer(text_in):
            n += 1
            if n == index0:
                return m.start(), m.end()
        return None

    def _set_tag_in_nth_block(text_in: str, block_tag: str, idx0: int, tag: str, val: str) -> tuple[str, bool]:
        span = _nth_block_span(text_in, block_tag, idx0)
        if span is None:
            return text_in, False
        st, ed = span
        block = text_in[st:ed]
        pat = re.compile(rf"(<{re.escape(tag)}>\s*)([^<]*)(\s*</{re.escape(tag)}>)")
        if pat.search(block):
            block2 = pat.sub(lambda m: f"{m.group(1)}{val}{m.group(3)}", block, count=1)
        else:
            close_i = block.rfind(f"</{block_tag}>")
            if close_i < 0:
                return text_in, False
            ins = f"          <{tag}>{val}</{tag}>\n"
            block2 = block[:close_i] + ins + block[close_i:]
        return text_in[:st] + block2 + text_in[ed:], True

    def _remove_tag_in_nth_block(text_in: str, block_tag: str, idx0: int, tag: str) -> tuple[str, bool]:
        span = _nth_block_span(text_in, block_tag, idx0)
        if span is None:
            return text_in, False
        st, ed = span
        block = text_in[st:ed]
        pat = re.compile(rf"\s*<{re.escape(tag)}>\s*[^<]*\s*</{re.escape(tag)}>\s*\n?", re.DOTALL)
        if not pat.search(block):
            return text_in, False
        block2 = pat.sub("\n", block, count=1)
        return text_in[:st] + block2 + text_in[ed:], True

    def _apply_compression_value(text_in: str, idx: int, user_val: str) -> tuple[str, list[str]]:
        w: list[str] = []
        out_local = text_in
        raw = (user_val or "").strip()
        if not raw:
            return out_local, w
        lv = raw.lower().replace("_", " ").replace("-", " ").strip()
        exp_num = "4"
        m_exp = re.match(r"^exponent(?:\s*[:=]\s*(\d+))?$", lv)
        if m_exp:
            if m_exp.group(1):
                exp_num = m_exp.group(1)
            # Ensure compression-type STATIC + exponent inside compression block.
            out_local, _ = _set_tag_in_compression(out_local, idx, "compression-type", "STATIC")
            out_local, _ = _set_tag_in_compression(out_local, idx, "exponent", exp_num)
            out_local, _ = _remove_tag_in_compression(out_local, idx, "compression-method")
            return out_local, w
        if lv in {"bfp", "block floating point", "block_floating_point"}:
            # Ensure compression-type STATIC + compression-method inside compression block.
            out_local, _ = _set_tag_in_compression(out_local, idx, "compression-type", "STATIC")
            out_local, _ = _set_tag_in_compression(out_local, idx, "compression-method", "BLOCK_FLOATING_POINT")
            out_local, _ = _remove_tag_in_compression(out_local, idx, "exponent")
            return out_local, w
        # Invalid value: ignore and keep previous XML value.
        w.append(f"{kind}: ignored invalid compression value '{raw}' at row {idx + 1}")
        return out_local, w

    for j, tag in enumerate(ordered_tags):
        if not tag or tag == endpoint_tag:
            continue

        # User rule:
        # 1) empty => delete tag
        # 2) value (including 0) => add/update tag
        if (kind_u == "PDSCH" and tag == "t-da-offset") or (kind_u == "PUSCH" and tag == "t-au-offset"):
            carrier_block = "tx-array-carriers" if kind_u == "PDSCH" else "rx-array-carriers"
            for i in range(row_n):
                if j >= len(rows[i]):
                    continue
                val = (rows[i][j] or "").strip()
                if val == "":
                    out, _ = _remove_tag_in_nth_block(out, carrier_block, i, tag)
                else:
                    out, ok = _set_tag_in_nth_block(out, carrier_block, i, tag, val)
                    if not ok:
                        warns.append(f"{kind}: could not set <{tag}> at row {i + 1}")
            continue

        target_tag = tag
        all_vals = extract_nth_tag_values(out, target_tag, limit=4096)
        if target_tag == "compression-method" and not all_vals:
            target_tag = "compression-type"
            all_vals = extract_nth_tag_values(out, target_tag, limit=4096)
        if not all_vals:
            warns.append(f"{kind}: missing tag <{target_tag}>")
            continue
        base = max(0, len(all_vals) - row_n)  # tail alignment
        for i in range(row_n):
            if j >= len(rows[i]):
                continue
            val = (rows[i][j] or "").strip()
            iq16_row = False
            if iq_col is not None and iq_col < len(rows[i]):
                iq16_row = (rows[i][iq_col] or "").strip() == "16"
            if tag == "compression-method":
                lv = val.strip().lower().replace("_", " ").replace("-", " ").strip()
                if iq16_row and lv in {"", "no comp", "none", "no compression", "static"}:
                    out, _ = _set_tag_in_compression(out, base + i, "compression-type", "STATIC")
                    out, _ = _remove_tag_in_compression(out, base + i, "compression-method")
                    out, _ = _remove_tag_in_compression(out, base + i, "exponent")
                    continue
                out, ws = _apply_compression_value(out, base + i, val)
                warns.extend(ws)
                continue
            if not val:
                # Safety: empty cell means "keep existing value" (do not delete),
                # to avoid accidental data loss from table clears.
                continue
            if target_tag == "center-of-channel-bandwidth":
                hz = mhz_to_hz_string(val)
                if not hz:
                    warns.append(f"{kind}: invalid center-freq at row {i+1}")
                    continue
                val = hz
            elif target_tag == "compression-method":
                if val.strip().upper() == "BFP":
                    val = "BLOCK_FLOATING_POINT"
            out, ok = replace_nth_tag(out, target_tag, val, base + i)
            if not ok:
                warns.append(f"{kind}: could not replace <{target_tag}> index {base + i}")
    return out, warns


def parse_baselines(xml_interfaces: str, xml_processing: str) -> dict[str, str]:
    """Extract default strings used for substitution (first occurrence)."""
    b: dict[str, str] = {}

    def one(pattern: str, text: str) -> str | None:
        m = re.search(pattern, text)
        return m.group(1) if m else None

    b["cu_if_name"] = one(r"<interface>\s*<name>([^<]+)</name>", xml_interfaces) or ""
    # More robust: interfaces block
    x = xml_interfaces
    if not b["cu_if_name"]:
        n = one(r"<name>([^<]+)</name>", x)
        b["cu_if_name"] = n or ""
    b["cu_base_if"] = one(r"<base-interface[^>]*>([^<]+)</base-interface>", xml_interfaces) or ""
    b["cu_vlan"] = one(r"<vlan-id[^>]*>([^<]+)</vlan-id>", xml_interfaces) or ""
    b["cu_mac"] = one(r"<mac-address[^>]*>([^<]+)</mac-address>", xml_interfaces) or ""

    p = xml_processing
    b["pe_name"] = one(r"<ru-elements>\s*<name>([^<]+)</name>", p) or one(r"<name>([^<]+)</name>", p) or ""
    b["odu_mac"] = one(r"<o-du-mac-address>([^<]+)</o-du-mac-address>", p) or ""
    b["ru_mac_pe"] = one(r"<ru-mac-address>([^<]+)</ru-mac-address>", p) or ""

    return b


def _is_cc_header_row(row: tuple[Any, ...]) -> bool:
    norms = []
    for c in row[:20]:
        norms.append(normalize_header(c) if c is not None else "")
    blob = "|".join(n for n in norms if n)
    return "cc" in blob and "number-of-prb" in blob and ("ant" in blob or "ant#" in blob)


def _looks_like_major_section_heading_b(bb: Any) -> bool:
    """Col B numbered sections like \"5. Title\" stop the CC grid; decimals like \"5.0\" do not."""
    if bb is None:
        return False
    s = str(bb).strip()
    if not s:
        return False
    m = re.match(r"^(\d+)\.(.*)$", s)
    if not m:
        return False
    rest = m.group(2).strip()
    if not rest:
        return True
    if rest[0].isdigit():
        return False
    return True


def _find_cc_column(cc_headers_norm: dict[str, int], synonyms: frozenset[str]) -> int | None:
    for hk, idx in cc_headers_norm.items():
        if hk in synonyms:
            return idx
        for syn in synonyms:
            if syn and (syn in hk or hk in syn):
                return idx
    return None


def parse_control_sheet(ws) -> tuple[dict[str, Any], list[dict[str, Any]], list[str]]:
    """
    Scan Control-Sheet for label=value pairs (cols C+D) plus 8CC table after header row.
    """
    warnings: list[str] = []
    fields: dict[str, Any] = {}
    rows = list(ws.iter_rows(values_only=True))

    cc_header_idx = None
    for i, row in enumerate(rows):
        if row and _is_cc_header_row(row):
            cc_header_idx = i
            break

    cc_headers_norm: dict[str, int] = {}
    cc_rows: list[dict[str, Any]] = []
    cc_end_idx = len(rows)
    if cc_header_idx is not None:
        hr = rows[cc_header_idx]
        for idx, cell in enumerate(hr):
            if cell is None:
                continue
            k = normalize_header(str(cell))
            if k:
                cc_headers_norm[k] = idx
        for canon, synonyms in CC_COL_MAP.items():
            if _find_cc_column(cc_headers_norm, synonyms) is None:
                warnings.append(f"[Control-Sheet] CC column missing for {canon}")

        for j in range(cc_header_idx + 1, len(rows)):
            r = rows[j]
            bb = r[1] if r and len(r) > 1 else None
            if _looks_like_major_section_heading_b(bb):
                cc_end_idx = j
                break

        for j in range(cc_header_idx + 1, cc_end_idx):
            row = rows[j]
            if not row:
                continue
            if not any(cell not in (None, "") for cell in row[:16]):
                continue
            rdict: dict[str, Any] = {}
            bad = False
            for canon, synonyms in CC_COL_MAP.items():
                coli = _find_cc_column(cc_headers_norm, synonyms)
                if coli is None or coli >= len(row):
                    bad = True
                    break
                rdict[canon] = row[coli]
            if bad:
                continue
            cc_rows.append(rdict)

    section = ""
    for i, row in enumerate(rows):
        if not row:
            continue
        if cc_header_idx is not None and cc_header_idx < i < cc_end_idx:
            continue

        b = row[1] if len(row) > 1 else None
        if b is not None and str(b).strip():
            sb = str(b).strip()
            if re.match(r"^\d+[\-.]", sb):
                section = sb

        c = row[2] if len(row) > 2 else None
        d = row[3] if len(row) > 3 else None
        lk = normalize_header(str(c).strip()) if c is not None and str(c).strip() else ""
        if lk == "vlan name":
            continue
        if lk == "name" and d:
            sec_l = normalize_header(section)
            if "cuplane" in sec_l or section.startswith("1."):
                fields["cu_if_name"] = d
            elif "processing" in sec_l or section.startswith("2"):
                fields["pe_name"] = d
            continue

        if lk and d is not None and str(d).strip() != "":
            canon = SIMPLE_LABELS.get(lk)
            if canon:
                fields[canon] = d

    return fields, cc_rows, warnings


# Control-Sheet 한글 헤더/영문 헤더 대응
SIMPLE_LABELS = {
    "base interface": "cu_base_if",
    "base-interface": "cu_base_if",
    "vlan id": "cu_vlan",
    # normalize_header turns hyphens into spaces ("oru mac address")
    "oru mac address": "cu_mac",
    "oru mac-address": "cu_mac",
    "odu mac address": "odu_mac",
    "odu mac-address": "odu_mac",
    "vlan name": "vlan_hint",
}


CC_COL_MAP: dict[str, frozenset[str]] = {
    "enabled": frozenset({"address(on/off)", "on/off"}),
    "cc": frozenset({"cc"}),
    "ant": frozenset({"ant#", "ant"}),
    "scs": frozenset({"scs"}),
    "prb": frozenset({"number-of-prb"}),
    "iq": frozenset({"iq-bitwidth"}),
    "compression": frozenset({"compression", "comp"}),
    "typ": frozenset({"type"}),
    "dl_mhz": frozenset({"dl center"}),
    "bw": frozenset({"bw"}),
    "ul_mhz": frozenset({"ul center"}),
    "t_au": frozenset({"t-au", "t-au-offset", "t au offset"}),
    "n_ta": frozenset({"n-ta", "n-ta-offset", "n ta offset"}),
    "t_du": frozenset({"t-du", "t-du-offset", "t-da-offset", "t du offset"}),
    "n_da": frozenset({"n-da", "n-da-offset", "n da offset", "t-du", "t-da-offset"}),
}


def _coerce_hz_string(mhz_any: Any) -> str | None:
    if mhz_any is None:
        return None
    s = str(mhz_any).strip()
    if not s:
        return None
    try:
        hz = float(s) * 1_000_000.0
        return str(int(round(hz))) if hz == int(round(hz)) else str(round(hz))
    except ValueError:
        return None


def mhz_to_hz_string(mhz_any: Any) -> str | None:
    """Public wrapper: MHz -> Hz string."""
    return _coerce_hz_string(mhz_any)


def _replace_nth_tag(xml: str, tag: str, new_inner: str, index: int) -> tuple[str, bool]:
    pat = re.compile(rf"<{re.escape(tag)}>([^<]*)</{re.escape(tag)}>")
    mi = None
    n = -1
    for m in pat.finditer(xml):
        n += 1
        if n == index:
            mi = m
            break
    if mi is None:
        return xml, False
    new_chunk = f"<{tag}>{new_inner}</{tag}>"
    return xml[: mi.start()] + new_chunk + xml[mi.end() :], True


def _remove_nth_tag(xml: str, tag: str, index: int) -> tuple[str, bool]:
    pat = re.compile(rf"<{re.escape(tag)}>([^<]*)</{re.escape(tag)}>")
    mi = None
    n = -1
    for m in pat.finditer(xml):
        n += 1
        if n == index:
            mi = m
            break
    if mi is None:
        return xml, False
    return xml[: mi.start()] + xml[mi.end() :], True


def extract_nth_tag_values(xml: str, tag: str, limit: int = 8) -> list[str]:
    """Extract up to N occurrences of <tag>inner</tag> in order."""
    pat = re.compile(rf"<{re.escape(tag)}>([^<]*)</{re.escape(tag)}>")
    out: list[str] = []
    for m in pat.finditer(xml):
        out.append((m.group(1) or "").strip())
        if len(out) >= limit:
            break
    return out


def replace_nth_tag(xml: str, tag: str, new_inner: str, index: int) -> tuple[str, bool]:
    """Public wrapper for replacing the Nth <tag>inner</tag>."""
    return _replace_nth_tag(xml, tag, new_inner, index)


def duplicate_pdsch_eaxc_id(
    pdsch_xml: str,
    from_index: int = 0,
    to_index: int = 1,
) -> tuple[str, list[str]]:
    """Copy eAxC-ID from CC `from_index` into CC `to_index` (negative conformance 3.1.10.2)."""
    warns: list[str] = []
    text = (pdsch_xml or "").strip()
    if not text:
        warns.append("duplicate_eaxc: empty PDSCH payload")
        return "", warns
    ids = extract_nth_tag_values(text, "eaxc-id", limit=64)
    if from_index >= len(ids):
        warns.append(f"duplicate_eaxc: source index {from_index} missing (have {len(ids)} eAxC-ID)")
        return "", warns
    if to_index >= len(ids):
        warns.append(f"duplicate_eaxc: target index {to_index} missing (have {len(ids)} eAxC-ID)")
        return "", warns
    dup_val = ids[from_index]
    out, ok = _replace_nth_tag(text, "eaxc-id", dup_val, to_index)
    if not ok:
        warns.append(f"duplicate_eaxc: could not replace eAxC-ID at index {to_index}")
        return "", warns
    return out, warns


def remove_nth_tag(xml: str, tag: str, index: int) -> tuple[str, bool]:
    """Public wrapper for removing the Nth <tag>inner</tag> node."""
    return _remove_nth_tag(xml, tag, index)


def hz_to_mhz_string(hz_any: Any) -> str:
    """Convert Hz string/number to MHz string (for UI)."""
    if hz_any is None:
        return ""
    s = str(hz_any).strip()
    if not s:
        return ""
    try:
        mhz = float(s) / 1_000_000.0
        # Keep compact, but not too lossy.
        return f"{mhz:.6f}".rstrip("0").rstrip(".")
    except ValueError:
        return ""


def bandwidth_to_hz_string(bw_any: Any) -> str | None:
    """Convert UI/Control-Sheet bandwidth to Hz; large values are already Hz."""
    if bw_any is None:
        return None
    s = str(bw_any).strip()
    if not s:
        return None
    try:
        val = float(s)
    except ValueError:
        return None
    hz = val if abs(val) >= 10000 else val * 1_000_000.0
    return str(int(round(hz))) if hz == int(round(hz)) else str(round(hz))


def apply_physical_cc_rows(pdsch: str, pusch: str, prach: str, cc_rows: list[dict[str, Any]]) -> tuple[str, str, str, list[str]]:
    """Apply Control-Sheet carrier table to PDSCH / PUSCH / PRACH numeric tags."""
    warns: list[str] = []
    for i in range(min(16, len(cc_rows))):
        row = cc_rows[i]

        hz_dl = _coerce_hz_string(row.get("dl_mhz"))
        if hz_dl:
            pdsch, ok = _replace_nth_tag(pdsch, "center-of-channel-bandwidth", hz_dl, i)
            if not ok:
                warns.append(f"PDSCH: could not replace center-of-channel-bandwidth index {i}")

        hz_ul = _coerce_hz_string(row.get("ul_mhz"))
        if hz_ul:
            pusch, ok = _replace_nth_tag(pusch, "center-of-channel-bandwidth", hz_ul, i)
            if not ok:
                warns.append(f"PUSCH: could not replace center-of-channel-bandwidth index {i}")

        prb = row.get("prb")
        if prb is not None and str(prb).strip() != "":
            pv = str(int(prb)) if isinstance(prb, (int, float)) else str(prb).strip()
            pdsch, ok = _replace_nth_tag(pdsch, "number-of-prb", pv, i)
            if not ok:
                warns.append(f"PDSCH: could not replace number-of-prb index {i}")
            pusch, ok = _replace_nth_tag(pusch, "number-of-prb", pv, i)
            if not ok:
                warns.append(f"PUSCH: could not replace number-of-prb index {i}")
            # PRACH per-endpoint PRB is owned by Control-Sheet section 5 (detail table).

    return pdsch, pusch, prach, warns


def strip_rpc_wrapper_for_netconf_cli(xml: str) -> str:
    """
    netopeer2-cli ``user-rpc --content=FILE`` expects the RPC operation body only
    (e.g. ``<edit-config>…</edit-config>``), not a full ``<rpc>`` envelope.
    """
    text = (xml or "").strip()
    if not text:
        return text
    text = re.sub(r"^\s*<\?xml[^>]*\?>\s*", "", text, flags=re.IGNORECASE)
    m = re.search(r"<rpc\b[^>]*>([\s\S]*)</rpc\s*>", text, flags=re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return text


def _upsert_first_tag(block: str, tag: str, value: str) -> str:
    """Insert or replace the first ``<tag>…</tag>`` inside a block fragment."""
    if not value:
        return block
    pat = re.compile(rf"<{re.escape(tag)}>\s*[^<]*\s*</{re.escape(tag)}>", re.IGNORECASE)
    if pat.search(block):
        return pat.sub(f"<{tag}>{value}</{tag}>", block, count=1)
    m_open = re.search(r"^(\s*<\w+[^>]*>\s*)", block)
    if not m_open:
        return block
    ins = f"{m_open.group(1)}<{tag}>{value}</{tag}>\n"
    return ins + block[len(m_open.group(1)) :]


def _upsert_nested_prb_scs(block: str, scs: str, prb: str) -> str:
    """Merge Control-Sheet PRB/SCS; empty PRB removes the whole ``number-of-prb-per-scs`` block."""
    scs = (scs or "").strip()
    prb = (prb or "").strip()
    pat = re.compile(r"\s*<number-of-prb-per-scs>\s*[\s\S]*?</number-of-prb-per-scs>\s*", re.IGNORECASE)
    if not prb:
        return pat.sub("", block) if pat.search(block) else block
    inner = ""
    if scs:
        inner += f"\n            <scs>{scs}</scs>"
    inner += f"\n            <number-of-prb>{prb}</number-of-prb>"
    repl = f"<number-of-prb-per-scs>{inner}\n          </number-of-prb-per-scs>"
    if pat.search(block):
        return pat.sub(repl, block)
    m_open = re.search(r"^(\s*<\w+[^>]*>\s*)", block)
    if not m_open:
        return block
    insert = f"{m_open.group(1)}{repl}\n"
    return insert + block[len(m_open.group(1)) :]


def _set_compression_leaf(block: str, tag: str, value: str) -> str:
    """Set/insert one leaf inside the first ``compression`` block."""
    if not value:
        return block
    comp_pat = re.compile(r"<compression>\s*[\s\S]*?</compression>", re.IGNORECASE)
    m = comp_pat.search(block)
    if not m:
        return block

    comp = m.group(0)
    leaf_pat = re.compile(
        rf"<{re.escape(tag)}\s*/>|<{re.escape(tag)}>\s*[^<]*\s*</{re.escape(tag)}>",
        re.IGNORECASE,
    )
    if leaf_pat.search(comp):
        comp2 = leaf_pat.sub(f"<{tag}>{value}</{tag}>", comp, count=1)
    else:
        close_i = comp.lower().rfind("</compression>")
        if close_i < 0:
            return block
        comp2 = comp[:close_i] + f"            <{tag}>{value}</{tag}>\n          " + comp[close_i:]
    return block[: m.start()] + comp2 + block[m.end() :]


def _normalize_compression_block(block: str) -> str:
    """Emit compression children in vendor order: type, iq-bitwidth, method, exponent."""
    comp_pat = re.compile(r"<compression>\s*([\s\S]*?)</compression>", re.IGNORECASE)
    m = comp_pat.search(block)
    if not m:
        return block

    inner = m.group(1)

    def _grab(tag: str) -> str:
        mm = re.search(rf"<{re.escape(tag)}\s*/>|<{re.escape(tag)}>\s*([^<]*?)\s*</{re.escape(tag)}>", inner, re.IGNORECASE)
        if not mm:
            return ""
        return (mm.group(1) or "").strip() if mm.lastindex else ""

    leaves: list[tuple[str, str]] = []
    for tag in ("compression-type", "iq-bitwidth", "compression-method", "exponent"):
        val = _grab(tag)
        if val:
            leaves.append((tag, val))
    if not leaves:
        return block
    body = "".join(f"\n            <{tag}>{val}</{tag}>" for tag, val in leaves)
    new_comp = f"<compression>{body}\n          </compression>"
    return block[: m.start()] + new_comp + block[m.end() :]


def _remove_compression_leaf(block: str, tag: str) -> str:
    """Remove one leaf from the first ``compression`` block."""
    comp_pat = re.compile(r"<compression>\s*[\s\S]*?</compression>", re.IGNORECASE)
    m = comp_pat.search(block)
    if not m:
        return block
    comp = m.group(0)
    leaf_pat = re.compile(
        rf"\s*<{re.escape(tag)}\s*/>\s*|\s*<{re.escape(tag)}>\s*[^<]*\s*</{re.escape(tag)}>\s*\n?",
        re.IGNORECASE,
    )
    comp2 = leaf_pat.sub("\n", comp, count=1)
    return block[: m.start()] + comp2 + block[m.end() :]


def _apply_acorn_compression(block: str, iq: str, compression: str) -> str:
    """
    Apply ACORN compression convention from Excel:
    - blank / no comp / static: <compression-type>STATIC</compression-type> + iq-bitwidth only
    - exponent / exponent:4: STATIC + <exponent>4</exponent>
    - bfp / block floating point: STATIC + BLOCK_FLOATING_POINT
    """
    iq_val = (iq or "").strip()
    comp_val = (compression or "").strip()
    norm = comp_val.lower().replace("_", " ").replace("-", " ").strip()

    plain_static = norm in {"", "no comp", "none", "no compression", "static"}
    if plain_static:
        block = _set_compression_leaf(block, "compression-type", "STATIC")
        block = _remove_compression_leaf(block, "compression-method")
        block = _remove_compression_leaf(block, "exponent")
    else:
        m_exp = re.match(r"^exponent(?:\s*[:=]\s*(\d+))?$", norm)
        if m_exp:
            exp = m_exp.group(1) or "4"
            block = _set_compression_leaf(block, "compression-type", "STATIC")
            block = _set_compression_leaf(block, "exponent", exp)
            block = _remove_compression_leaf(block, "compression-method")
        elif norm in {"bfp", "block floating point"}:
            block = _set_compression_leaf(block, "compression-type", "STATIC")
            block = _set_compression_leaf(block, "compression-method", "BLOCK_FLOATING_POINT")
            block = _remove_compression_leaf(block, "exponent")

    if iq_val:
        block = _set_compression_leaf(block, "iq-bitwidth", iq_val)
    return _normalize_compression_block(block)


def _upsert_nested_ul_fft_offsets(block: str, scs: str, offset: str) -> str:
    if not scs and not offset:
        return block
    inner = ""
    if scs:
        inner += f"\n            <scs>{scs}</scs>"
    if offset:
        inner += f"\n            <ul-fft-sampling-offset>{offset}</ul-fft-sampling-offset>"
    pat = re.compile(r"<ul-fft-sampling-offsets>\s*[\s\S]*?</ul-fft-sampling-offsets>", re.IGNORECASE)
    repl = f"<ul-fft-sampling-offsets>{inner}\n          </ul-fft-sampling-offsets>"
    if pat.search(block):
        return pat.sub(repl, block)
    m_open = re.search(r"^(\s*<\w+[^>]*>\s*)", block)
    if not m_open:
        return block
    insert = f"{m_open.group(1)}{repl}\n"
    return insert + block[len(m_open.group(1)) :]


def _merge_acorn_carrier_block(
    block: str,
    row: list[str],
    *,
    kind_u: str,
    c_center: int | None,
    c_bw: int | None,
    c_type: int | None,
    c_gain: int | None,
    c_gain_corr: int | None,
    c_n_ta: int | None,
    c_t_au: int | None,
    c_t_da: int | None,
    c_cell,
) -> str:
    """Inject mandatory carrier leaves from Control-Sheet detail row (ACORN)."""
    center = c_cell(row, c_center)
    if center:
        hz = mhz_to_hz_string(center)
        block = _upsert_first_tag(block, "center-of-channel-bandwidth", hz or center)
    bw = c_cell(row, c_bw)
    if bw:
        block = _upsert_first_tag(block, "channel-bandwidth", bandwidth_to_hz_string(bw) or bw)
    typ = c_cell(row, c_type)
    if typ:
        block = _upsert_first_tag(block, "type", typ)
    ku = kind_u.upper()
    if ku == "PDSCH":
        gain = c_cell(row, c_gain)
        if gain:
            block = _upsert_first_tag(block, "gain", gain)
        t_da = c_cell(row, c_t_da)
        if t_da:
            block = _upsert_first_tag(block, "t-da-offset", t_da)
    elif ku == "PUSCH":
        n_ta = c_cell(row, c_n_ta)
        if n_ta:
            block = _upsert_first_tag(block, "n-ta-offset", n_ta)
        t_au = c_cell(row, c_t_au)
        if t_au:
            block = _upsert_first_tag(block, "t-au-offset", t_au)
        gain_c = c_cell(row, c_gain_corr)
        if gain_c:
            block = _upsert_first_tag(block, "gain-correction", gain_c)
        else:
            gain = c_cell(row, c_gain)
            if gain:
                block = _upsert_first_tag(block, "gain-correction", gain)
    return block


def omit_prach_rx_endpoints_present_in_pusch(prach_xml: str, pusch_xml: str) -> tuple[str, list[str]]:
    """
    PRACH rows 0,2,4,6 reuse the same ``low-level-rx-endpoints`` names as PUSCH.
    Re-sending those blocks in a second edit-config (with different eAxC-ID) makes the RU
    reject ``number-of-prb-per-scs``; keep PRACH links only for those names.
    """
    warns: list[str] = []
    pusch_names: set[str] = set()
    for blk in _extract_closed_blocks(pusch_xml, "low-level-rx-endpoints"):
        m = re.search(r"<name>\s*([^<]+?)\s*</name>", blk, re.IGNORECASE)
        if m:
            pusch_names.add(m.group(1).strip())
    if not pusch_names:
        return prach_xml, warns

    def _sub(m: re.Match[str]) -> str:
        blk = m.group(0)
        nm = re.search(r"<name>\s*([^<]+?)\s*</name>", blk, re.IGNORECASE)
        if nm and nm.group(1).strip() in pusch_names:
            ep = nm.group(1).strip()
            warns.append(f"PRACH: omit <low-level-rx-endpoints> '{ep}' (already configured in PUSCH)")
            return ""
        return blk

    out = _closed_block_pattern("low-level-rx-endpoints").sub(_sub, prach_xml)
    return out, warns


def derive_uplane_link_name(endpoint_name: str, kind: str) -> str:
    """
    ACORN workbooks leave link ``<name>`` empty; derive from endpoint name.

    PDSCH: ``llte_*`` → ``lltl_*`` | PUSCH: same as endpoint ``llre_*`` |
    PRACH: ``llre_*`` → ``llrl_*`` (``…_0_0`` suffix → ``…_1_0``).
    """
    ep = (endpoint_name or "").strip()
    if not ep:
        return ""
    ku = kind.upper()
    if ku == "PDSCH":
        if ep.startswith("llte_"):
            return "lltl_" + ep[5:]
        return ep
    if ku == "PRACH":
        if ep.startswith("llre_"):
            link = "llrl_" + ep[5:]
            if link.endswith("_0_0"):
                return link[:-3] + "1_0"
            return link
        return ep
    return ep


def _upsert_eaxc_id(block: str, eaxc: str) -> str:
    if not eaxc:
        return block
    pat = re.compile(r"<e-axcid>\s*([\s\S]*?)</e-axcid>", re.IGNORECASE)
    m = pat.search(block)
    if not m:
        return block
    inner = m.group(1)
    if re.search(r"<eaxc-id>", inner, re.IGNORECASE):
        inner2 = re.sub(
            r"<eaxc-id>\s*[^<]*\s*</eaxc-id>",
            f"<eaxc-id>{eaxc}</eaxc-id>",
            inner,
            count=1,
            flags=re.IGNORECASE,
        )
    else:
        inner2 = inner.rstrip() + f"\n            <eaxc-id>{eaxc}</eaxc-id>\n          "
    return block[: m.start()] + f"<e-axcid>{inner2}</e-axcid>" + block[m.end() :]


def apply_acorn_control_details_to_rpc(
    xml: str,
    kind: str,
    headers: list[str],
    rows: list[list[str]],
) -> tuple[str, list[str]]:
    """
    ACORN workbooks keep U-Plane values on Control-Sheet and skeletal ``*-endpoints`` blocks.
    Merge detail rows into the Nth endpoint / carrier / link block.
    """
    warns: list[str] = []
    text = (xml or "").strip()
    if not text or not headers or not rows:
        return xml, warns

    def _col(*names: str) -> int | None:
        want = {normalize_header(n) for n in names}
        for i, h in enumerate(headers):
            nh = normalize_header(h)
            if nh in want:
                return i
            for w in want:
                if w and len(w) >= 4 and (w in nh or nh in w):
                    return i
        return None

    def _col_exact(*names: str) -> int | None:
        want = {normalize_header(n) for n in names}
        for i, h in enumerate(headers):
            if normalize_header(h) in want:
                return i
        return None

    def _cell(row: list[str], idx: int | None) -> str:
        if idx is None or idx >= len(row):
            return ""
        return (row[idx] or "").strip()

    kind_u = kind.upper()
    if kind_u == "PDSCH":
        ep_blk, car_blk, link_blk = "low-level-tx-endpoints", "tx-array-carriers", "low-level-tx-links"
    else:
        ep_blk, car_blk, link_blk = "low-level-rx-endpoints", "rx-array-carriers", "low-level-rx-links"

    ep_blocks = _extract_closed_blocks(text, ep_blk)
    car_blocks = _extract_closed_blocks(text, car_blk)
    link_blocks = _extract_closed_blocks(text, link_blk)
    if not ep_blocks:
        warns.append(f"{kind}: no <{ep_blk}> blocks for Control-Sheet merge")
        return xml, warns
    if len(rows) > len(ep_blocks):
        warns.append(
            f"{kind}: Control-Sheet has {len(rows)} rows but XML has {len(ep_blocks)} <{ep_blk}> block(s)"
        )

    if kind_u == "PDSCH":
        c_ep = _col(
            "low-level-tx-endpoint",
            "low level tx endpoint",
            "low-level-rx-endpoint",
            "end-point",
            "end point",
        )
        c_car = _col("tx-array-carrier", "rx-array-carrier", "array-carrier")
    else:
        c_ep = _col(
            "low-level-rx-endpoint",
            "low level rx endpoint",
            "end-point",
            "end point",
        )
        c_car = _col("rx-array-carrier", "tx-array-carrier", "array-carrier")
    c_scs = _col("scs")
    c_prb = _col("number-of-prb", "number of prb")
    c_iq = _col("iq-bitwidth", "iq bitwidth")
    c_comp = _col("compression", "compression-method", "compression method")
    c_ul_fft = _col("ul-fft-sampling-offset", "ul fft sampling offset", "ll fft sampling offset")
    c_eaxc = _col("eaxc-id", "eaxc id")
    c_link = _col_exact("link", "link name")
    c_pe = _col("processing-element", "processing element")
    c_center = _col("center-freq", "center freq", "center_freq")
    c_bw = _col("channel-bw", "channel bw", "channel-bandwidth", "channel bandwidth")
    c_type = _col("type")
    c_gain = _col("gain")
    c_gain_corr = _col("gain-correction", "gain correction")
    c_n_ta = _col("n-ta-offset", "n ta offset")
    c_t_au = _col("t-au-offset", "t au offset", "t-au")
    c_t_da = _col("t-da-offset", "t da offset", "t-du-offset", "t-du")

    new_ep: list[str] = []
    for i, row in enumerate(rows):
        if i >= len(ep_blocks):
            break
        seg = ep_blocks[i]
        ep_name = _cell(row, c_ep)
        if ep_name:
            seg = _upsert_first_tag(seg, "name", ep_name)
        seg = _apply_acorn_compression(seg, _cell(row, c_iq), _cell(row, c_comp))
        scs_val = _cell(row, c_scs)
        prb_val = _cell(row, c_prb)
        if prb_val and not scs_val:
            warns.append(f"{kind} row {i + 1}: number-of-prb set but scs is empty")
        seg = _upsert_nested_prb_scs(seg, scs_val, prb_val)
        if kind_u in ("PUSCH", "PRACH"):
            ul_val = _cell(row, c_ul_fft)
            if not scs_val and not ul_val:
                warns.append(
                    f"{kind} row {i + 1}: missing scs/ul-fft-sampling-offset for <ul-fft-sampling-offsets>"
                )
            seg = _upsert_nested_ul_fft_offsets(seg, scs_val, ul_val)
        seg = _upsert_eaxc_id(seg, _cell(row, c_eaxc))
        new_ep.append(seg)
    for j in range(len(new_ep), len(ep_blocks)):
        new_ep.append(ep_blocks[j])

    out = _replace_closed_blocks(text, ep_blk, new_ep)

    new_car: list[str] = []
    for i, row in enumerate(rows):
        if i >= len(car_blocks):
            break
        seg = car_blocks[i]
        car_name = _cell(row, c_car)
        if car_name:
            seg = _upsert_first_tag(seg, "name", car_name)
        center_val = _cell(row, c_center)
        if not center_val:
            warns.append(f"{kind} row {i + 1}: missing Center-FREQ for <center-of-channel-bandwidth>")
        if kind_u == "PUSCH" and not _cell(row, c_n_ta):
            warns.append(f"{kind} row {i + 1}: missing n-ta-offset for <rx-array-carriers>")
        seg = _merge_acorn_carrier_block(
            seg,
            row,
            kind_u=kind_u,
            c_center=c_center,
            c_bw=c_bw,
            c_type=c_type,
            c_gain=c_gain,
            c_gain_corr=c_gain_corr,
            c_n_ta=c_n_ta,
            c_t_au=c_t_au,
            c_t_da=c_t_da,
            c_cell=_cell,
        )
        new_car.append(seg)
    for j in range(len(new_car), len(car_blocks)):
        new_car.append(car_blocks[j])
    out = _replace_closed_blocks(out, car_blk, new_car)

    new_link: list[str] = []
    for i, row in enumerate(rows):
        if i >= len(link_blocks):
            break
        seg = link_blocks[i]
        epn = _cell(row, c_ep)
        lname = _cell(row, c_link)
        if not lname and epn:
            lname = derive_uplane_link_name(epn, kind_u)
        if lname:
            seg = _upsert_first_tag(seg, "name", lname)
        pe = _cell(row, c_pe)
        if pe:
            seg = _upsert_first_tag(seg, "processing-element", pe)
        if kind_u == "PDSCH":
            car = _cell(row, c_car)
            if car:
                seg = _upsert_first_tag(seg, "tx-array-carrier", car)
            if epn:
                seg = _upsert_first_tag(seg, "low-level-tx-endpoint", epn)
        else:
            car = _cell(row, c_car)
            if car:
                seg = _upsert_first_tag(seg, "rx-array-carrier", car)
            if epn:
                seg = _upsert_first_tag(seg, "low-level-rx-endpoint", epn)
        new_link.append(seg)
    for j in range(len(new_link), len(link_blocks)):
        new_link.append(link_blocks[j])
    out = _replace_closed_blocks(out, link_blk, new_link)
    return out, warns


def _is_control_uplane_data_row(cells: list[str]) -> bool:
    """True when the first grid column looks like an O-RAN endpoint name (not a note row)."""
    if not cells:
        return False
    name = (cells[0] or "").strip()
    if not name:
        return False
    upper = name.upper()
    if upper in ("NAME", "END-POINT", "END POINT"):
        return False
    return bool(_UPLANE_ENDPOINT_NAME_RE.match(name))


def parse_control_uplane_detail_tables(
    ws,
) -> tuple[dict[str, tuple[list[str], list[list[str]]]], list[str]]:
    """Read PDSCH/PUSCH/PRACH detail grids from Control-Sheet (ACORN layout)."""
    out: dict[str, tuple[list[str], list[list[str]]]] = {}
    parse_warnings: list[str] = []
    rows = list(ws.iter_rows(values_only=True))
    section_map = {
        "3. pdsch": "PDSCH",
        "4. pusch": "PUSCH",
        "5. prach": "PRACH",
    }

    def _section_start(key: str) -> int | None:
        sk = normalize_header(key)
        for i, row in enumerate(rows):
            b = row[1] if row and len(row) > 1 else None
            if b is not None and normalize_header(str(b)) == sk:
                return i
        return None

    for sec_key, kind in section_map.items():
        si = _section_start(sec_key)
        if si is None:
            continue
        header_i = None
        for j in range(si + 1, min(si + 12, len(rows))):
            row = rows[j] or ()
            c = row[2] if len(row) > 2 else None
            d = row[3] if len(row) > 3 else None
            if normalize_header(str(c or "")) == "name" and normalize_header(str(d or "")) == "scs":
                header_i = j
                break
        if header_i is None:
            continue
        hr = rows[header_i]
        j0 = 2
        j1 = max(2, len(hr) - 1)
        while j1 > j0 and (hr[j1] is None or str(hr[j1]).strip() == ""):
            j1 -= 1
        headers = ["" if hr[c] is None else str(hr[c]).strip() for c in range(j0, j1 + 1)]
        data_rows: list[list[str]] = []
        skipped_rows = 0
        for j in range(header_i + 1, len(rows)):
            row = rows[j] or ()
            b = row[1] if len(row) > 1 else None
            if b is not None and str(b).strip() and re.match(r"^\d+\.", str(b).strip()):
                break
            cells = ["" if (c >= len(row) or row[c] is None) else str(row[c]).strip() for c in range(j0, j1 + 1)]
            if not any(cells):
                break
            if cells[0].upper() in ("NAME", "END-POINT", "END POINT"):
                continue
            if not _is_control_uplane_data_row(cells):
                skipped_rows += 1
                continue
            data_rows.append(cells)
        if headers and data_rows:
            out[kind] = (headers, data_rows)
            if skipped_rows:
                parse_warnings.append(
                    f"[Control-Sheet] {kind}: skipped {skipped_rows} note/template row(s); "
                    "only llre_/llte_/rac_ names are loaded"
                )
    return out, parse_warnings


def resolve_l2vlan_interface_name(live: dict[str, str]) -> str:
    """
    ietf-interfaces ``<name>`` and PE ``interface-name`` — same value (ACORN: VLAN id, e.g. 1000).
    ``base-interface`` (sys) and ``vlan-id`` (1000) are separate leaves on the interface.
    """
    vlan = str(live.get("cu_vlan") or "").strip()
    if vlan:
        return vlan
    explicit = str(live.get("cu_if_name") or "").strip()
    if explicit:
        return explicit
    return ""


def resolve_pe_transport_interface_name(live: dict[str, str]) -> str:
    return resolve_l2vlan_interface_name(live)


def resolve_to_du_l2vlan_interface_name(live: dict[str, str]) -> str:
    """Alias for CUplane / PE naming (ACORN workbook)."""
    return resolve_l2vlan_interface_name(live)


def ensure_processing_element_fields(xml: str, live: dict[str, str]) -> str:
    """Inject mandatory ru-elements / eth-flow leaves from Control-Sheet + settings."""
    text = xml or ""
    pe = str(live.get("pe_name") or "").strip()
    pe_ifname = resolve_pe_transport_interface_name(live)
    vlan = str(live.get("cu_vlan") or "").strip()
    ru_mac = str(live.get("cu_mac") or live.get("ru_mac_pe") or "").strip()
    odu_mac = str(live.get("odu_mac") or "").strip()
    if not any((pe, pe_ifname, vlan, ru_mac, odu_mac)):
        return text

    def _has_tag(tag: str, block: str) -> bool:
        return bool(re.search(rf"<{re.escape(tag)}\b", block, flags=re.IGNORECASE))

    def _patch_ru(m: re.Match[str]) -> str:
        open_part, body, close_part = m.group(1), m.group(2), m.group(3)
        parts: list[str] = []
        if pe and not _has_tag("name", body):
            parts.append(f"<name>{pe}</name>")
        if parts:
            body = "\n          " + "\n          ".join(parts) + body

        def _patch_flow(fm: re.Match[str]) -> str:
            fo, fb, fc = fm.group(1), fm.group(2), fm.group(3)
            fp: list[str] = []
            if pe_ifname and not _has_tag("interface-name", fb):
                fp.append(f"<interface-name>{pe_ifname}</interface-name>")
            fb2 = fb
            if fp:
                fb2 = "\n            " + "\n            ".join(fp) + fb2

            def _patch_eth(em: re.Match[str]) -> str:
                eo, eb, ec = em.group(1), em.group(2), em.group(3)
                # eth-flow vlan-id: o-ran-processing-element leaf (no o-ran:interfaces xmlns).
                eb = re.sub(
                    r"\s*<vlan-id\b[^>]*>[^<]*</vlan-id>\s*",
                    "\n",
                    eb,
                    flags=re.IGNORECASE,
                )
                ep: list[str] = []
                if ru_mac and not _has_tag("ru-mac-address", eb):
                    ep.append(f"<ru-mac-address>{ru_mac}</ru-mac-address>")
                if vlan and not _has_tag("vlan-id", eb):
                    ep.append(f"<vlan-id>{vlan}</vlan-id>")
                if odu_mac and not _has_tag("o-du-mac-address", eb):
                    ep.append(f"<o-du-mac-address>{odu_mac}</o-du-mac-address>")
                if ep:
                    eb = "\n              " + "\n              ".join(ep) + eb
                return eo + eb + ec

            fb2 = re.sub(
                r"(<eth-flow>\s*)([\s\S]*?)(\s*</eth-flow>)",
                _patch_eth,
                fb2,
                count=1,
                flags=re.IGNORECASE,
            )
            return fo + fb2 + fc

        body = re.sub(
            r"(<transport-flow>\s*)([\s\S]*?)(\s*</transport-flow>)",
            _patch_flow,
            body,
            count=1,
            flags=re.IGNORECASE,
        )
        return open_part + body + close_part

    return re.sub(
        r"(<ru-elements>\s*)([\s\S]*?)(\s*</ru-elements>)",
        _patch_ru,
        text,
        count=1,
        flags=re.IGNORECASE,
    )


def ensure_cuplane_interface_fields(xml: str, live: dict[str, str]) -> str:
    """Workbook CUplane RPC often omits mandatory ietf-interfaces leaves; inject from settings."""
    text = xml or ""
    vlan = str(live.get("cu_vlan") or "").strip()
    base = str(live.get("cu_base_if") or "").strip()
    ifname = resolve_l2vlan_interface_name(live)
    mac = str(live.get("cu_mac") or "").strip()
    if not ifname and not base and not vlan and not mac:
        return text

    def _has_tag(tag: str, block: str) -> bool:
        return bool(re.search(rf"<{re.escape(tag)}\b", block, flags=re.IGNORECASE))

    def _patch_iface(m: re.Match[str]) -> str:
        open_part, body, close_part = m.group(1), m.group(2), m.group(3)
        parts: list[str] = []
        if ifname and not _has_tag("name", body):
            parts.append(f"<name>{ifname}</name>")
        if not _has_tag("type", body) and "l2vlan" not in body.lower():
            parts.append(
                '<type xmlns:ianaift="urn:ietf:params:xml:ns:yang:iana-if-type">ianaift:l2vlan</type>'
            )
        if base and not _has_tag("base-interface", body):
            parts.append(
                f'<base-interface xmlns="urn:o-ran:interfaces:1.0">{base}</base-interface>'
            )
        if vlan and not _has_tag("vlan-id", body):
            parts.append(f'<vlan-id xmlns="urn:o-ran:interfaces:1.0">{vlan}</vlan-id>')
        if mac and not _has_tag("mac-address", body):
            parts.append(f'<mac-address xmlns="urn:o-ran:interfaces:1.0">{mac}</mac-address>')
        if parts:
            body = "\n          " + "\n          ".join(parts) + body
        return open_part + body + close_part

    return re.sub(
        r"(<interface>\s*)([\s\S]*?)(\s*</interface>)",
        _patch_iface,
        text,
        count=1,
        flags=re.IGNORECASE,
    )


def extract_config_payload_from_edit_config(xml: str) -> tuple[str, str, str]:
    """
    Parse stripped ``<edit-config>`` XML for netopeer2-cli:
    ``edit-config --target … --defop … --config=FILE`` (FILE = config children only).
    """
    text = strip_rpc_wrapper_for_netconf_cli(xml)
    target = "running"
    defop = "merge"
    m_target = re.search(
        r"<target>\s*<([a-zA-Z0-9_\-:]+)\s*/>\s*</target>", text, flags=re.IGNORECASE | re.DOTALL
    )
    if m_target:
        target = m_target.group(1).split(":")[-1]
    m_defop = re.search(
        r"<default-operation>\s*([^<]+)\s*</default-operation>", text, flags=re.IGNORECASE | re.DOTALL
    )
    if m_defop:
        defop = m_defop.group(1).strip()
    m_cfg = re.search(r"(<config\b[\s\S]*?</config>)", text, flags=re.IGNORECASE)
    if m_cfg:
        cfg = m_cfg.group(1)
        inner = re.match(
            r"^\s*<config\b[^>]*>([\s\S]*)</config>\s*$", cfg, flags=re.IGNORECASE | re.DOTALL
        )
        if inner:
            return target, defop, inner.group(1).strip()
        return target, defop, cfg
    return target, defop, text


def finalize_mplane_rpc_payload(
    sheet: str,
    xml: str,
    baselines: dict[str, str],
    live: dict[str, str],
    *,
    for_netconf_cli: bool = True,
) -> str:
    """Apply baselines, CUplane fixes, and optional netopeer2-cli envelope stripping."""
    body = apply_global_baselines((xml or "").strip(), baselines, live)
    if sheet == "CUplane-interface":
        body = ensure_cuplane_interface_fields(body, live)
    if sheet == "Processing-element":
        body = ensure_processing_element_fields(body, live)
    if for_netconf_cli:
        body = strip_rpc_wrapper_for_netconf_cli(body)
    return body


def apply_global_baselines(xml: str, baselines: dict[str, str], live: dict[str, str]) -> str:
    def pick(*keys: str) -> str:
        for hk in keys:
            v = live.get(hk)
            if v is not None and str(v).strip() != "":
                return str(v).strip()
        return ""

    out = xml or ""

    def repl_tag(tag: str, new_val: str, count: int = 0) -> None:
        nonlocal out
        if not new_val:
            return
        pat = re.compile(rf"(<{re.escape(tag)}[^>]*>)([^<]*)(</{re.escape(tag)}>)")
        if count > 0:
            out = pat.sub(lambda m: f"{m.group(1)}{new_val}{m.group(3)}", out, count=count)
        else:
            out = pat.sub(lambda m: f"{m.group(1)}{new_val}{m.group(3)}", out)

    # l2vlan: <name>=1000, <vlan-id>=1000, <base-interface>=sys (CUplane + PE interface-name 동일)
    merged_live = {k: (baselines.get(k) or "") for k in ("cu_if_name", "cu_base_if", "cu_vlan")}
    for hk in ("cu_if_name", "cu_base_if", "cu_vlan"):
        v = pick(hk)
        if v:
            merged_live[hk] = v
    ifname = resolve_l2vlan_interface_name(merged_live)
    if not ifname:
        ifname = pick("cu_vlan", "cu_if_name")
    if ifname:
        out = re.sub(
            r"(<interface>\s*<name>)([^<]*)(</name>)",
            lambda m: f"{m.group(1)}{ifname}{m.group(3)}",
            out,
            count=1,
            flags=re.DOTALL,
        )

    base_if = pick("cu_base_if")
    if base_if:
        repl_tag("base-interface", base_if)
    vlan = pick("cu_vlan")
    if vlan and re.search(r"<processing-elements\b", out, flags=re.IGNORECASE):
        repl_tag("vlan-id", vlan)
        pe_if = resolve_pe_transport_interface_name(merged_live)
        if pe_if:
            repl_tag("interface-name", pe_if)
    elif vlan:
        repl_tag("vlan-id", vlan)
    cu_mac = pick("cu_mac")
    if cu_mac:
        repl_tag("mac-address", cu_mac)
        repl_tag("ru-mac-address", cu_mac)
    pe_name = pick("pe_name")
    if pe_name:
        repl_tag("processing-element", pe_name)
        out = re.sub(
            r"(<ru-elements>\s*<name>)([^<]*)(</name>)",
            lambda m: f"{m.group(1)}{pe_name}{m.group(3)}",
            out,
            count=1,
            flags=re.DOTALL,
        )
    odu_mac = pick("odu_mac")
    if odu_mac:
        repl_tag("o-du-mac-address", odu_mac)

    return out


def write_control_sheet_simple_fields(ws, live: dict[str, str]) -> None:
    """Persist CU-plane / Processing-element parameters to Control-Sheet (column D)."""
    live = dict(live)
    ifname = resolve_l2vlan_interface_name(live)
    if ifname:
        live["cu_if_name"] = ifname
    label_to_key = {
        "base interface": "cu_base_if",
        "base-interface": "cu_base_if",
        "vlan id": "cu_vlan",
        "oru mac address": "cu_mac",
        "odu mac address": "odu_mac",
    }
    section = ""

    def _safe_set(row: int, col: int, value: Any) -> None:
        try:
            cell = ws.cell(row=row, column=col)
            if cell.__class__.__name__ == "MergedCell":
                return
            cell.value = value
        except Exception:
            return

    for rr in range(1, (ws.max_row or 0) + 1):
        b = ws.cell(row=rr, column=2).value
        c = ws.cell(row=rr, column=3).value
        if b is not None and str(b).strip():
            sb = str(b).strip()
            if re.match(r"^\d+[\-.]", sb):
                section = normalize_header(sb)
        lk = normalize_header(str(c).strip()) if c is not None and str(c).strip() else ""
        if not lk:
            continue
        key: str | None
        if lk == "name":
            if "cuplane" in section or section.startswith("1"):
                key = "cu_if_name"
            elif "processing" in section or section.startswith("2"):
                key = "pe_name"
            else:
                continue
        else:
            key = label_to_key.get(lk)
        if not key:
            continue
        val = live.get(key)
        if val is None or str(val).strip() == "":
            continue
        _safe_set(rr, 4, str(val).strip())


def load_workbook_payloads(
    path: str | Path,
) -> tuple[
    dict[str, str],
    dict[str, str],
    dict[str, Any],
    list[dict[str, Any]],
    dict[str, tuple[list[str], list[list[str]]]],
    list[str],
]:
    """
    Returns:
      - rpc_text_by_sheet
      - baselines dict
      - merged_control_fields (from spreadsheet + inferred)
      - cc_rows (from Control-Sheet)
      - tables: { "PDSCH"/"PUSCH"/"PRACH": (headers, rows) } extracted from the Excel sheets
      - warnings
    """
    p = Path(path)
    warnings: list[str] = []
    if not p.exists():
        raise FileNotFoundError(str(p))

    wb = openpyxl.load_workbook(p, data_only=True)
    try:
        control_tables: dict[str, tuple[list[str], list[list[str]]]] = {}
        if "Control-Sheet" not in wb.sheetnames:
            warnings.append("[Workbook] No Control-Sheet found.")
            simple: dict[str, Any] = {}
            cc_rows: list[dict[str, Any]] = []
        else:
            ws_c = wb["Control-Sheet"]
            simple, cc_rows, cw = parse_control_sheet(ws_c)
            warnings.extend(cw)
            control_tables, table_parse_warns = parse_control_uplane_detail_tables(ws_c)
            warnings.extend(table_parse_warns)
            detail_cc_on = parse_control_detail_cc_on_flags(ws_c)
            if detail_cc_on:
                simple["_detail_cc_on_flags"] = detail_cc_on

        tables: dict[str, tuple[list[str], list[list[str]]]] = {}
        for sheet in ("PDSCH", "PUSCH", "PRACH"):
            ctrl = control_tables.get(sheet)
            if ctrl and ctrl[0] and ctrl[1]:
                tables[sheet] = ctrl
                headers, rows = ctrl
                prb_idx = None
                for i, h in enumerate(headers):
                    if normalize_header(h) in ("number of prb", "number-of-prb"):
                        prb_idx = i
                        break
                if prb_idx is not None:
                    empty_prb = sum(
                        1 for r in rows if prb_idx < len(r) and not (r[prb_idx] or "").strip()
                    )
                    if empty_prb:
                        warnings.append(
                            f"[Control-Sheet] {sheet}: {empty_prb} row(s) with empty number-of-prb "
                            "(RPC will omit <number-of-prb-per-scs>)"
                        )
                continue
            if sheet not in wb.sheetnames:
                continue
            tmp_xml = read_sheet_rpc_text(wb[sheet]).strip()
            try:
                headers, rows = derive_full_table_from_rpc(tmp_xml, sheet)
                if headers and rows:
                    tables[sheet] = (headers, rows)
                    warnings.append(
                        f"[Workbook] {sheet}: no Control-Sheet grid — using {sheet} RPC sheet values"
                    )
                else:
                    warnings.append(f"[Workbook] Could not derive table from sheet RPC: {sheet}")
            except Exception as exc:
                warnings.append(f"[Workbook] Failed to derive table from {sheet}: {exc}")

        rpc: dict[str, str] = {}
        for sheet in SEND_ORDER:
            if sheet not in wb.sheetnames:
                warnings.append(f"[Workbook] Missing sheet: {sheet}")
                continue
            if sheet in SKIP_SHEETS:
                continue
            rpc[sheet] = read_sheet_rpc_text(wb[sheet]).strip()

        cu = rpc.get("CUplane-interface", "")
        pe = rpc.get("Processing-element", "")
        baselines = parse_baselines(cu, pe)

        merged = dict(baselines)
        merged.update({k: v for k, v in simple.items() if v not in (None, "")})

        return rpc, baselines, merged, cc_rows, tables, warnings
    finally:
        wb.close()

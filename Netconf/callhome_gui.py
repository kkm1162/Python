import copy
import io
import os
import re
import shlex
import shutil
import signal
import sys
import subprocess
import threading
import tkinter as tk
import json
import base64
import hashlib
import tempfile
import time
import xml.dom.minidom
import xml.etree.ElementTree as ET
import importlib
from datetime import datetime
import tkinter.font as tkfont
from pathlib import Path, PurePosixPath
from tkinter import filedialog, messagebox, simpledialog, ttk
from typing import Any

import conformance_manifest as _conf_manifest
from conformance_mixin import ConformanceMixin


# CC table column keys (spreadsheet/grid); must match mplane_control.CC_COL_MAP round-trip.
_MPLANE_CC_TREE_COLS: tuple[str, ...] = (
    "enabled",
    "cc",
    "ant",
    "scs",
    "prb",
    "iq",
    "compression",
    "typ",
    "dl_mhz",
    "bw",
    "ul_mhz",
    "t_au",
    "n_ta",
    "t_du",
)

_MPLANE_CC_GRID_ROWS = 16
_MPLANE_CC_ENTRY_KEYS: tuple[str, ...] = _MPLANE_CC_TREE_COLS[1:]  # all carrier columns except "enabled"

_MPLANE_PHY_ROWS = 16

_NETCONF_LIB_EXTENSIONS = frozenset({".xml", ".rpc", ".txt"})


def _app_bundle_root() -> Path:
    """Directory containing the running script, or the folder with the .exe when frozen (PyInstaller)."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


_CONFIG_APP_NAME = "O-RAN-Netconf"
_CONFIG_FILENAME = "callhome_gui_config.json"


def _user_config_dir() -> Path:
    """Stable per-user config directory (independent of cwd / how the app was launched)."""
    base = os.environ.get("LOCALAPPDATA") or os.environ.get("APPDATA")
    if base:
        return Path(base) / _CONFIG_APP_NAME
    return Path.home() / f".{_CONFIG_APP_NAME.lower()}"


def _legacy_config_candidates() -> list[Path]:
    roots: list[Path] = [_app_bundle_root()]
    try:
        roots.append(Path.cwd().resolve())
    except OSError:
        pass
    out: list[Path] = []
    seen: set[str] = set()
    for root in roots:
        p = (root / _CONFIG_FILENAME).resolve()
        key = str(p).lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(p)
    return out


def _resolve_config_path() -> Path:
    """
    Use %LOCALAPPDATA%\\O-RAN-Netconf\\callhome_gui_config.json so settings survive
    restarts regardless of script path, cwd, or PyInstaller extract dir.
    Migrates an existing config from beside the .exe / .py on first run.
    """
    target = (_user_config_dir() / _CONFIG_FILENAME).resolve()
    if not target.exists():
        for legacy in _legacy_config_candidates():
            if legacy.exists() and legacy.resolve() != target:
                try:
                    target.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(legacy, target)
                    break
                except OSError:
                    continue
    target.parent.mkdir(parents=True, exist_ok=True)
    return target


# Default template next to the app: <bundle>/mplane/mplane.xlsx (works for exe + script).
_MPLANE_MASTER_XLSX = _app_bundle_root() / "mplane" / "mplane.xlsx"


class CallhomeGUI(tk.Tk, ConformanceMixin):
    def __init__(self) -> None:
        super().__init__()
        self.title("O-RAN Netconf")
        self.geometry("1180x820")

        self.proc: subprocess.Popen | None = None
        self.stop_event = threading.Event()
        self.is_running = False
        self.paramiko_client: Any | None = None
        self.paramiko_channel: Any | None = None
        self.session_established = False
        self.manual_send_ready = False
        self.message_tabs: list[dict[str, Any]] = []
        self.netconf_library_root: Path | None = None
        self.netconf_library_paths: dict[str, str] = {}
        self.netconf_library_selected_path: str = ""
        self.netconf_lib_tree: ttk.Treeview | None = None
        self.netconf_lib_status_var = tk.StringVar(value="XML library: no folder loaded")
        self.netconf_lib_get_btn: ttk.Button | None = None
        self.netconf_lib_set_btn: ttk.Button | None = None
        self.netconf_lib_send_btn: ttk.Button | None = None
        self._netconf_library_path_to_restore: str = ""
        self.remote_cfg_cache: dict[str, str] = {}
        self.log_buffer: list[str] = []
        self.hidden_log_chunks: list[str] = []
        self._log_stream_carry = ""
        self._rpc_exchange_collecting = False
        self._rpc_exchange_buf: list[str] = []
        self.hidden_render_active = False
        self.log_lock = threading.Lock()
        self.max_log_lines = 500
        self.gui_log_max_lines_var = tk.StringVar(value="500")
        self.flush_tick = 0
        self.log_window: tk.Toplevel | None = None
        self.log: tk.Text | None = None
        self.rpc_error_list: tk.Listbox | None = None
        self.rpc_event_window: tk.Toplevel | None = None
        self.rpc_event_list: tk.Listbox | None = None
        self.rpc_error_items: list[dict[str, str]] = []
        self.rpc_error_seen: set[str] = set()
        self.rpc_event_window_geometry = "980x520"
        self.log_window_geometry = "1100x720"
        self.remote_control_fifo = "/var/tmp/netconf_tmp/netconf_control.fifo"
        self.script_path = self._resolve_local_script_path()
        self._config_hydrating = False
        self.config_path = _resolve_config_path()
        self.exec_mode = tk.StringVar(value="remote")
        self.remote_user = tk.StringVar(value="oranuser")
        self.remote_host = tk.StringVar(value="10.0.20.128")
        self.remote_port = tk.StringVar(value="22")
        self.remote_password = tk.StringVar(value="")
        self.remote_script_path = tk.StringVar(value="/var/tmp/miniDU_callhome.sh")
        self.remote_key_path = tk.StringVar(value="")
        # Lab profiles (e.g. 10.0.20.x ↔ 10.0.60.x): named snapshots of network-related settings
        self._profiles: dict[str, Any] = {}
        self.profile_name_var = tk.StringVar(value="lab-20")
        self.profile_combo: ttk.Combobox | None = None
        self.field_notes_widget: tk.Text | None = None  # legacy; Guardrails tab replaced notes
        self.guardrails_check_vars: dict[str, tk.BooleanVar] = {}
        self.guardrails_result_vars: dict[str, tk.StringVar] = {}
        self.guardrails_busy = False
        self._guardrails_cancel = threading.Event()
        self._guardrails_user_items: list[dict[str, str]] = []
        self._guardrails_per_test_settings: dict[str, dict[str, str]] = {}
        self.guardrails_list_tree: ttk.Treeview | None = None
        self._guardrails_settings_item_id: str = "dhcp_v4"
        self.guardrails_run_repeat_var = tk.StringVar(value="1")
        self._guardrails_detail_by_id: dict[str, str] = {}
        self._guardrails_detail_text: tk.Text | None = None
        self.guardrails_detail_win_geometry = "900x520"
        # VLAN Discovery: 시험 후 untag 원복을 미룬 경우 (버튼으로 추후 실행)
        self._guardrails_vlan_restore_pending: dict[str, Any] | None = None
        self.guardrails_vlan_restore_btn: ttk.Button | None = None

        self.fields: dict[str, tk.StringVar] = {}
        self.status_var = tk.StringVar(value="Idle")
        self.log_target_hint_var = tk.StringVar(value="")
        self._start_log_path_hint: str = ""
        self._conformance_active_host_log: str | None = None
        self._conformance_last_host_log: str | None = None
        self.send_mode_var = tk.StringVar(value="raw_rpc")
        self.perf_debug_var = tk.BooleanVar(value=False)
        self.auto_xml_send_var = tk.BooleanVar(value=False)
        self.auto_xml_send_done = False
        self.auto_start_var = tk.BooleanVar(value=True)
        self._user_stop_requested = False
        self._auto_start_retry_job: str | None = None
        self._auto_restart_delay_ms = 5_000
        self._transport_reconnect_pending = False
        self._session_lost_logged = False
        self.perf_stats: dict[str, float] = {}
        self.perf_counts: dict[str, int] = {}
        self.perf_max: dict[str, float] = {}
        self.perf_last: dict[str, float] = {}
        self.perf_last_report = time.time()
        self.perf_text_var = tk.StringVar(value="PERF: off")
        self.trace_seq = 0
        self._all_send_shortcut_pending_until = 0.0
        self._geometry_save_job: str | None = None
        self._log_geometry_save_job: str | None = None
        self._config_save_job: str | None = None
        self._last_find_text = ""
        self.find_replace_window: tk.Toplevel | None = None
        self.find_replace_target: tk.Text | None = None
        self.find_var = tk.StringVar(value="")
        self.replace_var = tk.StringVar(value="")
        self.find_ignore_case_var = tk.BooleanVar(value=True)
        self.find_preview_list: tk.Listbox | None = None
        self.find_matches: list[tuple[str, str, str]] = []
        self.mplane_xlsx_path = tk.StringVar(value=str(_MPLANE_MASTER_XLSX))
        self._mplane_loaded_xlsx_path: str = ""
        self._mplane_rpc_raw: dict[str, str] = {}
        self._mplane_baselines: dict[str, str] = {}
        self._mplane_merged: dict[str, Any] = {}
        self._mplane_cc_rows: list[dict[str, Any]] = []
        self.mplane_fields: dict[str, tk.StringVar] = {}
        self.mplane_reload_btn: ttk.Button | None = None
        self.mplane_apply_btn: ttk.Button | None = None
        self.mplane_save_btn: ttk.Button | None = None
        self.mplane_warnings_var = tk.StringVar(value="")
        self.mplane_cc_cell_vars: list[list[tk.StringVar]] = []
        self.mplane_cc_enabled_vars: list[tk.BooleanVar] = []
        # Physical (per-sheet) carrier grids
        self.mplane_phy_enabled: dict[str, list[tk.BooleanVar]] = {}
        self.mplane_phy_prb: dict[str, list[tk.StringVar]] = {}
        self.mplane_phy_center_mhz: dict[str, list[tk.StringVar]] = {}
        self._recent_log_for_session: str = ""
        self._session_watch_job: str | None = None
        self._session_watch_rounds: int = 0
        self._last_session_activity_mono: float = 0.0
        self._session_lost_at_mono: float = 0.0
        self._session_lost_force_reconnect_sec: float = 180.0
        self._mplane_tables: dict[str, tuple[list[str], list[list[str]]]] = {}
        self._mplane_table_widgets: dict[str, dict[str, Any]] = {}
        self._mplane_table_vars: dict[str, dict[str, Any]] = {}
        self._mplane_selection_sheet: str | None = None
        self._mplane_selection_anchor: tuple[int, int] | None = None
        self._mplane_selection_cells: set[tuple[int, int]] = set()
        self._mplane_drag_selecting: bool = False
        self._mplane_cell_index: dict[str, dict[tk.Entry, tuple[int, int]]] = {}
        self._mplane_undo_stack: dict[str, list[dict[tuple[int, int], str]]] = {}
        self._mplane_redo_stack: dict[str, list[dict[tuple[int, int], str]]] = {}
        self._mplane_sync_lock: bool = False
        self._mplane_last_valid_cell_values: dict[tuple[str, int, int], str] = {}
        self._mplane_prev_selection_sheet: str | None = None
        self._mplane_prev_selection_cells: set[tuple[int, int]] = set()
        self.mplane_find_window: tk.Toplevel | None = None
        self._mplane_find_sheet: str | None = None
        self._mplane_find_hits: list[tuple[int, int, int, int]] = []
        self._mplane_find_idx: int = -1
        self.mplane_cc_on_vars: list[tk.BooleanVar] = []
        self.mplane_debug_log_path = Path(__file__).with_name("mplane_ui_debug.log")
        self.mplane_scroll_canvas: tk.Canvas | None = None
        self.mplane_scroll_wrap: ttk.Frame | None = None
        self.mplane_scroll_inner: ttk.Frame | None = None
        self._mplane_wheel_bound: bool = False
        self.conformance_check_vars: dict[str, tk.BooleanVar] = {}
        self.conformance_reboot_vars: dict[str, tk.BooleanVar] = {}
        self.conformance_scroll_canvas: tk.Canvas | None = None
        self.conformance_list_tree: ttk.Treeview | None = None
        self._conformance_run_active_targets: set[str] | None = None
        self._conformance_run_labels: dict[str, ttk.Label] = {}
        self._conformance_run_busy: bool = False
        self._conformance_stop_idle_wait: bool = False
        self._conformance_cancel_event = threading.Event()
        self._conformance_run_transport_lock = threading.Lock()
        self._conformance_run_ssh_client: Any | None = None
        self._conformance_run_script_channel: Any | None = None
        self._conformance_progress: dict[str, dict[str, Any]] = {}
        self._conformance_script_meta_cache: dict[str, dict[str, Any]] = {}
        self._conformance_detail_win: tk.Toplevel | None = None
        self._conformance_detail_text: tk.Text | None = None
        self._conformance_detail_refresh_job: str | None = None
        self._conformance_detail_last_body: str | None = None
        self._conformance_item_detail_wins: dict[str, tk.Toplevel] = {}
        self._conformance_item_detail_texts: dict[str, tk.Text] = {}
        self._conformance_item_detail_refresh_jobs: dict[str, str | None] = {}
        self._conformance_item_detail_last_body: dict[str, str] = {}
        self._conformance_detail_lines: dict[str, list[str]] = {}
        self._conformance_detail_capture_key: str | None = None
        self._conformance_detail_lock = threading.Lock()
        self._conformance_per_test_settings: dict[str, dict[str, str]] = {}
        self._conformance_oru_boost_active: bool = False
        self._conformance_oru_boost_remote_dir: str | None = None
        self._conformance_detail_run_started_wall: dict[str, str] = {}
        self._conformance_detail_run_started_mono: dict[str, float] = {}
        self._conformance_detail_run_ended_wall: dict[str, str] = {}
        self._conformance_detail_run_ended_mono: dict[str, float] = {}
        self._conformance_auto_sync_scheduled = False
        self._conformance_tab_frame: ttk.Frame | None = None
        self.conformance_path_hint_var: tk.StringVar | None = None
        self.conformance_run_remote_dir_var = tk.StringVar(value=_conf_manifest.CONFORMANCE_REMOTE_DIR)
        self.conformance_run_sw_pkg_var = tk.StringVar(value="")
        self.conformance_run_sw_remote_dir_var = tk.StringVar(value="/var/tmp/conformance/sw_pkg")
        self.conformance_restart_start_after_run_var = tk.BooleanVar(value=True)
        self.conformance_debug_var = tk.BooleanVar(value=True)
        self._conformance_stopped_start_for_run: bool = False
        self.conformance_run_rpc_timeout_var = tk.StringVar(value="30")
        self.conformance_run_idle_timeout_var = tk.StringVar(value="120")
        self.conformance_run_supervision_interval_var = tk.StringVar(value="60")
        self.conformance_run_supervision_reset_cycles_var = tk.StringVar(value="30")
        self.conformance_run_supervision_negative_fail_cycle_var = tk.StringVar(value="3")
        self.conformance_run_conn_delay_var = tk.StringVar(value="3")
        self.conformance_post_listen_wait_var = tk.StringVar(value="0")
        self.conformance_run_repeat_var = tk.StringVar(value="1")
        self.conformance_reboot_wait_var = tk.StringVar(value="360")
        self.conformance_last_run_hint_var = tk.StringVar(value="")
        self._conformance_last_run_snapshot_cache: dict[str, Any] | None = None
        self._conformance_final_results: dict[str, dict[str, Any]] = {}
        self._conformance_session_run_stats: dict[str, Any] = self._conformance_new_session_run_stats()
        self._conformance_run_stats_mode: str | None = None
        self._conformance_results_summary_win: tk.Toplevel | None = None
        self._conformance_results_summary_tree: ttk.Treeview | None = None
        self._conformance_results_summary_summary_var: tk.StringVar | None = None
        self._conformance_omit_last_run_from_config_save: bool = False
        self._conformance_extra_uploads: list[tuple[str, str]] = []
        self._build_ui()
        self._setup_shortcuts()
        self._load_saved_config()
        self._setup_auto_persist()
        self._fix_mplane_path_after_load()
        self.after(400, self._netconf_restore_library_from_config)
        self.after(250, self._conformance_refresh_status_labels)
        self.after(300, self._refresh_log_target_hint_line)

    def _fix_mplane_path_after_load(self) -> None:
        """If JSON points to a missing file, use <bundle>/mplane/mplane.xlsx when deployed next to exe."""
        cur = self.mplane_xlsx_path.get().strip()
        if cur:
            self.mplane_xlsx_path.set(self._normalize_mplane_workbook_path(cur))
            cur = self.mplane_xlsx_path.get().strip()
        try:
            exists = bool(cur) and Path(cur).expanduser().resolve().exists()
        except OSError:
            exists = False
        if not exists and _MPLANE_MASTER_XLSX.exists():
            self.mplane_xlsx_path.set(str(_MPLANE_MASTER_XLSX.resolve()))

    def _setup_xml_editor_theme(self, text_widget: tk.Text) -> None:
        self._apply_code_text_theme(text_widget)
        base_font = tkfont.Font(font=text_widget.cget("font"))
        italic_font = base_font.copy()
        italic_font.configure(slant="italic")
        text_widget.tag_configure("xml_tag", foreground="#93c5fd")
        text_widget.tag_configure("xml_attr", foreground="#f0abfc")
        text_widget.tag_configure("xml_comment", foreground="#a5b4fc", font=italic_font)

    @staticmethod
    def _apply_code_text_theme(text_widget: tk.Text) -> None:
        text_widget.configure(
            bg="#0b1220",
            fg="#f8fafc",
            insertbackground="#f8fafc",
            selectbackground="#1d4ed8",
            selectforeground="#f8fafc",
            inactiveselectbackground="#334155",
            font=("Consolas", 11),
            padx=8,
            pady=6,
            spacing1=1,
            spacing2=1,
            spacing3=1,
        )

    def _highlight_xml(self, text_widget: tk.Text) -> None:
        for tag in ("xml_tag", "xml_attr", "xml_comment"):
            text_widget.tag_remove(tag, "1.0", "end")
        content = text_widget.get("1.0", "end-1c")
        if not content:
            return
        for m in re.finditer(r"<!--[\s\S]*?-->", content):
            text_widget.tag_add("xml_comment", f"1.0+{m.start()}c", f"1.0+{m.end()}c")
        for m in re.finditer(r"</?[\w:\-\.]+(?:\s+[\w:\-\.]+(?:\s*=\s*\"[^\"]*\")?)*\s*/?>", content):
            text_widget.tag_add("xml_tag", f"1.0+{m.start()}c", f"1.0+{m.end()}c")
            seg = m.group(0)
            seg_offset = m.start()
            for am in re.finditer(r"\s([\w:\-\.]+)\s*=", seg):
                a_start = seg_offset + am.start(1)
                a_end = seg_offset + am.end(1)
                text_widget.tag_add("xml_attr", f"1.0+{a_start}c", f"1.0+{a_end}c")

    def _on_xml_editor_changed(self, text_widget: tk.Text) -> None:
        self.after_idle(lambda: self._highlight_xml(text_widget))

    def _setup_style(self) -> None:
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        self.configure(bg="#f1f5f9")
        style.configure(".", font=("Segoe UI", 10), background="#f1f5f9")
        style.configure("TFrame", background="#f1f5f9")
        style.configure("TLabelframe.Label", font=("Segoe UI", 10, "bold"))
        style.configure("Header.TLabel", font=("Segoe UI", 10, "bold"))
        style.configure("Status.TLabel", font=("Segoe UI", 10, "bold"), foreground="#0b5cab")
        style.configure("TButton", padding=7)
        style.configure("Big.TButton", padding=9, font=("Segoe UI", 10, "bold"))
        style.configure("TLabelframe", background="#f8fafc", borderwidth=1, relief="solid")
        style.configure("TNotebook", background="#f1f5f9", borderwidth=0)
        style.configure("TNotebook.Tab", padding=(16, 9), font=("Segoe UI", 10, "bold"))
        style.map(
            "TNotebook.Tab",
            background=[("selected", "#dbeafe"), ("!selected", "#e2e8f0")],
            foreground=[("selected", "#0f172a"), ("!selected", "#334155")],
        )

        # M-Plane combobox styling (compression column): normal/selected/disabled backgrounds.
        style.configure("MPlane.TCombobox", fieldbackground="#ffffff", background="#ffffff", foreground="#111827")
        style.configure("MPlaneSel.TCombobox", fieldbackground="#bfdbfe", background="#bfdbfe", foreground="#111827")
        style.configure("MPlaneDis.TCombobox", fieldbackground="#e5e7eb", background="#e5e7eb", foreground="#6b7280")
        style.map(
            "MPlane.TCombobox",
            fieldbackground=[("disabled", "#e5e7eb")],
            foreground=[("disabled", "#6b7280")],
        )

    def _setup_shortcuts(self) -> None:
        for seq in ("<Control-Shift-KeyPress-N>", "<Control-Shift-n>", "<Control-Shift-N>"):
            self.bind_all(seq, self._on_shortcut_connect, add="+")
        for seq in ("<Control-Shift-KeyPress-B>", "<Control-Shift-b>", "<Control-Shift-B>"):
            self.bind_all(seq, self._on_shortcut_disconnect, add="+")
        for seq in ("<Control-Shift-KeyPress-Q>", "<Control-Shift-q>", "<Control-Shift-Q>"):
            self.bind_all(seq, self._on_shortcut_open_logs, add="+")
        # Extra global logs shortcut (works consistently from M-Plane grid editors too).
        for seq in ("<Control-l>", "<Control-L>", "<Control-KeyPress-l>", "<Control-KeyPress-L>"):
            self.bind_all(seq, self._on_shortcut_open_logs, add="+")
        for seq in ("<Control-Shift-KeyPress-A>", "<Control-Shift-a>", "<Control-Shift-A>"):
            self.bind_all(seq, self._on_shortcut_all_armed, add="+")
        # Some layouts report Ctrl+Shift+Space as Ctrl+Space; bind both.
        for seq in ("<Control-Shift-KeyPress-space>", "<Control-Shift-space>", "<Control-space>"):
            self.bind_all(seq, self._on_shortcut_send, add="+")
        for seq in ("<Control-f>", "<Control-F>", "<Control-KeyPress-f>", "<Control-KeyPress-F>"):
            self.bind_all(seq, self._on_shortcut_find, add="+")
        for seq in ("<Control-h>", "<Control-H>", "<Control-KeyPress-h>", "<Control-KeyPress-H>"):
            self.bind_all(seq, self._on_shortcut_replace, add="+")

    def _setup_auto_persist(self) -> None:
        """Persist common settings automatically on change."""
        watched_vars: list[tk.Variable] = [
            self.remote_user,
            self.remote_host,
            self.remote_port,
            self.remote_password,
            self.remote_script_path,
            self.remote_key_path,
            self.gui_log_max_lines_var,
            self.auto_xml_send_var,
            self.auto_start_var,
            self.perf_debug_var,
            self.mplane_xlsx_path,
        ]
        watched_vars.extend(self.fields.values())
        watched_vars.extend(self.mplane_fields.values())
        for var in watched_vars:
            try:
                var.trace_add("write", self._on_any_setting_changed)
            except Exception:
                pass
        for bv in self.mplane_cc_on_vars:
            try:
                bv.trace_add("write", self._on_any_setting_changed)
            except Exception:
                pass

    def _on_any_setting_changed(self, *_args: Any) -> None:
        if getattr(self, "_config_hydrating", False):
            return
        if self._config_save_job is not None:
            try:
                self.after_cancel(self._config_save_job)
            except Exception:
                pass
        # Debounce frequent keypress updates.
        self._config_save_job = self.after(700, self._save_current_config)

    def _on_shortcut_connect(self, _event: tk.Event) -> str:
        self.manual_start_script()
        return "break"

    def _on_shortcut_disconnect(self, _event: tk.Event) -> str:
        self.stop_script()
        return "break"

    def _on_shortcut_open_logs(self, _event: tk.Event) -> str:
        self.open_log_window()
        return "break"

    def _on_shortcut_all_armed(self, _event: tk.Event) -> str:
        # Arm "all send" for a short window: Ctrl+Shift+A then Ctrl+Shift+Space.
        self._all_send_shortcut_pending_until = time.time() + 1.2
        self.append_log("[GUI] Shortcut armed: Ctrl+Shift+A+Space => All Tabs Send (1x)\n")
        return "break"

    def _on_shortcut_send(self, _event: tk.Event) -> str:
        if time.time() <= self._all_send_shortcut_pending_until:
            self._all_send_shortcut_pending_until = 0.0
            self.send_all_once()
        else:
            self.send_selected_once()
        return "break"

    def _on_shortcut_find(self, event: tk.Event) -> str | None:
        w = event.widget
        if isinstance(w, tk.Text) and any(tab.get("text") is w for tab in self.message_tabs):
            return self._on_xml_find(event)
        sheet = self._mplane_sheet_from_widget(w)
        if sheet is None and self._mplane_selection_sheet is not None:
            sheet = self._mplane_selection_sheet
        if sheet is None:
            text = self._selected_message_text_widget()
            if text is not None:
                self._open_find_replace_panel(text, focus_replace=False)
                return "break"
            return None
        self._open_mplane_find_panel(sheet, focus_replace=False)
        return "break"

    def _on_shortcut_replace(self, event: tk.Event) -> str | None:
        w = event.widget
        if isinstance(w, tk.Text) and any(tab.get("text") is w for tab in self.message_tabs):
            return self._on_xml_replace(event)
        sheet = self._mplane_sheet_from_widget(w)
        if sheet is None and self._mplane_selection_sheet is not None:
            sheet = self._mplane_selection_sheet
        if sheet is None:
            text = self._selected_message_text_widget()
            if text is not None:
                self._open_find_replace_panel(text, focus_replace=True)
                return "break"
            return None
        self._open_mplane_find_panel(sheet, focus_replace=True)
        return "break"

    def _mplane_sheet_from_widget(self, widget: Any) -> str | None:
        for sheet, mapping in self._mplane_cell_index.items():
            if widget in mapping:
                return sheet
        return None

    def _selected_message_text_widget(self) -> tk.Text | None:
        if not self.message_tabs:
            return None
        try:
            idx = self.msg_notebook.index("current")
        except Exception:
            return None
        if idx < 0 or idx >= len(self.message_tabs):
            return None
        w = self.message_tabs[idx].get("text")
        return w if isinstance(w, tk.Text) else None

    def _target_message_text_widget(self) -> tk.Text | None:
        # Priority: widget under mouse -> focused widget -> currently selected tab.
        try:
            under_mouse = self.winfo_containing(self.winfo_pointerx(), self.winfo_pointery())
        except Exception:
            under_mouse = None
        if isinstance(under_mouse, tk.Text) and any(tab.get("text") is under_mouse for tab in self.message_tabs):
            return under_mouse

        w = self.focus_get()
        if isinstance(w, tk.Text) and any(tab.get("text") is w for tab in self.message_tabs):
            return w

        selected = self._selected_message_text_widget()
        if selected is not None:
            return selected
        return None

    @staticmethod
    def _line_is_commented(line: str) -> bool:
        stripped = line.lstrip()
        return bool(stripped) and stripped.startswith("<!--") and stripped.endswith("-->")

    @staticmethod
    def _comment_line(line: str) -> str:
        if not line.strip():
            return line
        m = re.match(r"^(\s*)", line)
        indent = m.group(1) if m else ""
        body = line[len(indent):]
        return f"{indent}<!-- {body} -->"

    @staticmethod
    def _uncomment_line(line: str) -> str:
        m = re.match(r"^(\s*)<!--\s?(.*?)\s?-->\s*$", line)
        if not m:
            return line
        indent = m.group(1) or ""
        body = m.group(2) or ""
        return f"{indent}{body}"

    def _on_shortcut_toggle_comment(self, event: tk.Event) -> str | None:
        text = event.widget if isinstance(event.widget, tk.Text) else None
        if text is None or not any(tab.get("text") is text for tab in self.message_tabs):
            return None

        sel_ranges = text.tag_ranges("sel")
        if len(sel_ranges) >= 2:
            sel_start = text.index(sel_ranges[0])
            sel_end = text.index(sel_ranges[1])
        else:
            sel_start = text.index("insert")
            sel_end = text.index("insert")

        start_line = int(sel_start.split(".")[0])
        end_line = int(sel_end.split(".")[0])
        if sel_end.endswith(".0") and sel_end != sel_start:
            end_line = max(start_line, end_line - 1)

        start_idx = f"{start_line}.0"
        end_idx = f"{end_line}.end"
        selected_block = text.get(start_idx, end_idx)
        stripped = selected_block.strip()

        # Toggle as ONE XML comment block: <!-- ... -->
        should_uncomment = stripped.startswith("<!--") and stripped.endswith("-->")

        old_auto = bool(int(text.cget("autoseparators")))
        text.configure(autoseparators=False)
        text.edit_separator()
        text.mark_set("_orig_insert", "insert")
        try:
            if should_uncomment:
                m = re.match(r"^\s*<!--\s*([\s\S]*?)\s*-->\s*$", selected_block)
                new_block = m.group(1) if m else selected_block
            else:
                new_block = f"<!--\n{selected_block}\n-->"
            text.delete(start_idx, end_idx)
            text.insert(start_idx, new_block)
            text.mark_set("insert", "_orig_insert")
            text.tag_remove("sel", "1.0", "end")
        finally:
            text.mark_unset("_orig_insert")
            text.edit_separator()
            text.configure(autoseparators=old_auto)

        self._on_xml_editor_changed(text)
        return "break"

    def _on_xml_undo(self, event: tk.Event) -> str | None:
        text = event.widget if isinstance(event.widget, tk.Text) else None
        if text is None or not any(tab.get("text") is text for tab in self.message_tabs):
            return None
        try:
            text.edit_undo()
        except tk.TclError:
            return "break"
        self._on_xml_editor_changed(text)
        return "break"

    def _on_xml_redo(self, event: tk.Event) -> str | None:
        text = event.widget if isinstance(event.widget, tk.Text) else None
        if text is None or not any(tab.get("text") is text for tab in self.message_tabs):
            return None
        try:
            text.edit_redo()
        except tk.TclError:
            return "break"
        self._on_xml_editor_changed(text)
        return "break"

    def _on_xml_find(self, event: tk.Event) -> str | None:
        text = event.widget if isinstance(event.widget, tk.Text) else None
        if text is None or not any(tab.get("text") is text for tab in self.message_tabs):
            return None
        self._open_find_replace_panel(text, focus_replace=False)
        return "break"

    def _on_xml_replace(self, event: tk.Event) -> str | None:
        text = event.widget if isinstance(event.widget, tk.Text) else None
        if text is None or not any(tab.get("text") is text for tab in self.message_tabs):
            return None
        self._open_find_replace_panel(text, focus_replace=True)
        return "break"

    def _open_find_replace_panel(self, text: tk.Text, focus_replace: bool = False) -> None:
        self.find_replace_target = text
        if self.find_replace_window is None or not self.find_replace_window.winfo_exists():
            win = tk.Toplevel(self)
            win.title("Find / Replace")
            win.geometry("700x420")
            self.find_replace_window = win

            top = ttk.Frame(win, padding=10)
            top.pack(fill="x")
            ttk.Label(top, text="Find").grid(row=0, column=0, padx=(0, 8), pady=4, sticky="w")
            find_entry = ttk.Entry(top, textvariable=self.find_var)
            find_entry.grid(row=0, column=1, padx=(0, 8), pady=4, sticky="we")
            ttk.Label(top, text="Replace").grid(row=1, column=0, padx=(0, 8), pady=4, sticky="w")
            replace_entry = ttk.Entry(top, textvariable=self.replace_var)
            replace_entry.grid(row=1, column=1, padx=(0, 8), pady=4, sticky="we")
            ttk.Checkbutton(top, text="Ignore case", variable=self.find_ignore_case_var).grid(
                row=0, column=2, padx=6, pady=4, sticky="w"
            )
            ttk.Button(top, text="Preview", command=self._refresh_find_preview).grid(row=0, column=3, padx=4, pady=4)
            ttk.Button(top, text="Find Next", command=self._find_next_from_panel).grid(row=0, column=4, padx=4, pady=4)
            ttk.Button(top, text="Replace Current", command=self._replace_current_from_panel).grid(
                row=1, column=3, padx=4, pady=4
            )
            ttk.Button(top, text="Replace All", command=self._replace_all_from_panel).grid(row=1, column=4, padx=4, pady=4)
            top.columnconfigure(1, weight=1)

            preview_box = ttk.LabelFrame(win, text="Preview (double-click to jump)", padding=8)
            preview_box.pack(fill="both", expand=True, padx=10, pady=(0, 10))
            self.find_preview_list = tk.Listbox(preview_box, activestyle="none")
            self.find_preview_list.pack(side="left", fill="both", expand=True)
            self.find_preview_list.bind("<Double-Button-1>", self._jump_to_preview_selection)
            self.find_preview_list.bind("<<ListboxSelect>>", self._jump_to_preview_selection)
            pv_scroll = ttk.Scrollbar(preview_box, orient="vertical", command=self.find_preview_list.yview)
            pv_scroll.pack(side="right", fill="y")
            self.find_preview_list.configure(yscrollcommand=pv_scroll.set)

            def _on_close() -> None:
                if self.find_replace_target is not None and self.find_replace_target.winfo_exists():
                    self.find_replace_target.tag_remove("xml_find_hit", "1.0", "end")
                self.find_replace_window = None
                self.find_preview_list = None
                self.find_matches.clear()
                win.destroy()

            win.protocol("WM_DELETE_WINDOW", _on_close)
        else:
            self.find_replace_window.deiconify()
            self.find_replace_window.lift()

        if not self.find_var.get().strip() and self._last_find_text:
            self.find_var.set(self._last_find_text)
        self._refresh_find_preview()
        if focus_replace:
            self.find_replace_window.focus_force()

    def _open_find_panel_for_current_tab(self) -> None:
        text = self._selected_message_text_widget()
        if text is None:
            return
        self._open_find_replace_panel(text, focus_replace=False)

    def _open_replace_panel_for_current_tab(self) -> None:
        text = self._selected_message_text_widget()
        if text is None:
            return
        self._open_find_replace_panel(text, focus_replace=True)

    def _refresh_find_preview(self) -> None:
        text = self.find_replace_target
        if text is None or not text.winfo_exists():
            return
        keyword = self.find_var.get()
        self._last_find_text = keyword.strip() or self._last_find_text
        text.tag_remove("xml_find_hit", "1.0", "end")
        self.find_matches.clear()
        if self.find_preview_list is not None:
            self.find_preview_list.delete(0, "end")
        if not keyword:
            self.status_var.set("Find: enter text")
            return

        nocase = bool(self.find_ignore_case_var.get())
        pos = "1.0"
        while True:
            hit = text.search(keyword, pos, stopindex="end", nocase=nocase)
            if not hit:
                break
            end = f"{hit}+{len(keyword)}c"
            line_no = int(hit.split(".")[0])
            line = text.get(f"{line_no}.0", f"{line_no}.end").strip()
            preview = f"L{line_no}: {line[:180]}"
            self.find_matches.append((hit, end, preview))
            pos = end
            if len(self.find_matches) >= 1000:
                break

        for hit, end, _ in self.find_matches:
            text.tag_add("xml_find_hit", hit, end)
        text.tag_configure("xml_find_hit", background="#78350f", foreground="#fffbeb")
        if self.find_preview_list is not None:
            for _, _, preview in self.find_matches:
                self.find_preview_list.insert("end", preview)
        self.status_var.set(f"Find matches: {len(self.find_matches)}")

    def _jump_to_preview_selection(self, _event: tk.Event | None = None) -> None:
        text = self.find_replace_target
        if text is None or self.find_preview_list is None:
            return
        sel = self.find_preview_list.curselection()
        if not sel:
            return
        idx = int(sel[0])
        if idx < 0 or idx >= len(self.find_matches):
            return
        hit, end, _ = self.find_matches[idx]
        text.tag_remove("sel", "1.0", "end")
        text.tag_add("sel", hit, end)
        text.mark_set("insert", end)
        text.see(hit)

    def _find_next_from_panel(self) -> None:
        text = self.find_replace_target
        if text is None or not text.winfo_exists():
            return
        keyword = self.find_var.get()
        if not keyword:
            return
        nocase = bool(self.find_ignore_case_var.get())
        start_at = text.index("insert")
        hit = text.search(keyword, start_at, stopindex="end", nocase=nocase)
        if not hit:
            hit = text.search(keyword, "1.0", stopindex=start_at, nocase=nocase)
        if not hit:
            self.status_var.set("Find: no match")
            return
        end = f"{hit}+{len(keyword)}c"
        text.tag_remove("sel", "1.0", "end")
        text.tag_add("sel", hit, end)
        text.mark_set("insert", end)
        text.see(hit)
        self._refresh_find_preview()

    def _replace_current_from_panel(self) -> None:
        text = self.find_replace_target
        if text is None or not text.winfo_exists():
            return
        keyword = self.find_var.get()
        if not keyword:
            return
        replace_text = self.replace_var.get()
        sel = text.tag_ranges("sel")
        if len(sel) >= 2:
            s0, s1 = text.index(sel[0]), text.index(sel[1])
            selected = text.get(s0, s1)
            same = selected.lower() == keyword.lower() if self.find_ignore_case_var.get() else selected == keyword
            if same:
                text.delete(s0, s1)
                text.insert(s0, replace_text)
                text.mark_set("insert", f"{s0}+{len(replace_text)}c")
                self._on_xml_editor_changed(text)
                self._refresh_find_preview()
                self._find_next_from_panel()
                return
        self._find_next_from_panel()

    def _replace_all_from_panel(self) -> None:
        text = self.find_replace_target
        if text is None or not text.winfo_exists():
            return
        keyword = self.find_var.get()
        if not keyword:
            return
        replace_text = self.replace_var.get()
        nocase = bool(self.find_ignore_case_var.get())
        count = 0
        pos = "1.0"
        old_auto = bool(int(text.cget("autoseparators")))
        text.configure(autoseparators=False)
        text.edit_separator()
        try:
            while True:
                hit = text.search(keyword, pos, stopindex="end", nocase=nocase)
                if not hit:
                    break
                end = f"{hit}+{len(keyword)}c"
                text.delete(hit, end)
                text.insert(hit, replace_text)
                pos = f"{hit}+{len(replace_text)}c"
                count += 1
        finally:
            text.edit_separator()
            text.configure(autoseparators=old_auto)
        self._on_xml_editor_changed(text)
        self._refresh_find_preview()
        self.status_var.set(f"Replace all: {count} changed")

    @staticmethod
    def _ssh_options(use_password: bool) -> list[str]:
        # Fail fast when auth/network is not ready, instead of hanging UI.
        options = [
            "-o",
            "ConnectTimeout=8",
            "-o",
            "ConnectionAttempts=1",
            "-o",
            "ServerAliveInterval=5",
            "-o",
            "ServerAliveCountMax=1",
        ]
        if not use_password:
            options = ["-o", "BatchMode=yes", *options]
        return options

    def _build_ui(self) -> None:
        self._setup_style()
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)

        settings_tab = ttk.Frame(self.notebook)
        scheduler_tab = ttk.Frame(self.notebook)
        mplane_tab = ttk.Frame(self.notebook)
        conformance_tab = ttk.Frame(self.notebook)
        self._conformance_tab_frame = conformance_tab
        field_notes_tab = ttk.Frame(self.notebook)
        shortcuts_tab = ttk.Frame(self.notebook)
        self.notebook.add(settings_tab, text="Settings")
        self.notebook.add(mplane_tab, text="M-Plane Control")
        self.notebook.add(scheduler_tab, text="Netconf Client")
        self.notebook.add(conformance_tab, text="Conformance")
        self.notebook.add(field_notes_tab, text="M-Plane Test")
        self.notebook.add(shortcuts_tab, text="Shortcuts")

        mode_frame = ttk.LabelFrame(settings_tab, text="Execution Mode", padding=8)
        mode_frame.pack(fill="x", padx=8, pady=(8, 0))
        ttk.Label(mode_frame, text="Remote Linux (SSH)", style="Header.TLabel").grid(
            row=0, column=0, padx=8, pady=6, sticky="w"
        )
        ttk.Label(
            mode_frame,
            text="Run script on Linux via SSH.",
            foreground="#555555",
        ).grid(row=1, column=0, columnspan=4, padx=8, pady=(0, 6), sticky="w")

        remote_frame = ttk.LabelFrame(settings_tab, text="Remote SSH (used in remote mode)", padding=8)
        remote_frame.pack(fill="x", padx=8, pady=10)
        ttk.Label(remote_frame, text="SSH_USER", width=24).grid(row=0, column=0, padx=8, pady=6, sticky="w")
        ttk.Entry(remote_frame, textvariable=self.remote_user, width=36).grid(row=0, column=1, padx=8, pady=6, sticky="we")
        ttk.Label(remote_frame, text="SSH_HOST", width=24).grid(row=0, column=2, padx=8, pady=6, sticky="w")
        ttk.Entry(remote_frame, textvariable=self.remote_host, width=36).grid(row=0, column=3, padx=8, pady=6, sticky="we")

        ttk.Label(remote_frame, text="SSH_PORT", width=24).grid(row=1, column=0, padx=8, pady=6, sticky="w")
        ttk.Entry(remote_frame, textvariable=self.remote_port, width=36).grid(row=1, column=1, padx=8, pady=6, sticky="we")
        ttk.Label(remote_frame, text="SSH_PASSWORD(optional)", width=24).grid(row=1, column=2, padx=8, pady=6, sticky="w")
        ttk.Entry(remote_frame, textvariable=self.remote_password, width=36, show="*").grid(
            row=1, column=3, padx=8, pady=6, sticky="we"
        )
        # REMOTE_SCRIPT_PATH 고정(/var/tmp/miniDU_callhome.sh) — UI 숨김, 값은 유지
        self.remote_script_path.set("/var/tmp/miniDU_callhome.sh")

        ttk.Label(
            remote_frame,
            text="SSH_USER/HOST는 Linux 접속 계정 정보입니다. PASSWORD 기반 접속을 사용합니다. "
            "(REMOTE_SCRIPT_PATH=/var/tmp/miniDU_callhome.sh 고정)",
            foreground="#555555",
        ).grid(row=2, column=0, columnspan=4, padx=8, pady=(0, 6), sticky="w")
        remote_frame.columnconfigure(1, weight=1)
        remote_frame.columnconfigure(3, weight=1)

        profile_frame = ttk.LabelFrame(
            settings_tab,
            text="Lab Profile — 대역/현장 설정 묶음 (예: lab-20 / lab-60)",
            padding=8,
        )
        profile_frame.pack(fill="x", padx=8, pady=(0, 10))
        ttk.Label(profile_frame, text="Profile").grid(row=0, column=0, padx=8, pady=6, sticky="w")
        self.profile_combo = ttk.Combobox(
            profile_frame,
            textvariable=self.profile_name_var,
            width=28,
            values=("lab-20", "lab-60"),
        )
        self.profile_combo.grid(row=0, column=1, padx=8, pady=6, sticky="we")
        ttk.Button(profile_frame, text="Load", command=self._profile_load_selected).grid(
            row=0, column=2, padx=4, pady=6
        )
        ttk.Button(profile_frame, text="Save", command=self._profile_save_selected).grid(
            row=0, column=3, padx=4, pady=6
        )
        ttk.Button(profile_frame, text="Save As…", command=self._profile_save_as).grid(
            row=0, column=4, padx=4, pady=6
        )
        ttk.Button(profile_frame, text="Delete", command=self._profile_delete_selected).grid(
            row=0, column=5, padx=4, pady=6
        )
        ttk.Label(
            profile_frame,
            text=(
                "Load: 선택한 프로파일로 Settings/SSH/M-Plane·SWM 관련 값을 채웁니다.  "
                "Save: 현재 화면 값을 그 이름에 덮어씁니다.  "
                "10.0.20 ↔ 10.0.60 전환 시 프로파일만 바꾸면 됩니다."
            ),
            foreground="#555555",
            wraplength=980,
        ).grid(row=1, column=0, columnspan=6, padx=8, pady=(0, 4), sticky="w")
        profile_frame.columnconfigure(1, weight=1)

        form = ttk.LabelFrame(settings_tab, text="ORU Netconf configration", padding=8)
        form.pack(fill="x", padx=8, pady=10)

        settings_labels = {
            "LOCAL_IF": "Server NIC (ethping -i)",
            "CONN_DELAY": "Start delay before listen (s)",
            "POST_LISTEN_WAIT": "Wait after listen cmd (s)",
            "NP2_BOOT_WAIT": "netopeer2-cli boot wait (s)",
            "NP2_YANG_WAIT": "netopeer YANG preload max (s)",
            "LOGIN_WAIT_SEC": "Call Home login wait (s)",
            "ALLOWED_IP": "★ RU IPv4 (CallHome/SSH)",
            "LOCAL_IP": "Controller IPv4 (listen)",
            "ALLOWED_IP_V6": "★ RU IPv6 global (CallHome/SSH)",
            "LOCAL_IP_V6": "Controller IPv6 global (listen)",
            "CLI-ID": "★ RU SSH ID",
            "CLI-PW": "★ RU SSH PW",
            "LOG_PATH": "LOG_PATH (= /var/tmp/log/PRODUCT)",
        }
        defaults = {
            "USER": "oranuser",
            "PASSWORD": "o-ran-password",
            "ALLOWED_IP": "10.0.60.144",
            "LOCAL_IP": "10.0.60.253",
            "ALLOWED_IP_V6": "",
            "LOCAL_IP_V6": "2001:1200:1100:1000::253",
            "CLI-ID": "",
            "CLI-PW": "",
            "CALLHOME_PORT": "4334",
            "CONN_DELAY": "1",
            "POST_LISTEN_WAIT": "0",
            "NP2_BOOT_WAIT": "2",
            "NP2_YANG_WAIT": "90",
            "LOGIN_WAIT_SEC": "120",
            "NETCONF_PORT": "830",
            "PRODUCT": "nDLPU",
            "LOG_PATH": "/var/tmp/log/nDLPU",
            # 숨김(미사용): 값은 유지·전달만
            "LOCAL_IF": "",
        }
        settings_hidden = {
            "LOCAL_IF",
            "NETCONF_PORT",
            "CONN_DELAY",
            "POST_LISTEN_WAIT",
            "NP2_BOOT_WAIT",
            "NP2_YANG_WAIT",
            "LOCAL_IP",
            "LOCAL_IP_V6",
        }

        visible_i = 0
        for key, value in defaults.items():
            var = tk.StringVar(value=value)
            self.fields[key] = var
            if key in settings_hidden:
                continue
            label = settings_labels.get(key, key)
            lbl_kw: dict[str, Any] = {"text": label, "width": 28}
            if key in ("ALLOWED_IP", "ALLOWED_IP_V6", "CLI-ID", "CLI-PW"):
                lbl_kw["foreground"] = "#b45309"
                lbl_kw["font"] = ("", 9, "bold")
            ttk.Label(form, **lbl_kw).grid(
                row=visible_i // 2, column=(visible_i % 2) * 2, padx=8, pady=6, sticky="w"
            )
            show = "*" if key in ("PASSWORD", "CLI-PW") else ""
            state = "readonly" if key == "LOG_PATH" else "normal"
            ttk.Entry(form, textvariable=var, width=36, show=show, state=state).grid(
                row=visible_i // 2,
                column=(visible_i % 2) * 2 + 1,
                padx=8,
                pady=6,
                sticky="we",
            )
            visible_i += 1

        def _sync_log_path_from_product(*_args: Any) -> None:
            try:
                prod = (self.fields.get("PRODUCT").get() or "").strip()  # type: ignore[union-attr]
            except Exception:
                return
            if not prod:
                return
            safe = re.sub(r"[^0-9A-Za-z._-]+", "_", prod).strip("._") or "product"
            path = f"/var/tmp/log/{safe}"
            try:
                lp = self.fields.get("LOG_PATH")
                if lp is not None and (lp.get() or "").strip() != path:
                    lp.set(path)
            except Exception:
                pass

        try:
            self.fields["PRODUCT"].trace_add("write", _sync_log_path_from_product)
            self._sync_log_path_from_product = _sync_log_path_from_product
            _sync_log_path_from_product()
        except Exception:
            self._sync_log_path_from_product = lambda *_a: None  # type: ignore[assignment]
            pass
        # Controller listen IP: lab untag/tag 자동 (Settings UI 숨김)
        try:
            self._apply_lab_controller_listen_ips("untag")
        except Exception:
            pass

        form.columnconfigure(1, weight=1)
        form.columnconfigure(3, weight=1)
        ttk.Label(
            form,
            text=(
                "이 값들은 miniDU_callhome.sh 실행 시 환경변수로 전달됩니다. "
                "LOG_PATH 는 PRODUCT 변경 시 /var/tmp/log/<PRODUCT> 로 자동 설정됩니다. "
                "LOGIN_WAIT_SEC: RU Call Home·로그인 대기 (기본 120s). "
                "★ ALLOWED_IP / ALLOWED_IP_V6 · ★ RU SSH ID/PW(CLI-ID/PW) 는 전 시험 공용. "
                "Controller LOCAL_IP*: untag 10.0.60.253 / 2001:1200:1100:1000::253, "
                "tag 10.0.61.253 / 2001:1300:1100:1000::253 (VLAN Discovery 시 자동)."
            ),
            foreground="#555555",
        ).grid(row=(visible_i + 1) // 2 + 1, column=0, columnspan=4, padx=8, pady=(0, 6), sticky="w")

        btn_frame = ttk.Frame(settings_tab)
        btn_frame.pack(fill="x", padx=8, pady=8)

        self.start_btn = ttk.Button(btn_frame, text="Start", command=self.manual_start_script, style="Big.TButton")
        self.start_btn.pack(side="left")
        self.stop_btn = ttk.Button(btn_frame, text="Stop", command=self.stop_script, state="disabled", style="Big.TButton")
        self.stop_btn.pack(side="left", padx=8)
        ttk.Button(btn_frame, text="Logs", command=self.open_log_window, style="Big.TButton").pack(side="left", padx=8)
        ttk.Checkbutton(btn_frame, text="Auto XML Send", variable=self.auto_xml_send_var).pack(side="left", padx=10)
        ttk.Checkbutton(
            btn_frame,
            text="Auto reconnect on disconnect (until Stop)",
            variable=self.auto_start_var,
            command=self._on_auto_start_toggled,
        ).pack(side="left", padx=10)
        ttk.Label(btn_frame, text="Status:", style="Header.TLabel").pack(side="right", padx=(0, 6))
        ttk.Label(btn_frame, textvariable=self.status_var, style="Status.TLabel").pack(side="right")

        sched_ctrl = ttk.LabelFrame(scheduler_tab, text="Scheduler Control", padding=8)
        sched_ctrl.pack(fill="x", padx=8, pady=(8, 6))
        self.send_selected_btn = ttk.Button(
            sched_ctrl, text="Selected Tab Send (1x)", command=self.send_selected_once, style="Big.TButton"
        )
        self.send_selected_btn.pack(side="left", padx=(8, 6), pady=8)
        self.send_all_btn = ttk.Button(
            sched_ctrl, text="All Tabs Send (1x)", command=self.send_all_once, style="Big.TButton"
        )
        self.send_all_btn.pack(side="left", padx=6, pady=8)
        ttk.Button(sched_ctrl, text="Logs", command=self.open_log_window, style="Big.TButton").pack(
            side="left", padx=6, pady=8
        )
        ttk.Button(sched_ctrl, text="RPC Events", command=self.open_rpc_event_window, style="Big.TButton").pack(
            side="left", padx=6, pady=8
        )
        self.send_selected_btn.config(state="disabled")
        self.send_all_btn.config(state="disabled")
        ttk.Label(sched_ctrl, text="RAW RPC (ATOM) Mode", style="Header.TLabel").pack(side="left", padx=(16, 6), pady=8)
        ttk.Checkbutton(sched_ctrl, text="Perf DEBUG", variable=self.perf_debug_var).pack(
            side="left", padx=(16, 6), pady=8
        )
        ttk.Label(
            sched_ctrl,
            text="RAW RPC 모드에서는 <rpc>...</rpc> 원문을 그대로 전송합니다.",
            foreground="#555555",
        ).pack(side="right", padx=8)

        client_body = ttk.Frame(scheduler_tab)
        client_body.pack(fill="both", expand=True)
        client_paned = tk.PanedWindow(client_body, orient=tk.HORIZONTAL, sashwidth=6, bg="#e2e8f0")
        client_paned.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        lib_panel = ttk.Frame(client_paned, width=300)
        editor_panel = ttk.Frame(client_paned)
        client_paned.add(lib_panel, minsize=220)
        client_paned.add(editor_panel, minsize=420)

        self._build_netconf_library_panel(lib_panel)

        msg_tools = ttk.Frame(editor_panel)
        msg_tools.pack(fill="x", pady=4)
        ttk.Button(msg_tools, text="Add Message Tab", command=self.add_message_tab).pack(side="left")
        ttk.Button(msg_tools, text="Load XML/Text File", command=self.load_file_into_selected_tab).pack(side="left", padx=6)
        ttk.Button(msg_tools, text="Load Excel Tabs", command=self.load_excel_tabs).pack(side="left", padx=6)
        ttk.Button(msg_tools, text="Find", command=self._open_find_panel_for_current_tab).pack(side="left", padx=6)
        ttk.Button(msg_tools, text="Replace", command=self._open_replace_panel_for_current_tab).pack(side="left", padx=6)
        ttk.Button(msg_tools, text="Move Left", command=self.move_selected_tab_left).pack(side="left", padx=6)
        ttk.Button(msg_tools, text="Move Right", command=self.move_selected_tab_right).pack(side="left", padx=6)
        ttk.Button(msg_tools, text="Remove Selected Tab", command=self.remove_selected_message_tab).pack(side="left", padx=6)

        self.msg_notebook = ttk.Notebook(editor_panel)
        self.msg_notebook.pack(fill="both", expand=True)
        self.add_message_tab(initial_title="MSG-1")

        self._build_mplane_tab(mplane_tab)
        self._build_conformance_tab(conformance_tab)
        self._build_guardrails_tab(field_notes_tab)

        shortcuts_frame = ttk.LabelFrame(shortcuts_tab, text="Keyboard Shortcuts", padding=10)
        shortcuts_frame.pack(fill="both", expand=True, padx=8, pady=8)
        shortcut_guide = tk.Text(
            shortcuts_frame,
            wrap="word",
            height=16,
            relief="flat",
            state="normal",
        )
        self._apply_code_text_theme(shortcut_guide)
        shortcut_guide.insert(
            "1.0",
            "\n".join(
                [
                    "Connection / Control",
                    "  - Ctrl + Shift + N : Connect (Start)",
                    "  - Ctrl + Shift + B : Disconnect (Stop)",
                    "  - Settings: Auto Start — manual Start only; if checked, retry every 10s after failure/exit until Stop or unchecked.",
                    "  - Ctrl + Shift + Q : Open Logs Window",
                    "    * Same action as the Logs button in Settings/Netconf Client.",
                    "    * In Logs: set \"GUI keep lines\" + Apply to change how much live log is kept.",
                    "",
                    "XML Send",
                    "  - Ctrl + Shift + Space : Selected Tab Send (1x)",
                    "  - Ctrl + Shift + A + Space : All Tabs Send (1x)",
                    "",
                    "Netconf Client — XML Library (left tree)",
                    "  - Load Folder… : scan .xml/.rpc/.txt into tree",
                    "  - Double-click file : open in editor tab",
                    "  - Send GET / Send SET : send selected library file",
                    "  - Send (library) / Ctrl+Shift+Space : send selected editor tab (1x)",
                    "  - Right-click tree item for context menu",
                    "",
                    "M-Plane Control tab",
                    "  - Reload workbook: ON column is checkboxes; edit grid, Apply when session is ready.",
                    "",
                    "XML Edit",
                    "  - Ctrl + / : Comment/Uncomment selected lines",
                    "  - Ctrl + Z : Undo in XML editor",
                    "  - Ctrl + Shift + Z : Redo in XML editor",
                    "  - Ctrl + F : Open Find/Replace panel (Find)",
                    "  - Ctrl + H : Open Find/Replace panel (Replace)",
                    "    * Panel supports Preview, Find Next, Replace Current, Replace All.",
                    "    * If no selection, toggles comment on current line.",
                    "    * Works on full line units for selected range.",
                ]
            ),
        )
        shortcut_guide.configure(state="disabled")
        shortcut_guide.pack(fill="both", expand=True)

        self.after(120, self._flush_log_buffer)
        self.bind("<Configure>", self._on_main_window_configure)
        self.protocol("WM_DELETE_WINDOW", self.on_close)

        try:
            self.notebook.bind("<<NotebookTabChanged>>", self._on_main_notebook_tab_changed)
        except tk.TclError:
            pass
        self._sync_manual_send_widgets()

    def _on_main_notebook_tab_changed(self, _event: tk.Event | None = None) -> None:
        self.after_idle(self._sync_manual_send_widgets)
        try:
            if self._conformance_tab_frame is not None and self.notebook.select() == str(self._conformance_tab_frame):
                self.after_idle(self._conformance_refresh_status_labels)
        except tk.TclError:
            pass

    @staticmethod
    def _resolve_local_script_path() -> Path:
        name = "miniDU_callhome.sh"
        candidates: list[Path] = []
        app_dir = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
        candidates.append(app_dir / name)
        candidates.append(Path(__file__).resolve().with_name(name))
        meipass = getattr(sys, "_MEIPASS", "")
        if meipass:
            candidates.append(Path(meipass) / name)
        # De-duplicate while preserving order.
        seen: set[str] = set()
        uniq: list[Path] = []
        for p in candidates:
            ps = str(p)
            if ps in seen:
                continue
            seen.add(ps)
            uniq.append(p)
        for p in uniq:
            if p.exists():
                return p
        return uniq[0]

    def _sync_manual_send_widgets(self) -> None:
        ok = bool(self.session_established and self.manual_send_ready)
        st = "normal" if ok else "disabled"
        try:
            self.send_selected_btn.configure(state=st)
            self.send_all_btn.configure(state=st)
            for btn in (
                self.netconf_lib_get_btn,
                self.netconf_lib_set_btn,
                self.netconf_lib_send_btn,
            ):
                if btn is not None:
                    btn.configure(state=st)
        except tk.TclError:
            pass
        pb = self.mplane_apply_btn
        if pb is not None:
            try:
                # Apply for M-Plane just queues tabs into Netconf Client;
                # it should not depend on live connection/session state.
                pb.configure(state="normal")
                try:
                    pb.state(["!disabled"])
                except tk.TclError:
                    pass
            except tk.TclError:
                pass

    def _append_session_log_tail(self, text: str) -> None:
        if not text or not self.is_running:
            return
        with self.log_lock:
            self._recent_log_for_session += text
            if len(self._recent_log_for_session) > 128_000:
                self._recent_log_for_session = self._recent_log_for_session[-128_000:]

    @staticmethod
    def _recent_session_log_chunk(tail: str, *, max_chars: int = 4500, max_lines: int = 30) -> str:
        if not tail:
            return ""
        lines = tail.splitlines()
        if len(lines) > max_lines:
            lines = lines[-max_lines:]
        chunk = "\n".join(lines)
        if len(chunk) > max_chars:
            chunk = chunk[-max_chars:]
        return chunk

    @staticmethod
    def _log_indicates_session_alive(upper: str) -> bool:
        """True only when supervision/login is actively working — not generic rpc-reply."""
        if not upper:
            return False
        if any(
            m in upper
            for m in (
                "SUPERVISION RESET SUCCESSFUL",
                "HEARTBEAT SUCCESSFUL, SESSION ALIVE",
                "LOGIN SUCCESSFUL",
            )
        ):
            return True
        if "AUTHENTICATION SUCCESSFUL" in upper and "SUPERVISION ACTIVE" in upper:
            return True
        return False

    @staticmethod
    def _log_indicates_minidu_reconnect_cycle(upper: str) -> bool:
        """miniDU is already reconnecting internally — not a new user-visible disconnect."""
        if not upper:
            return False
        return any(
            m in upper
            for m in (
                "RECONNECTING IN",
                "CALL HOME SESSION ROUND",
                "NETOPEER2-CLI BOOT WAIT",
                "LOGIN NOT ESTABLISHED THIS ROUND",
                "WAITING FOR CLIENT CONNECTION",
                "STARTING CALLHOME LISTENER",
                "NETOPEER2-CLI 세션 종료",
                "CLOSE-SESSION",
            )
        )

    @classmethod
    def _log_indicates_session_lost(cls, upper: str, *, recent_only: str = "") -> bool:
        """Session lost: active NETCONF ended or supervision failed (ORU reboot)."""
        recent = (recent_only or upper).upper()
        if cls._log_indicates_session_alive(recent):
            return False
        if cls._log_indicates_minidu_reconnect_cycle(recent):
            return False

        if any(
            m in recent
            for m in (
                "NETOPEER2-CLI 프로세스 종료",
                "NETOPEER2-CLI DIED",
                "NETOPEER2-CLI EXITED",
                "NETCONF SESSION ENDED",
                "M-PLANE ACTIVATION END",
                "MAX SESSION ERRORS",
            )
        ):
            return True

        if "BOTH SUPERVISION RESET AND HEARTBEAT FAILED" in recent:
            return True

        if "SUPERVISION RESET TIMEOUT" in recent:
            after_timeout = recent.split("SUPERVISION RESET TIMEOUT")[-1]
            if not cls._log_indicates_session_alive(after_timeout):
                return True

        return False

    def _update_session_activity_from_log(self, text: str) -> None:
        upper = text.upper()
        if any(
            m in upper
            for m in (
                "SUPERVISION RESET SUCCESSFUL",
                "HEARTBEAT SUCCESSFUL, SESSION ALIVE",
                "AUTHENTICATION SUCCESSFUL",
                "ACCEPTED A CONNECTION ON",
                "SUPERVISION ACTIVE",
            )
        ):
            self._last_session_activity_mono = time.monotonic()

    def _cancel_session_watch(self) -> None:
        jid = self._session_watch_job
        self._session_watch_job = None
        if jid is not None:
            try:
                self.after_cancel(jid)
            except Exception:
                pass

    def _schedule_session_watch(self) -> None:
        self._cancel_session_watch()
        self._session_watch_rounds = 0
        self._session_watch_job = self.after(2000, self._session_watch_tick)

    def _session_watch_tick(self) -> None:
        self._session_watch_job = None
        if not self.is_running:
            return
        if self._user_stop_requested:
            return

        ch = self.paramiko_channel
        if ch is not None and getattr(ch, "closed", False):
            self._handle_transport_ended("ssh channel closed (watch)", -1)
            return
        client = self.paramiko_client
        if client is not None:
            try:
                tr = client.get_transport()
                if tr is None or not tr.is_active():
                    self._handle_transport_ended("ssh transport inactive (watch)", -1)
                    return
            except Exception:
                self._handle_transport_ended("ssh transport error (watch)", -1)
                return
        proc = self.proc
        if proc is not None and proc.poll() is not None:
            rc = proc.returncode if proc.returncode is not None else -1
            self._handle_transport_ended(f"process exited (watch rc={rc})", rc)
            return

        with self.log_lock:
            tail = self._recent_log_for_session
        if tail:
            tail_slice = tail[-32_000:]
            recent = self._recent_session_log_chunk(tail_slice)
            self._update_session_activity_from_log(recent)
            self._detect_session_lost(recent)
            if not (self.session_established and self.manual_send_ready):
                self._detect_session_established(tail_slice)

        if (
            getattr(self, "_session_lost_logged", False)
            and not self.session_established
            and self._session_lost_at_mono > 0
            and (time.monotonic() - self._session_lost_at_mono) > self._session_lost_force_reconnect_sec
            and not self._transport_reconnect_pending
        ):
            if self._conformance_run_busy or self._conformance_stop_idle_wait:
                self.append_log(
                    "[GUI] NETCONF session still down during Conformance run — "
                    "will retry Start after tests finish (or click Stop).\n"
                )
                self._session_lost_at_mono = time.monotonic()
            elif self.auto_start_var.get():
                with self.log_lock:
                    tail_chk = self._recent_log_for_session[-8000:]
                if self._log_indicates_minidu_reconnect_cycle(tail_chk.upper()):
                    self._session_lost_at_mono = time.monotonic()
                else:
                    self.append_log(
                        f"[GUI] Session lost >{int(self._session_lost_force_reconnect_sec)}s — forcing Start reconnect.\n"
                    )
                    self._handle_transport_ended("session lost timeout (watch)", -1)
                    return

        self._session_watch_rounds += 1
        if self._session_watch_rounds > 500:
            self._session_watch_rounds = 0
        self._session_watch_job = self.after(2500, self._session_watch_tick)

    def _build_mplane_cc_sheet(self, parent: ttk.Widget) -> None:
        self.mplane_cc_cell_vars.clear()
        self.mplane_cc_enabled_vars.clear()
        wrap = ttk.Frame(parent)
        wrap.pack(fill="both", expand=True)
        canv = tk.Canvas(wrap, highlightthickness=0, bg="#f8fafc", height=280)
        ys = ttk.Scrollbar(wrap, orient="vertical", command=canv.yview)
        xs = ttk.Scrollbar(wrap, orient="horizontal", command=canv.xview)
        canv.configure(yscrollcommand=ys.set, xscrollcommand=xs.set)
        canv.grid(row=0, column=0, sticky="nsew")
        ys.grid(row=0, column=1, sticky="ns")
        xs.grid(row=1, column=0, sticky="ew")
        wrap.rowconfigure(0, weight=1)
        wrap.columnconfigure(0, weight=1)

        inner = ttk.Frame(canv, padding=2)
        win_id = canv.create_window((0, 0), window=inner, anchor="nw")

        def _on_inner_configure(_e: tk.Event | None = None) -> None:
            canv.configure(scrollregion=canv.bbox("all"))

        inner.bind("<Configure>", _on_inner_configure)

        def _on_canvas_configure(e: tk.Event) -> None:
            # Always follow viewport width (do not lock to requested width).
            canv.itemconfigure(win_id, width=max(1, int(e.width)))

        canv.bind("<Configure>", _on_canvas_configure)
        # Ensure initial width sync without requiring manual resize.
        self.after_idle(lambda c=canv, w=win_id: c.itemconfigure(w, width=max(1, c.winfo_width())))

        cols = list(_MPLANE_CC_TREE_COLS)
        entry_cols = cols[1:]
        hdr = ("#", "ON") + tuple(c.replace("_", " ") for c in entry_cols)
        for j, h in enumerate(hdr):
            ttk.Label(inner, text=h, font=("Segoe UI", 9, "bold")).grid(row=0, column=j, padx=1, pady=1, sticky="w")

        for ri in range(_MPLANE_CC_GRID_ROWS):
            ttk.Label(inner, text=str(ri + 1), width=3).grid(row=ri + 1, column=0, padx=1, pady=1, sticky="e")
            bv = tk.BooleanVar(value=False)
            self.mplane_cc_enabled_vars.append(bv)
            ttk.Checkbutton(inner, variable=bv).grid(row=ri + 1, column=1, padx=2, pady=1)
            row_vars: list[tk.StringVar] = []
            for j, _c in enumerate(entry_cols, start=2):
                sv = tk.StringVar(value="")
                row_vars.append(sv)
                ttk.Entry(inner, textvariable=sv, width=9).grid(row=ri + 1, column=j, padx=1, pady=1, sticky="we")
            self.mplane_cc_cell_vars.append(row_vars)
        for j in range(len(hdr)):
            inner.columnconfigure(j, weight=1)

    def _build_mplane_phy_grid(self, parent: ttk.Widget, sheet: str, include_center: bool) -> None:
        """Simple 8-row grid for per-sheet physical params (PRB + optional Center MHz)."""
        self.mplane_phy_enabled[sheet] = []
        self.mplane_phy_prb[sheet] = []
        self.mplane_phy_center_mhz[sheet] = []

        wrap = ttk.Frame(parent)
        wrap.pack(fill="x", expand=False)
        hdr = ["#", "ON", "PRB"]
        if include_center:
            hdr.append("Center(MHz)")
        for j, h in enumerate(hdr):
            ttk.Label(wrap, text=h, font=("Segoe UI", 9, "bold")).grid(row=0, column=j, padx=3, pady=2, sticky="w")

        for ri in range(_MPLANE_PHY_ROWS):
            ttk.Label(wrap, text=str(ri + 1), width=3).grid(row=ri + 1, column=0, padx=3, pady=1, sticky="e")
            bv = tk.BooleanVar(value=False)
            self.mplane_phy_enabled[sheet].append(bv)
            ttk.Checkbutton(wrap, variable=bv).grid(row=ri + 1, column=1, padx=3, pady=1)
            prb = tk.StringVar(value="")
            self.mplane_phy_prb[sheet].append(prb)
            ttk.Entry(wrap, textvariable=prb, width=8).grid(row=ri + 1, column=2, padx=3, pady=1, sticky="we")
            if include_center:
                mhz = tk.StringVar(value="")
                self.mplane_phy_center_mhz[sheet].append(mhz)
                ttk.Entry(wrap, textvariable=mhz, width=12).grid(row=ri + 1, column=3, padx=3, pady=1, sticky="we")
        for j in range(len(hdr)):
            wrap.columnconfigure(j, weight=1 if j in (2, 3) else 0)

    def _mplane_fill_phy_grids_from_rpc(self) -> None:
        """Populate per-sheet PRB/Center grids from current rpc payloads."""
        try:
            import mplane_control as mp
        except Exception:
            return

        def _fill(sheet: str, include_center: bool) -> None:
            enabled = self.mplane_phy_enabled.get(sheet, [])
            prb_vars = self.mplane_phy_prb.get(sheet, [])
            mhz_vars = self.mplane_phy_center_mhz.get(sheet, [])
            for i in range(_MPLANE_PHY_ROWS):
                if i < len(enabled):
                    enabled[i].set(False)
                if i < len(prb_vars):
                    prb_vars[i].set("")
                if include_center and i < len(mhz_vars):
                    mhz_vars[i].set("")

            xml = (self._mplane_rpc_raw.get(sheet) or "").strip()
            if not xml:
                return
            prbs = mp.extract_nth_tag_values(xml, "number-of-prb", limit=_MPLANE_PHY_ROWS)
            centers = mp.extract_nth_tag_values(xml, "center-of-channel-bandwidth", limit=_MPLANE_PHY_ROWS) if include_center else []
            for i in range(_MPLANE_PHY_ROWS):
                pv = prbs[i] if i < len(prbs) else ""
                cv = centers[i] if i < len(centers) else ""
                if i < len(prb_vars):
                    prb_vars[i].set(pv)
                if include_center and i < len(mhz_vars):
                    mhz_vars[i].set(mp.hz_to_mhz_string(cv))
                if i < len(enabled):
                    enabled[i].set(bool(pv) or bool(cv))

        _fill("PDSCH", include_center=True)
        _fill("PUSCH", include_center=True)
        _fill("PRACH", include_center=False)

    def _mplane_fill_cc_sheet(self, cc_rows: list[dict[str, Any]]) -> None:
        for bv in self.mplane_cc_enabled_vars:
            bv.set(False)
        for sv_row in self.mplane_cc_cell_vars:
            for sv in sv_row:
                sv.set("")
        for ri, crow in enumerate(cc_rows[:_MPLANE_CC_GRID_ROWS]):
            self.mplane_cc_enabled_vars[ri].set(CallhomeGUI._mplane_enabled_from_cell(crow.get("enabled")))
            for j, ckey in enumerate(_MPLANE_CC_ENTRY_KEYS):
                raw = crow.get(ckey)
                self.mplane_cc_cell_vars[ri][j].set("" if raw is None else str(raw))

    @staticmethod
    def _mplane_enabled_from_cell(v: Any) -> bool:
        if v is None:
            return False
        if isinstance(v, bool):
            return bool(v)
        s = str(v).strip().upper()
        return s in ("ON", "1", "TRUE", "YES", "Y", "O", "ENABLE", "ENABLED")

    def _mplane_cc_rows_from_grid(self) -> list[dict[str, Any]]:
        rows_out: list[dict[str, Any]] = []
        if not self.mplane_cc_enabled_vars or not self.mplane_cc_cell_vars:
            return rows_out
        for ri in range(_MPLANE_CC_GRID_ROWS):
            if ri >= len(self.mplane_cc_enabled_vars) or ri >= len(self.mplane_cc_cell_vars):
                break
            d: dict[str, Any] = {}
            on = bool(self.mplane_cc_enabled_vars[ri].get())
            d["enabled"] = "ON" if on else "OFF"
            nonempty = on
            for j, key in enumerate(_MPLANE_CC_ENTRY_KEYS):
                if j >= len(self.mplane_cc_cell_vars[ri]):
                    break
                s = (self.mplane_cc_cell_vars[ri][j].get() or "").strip()
                d[key] = s if s else None
                if s:
                    nonempty = True
            if nonempty:
                rows_out.append(d)
        return rows_out

    def _mplane_apply_cc_grid_overrides_to_tables(self) -> None:
        """
        Apply Control-Sheet grid overrides into visible PDSCH/PUSCH/PRACH table vars.
        Targets:
          - PDSCH/PUSCH/PRACH compression <= compression
          - PUSCH t-au-offset <= t_au
          - PUSCH n-ta-offset <= n_ta
          - PDSCH t-da-offset (or t-du-offset header alias) <= t_du
        """
        cc_rows = self._mplane_cc_rows_from_grid()
        if not cc_rows:
            return

        def _col_idx(sheet: str, keys: tuple[str, ...]) -> int | None:
            meta = self._mplane_table_vars.get(sheet) or {}
            headers: list[str] = meta.get("headers") or []
            norm_headers = [self._mplane_tagify_header(h).replace("_", "-") for h in headers]
            for k in keys:
                nk = self._mplane_tagify_header(k).replace("_", "-")
                for i, hh in enumerate(norm_headers):
                    if hh == nk:
                        return i
            return None

        pdsch_comp = _col_idx("PDSCH", ("compression",))
        pusch_comp = _col_idx("PUSCH", ("compression",))
        prach_comp = _col_idx("PRACH", ("compression",))
        pusch_t_au = _col_idx("PUSCH", ("t-au-offset", "t_au_offset"))
        pusch_n_ta = _col_idx("PUSCH", ("n-ta-offset", "n_ta_offset"))
        pdsch_t_du = _col_idx("PDSCH", ("t-da-offset", "t-du-offset", "t_da_offset", "t_du_offset"))

        for i, row in enumerate(cc_rows[:_MPLANE_CC_GRID_ROWS]):
            comp_v = row.get("compression")
            if comp_v is not None and str(comp_v).strip() != "":
                for sheet_name, c_idx in (("PDSCH", pdsch_comp), ("PUSCH", pusch_comp), ("PRACH", prach_comp)):
                    if c_idx is None:
                        continue
                    meta = self._mplane_table_vars.get(sheet_name) or {}
                    vars_rows: list[list[tk.StringVar]] = meta.get("vars") or []
                    if i < len(vars_rows) and c_idx < len(vars_rows[i]):
                        vars_rows[i][c_idx].set(str(comp_v))
            if pusch_t_au is not None:
                v = row.get("t_au")
                if v is not None and str(v).strip() != "":
                    meta = self._mplane_table_vars.get("PUSCH") or {}
                    vars_rows: list[list[tk.StringVar]] = meta.get("vars") or []
                    if i < len(vars_rows) and pusch_t_au < len(vars_rows[i]):
                        vars_rows[i][pusch_t_au].set(str(v))
            if pusch_n_ta is not None:
                v = row.get("n_ta")
                if v is not None and str(v).strip() != "":
                    meta = self._mplane_table_vars.get("PUSCH") or {}
                    vars_rows: list[list[tk.StringVar]] = meta.get("vars") or []
                    if i < len(vars_rows) and pusch_n_ta < len(vars_rows[i]):
                        vars_rows[i][pusch_n_ta].set(str(v))
            if pdsch_t_du is not None:
                v = row.get("t_du")
                if (v is None or str(v).strip() == ""):
                    v = row.get("n_da")
                if v is not None and str(v).strip() != "":
                    meta = self._mplane_table_vars.get("PDSCH") or {}
                    vars_rows: list[list[tk.StringVar]] = meta.get("vars") or []
                    if i < len(vars_rows) and pdsch_t_du < len(vars_rows[i]):
                        vars_rows[i][pdsch_t_du].set(str(v))

    def _build_mplane_tab(self, parent: ttk.Frame) -> None:
        fld_meta: tuple[tuple[str, str], ...] = (
            ("cu_if_name", "Interface name"),
            ("cu_base_if", "base-interface"),
            ("cu_vlan", "VLAN ID"),
            ("cu_mac", "MAC (CU-plane sheet)"),
            ("pe_name", "Processing-element name"),
            ("odu_mac", "O-DU MAC"),
        )
        top = ttk.LabelFrame(parent, text="Workbook (.xlsx)", padding=8)
        top.pack(fill="x", padx=8, pady=(8, 6))
        top.columnconfigure(0, weight=1)
        ttk.Entry(top, textvariable=self.mplane_xlsx_path).grid(row=0, column=0, padx=(0, 8), pady=4, sticky="we")
        ttk.Button(top, text="Browse…", command=self._mplane_browse_workbook).grid(row=0, column=1, pady=4)
        self.mplane_reload_btn = ttk.Button(top, text="Reload from Excel", command=self.reload_mplane_from_excel)
        self.mplane_reload_btn.grid(row=0, column=2, padx=(8, 0), pady=4)
        self.mplane_save_btn = ttk.Button(top, text="Save as Excel", command=self.save_mplane_to_excel)
        self.mplane_save_btn.grid(row=0, column=3, padx=(8, 0), pady=4)

        gridf = ttk.LabelFrame(parent, text="Parameters (edit then Apply; LLTE / LLRE / RSRP_EffBW not loaded)", padding=8)
        gridf.pack(fill="x", padx=8, pady=6)
        self.mplane_fields.clear()
        for i, (key, label) in enumerate(fld_meta):
            self.mplane_fields[key] = tk.StringVar(value="")
            r, c = divmod(i, 2)
            c0 = c * 2
            ttk.Label(gridf, text=label).grid(row=r, column=c0, padx=8, pady=4, sticky="w")
            ttk.Entry(gridf, textvariable=self.mplane_fields[key], width=40).grid(
                row=r, column=c0 + 1, padx=8, pady=4, sticky="we"
            )
        gridf.columnconfigure(1, weight=1)
        gridf.columnconfigure(3, weight=1)

        cc_on_line = ttk.LabelFrame(parent, text="CC ON/OFF (global for DL/UL/PRACH)", padding=6)
        cc_on_line.pack(fill="x", padx=8, pady=(0, 6))
        self.mplane_cc_on_vars = []
        for i in range(_MPLANE_CC_GRID_ROWS):
            bv = tk.BooleanVar(value=True)
            self.mplane_cc_on_vars.append(bv)
            ttk.Checkbutton(
                cc_on_line,
                text=f"CC{i+1}",
                variable=bv,
                command=self._mplane_apply_off_row_styles,
            ).pack(side="left", padx=6)

        # One-screen scroll area (all tables stacked, like Excel view)
        scroll_wrap = ttk.Frame(parent)
        scroll_wrap.pack(fill="both", expand=True, padx=8, pady=6)
        canv = tk.Canvas(scroll_wrap, highlightthickness=0, bg="#f1f5f9")
        ys = ttk.Scrollbar(scroll_wrap, orient="vertical", command=canv.yview)
        xs = ttk.Scrollbar(scroll_wrap, orient="horizontal", command=canv.xview)
        canv.configure(yscrollcommand=ys.set, xscrollcommand=xs.set)
        canv.grid(row=0, column=0, sticky="nsew")
        ys.grid(row=0, column=1, sticky="ns")
        xs.grid(row=1, column=0, sticky="ew")
        scroll_wrap.rowconfigure(0, weight=1)
        scroll_wrap.columnconfigure(0, weight=1)

        inner = ttk.Frame(canv, padding=2)
        win_id = canv.create_window((0, 0), window=inner, anchor="nw")
        self.mplane_scroll_wrap = scroll_wrap
        self.mplane_scroll_canvas = canv
        self.mplane_scroll_inner = inner

        def _on_inner_configure(_e: tk.Event | None = None) -> None:
            canv.configure(scrollregion=canv.bbox("all"))

        inner.bind("<Configure>", _on_inner_configure)

        def _on_canvas_configure(e: tk.Event) -> None:
            # Keep content width locked to current viewport width.
            # Using requested width here causes runaway growth after window resize.
            canv.itemconfigure(win_id, width=max(1, int(e.width)))

        canv.bind("<Configure>", _on_canvas_configure)
        self._mplane_enable_mousewheel_scroll()

        # Control-Sheet edit grid is intentionally hidden.
        # Values are written back to Control-Sheet cells during Excel Save.

        # PDSCH/PUSCH/PRACH full tables (editable grid; changes are applied into RPC)
        for sheet, title, include_center in (
            ("PDSCH", "PDSCH (DL) — 전체 표", True),
            ("PUSCH", "PUSCH (UL) — 전체 표", True),
            ("PRACH", "PRACH — 전체 표", False),
        ):
            sec = ttk.LabelFrame(inner, text=title, padding=8)
            sec.pack(fill="both", expand=True, padx=4, pady=(0, 10))
            hint = (
                "값은 Control-Sheet 3/4/5 절 표 셀만 사용 (A열 XML 직접 편집은 무시). "
                "Apply는 GUI 편집값 반영. Excel 보관은 Save."
            )
            ttk.Label(sec, text=hint, foreground="#475569").pack(anchor="w", padx=4, pady=(0, 6))
            grid_host = ttk.Frame(sec)
            grid_host.pack(fill="both", expand=True, padx=2, pady=(8, 2))
            self._mplane_table_widgets[sheet] = {"host": grid_host, "include_center": include_center}

        ttk.Label(parent, textvariable=self.mplane_warnings_var, foreground="#92400e", wraplength=1020).pack(
            fill="x", padx=10, pady=(0, 2)
        )
        btm = ttk.Frame(parent)
        btm.pack(fill="x", padx=8, pady=(4, 8))
        self.mplane_apply_btn = ttk.Button(
            btm,
            text="Apply to device (send in workbook order)",
            command=self.apply_mplane_workbook_once,
            style="Big.TButton",
        )
        self.mplane_apply_btn.pack(side="left")
        ttk.Label(
            btm,
            text="Save / Apply / Conformance: M-Plane Control 탭(GUI)이 기준. xlsx는 Reload·보관용.",
            foreground="#64748b",
        ).pack(side="left", padx=14)

    def _mplane_pointer_over_scroll_area(self, event: tk.Event | None = None) -> bool:
        """True when the pointer is over the M-Plane table scroll canvas (or its children)."""
        canv = self.mplane_scroll_canvas
        wrap = self.mplane_scroll_wrap
        if canv is None or wrap is None:
            return False
        try:
            if event is not None and hasattr(event, "x_root") and hasattr(event, "y_root"):
                w: Any = self.winfo_containing(int(event.x_root), int(event.y_root))
            else:
                x, y = self.winfo_pointerxy()
                w = self.winfo_containing(x, y)
        except Exception:
            return False
        while w is not None:
            if w in (canv, wrap, self.mplane_scroll_inner):
                return True
            try:
                w = w.master
            except Exception:
                break
        return False

    def _mplane_on_mousewheel(self, event: tk.Event) -> str | None:
        """Scroll M-Plane tables with the mouse wheel while the pointer is over that area."""
        if not self._mplane_pointer_over_scroll_area(event):
            return None
        canv = self.mplane_scroll_canvas
        if canv is None:
            return None
        # Windows / macOS: event.delta; Linux often uses Button-4/5 separately.
        delta = getattr(event, "delta", 0) or 0
        if delta:
            steps = int(-1 * (delta / 120))
            if steps == 0:
                steps = -1 if delta > 0 else 1
            # Shift+wheel → horizontal scroll when content is wider than viewport.
            state = int(getattr(event, "state", 0) or 0)
            if state & 0x0001:  # Shift
                canv.xview_scroll(steps, "units")
            else:
                canv.yview_scroll(steps, "units")
            return "break"
        return None

    def _mplane_on_mousewheel_linux(self, event: tk.Event) -> str | None:
        if not self._mplane_pointer_over_scroll_area(event):
            return None
        canv = self.mplane_scroll_canvas
        if canv is None:
            return None
        num = int(getattr(event, "num", 0) or 0)
        state = int(getattr(event, "state", 0) or 0)
        steps = -1 if num == 4 else 1 if num == 5 else 0
        if steps == 0:
            return None
        if state & 0x0001:
            canv.xview_scroll(steps, "units")
        else:
            canv.yview_scroll(steps, "units")
        return "break"

    def _mplane_enable_mousewheel_scroll(self) -> None:
        """Bind wheel once; works for nested Entry cells created after Reload."""
        if self._mplane_wheel_bound:
            return
        self.bind_all("<MouseWheel>", self._mplane_on_mousewheel, add="+")
        self.bind_all("<Button-4>", self._mplane_on_mousewheel_linux, add="+")
        self.bind_all("<Button-5>", self._mplane_on_mousewheel_linux, add="+")
        self._mplane_wheel_bound = True

    def _mplane_browse_workbook(self) -> None:
        p = filedialog.askopenfilename(filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")])
        if p:
            self.mplane_xlsx_path.set(p)
            self._save_current_config()

    def _store_mplane_workbook_for_gui(self, source_path: str) -> str:
        """
        Keep a GUI-managed copy at ./mplane/mplane_gui_cache.xlsx and return that path.
        Never overwrite user-maintained ./mplane/mplane.xlsx automatically.
        If copy fails, return original source path.
        """
        try:
            src = Path(source_path).expanduser().resolve()
            if not src.exists():
                return source_path
            dst_dir = _app_bundle_root() / "mplane"
            dst_dir.mkdir(parents=True, exist_ok=True)
            user_main = dst_dir / "mplane.xlsx"
            # If user selected the main workbook itself, do not copy/overwrite.
            if src == user_main:
                return str(user_main)
            dst = dst_dir / "mplane_gui_cache.xlsx"
            # Cache copy only; never touches user_main.
            if src != dst:
                shutil.copy2(src, dst)
                self.append_log(f"[GUI] M-Plane workbook cached: {dst}\n")
            return str(dst)
        except Exception as exc:
            self.append_log(f"[GUI] M-Plane cache warning: {exc}\n")
            return source_path

    def _resolve_mplane_template_path(self) -> tuple[Path | None, list[Path]]:
        """
        Pick a template .xlsx for Save.
        Save should always start from the clean master template when available:
        <bundle>/mplane/mplane.xlsx -> <bundle>/mplane.xlsx -> path box -> last Reload path.
        Relative paths resolve against the bundle root (folder containing .exe or the script).
        """
        root = _app_bundle_root()
        ordered: list[Path] = []

        ordered.extend([root / "mplane" / "mplane.xlsx", root / "mplane.xlsx"])

        cfg = self.mplane_xlsx_path.get().strip()
        if cfg:
            p = Path(cfg).expanduser()
            ordered.append((root / p).resolve() if not p.is_absolute() else p.resolve())

        if self._mplane_loaded_xlsx_path:
            p = Path(self._mplane_loaded_xlsx_path).expanduser()
            ordered.append((root / p).resolve() if not p.is_absolute() else p.resolve())

        seen: set[str] = set()
        tried: list[Path] = []
        for p in ordered:
            key = str(p)
            if key in seen:
                continue
            seen.add(key)
            tried.append(p)
            try:
                if p.exists():
                    return p, tried
            except OSError:
                continue
        return None, tried

    def _normalize_mplane_workbook_path(self, raw: str) -> str:
        """Resolve workbook path for Reload/Save: absolute paths, then bundle-root-relative."""
        s = (raw or "").strip()
        if not s:
            return ""
        p = Path(s).expanduser()
        try:
            if p.is_absolute() and p.exists():
                return str(p.resolve())
            root = _app_bundle_root()
            rel = (root / p).resolve() if not p.is_absolute() else p.resolve()
            if rel.exists():
                return str(rel)
            if p.is_absolute():
                return str(p.resolve())
            return str(rel)
        except OSError:
            return s

    def save_mplane_to_excel(self) -> None:
        """Export current M-Plane UI state to a new .xlsx file."""
        self._mplane_sync_visible_widget_values()
        has_rpc = bool(self._mplane_rpc_raw)
        has_table_data = any(
            bool((self._mplane_table_vars.get(s) or {}).get("vars"))
            for s in ("PDSCH", "PUSCH", "PRACH")
        )
        if not has_rpc or not has_table_data:
            messagebox.showerror("M-Plane Save", 'No loaded data. Click "Reload from Excel" first.')
            return

        def _safe_part(text: str, fallback: str) -> str:
            s = (text or "").strip()
            if not s:
                return fallback
            s = re.sub(r'[\\/:*?"<>|]+', "_", s)
            s = re.sub(r"\s+", "_", s)
            s = s.strip("._")
            return s or fallback

        product_raw = ""
        if "PRODUCT" in self.fields:
            product_raw = self.fields["PRODUCT"].get()
        if not (product_raw or "").strip() and "pe_name" in self.mplane_fields:
            product_raw = self.mplane_fields["pe_name"].get()
        if not (product_raw or "").strip() and "cu_if_name" in self.mplane_fields:
            product_raw = self.mplane_fields["cu_if_name"].get()
        mac_raw = ""
        if "cu_mac" in self.mplane_fields:
            mac_raw = self.mplane_fields["cu_mac"].get()
        if not (mac_raw or "").strip() and "odu_mac" in self.mplane_fields:
            mac_raw = self.mplane_fields["odu_mac"].get()
        mac_raw = (mac_raw or "").replace(":", "-")
        base_name = f"{_safe_part(product_raw, 'product')}_{_safe_part(mac_raw, 'mac')}"

        # Save directly under "<exe_or_script_dir>/mplane" with auto-increment suffix.
        if getattr(sys, "frozen", False):
            root_dir = Path(sys.executable).resolve().parent
        else:
            root_dir = Path(__file__).resolve().parent
        target_dir = root_dir / "mplane"
        target_dir.mkdir(parents=True, exist_ok=True)

        candidate = target_dir / f"{base_name}.xlsx"
        if candidate.exists():
            idx = 1
            while True:
                probe = target_dir / f"{base_name}_{idx:03d}.xlsx"
                if not probe.exists():
                    candidate = probe
                    break
                idx += 1
        out_path = str(candidate)
        try:
            import openpyxl
        except Exception as exc:
            messagebox.showerror("M-Plane Save", f"openpyxl is required:\n{exc}")
            return

        # Template resolution for exe/script:
        # 1) Prefer <exe_or_script_dir>/mplane/mplane.xlsx when present (dist deployment).
        # 2) Then path box / last-loaded (relative paths resolve against bundle root).
        template_path, tried_paths = self._resolve_mplane_template_path()
        if template_path is None:
            tried = "\n".join(f"- {p}" for p in tried_paths[:12])
            messagebox.showerror(
                "M-Plane Save",
                "Template workbook not found.\n\n"
                "Place mplane.xlsx next to the app:\n"
                f"  {_app_bundle_root() / 'mplane' / 'mplane.xlsx'}\n\n"
                "Or set the full path in the M-Plane workbook field.\n\n"
                "Checked:\n"
                f"{tried}",
            )
            return
        try:
            wb = openpyxl.load_workbook(template_path)
        except Exception as exc:
            messagebox.showerror("M-Plane Save", f"Failed to open template workbook:\n{exc}")
            return
        self.append_log(f"[GUI] M-Plane Save template: {template_path}\n")

        try:
            import mplane_control as mp
            mp = importlib.reload(mp)
        except Exception as exc:
            messagebox.showerror("M-Plane Save", f"mplane_control load failed:\n{exc}")
            return

        gui_state = self._mplane_gui_state_snapshot()
        try:
            rpc, warns = self._mplane_build_rpc_from_gui(
                mp,
                gui_tables=gui_state["tables"],
                gui_live=gui_state["live"],
                gui_off_rows=gui_state["off_rows"],
                comment_off_rows=True,
            )
        except RuntimeError as exc:
            messagebox.showerror("M-Plane Save", str(exc))
            return
        live = gui_state["live"]
        if warns:
            self.append_log(f"[GUI] M-Plane Save warnings: {' | '.join(warns)}\n")

        # Control-Sheet upper CC table is untouched.
        # But detail sections (3.PDSCH / 4.PUSCH / 5.PRACH) are replaced from GUI values.
        ws_c = wb["Control-Sheet"] if "Control-Sheet" in wb.sheetnames else None
        if ws_c is not None:
            mp.write_control_sheet_simple_fields(ws_c, live)
            def _safe_set(row: int, col: int, value: Any) -> None:
                try:
                    cell = ws_c.cell(row=row, column=col)
                    if cell.__class__.__name__ == "MergedCell":
                        return
                    cell.value = value
                except Exception:
                    return

            def _gui_rows_for(sheet_name: str) -> tuple[list[str], list[list[str]]]:
                meta = self._mplane_table_vars.get(sheet_name) or {}
                headers: list[str] = meta.get("headers") or []
                vars_rows: list[list[tk.StringVar]] = meta.get("vars") or []
                rows: list[list[str]] = []
                for rr in vars_rows:
                    rows.append([(v.get() or "") for v in rr])
                return headers, rows

            # Header aliases: Control-Sheet label -> GUI table label.
            alias_keys = {
                "end point name": ("end point", "low level tx endpoint", "low level rx endpoint"),
                "link name": ("low level tx endpoint", "low level rx endpoint"),
                "array carrier name": ("tx array carrier", "rx array carrier"),
                "center freq": ("center freq",),
                "channel bw": ("channel bw",),
                "t da offset": ("t-da-offset", "t-du-offset"),
                "t au offset": ("t-au-offset",),
                "n ta offset": ("n-ta-offset",),
                "ul fft sampling offset": ("ul-fft-sampling-offset",),
                "ll fft sampling offset": ("ul-fft-sampling-offset",),
                "number of prb": ("number-of-prb",),
                "eaxc id": ("eaxc-id",),
                "processing element": ("processing-element",),
                "iq bitwidth": ("iq-bitwidth",),
                "gain correction": ("gain-correction",),
                "gain": ("gain",),
                "low level rx endpoint": ("low-level-rx-endpoint",),
                "low level tx endpoint": ("low-level-tx-endpoint",),
            }

            def _find_section_row(section_key: str) -> int | None:
                sk = mp.normalize_header(section_key)
                for rr in range(1, ws_c.max_row + 1):
                    vals = []
                    for cc in range(1, 10):
                        vv = ws_c.cell(row=rr, column=cc).value
                        if vv is not None:
                            vals.append(mp.normalize_header(str(vv)))
                    if sk in "|".join(vals):
                        return rr
                return None

            def _header_columns(header_row: int) -> list[tuple[int, str]]:
                """
                Build normalized header names using marker row (header_row-1) + header_row.
                Ex: "end-point" + "name" => "end-point name".
                """
                out_cols: list[tuple[int, str]] = []
                for cc in range(1, 80):
                    top = ws_c.cell(row=header_row - 1, column=cc).value if header_row - 1 >= 1 else None
                    bot = ws_c.cell(row=header_row, column=cc).value
                    t1 = mp.normalize_header(str(top)) if top is not None else ""
                    t2 = mp.normalize_header(str(bot)) if bot is not None else ""
                    if t1 and t2:
                        hn = f"{t1} {t2}".strip()
                    else:
                        hn = t1 or t2
                    if not hn:
                        continue
                    out_cols.append((cc, hn))
                return out_cols

            def _fill_control_detail_block(header_row: int, sheet_name: str) -> None:
                gui_headers, gui_rows = _gui_rows_for(sheet_name)
                if not gui_headers or not gui_rows:
                    return
                gui_norm = [mp.normalize_header(h) for h in gui_headers]
                ctrl_cols = _header_columns(header_row)
                if not ctrl_cols:
                    return

                # Fill up to GUI row count, keep template layout.
                for i in range(len(gui_rows)):
                    rr = header_row + 1 + i
                    row_vals = gui_rows[i]
                    row_map: dict[str, str] = {}
                    for j, kn in enumerate(gui_norm):
                        if j < len(row_vals):
                            row_map[kn] = row_vals[j]
                    # For 3.PDSCH table, sync CC ON/OFF to column A (e.g. A23~A30).
                    if sheet_name == "PDSCH":
                        on = i < len(self.mplane_cc_on_vars) and self.mplane_cc_on_vars[i].get()
                        _safe_set(rr, 1, "ON" if on else "OFF")
                    # Clear existing formula/text in target columns first,
                    # so template references do not remain.
                    for cc, _nk in ctrl_cols:
                        _safe_set(rr, cc, "")
                    # likely CC label column if present
                    for cc, nk in ctrl_cols:
                        if nk == "cc":
                            _safe_set(rr, cc, f"C{i}")
                            continue
                        # Explicit endpoint/array-carrier mapping for template header variants.
                        if ("rx array carrier" in nk) or ("tx array carrier" in nk):
                            _safe_set(
                                rr,
                                cc,
                                row_map.get("rx array carrier", "")
                                or row_map.get("tx array carrier", "")
                                or row_map.get("rx-array-carrier", "")
                                or row_map.get("tx-array-carrier", "")
                                or row_map.get("rx array carrier", "")
                                or row_map.get("tx array carrier", ""),
                            )
                            continue
                        if ("low level rx endpoint" in nk) or ("low level tx endpoint" in nk):
                            _safe_set(
                                rr,
                                cc,
                                row_map.get("low level rx endpoint", "")
                                or row_map.get("low level tx endpoint", "")
                                or row_map.get("low-level-rx-endpoint", "")
                                or row_map.get("low-level-tx-endpoint", "")
                                or row_map.get("low level rx endpoint", "")
                                or row_map.get("low level tx endpoint", ""),
                            )
                            continue
                        # DL tail columns explicit map: avoid template variant misses.
                        if sheet_name == "PDSCH":
                            if "gain" == nk or nk.endswith(" gain") or " gain " in f" {nk} ":
                                _safe_set(rr, cc, row_map.get("gain", ""))
                                continue
                            if ("t da offset" in nk) or ("t-du" in nk) or ("t du offset" in nk):
                                _safe_set(
                                    rr,
                                    cc,
                                    row_map.get("t da offset", "")
                                    or row_map.get("t du offset", "")
                                    or row_map.get("t-da-offset", "")
                                    or row_map.get("t-du-offset", ""),
                                )
                                continue
                        val = row_map.get(nk, "")
                        if not val and nk in alias_keys:
                            for ak in alias_keys[nk]:
                                val = row_map.get(mp.normalize_header(ak), "")
                                if val:
                                    break
                        if not val:
                            # Fallback for template/header typos and variants.
                            for rk, rv in row_map.items():
                                if rv and (rk in nk or nk in rk):
                                    val = rv
                                    break
                        _safe_set(rr, cc, val)

            # Fill by explicit section anchors to avoid fragile global pattern matching.
            sec_pd = _find_section_row("3. pdsch")
            sec_pu = _find_section_row("4. pusch")
            sec_pr = _find_section_row("5. prach")
            # In this workbook template, detail headers are located at section_row + 3.
            if sec_pd is not None:
                _fill_control_detail_block(sec_pd + 3, "PDSCH")
            if sec_pu is not None:
                _fill_control_detail_block(sec_pu + 3, "PUSCH")
            if sec_pr is not None:
                _fill_control_detail_block(sec_pr + 3, "PRACH")

        # IMPORTANT:
        # Do NOT rewrite whole RPC sheets here. Users rely on the original template layout
        # (tables/formulas/formatting). We only update Control-Sheet values in place.

        try:
            out = Path(out_path)
            wb.save(str(out))
            self.mplane_xlsx_path.set(str(out.resolve()))
            self._mplane_loaded_xlsx_path = str(out.resolve())
            self._save_current_config()
            self.append_log(f"[GUI] M-Plane saved: {out} (workbook path updated — Reload to refresh)\n")
            messagebox.showinfo(
                "M-Plane Save",
                f"Saved:\n{out}\n\nWorkbook path updated. Click \"Reload from Excel\" to load MAC/VLAN from the saved file.",
            )
        except Exception as exc:
            messagebox.showerror("M-Plane Save", f"Failed to save file:\n{exc}")
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def reload_mplane_from_excel(self) -> None:
        path = self._normalize_mplane_workbook_path(self.mplane_xlsx_path.get())
        if not path:
            messagebox.showwarning("M-Plane", "Choose an .xlsx path first.")
            return
        self.mplane_xlsx_path.set(path)
        if self.mplane_reload_btn is not None:
            self.mplane_reload_btn.configure(state="disabled")
        self.mplane_warnings_var.set("Loading workbook...")
        self.append_log(f"[GUI] M-Plane: loading workbook in background: {path}\n")
        threading.Thread(target=self._reload_mplane_from_excel_worker, args=(path,), daemon=True).start()

    def _reload_mplane_from_excel_worker(self, path: str) -> None:
        try:
            import mplane_control as mp
            mp = importlib.reload(mp)
        except ImportError as exc:
            self.after(0, lambda: messagebox.showerror("M-Plane", f"Cannot load mplane_control / openpyxl:\n{exc}"))
            self.after(0, self._mplane_reload_done)
            return
        try:
            rpc, baselines, merged, cc_rows, tables, warnings = mp.load_workbook_payloads(path)
            payload = (dict(rpc), dict(baselines), dict(merged), list(cc_rows), dict(tables), list(warnings))
            self.after(0, lambda p=payload, src=path: self._apply_mplane_reload_result(src, p))
        except FileNotFoundError:
            self.after(0, lambda: messagebox.showerror("M-Plane", f"File not found:\n{path}"))
            self.after(0, self._mplane_reload_done)
        except Exception as exc:
            self.after(0, lambda: messagebox.showerror("M-Plane", f"Failed to load workbook:\n{exc}"))
            self.after(0, self._mplane_reload_done)

    def _apply_mplane_reload_result(
        self,
        path: str,
        payload: tuple[dict[str, str], dict[str, str], dict[str, Any], list[dict[str, Any]], dict[str, tuple[list[str], list[list[str]]]], list[str]],
    ) -> None:
        rpc, baselines, merged, cc_rows, tables, warnings = payload
        self._dbg_mplane(
            "reload_result "
            + ", ".join(
                f"{k}:h={len(v[0])},r={len(v[1])}" for k, v in sorted(tables.items())
            )
        )
        self._mplane_rpc_raw = rpc
        self._mplane_baselines = baselines
        self._mplane_merged = dict(merged)
        self._mplane_cc_rows = list(cc_rows)
        self._mplane_tables = tables
        self._mplane_loaded_xlsx_path = path
        self.mplane_xlsx_path.set(path)
        for key, var in self.mplane_fields.items():
            v = merged.get(key, "")
            var.set("" if v is None else str(v))
        try:
            import mplane_control as mp

            live = {k: (self.mplane_fields[k].get() or "").strip() for k in ("cu_if_name", "cu_base_if", "cu_vlan") if k in self.mplane_fields}
            ifname = mp.resolve_l2vlan_interface_name(live)
            if ifname and "cu_if_name" in self.mplane_fields:
                self.mplane_fields["cu_if_name"].set(ifname)
        except Exception:
            pass
        self._mplane_fill_cc_sheet(cc_rows)
        # Sync PDSCH detail column-A / RPC comment state into CC toggles (comment-out index).
        detail_flags = merged.get("_detail_cc_on_flags")
        if isinstance(detail_flags, list) and detail_flags:
            for i, bv in enumerate(self.mplane_cc_on_vars):
                bv.set(detail_flags[i] if i < len(detail_flags) else True)
        else:
            for i, bv in enumerate(self.mplane_cc_on_vars):
                if i < len(cc_rows):
                    bv.set(self._mplane_enabled_from_cell(cc_rows[i].get("enabled")))
                else:
                    bv.set(True)
        self._mplane_render_full_tables()
        wtext = " | ".join(warnings) if warnings else ""
        self.mplane_warnings_var.set(wtext[:2000] if len(wtext) > 2000 else wtext)
        tbl_info = ", ".join(
            f"{k}={len(v[1])}rows" for k, v in sorted(tables.items()) if v and v[1]
        )
        self.append_log(
            f"[GUI] M-Plane workbook loaded: {path} ({len(cc_rows)} CC row(s); Control-Sheet tables: {tbl_info})\n"
        )
        for w in warnings:
            self.append_log(f"[GUI] M-Plane: {w}\n")
        self._save_current_config()
        self.after_idle(self._sync_manual_send_widgets)
        self._mplane_reload_done()

    def _mplane_reload_done(self) -> None:
        if self.mplane_reload_btn is not None:
            self.mplane_reload_btn.configure(state="normal")

    def _dbg_mplane(self, msg: str) -> None:
        """Write debug traces for M-Plane table rendering."""
        try:
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            with self.mplane_debug_log_path.open("a", encoding="utf-8") as f:
                f.write(f"[{ts}] {msg}\n")
        except Exception:
            pass

    def _mplane_render_full_tables(self) -> None:
        """Render extracted Excel tables for PDSCH/PUSCH/PRACH in the one-screen layout (editable)."""
        self._dbg_mplane("render_start")
        for sheet, meta in self._mplane_table_widgets.items():
            host = meta.get("host")
            if host is None or not host.winfo_exists():
                continue
            for child in list(host.winfo_children()):
                try:
                    child.destroy()
                except Exception:
                    pass
            headers, rows = self._mplane_tables.get(sheet, ([], []))
            self._mplane_table_vars[sheet] = {"headers": headers, "vars": [], "entries": []}
            self._mplane_cell_index[sheet] = {}
            self._mplane_undo_stack[sheet] = []
            self._mplane_redo_stack[sheet] = []
            self._dbg_mplane(
                f"{sheet}: headers={len(headers)} rows={len(rows)} first_row_len={(len(rows[0]) if rows else 0)}"
            )
            if not headers:
                ttk.Label(host, text="(표를 찾지 못했습니다. 엑셀에서 헤더 행에 end_point / number-of-prb 가 있는지 확인)", foreground="#b45309").pack(anchor="w")
                continue
            ttk.Label(host, text=f"rows: {len(rows)}", foreground="#64748b").pack(anchor="w", padx=2, pady=(0, 2))
            # Use direct frame-grid rendering (no nested canvas) to avoid width=1 timing bugs.
            inner = ttk.Frame(host, padding=1)
            inner.pack(fill="x", expand=True)

            # Header row with row numbers
            ttk.Label(inner, text="#", font=("Segoe UI", 9, "bold")).grid(row=0, column=0, padx=1, pady=1, sticky="w")
            for j, h in enumerate(headers):
                ttk.Label(inner, text=h, font=("Segoe UI", 9, "bold")).grid(
                    row=0, column=j + 1, padx=1, pady=1, sticky="ew"
                )

            # Data rows (editable)
            for i, r in enumerate(rows):
                ttk.Label(inner, text=str(i + 1), foreground="#64748b", font=("Segoe UI", 9)).grid(
                    row=i + 1, column=0, padx=1, pady=1, sticky="e"
                )
                row_vars: list[tk.StringVar] = []
                row_entries: list[tk.Widget] = []
                c_comp = self._mplane_col_index(sheet, "compression")
                c_iq = self._mplane_col_index(sheet, "iq-bitwidth")
                for j in range(len(headers)):
                    v = r[j] if j < len(r) else ""
                    if c_comp is not None and j == c_comp:
                        iqv = ""
                        if c_iq is not None and c_iq < len(r):
                            iqv = (r[c_iq] or "").strip()
                        if iqv in {"16", "16.0"}:
                            v = "no comp"
                        else:
                            if (v or "").strip().lower() in {"", "no comp", "nocomp", "static"}:
                                v = "BFP"
                            nv = self._mplane_normalize_compression_value(v)
                            if nv is not None:
                                v = nv
                    sv = tk.StringVar(value=v)
                    sv.trace_add(
                        "write",
                        lambda *_a, s=sheet, rr=i, cc=j: self._mplane_on_cell_var_changed(s, rr, cc),
                    )
                    if c_iq is not None and j == c_iq:
                        sv.trace_add("write", lambda *_a: self.after_idle(self._mplane_apply_off_row_styles))
                    row_vars.append(sv)
                    self._mplane_last_valid_cell_values[(sheet, i, j)] = v
                    # Compression cell: editable dropdown + direct typing.
                    if c_comp is not None and j == c_comp:
                        ent: tk.Widget = ttk.Combobox(
                            inner,
                            textvariable=sv,
                            values=("BFP", "exponent"),
                            width=8,
                            state="normal",
                        )
                    else:
                        # Use compact tk.Entry so DL/UL can shrink like PRACH with many columns.
                        ent = tk.Entry(
                            inner,
                            textvariable=sv,
                            width=6,
                            font=("Consolas", 10),
                            relief="solid",
                            borderwidth=1,
                        )
                    # Keep endpoint key immutable to prevent accidental carrier row loss.
                    if j == 0:
                        try:
                            ent.configure(state="readonly", readonlybackground="#eef2ff")
                        except Exception:
                            ent.configure(state="disabled")
                    ent.grid(row=i + 1, column=j + 1, padx=1, pady=1, sticky="we")
                    ent.bind(
                        "<Button-1>",
                        lambda e, s=sheet, rr=i, cc=j: self._mplane_on_cell_press(e, s, rr, cc),
                    )
                    ent.bind(
                        "<B1-Motion>",
                        lambda e, s=sheet, rr=i, cc=j: self._mplane_on_cell_drag(e, s, rr, cc),
                    )
                    ent.bind(
                        "<Enter>",
                        lambda e, s=sheet, rr=i, cc=j: self._mplane_on_cell_enter(e, s, rr, cc),
                    )
                    ent.bind("<ButtonRelease-1>", self._mplane_on_cell_release)
                    ent.bind("<Control-v>", lambda e, s=sheet: self._mplane_on_paste(e, s))
                    ent.bind("<Control-V>", lambda e, s=sheet: self._mplane_on_paste(e, s))
                    ent.bind("<Control-c>", lambda e, s=sheet: self._mplane_on_copy(e, s))
                    ent.bind("<Control-C>", lambda e, s=sheet: self._mplane_on_copy(e, s))
                    ent.bind("<Control-z>", lambda e, s=sheet: self._mplane_on_undo(e, s))
                    ent.bind("<Control-Z>", lambda e, s=sheet: self._mplane_on_undo(e, s))
                    ent.bind("<Control-Shift-z>", lambda e, s=sheet: self._mplane_on_redo(e, s))
                    ent.bind("<Control-Shift-Z>", lambda e, s=sheet: self._mplane_on_redo(e, s))
                    ent.bind("<Control-y>", lambda e, s=sheet: self._mplane_on_redo(e, s))
                    ent.bind("<Control-Y>", lambda e, s=sheet: self._mplane_on_redo(e, s))
                    ent.bind("<Delete>", lambda e, s=sheet: self._mplane_on_multi_clear(e, s))
                    ent.bind("<BackSpace>", lambda e, s=sheet: self._mplane_on_multi_clear(e, s))
                    ent.bind("<FocusOut>", lambda _e, s=sheet, rr=i, cc=j: self._mplane_on_cell_focus_out(s, rr, cc))
                    ent.bind("<KeyRelease>", lambda _e, s=sheet, rr=i, cc=j: self._mplane_on_cell_keyrelease(s, rr, cc))
                    row_entries.append(ent)
                    self._mplane_cell_index[sheet][ent] = (i, j)
                self._mplane_table_vars[sheet]["vars"].append(row_vars)
                self._mplane_table_vars[sheet]["entries"].append(row_entries)
            # Responsive column sizing in direct-grid mode.
            # Keep a very small minimum so tables do not hard-clip on narrow windows.
            inner.columnconfigure(0, weight=0, minsize=20)
            for j in range(1, len(headers) + 1):
                inner.columnconfigure(j, weight=1, minsize=6, uniform=f"mplane_{sheet}")
            self._dbg_mplane(
                f"{sheet}: rendered var_rows={len(self._mplane_table_vars[sheet]['vars'])} "
                f"inner_w={inner.winfo_width()}"
            )
        self._mplane_apply_off_row_styles()
        self.bind_all("<B1-Motion>", self._mplane_on_global_drag, add="+")
        self.bind_all("<ButtonRelease-1>", self._mplane_on_global_release, add="+")
        self.bind_all("<Control-c>", self._mplane_on_global_copy, add="+")
        self.bind_all("<Control-C>", self._mplane_on_global_copy, add="+")
        self.bind_all("<Control-v>", self._mplane_on_global_paste, add="+")
        self.bind_all("<Control-V>", self._mplane_on_global_paste, add="+")
        canv = self.mplane_scroll_canvas
        if canv is not None:
            try:
                canv.update_idletasks()
                canv.configure(scrollregion=canv.bbox("all"))
                canv.yview_moveto(0)
            except Exception:
                pass
        self._dbg_mplane("render_done")

    def _mplane_apply_off_row_styles(self) -> None:
        """When CC is OFF, lock and shade corresponding table rows."""
        for sheet in ("PDSCH", "PUSCH", "PRACH"):
            meta = self._mplane_table_vars.get(sheet) or {}
            entries: list[list[tk.Widget]] = meta.get("entries") or []
            c_comp = self._mplane_col_index(sheet, "compression")
            for r_idx, row_entries in enumerate(entries):
                is_off = r_idx < len(self.mplane_cc_on_vars) and (not self.mplane_cc_on_vars[r_idx].get())
                iq16 = self._mplane_row_iq_is_16(sheet, r_idx)
                for c_idx, ent in enumerate(row_entries):
                    if is_off:
                        self._mplane_set_widget_state(ent, "readonly", "#e5e7eb")
                    else:
                        if c_idx == 0:
                            self._mplane_set_widget_state(ent, "readonly", "#eef2ff")
                        else:
                            if c_comp is not None and c_idx == c_comp and iq16:
                                self._mplane_set_widget_state(ent, "readonly", "#e5e7eb")
                            else:
                                self._mplane_set_widget_state(ent, "normal", "#ffffff")
        self._mplane_refresh_selection_visuals()

    def _mplane_on_cell_press(self, event: tk.Event, sheet: str, row: int, col: int) -> None:
        self._mplane_drag_selecting = True
        self._mplane_selection_sheet = sheet
        self._mplane_selection_anchor = (row, col)
        self._mplane_select_range(sheet, row, col, row, col)
        try:
            w = event.widget
            if isinstance(w, tk.Entry):
                w.focus_set()
                w.icursor("end")
        except Exception:
            pass

    def _mplane_on_cell_drag(self, _event: tk.Event, sheet: str, row: int, col: int) -> None:
        anchor = self._mplane_selection_anchor
        if not anchor or self._mplane_selection_sheet != sheet:
            return
        self._mplane_select_range(sheet, anchor[0], anchor[1], row, col)

    def _mplane_on_cell_enter(self, event: tk.Event, sheet: str, row: int, col: int) -> None:
        # While mouse button is pressed, extending selection must work across cell boundaries.
        if not self._mplane_drag_selecting:
            return
        if int(getattr(event, "state", 0)) & 0x0100 == 0:
            return
        anchor = self._mplane_selection_anchor
        if not anchor or self._mplane_selection_sheet != sheet:
            return
        self._mplane_select_range(sheet, anchor[0], anchor[1], row, col)

    def _mplane_on_cell_release(self, _event: tk.Event) -> None:
        self._mplane_drag_selecting = False

    def _mplane_on_global_drag(self, _event: tk.Event) -> None:
        """Global drag tracking so selection continues across entry boundaries."""
        if not self._mplane_drag_selecting:
            return
        try:
            ptr_x, ptr_y = self.winfo_pointerx(), self.winfo_pointery()
            target = self.winfo_containing(ptr_x, ptr_y)
        except Exception:
            # Combobox popdown can raise KeyError("popdown") in winfo_containing.
            return
        if target is None:
            return
        for sheet, mapping in self._mplane_cell_index.items():
            cell = target
            while cell is not None:
                if cell in mapping:
                    anchor = self._mplane_selection_anchor
                    if not anchor or self._mplane_selection_sheet != sheet:
                        return
                    rr, cc = mapping[cell]
                    self._mplane_select_range(sheet, anchor[0], anchor[1], rr, cc)
                    return
                parent = getattr(cell, "master", None)
                if parent is None:
                    break
                cell = parent

    def _mplane_on_global_release(self, _event: tk.Event) -> None:
        self._mplane_drag_selecting = False

    def _mplane_on_global_copy(self, event: tk.Event) -> str | None:
        w = event.widget
        if isinstance(w, tk.Text) and any(tab.get("text") is w for tab in self.message_tabs):
            return None
        sheet = self._mplane_selection_sheet
        if not sheet or not self._mplane_selection_cells:
            return
        return self._mplane_on_copy(event, sheet)

    def _mplane_on_global_paste(self, event: tk.Event) -> str | None:
        w = event.widget
        if isinstance(w, tk.Text) and any(tab.get("text") is w for tab in self.message_tabs):
            return None
        sheet = self._mplane_selection_sheet
        if not sheet:
            return
        return self._mplane_on_paste(event, sheet)

    def _open_mplane_find_panel(self, sheet: str, focus_replace: bool = False) -> None:
        self._mplane_find_sheet = sheet
        self._mplane_find_hits = []
        self._mplane_find_idx = -1
        if self.mplane_find_window is None or not self.mplane_find_window.winfo_exists():
            win = tk.Toplevel(self)
            win.title("M-Plane Find / Replace")
            win.geometry("520x180")
            self.mplane_find_window = win

            top = ttk.Frame(win, padding=10)
            top.pack(fill="both", expand=True)
            ttk.Label(top, text="Find").grid(row=0, column=0, padx=(0, 8), pady=6, sticky="w")
            find_entry = ttk.Entry(top, textvariable=self.find_var)
            find_entry.grid(row=0, column=1, padx=(0, 8), pady=6, sticky="we")
            ttk.Label(top, text="Replace").grid(row=1, column=0, padx=(0, 8), pady=6, sticky="w")
            replace_entry = ttk.Entry(top, textvariable=self.replace_var)
            replace_entry.grid(row=1, column=1, padx=(0, 8), pady=6, sticky="we")
            ttk.Checkbutton(top, text="Ignore case", variable=self.find_ignore_case_var).grid(
                row=0, column=2, padx=6, pady=6, sticky="w"
            )
            ttk.Button(top, text="Find Next", command=self._mplane_find_next).grid(row=0, column=3, padx=4, pady=6)
            ttk.Button(top, text="Replace Current", command=self._mplane_replace_current).grid(row=1, column=2, padx=4, pady=6)
            ttk.Button(top, text="Replace All", command=self._mplane_replace_all).grid(row=1, column=3, padx=4, pady=6)
            top.columnconfigure(1, weight=1)

            def _on_close() -> None:
                self.mplane_find_window = None
                self._mplane_find_hits = []
                self._mplane_find_idx = -1
                win.destroy()

            win.protocol("WM_DELETE_WINDOW", _on_close)
        else:
            self.mplane_find_window.deiconify()
            self.mplane_find_window.lift()
        if focus_replace:
            self.mplane_find_window.focus_force()

    def _mplane_collect_find_hits(self) -> list[tuple[int, int, int, int]]:
        sheet = self._mplane_find_sheet
        if not sheet:
            return []
        keyword = self.find_var.get()
        if not keyword:
            return []
        nocase = bool(self.find_ignore_case_var.get())
        key_cmp = keyword.casefold() if nocase else keyword
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        hits: list[tuple[int, int, int, int]] = []
        for rr, row in enumerate(var_rows):
            for cc, sv in enumerate(row):
                if cc == 0:
                    continue
                txt = sv.get() or ""
                src = txt.casefold() if nocase else txt
                pos = src.find(key_cmp)
                if pos >= 0:
                    hits.append((rr, cc, pos, pos + len(keyword)))
        return hits

    def _mplane_focus_hit(self, hit: tuple[int, int, int, int]) -> None:
        sheet = self._mplane_find_sheet
        if not sheet:
            return
        rr, cc, s_idx, e_idx = hit
        self._mplane_select_range(sheet, rr, cc, rr, cc)
        meta = self._mplane_table_vars.get(sheet) or {}
        entries: list[list[tk.Widget]] = meta.get("entries") or []
        if rr >= len(entries) or cc >= len(entries[rr]):
            return
        ent = entries[rr][cc]
        try:
            ent.focus_set()
            if hasattr(ent, "icursor"):
                ent.icursor(e_idx)  # type: ignore[attr-defined]
            if hasattr(ent, "selection_clear"):
                ent.selection_clear()  # type: ignore[attr-defined]
            if hasattr(ent, "selection_range"):
                ent.selection_range(s_idx, e_idx)  # type: ignore[attr-defined]
        except Exception:
            pass

    def _mplane_find_next(self) -> None:
        hits = self._mplane_collect_find_hits()
        self._mplane_find_hits = hits
        if not hits:
            self.status_var.set("M-Plane Find: no matches")
            return
        self._mplane_find_idx = (self._mplane_find_idx + 1) % len(hits)
        self._mplane_focus_hit(hits[self._mplane_find_idx])
        self.status_var.set(f"M-Plane Find: {self._mplane_find_idx + 1}/{len(hits)}")

    def _mplane_replace_current(self) -> None:
        sheet = self._mplane_find_sheet
        if not sheet:
            return
        hits = self._mplane_collect_find_hits()
        if not hits:
            self.status_var.set("M-Plane Replace: no matches")
            return
        idx = self._mplane_find_idx if 0 <= self._mplane_find_idx < len(hits) else 0
        rr, cc, _s_idx, _e_idx = hits[idx]
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        if rr >= len(var_rows) or cc >= len(var_rows[rr]):
            return
        before = self._mplane_snapshot_values(sheet)
        keyword = self.find_var.get()
        repl = self.replace_var.get()
        nocase = bool(self.find_ignore_case_var.get())
        text = var_rows[rr][cc].get()
        if nocase:
            pattern = re.compile(re.escape(keyword), re.IGNORECASE)
            new_text, n = pattern.subn(repl, text, count=1)
        else:
            new_text = text.replace(keyword, repl, 1)
            n = 1 if new_text != text else 0
        if n > 0:
            var_rows[rr][cc].set(new_text)
            self._mplane_push_undo(sheet, before)
        self._mplane_find_hits = []
        self._mplane_find_idx = -1
        self._mplane_find_next()

    def _mplane_replace_all(self) -> None:
        sheet = self._mplane_find_sheet
        if not sheet:
            return
        keyword = self.find_var.get()
        if not keyword:
            return
        repl = self.replace_var.get()
        nocase = bool(self.find_ignore_case_var.get())
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        before = self._mplane_snapshot_values(sheet)
        changed = 0
        for rr, row in enumerate(var_rows):
            for cc, sv in enumerate(row):
                if cc == 0:
                    continue
                txt = sv.get() or ""
                if nocase:
                    pattern = re.compile(re.escape(keyword), re.IGNORECASE)
                    new_txt, n = pattern.subn(repl, txt)
                else:
                    n = txt.count(keyword)
                    new_txt = txt.replace(keyword, repl)
                if n > 0:
                    sv.set(new_txt)
                    changed += n
        if changed:
            self._mplane_push_undo(sheet, before)
        self.status_var.set(f"M-Plane Replace All: {changed} replacement(s)")

    def _mplane_select_range(self, sheet: str, r1: int, c1: int, r2: int, c2: int) -> None:
        self._mplane_selection_sheet = sheet
        r_min, r_max = sorted((r1, r2))
        c_min, c_max = sorted((c1, c2))
        new_sel: set[tuple[int, int]] = set()
        for rr in range(r_min, r_max + 1):
            for cc in range(c_min, c_max + 1):
                new_sel.add((rr, cc))
        if self._mplane_selection_sheet == self._mplane_prev_selection_sheet and new_sel == self._mplane_selection_cells:
            return
        self._mplane_selection_cells = new_sel
        self._mplane_refresh_selection_visuals()

    def _mplane_refresh_selection_visuals(self) -> None:
        old_sheet = self._mplane_prev_selection_sheet
        old_cells = self._mplane_prev_selection_cells
        new_sheet = self._mplane_selection_sheet
        new_cells = set(self._mplane_selection_cells)

        # Fast path: update only changed cells when staying on same sheet.
        if old_sheet == new_sheet and new_sheet is not None:
            to_clear = old_cells - new_cells
            to_set = new_cells - old_cells
            for rr, cc in to_clear:
                self._mplane_style_cell(new_sheet, rr, cc, selected=False)
            for rr, cc in to_set:
                self._mplane_style_cell(new_sheet, rr, cc, selected=True)
        else:
            # Sheet changed (or initial render): repaint previous + current selection sets only.
            if old_sheet is not None:
                for rr, cc in old_cells:
                    self._mplane_style_cell(old_sheet, rr, cc, selected=False)
            if new_sheet is not None:
                for rr, cc in new_cells:
                    self._mplane_style_cell(new_sheet, rr, cc, selected=True)

        self._mplane_prev_selection_sheet = new_sheet
        self._mplane_prev_selection_cells = new_cells

    def _mplane_style_cell(self, sheet: str, r_idx: int, c_idx: int, selected: bool) -> None:
        meta = self._mplane_table_vars.get(sheet) or {}
        entries: list[list[tk.Widget]] = meta.get("entries") or []
        if r_idx >= len(entries) or c_idx >= len(entries[r_idx]):
            return
        ent = entries[r_idx][c_idx]
        is_off = r_idx < len(self.mplane_cc_on_vars) and (not self.mplane_cc_on_vars[r_idx].get())
        c_comp = self._mplane_col_index(sheet, "compression")
        iq16 = self._mplane_row_iq_is_16(sheet, r_idx)
        if selected:
            self._mplane_set_widget_state(ent, "normal", "#bfdbfe")
            return
        if is_off:
            self._mplane_set_widget_state(ent, "readonly", "#e5e7eb")
        elif c_idx == 0:
            self._mplane_set_widget_state(ent, "readonly", "#eef2ff")
        elif c_comp is not None and c_idx == c_comp and iq16:
            self._mplane_set_widget_state(ent, "readonly", "#e5e7eb")
        else:
            self._mplane_set_widget_state(ent, "normal", "#ffffff")

    def _mplane_set_widget_state(self, widget: tk.Widget, mode: str, bg: str) -> None:
        """Best-effort state/style setter for Entry and Combobox widgets."""
        if isinstance(widget, ttk.Combobox):
            try:
                if mode == "readonly":
                    widget.configure(state="disabled", style="MPlaneDis.TCombobox")
                else:
                    widget.configure(
                        state="normal",
                        style="MPlaneSel.TCombobox" if bg == "#bfdbfe" else "MPlane.TCombobox",
                    )
            except Exception:
                pass
            return
        try:
            if mode == "readonly":
                widget.configure(state="readonly", readonlybackground=bg, fg="#111827", disabledforeground="#111827")
            else:
                widget.configure(state="normal", bg=bg, fg="#111827", disabledforeground="#111827")
        except Exception:
            pass

    def _mplane_on_paste(self, event: tk.Event, sheet: str) -> str:
        before = self._mplane_snapshot_values(sheet)
        try:
            clip = self.clipboard_get()
        except Exception:
            return "break"
        if not clip:
            return "break"
        rows_in = [ln for ln in clip.replace("\r\n", "\n").replace("\r", "\n").split("\n") if ln != ""]
        if not rows_in:
            return "break"
        matrix = [line.split("\t") for line in rows_in]

        # Excel-like: if clipboard is single cell and a range is selected,
        # fill the entire selection.
        if (
            self._mplane_selection_sheet == sheet
            and len(self._mplane_selection_cells) > 1
            and len(matrix) == 1
            and len(matrix[0]) == 1
        ):
            meta = self._mplane_table_vars.get(sheet) or {}
            var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
            entries: list[list[tk.Widget]] = meta.get("entries") or []
            val0 = matrix[0][0]
            for rr, cc in sorted(self._mplane_selection_cells):
                if rr >= len(var_rows) or cc >= len(var_rows[rr]) or cc == 0:
                    continue
                if rr < len(self.mplane_cc_on_vars) and (not self.mplane_cc_on_vars[rr].get()):
                    continue
                ent = entries[rr][cc] if rr < len(entries) and cc < len(entries[rr]) else None
                if ent is not None and str(ent.cget("state")) in {"readonly", "disabled"}:
                    continue
                vset = val0
                if self._mplane_is_compression_col(sheet, cc):
                    if self._mplane_row_iq_is_16(sheet, rr):
                        continue
                    norm = self._mplane_normalize_compression_value(vset)
                    if norm is None:
                        continue
                    vset = norm
                var_rows[rr][cc].set(vset)
                self._mplane_last_valid_cell_values[(sheet, rr, cc)] = vset
            self._mplane_push_undo(sheet, before)
            return "break"

        start_row, start_col = 0, 0
        if self._mplane_selection_sheet == sheet and self._mplane_selection_cells:
            start_row = min(r for r, _ in self._mplane_selection_cells)
            start_col = min(c for _, c in self._mplane_selection_cells)
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        entries: list[list[tk.Widget]] = meta.get("entries") or []
        for dr, vals in enumerate(matrix):
            rr = start_row + dr
            if rr >= len(var_rows):
                break
            for dc, val in enumerate(vals):
                cc = start_col + dc
                if cc >= len(var_rows[rr]):
                    break
                if cc == 0:
                    continue
                if rr < len(self.mplane_cc_on_vars) and (not self.mplane_cc_on_vars[rr].get()):
                    continue
                ent = entries[rr][cc] if rr < len(entries) and cc < len(entries[rr]) else None
                if ent is not None and str(ent.cget("state")) == "readonly":
                    continue
                if self._mplane_is_compression_col(sheet, cc):
                    if self._mplane_row_iq_is_16(sheet, rr):
                        continue
                    norm = self._mplane_normalize_compression_value(val)
                    if norm is None:
                        # Invalid pasted compression value: ignore and keep prior value.
                        continue
                    val = norm
                var_rows[rr][cc].set(val)
                self._mplane_last_valid_cell_values[(sheet, rr, cc)] = val
        end_row = min(len(var_rows) - 1, start_row + len(matrix) - 1)
        end_col = start_col
        for vals in matrix:
            end_col = max(end_col, start_col + max(0, len(vals) - 1))
        if var_rows:
            end_col = min(len(var_rows[0]) - 1, end_col)
        self._mplane_select_range(sheet, start_row, start_col, end_row, end_col)
        self._mplane_push_undo(sheet, before)
        return "break"

    def _mplane_on_multi_clear(self, event: tk.Event, sheet: str) -> str | None:
        if self._mplane_selection_sheet != sheet or len(self._mplane_selection_cells) <= 1:
            return
        before = self._mplane_snapshot_values(sheet)
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        for rr, cc in sorted(self._mplane_selection_cells):
            if rr >= len(var_rows) or cc >= len(var_rows[rr]) or cc == 0:
                continue
            if rr < len(self.mplane_cc_on_vars) and (not self.mplane_cc_on_vars[rr].get()):
                continue
            var_rows[rr][cc].set("")
        self._mplane_push_undo(sheet, before)
        return "break"

    def _mplane_on_cell_var_changed(self, sheet: str, src_r: int, src_c: int) -> None:
        """Mirror active-cell edits to the full selected range."""
        if self._mplane_sync_lock:
            return
        meta0 = self._mplane_table_vars.get(sheet) or {}
        var_rows0: list[list[tk.StringVar]] = meta0.get("vars") or []
        if src_r < len(var_rows0) and src_c < len(var_rows0[src_r]):
            cur_val = var_rows0[src_r][src_c].get()
            if self._mplane_is_compression_col(sheet, src_c):
                if self._mplane_row_iq_is_16(sheet, src_r):
                    prev = self._mplane_last_valid_cell_values.get((sheet, src_r, src_c), "")
                    self._mplane_sync_lock = True
                    try:
                        var_rows0[src_r][src_c].set(prev)
                    finally:
                        self._mplane_sync_lock = False
                    return
                norm = self._mplane_normalize_compression_value(cur_val)
                if norm is not None:
                    if norm != cur_val:
                        self._mplane_sync_lock = True
                        try:
                            var_rows0[src_r][src_c].set(norm)
                        finally:
                            self._mplane_sync_lock = False
                    self._mplane_last_valid_cell_values[(sheet, src_r, src_c)] = norm
                # invalid typed text is reverted on FocusOut; keep editing flow.
            else:
                self._mplane_last_valid_cell_values[(sheet, src_r, src_c)] = cur_val
                headers0: list[str] = meta0.get("headers") or []
                if src_c < len(headers0) and self._mplane_tagify_header(headers0[src_c]) == "iq-bitwidth":
                    c_comp = self._mplane_col_index(sheet, "compression")
                    if cur_val.strip() == "16":
                        if c_comp is not None and src_r < len(var_rows0) and c_comp < len(var_rows0[src_r]):
                            self._mplane_sync_lock = True
                            try:
                                var_rows0[src_r][c_comp].set("no comp")
                            finally:
                                self._mplane_sync_lock = False
                            self._mplane_last_valid_cell_values[(sheet, src_r, c_comp)] = "no comp"
                    else:
                        if c_comp is not None and src_r < len(var_rows0) and c_comp < len(var_rows0[src_r]):
                            cv = (var_rows0[src_r][c_comp].get() or "").strip().lower()
                            if cv in {"", "no comp", "nocomp", "static"}:
                                self._mplane_sync_lock = True
                                try:
                                    var_rows0[src_r][c_comp].set("BFP")
                                finally:
                                    self._mplane_sync_lock = False
                                self._mplane_last_valid_cell_values[(sheet, src_r, c_comp)] = "BFP"
                    self._mplane_apply_off_row_styles()
        if self._mplane_selection_sheet != sheet or len(self._mplane_selection_cells) <= 1:
            return
        if (src_r, src_c) not in self._mplane_selection_cells:
            return
        if src_c == 0:
            return
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        if src_r >= len(var_rows) or src_c >= len(var_rows[src_r]):
            return
        value = var_rows[src_r][src_c].get()
        self._mplane_sync_lock = True
        try:
            for rr, cc in sorted(self._mplane_selection_cells):
                if (rr, cc) == (src_r, src_c):
                    continue
                if rr >= len(var_rows) or cc >= len(var_rows[rr]) or cc == 0:
                    continue
                if rr < len(self.mplane_cc_on_vars) and (not self.mplane_cc_on_vars[rr].get()):
                    continue
                if self._mplane_is_compression_col(sheet, cc):
                    norm = self._mplane_normalize_compression_value(value)
                    if norm is None:
                        continue
                    var_rows[rr][cc].set(norm)
                    self._mplane_last_valid_cell_values[(sheet, rr, cc)] = norm
                    continue
                var_rows[rr][cc].set(value)
                self._mplane_last_valid_cell_values[(sheet, rr, cc)] = value
        finally:
            self._mplane_sync_lock = False

    def _mplane_on_cell_focus_out(self, sheet: str, row: int, col: int) -> None:
        # Re-evaluate row lock state after editor commit.
        if self._mplane_is_iq_col(sheet, col):
            self.after_idle(self._mplane_apply_off_row_styles)
        if not self._mplane_is_compression_col(sheet, col):
            return
        if self._mplane_row_iq_is_16(sheet, row):
            return
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        if row >= len(var_rows) or col >= len(var_rows[row]):
            return
        val = var_rows[row][col].get()
        norm = self._mplane_normalize_compression_value(val)
        if norm is None:
            prev = self._mplane_last_valid_cell_values.get((sheet, row, col), "")
            self._mplane_sync_lock = True
            try:
                var_rows[row][col].set(prev)
            finally:
                self._mplane_sync_lock = False
            return
        if norm != val:
            self._mplane_sync_lock = True
            try:
                var_rows[row][col].set(norm)
            finally:
                self._mplane_sync_lock = False
        self._mplane_last_valid_cell_values[(sheet, row, col)] = norm

    def _mplane_on_cell_keyrelease(self, sheet: str, row: int, col: int) -> None:
        # Keep compression lock/unlock responsive while typing iq-bitwidth.
        if self._mplane_is_iq_col(sheet, col):
            self.after_idle(self._mplane_apply_off_row_styles)

    def _mplane_is_compression_col(self, sheet: str, col: int) -> bool:
        meta = self._mplane_table_vars.get(sheet) or {}
        headers: list[str] = meta.get("headers") or []
        if col < 0 or col >= len(headers):
            return False
        return self._mplane_tagify_header(headers[col]) == "compression"

    def _mplane_is_iq_col(self, sheet: str, col: int) -> bool:
        meta = self._mplane_table_vars.get(sheet) or {}
        headers: list[str] = meta.get("headers") or []
        if col < 0 or col >= len(headers):
            return False
        return self._mplane_tagify_header(headers[col]) == "iq-bitwidth"

    @staticmethod
    def _mplane_normalize_compression_value(v: str) -> str | None:
        s = (v or "").strip()
        if not s:
            return ""
        t = s.lower().replace("_", " ").replace("-", " ").strip()
        if t in {"bfp", "block floating point"}:
            return "BFP"
        if re.match(r"^exponent(\s*[:=]\s*\d+)?$", t):
            return "exponent"
        return None

    def _mplane_col_index(self, sheet: str, header_key: str) -> int | None:
        meta = self._mplane_table_vars.get(sheet) or {}
        headers: list[str] = meta.get("headers") or []
        want = self._mplane_tagify_header(header_key)
        for i, h in enumerate(headers):
            if self._mplane_tagify_header(h) == want:
                return i
        return None

    def _mplane_row_iq_is_16(self, sheet: str, row: int) -> bool:
        c_iq = self._mplane_col_index(sheet, "iq-bitwidth")
        if c_iq is None:
            return False
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        if row >= len(var_rows) or c_iq >= len(var_rows[row]):
            return False
        s = (var_rows[row][c_iq].get() or "").strip()
        if s in {"16", "16.0"}:
            return True
        try:
            return float(s) == 16.0
        except Exception:
            return False

    def _mplane_on_copy(self, _event: tk.Event, sheet: str) -> str | None:
        if self._mplane_selection_sheet != sheet or not self._mplane_selection_cells:
            return
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        if not var_rows:
            return "break"
        r0 = min(r for r, _ in self._mplane_selection_cells)
        r1 = max(r for r, _ in self._mplane_selection_cells)
        c0 = min(c for _, c in self._mplane_selection_cells)
        c1 = max(c for _, c in self._mplane_selection_cells)
        out_lines: list[str] = []
        for rr in range(r0, r1 + 1):
            vals: list[str] = []
            for cc in range(c0, c1 + 1):
                if rr < len(var_rows) and cc < len(var_rows[rr]):
                    vals.append(var_rows[rr][cc].get())
                else:
                    vals.append("")
            out_lines.append("\t".join(vals))
        text = "\n".join(out_lines)
        try:
            self.clipboard_clear()
            self.clipboard_append(text)
        except Exception:
            pass
        return "break"

    def _mplane_on_undo(self, _event: tk.Event, sheet: str) -> str:
        stack = self._mplane_undo_stack.get(sheet) or []
        if not stack:
            return "break"
        current = self._mplane_snapshot_values(sheet)
        snap = stack.pop()
        self._mplane_redo_stack.setdefault(sheet, []).append(current)
        self._mplane_restore_snapshot(sheet, snap)
        return "break"

    def _mplane_on_redo(self, _event: tk.Event, sheet: str) -> str:
        stack = self._mplane_redo_stack.get(sheet) or []
        if not stack:
            return "break"
        current = self._mplane_snapshot_values(sheet)
        snap = stack.pop()
        self._mplane_undo_stack.setdefault(sheet, []).append(current)
        self._mplane_restore_snapshot(sheet, snap)
        return "break"

    def _mplane_snapshot_values(self, sheet: str) -> dict[tuple[int, int], str]:
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        snap: dict[tuple[int, int], str] = {}
        for rr, row in enumerate(var_rows):
            for cc, sv in enumerate(row):
                snap[(rr, cc)] = sv.get()
        return snap

    def _mplane_restore_snapshot(self, sheet: str, snap: dict[tuple[int, int], str]) -> None:
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        for (rr, cc), val in snap.items():
            if rr < len(var_rows) and cc < len(var_rows[rr]):
                var_rows[rr][cc].set(val)

    def _mplane_push_undo(self, sheet: str, before: dict[tuple[int, int], str]) -> None:
        after = self._mplane_snapshot_values(sheet)
        if before == after:
            return
        stack = self._mplane_undo_stack.setdefault(sheet, [])
        stack.append(before)
        if len(stack) > 100:
            del stack[0 : len(stack) - 100]
        self._mplane_redo_stack[sheet] = []

    @staticmethod
    def _mplane_tagify_header(h: str) -> str:
        s = (h or "").strip().lower()
        s = re.sub(r"\s+", "-", s)
        s = re.sub(r"[^a-z0-9\\-_/()]+", "", s)
        return s

    def _load_mplane_workbook_from_disk(
        self, path: str, mp: Any
    ) -> tuple[
        dict[str, str],
        dict[str, str],
        dict[str, Any],
        list[dict[str, Any]],
        dict[str, tuple[list[str], list[list[str]]]],
        list[str],
        Any,
    ]:
        rpc, baselines, merged, cc_rows, tables, warnings = mp.load_workbook_payloads(path)
        return rpc, baselines, merged, cc_rows, tables, warnings, mp

    @staticmethod
    def _mplane_apply_tables_from_workbook(
        rpc: dict[str, str],
        tables: dict[str, tuple[list[str], list[list[str]]]],
        mp: Any,
    ) -> list[str]:
        """Merge Control-Sheet tables from a workbook into RPC XML (no GUI table vars)."""
        warns: list[str] = []
        for sheet in ("PDSCH", "PUSCH", "PRACH"):
            headers, rows = tables.get(sheet, ([], []))
            xml = (rpc.get(sheet) or "").strip()
            if not headers or not xml or not rows:
                continue
            if "low-level-tx-endpoints" in xml or "low-level-rx-endpoints" in xml:
                new_xml, ws = mp.apply_acorn_control_details_to_rpc(xml, sheet, headers, rows)
            else:
                new_xml, ws = mp.apply_full_table_to_rpc(xml, sheet, headers, rows)
            rpc[sheet] = new_xml
            warns.extend(ws)
        for sheet in ("PDSCH", "PUSCH", "PRACH"):
            body = (rpc.get(sheet) or "").strip()
            if body:
                rpc[sheet] = mp.normalize_uplane_rpc_element_order(body, sheet)
        return warns

    @staticmethod
    def _mplane_cc_off_rows_from_workbook(cc_rows: list[dict[str, Any]]) -> list[int]:
        return [
            i + 1
            for i, row in enumerate(cc_rows)
            if not CallhomeGUI._mplane_enabled_from_cell(row.get("enabled"))
        ]

    def _mplane_apply_full_tables_to_rpc(self, rpc: dict[str, str]) -> list[str]:
        """Apply edited full-table values into RPC xml strings by nth-occurrence replacement."""
        warns: list[str] = []
        try:
            import mplane_control as mp
        except Exception as exc:
            return [f"mplane_control missing: {exc}"]

        for sheet in ("PDSCH", "PUSCH", "PRACH"):
            meta = self._mplane_table_vars.get(sheet) or {}
            headers: list[str] = meta.get("headers") or []
            var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
            xml = (rpc.get(sheet) or "").strip()
            if not headers or not xml or not var_rows:
                headers, rows_fb = self._mplane_tables.get(sheet, ([], []))
                if headers and rows_fb and xml:
                    if "low-level-tx-endpoints" in xml or "low-level-rx-endpoints" in xml:
                        new_xml, ws = mp.apply_acorn_control_details_to_rpc(xml, sheet, headers, rows_fb)
                    else:
                        new_xml, ws = mp.apply_full_table_to_rpc(xml, sheet, headers, rows_fb)
                    rpc[sheet] = new_xml
                    warns.extend(ws)
                continue
            rows: list[list[str]] = []
            for r in var_rows:
                rows.append([(v.get() or "") for v in r])
            if "low-level-tx-endpoints" in xml or "low-level-rx-endpoints" in xml:
                new_xml, ws = mp.apply_acorn_control_details_to_rpc(xml, sheet, headers, rows)
            else:
                new_xml, ws = mp.apply_full_table_to_rpc(xml, sheet, headers, rows)
            rpc[sheet] = new_xml
            warns.extend(ws)

        for sheet in ("PDSCH", "PUSCH", "PRACH"):
            body = (rpc.get(sheet) or "").strip()
            if body:
                rpc[sheet] = mp.normalize_uplane_rpc_element_order(body, sheet)

        return warns

    def _mplane_collect_live(self) -> dict[str, str]:
        return {k: (v.get() or "").strip() for k, v in self.mplane_fields.items()}

    _MPLANE_LIVE_KEYS: tuple[str, ...] = (
        "cu_if_name",
        "cu_base_if",
        "cu_vlan",
        "cu_mac",
        "odu_mac",
        "ru_mac_pe",
        "pe_name",
    )

    def _mplane_gui_state_snapshot(self) -> dict[str, Any]:
        """Capture M-Plane Control tab state (main thread only)."""
        self._mplane_sync_visible_widget_values()
        tables = self._mplane_gui_tables_snapshot()
        has_tables = any(bool(v[0]) and bool(v[1]) for v in tables.values())
        return {
            "has_data": bool(self._mplane_rpc_raw) and has_tables,
            "tables": tables,
            "live": self._mplane_collect_live(),
            "off_rows": [i + 1 for i, v in enumerate(self.mplane_cc_on_vars) if not v.get()],
        }

    def _mplane_build_rpc_from_gui(
        self,
        mp: Any,
        *,
        rpc_raw: dict[str, str] | None = None,
        baselines: dict[str, str] | None = None,
        gui_tables: dict[str, tuple[list[str], list[list[str]]]] | None = None,
        gui_live: dict[str, str] | None = None,
        gui_off_rows: list[int] | None = None,
        comment_off_rows: bool = True,
    ) -> tuple[dict[str, str], list[str]]:
        """
        Build final RPC payloads: in-memory workbook templates + M-Plane Control GUI edits.
        Used by Save, Apply, and Conformance.
        """
        warns: list[str] = []
        rpc = dict(rpc_raw if rpc_raw is not None else self._mplane_rpc_raw)
        if not rpc:
            raise RuntimeError("No M-Plane RPC loaded. Click Reload from Excel first.")
        baselines_in = dict(baselines if baselines is not None else self._mplane_baselines)
        tables = gui_tables if gui_tables is not None else self._mplane_gui_tables_snapshot()
        live_src = gui_live if gui_live is not None else self._mplane_collect_live()
        live = {k: (live_src.get(k) or "").strip() for k in self._MPLANE_LIVE_KEYS}
        off_rows = (
            list(gui_off_rows)
            if gui_off_rows is not None
            else [i + 1 for i, v in enumerate(self.mplane_cc_on_vars) if not v.get()]
        )

        for sheet in ("PDSCH", "PUSCH", "PRACH"):
            rpc[sheet] = mp.uncomment_endpoint_rows((rpc.get(sheet) or ""), sheet)
        rpc["ACTIVE"] = mp.uncomment_active_rows((rpc.get("ACTIVE") or ""))

        for name in mp.SEND_ORDER:
            body = (rpc.get(name) or "").strip()
            if not body:
                continue
            rpc[name] = mp.apply_global_baselines(body, baselines_in, live)
            if name == "CUplane-interface":
                rpc[name] = mp.ensure_cuplane_interface_fields(rpc[name], live)
            elif name == "Processing-element":
                rpc[name] = mp.ensure_processing_element_fields(rpc[name], live)

        table_warns = self._mplane_apply_tables_from_workbook(rpc, tables, mp)
        warns.extend(table_warns)

        pusch_xml = (rpc.get("PUSCH") or "").strip()
        prach_xml = (rpc.get("PRACH") or "").strip()
        if pusch_xml and prach_xml:
            prach_xml, pr_warns = mp.omit_prach_rx_endpoints_present_in_pusch(prach_xml, pusch_xml)
            rpc["PRACH"] = prach_xml
            warns.extend(pr_warns)

        active_body, act_warns = mp.sync_active_carrier_names_from_tables(
            rpc.get("ACTIVE", ""),
            tables.get("PDSCH", ([], [])),
            tables.get("PUSCH", ([], [])),
        )
        rpc["ACTIVE"] = active_body
        warns.extend(act_warns)

        if comment_off_rows and off_rows:
            for sheet in ("PDSCH", "PUSCH", "PRACH"):
                rpc[sheet] = mp.comment_out_endpoint_rows((rpc.get(sheet) or ""), sheet, off_rows)
            rpc["ACTIVE"] = mp.comment_out_active_rows((rpc.get("ACTIVE") or ""), off_rows)

        return rpc, warns

    def _mplane_gui_tables_snapshot(self) -> dict[str, tuple[list[str], list[list[str]]]]:
        """Capture current PDSCH/PUSCH/PRACH grid edits (main thread only)."""
        out: dict[str, tuple[list[str], list[list[str]]]] = {}
        for sheet in ("PDSCH", "PUSCH", "PRACH"):
            meta = self._mplane_table_vars.get(sheet) or {}
            headers: list[str] = meta.get("headers") or []
            vars_rows: list[list[tk.StringVar]] = meta.get("vars") or []
            if not headers or not vars_rows:
                continue
            rows = [[(v.get() or "") for v in r] for r in vars_rows]
            out[sheet] = (headers, rows)
        return out

    def apply_mplane_workbook_once(self) -> None:
        gui_state = self._mplane_gui_state_snapshot()
        path = self._normalize_mplane_workbook_path(self.mplane_xlsx_path.get())
        if gui_state["has_data"]:
            threading.Thread(
                target=self._prepare_mplane_payload_tabs_worker,
                args=(path, gui_state),
                daemon=True,
            ).start()
            return
        if not path:
            messagebox.showwarning("M-Plane", "Reload from Excel first, or choose an .xlsx path.")
            return
        if not os.path.isfile(path):
            messagebox.showwarning("M-Plane", f"Workbook not found:\n{path}")
            return
        threading.Thread(
            target=self._prepare_mplane_payload_tabs_worker,
            args=(path, gui_state),
            daemon=True,
        ).start()

    def _mplane_sync_visible_widget_values(self) -> None:
        """Force-sync current widget text into StringVars before Save/Apply."""
        for sheet in ("PDSCH", "PUSCH", "PRACH"):
            meta = self._mplane_table_vars.get(sheet) or {}
            vars_rows: list[list[tk.StringVar]] = meta.get("vars") or []
            entries_rows: list[list[tk.Widget]] = meta.get("entries") or []
            for r_idx, row_entries in enumerate(entries_rows):
                for c_idx, w in enumerate(row_entries):
                    if r_idx >= len(vars_rows) or c_idx >= len(vars_rows[r_idx]):
                        continue
                    try:
                        cur = w.get() if hasattr(w, "get") else vars_rows[r_idx][c_idx].get()
                    except Exception:
                        cur = vars_rows[r_idx][c_idx].get()
                    if cur != vars_rows[r_idx][c_idx].get():
                        vars_rows[r_idx][c_idx].set(cur)

    def _prepare_mplane_payload_tabs_worker(
        self,
        xlsx_path: str,
        gui_state: dict[str, Any] | None = None,
    ) -> None:
        try:
            import mplane_control as mp
            mp = importlib.reload(mp)
        except ImportError as exc:
            self.after(0, messagebox.showerror, "M-Plane", str(exc))
            return

        apply_btn = self.mplane_apply_btn
        if apply_btn:
            self.after(0, lambda b=apply_btn: b.config(state="disabled"))
        try:
            state = gui_state or {}
            use_gui = bool(state.get("has_data"))
            rpc_raw = dict(self._mplane_rpc_raw)
            baselines = dict(self._mplane_baselines)
            load_warns: list[str] = []

            if not rpc_raw and xlsx_path and os.path.isfile(xlsx_path):
                rpc_raw, baselines, _merged, _cc_rows, _tables, load_warns, mp = self._load_mplane_workbook_from_disk(
                    xlsx_path, mp
                )
            elif not rpc_raw:
                self.after(
                    0,
                    messagebox.showerror,
                    "M-Plane",
                    "No M-Plane data loaded.\n\nClick Reload from Excel first.",
                )
                return

            if use_gui:
                self.after(0, self.append_log, "[GUI] Apply: M-Plane Control tab (GUI) → M-* tabs\n")
                try:
                    rpc, build_warns = self._mplane_build_rpc_from_gui(
                        mp,
                        rpc_raw=rpc_raw,
                        baselines=baselines,
                        gui_tables=state.get("tables"),
                        gui_live=state.get("live"),
                        gui_off_rows=state.get("off_rows"),
                        comment_off_rows=True,
                    )
                except RuntimeError as exc:
                    self.after(0, messagebox.showerror, "M-Plane", str(exc))
                    return
                for w in load_warns + build_warns:
                    self.after(0, self.append_log, f"[GUI] M-Plane Apply: {w}\n")
                off_rows = list(state.get("off_rows") or [])
                if off_rows:
                    self.after(0, self.append_log, f"[GUI] M-Plane OFF rows commented: {off_rows}\n")
            else:
                self.after(0, self.append_log, f"[GUI] Apply: xlsx fallback ({xlsx_path})\n")
                rpc_raw2, baselines2, merged, cc_rows, tables, load_warns, mp = self._load_mplane_workbook_from_disk(
                    xlsx_path, mp
                )
                rpc = dict(rpc_raw2)
                baselines = dict(baselines2)
                for w in load_warns:
                    self.after(0, self.append_log, f"[GUI] M-Plane Apply: {w}\n")
                for sheet in ("PDSCH", "PUSCH", "PRACH"):
                    rpc[sheet] = mp.uncomment_endpoint_rows((rpc.get(sheet) or ""), sheet)
                rpc["ACTIVE"] = mp.uncomment_active_rows((rpc.get("ACTIVE") or ""))
                live = {
                    k: ("" if merged.get(k) is None else str(merged.get(k)).strip())
                    for k in self._MPLANE_LIVE_KEYS
                }
                for name in mp.SEND_ORDER:
                    body = (rpc.get(name) or "").strip()
                    if not body:
                        self.after(0, self.append_log, f"[GUI] M-Plane skip (empty payload): {name}\n")
                        continue
                    rpc[name] = mp.apply_global_baselines(body, baselines, live)
                    if name == "CUplane-interface":
                        rpc[name] = mp.ensure_cuplane_interface_fields(rpc[name], live)
                    elif name == "Processing-element":
                        rpc[name] = mp.ensure_processing_element_fields(rpc[name], live)
                table_warns = self._mplane_apply_tables_from_workbook(rpc, tables, mp)
                for w in table_warns:
                    self.after(0, self.append_log, f"[GUI] M-Plane table: {w}\n")
                pusch_xml = (rpc.get("PUSCH") or "").strip()
                prach_xml = (rpc.get("PRACH") or "").strip()
                if pusch_xml and prach_xml:
                    prach_xml, pr_warns = mp.omit_prach_rx_endpoints_present_in_pusch(prach_xml, pusch_xml)
                    rpc["PRACH"] = prach_xml
                    for w in pr_warns:
                        self.after(0, self.append_log, f"[GUI] M-Plane PRACH: {w}\n")
                off_rows = self._mplane_cc_off_rows_from_workbook(cc_rows)
                if not off_rows:
                    flags = merged.get("_detail_cc_on_flags")
                    if isinstance(flags, list):
                        off_rows = [i + 1 for i, on in enumerate(flags) if not on]
                active_body, act_warns = mp.sync_active_carrier_names_from_tables(
                    rpc.get("ACTIVE", ""),
                    tables.get("PDSCH", ([], [])),
                    tables.get("PUSCH", ([], [])),
                )
                rpc["ACTIVE"] = active_body
                for w in act_warns:
                    self.after(0, self.append_log, f"[GUI] M-Plane ACTIVE: {w}\n")
                if off_rows:
                    for sheet in ("PDSCH", "PUSCH", "PRACH"):
                        rpc[sheet] = mp.comment_out_endpoint_rows((rpc.get(sheet) or ""), sheet, off_rows)
                    rpc["ACTIVE"] = mp.comment_out_active_rows((rpc.get("ACTIVE") or ""), off_rows)
                    self.after(0, self.append_log, f"[GUI] M-Plane OFF rows commented: {off_rows}\n")

            for name in mp.SEND_ORDER:
                if not (rpc.get(name) or "").strip():
                    self.after(0, self.append_log, f"[GUI] M-Plane skip (empty payload): {name}\n")
            self.after(0, self._enqueue_mplane_payload_tabs, rpc, mp.SEND_ORDER)
        finally:
            if apply_btn:
                self.after(0, lambda b=apply_btn: b.config(state="normal"))

    def _enqueue_mplane_payload_tabs(self, rpc: dict[str, str], send_order: list[str]) -> None:
        created = 0
        for name in send_order:
            payload = (rpc.get(name) or "").strip()
            if not payload:
                continue
            payload = self._format_rpc_for_message_tab(name, payload)
            self.add_message_tab(initial_title=f"M-{name}")
            tab = self.message_tabs[-1]
            text_widget = tab["text"]
            text_widget.delete("1.0", "end")
            text_widget.insert("1.0", payload)
            self._highlight_xml(text_widget)
            created += 1
        if created == 0:
            self.append_log("[GUI] M-Plane: no payload to queue.\n")
            return
        # Focus Netconf Client tab and the last created message tab.
        try:
            self.notebook.select(2)  # Settings(0), M-Plane(1), Netconf(2), Conformance(3), Shortcuts(4)
        except Exception:
            pass
        try:
            self.msg_notebook.select(len(self.message_tabs) - 1)
        except Exception:
            pass
        self.append_log(f"[GUI] M-Plane queued {created} payload tab(s). Use Netconf Client send buttons.\n")

    def _conformance_prepare_mplane_bundle(self, fname: str, log_line: Any) -> bool:
        """Conformance M-Plane scripts: prefer M-Plane Control tab (GUI), fall back to xlsx."""
        from conformance_mixin import _CONFORMANCE_MPLANE_SCRIPTS

        if fname not in _CONFORMANCE_MPLANE_SCRIPTS:
            return True
        cache = self._conformance_mplane_bundle_cache()
        cache.pop(fname, None)

        gui_state = self._mplane_gui_state_snapshot()
        if gui_state.get("has_data"):
            try:
                import mplane_control as mp
                import mplane_conformance as mc

                mp = importlib.reload(mp)
                rpc, warns = self._mplane_build_rpc_from_gui(
                    mp,
                    gui_tables=gui_state["tables"],
                    gui_live=gui_state["live"],
                    gui_off_rows=gui_state["off_rows"],
                    comment_off_rows=True,
                )
                merged = dict(self._mplane_merged)
                for k, v in gui_state["live"].items():
                    if v:
                        merged[k] = v
                bundle = mc.prepare_mplane_conformance_bundle_from_gui(
                    rpc,
                    merged=merged,
                    duplicate_eaxc=(fname == "conformance_31102.sh"),
                    warnings=warns,
                )
            except Exception as exc:
                log_line(f"[WARN] M-Plane GUI 빌드 실패, xlsx fallback: {exc}")
            else:
                cache[fname] = bundle
                log_line(f"M-Plane GUI 준비 ({len(bundle.remote_files)} RPC)")
                for w in bundle.warnings[:12]:
                    log_line(f"  [M-Plane] {w}")
                if len(bundle.warnings) > 12:
                    log_line(f"  [M-Plane] … 외 {len(bundle.warnings) - 12}건")
                return True

        return ConformanceMixin._conformance_prepare_mplane_bundle(self, fname, log_line)

    def _format_rpc_for_message_tab(self, name: str, payload: str) -> str:
        """Pretty-format RPC payload for editor readability."""
        text = (payload or "").strip()
        if not text:
            return payload
        out = text
        try:
            pretty = xml.dom.minidom.parseString(text).toprettyxml(indent="  ")
            # Drop empty lines produced by minidom.
            pretty = "\n".join(line for line in pretty.splitlines() if line.strip())
            out = pretty
        except Exception:
            out = text

        # Improve carrier block readability for multi-CC payloads.
        if name in {"PDSCH", "PUSCH", "PRACH", "ACTIVE"}:
            # Add block spacing only when payload actually has multiple CC sections.
            cc_blocks = len(re.findall(r"<\s*low-level-(?:tx|rx)-endpoints\s*>", out))
            if cc_blocks > 1:
                out = re.sub(
                    r"\n(\s*<(?:low-level-(?:tx|rx)-endpoints|tx-array-carriers|rx-array-carriers|low-level-(?:tx|rx)-links)>)",
                    r"\n\n\n\1",
                    out,
                )
                out = re.sub(r"\n{5,}", "\n\n\n\n", out)
            else:
                # Single-CC payload: keep blocks compact (no extra blank gaps).
                out = re.sub(
                    r"\n{2,}(\s*<(?:tx-array-carriers|rx-array-carriers|low-level-(?:tx|rx)-links)>)",
                    r"\n\1",
                    out,
                )
            # Same CC: keep endpoints/array-carriers/links attached without blank lines.
            out = re.sub(
                r"</(low-level-(?:tx|rx)-endpoints|tx-array-carriers|rx-array-carriers)>\n\s*\n(\s*<(?:tx-array-carriers|rx-array-carriers|low-level-(?:tx|rx)-links)>)",
                r"</\1>\n\2",
                out,
            )
            # Keep indentation around comment wrappers readable.
            out = re.sub(r"(\n\s*-->)\n(?=\s*<)", r"\1\n\n", out)
            out = out.replace("<!--\n\n", "<!--\n").replace("\n\n-->", "\n-->")
        return out

    def _on_main_window_configure(self, event: tk.Event) -> None:
        if event.widget is not self:
            return
        if self._geometry_save_job is not None:
            try:
                self.after_cancel(self._geometry_save_job)
            except Exception:
                pass
        self._geometry_save_job = self.after(900, self._save_current_config)

    def _on_log_window_configure(self, event: tk.Event) -> None:
        if self.log_window is None or event.widget is not self.log_window:
            return
        self.log_window_geometry = self.log_window.geometry()
        if self._log_geometry_save_job is not None:
            try:
                self.after_cancel(self._log_geometry_save_job)
            except Exception:
                pass
        self._log_geometry_save_job = self.after(900, self._save_current_config)

    def _on_log_mousewheel(self, event: tk.Event) -> str:
        if self.log is None:
            return "break"
        delta = int(-1 * (event.delta / 120))
        self.log.yview_scroll(delta, "units")
        return "break"

    def _extract_rpc_error_items(self, text: str) -> list[dict[str, str]]:
        items: list[dict[str, str]] = []
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for block in re.findall(r"<rpc-error\b[\s\S]*?</rpc-error>", text, flags=re.IGNORECASE):
            block_upper = block.upper()
            if "OPERATION-FAILED" not in block_upper:
                continue
            m_msg = re.search(r"<error-message>\s*([\s\S]*?)\s*</error-message>", block, flags=re.IGNORECASE)
            m_path = re.search(r"<error-path>\s*([\s\S]*?)\s*</error-path>", block, flags=re.IGNORECASE)
            m_info = re.search(r"<error-info>\s*([\s\S]*?)\s*</error-info>", block, flags=re.IGNORECASE)
            msg = re.sub(r"\s+", " ", m_msg.group(1)).strip() if m_msg else ""
            path = re.sub(r"\s+", " ", m_path.group(1)).strip() if m_path else ""
            info = re.sub(r"\s+", " ", m_info.group(1)).strip() if m_info else ""
            detail = msg or path or info or "operation-failed"
            items.append(
                {
                    "ts": now,
                    "kind": "rpc-error",
                    "summary": f"operation-failed | {detail[:180]}",
                    "anchor": msg or path or "operation-failed",
                    "raw": re.sub(r"\s+", " ", block).strip(),
                }
            )
        # Generic signal keywords requested by user.
        # Capture line-level snippets for fail/error/alarm (latest-first UI later).
        for ln in text.splitlines():
            s = ln.strip()
            if not s:
                continue
            m = re.search(r"\b(fail(?:ed|ure)?|error|alarm)\b", s, flags=re.IGNORECASE)
            if not m:
                continue
            kw = m.group(1).lower()
            flat = re.sub(r"\s+", " ", s).strip()
            items.append(
                {
                    "ts": now,
                    "kind": kw,
                    "summary": f"{kw} | {flat[:180]}",
                    "anchor": flat[:180] or kw,
                    "raw": flat,
                }
            )
        return items

    def _register_rpc_error_summary(self, text: str) -> None:
        for item in self._extract_rpc_error_items(text):
            sig = hashlib.sha1(item["raw"].encode("utf-8", errors="ignore")).hexdigest()
            if sig in self.rpc_error_seen:
                continue
            self.rpc_error_seen.add(sig)
            self.rpc_error_items.append(item)
            self._refresh_rpc_event_list_widget()

    def _jump_to_selected_rpc_error(self, _event: tk.Event | None = None) -> None:
        if self.log is None or self.rpc_error_list is None:
            return
        sel = self.rpc_error_list.curselection()
        if not sel:
            return
        idx = int(sel[0])
        if idx < 0 or idx >= len(self.rpc_error_items):
            return
        item = self.rpc_error_items[idx]
        anchor = item.get("anchor", "operation-failed").strip() or "operation-failed"
        self.log.configure(state="normal")
        self.log.tag_remove("rpc_err_hit", "1.0", "end")
        pos = self.log.search(anchor, "1.0", stopindex="end", nocase=True)
        if not pos:
            pos = self.log.search("operation-failed", "1.0", stopindex="end", nocase=True)
        if pos:
            end = f"{pos}+{max(1, len(anchor))}c"
            self.log.tag_add("rpc_err_hit", pos, end)
            self.log.tag_configure("rpc_err_hit", background="#7f1d1d", foreground="#fee2e2")
            self.log.see(pos)
            self.status_var.set(f"RPC error line: {pos}")
        self.log.configure(state="disabled")

    def _refresh_rpc_error_summary_widget(self) -> None:
        if self.rpc_error_list is None or not self.rpc_error_list.winfo_exists():
            return
        self.rpc_error_list.delete(0, "end")
        for item in self.rpc_error_items[-300:]:
            self.rpc_error_list.insert("end", item["summary"])

    def _refresh_rpc_event_list_widget(self) -> None:
        if self.rpc_event_list is None or not self.rpc_event_list.winfo_exists():
            return
        self.rpc_event_list.delete(0, "end")
        # Latest first
        for item in reversed(self.rpc_error_items[-600:]):
            ts = item.get("ts", "")
            kind = item.get("kind", "").upper()
            summary = item.get("summary", "")
            self.rpc_event_list.insert("end", f"[{ts}] [{kind}] {summary}")

    def open_rpc_event_window(self) -> None:
        if self.rpc_event_window is not None and self.rpc_event_window.winfo_exists():
            self.rpc_event_window.deiconify()
            self.rpc_event_window.lift()
            self._refresh_rpc_event_list_widget()
            return

        win = tk.Toplevel(self)
        win.title("RPC Events")
        win.geometry(self.rpc_event_window_geometry or "980x520")
        self.rpc_event_window = win

        top = ttk.Frame(win, padding=8)
        top.pack(fill="x")
        ttk.Label(
            top,
            text="Keywords: rpc-error(operation-failed), fail, error, alarm | Newest first",
            foreground="#64748b",
        ).pack(side="left")
        ttk.Button(top, text="Refresh", command=self._refresh_rpc_event_list_widget).pack(side="right")

        body = ttk.Frame(win, padding=8)
        body.pack(fill="both", expand=True)
        self.rpc_event_list = tk.Listbox(body, activestyle="none")
        self.rpc_event_list.pack(side="left", fill="both", expand=True)
        self.rpc_event_list.bind("<<ListboxSelect>>", self._jump_to_selected_rpc_event)
        self.rpc_event_list.bind("<Double-Button-1>", self._jump_to_selected_rpc_event)
        ys = ttk.Scrollbar(body, orient="vertical", command=self.rpc_event_list.yview)
        ys.pack(side="right", fill="y")
        self.rpc_event_list.configure(yscrollcommand=ys.set)
        self._refresh_rpc_event_list_widget()

        def _on_close() -> None:
            if self.rpc_event_window is not None:
                try:
                    self.rpc_event_window_geometry = self.rpc_event_window.geometry()
                except Exception:
                    pass
            self.rpc_event_list = None
            self._save_current_config()
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", _on_close)

    def _jump_to_selected_rpc_event(self, _event: tk.Event | None = None) -> None:
        if self.rpc_event_list is None or self.log is None:
            self.open_log_window()
        if self.rpc_event_list is None or self.log is None:
            return
        sel = self.rpc_event_list.curselection()
        if not sel:
            return
        # rpc_event_list is latest-first view of rpc_error_items tail.
        view = self.rpc_error_items[-600:]
        rev = list(reversed(view))
        idx = int(sel[0])
        if idx < 0 or idx >= len(rev):
            return
        item = rev[idx]
        anchor = (item.get("anchor") or item.get("raw") or item.get("kind") or "error").strip()
        self.open_log_window()
        if self.log is None:
            return
        self.log.configure(state="normal")
        self.log.tag_remove("rpc_err_hit", "1.0", "end")
        # Search from bottom to jump to latest occurrence.
        pos = self.log.search(anchor, "end-1c", stopindex="1.0", nocase=True, backwards=True)
        if not pos:
            k = (item.get("kind") or "error").strip() or "error"
            pos = self.log.search(k, "end-1c", stopindex="1.0", nocase=True, backwards=True)
        if pos:
            end = f"{pos}+{max(1, min(len(anchor), 180))}c"
            self.log.tag_add("rpc_err_hit", pos, end)
            self.log.tag_configure("rpc_err_hit", background="#7f1d1d", foreground="#fee2e2")
            self.log.see(pos)
            self.status_var.set(f"RPC event line: {pos}")
        self.log.configure(state="disabled")

    def open_log_window(self) -> None:
        if self.log_window is not None and self.log_window.winfo_exists():
            self.log_window.deiconify()
            self.log_window.lift()
            self._flush_hidden_logs_to_widget()
            return

        win = tk.Toplevel(self)
        win.title("Netconf Logs")
        win.geometry(self.log_window_geometry or "1100x720")
        self.log_window = win
        win.bind("<Configure>", self._on_log_window_configure)

        tools = ttk.LabelFrame(win, text="Remote Log Tools", padding=8)
        tools.pack(fill="x", padx=8, pady=(8, 6))
        self.fetch_lines_var = tk.StringVar(value="300")
        ttk.Button(tools, text="Log Load", command=self.load_full_remote_log).pack(side="left", padx=(8, 6), pady=8)
        ttk.Button(tools, text="Clear Logs", command=self.clear_log).pack(side="left", padx=6, pady=8)
        ttk.Button(tools, text="RPC Events", command=self.open_rpc_event_window).pack(side="left", padx=6, pady=8)
        gui_cap = ttk.Frame(tools)
        gui_cap.pack(side="left", padx=(16, 4), pady=4)
        ttk.Label(gui_cap, text="GUI keep lines").pack(side="left", padx=(0, 4))
        ttk.Entry(gui_cap, textvariable=self.gui_log_max_lines_var, width=8).pack(side="left", padx=2)
        ttk.Button(gui_cap, text="Apply", command=self.apply_gui_log_line_limit).pack(side="left", padx=6)
        ttk.Label(gui_cap, foreground="#64748b", text="Oldest trimmed when exceeded.").pack(side="left", padx=(8, 0))
        ttk.Label(tools, textvariable=self.perf_text_var, foreground="#666666").pack(side="right", padx=8)

        log_frame = ttk.LabelFrame(win, text="Live Output / Remote Log", padding=8)
        log_frame.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        ttk.Label(
            log_frame,
            textvariable=self.log_target_hint_var,
            foreground="#0369a1",
            justify="left",
            wraplength=1040,
            font=("Segoe UI", 9),
        ).pack(anchor="w", fill="x", pady=(0, 6))
        self._refresh_log_target_hint_line()

        log_inner = ttk.Frame(log_frame)
        log_inner.pack(fill="both", expand=True)
        self.log = tk.Text(
            log_inner,
            wrap="none",
            relief="flat",
        )
        self._apply_code_text_theme(self.log)
        log_scroll_y = ttk.Scrollbar(log_inner, orient="vertical", command=self.log.yview)
        log_scroll_x = ttk.Scrollbar(log_inner, orient="horizontal", command=self.log.xview)
        self.log.configure(yscrollcommand=log_scroll_y.set, xscrollcommand=log_scroll_x.set)
        self.log.grid(row=0, column=0, sticky="nsew")
        log_scroll_y.grid(row=0, column=1, sticky="ns")
        log_scroll_x.grid(row=1, column=0, sticky="ew")
        log_inner.rowconfigure(0, weight=1)
        log_inner.columnconfigure(0, weight=1)
        self.log.configure(state="disabled")
        self.log.bind("<MouseWheel>", self._on_log_mousewheel)

        def _on_close() -> None:
            if self.log_window is not None:
                try:
                    self.log_window_geometry = self.log_window.geometry()
                except Exception:
                    pass
            self.log = None
            self.rpc_error_list = None
            self._save_current_config()
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", _on_close)
        self._flush_hidden_logs_to_widget()

    def add_message_tab(self, initial_title: str | None = None) -> None:
        idx = len(self.message_tabs) + 1
        title_var = tk.StringVar(value=initial_title or f"MSG-{idx}")
        frame = ttk.Frame(self.msg_notebook)
        top = ttk.Frame(frame)
        top.pack(fill="x", padx=8, pady=6)
        ttk.Label(top, text="Title").pack(side="left")
        ttk.Entry(top, textvariable=title_var, width=18).pack(side="left", padx=(6, 10))
        ttk.Label(top, text="(XML 또는 CLI text)").pack(side="left")

        text = tk.Text(frame, wrap="none")
        text.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        text.configure(undo=True, autoseparators=True, maxundo=-1)
        self._setup_xml_editor_theme(text)
        text.bind("<KeyRelease>", lambda _e, tw=text: self._on_xml_editor_changed(tw))
        text.bind("<FocusIn>", lambda _e, tw=text: self._on_xml_editor_changed(tw))
        for seq in ("<Control-slash>", "<Control-question>", "<Control-KeyPress-slash>", "<Control-KeyPress-KP_Divide>"):
            text.bind(seq, self._on_shortcut_toggle_comment, add="+")
        for seq in ("<Control-z>", "<Control-Z>", "<Control-KeyPress-z>", "<Control-KeyPress-Z>"):
            text.bind(seq, self._on_xml_undo, add="+")
        for seq in ("<Control-Shift-z>", "<Control-Shift-Z>", "<Control-y>", "<Control-Y>"):
            text.bind(seq, self._on_xml_redo, add="+")
        for seq in ("<Control-f>", "<Control-F>", "<Control-KeyPress-f>", "<Control-KeyPress-F>"):
            text.bind(seq, self._on_xml_find, add="+")
        for seq in ("<Control-h>", "<Control-H>", "<Control-KeyPress-h>", "<Control-KeyPress-H>"):
            text.bind(seq, self._on_xml_replace, add="+")
        tab_data = {"frame": frame, "title_var": title_var, "text": text}
        self.message_tabs.append(tab_data)
        self.msg_notebook.add(frame, text=title_var.get())
        title_var.trace_add("write", lambda *_: self._refresh_message_tab_titles())
        self._refresh_message_tab_titles()

    def _refresh_message_tab_titles(self) -> None:
        for i, tab in enumerate(self.message_tabs):
            self.msg_notebook.tab(i, text=tab["title_var"].get().strip() or f"MSG-{i+1}")

    def remove_selected_message_tab(self) -> None:
        if not self.message_tabs:
            return
        idx = self.msg_notebook.index("current")
        self.msg_notebook.forget(idx)
        self.message_tabs.pop(idx)
        if not self.message_tabs:
            self.add_message_tab(initial_title="MSG-1")
        self._refresh_message_tab_titles()

    def move_selected_tab_left(self) -> None:
        if len(self.message_tabs) < 2:
            return
        idx = self.msg_notebook.index("current")
        if idx <= 0:
            return
        tab = self.message_tabs.pop(idx)
        self.message_tabs.insert(idx - 1, tab)
        self.msg_notebook.insert(idx - 1, tab["frame"])
        self.msg_notebook.select(idx - 1)
        self._refresh_message_tab_titles()

    def move_selected_tab_right(self) -> None:
        if len(self.message_tabs) < 2:
            return
        idx = self.msg_notebook.index("current")
        if idx >= len(self.message_tabs) - 1:
            return
        tab = self.message_tabs.pop(idx)
        self.message_tabs.insert(idx + 1, tab)
        self.msg_notebook.insert(idx + 1, tab["frame"])
        self.msg_notebook.select(idx + 1)
        self._refresh_message_tab_titles()

    def _build_netconf_library_panel(self, parent: ttk.Widget) -> None:
        """Left-side folder tree for XML/RPC templates with Get/Set send actions."""
        wrap = ttk.LabelFrame(parent, text="XML Library", padding=6)
        wrap.pack(fill="both", expand=True)

        tools = ttk.Frame(wrap)
        tools.pack(fill="x", pady=(0, 6))
        ttk.Button(tools, text="Load Folder…", command=self._netconf_browse_library_folder).pack(
            side="left", padx=(0, 4)
        )
        ttk.Button(tools, text="Refresh", command=self._netconf_refresh_library_tree).pack(side="left")

        tree_wrap = ttk.Frame(wrap)
        tree_wrap.pack(fill="both", expand=True)
        tree_wrap.rowconfigure(0, weight=1)
        tree_wrap.columnconfigure(0, weight=1)
        ys = ttk.Scrollbar(tree_wrap, orient="vertical")
        self.netconf_lib_tree = ttk.Treeview(tree_wrap, show="tree", selectmode="browse", yscrollcommand=ys.set)
        ys.configure(command=self.netconf_lib_tree.yview)
        self.netconf_lib_tree.grid(row=0, column=0, sticky="nsew")
        ys.grid(row=0, column=1, sticky="ns")
        self.netconf_lib_tree.bind("<<TreeviewSelect>>", self._on_netconf_library_select, add="+")
        self.netconf_lib_tree.bind("<Double-1>", self._on_netconf_library_double_click, add="+")
        self.netconf_lib_tree.bind(
            "<Button-3>",
            self._on_netconf_library_right_click,
            add="+",
        )

        ops = ttk.Frame(wrap)
        ops.pack(fill="x", pady=(8, 4))
        self.netconf_lib_get_btn = ttk.Button(
            ops, text="Send GET", command=lambda: self._netconf_send_library_selection("get"), state="disabled"
        )
        self.netconf_lib_get_btn.pack(side="left", padx=(0, 4))
        self.netconf_lib_set_btn = ttk.Button(
            ops, text="Send SET", command=lambda: self._netconf_send_library_selection("set"), state="disabled"
        )
        self.netconf_lib_set_btn.pack(side="left", padx=(0, 4))
        self.netconf_lib_send_btn = ttk.Button(
            ops, text="Send", command=self.send_selected_once, state="disabled"
        )
        self.netconf_lib_send_btn.pack(side="left", padx=(0, 4))
        ttk.Button(ops, text="Open Tab", command=self._netconf_open_library_selection_in_tab).pack(
            side="right"
        )

        ttk.Label(
            wrap,
            textvariable=self.netconf_lib_status_var,
            foreground="#64748b",
            wraplength=260,
            justify="left",
        ).pack(fill="x", pady=(6, 0))
        ttk.Label(
            wrap,
            text="Double-click → 편집 탭. GET/SET는 XML 원문(user-rpc). Send는 선택된 탭 전송 (Ctrl+Shift+Space).",
            foreground="#94a3b8",
            wraplength=260,
            justify="left",
            font=("Segoe UI", 8),
        ).pack(fill="x", pady=(4, 0))

    def _netconf_restore_library_from_config(self) -> None:
        path = (getattr(self, "_netconf_library_path_to_restore", "") or "").strip()
        if not path:
            return
        p = Path(path).expanduser()
        if p.is_dir():
            self._netconf_populate_library_tree(p)

    def _netconf_browse_library_folder(self) -> None:
        initial = ""
        if self.netconf_library_root and self.netconf_library_root.exists():
            initial = str(self.netconf_library_root)
        path = filedialog.askdirectory(title="Select XML template folder", initialdir=initial or None)
        if not path:
            return
        self._netconf_populate_library_tree(Path(path))
        self._save_current_config()

    def _netconf_refresh_library_tree(self) -> None:
        if self.netconf_library_root and self.netconf_library_root.exists():
            self._netconf_populate_library_tree(self.netconf_library_root)
        else:
            messagebox.showinfo("XML Library", "Load a folder first.")

    def _netconf_populate_library_tree(self, root_path: Path) -> None:
        tree = self.netconf_lib_tree
        if tree is None:
            return
        root_path = root_path.expanduser().resolve()
        if not root_path.is_dir():
            messagebox.showerror("XML Library", f"Not a folder:\n{root_path}")
            return
        self.netconf_library_root = root_path
        self.netconf_library_paths.clear()
        self.netconf_library_selected_path = ""
        for item in tree.get_children(""):
            tree.delete(item)
        root_iid = tree.insert("", "end", text=root_path.name, open=True)
        self.netconf_library_paths[root_iid] = str(root_path)
        file_count = self._netconf_add_library_tree_children(tree, root_iid, root_path)
        self.netconf_lib_status_var.set(f"{root_path} — {file_count} file(s)")
        self.append_log(f"[GUI] XML library loaded: {root_path} ({file_count} files)\n")

    def _netconf_add_library_tree_children(self, tree: ttk.Treeview, parent_iid: str, dir_path: Path) -> int:
        count = 0
        try:
            entries = sorted(dir_path.iterdir(), key=lambda p: (not p.is_dir(), p.name.lower()))
        except OSError as exc:
            self.append_log(f"[GUI] XML library scan failed: {dir_path}: {exc}\n")
            return 0
        for entry in entries:
            if entry.name.startswith("."):
                continue
            if entry.is_dir():
                child_iid = tree.insert(parent_iid, "end", text=entry.name, open=False)
                self.netconf_library_paths[child_iid] = str(entry.resolve())
                count += self._netconf_add_library_tree_children(tree, child_iid, entry)
            elif entry.suffix.lower() in _NETCONF_LIB_EXTENSIONS:
                kind = self._netconf_guess_file_kind(entry)
                label = entry.name if not kind else f"{entry.name}  [{kind}]"
                leaf_iid = tree.insert(parent_iid, "end", text=label)
                self.netconf_library_paths[leaf_iid] = str(entry.resolve())
                count += 1
        return count

    @staticmethod
    def _netconf_read_text_file(path: Path) -> str:
        try:
            return path.read_text(encoding="utf-8")
        except Exception:
            return path.read_text(errors="ignore")

    def _netconf_guess_file_kind(self, path: Path) -> str:
        try:
            head = self._netconf_read_text_file(path)[:4000].lower()
        except OSError:
            return ""
        if "<get-config" in head or re.search(r"<get\b", head):
            return "GET"
        if "<edit-config" in head or head.lstrip().startswith("<config"):
            return "SET"
        if head.lstrip().startswith("get"):
            return "GET"
        if "edit-config" in head:
            return "SET"
        return ""

    def _netconf_library_path_for_iid(self, iid: str) -> str | None:
        path = self.netconf_library_paths.get(iid, "")
        if not path:
            return None
        p = Path(path)
        if p.is_file():
            return path
        return None

    def _netconf_selected_library_file(self) -> tuple[str, str] | None:
        tree = self.netconf_lib_tree
        if tree is None:
            return None
        sel = tree.selection()
        if not sel:
            return None
        path = self._netconf_library_path_for_iid(sel[0])
        if not path:
            return None
        return path, self._netconf_read_text_file(Path(path))

    def _on_netconf_library_select(self, _event: tk.Event | None = None) -> None:
        picked = self._netconf_selected_library_file()
        if not picked:
            self.netconf_library_selected_path = ""
            if self.netconf_library_root:
                self.netconf_lib_status_var.set(str(self.netconf_library_root))
            return
        path, content = picked
        self.netconf_library_selected_path = path
        kind = self._netconf_classify_payload(content)
        self.netconf_lib_status_var.set(f"{Path(path).name} — detected: {kind.upper()}")

    def _on_netconf_library_double_click(self, _event: tk.Event | None = None) -> None:
        self._netconf_open_library_selection_in_tab(new_tab=True)

    def _on_netconf_library_right_click(self, event: tk.Event) -> None:
        tree = self.netconf_lib_tree
        if tree is None:
            return
        iid = tree.identify_row(event.y)
        if iid:
            tree.selection_set(iid)
            self._on_netconf_library_select()
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="Open in Tab", command=self._netconf_open_library_selection_in_tab)
        menu.add_command(label="Open in New Tab", command=lambda: self._netconf_open_library_selection_in_tab(new_tab=True))
        menu.add_separator()
        menu.add_command(label="Send GET", command=lambda: self._netconf_send_library_selection("get"))
        menu.add_command(label="Send SET", command=lambda: self._netconf_send_library_selection("set"))
        try:
            menu.tk_popup(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _netconf_open_library_selection_in_tab(self, new_tab: bool = False) -> None:
        picked = self._netconf_selected_library_file()
        if not picked:
            messagebox.showinfo("XML Library", "Select an XML/text file in the tree.")
            return
        path, content = picked
        title = Path(path).stem
        if new_tab or not self.message_tabs:
            self.add_message_tab(initial_title=title)
            idx = len(self.message_tabs) - 1
        else:
            idx = self.msg_notebook.index("current")
        text_widget = self.message_tabs[idx]["text"]
        self.message_tabs[idx]["title_var"].set(title)
        text_widget.delete("1.0", "end")
        text_widget.insert("1.0", content)
        self._highlight_xml(text_widget)
        self.msg_notebook.select(idx)
        self._refresh_message_tab_titles()
        self.append_log(f"[GUI] Opened library file in tab: {path}\n")

    def _netconf_classify_payload(self, content: str) -> str:
        text = (content or "").lstrip("\ufeff").strip()
        if not text:
            return "empty"
        lower = text.lower()
        if lower.startswith("get") and not lower.startswith("<"):
            return "get"
        if "edit-config" in lower and not lower.startswith("<rpc"):
            return "set"
        body = re.sub(r"^\s*<\?xml[^>]*\?>\s*", "", text, flags=re.IGNORECASE)
        if body.lower().startswith("<rpc"):
            if "<get-config" in lower or re.search(r"<get\b", lower):
                return "get"
            if "<edit-config" in lower:
                return "set"
            return "rpc"
        if body.lower().startswith("<edit-config") or body.lower().startswith("<config"):
            return "set"
        return "xml"

    def _netconf_prepare_library_payload(self, content: str, operation: str) -> str:
        op = operation.lower()
        kind = self._netconf_classify_payload(content)
        text = content.lstrip("\ufeff").strip()
        if op == "auto":
            return text
        if op == "get":
            if kind == "set":
                raise ValueError("Selected file looks like SET (edit-config). Choose a GET template.")
            return text
        if op == "set":
            if kind == "get":
                raise ValueError("Selected file looks like GET. Choose a SET (edit-config) template.")
            return text
        return text

    def _netconf_wrap_payload_for_send(self, text: str, kind: str) -> str:
        """Prepare XML Library payload using legacy GET/SET mapping behavior."""
        body = re.sub(r"^\s*<\?xml[^>]*\?>\s*", "", text.strip(), flags=re.IGNORECASE)
        lower = body.lower()
        if kind == "get":
            if lower.startswith("get"):
                return text.strip()
            xpath = self._extract_rpc_xpath_filter(body)
            if xpath:
                return f"get --filter-xpath {shlex.quote(xpath)}"
            if lower.startswith("<rpc"):
                return body
            if lower.startswith("<get-config") or re.match(r"^\s*<get\b", body, flags=re.IGNORECASE):
                return body
            if lower.startswith("<get-config") or re.match(r"^\s*<get\b", body, flags=re.IGNORECASE):
                return f"<rpc message-id=\"lib-get\">{body}</rpc>" if not lower.startswith("<rpc") else body
            remote_path = f"/var/tmp/netconf_tmp/gui_filter_{int(time.time() * 1000)}.xml"
            cache_key = self._cache_key("lib_get_filter", body)
            if cache_key not in self.remote_cfg_cache:
                ok, msg = self._upload_text_to_remote(remote_path, body)
                if not ok:
                    raise RuntimeError(msg)
                self.remote_cfg_cache[cache_key] = remote_path
            else:
                remote_path = self.remote_cfg_cache[cache_key]
            return f"get --filter={remote_path}"
        # set
        if lower.startswith("edit-config") or "edit-config" in lower.split("\n", 1)[0]:
            return body
        if lower.startswith("<rpc"):
            return body
        if lower.startswith("<config"):
            return body
        if lower.startswith("<edit-config"):
            return f"<rpc message-id=\"lib-set\">{body}</rpc>"
        return text.strip()

    def _netconf_send_library_selection(self, operation: str) -> None:
        if not self.session_established:
            messagebox.showwarning("Warning", "Session is not established yet.")
            return
        if not self.manual_send_ready:
            messagebox.showwarning("Warning", "Manual send is enabled after session login.")
            return
        picked = self._netconf_selected_library_file()
        if not picked:
            messagebox.showinfo("XML Library", "Select an XML/text file in the tree.")
            return
        path, content = picked
        title = Path(path).name
        try:
            payload = self._netconf_prepare_library_payload(content, operation)
        except ValueError as exc:
            messagebox.showwarning("XML Library", str(exc))
            return
        except Exception as exc:
            messagebox.showerror("XML Library", str(exc))
            return
        threading.Thread(
            target=self._netconf_send_library_payload_worker,
            args=(title, path, payload, operation),
            daemon=True,
        ).start()

    def _netconf_send_library_payload_worker(
        self, title: str, path: str, payload: str, operation: str
    ) -> None:
        self.after(0, lambda: self.send_selected_btn.config(state="disabled"))
        self.after(0, lambda: self.send_all_btn.config(state="disabled"))
        mb = self.mplane_apply_btn
        if mb is not None:
            self.after(0, lambda b=mb: b.config(state="disabled"))
        try:
            to_send = payload
            # XML Library는 요청대로 XML 원문을 그대로 user-rpc로 전송한다.
            if self.send_mode_var.get() == "raw_rpc" and self._looks_like_netconf_xml_document(payload):
                to_send = self._upload_and_build_user_rpc_command(
                    payload,
                    cache_kind="lib_user_rpc",
                    with_out=self._is_rpc_get_like(payload),
                )
                self.after(0, self.append_log, f"[GUI] XML library RAW(as-is) -> {to_send}\n")
            self._send_scheduler_payload(to_send)
            self.after(
                0,
                self.append_log,
                f"[GUI] XML library {operation.upper()} sent: {title} ({path})\n",
            )
        except Exception as exc:
            self.after(0, self.append_log, f"[GUI] XML library send failed ({title}): {exc}\n")
            self.after(0, lambda e=exc: messagebox.showerror("XML Library", str(e)))
        finally:
            self.after(0, self._sync_manual_send_widgets)

    def load_file_into_selected_tab(self) -> None:
        if not self.message_tabs:
            return
        path = filedialog.askopenfilename(
            title="Select XML/Text file",
            filetypes=[("XML/Text", "*.xml *.txt *.rpc"), ("All files", "*.*")],
        )
        if not path:
            return
        idx = self.msg_notebook.index("current")
        text_widget = self.message_tabs[idx]["text"]
        try:
            content = Path(path).read_text(encoding="utf-8")
        except Exception:
            content = Path(path).read_text(errors="ignore")
        text_widget.delete("1.0", "end")
        text_widget.insert("1.0", content)
        self._highlight_xml(text_widget)
        self.append_log(f"[GUI] Loaded file into tab: {path}\n")

    def load_excel_tabs(self) -> None:
        path = filedialog.askopenfilename(
            title="Select Excel file",
            initialdir="F:/ORAN/nDIU",
            initialfile="nDIU_n78_Mplane정리.xlsx",
            filetypes=[("Excel Workbook", "*.xlsx *.xlsm"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            import openpyxl  # type: ignore
        except Exception:
            messagebox.showerror("Error", "openpyxl is required. Install with: py -m pip install openpyxl")
            return
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as exc:
            messagebox.showerror("Error", f"Failed to open Excel file:\n{exc}")
            return

        # Use last 6 sheet tabs (matches "bottom tabs 6개" request).
        target_sheets = wb.sheetnames[-6:]
        if not target_sheets:
            messagebox.showwarning("Warning", "No sheets found in workbook.")
            return

        while self.message_tabs:
            self.msg_notebook.forget(0)
            self.message_tabs.pop(0)

        for sname in target_sheets:
            ws = wb[sname]
            # Preserve worksheet row structure first, then normalize line endings.
            row_lines: list[str] = []
            for row in ws.iter_rows(values_only=True):
                vals = [v for v in row if v is not None]
                if not vals:
                    row_lines.append("")
                    continue
                # Use the first non-empty cell in each row to preserve visual XML layout.
                s = str(vals[0])
                s = s.replace("\r\n", "\n").replace("_x000D_", "")
                s = s.replace("\\r\\n", "\n").replace("\\n", "\n").replace("\\t", "\t")
                row_lines.append(s)

            content = "\n".join(row_lines)
            # Remove only trailing empty lines.
            content = re.sub(r"\n+\Z", "", content)
            # If XML is still one-line, pretty-print for readability.
            if content.startswith("<") and "\n" not in content and "><" in content:
                try:
                    pretty = xml.dom.minidom.parseString(content).toprettyxml(indent="  ")
                    content = "\n".join(
                        l for l in pretty.splitlines() if l.strip() and not l.strip().startswith("<?xml")
                    )
                except Exception:
                    pass
            self.add_message_tab(initial_title=sname)
            self.message_tabs[-1]["text"].delete("1.0", "end")
            self.message_tabs[-1]["text"].insert("1.0", content)
            self._highlight_xml(self.message_tabs[-1]["text"])

        self.append_log(f"[GUI] Loaded {len(target_sheets)} Excel tabs from: {path}\n")

    def _ssh_base_cmd(
        self,
        user: str,
        host: str,
        port: str,
        key_path: str,
        use_password: bool,
        auth_mode: str,
        ssh_password: str,
    ) -> list[str]:
        cmd: list[str] = []
        if auth_mode == "sshpass":
            cmd.extend(["sshpass", "-e"])
            cmd.extend(["ssh", "-p", port, *self._ssh_options(use_password)])
        elif auth_mode == "putty":
            cmd.extend(["plink", "-ssh", "-P", port])
            if use_password:
                cmd.extend(["-pw", ssh_password])
        else:
            cmd.extend(["ssh", "-p", port, *self._ssh_options(use_password)])
        if key_path:
            cmd.extend(["-i", key_path])
        cmd.append(f"{user}@{host}")
        return cmd

    def _scp_base_cmd(
        self, port: str, key_path: str, use_password: bool, auth_mode: str, ssh_password: str
    ) -> list[str]:
        cmd: list[str] = []
        if auth_mode == "sshpass":
            cmd.extend(["sshpass", "-e"])
            cmd.extend([
                "scp",
                "-P",
                port,
                "-o",
                "ConnectTimeout=8",
                "-o",
                "ConnectionAttempts=1",
            ])
            if not use_password:
                cmd.extend(["-o", "BatchMode=yes"])
        elif auth_mode == "putty":
            cmd.extend(["pscp", "-P", port])
            if use_password:
                cmd.extend(["-pw", ssh_password])
        else:
            cmd.extend([
                "scp",
                "-P",
                port,
                "-o",
                "ConnectTimeout=8",
                "-o",
                "ConnectionAttempts=1",
            ])
            if not use_password:
                cmd.extend(["-o", "BatchMode=yes"])
        if key_path:
            cmd.extend(["-i", key_path])
        return cmd

    def _ensure_remote_script(
        self,
        ssh_user: str,
        ssh_host: str,
        ssh_port: str,
        key_path: str,
        ssh_password: str,
        auth_mode: str,
        remote_script: str,
        remote_log_path: str,
    ) -> tuple[bool, str]:
        local_script = self._resolve_local_script_path()
        self.script_path = local_script
        local_exists = local_script.exists()

        if auth_mode == "paramiko":
            return self._ensure_remote_script_paramiko(
                ssh_user, ssh_host, ssh_port, key_path, ssh_password, remote_script, remote_log_path, local_exists
            )

        use_password = bool(ssh_password)
        ssh_cmd = self._ssh_base_cmd(
            ssh_user, ssh_host, ssh_port, key_path, use_password, auth_mode, ssh_password
        )
        run_env = os.environ.copy()
        if use_password and auth_mode == "sshpass":
            run_env["SSHPASS"] = ssh_password
        check_cmd = ssh_cmd + [f"test -f {shlex.quote(remote_script)}"]
        check_result = subprocess.run(check_cmd, capture_output=True, text=True, env=run_env)
        script_exists = check_result.returncode == 0

        remote_parent = str(Path(remote_script).parent).replace("\\", "/")
        mkdir_cmd = ssh_cmd + [f"mkdir -p {shlex.quote(remote_parent)}"]
        mkdir_result = subprocess.run(mkdir_cmd, capture_output=True, text=True, env=run_env)
        if mkdir_result.returncode != 0:
            err = (mkdir_result.stderr or mkdir_result.stdout).strip()
            return False, f"Failed to create remote directory: {err}"

        # If remote already has the script, allow run even when local file is missing.
        if script_exists and not local_exists:
            log_mkdir_cmd = ssh_cmd + [f"mkdir -p {shlex.quote(remote_log_path)}"]
            log_mkdir_result = subprocess.run(log_mkdir_cmd, capture_output=True, text=True, env=run_env)
            if log_mkdir_result.returncode != 0:
                err = (log_mkdir_result.stderr or log_mkdir_result.stdout).strip()
                return False, f"Failed to create remote LOG_PATH directory: {err}"
            return True, f"Remote script reused: {remote_script}, LOG_PATH ready: {remote_log_path}"

        if not local_exists:
            app_dir = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
            return (
                False,
                "Local script not found.\n"
                f"- tried: {local_script}\n"
                f"- expected near exe: {app_dir / 'miniDU_callhome.sh'}",
            )

        scp_cmd = self._scp_base_cmd(ssh_port, key_path, use_password, auth_mode, ssh_password) + [
            str(self.script_path),
            f"{ssh_user}@{ssh_host}:{remote_script}",
        ]
        scp_result = subprocess.run(scp_cmd, capture_output=True, text=True, env=run_env)
        if scp_result.returncode != 0:
            err = (scp_result.stderr or scp_result.stdout).strip()
            return False, f"Failed to upload script via scp: {err}"

        chmod_cmd = ssh_cmd + [f"chmod +x {shlex.quote(remote_script)}"]
        chmod_result = subprocess.run(chmod_cmd, capture_output=True, text=True, env=run_env)
        if chmod_result.returncode != 0:
            err = (chmod_result.stderr or chmod_result.stdout).strip()
            return False, f"Uploaded, but chmod failed: {err}"

        log_mkdir_cmd = ssh_cmd + [f"mkdir -p {shlex.quote(remote_log_path)}"]
        log_mkdir_result = subprocess.run(log_mkdir_cmd, capture_output=True, text=True, env=run_env)
        if log_mkdir_result.returncode != 0:
            err = (log_mkdir_result.stderr or log_mkdir_result.stdout).strip()
            return False, f"Failed to create remote LOG_PATH directory: {err}"

        if script_exists:
            return True, f"Remote script updated: {remote_script}, LOG_PATH ready: {remote_log_path}"
        return True, f"Remote script created: {remote_script}, LOG_PATH ready: {remote_log_path}"

    def _ensure_remote_script_paramiko(
        self,
        ssh_user: str,
        ssh_host: str,
        ssh_port: str,
        key_path: str,
        ssh_password: str,
        remote_script: str,
        remote_log_path: str,
        local_exists: bool,
    ) -> tuple[bool, str]:
        try:
            import paramiko  # type: ignore
        except Exception:
            return False, "paramiko is not installed. Run: py -m pip install paramiko"

        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:
            client.connect(
                hostname=ssh_host,
                port=int(ssh_port),
                username=ssh_user,
                password=ssh_password if ssh_password else None,
                key_filename=key_path if key_path else None,
                timeout=8,
                auth_timeout=8,
                banner_timeout=8,
                look_for_keys=not bool(ssh_password),
                allow_agent=True,
            )

            sftp = client.open_sftp()
            remote_parent = str(Path(remote_script).parent).replace("\\", "/")
            _, stdout, stderr = client.exec_command(f"mkdir -p {shlex.quote(remote_parent)}")
            if stdout.channel.recv_exit_status() != 0:
                return False, f"Failed to create remote directory: {stderr.read().decode(errors='ignore').strip()}"
            remote_exists = False
            try:
                sftp.stat(remote_script)
                remote_exists = True
            except Exception:
                remote_exists = False

            if local_exists:
                # Prefer fresh upload when local script exists.
                sftp.put(str(self.script_path), remote_script)
                sftp.chmod(remote_script, 0o755)
            elif not remote_exists:
                app_dir = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
                return (
                    False,
                    "Local script not found.\n"
                    f"- tried: {self.script_path}\n"
                    f"- expected near exe: {app_dir / 'miniDU_callhome.sh'}",
                )

            _, stdout, stderr = client.exec_command(f"mkdir -p {shlex.quote(remote_log_path)}")
            if stdout.channel.recv_exit_status() != 0:
                return False, f"Failed to create remote LOG_PATH directory: {stderr.read().decode(errors='ignore').strip()}"
            if local_exists:
                return True, f"Remote script ready: {remote_script}, LOG_PATH ready: {remote_log_path}"
            return True, f"Remote script reused: {remote_script}, LOG_PATH ready: {remote_log_path}"
        except Exception as exc:
            return False, f"Remote prep failed: {exc}"
        finally:
            try:
                client.close()
            except Exception:
                pass

    @staticmethod
    def _sanitize_log_text(text: str) -> str:
        # Remove control chars (e.g., BEL) that render as odd symbols in Tk.
        return re.sub(r"[\x00-\x08\x0b-\x1f\x7f]", "", text)

    @staticmethod
    def _pretty_xml_if_possible(text: str) -> str:
        raw = text.strip()
        if not raw.startswith("<"):
            return text
        try:
            pretty = xml.dom.minidom.parseString(raw).toprettyxml(indent="  ")
            pretty = "\n".join(
                l for l in pretty.splitlines() if l.strip() and not l.strip().startswith("<?xml")
            )
            return pretty + "\n"
        except Exception:
            return text if text.endswith("\n") else text + "\n"

    def _format_rpc_exchange_xml(self, blob: str) -> str:
        """Pretty-print rpc + rpc-reply inside one NETCONF RPC exchange block."""
        # Drop netopeer/libssh noise that may have leaked into the capture window.
        cleaned_lines: list[str] = []
        for ln in blob.splitlines():
            if self._should_hide_line(ln + "\n"):
                continue
            cleaned_lines.append(ln)
        blob = "\n".join(cleaned_lines).strip()
        if not blob:
            return ""
        limit = self._max_log_pretty_chars
        m = re.search(r"</rpc\s*>", blob, flags=re.IGNORECASE)
        pieces: list[str] = []
        if m:
            head = blob[: m.end()].strip()
            tail = blob[m.end() :].strip()
            if head:
                pieces.append(head)
            if tail:
                pieces.append(tail)
        else:
            pieces = [blob]
        out: list[str] = []
        for piece in pieces:
            if not piece:
                continue
            if len(piece) <= limit:
                out.append(self._pretty_xml_if_possible(piece))
            else:
                out.append(piece if piece.endswith("\n") else piece + "\n")
        return "".join(out)

    def _should_hide_line(self, line: str) -> bool:
        line_upper = line.upper()
        stripped_upper = line_upper.lstrip()

        # Strip GUI / Conformance prefixes so content rules still apply.
        content_upper = stripped_upper
        for pref in (
            "[CONFORMANCE-RUN] ",
            "[CONFORMANCE-SYNC] ",
            "[CONFORMANCE] ",
            "[CONFORMANCE-DEBUG] ",
            "[GUI] ",
        ):
            if content_upper.startswith(pref):
                content_upper = content_upper[len(pref) :].lstrip()
                break

        # Drop netopeer2 / libssh DEBUG·VERBOSE noise (and SSH packet chatter).
        # Also catch concatenated lines like "...packet type 94[GUI] ..."
        if (
            "NC DEBUG:" in line_upper
            or "NC VERBOSE:" in line_upper
            or content_upper.startswith("NC DEBUG:")
            or content_upper.startswith("NC VERBOSE:")
            or content_upper.startswith("DEBUG:")
            or content_upper.startswith("VERBOSE:")
            or "SSH_PACKET_" in line_upper
            or "SSH_SOCKET_" in line_upper
            or "SSH_CHANNEL_" in line_upper
            or "CHANNEL_RCV_DATA" in line_upper
            or "CHANNEL_DEFAULT_BUFFERIZE" in line_upper
            or "CHANNEL_WRITE_COMMON" in line_upper
            or "CHANNEL WINDOWS ARE NOW" in line_upper
            or "PACKET_SEND2" in line_upper
            or "BYTES LEFT IN SOCKET BUFFER" in line_upper
            or "DISPATCHING HANDLER FOR PACKET TYPE" in line_upper
            or "SSH_PACKET_NEED_REKEY" in line_upper
            or "SSH_PACKET_SOCKET_CALLBACK" in line_upper
            or "SSH_PACKET_PROCESS" in line_upper
            or "ENABLING POLLOUT FOR SOCKET" in line_upper
            or content_upper.startswith("SENDING MESSAGE:")
        ):
            return True

        # Hide [GUI] / [INFO] lines from the log window.
        if (
            stripped_upper.startswith("[GUI]")
            or content_upper.startswith("[GUI]")
            or "[GUI] NETCONF RPC EXCHANGE" in line_upper
            or "[GUI] RPC REPLY VERDICT" in line_upper
            or "[GUI] EDIT-CONFIG" in line_upper
            or "[GUI] USER-RPC" in line_upper
            or "[GUI] RAW RPC" in line_upper
            or "[GUI] UPLOADED" in line_upper
            or "[GUI] REUSED CACHED" in line_upper
        ):
            return True
        # Keep CallHome / host-key progress (otherwise Start looks "stuck" after listen).
        if any(
            t in line_upper
            for t in (
                "HOST-KEY",
                "CALLHOME LISTEN",
                "CALLHOME TCP ACCEPTED",
                "LOGIN SUCCESSFUL",
                "LOGIN NOT ESTABLISHED",
                "ARE YOU SURE YOU WANT TO CONTINUE CONNECTING",
                "AUTHENTICITY OF THE HOST",
                "WAITING FOR NETOPEER AUTH",
            )
        ):
            return False

        if content_upper.startswith("[INFO]") or stripped_upper.startswith("[INFO]"):
            return True

        # Keep Conformance-runner / important verdict lines.
        if "[CONFORMANCE" in line_upper or "[CONFORMANCE-RUN]" in line_upper:
            return False
        if stripped_upper.startswith(("[TRACE:", "[WARN]", "[ERROR]", "[FAIL]", "[OK]")):
            return False
        if content_upper.startswith(("[WARN]", "[ERROR]", "[FAIL]", "[OK]")):
            return False
        if "CLIENT SENT" in line_upper:
            return False
        if "STEP " in line_upper and any(
            k in line_upper for k in ("CALLHOME", "LOGIN", "CRITERIA", "SUBSCRIPTION", "SUPERVISION")
        ):
            return False
        if "AUTHENTICATION SUCCESSFUL" in line_upper or "ACCEPTED A CONNECTION ON" in line_upper:
            return False
        if stripped_upper.strip() in ("OK", "NOK", "ERROR"):
            return False
        if re.search(
            r"\b(ERROR-TAG|ERROR-MESSAGE|ERROR-TYPE|ERROR-SEVERITY|BAD-ELEMENT|BAD-ATTRIBUTE|OPERATION-FAILED)\b",
            line_upper,
        ):
            return False
        if re.match(r"^\s*(type|tag|severity|path|message)\s*:", line, flags=re.IGNORECASE):
            return False

        # Otherwise keep only NETCONF-centric lines (historic log filter behaviour).
        keep_tokens = [
            "CLIENT_SENT:",
            "<RPC",
            "</RPC",
            "<RPC-REPLY",
            "</RPC-REPLY",
            "<HELLO",
            "</HELLO",
            "<NOTIFICATION",
            "</NOTIFICATION",
            "<DATA",
            "</DATA",
            "<OK/>",
            "<RPC-ERROR",
            "</RPC-ERROR",
            "<ERROR-",
            "</ERROR-",
            "<BAD-ELEMENT",
            "<BAD-ATTRIBUTE",
            "<ALARM-NOTIF",
            "</ALARM-NOTIF",
            "<FAULT-",
            "<IS-CLEARED",
            "<EVENT-TIME",
            "<AFFECTED-",
            "<FAULT-SEVERITY",
            "<FAULT-TEXT",
            "<SYNC",
            "<SYNCHRONIZATION",
            "</SYNC",
            "</SYNCHRONIZATION",
            "<SUPERVISION",
            "</SUPERVISION",
        ]
        if any(token in line_upper for token in keep_tokens):
            return False
        # Keep lines that look like XML element content (indented tags inside a block).
        s = line.strip()
        if s.startswith("<") and s.endswith(">"):
            return False
        return True

    def _refresh_log_target_hint_line(self) -> None:
        """로그 창 상단: Start( LOG_PATH ) vs Conformance( …/var/tmp/conformance/logs/… ) 안내."""
        v = self.log_target_hint_var
        if getattr(self, "_conformance_run_busy", False):
            cur = getattr(self, "_conformance_active_host_log", None)
            if cur:
                v.set(f"[Conformance] 원격 세션 로그 (tee): {cur}")
            else:
                v.set("[Conformance] 업로드/준비 중… (세션 로그 경로는 첫 스크립트 시작 시 표시)")
            return
        if getattr(self, "is_running", False):
            lp = (getattr(self, "_start_log_path_hint", None) or "").strip()
            if not lp:
                try:
                    w = self.fields.get("LOG_PATH")
                    lp = w.get().strip() if w is not None else ""
                except Exception:
                    lp = ""
            if lp:
                v.set(f"[Start / miniDU] 원격 로그 디렉터리 (LOG_PATH): {lp}")
            else:
                v.set("[Start / miniDU] 실행 중 — Settings의 LOG_PATH를 확인하세요.")
            return
        lp = ""
        try:
            w = self.fields.get("LOG_PATH")
            lp = w.get().strip() if w is not None else ""
        except Exception:
            pass
        prod = ""
        try:
            w = self.fields.get("PRODUCT")
            prod = w.get().strip() if w is not None else ""
        except Exception:
            pass
        safe = re.sub(r"[^0-9A-Za-z._-]+", "_", prod).strip("_") or "PRODUCT"
        lastc = getattr(self, "_conformance_last_host_log", None)
        last_txt = f"  |  마지막 Conf. 세션 로그: {lastc}" if lastc else ""
        if lp:
            v.set(
                f"[대기] Start → {lp}  |  Conformance tee 로그 → /var/tmp/conformance/logs/{safe}/CONF_{safe}_<yymmdd_HHMMSS>_<script>.log{last_txt}"
            )
        else:
            v.set(
                f"[대기] Start는 Settings LOG_PATH 필요  |  Conformance tee → /var/tmp/conformance/logs/{safe}/CONF_*.log (실행 시 생성){last_txt}"
            )

    # Pretty-print NETCONF XML in the log viewer; skip only on huge payloads to avoid UI freezes.
    _max_log_pretty_chars = 2 * 1024 * 1024

    def append_log(self, line: str) -> None:
        t0 = time.perf_counter()
        line = self._log_stream_carry + line
        self._log_stream_carry = ""
        if not line:
            return
        # Paramiko/tail can split one long <rpc-reply> line across reads; buffer until newline.
        if not line.endswith("\n"):
            last_nl = line.rfind("\n")
            if last_nl < 0:
                self._log_stream_carry = line
                return
            self._log_stream_carry = line[last_nl + 1 :]
            line = line[: last_nl + 1]

        line = self._sanitize_log_text(line)
        if not line:
            return

        kept_parts: list[str] = []
        for part in line.splitlines(keepends=True):
            part_u = part.upper()
            # M-Plane Test progress must always show (never swallow into RPC exchange buffer).
            if "[M-PLANE TEST]" in part_u:
                if self._rpc_exchange_collecting:
                    # keep collecting in background but still display this line
                    pass
                kept_parts.append(part)
                continue
            if "[GUI] NETCONF RPC EXCHANGE BEGIN" in part_u:
                self._rpc_exchange_collecting = True
                self._rpc_exchange_buf = []
                # Do not show the [GUI] begin marker.
                continue
            if self._rpc_exchange_collecting:
                if "[GUI] NETCONF RPC EXCHANGE END" in part_u:
                    if self._rpc_exchange_buf:
                        formatted = self._format_rpc_exchange_xml("".join(self._rpc_exchange_buf))
                        if formatted:
                            kept_parts.append(formatted)
                        self._rpc_exchange_buf = []
                    self._rpc_exchange_collecting = False
                    # Do not show the [GUI] end marker (may be glued onto a DEBUG line).
                    continue
                # Filter DEBUG/noise while collecting the exchange window.
                if self._should_hide_line(part):
                    continue
                self._rpc_exchange_buf.append(part)
                continue
            if self._should_hide_line(part):
                continue
            p = part
            s = p.strip()
            su = s.upper()
            if len(s) <= self._max_log_pretty_chars and s.startswith("<") and any(
                m in su
                for m in (
                    "<RPC",
                    "<RPC-REPLY",
                    "<NOTIFICATION",
                    "<DATA",
                    "<HELLO",
                    "<EDIT-CONFIG",
                )
            ):
                p = self._pretty_xml_if_possible(s)
            kept_parts.append(p)

        if not kept_parts:
            return
        merged = "".join(kept_parts)
        self._register_rpc_error_summary(merged)
        self._detect_remote_start_failure(merged)
        if self.is_running and not self._user_stop_requested:
            self._update_session_activity_from_log(merged)
            self._detect_session_lost(merged)
            self._detect_session_established(merged)
        with self.log_lock:
            self.log_buffer.append(merged)
            self._recent_log_for_session += merged
            if len(self._recent_log_for_session) > 128_000:
                self._recent_log_for_session = self._recent_log_for_session[-128_000:]
        self._perf_record("append_log", time.perf_counter() - t0)

    def _flush_log_buffer(self) -> None:
        t0 = time.perf_counter()
        chunk = ""
        with self.log_lock:
            if self.log_buffer:
                chunk = "".join(self.log_buffer)
                self.log_buffer.clear()
        if chunk:
            # If log window is not visible, defer rendering to keep UI responsive.
            if self.log is None or self.log_window is None or not self.log_window.winfo_exists():
                with self.log_lock:
                    self.hidden_log_chunks.append(chunk)
                self._perf_record("flush_log_buffer", time.perf_counter() - t0)
                self.after(400, self._flush_log_buffer)
                return
            if self.hidden_log_chunks:
                with self.log_lock:
                    chunk = "".join(self.hidden_log_chunks) + chunk
                    self.hidden_log_chunks.clear()
            self.log.configure(state="normal")
            at_bottom_before = self.log.yview()[1] >= 0.97
            self.log.insert("end", chunk)
            # Cap line count periodically (not every flush) for smoother rendering.
            self.flush_tick += 1
            if self.flush_tick % 12 == 0:
                self._trim_log_widget_to_limit()
            if at_bottom_before:
                self.log.see("end")
            self.log.configure(state="disabled")
        self._perf_record("flush_log_buffer", time.perf_counter() - t0)
        self.after(200, self._flush_log_buffer)

    def _flush_hidden_logs_to_widget(self) -> None:
        if self.hidden_render_active:
            return
        self.hidden_render_active = True
        self.after(0, self._render_hidden_logs_batch)

    def _render_hidden_logs_batch(self) -> None:
        if self.log is None or self.log_window is None or not self.log_window.winfo_exists():
            self.hidden_render_active = False
            return
        batch = ""
        with self.log_lock:
            if self.hidden_log_chunks:
                # Render in smaller batches to avoid tab-switch stalls.
                while self.hidden_log_chunks and len(batch) < 16000:
                    batch += self.hidden_log_chunks.pop(0)
        if not batch:
            self.hidden_render_active = False
            return
        self.log.configure(state="normal")
        self.log.insert("end", batch)
        self._trim_log_widget_to_limit()
        self.log.see("end")
        self.log.configure(state="disabled")
        self.after(18, self._render_hidden_logs_batch)

    def _perf_record(self, key: str, elapsed_sec: float) -> None:
        if not self.perf_debug_var.get():
            return
        self.perf_stats[key] = self.perf_stats.get(key, 0.0) + elapsed_sec
        self.perf_counts[key] = self.perf_counts.get(key, 0) + 1
        self.perf_last[key] = elapsed_sec
        self.perf_max[key] = max(self.perf_max.get(key, 0.0), elapsed_sec)
        now = time.time()
        if now - self.perf_last_report < 5.0:
            return
        self.perf_last_report = now
        parts = []
        for k in sorted(self.perf_stats.keys()):
            count = max(1, self.perf_counts.get(k, 1))
            avg_ms = (self.perf_stats[k] / count) * 1000.0
            last_ms = self.perf_last.get(k, 0.0) * 1000.0
            max_ms = self.perf_max.get(k, 0.0) * 1000.0
            parts.append(f"{k}:avg={avg_ms:.2f}ms last={last_ms:.2f}ms max={max_ms:.2f}ms n={count}")
        if parts:
            text = "[PERF] " + " | ".join(parts)
            self.after(0, lambda t=text: self.perf_text_var.set(t))

    def _trace(self, event: str, **fields: str) -> int:
        if not self.perf_debug_var.get():
            return -1
        self.trace_seq += 1
        trace_id = self.trace_seq
        kv = " ".join(f"{k}={v}" for k, v in fields.items())
        self.append_log(f"[TRACE:{trace_id}] {event} {kv}\n")
        return trace_id

    def clear_log(self) -> None:
        if self.log is None:
            self.open_log_window()
        with self.log_lock:
            self.log_buffer.clear()
            self.hidden_log_chunks.clear()
            self._log_stream_carry = ""
            self._rpc_exchange_collecting = False
            self._rpc_exchange_buf = []
        self.rpc_error_items.clear()
        self.rpc_error_seen.clear()
        self._refresh_rpc_error_summary_widget()
        self._refresh_rpc_event_list_widget()
        if self.log is not None:
            self.log.configure(state="normal")
            self.log.delete("1.0", "end")
            self.log.configure(state="disabled")

    def _trim_log_widget_to_limit(self) -> None:
        if self.log is None or not self.log.winfo_exists():
            return
        limit = max(50, min(int(self.max_log_lines), 200_000))
        self.log.configure(state="normal")
        current_lines = int(self.log.index("end-1c").split(".")[0])
        if current_lines > limit:
            delete_to = current_lines - limit
            self.log.delete("1.0", f"{delete_to}.0")
        self.log.configure(state="disabled")

    def apply_gui_log_line_limit(self) -> None:
        raw = self.gui_log_max_lines_var.get().strip().replace(",", "")
        try:
            n = int(raw)
        except ValueError:
            messagebox.showerror("Invalid value", "Enter an integer line count (e.g. 5000).")
            return
        n = max(100, min(n, 100_000))
        self.max_log_lines = n
        self.gui_log_max_lines_var.set(str(n))
        self._trim_log_widget_to_limit()
        self._save_current_config()
        self.status_var.set(f"GUI log buffer: keep last {n} lines")

    def _load_log_text_direct(self, text: str) -> None:
        # Used by "Log Load": render fetched remote log as-is (no NETCONF filter).
        self.clear_log()
        self.open_log_window()
        self._register_rpc_error_summary(text)
        if self.log is None:
            with self.log_lock:
                self.hidden_log_chunks.append(text)
            return
        self.log.configure(state="normal")
        self.log.insert("1.0", text)
        self._trim_log_widget_to_limit()
        self.log.see("end")
        self.log.configure(state="disabled")

    def _open_text_in_notepad(self, text: str) -> None:
        try:
            tmp_dir = Path(tempfile.gettempdir())
            out_path = tmp_dir / f"netconf_log_load_{int(time.time() * 1000)}.log"
            out_path.write_text(text, encoding="utf-8", errors="ignore")
            if os.name == "nt":
                subprocess.Popen(["notepad.exe", str(out_path)])
            else:
                messagebox.showinfo("Log Load", f"Saved log file:\n{out_path}")
            self.status_var.set(f"Log opened: {out_path.name}")
        except Exception as exc:
            messagebox.showerror("Error", f"Failed to open log in notepad:\n{exc}")

    def _collect_current_config(self) -> dict[str, Any]:
        message_data = []
        for tab in self.message_tabs:
            message_data.append(
                {
                    "title": tab["title_var"].get(),
                    "content": tab["text"].get("1.0", "end").strip(),
                }
            )
        omit = getattr(self, "_conformance_omit_last_run_from_config_save", False)
        if not omit:
            self._conformance_refresh_last_run_cache_from_progress()
        else:
            self._conformance_last_run_snapshot_cache = None
        payload: dict[str, Any] = {
            "window_geometry": self.geometry(),
            "log_window_geometry": self.log_window_geometry,
            "rpc_event_window_geometry": self.rpc_event_window_geometry,
            "remote_user": self.remote_user.get(),
            "remote_host": self.remote_host.get(),
            "remote_port": self.remote_port.get(),
            "remote_password": self.remote_password.get(),
            "remote_script_path": self.remote_script_path.get(),
            "remote_key_path": self.remote_key_path.get(),
            "perf_debug": bool(self.perf_debug_var.get()),
            "auto_xml_send": bool(self.auto_xml_send_var.get()),
            "auto_start": bool(self.auto_start_var.get()),
            "gui_max_log_lines": int(self.max_log_lines),
            "runtime_fields": {k: v.get() for k, v in self.fields.items()},
            "mplane_xlsx_path": self.mplane_xlsx_path.get(),
            "mplane_fields": {k: v.get() for k, v in self.mplane_fields.items()},
            "mplane_cc_on": [bool(v.get()) for v in self.mplane_cc_on_vars],
            "netconf_library_path": (
                str(self.netconf_library_root.resolve())
                if self.netconf_library_root and self.netconf_library_root.exists()
                else ""
            ),
            "conformance_checked": {k: bool(v.get()) for k, v in self.conformance_check_vars.items()},
            "conformance_reboot_checked": {k: bool(v.get()) for k, v in self.conformance_reboot_vars.items()},
            "conformance_reboot_wait_sec": self.conformance_reboot_wait_var.get(),
            "conformance_run_remote_dir": self.conformance_run_remote_dir_var.get(),
            "conformance_run_sw_pkg": self.conformance_run_sw_pkg_var.get(),
            "conformance_run_sw_remote_dir": self.conformance_run_sw_remote_dir_var.get(),
            "conformance_restart_start_after_run": bool(self.conformance_restart_start_after_run_var.get()),
            "conformance_debug": bool(self.conformance_debug_var.get()),
            "conformance_run_rpc_timeout": self.conformance_run_rpc_timeout_var.get(),
            "conformance_run_idle_timeout": self.conformance_run_idle_timeout_var.get(),
            "conformance_run_supervision_interval": self.conformance_run_supervision_interval_var.get(),
            "conformance_run_supervision_reset_cycles": self.conformance_run_supervision_reset_cycles_var.get(),
            "conformance_run_supervision_negative_fail_cycle": self.conformance_run_supervision_negative_fail_cycle_var.get(),
            "conformance_run_conn_delay": self.conformance_run_conn_delay_var.get(),
            "conformance_post_listen_wait": self.conformance_post_listen_wait_var.get(),
            "conformance_run_repeat": self.conformance_run_repeat_var.get(),
            "conformance_per_test_settings": copy.deepcopy(self._conformance_per_test_settings),
            "conformance_extra_uploads": [
                {"local": a, "remote": b} for a, b in self._conformance_extra_uploads
            ],
            "message_tabs": message_data,
            "profiles": copy.deepcopy(getattr(self, "_profiles", {}) or {}),
            "active_profile": (self.profile_name_var.get() or "").strip(),
            "guardrails_checked": {k: bool(v.get()) for k, v in self.guardrails_check_vars.items()},
            "guardrails_last_results": {k: v.get() for k, v in self.guardrails_result_vars.items()},
            "guardrails_run_repeat": self.guardrails_run_repeat_var.get(),
            "guardrails_per_test_settings": copy.deepcopy(
                getattr(self, "_guardrails_per_test_settings", {}) or {}
            ),
        }
        cl = self._conformance_last_run_snapshot_cache
        if not omit and isinstance(cl, dict) and cl.get("by_script"):
            payload["conformance_last_run"] = cl
        fr = getattr(self, "_conformance_final_results", None)
        if not omit and isinstance(fr, dict) and fr:
            payload["conformance_final_results"] = copy.deepcopy(fr)
        return payload

    def _apply_config(self, data: dict[str, Any]) -> None:
        geometry = data.get("window_geometry")
        if isinstance(geometry, str) and geometry.strip():
            try:
                self.geometry(geometry.strip())
            except Exception:
                pass
        log_geometry = data.get("log_window_geometry")
        if isinstance(log_geometry, str) and log_geometry.strip():
            self.log_window_geometry = log_geometry.strip()
        rpc_event_geometry = data.get("rpc_event_window_geometry")
        if isinstance(rpc_event_geometry, str) and rpc_event_geometry.strip():
            self.rpc_event_window_geometry = rpc_event_geometry.strip()
        self.exec_mode.set("remote")
        self.remote_user.set(str(data.get("remote_user", self.remote_user.get())))
        self.remote_host.set(str(data.get("remote_host", self.remote_host.get())))
        self.remote_port.set(str(data.get("remote_port", self.remote_port.get())))
        self.remote_password.set(str(data.get("remote_password", self.remote_password.get())))
        self.remote_script_path.set("/var/tmp/miniDU_callhome.sh")
        self.remote_key_path.set(str(data.get("remote_key_path", self.remote_key_path.get())))
        self.send_mode_var.set("raw_rpc")
        self.perf_debug_var.set(bool(data.get("perf_debug", self.perf_debug_var.get())))
        self.auto_xml_send_var.set(bool(data.get("auto_xml_send", self.auto_xml_send_var.get())))
        self.auto_start_var.set(bool(data.get("auto_start", self.auto_start_var.get())))
        glm = data.get("gui_max_log_lines")
        if glm is not None:
            try:
                n = int(glm)
                n = max(100, min(n, 100_000))
                self.max_log_lines = n
                self.gui_log_max_lines_var.set(str(n))
            except (TypeError, ValueError):
                pass
        if not self.perf_debug_var.get():
            self.perf_text_var.set("PERF: off")

        runtime_fields = data.get("runtime_fields", {})
        if isinstance(runtime_fields, dict):
            for key, value in runtime_fields.items():
                if key in self.fields:
                    self.fields[key].set(str(value))
        mxp = data.get("mplane_xlsx_path")
        if isinstance(mxp, str) and mxp.strip():
            self.mplane_xlsx_path.set(mxp.strip())
        mplane_saved = data.get("mplane_fields")
        if isinstance(mplane_saved, dict):
            for key, value in mplane_saved.items():
                var = self.mplane_fields.get(str(key))
                if var is not None:
                    var.set("" if value is None else str(value))
        cc_on_saved = data.get("mplane_cc_on")
        if isinstance(cc_on_saved, list) and self.mplane_cc_on_vars:
            for i, on in enumerate(cc_on_saved):
                if i < len(self.mplane_cc_on_vars):
                    self.mplane_cc_on_vars[i].set(bool(on))
            try:
                self._mplane_apply_off_row_styles()
            except Exception:
                pass
        nlp = data.get("netconf_library_path")
        if isinstance(nlp, str) and nlp.strip():
            self._netconf_library_path_to_restore = nlp.strip()
        conf_chk = data.get("conformance_checked")
        if isinstance(conf_chk, dict) and self.conformance_check_vars:
            for fname, val in conf_chk.items():
                bv = self.conformance_check_vars.get(str(fname))
                if bv is not None:
                    bv.set(bool(val))
        conf_rb = data.get("conformance_reboot_checked")
        if isinstance(conf_rb, dict) and self.conformance_reboot_vars:
            for fname, val in conf_rb.items():
                bv = self.conformance_reboot_vars.get(str(fname))
                if bv is not None:
                    bv.set(bool(val))
        crw = data.get("conformance_reboot_wait_sec")
        if isinstance(crw, str) and crw.strip():
            self.conformance_reboot_wait_var.set(crw.strip())
        elif crw is not None:
            try:
                self.conformance_reboot_wait_var.set(str(int(crw)))
            except (TypeError, ValueError):
                pass
        self._conformance_apply_last_run_from_config(data.get("conformance_last_run"))
        self._conformance_apply_final_results_from_config(data.get("conformance_final_results"))
        crd = data.get("conformance_run_remote_dir")
        if isinstance(crd, str) and crd.strip():
            self.conformance_run_remote_dir_var.set(crd.strip())
        csp = data.get("conformance_run_sw_pkg")
        if isinstance(csp, str):
            self.conformance_run_sw_pkg_var.set(csp)
        csr = data.get("conformance_run_sw_remote_dir")
        if isinstance(csr, str) and csr.strip():
            self.conformance_run_sw_remote_dir_var.set(csr.strip())
        crs2 = data.get("conformance_restart_start_after_run")
        if crs2 is not None:
            try:
                self.conformance_restart_start_after_run_var.set(bool(crs2))
            except (tk.TclError, ValueError):
                pass
        cd = data.get("conformance_debug")
        if cd is not None:
            try:
                self.conformance_debug_var.set(bool(cd))
            except (tk.TclError, ValueError):
                pass
        for var_key, cfg_key in (
            (self.conformance_run_rpc_timeout_var, "conformance_run_rpc_timeout"),
            (self.conformance_run_idle_timeout_var, "conformance_run_idle_timeout"),
            (self.conformance_run_supervision_interval_var, "conformance_run_supervision_interval"),
            (self.conformance_run_supervision_reset_cycles_var, "conformance_run_supervision_reset_cycles"),
            (self.conformance_run_supervision_negative_fail_cycle_var, "conformance_run_supervision_negative_fail_cycle"),
            (self.conformance_run_conn_delay_var, "conformance_run_conn_delay"),
            (self.conformance_post_listen_wait_var, "conformance_post_listen_wait"),
            (self.conformance_run_repeat_var, "conformance_run_repeat"),
        ):
            v = data.get(cfg_key)
            if isinstance(v, str) and v.strip():
                var_key.set(v.strip())
        pts = data.get("conformance_per_test_settings")
        if isinstance(pts, dict):
            for fname, vals in pts.items():
                if isinstance(vals, dict):
                    self._conformance_per_test_settings[str(fname)] = {
                        str(k): str(v) for k, v in vals.items()
                    }
            try:
                self._conformance_reconcile_per_test_settings()
            except Exception:
                pass
        ex = data.get("conformance_extra_uploads")
        if isinstance(ex, list):
            self._conformance_extra_uploads = []
            for it in ex:
                if isinstance(it, dict) and "local" in it and "remote" in it:
                    self._conformance_extra_uploads.append((str(it["local"]), str(it["remote"])))
        tabs = data.get("message_tabs", [])
        if isinstance(tabs, list) and tabs:
            while self.message_tabs:
                self.msg_notebook.forget(0)
                self.message_tabs.pop(0)
            for t in tabs:
                self.add_message_tab(initial_title=str(t.get("title", "MSG")))
                self.message_tabs[-1]["text"].delete("1.0", "end")
                self.message_tabs[-1]["text"].insert("1.0", str(t.get("content", "")))
        profs = data.get("profiles")
        if isinstance(profs, dict):
            cleaned: dict[str, Any] = {}
            for name, body in profs.items():
                n = str(name).strip()
                if n and isinstance(body, dict):
                    cleaned[n] = body
            self._profiles = cleaned
        ap = data.get("active_profile")
        if isinstance(ap, str) and ap.strip():
            self.profile_name_var.set(ap.strip())
        self._profile_refresh_combo()
        gchk = data.get("guardrails_checked")
        if isinstance(gchk, dict) and self.guardrails_check_vars:
            for key, val in gchk.items():
                bv = self.guardrails_check_vars.get(str(key))
                if bv is not None:
                    bv.set(bool(val))
        gres = data.get("guardrails_last_results")
        if isinstance(gres, dict) and self.guardrails_result_vars:
            for key, val in gres.items():
                rv = self.guardrails_result_vars.get(str(key))
                if rv is not None:
                    rv.set("" if val is None else str(val))
        grr = data.get("guardrails_run_repeat")
        if isinstance(grr, (str, int)) and str(grr).strip() != "":
            try:
                self.guardrails_run_repeat_var.set(str(grr).strip())
            except Exception:
                pass
        gfields = data.get("guardrails_per_test_settings")
        if not isinstance(gfields, dict):
            # migrate legacy flat guardrails_fields → dhcp_boot_shared
            legacy = data.get("guardrails_fields")
            if isinstance(legacy, dict):
                gfields = {"dhcp_boot_shared": {str(k): str(v) for k, v in legacy.items() if v is not None}}
        if isinstance(gfields, dict):
            store = getattr(self, "_guardrails_per_test_settings", None)
            if store is None:
                self._guardrails_per_test_settings = {}
                store = self._guardrails_per_test_settings
            for iid, vals in gfields.items():
                if isinstance(vals, dict):
                    store[str(iid)] = {str(k): str(v) for k, v in vals.items() if v is not None}
        for iid in list(getattr(self, "guardrails_check_vars", {}) or {}):
            try:
                self._guardrails_sync_tree_row(str(iid))
            except Exception:
                pass
        try:
            sync = getattr(self, "_sync_log_path_from_product", None)
            if callable(sync):
                sync()
        except Exception:
            pass
        try:
            self._apply_lab_controller_listen_ips("untag")
        except Exception:
            pass

    def _profile_snapshot(self) -> dict[str, Any]:
        """Network/lab-related settings switched between 10.0.20 ↔ 10.0.60 (and similar)."""
        return {
            "remote_user": self.remote_user.get(),
            "remote_host": self.remote_host.get(),
            "remote_port": self.remote_port.get(),
            "remote_password": self.remote_password.get(),
            "remote_script_path": self.remote_script_path.get(),
            "runtime_fields": {k: v.get() for k, v in self.fields.items()},
            "mplane_fields": {k: v.get() for k, v in self.mplane_fields.items()},
            "conformance_run_remote_dir": self.conformance_run_remote_dir_var.get(),
            "conformance_run_sw_pkg": self.conformance_run_sw_pkg_var.get(),
            "conformance_run_sw_remote_dir": self.conformance_run_sw_remote_dir_var.get(),
            "conformance_per_test_settings": copy.deepcopy(
                getattr(self, "_conformance_per_test_settings", {}) or {}
            ),
            "guardrails_per_test_settings": copy.deepcopy(
                getattr(self, "_guardrails_per_test_settings", {}) or {}
            ),
        }

    def _profile_apply_snapshot(self, snap: dict[str, Any]) -> None:
        if not isinstance(snap, dict):
            return
        was = getattr(self, "_config_hydrating", False)
        self._config_hydrating = True
        try:
            if "remote_user" in snap:
                self.remote_user.set(str(snap.get("remote_user", "")))
            if "remote_host" in snap:
                self.remote_host.set(str(snap.get("remote_host", "")))
            if "remote_port" in snap:
                self.remote_port.set(str(snap.get("remote_port", "")))
            if "remote_password" in snap:
                self.remote_password.set(str(snap.get("remote_password", "")))
            # REMOTE_SCRIPT_PATH 고정
            self.remote_script_path.set("/var/tmp/miniDU_callhome.sh")
            rf = snap.get("runtime_fields")
            if isinstance(rf, dict):
                for key, value in rf.items():
                    if key in self.fields:
                        self.fields[key].set("" if value is None else str(value))
            mf = snap.get("mplane_fields")
            if isinstance(mf, dict):
                for key, value in mf.items():
                    var = self.mplane_fields.get(str(key))
                    if var is not None:
                        var.set("" if value is None else str(value))
            for var, key in (
                (self.conformance_run_remote_dir_var, "conformance_run_remote_dir"),
                (self.conformance_run_sw_pkg_var, "conformance_run_sw_pkg"),
                (self.conformance_run_sw_remote_dir_var, "conformance_run_sw_remote_dir"),
            ):
                v = snap.get(key)
                if isinstance(v, str):
                    var.set(v)
            pts = snap.get("conformance_per_test_settings")
            if isinstance(pts, dict):
                for fname, vals in pts.items():
                    if isinstance(vals, dict):
                        self._conformance_per_test_settings[str(fname)] = {
                            str(k): str(v) for k, v in vals.items()
                        }
                try:
                    self._conformance_reconcile_per_test_settings()
                except Exception:
                    pass
            gf = snap.get("guardrails_per_test_settings")
            if not isinstance(gf, dict):
                legacy = snap.get("guardrails_fields")
                if isinstance(legacy, dict):
                    gf = {"dhcp_boot_shared": {str(k): str(v) for k, v in legacy.items() if v is not None}}
            if isinstance(gf, dict):
                store = getattr(self, "_guardrails_per_test_settings", None)
                if store is None:
                    self._guardrails_per_test_settings = {}
                    store = self._guardrails_per_test_settings
                for iid, vals in gf.items():
                    if isinstance(vals, dict):
                        store[str(iid)] = {str(k): str(v) for k, v in vals.items() if v is not None}
        finally:
            self._config_hydrating = was
            try:
                sync = getattr(self, "_sync_log_path_from_product", None)
                if callable(sync):
                    sync()
            except Exception:
                pass
            try:
                self._apply_lab_controller_listen_ips("untag")
            except Exception:
                pass

    def _profile_refresh_combo(self) -> None:
        names = sorted(self._profiles.keys(), key=str.lower)
        if "lab-20" not in names:
            names = ["lab-20", *names]
        if "lab-60" not in names:
            # keep lab-60 near top for 20/60 switching
            if names and names[0] == "lab-20":
                names = ["lab-20", "lab-60", *[n for n in names if n not in ("lab-20", "lab-60")]]
            else:
                names = ["lab-60", *names]
        # unique preserve order
        seen: set[str] = set()
        ordered: list[str] = []
        for n in names:
            if n not in seen:
                seen.add(n)
                ordered.append(n)
        if self.profile_combo is not None:
            self.profile_combo["values"] = tuple(ordered)
        cur = (self.profile_name_var.get() or "").strip()
        if cur not in ordered and ordered:
            self.profile_name_var.set(ordered[0])

    def _profile_load_selected(self) -> None:
        name = (self.profile_name_var.get() or "").strip()
        if not name:
            messagebox.showwarning("Profile", "프로파일 이름을 입력하거나 선택하세요.")
            return
        snap = self._profiles.get(name)
        if not isinstance(snap, dict):
            messagebox.showwarning(
                "Profile",
                f"'{name}' 프로파일이 아직 없습니다.\n현재 설정을 맞춘 뒤 Save / Save As로 저장하세요.",
            )
            return
        self._profile_apply_snapshot(snap)
        self.append_log(f"[GUI] Profile loaded: {name}\n")
        self._save_current_config()
        messagebox.showinfo("Profile", f"프로파일 '{name}' 을(를) 불러왔습니다.")

    def _profile_save_selected(self) -> None:
        name = (self.profile_name_var.get() or "").strip()
        if not name:
            messagebox.showwarning("Profile", "프로파일 이름을 입력하세요.")
            return
        self._profiles[name] = self._profile_snapshot()
        self._profile_refresh_combo()
        self._save_current_config()
        self.append_log(f"[GUI] Profile saved: {name}\n")
        messagebox.showinfo("Profile", f"프로파일 '{name}' 에 현재 설정을 저장했습니다.")

    def _profile_save_as(self) -> None:
        name = simpledialog.askstring("Save Profile As", "새 프로파일 이름:", parent=self)
        if name is None:
            return
        name = name.strip()
        if not name:
            messagebox.showwarning("Profile", "이름이 비어 있습니다.")
            return
        if name in self._profiles:
            if not messagebox.askyesno("Profile", f"'{name}' 이(가) 이미 있습니다. 덮어쓸까요?"):
                return
        self.profile_name_var.set(name)
        self._profiles[name] = self._profile_snapshot()
        self._profile_refresh_combo()
        self._save_current_config()
        self.append_log(f"[GUI] Profile saved as: {name}\n")
        messagebox.showinfo("Profile", f"프로파일 '{name}' 으로 저장했습니다.")

    def _profile_delete_selected(self) -> None:
        name = (self.profile_name_var.get() or "").strip()
        if not name:
            return
        if name not in self._profiles:
            messagebox.showwarning("Profile", f"'{name}' 은(는) 저장된 프로파일이 아닙니다.")
            return
        if not messagebox.askyesno("Profile", f"프로파일 '{name}' 을(를) 삭제할까요?"):
            return
        del self._profiles[name]
        self._profile_refresh_combo()
        self._save_current_config()
    _GUARDRAILS_DHCP_SHARED_KEY = "dhcp_boot_shared"
    _GUARDRAILS_RU_SSH_FIELDS: list[dict[str, Any]] = [
        {
            "key": "oru_cli_id",
            "label": "RU SSH ID",
            "default": "",
            "hint": "Settings ★ RU SSH ID 사용 (숨김)",
            "wide": False,
            "hidden": True,
        },
        {
            "key": "oru_cli_pw",
            "label": "RU SSH PW",
            "default": "",
            "hint": "Settings ★ RU SSH PW 사용 (숨김)",
            "password": True,
            "wide": False,
            "hidden": True,
        },
        {
            "key": "probe_v4",
            "label": "RU IPv4",
            "default": "",
            "hint": "Settings ALLOWED_IP 사용 (숨김)",
            "wide": True,
            "hidden": True,
        },
        {
            "key": "probe_v6",
            "label": "RU IPv6",
            "default": "",
            "hint": "Settings ALLOWED_IP_V6 사용 (숨김)",
            "wide": True,
            "hidden": True,
        },
        {
            "key": "ru_mac",
            "label": "RU MAC",
            "default": "",
            "hint": "시험 family IP ping→neigh 자동 (v4=ALLOWED_IP / v6=ALLOWED_IP_V6, 숨김)",
            "wide": True,
            "hidden": True,
        },
        {
            "key": "mplane_if",
            "label": "M-Plane NIC",
            "default": "",
            "hint": "solid NIC (fe80%scope). 비우면 LOCAL_IF",
            "wide": False,
        },
        {
            "key": "ssh_family",
            "label": "SSH family",
            "default": "v4",
            "hint": "v4 또는 v6",
            "wide": False,
        },
    ]
    _GUARDRAILS_DHCP_COMMON_FIELDS: list[dict[str, Any]] = [
        {"key": "l2sw_ip", "label": "L2SW IP", "default": "", "hint": "다산 M3500 관리 IP", "wide": True},
        {
            "key": "l2sw_id",
            "label": "L2SW ID",
            "default": "",
            "hint": "고정 사용 — UI 숨김",
            "wide": False,
            "hidden": True,
        },
        {
            "key": "l2sw_pw",
            "label": "L2SW PW",
            "default": "",
            "hint": "고정 사용 — UI 숨김",
            "password": True,
            "wide": False,
            "hidden": True,
        },
        {"key": "l2sw_if", "label": "L2SW IF", "default": "", "hint": "RU MAC → show mac 자동 (숨김)", "wide": False, "hidden": True},
        {"key": "acl_num", "label": "ACL #", "default": "110", "hint": "포트에만 bind (ACL 본문은 스위치에 유지)", "wide": False},
        {
            "key": "ru_mac",
            "label": "RU MAC",
            "default": "",
            "hint": "시험 family IP ping→neigh 자동 (v4=ALLOWED_IP / v6=ALLOWED_IP_V6, 숨김)",
            "wide": True,
            "hidden": True,
        },
        {
            "key": "mplane_if",
            "label": "M-Plane NIC",
            "default": "",
            "hint": "solid NIC (fe80%scope). 비우면 LOCAL_IF",
            "wide": False,
        },
        {
            "key": "oru_cli_id",
            "label": "RU SSH ID",
            "default": "",
            "hint": "Settings ★ RU SSH ID 사용 (숨김)",
            "wide": False,
            "hidden": True,
        },
        {
            "key": "oru_cli_pw",
            "label": "RU SSH PW",
            "default": "",
            "hint": "Settings ★ RU SSH PW 사용 (숨김)",
            "password": True,
            "wide": False,
            "hidden": True,
        },
        {
            "key": "ru_if_name",
            "label": "RU IF name",
            "default": "",
            "hint": "검사 iface (예: eth0). 비우면 전체",
            "wide": False,
        },
        {
            "key": "probe_v4",
            "label": "RU IPv4",
            "default": "",
            "hint": "Settings ALLOWED_IP 사용 (숨김)",
            "wide": True,
            "hidden": True,
        },
        {
            "key": "probe_v6",
            "label": "RU IPv6",
            "default": "",
            "hint": "Settings ALLOWED_IP_V6 사용 (숨김)",
            "wide": True,
            "hidden": True,
        },
        {"key": "pass_sec", "label": "PASS ≤초", "default": "240", "hint": "이하면 PASS", "wide": False},
        {"key": "timeout_sec", "label": "Timeout 초", "default": "540", "hint": "복구 대기 한도", "wide": False},
        {"key": "poll_sec", "label": "RU확인주기", "default": "5", "hint": "초 · SSH up/down 검사 주기", "wide": False},
        {
            "key": "stable_sec",
            "label": "복구유지 초",
            "default": "10",
            "hint": "SSH healthy 연속 유지 초 (한 번 성공만으로 PASS 안 함)",
            "wide": False,
            "hidden": True,
        },
        {"key": "down_detect_sec", "label": "Down감지 초", "default": "180", "hint": "재부팅(비정상) 대기", "wide": False},
    ]
    _GUARDRAILS_DHCP_CAPTURE_FIELDS: list[dict[str, Any]] = [
        {
            "key": "dhcp_host",
            "label": "DHCP SSH host",
            "default": "",
            "hint": "비우면 Settings 원격(solid)에서 tcpdump. 있으면 solid→이 호스트 SSH",
            "wide": True,
        },
        {
            "key": "dhcp_id",
            "label": "DHCP SSH ID",
            "default": "",
            "hint": "dhcp_host 쓸 때 필수",
            "wide": False,
        },
        {
            "key": "dhcp_pw",
            "label": "DHCP SSH PW",
            "default": "",
            "hint": "SSH + sudo -S 공통(vlan100 ip addr add). 비우면 sudo -n만 시도→실패 많음",
            "password": True,
            "wide": False,
        },
        {
            "key": "dhcp_port",
            "label": "DHCP SSH port",
            "default": "22",
            "hint": "22 고정 — UI 숨김",
            "wide": False,
            "hidden": True,
        },
        {
            "key": "dhcp_if",
            "label": "Capture IF",
            "default": "",
            "hint": "dhcp_host ifconfig 10.0.60/61 자동 (숨김)",
            "wide": False,
            "hidden": True,
        },
        {
            "key": "use_sudo",
            "label": "sudo tcpdump",
            "default": "1",
            "hint": "1=sudo -n tcpdump (NOPASSWD 필요)",
            "wide": False,
            "hidden": True,
        },
        {
            "key": "capture_sec",
            "label": "Capture 초",
            "default": "300",
            "hint": "기본 300 — UI 숨김",
            "wide": False,
            "hidden": True,
        },
        {
            "key": "require_reboot",
            "label": "Reboot prompt",
            "default": "1",
            "hint": "1=캡처 시작 후 재부팅 안내",
            "wide": False,
            "hidden": True,  # dhcp 통합시험은 자동 reset — UI 불필요
        },
        {
            "key": "ru_mac",
            "label": "RU MAC",
            "default": "",
            "hint": "시험 family IP ping→neigh 자동 (v4=ALLOWED_IP / v6=ALLOWED_IP_V6, 숨김)",
            "wide": True,
            "hidden": True,
        },
        {
            "key": "reset_mode",
            "label": "Reset mode",
            "default": "auto",
            "hint": "auto|mplane=기존 conformance_oru_reboot.sh 호출(수정 없음) / manual",
            "wide": False,
            "hidden": True,  # 현장은 항상 auto — UI 불필요
        },
    ]
    _GUARDRAILS_VLAN_DISC_FIELDS: list[dict[str, Any]] = [
        {
            "key": "vlan_discovery_vid",
            "label": "★ 시험 VLAN ID",
            "default": "100",
            "hint": "RU 변경 시 필수 — 예: 100",
            "wide": False,
            "emphasize": True,
        },
        {
            "key": "vlan_discovery_base_vid",
            "label": "원복(base) VLAN ID",
            "default": "1",
            "hint": "reset 직후 trunk remove, 원복 시 다시 add (기본 1)",
            "wide": False,
            "hidden": True,
        },
        {
            "key": "vlan_discovery_solid_parent",
            "label": "DHCP호스트 VLAN parent",
            "default": "",
            "hint": "비우면 Capture IF → vlan<VID> 임시 생성",
            "wide": False,
            "hidden": True,
        },
        {
            "key": "vlan_discovery_solid_cidr",
            "label": "DHCP호스트 VLAN IP",
            "default": "10.0.61.252/24",
            "hint": "반드시 /24 (10.0.61.252/24). /32·마스크생략 시 ping이 GW로 나감",
            "wide": True,
            "hidden": True,
        },
        {
            "key": "vlan_discovery_solid_cidr_v6",
            "label": "DHCP호스트 VLAN IPv6",
            "default": "2001:1300:1100:1000::252/64",
            "hint": "GUI 자동: tag=1300::252 / untag=1200::252 (대상 /64 우선)",
            "wide": True,
            "hidden": True,  # 시험 상황별 자동 — UI 불필요
        },
        {
            "key": "vlan_discovery_name",
            "label": "시험 VLAN 이름",
            "default": "",
            "hint": "비우면 이름 생략",
            "wide": False,
            "hidden": True,
        },
        {
            "key": "dhcp_renew_cmd",
            "label": "DHCP renew 명령",
            "default": "dhcp vlan-discovery renew force",
            "hint": "원복 시 RU SSH 명령 (소문자)",
            "wide": True,
        },
    ]
    _GUARDRAILS_DHCP_L2SW_FIELDS: list[dict[str, Any]] = [
        {"key": "l2sw_ip", "label": "L2SW IP", "default": "", "hint": "다산 M3500", "wide": True},
        {
            "key": "l2sw_id",
            "label": "L2SW ID",
            "default": "",
            "hint": "고정 — 숨김",
            "wide": False,
            "hidden": True,
        },
        {
            "key": "l2sw_pw",
            "label": "L2SW PW",
            "default": "",
            "hint": "고정 — 숨김",
            "password": True,
            "wide": False,
            "hidden": True,
        },
        {
            "key": "l2sw_if",
            "label": "L2SW IF",
            "default": "",
            "hint": "RU MAC → show mac 자동 (숨김)",
            "wide": False,
            "hidden": True,
        },
        {
            "key": "acl_num",
            "label": "ACL #",
            "default": "110",
            "hint": "extended 100-199",
            "wide": False,
            "hidden": True,
        },
    ]
    _GUARDRAILS_DHCP_COMMON_HIDDEN: list[dict[str, Any]] = [
        {
            "key": "oru_cli_id",
            "label": "RU SSH ID",
            "default": "",
            "hint": "Settings ★ RU SSH ID 사용 (숨김)",
            "wide": False,
            "hidden": True,
        },
        {
            "key": "oru_cli_pw",
            "label": "RU SSH PW",
            "default": "",
            "hint": "Settings ★ RU SSH PW 사용 (숨김)",
            "password": True,
            "wide": False,
            "hidden": True,
        },
        {
            "key": "probe_v4",
            "label": "RU IPv4",
            "default": "",
            "hint": "Settings ALLOWED_IP 사용 (숨김)",
            "wide": True,
            "hidden": True,
        },
        {
            "key": "probe_v6",
            "label": "RU IPv6",
            "default": "",
            "hint": "Settings ALLOWED_IP_V6 사용 (숨김)",
            "wide": True,
            "hidden": True,
        },
        {
            "key": "mplane_if",
            "label": "M-Plane NIC",
            "default": "",
            "hint": "fe80%scope (비우면 LOCAL_IF)",
            "wide": False,
            "hidden": True,
        },
        {"key": "pass_sec", "label": "PASS ≤초", "default": "240", "hint": "주소 복구 PASS 한도", "wide": False},
        {"key": "timeout_sec", "label": "Timeout 초", "default": "540", "hint": "복구 대기 한도", "wide": False},
        {"key": "poll_sec", "label": "RU확인주기", "default": "5", "hint": "초 · SSH up/down 검사 주기", "wide": False},
        {
            "key": "stable_sec",
            "label": "복구유지 초",
            "default": "10",
            "hint": "SSH healthy 연속 유지 (기본 10s)",
            "wide": False,
            "hidden": True,
        },
        {"key": "down_detect_sec", "label": "Down감지 초", "default": "180", "hint": "재부팅 감지", "wide": False},
    ]
    _GUARDRAILS_PER_TEST_SCHEMA: dict[str, dict[str, Any]] = {
        "dhcp_v4": {
            "title": "DHCP v4 Boot — 재시작 후 IPv4 재수신",
            "settings_key": "dhcp_v4",
            "family": "v4",
            "mode": "boot",
            "fields": list(_GUARDRAILS_DHCP_CAPTURE_FIELDS) + list(_GUARDRAILS_DHCP_COMMON_HIDDEN) + [
                {
                    "key": "healthy_regex_v4",
                    "label": "Healthy regex v4",
                    "default": r"inet\s+\d+\.\d+\.\d+\.\d+/",
                    "hint": "있으면 표시용. 없어도 SSH 성공이면 PASS",
                    "wide": True,
                    "hidden": True,
                },
                {
                    "key": "option_regex",
                    "label": "Option regex (60)",
                    "default": r"(?i)(Option\s*60|Vendor-Class|vendor.class)",
                    "hint": "tcpdump -vv",
                    "wide": True,
                    "hidden": True,
                },
                {
                    "key": "enable_discovery",
                    "label": "Enable Discovery",
                    "default": "0",
                    "hint": "1=Option 43/Controller도",
                    "wide": False,
                    "hidden": True,
                },
                {
                    "key": "discovery_regex",
                    "label": "Discovery regex (43)",
                    "default": r"(?i)(Option\s*43|Vendor-Specific|vendor.specific)",
                    "hint": "",
                    "wide": True,
                    "hidden": True,
                },
                {
                    "key": "expected_controllers",
                    "label": "Controller IPs",
                    "default": "",
                    "hint": "비우면 LOCAL_IP",
                    "wide": True,
                    "hidden": True,
                },
            ],
        },
        "dhcp_v4_vlan": {
            "title": "DHCP v4 VLAN Discovery — L2SW trunk/ACL",
            "settings_key": "dhcp_v4_vlan",
            "family": "v4",
            "mode": "vlan",
            "fields": list(_GUARDRAILS_DHCP_CAPTURE_FIELDS)
            + list(_GUARDRAILS_VLAN_DISC_FIELDS)
            + list(_GUARDRAILS_DHCP_L2SW_FIELDS)
            + list(_GUARDRAILS_DHCP_COMMON_HIDDEN)
            + [
                {
                    "key": "healthy_regex_v4",
                    "label": "Healthy regex v4",
                    "default": r"inet\s+\d+\.\d+\.\d+\.\d+/",
                    "hint": "있으면 표시용. 없어도 SSH 성공이면 PASS",
                    "wide": True,
                    "hidden": True,
                },
                {
                    "key": "option_regex",
                    "label": "Option regex (60)",
                    "default": r"(?i)(Option\s*60|Vendor-Class|vendor.class)",
                    "hint": "tcpdump -vv",
                    "wide": True,
                    "hidden": True,
                },
                {
                    "key": "enable_discovery",
                    "label": "Enable Discovery",
                    "default": "0",
                    "hint": "1=Option 43/Controller도",
                    "wide": False,
                    "hidden": True,
                },
                {
                    "key": "discovery_regex",
                    "label": "Discovery regex (43)",
                    "default": r"(?i)(Option\s*43|Vendor-Specific|vendor.specific)",
                    "hint": "",
                    "wide": True,
                    "hidden": True,
                },
                {
                    "key": "expected_controllers",
                    "label": "Controller IPs",
                    "default": "",
                    "hint": "비우면 LOCAL_IP",
                    "wide": True,
                    "hidden": True,
                },
            ],
        },
        "dhcp_v6": {
            "title": "DHCP v6 Boot — 재시작 후 IPv6 재수신",
            "settings_key": "dhcp_v6",
            "family": "v6",
            "mode": "boot",
            "fields": list(_GUARDRAILS_DHCP_CAPTURE_FIELDS) + list(_GUARDRAILS_DHCP_COMMON_HIDDEN) + [
                {
                    "key": "healthy_regex",
                    "label": "Healthy regex v6",
                    "default": r"inet6\s+[0-9a-fA-F:]+/",
                    "hint": "있으면 표시용. 없어도 SSH 성공이면 PASS",
                    "wide": True,
                    "hidden": True,
                },
                {
                    "key": "option_regex",
                    "label": "Option regex (16)",
                    "default": r"(?i)(Option\s*16|vendor.class|Vendor Class)",
                    "hint": "tcpdump -vv",
                    "wide": True,
                    "hidden": True,
                },
                {
                    "key": "enable_ia_na",
                    "label": "Check IA_NA (Opt3)",
                    "default": "1",
                    "hint": "v6 필수 검사 — UI 숨김",
                    "wide": False,
                    "hidden": True,
                },
                {
                    "key": "ia_na_regex",
                    "label": "IA_NA regex",
                    "default": r"(?i)(IA[_-]?NA|Identity Association for Non-temporary)",
                    "hint": "tcpdump/Wireshark Solicit Option 3",
                    "wide": True,
                    "hidden": True,
                },
                {
                    "key": "enable_discovery",
                    "label": "Enable Discovery",
                    "default": "0",
                    "hint": "1=Option 17/Controller도",
                    "wide": False,
                    "hidden": True,
                },
                {
                    "key": "discovery_regex",
                    "label": "Discovery regex (17)",
                    "default": r"(?i)(Option\s*17|vendor.opts|Vendor-opts|vendor.options)",
                    "hint": "",
                    "wide": True,
                    "hidden": True,
                },
                {
                    "key": "expected_controllers",
                    "label": "Controller IPs",
                    "default": "",
                    "hint": "비우면 LOCAL_IP",
                    "wide": True,
                    "hidden": True,
                },
            ],
        },
        "dhcp_v6_vlan": {
            "title": "DHCP v6 VLAN Discovery — L2SW trunk/ACL",
            "settings_key": "dhcp_v6_vlan",
            "family": "v6",
            "mode": "vlan",
            "fields": list(_GUARDRAILS_DHCP_CAPTURE_FIELDS)
            + list(_GUARDRAILS_VLAN_DISC_FIELDS)
            + list(_GUARDRAILS_DHCP_L2SW_FIELDS)
            + list(_GUARDRAILS_DHCP_COMMON_HIDDEN)
            + [
                {
                    "key": "healthy_regex",
                    "label": "Healthy regex v6",
                    "default": r"inet6\s+[0-9a-fA-F:]+/",
                    "hint": "있으면 표시용. 없어도 SSH 성공이면 PASS",
                    "wide": True,
                    "hidden": True,
                },
                {
                    "key": "option_regex",
                    "label": "Option regex (16)",
                    "default": r"(?i)(Option\s*16|vendor.class|Vendor Class)",
                    "hint": "tcpdump -vv",
                    "wide": True,
                    "hidden": True,
                },
                {
                    "key": "enable_ia_na",
                    "label": "Check IA_NA (Opt3)",
                    "default": "1",
                    "hint": "v6 필수 검사 — UI 숨김",
                    "wide": False,
                    "hidden": True,
                },
                {
                    "key": "ia_na_regex",
                    "label": "IA_NA regex",
                    "default": r"(?i)(IA[_-]?NA|Identity Association for Non-temporary)",
                    "hint": "tcpdump/Wireshark Solicit Option 3",
                    "wide": True,
                    "hidden": True,
                },
                {
                    "key": "enable_discovery",
                    "label": "Enable Discovery",
                    "default": "0",
                    "hint": "1=Option 17/Controller도",
                    "wide": False,
                    "hidden": True,
                },
                {
                    "key": "discovery_regex",
                    "label": "Discovery regex (17)",
                    "default": r"(?i)(Option\s*17|vendor.opts|Vendor-opts|vendor.options)",
                    "hint": "",
                    "wide": True,
                    "hidden": True,
                },
                {
                    "key": "expected_controllers",
                    "label": "Controller IPs",
                    "default": "",
                    "hint": "비우면 LOCAL_IP",
                    "wide": True,
                    "hidden": True,
                },
            ],
        },
        # legacy
        "dhcp_options": {
            "title": "DHCP Options (legacy)",
            "settings_key": "dhcp_options",
            "family": "v4",
            "fields": list(_GUARDRAILS_DHCP_CAPTURE_FIELDS),
        },
        "dhcp_boot": {
            "title": "DHCP boot (legacy combined)",
            "settings_key": "dhcp_boot_shared",
            "family": "v4",
            "fields": list(_GUARDRAILS_DHCP_COMMON_FIELDS),
        },
        "dhcp_v4_only_boot": {
            "title": "DHCPv4-only (포트 DHCPv6 drop)",
            "settings_key": "dhcp_boot_shared",
            "family": "v4",
            "fields": list(_GUARDRAILS_DHCP_COMMON_FIELDS) + [
                {
                    "key": "healthy_regex_v4",
                    "label": "Healthy regex",
                    "default": r"inet\s+\d+\.\d+\.\d+\.\d+/",
                    "hint": "SSH 출력 — IPv4 inet",
                    "wide": True,
                },
            ],
        },
        "dhcp_v6_only_boot": {
            "title": "DHCPv6-only (포트 DHCPv4 drop)",
            "settings_key": "dhcp_boot_shared",
            "family": "v6",
            "fields": list(_GUARDRAILS_DHCP_COMMON_FIELDS) + [
                {
                    "key": "healthy_regex",
                    "label": "Healthy regex",
                    "default": r"inet6\s+[0-9a-fA-F:]+/",
                    "hint": "SSH 출력 — global inet6",
                    "wide": True,
                },
            ],
        },
        "vlan_discovery": {
            "title": "M-Plane VLAN Discovery",
            "settings_key": "vlan_discovery",
            "family": "v4",
            "fields": list(_GUARDRAILS_RU_SSH_FIELDS) + [
                {
                    "key": "vid_min",
                    "label": "VID min",
                    "default": "1",
                    "hint": "vlan-discovery 범위 하한",
                    "wide": False,
                },
                {
                    "key": "vid_max",
                    "label": "VID max",
                    "default": "4094",
                    "hint": "vlan-discovery 범위 상한",
                    "wide": False,
                },
                {
                    "key": "expected_vid",
                    "label": "Expected VID",
                    "default": "",
                    "hint": "부팅 후 확인할 VLAN ID (필수)",
                    "wide": False,
                },
                {
                    "key": "pre_cmds",
                    "label": "Pre cmds (재부팅 전)",
                    "default": "",
                    "hint": "RU에서 vid range 설정 CLI",
                    "wide": True,
                },
                {
                    "key": "check_cmd",
                    "label": "RU check cmd",
                    "default": (
                        "ip -d link; ip -4 -o addr; "
                        "show yang-module-data o-ran-interfaces; "
                        "show interfaces"
                    ),
                    "hint": "재부팅 후 VLAN/주소 확인",
                    "wide": True,
                },
                {
                    "key": "expect_regex",
                    "label": "Expect regex",
                    "default": "",
                    "hint": "비우면 expected_vid 자동 패턴",
                    "wide": True,
                },
                {
                    "key": "require_reboot",
                    "label": "Reboot wait",
                    "default": "1",
                    "hint": "1=SSH down→up 대기 후 검사, 0=즉시 검사",
                    "wide": False,
                },
                {"key": "pass_sec", "label": "PASS ≤초", "default": "240", "hint": "복구 한도(재부팅 시)", "wide": False},
                {"key": "timeout_sec", "label": "Timeout 초", "default": "540", "hint": "복구 타임아웃", "wide": False},
                {"key": "poll_sec", "label": "RU확인주기", "default": "5", "hint": "초 · SSH up/down 검사 주기", "wide": False},
                {"key": "down_detect_sec", "label": "Down감지 초", "default": "180", "hint": "", "wide": False},
            ],
        },
        "netconf_capability": {
            "title": "NETCONF capability discovery",
            "settings_key": "netconf_capability",
            "family": "v4",
            "fields": list(_GUARDRAILS_RU_SSH_FIELDS) + [
                {
                    "key": "note",
                    "label": "상태",
                    "default": "추후 구현 (목록 등록만)",
                    "hint": "Conformance 외 capability 세부 검증 예정",
                    "wide": True,
                },
            ],
        },
        "config_admin_state": {
            "title": "admin-state",
            "settings_key": "config_admin_state",
            "family": "v4",
            "fields": list(_GUARDRAILS_RU_SSH_FIELDS) + [
                {
                    "key": "note",
                    "label": "상태",
                    "default": "추후 구현 (목록 등록만)",
                    "hint": "admin-state 전이 검증 예정",
                    "wide": True,
                },
            ],
        },
        "config_oper_state": {
            "title": "oper-state",
            "settings_key": "config_oper_state",
            "family": "v4",
            "fields": list(_GUARDRAILS_RU_SSH_FIELDS) + [
                {
                    "key": "note",
                    "label": "상태",
                    "default": "추후 구현 (목록 등록만)",
                    "hint": "oper-state 검증 예정",
                    "wide": True,
                },
            ],
        },
        "config_availability_state": {
            "title": "availability-state",
            "settings_key": "config_availability_state",
            "family": "v4",
            "fields": list(_GUARDRAILS_RU_SSH_FIELDS) + [
                {
                    "key": "note",
                    "label": "상태",
                    "default": "추후 구현 (목록 등록만)",
                    "hint": "availability-state 검증 예정",
                    "wide": True,
                },
            ],
        },
        "config_usage_state": {
            "title": "usage-state",
            "settings_key": "config_usage_state",
            "family": "v4",
            "fields": list(_GUARDRAILS_RU_SSH_FIELDS) + [
                {
                    "key": "note",
                    "label": "상태",
                    "default": "추후 구현 (목록 등록만)",
                    "hint": "usage-state 검증 예정",
                    "wide": True,
                },
            ],
        },
        # legacy combined (catalog 미등록)
        "config_states": {
            "title": "admin/oper/availability/usage-state (legacy)",
            "settings_key": "config_states",
            "family": "v4",
            "fields": list(_GUARDRAILS_RU_SSH_FIELDS) + [
                {
                    "key": "note",
                    "label": "상태",
                    "default": "분리됨 — config_admin/oper/availability/usage_state 사용",
                    "hint": "legacy",
                    "wide": True,
                },
            ],
        },
        "fault_alarm": {
            "title": "Fault Alarm (test-alarm batch)",
            "settings_key": "fault_alarm",
            "family": "v4",
            "fields": list(_GUARDRAILS_RU_SSH_FIELDS) + [
                {
                    "key": "source_id",
                    "label": "source-id",
                    "default": "0",
                    "hint": "CLI source-id (템플릿 {source_id})",
                    "wide": False,
                },
                {
                    "key": "show_cmd",
                    "label": "조회 CLI",
                    "default": "show alarm information oran",
                    "hint": "전체 Alarm Id 목록 조회 (전부 시험)",
                    "wide": True,
                },
                {
                    "key": "active_show_cmd",
                    "label": "Active 조회 CLI",
                    "default": "show alarm active-alarms",
                    "hint": "시험 전 기존 active 요약만 로그",
                    "wide": True,
                },
                {
                    "key": "raise_tmpl",
                    "label": "발생 CLI",
                    "default": "test alarm alarm-id {alarm_id} source-id {source_id} start-alarm",
                    "hint": "{alarm_id}|{fault_id} {source_id} 치환",
                    "wide": True,
                },
                {
                    "key": "clear_tmpl",
                    "label": "클리어 CLI",
                    "default": "no test alarm alarm-id {alarm_id} source-id {source_id}",
                    "hint": "cancel = no test alarm …",
                    "wide": True,
                },
                {
                    "key": "alarm_timeout_sec",
                    "label": "noti 대기 초",
                    "default": "60",
                    "hint": "alarm당 raise/clear 각각 noti 대기",
                    "wide": False,
                },
                {
                    "key": "listen_timeout_sec",
                    "label": "CallHome 대기 초",
                    "default": "180",
                    "hint": "세션 listen 타임아웃",
                    "wide": False,
                },
                {
                    "key": "require_noti",
                    "label": "noti 필수",
                    "default": "1",
                    "hint": "1=raise/clear alarm-notif 없으면 FAIL",
                    "wide": False,
                },
                {
                    "key": "skip_normal",
                    "label": "NORMAL 제외",
                    "default": "1",
                    "hint": "1=Fault Id 0 / NORMAL 행 스킵",
                    "wide": False,
                },
            ],
        },
        "performance_mgmt": {
            "title": "Performance Management",
            "settings_key": "performance_mgmt",
            "family": "v4",
            "fields": list(_GUARDRAILS_RU_SSH_FIELDS) + [
                {
                    "key": "note",
                    "label": "상태",
                    "default": "추후 구현 (O-RAN Player 연동)",
                    "hint": "PM activation/reporting — Player와 함께 실행",
                    "wide": True,
                },
            ],
        },
    }

    def _guardrails_catalog(self) -> list[dict[str, str]]:
        """M-Plane Test items (TLS/CCM/802.1X/Management-8 제외)."""
        items: list[dict[str, str]] = [
            {
                "id": "dhcp_v4",
                "scope": "DHCP",
                "ref": "MP-DHCPv4-Boot",
                "title": "DHCP v4 Boot (재시작 후 IPv4 재수신)",
                "detail": "ACL로 v6 DHCP 차단(단일 v4) → 재부팅→IPv4 복구 + Option 60. 연속 시험 가능.",
            },
            {
                "id": "dhcp_v4_vlan",
                "scope": "DHCP",
                "ref": "MP-DHCPv4-VLAN",
                "title": "DHCP v4 VLAN Discovery (L2SW trunk/ACL)",
                "detail": "ACL + vlan DB/trunk → Discovery → renew/원복. 연속 시험 가능.",
            },
            {
                "id": "dhcp_v6",
                "scope": "DHCP",
                "ref": "MP-DHCPv6-Boot",
                "title": "DHCP v6 Boot (재시작 후 IPv6 재수신)",
                "detail": "ACL로 v4 DHCP 차단(단일 v6) → 재부팅→inet6 복구 + Option 16·IA_NA. 연속 시험 가능.",
            },
            {
                "id": "dhcp_v6_vlan",
                "scope": "DHCP",
                "ref": "MP-DHCPv6-VLAN",
                "title": "DHCP v6 VLAN Discovery (L2SW trunk/ACL)",
                "detail": "ACL + vlan DB/trunk → Discovery → renew/원복. 연속 시험 가능.",
            },
            {
                "id": "netconf_capability",
                "scope": "NETCONF",
                "ref": "MP-CAP-1",
                "title": "NETCONF capability (세부)",
                "detail": "목록 등록. 실행은 추후 (yang-library/xpath/rollback 등).",
            },
            {
                "id": "config_admin_state",
                "scope": "Config",
                "ref": "MP-STATE-admin",
                "title": "admin-state",
                "detail": "목록 등록. 실행은 추후. (필요 시 다른 state와 통합 가능)",
            },
            {
                "id": "config_oper_state",
                "scope": "Config",
                "ref": "MP-STATE-oper",
                "title": "oper-state",
                "detail": "목록 등록. 실행은 추후.",
            },
            {
                "id": "config_availability_state",
                "scope": "Config",
                "ref": "MP-STATE-avail",
                "title": "availability-state",
                "detail": "목록 등록. 실행은 추후.",
            },
            {
                "id": "config_usage_state",
                "scope": "Config",
                "ref": "MP-STATE-usage",
                "title": "usage-state",
                "detail": "목록 등록. 실행은 추후.",
            },
            {
                "id": "fault_alarm",
                "scope": "FM",
                "ref": "MP-FM-alarm",
                "title": "Fault Alarm (세션 유지 · 일괄 raise/clear)",
                "detail": "show alarm information oran → Alarm Id로 test start/cancel + noti. 상세에 조회·발생 정리.",
            },
            {
                "id": "performance_mgmt",
                "scope": "PM",
                "ref": "MP-PM-1",
                "title": "Performance Management (O-RAN Player)",
                "detail": "맨 마지막 실행. O-RAN Player와 연동 후 구현.",
            },
        ]
        items.extend(getattr(self, "_guardrails_user_items", None) or [])
        return items

    def _guardrails_store_key(self, item_id: str) -> str:
        schema = self._GUARDRAILS_PER_TEST_SCHEMA.get(item_id) or {}
        return str(schema.get("settings_key") or item_id)

    def _guardrails_item_family(self, item_id: str) -> str:
        schema = self._GUARDRAILS_PER_TEST_SCHEMA.get(item_id) or {}
        return str(schema.get("family") or "v6").lower()

    def _guardrails_item_mode(self, item_id: str) -> str:
        """boot = L2SW 없이 재시작+IP / vlan = L2SW VLAN Discovery."""
        schema = self._GUARDRAILS_PER_TEST_SCHEMA.get(item_id) or {}
        mode = str(schema.get("mode") or "").strip().lower()
        if mode in ("boot", "vlan"):
            return mode
        if str(item_id or "").endswith("_vlan"):
            return "vlan"
        return "boot"

    _GUARDRAILS_DHCP_ITEM_IDS: tuple[str, ...] = (
        "dhcp_v4",
        "dhcp_v4_vlan",
        "dhcp_v6",
        "dhcp_v6_vlan",
    )

    def _guardrails_ssh_exec(self, command: str, timeout: int = 25) -> tuple[bool, str]:
        ssh_user = self.remote_user.get().strip()
        ssh_host = self.remote_host.get().strip()
        ssh_port = self.remote_port.get().strip() or "22"
        ssh_password = self.remote_password.get()
        if not ssh_user or not ssh_host:
            return False, "Settings SSH_USER/SSH_HOST 필요"
        try:
            import paramiko  # type: ignore
        except Exception as exc:
            return False, f"paramiko 필요: {exc}"
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:
            client.connect(
                hostname=ssh_host,
                port=int(ssh_port),
                username=ssh_user,
                password=ssh_password or None,
                timeout=12,
                allow_agent=True,
                look_for_keys=True,
            )
            _stdin, stdout, stderr = client.exec_command(command, timeout=timeout)
            out = (stdout.read() or b"").decode(errors="replace")
            err = (stderr.read() or b"").decode(errors="replace")
            rc = stdout.channel.recv_exit_status()
            text = (out + (("\n" + err) if err.strip() else "")).strip()
            return rc == 0, text or f"(exit {rc})"
        except Exception as exc:
            return False, str(exc)
        finally:
            try:
                client.close()
            except Exception:
                pass

    # Keys that must stay family-specific (v4 Option60 vs v6 Option16 / IA_NA 등)
    _GUARDRAILS_DHCP_FAMILY_LOCAL_KEYS = frozenset(
        {
            "option_regex",
            "discovery_regex",
            "healthy_regex",
            "healthy_regex_v4",
            "enable_ia_na",
            "ia_na_regex",
            "enable_discovery",
            "expected_controllers",
        }
    )

    def _guardrails_dhcp_sibling_key(self, item_id: str) -> str:
        """Opposite-family primary key (boot) — UI note용. 미러는 peer_keys 사용."""
        fam = self._guardrails_item_family(item_id)
        return "dhcp_v6" if fam == "v4" else "dhcp_v4"

    def _guardrails_dhcp_peer_keys(self, item_id: str) -> list[str]:
        """All other DHCP store keys for shared-field mirroring."""
        sk = self._guardrails_store_key(item_id)
        peers: list[str] = []
        for iid in self._GUARDRAILS_DHCP_ITEM_IDS:
            psk = self._guardrails_store_key(iid)
            if psk and psk != sk and psk not in peers:
                peers.append(psk)
        return peers

    def _guardrails_dhcp_same_family_keys(self, item_id: str) -> list[str]:
        fam = self._guardrails_item_family(item_id)
        sk = self._guardrails_store_key(item_id)
        out: list[str] = []
        for iid in self._GUARDRAILS_DHCP_ITEM_IDS:
            if self._guardrails_item_family(iid) != fam:
                continue
            psk = self._guardrails_store_key(iid)
            if psk and psk != sk and psk not in out:
                out.append(psk)
        return out

    def _guardrails_get_val(self, item_id: str, key: str, default: str = "") -> str:
        store = getattr(self, "_guardrails_per_test_settings", None) or {}
        sk = self._guardrails_store_key(item_id)
        family_local = key in self._GUARDRAILS_DHCP_FAMILY_LOCAL_KEYS
        candidates: tuple[str, ...] = (sk, item_id)
        if family_local:
            # same-family boot↔vlan 공유 (Option regex 등)
            extra_f = self._guardrails_dhcp_same_family_keys(item_id)
            if extra_f:
                candidates = (sk, item_id, *extra_f)
        else:
            peers = self._guardrails_dhcp_peer_keys(item_id)
            extra = [*peers, "dhcp_capture_shared", "dhcp_boot_shared", "dhcp_v4", "dhcp_v6"]
            # de-dupe preserve order
            seen: set[str] = set()
            ordered: list[str] = []
            for c in (sk, item_id, *extra):
                if c and c not in seen:
                    seen.add(c)
                    ordered.append(c)
            candidates = tuple(ordered)
        for candidate in candidates:
            if not candidate:
                continue
            cur = store.get(candidate) or {}
            raw = cur.get(key)
            if raw is not None and str(raw).strip() != "":
                return str(raw).strip()
        schema = self._GUARDRAILS_PER_TEST_SCHEMA.get(item_id) or {}
        for field in schema.get("fields") or []:
            if field.get("key") == key:
                return str(field.get("default") or default).strip() or default
        return default

    def _guardrails_set_vals(self, item_id: str, vals: dict[str, str]) -> None:
        if not hasattr(self, "_guardrails_per_test_settings") or self._guardrails_per_test_settings is None:
            self._guardrails_per_test_settings = {}
        clean = {str(k): str(v) for k, v in vals.items() if str(v).strip() != ""}
        sk = self._guardrails_store_key(item_id)
        self._guardrails_per_test_settings[sk] = clean
        # keep legacy key in sync for older configs
        if sk != item_id:
            self._guardrails_per_test_settings[item_id] = dict(clean)
        # DHCP 4항목: shared → 전체 peer / family-local → 같은 family boot↔vlan
        for peer in self._guardrails_dhcp_peer_keys(item_id):
            peer_blob = dict(self._guardrails_per_test_settings.get(peer) or {})
            for k, v in clean.items():
                if k in self._GUARDRAILS_DHCP_FAMILY_LOCAL_KEYS:
                    continue
                peer_blob[k] = v
            self._guardrails_per_test_settings[peer] = {
                kk: vv for kk, vv in peer_blob.items() if str(vv).strip() != ""
            }
        for peer in self._guardrails_dhcp_same_family_keys(item_id):
            peer_blob = dict(self._guardrails_per_test_settings.get(peer) or {})
            for k, v in clean.items():
                if k not in self._GUARDRAILS_DHCP_FAMILY_LOCAL_KEYS:
                    continue
                peer_blob[k] = v
            self._guardrails_per_test_settings[peer] = {
                kk: vv for kk, vv in peer_blob.items() if str(vv).strip() != ""
            }
        # Settings ALLOWED_IP* / CLI-ID·PW 가 SSOT — ⚙ 로 Settings 를 덮지 않음
        try:
            self._guardrails_sync_probe_from_settings(item_id)
        except Exception:
            pass
        try:
            self._guardrails_sync_oru_cli_from_settings(item_id)
        except Exception:
            pass
        try:
            self._on_any_setting_changed()
        except Exception:
            pass
        try:
            self._save_current_config()
        except Exception:
            pass

    def _guardrails_sync_probe_from_settings(self, item_id: str | None = None) -> None:
        """Settings ALLOWED_IP* → 내부 probe (항상 Settings 우선)."""
        iid = item_id or getattr(self, "_guardrails_settings_item_id", None) or "dhcp_v4"
        sk = self._guardrails_store_key(iid)
        if not hasattr(self, "_guardrails_per_test_settings") or self._guardrails_per_test_settings is None:
            self._guardrails_per_test_settings = {}
        cur = dict(self._guardrails_per_test_settings.get(sk) or {})
        try:
            a4 = self._guardrails_strip_ip_cidr(
                (self.fields.get("ALLOWED_IP").get() or "").strip()  # type: ignore[union-attr]
            )
        except Exception:
            a4 = ""
        try:
            a6 = self._guardrails_strip_ip_cidr(
                (self.fields.get("ALLOWED_IP_V6").get() or "").strip()  # type: ignore[union-attr]
            )
        except Exception:
            a6 = ""
        if a4:
            old = self._guardrails_strip_ip_cidr(cur.get("probe_v4") or "")
            cur["probe_v4"] = a4
            if old != a4:
                self._guardrails_log(
                    f"{iid}: RU IPv4 ← Settings ALLOWED_IP ({old or '-'} → {a4})"
                )
        if a6 and ":" in a6:
            old = self._guardrails_strip_ip_cidr(cur.get("probe_v6") or "")
            cur["probe_v6"] = a6
            if old != a6:
                self._guardrails_log(
                    f"{iid}: RU IPv6 ← Settings ALLOWED_IP_V6 ({old or '-'} → {a6})"
                )
        if not (a4 or (a6 and ":" in a6)):
            return
        self._guardrails_per_test_settings[sk] = {
            k: v for k, v in cur.items() if str(v).strip()
        }
        sib = self._guardrails_dhcp_sibling_key(iid)
        if sib:
            sib_blob = dict(self._guardrails_per_test_settings.get(sib) or {})
            if a4:
                sib_blob["probe_v4"] = a4
            if a6 and ":" in a6:
                sib_blob["probe_v6"] = a6
            self._guardrails_per_test_settings[sib] = {
                k: v for k, v in sib_blob.items() if str(v).strip()
            }

    def _guardrails_sync_oru_cli_from_settings(self, item_id: str | None = None) -> None:
        """Settings ★ RU SSH ID/PW (CLI-ID/CLI-PW) → 내부 oru_cli_* (전 시험 공용 SSOT)."""
        iid = item_id or getattr(self, "_guardrails_settings_item_id", None) or "dhcp_v4"
        sk = self._guardrails_store_key(iid)
        if not hasattr(self, "_guardrails_per_test_settings") or self._guardrails_per_test_settings is None:
            self._guardrails_per_test_settings = {}
        cur = dict(self._guardrails_per_test_settings.get(sk) or {})
        try:
            cid = (self.fields.get("CLI-ID").get() or "").strip()  # type: ignore[union-attr]
        except Exception:
            cid = ""
        try:
            cpw = (self.fields.get("CLI-PW").get() or "").strip()  # type: ignore[union-attr]
        except Exception:
            cpw = ""
        if not cid and not cpw:
            return
        if cid:
            old = (cur.get("oru_cli_id") or "").strip()
            cur["oru_cli_id"] = cid
            if old != cid:
                self._guardrails_log(
                    f"{iid}: RU SSH ID ← Settings CLI-ID ({old or '-'} → {cid})"
                )
        if cpw:
            old = (cur.get("oru_cli_pw") or "").strip()
            cur["oru_cli_pw"] = cpw
            if old != cpw:
                self._guardrails_log(f"{iid}: RU SSH PW ← Settings CLI-PW (updated)")
        self._guardrails_per_test_settings[sk] = {
            k: v for k, v in cur.items() if str(v).strip()
        }
        sib = self._guardrails_dhcp_sibling_key(iid)
        if sib:
            sib_blob = dict(self._guardrails_per_test_settings.get(sib) or {})
            if cid:
                sib_blob["oru_cli_id"] = cid
            if cpw:
                sib_blob["oru_cli_pw"] = cpw
            self._guardrails_per_test_settings[sib] = {
                k: v for k, v in sib_blob.items() if str(v).strip()
            }

    def _guardrails_gf(self, key: str, default: str = "") -> str:
        iid = getattr(self, "_guardrails_settings_item_id", None) or "dhcp_v4"
        return self._guardrails_get_val(iid, key, default)

    def _guardrails_int(self, key: str, default: int) -> int:
        raw = self._guardrails_gf(key, str(default))
        try:
            return max(1, int(raw))
        except ValueError:
            return default

    def _guardrails_fill_defaults_from_context(self, item_id: str | None = None) -> None:
        """Prefill empty per-test fields from Settings / Conformance 3151 (in memory only)."""
        iid = item_id or getattr(self, "_guardrails_settings_item_id", None) or "dhcp_v4"
        if not hasattr(self, "_guardrails_per_test_settings") or self._guardrails_per_test_settings is None:
            self._guardrails_per_test_settings = {}
        sk = self._guardrails_store_key(iid)
        # migrate old per-item blob → shared / merged keys
        if sk not in self._guardrails_per_test_settings:
            for legacy_key in ("dhcp_v6_only_boot", "dhcp_v4_only_boot", "dhcp_boot_shared"):
                legacy = self._guardrails_per_test_settings.get(legacy_key)
                if isinstance(legacy, dict) and legacy and sk == "dhcp_boot_shared":
                    self._guardrails_per_test_settings[sk] = dict(legacy)
                    break
            if sk in ("dhcp_v4", "dhcp_v6", "dhcp_v4_vlan", "dhcp_v6_vlan", "dhcp_capture_shared"):
                merged: dict[str, str] = {}
                for legacy_key in (
                    "dhcp_capture_shared",
                    "dhcp_v4",
                    "dhcp_v6",
                    "dhcp_v4_vlan",
                    "dhcp_v6_vlan",
                    "dhcp_options",
                    "dhcp_option_identity",
                    "dhcp_controller_discovery",
                ):
                    if legacy_key == sk:
                        continue
                    legacy = self._guardrails_per_test_settings.get(legacy_key)
                    if isinstance(legacy, dict):
                        for k, v in legacy.items():
                            if v is not None and k not in merged:
                                merged[str(k)] = str(v)
                if merged.get("identity_regex") and not merged.get("option_regex"):
                    merged["option_regex"] = merged["identity_regex"]
                if merged.get("expect_regex") and not merged.get("option_regex"):
                    merged["option_regex"] = merged["expect_regex"]
                if merged.get("controller_regex") and not merged.get("discovery_regex"):
                    merged["discovery_regex"] = merged["controller_regex"]
                if merged and sk not in self._guardrails_per_test_settings:
                    self._guardrails_per_test_settings[sk] = merged
        cur = dict(self._guardrails_per_test_settings.get(sk) or {})
        changed = False

        def _field(name: str) -> str:
            try:
                return (self.fields.get(name).get() or "").strip()  # type: ignore[union-attr]
            except Exception:
                return ""

        # sync capture/L2SW/VLAN/timing across dhcp_* items (+ legacy blobs)
        if sk in self._GUARDRAILS_DHCP_ITEM_IDS or sk in ("dhcp_v4", "dhcp_v6", "dhcp_v4_vlan", "dhcp_v6_vlan"):
            srcs = list(self._GUARDRAILS_DHCP_ITEM_IDS) + [
                "dhcp_capture_shared",
                "dhcp_boot_shared",
                "dhcp_v4_only_boot",
                "dhcp_v6_only_boot",
            ]
            for src in srcs:
                if src == sk:
                    continue
                other = dict(self._guardrails_per_test_settings.get(src) or {})
                for k in (
                    "dhcp_host",
                    "dhcp_id",
                    "dhcp_pw",
                    "dhcp_port",
                    "dhcp_if",
                    "use_sudo",
                    "capture_sec",
                    "require_reboot",
                    "ru_mac",
                    "l2sw_ip",
                    "l2sw_id",
                    "l2sw_pw",
                    "l2sw_if",
                    "acl_num",
                    "vlan_discovery_vid",
                    "vlan_discovery_base_vid",
                    "vlan_discovery_solid_parent",
                    "vlan_discovery_solid_cidr",
                    "vlan_discovery_solid_cidr_v6",
                    "vlan_discovery_name",
                    "dhcp_renew_cmd",
                    "oru_cli_id",
                    "oru_cli_pw",
                    "probe_v4",
                    "probe_v6",
                    "mplane_if",
                    "pass_sec",
                    "timeout_sec",
                    "poll_sec",
                    "stable_sec",
                    "down_detect_sec",
                    "healthy_regex",
                    "healthy_regex_v4",
                ):
                    if not (cur.get(k) or "").strip() and (other.get(k) or "").strip():
                        cur[k] = other[k]
                        changed = True

        if sk in ("dhcp_v4_vlan", "dhcp_v6_vlan") and (cur.get("acl_num") or "").strip() != "110":
            cur["acl_num"] = "110"
            changed = True
        if not (cur.get("dhcp_if") or "").strip():
            lif = _field("LOCAL_IF")
            if lif:
                cur["dhcp_if"] = lif
                changed = True
        if not (cur.get("mplane_if") or "").strip():
            lif = _field("LOCAL_IF")
            if lif:
                cur["mplane_if"] = lif
                changed = True
        # Settings ALLOWED_IP* 가 SSOT — 비어 있을 때만이 아니라 항상 Settings 로 probe 맞춤
        allowed = self._guardrails_strip_ip_cidr(_field("ALLOWED_IP"))
        if allowed:
            old = self._guardrails_strip_ip_cidr(cur.get("probe_v4") or "")
            if old != allowed:
                cur["probe_v4"] = allowed
                changed = True
            else:
                cur["probe_v4"] = allowed
        allowed6 = self._guardrails_strip_ip_cidr(_field("ALLOWED_IP_V6"))
        if allowed6 and ":" in allowed6:
            old = self._guardrails_strip_ip_cidr(cur.get("probe_v6") or "")
            if old != allowed6:
                cur["probe_v6"] = allowed6
                changed = True
            else:
                cur["probe_v6"] = allowed6
        # dhcp_v6: never keep DHCPv4 Option 60/43 regex (common after v4↔v6 ⚙ 공유)
        fam_now = ""
        try:
            fam_now = str((self._GUARDRAILS_PER_TEST_SCHEMA.get(iid) or {}).get("family") or "")
        except Exception:
            fam_now = ""
        if fam_now == "v6" or sk in ("dhcp_v6", "dhcp_v6_vlan") or iid in ("dhcp_v6", "dhcp_v6_vlan"):
            ore = (cur.get("option_regex") or "").strip()
            looks_v4_opt = bool(re.search(r"Option\s*60", ore, re.I)) or (
                bool(ore) and bool(re.search(r"Vendor-Class", ore, re.I)) and not re.search(r"Option\s*16", ore, re.I)
            )
            if (not ore) or looks_v4_opt:
                cur["option_regex"] = r"(?i)(Option\s*16|vendor.class|Vendor Class|VENDOR_CLASS)"
                changed = True
            dre = (cur.get("discovery_regex") or "").strip()
            if (not dre) or re.search(r"Option\s*43", dre, re.I):
                cur["discovery_regex"] = r"(?i)(Option\s*17|vendor.opts|Vendor-opts|vendor.options)"
                changed = True
            pv6 = (cur.get("probe_v6") or "").strip()
            if pv6 and "/" in pv6:
                cur["probe_v6"] = self._guardrails_strip_ip_cidr(pv6)
                changed = True
        # Settings CLI-ID/PW 가 SSOT — 항상 Settings 로 oru_cli 맞춤
        cid = _field("CLI-ID")
        if cid:
            old = (cur.get("oru_cli_id") or "").strip()
            if old != cid:
                cur["oru_cli_id"] = cid
                changed = True
            else:
                cur["oru_cli_id"] = cid
        cpw = _field("CLI-PW")
        if cpw:
            old = (cur.get("oru_cli_pw") or "").strip()
            if old != cpw:
                cur["oru_cli_pw"] = cpw
                changed = True
            else:
                cur["oru_cli_pw"] = cpw
        # pull common SSH fields from dhcp_boot_shared if empty
        shared = dict(self._guardrails_per_test_settings.get("dhcp_boot_shared") or {})
        for k in ("ru_mac", "mplane_if"):
            if not (cur.get(k) or "").strip() and (shared.get(k) or "").strip():
                cur[k] = shared[k]
                changed = True
        for src_key in ("l2sw_ip", "l2sw_id", "l2sw_pw"):
            if (cur.get(src_key) or "").strip():
                continue
            val = ""
            try:
                val = (self._conformance_get_per_test_val("conformance_3151.sh", src_key) or "").strip()
            except Exception:
                val = ""
            if val:
                cur[src_key] = val
                changed = True
        if changed:
            self._guardrails_per_test_settings[sk] = cur

    def _guardrails_resolve_ssh_family(self, item_id: str | None = None) -> str:
        iid = item_id or getattr(self, "_guardrails_settings_item_id", None) or "dhcp_v4"
        if iid in ("dhcp_v4", "dhcp_v4_vlan", "dhcp_v4_only_boot"):
            return "v4"
        if iid in ("dhcp_v6", "dhcp_v6_vlan", "dhcp_v6_only_boot"):
            return "v6"
        fam = (self._guardrails_gf("ssh_family") or self._guardrails_item_family(iid) or "v4").lower()
        return "v6" if fam.startswith("v6") else "v4"

    def _guardrails_ru_ssh_run(
        self,
        command: str,
        family: str | None = None,
        timeout: int = 40,
        host_override: str | None = None,
    ) -> tuple[bool, str, str]:
        """Run command on RU via solid→ssh. Returns (ok, text, how)."""
        fam = (family or self._guardrails_resolve_ssh_family()).lower()
        if host_override:
            host = self._guardrails_strip_ip_cidr(host_override)
            how = f"override {host}"
        else:
            host, how = self._guardrails_ru_ssh_target(fam)
        if not host:
            return False, how, how
        user = self._guardrails_gf("oru_cli_id")
        pw = self._guardrails_gf("oru_cli_pw")
        if not user:
            return False, "RU SSH ID 필요", how
        # support multiple cmds separated by ;
        parts = [c.strip() for c in re.split(r"\s*;\s*", command or "") if c.strip()]
        if not parts:
            return False, "check_cmd 비어 있음", how
        remote_body = " ; ".join(parts)
        ssh_flag = "-4" if fam == "v4" else "-6"
        remote = (
            "export SSHPASS=" + shlex.quote(pw) + "; "
            f"sshpass -e ssh {ssh_flag} -n "
            "-o StrictHostKeyChecking=no -o UserKnownHostsFile=/dev/null -o LogLevel=ERROR "
            "-o ConnectTimeout=8 -o BatchMode=no "
            + shlex.quote(f"{user}@{host}")
            + " "
            + shlex.quote(f"bash -lc {shlex.quote(remote_body)} || ({remote_body})")
        )
        ok, text = self._guardrails_ssh_exec(f"bash -lc {shlex.quote(remote)}", timeout=timeout)
        return ok, text, how

    @staticmethod
    def _guardrails_parse_dhcp_assigned_ips(
        decode: str, family: str = "v4", mac_colon: str = ""
    ) -> list[str]:
        """Extract leased addresses from tcpdump -vv -e decode (Your-IP / IAADDR).

        mac_colon 이 있으면 해당 MAC 이 등장한 패킷(청크)의 lease 만 채택.
        (pcap 에 다른 클라/서버 IAADDR 가 섞여 ::456 고아 IP 잡히는 것 방지)
        """
        text = decode or ""
        found: list[str] = []
        fam = (family or "v4").lower()
        mac = (mac_colon or "").strip().lower()
        mac_flex = ""
        if mac:
            mac_flex = mac.replace(":", "[-:]")
            # tcpdump -e 패킷 단위로 나눠 MAC 없는 청크 제외
            chunks = re.split(r"(?=^\d{2}:\d{2}:\d{2}\.\d+)", text, flags=re.M)
            if len(chunks) <= 1:
                chunks = [text]
            text = "\n".join(
                c for c in chunks if c and re.search(mac_flex, c, re.I)
            )
            if not text.strip():
                return []
        if fam == "v4":
            pats = (
                r"Your-IP\s+(\d{1,3}(?:\.\d{1,3}){3})",
                r"yiaddr[=:\s]+(\d{1,3}(?:\.\d{1,3}){3})",
                r"Client-IP\s+(\d{1,3}(?:\.\d{1,3}){3})",
                r"Requested-IP(?:\s+Address)?\s+(\d{1,3}(?:\.\d{1,3}){3})",
                r"(?:ACK|OFFER)[^\n]{0,80}?(\d{1,3}(?:\.\d{1,3}){3})",
            )
            for pat in pats:
                for m in re.finditer(pat, text, re.I):
                    ip = m.group(1)
                    if ip.startswith("0.") or ip.startswith("255.") or ip.startswith("127."):
                        continue
                    if ip not in found:
                        found.append(ip)
            return found
        pats = (
            r"IAADDR\s+([0-9a-fA-F:]+)",
            r"(?i)iaaddr[=:\s]+([0-9a-fA-F:]+)",
            r"Identity Association.*?address[:\s]+([0-9a-fA-F:]+)",
            r"(?i)(?:IA_NA|iaprefix|address)\s+([0-9a-fA-F]{0,4}(?::[0-9a-fA-F]{0,4}){2,})",
            r"\b((?:2001|240|fc00|fd[0-9a-f]{0,2})[0-9a-fA-F:]*)\b",
        )
        for pat in pats:
            for m in re.finditer(pat, text, re.I | re.S):
                ip = m.group(1).rstrip(",.;")
                if not ip:
                    continue
                ip = re.split(r"[^\w:]", ip, maxsplit=1)[0]
                if not CallhomeGUI._guardrails_is_plausible_ip(ip, "v6"):
                    continue
                if ip not in found:
                    found.append(ip)
        return found

    @staticmethod
    def _guardrails_is_plausible_ip(ip: str, family: str = "v4") -> bool:
        """MAC(dc:c1:…) / link-local 등을 lease IP 후보에서 제외."""
        s = (ip or "").strip().split("%")[0].split("/")[0]
        if not s:
            return False
        fam = "v6" if str(family).lower().startswith("v6") else "v4"
        if fam == "v4":
            return bool(re.fullmatch(r"\d{1,3}(?:\.\d{1,3}){3}", s))
        # Ethernet MAC 오인 방지
        if re.fullmatch(r"(?i)[0-9a-f]{2}(?::[0-9a-f]{2}){5}", s):
            return False
        try:
            import ipaddress

            a = ipaddress.IPv6Address(s)
            if a.is_link_local or a.is_multicast or a.is_loopback or a.is_unspecified:
                return False
            return True
        except Exception:
            return False

    @staticmethod
    def _guardrails_v6_local_cidr_for_target(target: str) -> str:
        """Same /64 as target, host ::252 (for vlan100 ping6)."""
        try:
            import ipaddress

            t = (target or "").strip().split("%")[0]
            # MAC 형태는 IPv6 로 파싱하지 않음
            if re.fullmatch(r"(?i)[0-9a-f]{2}(?::[0-9a-f]{2}){5}", t.split("/")[0]):
                return ""
            if "/" in t:
                iface = ipaddress.IPv6Interface(t)
            else:
                iface = ipaddress.IPv6Interface(f"{t}/64")
            local = iface.network.network_address + 0x252
            if local == iface.ip:
                local = iface.network.network_address + 0x253
            return f"{local}/64"
        except Exception:
            return ""

    def _lab_controller_listen_ip(self, family: str = "v4", plane: str = "untag") -> str:
        """Lab CallHome/Controller listen IP (Settings LOCAL_IP* 대체).

        untag: v4 10.0.60.253 / v6 2001:1200:1100:1000::253
        tag:   v4 10.0.61.253 / v6 2001:1300:1100:1000::253
        """
        fam = "v6" if str(family or "").lower().startswith("v6") else "v4"
        pl = str(plane or "untag").strip().lower()
        if pl in ("tag", "vlan", "discovery", "1300", "61"):
            return "10.0.61.253" if fam == "v4" else "2001:1300:1100:1000::253"
        return "10.0.60.253" if fam == "v4" else "2001:1200:1100:1000::253"

    def _apply_lab_controller_listen_ips(self, plane: str = "untag") -> None:
        """Settings LOCAL_IP / LOCAL_IP_V6 를 lab untag|tag 값으로 강제."""
        v4 = self._lab_controller_listen_ip("v4", plane)
        v6 = self._lab_controller_listen_ip("v6", plane)
        try:
            if "LOCAL_IP" in self.fields:
                self.fields["LOCAL_IP"].set(v4)
        except Exception:
            pass
        try:
            if "LOCAL_IP_V6" in self.fields:
                self.fields["LOCAL_IP_V6"].set(v6)
        except Exception:
            pass

    def _guardrails_v6_solid_cidr(self, target: str = "", *, plane: str = "") -> str:
        """dhcp_host IF에 올릴 IPv6 CIDR — UI 없이 시험 상황별 자동.

        - target 있으면 그 /64 의 ::252
        - plane=untag → 2001:1200:1100:1000::252/64 (parent enx 임시)
        - plane=tag/빈값 → 2001:1300:1100:1000::252/64 (vlan IF)
        """
        t = self._guardrails_strip_ip_cidr(target)
        if t and self._guardrails_is_plausible_ip(t, "v6"):
            derived = self._guardrails_v6_local_cidr_for_target(t)
            if derived:
                return derived
        pl = (plane or "").strip().lower()
        if pl in ("untag", "pre", "parent", "1200"):
            return "2001:1200:1100:1000::252/64"
        # tag / vlan Discovery IF (기본)
        if pl in ("tag", "vlan", "discovery", "1300", ""):
            # probe 가 이미 tag(1300) 이면 그 /64, 아니면 lab tag 기본
            probe6 = self._guardrails_strip_ip_cidr(self._guardrails_gf("probe_v6"))
            if not probe6 or ":" not in probe6:
                try:
                    probe6 = self._guardrails_strip_ip_cidr(
                        (self.fields.get("ALLOWED_IP_V6").get() or "").strip()  # type: ignore[union-attr]
                    )
                except Exception:
                    probe6 = ""
            if (
                probe6
                and self._guardrails_is_plausible_ip(probe6, "v6")
                and not self._guardrails_is_pre_discovery_ip(probe6, "v6")
            ):
                derived = self._guardrails_v6_local_cidr_for_target(probe6)
                if derived:
                    return derived
            return "2001:1300:1100:1000::252/64"
        return "2001:1300:1100:1000::252/64"

    def _guardrails_dhcp_host_ensure_v6_addr(
        self, ifname: str, cidr6: str
    ) -> tuple[bool, str, bool]:
        """dhcp_host IF에 IPv6 부여 + connected /64 라우트.

        Returns (ok, detail, newly_added). newly_added=True 이면 시험 후 del 대상.
        """
        name = (ifname or "").strip()
        cidr = (cidr6 or "").strip()
        if not name or not cidr or ":" not in cidr:
            return False, "no if/cidr6", False
        if "/" not in cidr:
            cidr = f"{cidr}/64"
        lip = cidr.split("/")[0].strip()
        try:
            import ipaddress

            net6 = str(ipaddress.IPv6Interface(cidr).network)
        except Exception:
            return False, f"bad cidr6 {cidr}", False
        s = self._guardrails_dhcp_host_sudo_line
        script = "\n".join(
            [
                "set +e",
                f"{s(f'ip link set {name} up')} 2>/dev/null || true",
                f"{s(f'sysctl -w net.ipv6.conf.{name}.disable_ipv6=0')} 2>/dev/null || true",
                f"{s('sysctl -w net.ipv6.conf.all.disable_ipv6=0')} 2>/dev/null || true",
                f"if ! ip -6 -o addr show dev {shlex.quote(name)} 2>/dev/null "
                f"| grep -Fq {shlex.quote(lip)}; then",
                f"  {s(f'ip -6 addr add {cidr} dev {name} nodad')} 2>&1 || "
                f"  {s(f'ip -6 addr add {cidr} dev {name}')} 2>&1 || true",
                "  echo CAP_V6_ADD",
                "else",
                "  echo CAP_V6_EXISTS",
                "fi",
                f"{s(f'ip -6 route replace {net6} dev {name} metric 0')} 2>&1 || true",
                f"echo ADDR6=$(ip -6 -o addr show dev {shlex.quote(name)} 2>&1 | tr ' ' '_' | head -c 200)",
                f"if ip -6 -o addr show dev {shlex.quote(name)} 2>/dev/null "
                f"| grep -Fq {shlex.quote(lip)}; then echo CAP_V6_OK; else echo CAP_V6_FAIL; fi",
            ]
        )
        _ok, text = self._guardrails_dhcp_server_run(script, timeout=25)
        out = text or ""
        if "CAP_V6_OK" in out:
            # EXISTS여도 시험용 임시주소면 종료 시 삭제(과거 잔존 포함)
            return True, f"{name} {cidr}", True
        return False, out.replace("\n", " | ")[:220], False

    def _guardrails_arm_parent_v6_cleanup(self, ifname: str, cidr6: str) -> None:
        """Deprecated no-op — v6 untag ::252 는 netplan 고정. 시험 후 삭제하지 않음."""
        return

    def _guardrails_cleanup_parent_v6_pending(self) -> None:
        """Deprecated no-op — parent enx 임시 IPv6 삭제 안 함 (netplan 고정)."""
        self._guardrails_parent_v6_pending = None
        return

    def _guardrails_dhcp_host_del_v6_addr(
        self, ifname: str, cidr6: str
    ) -> tuple[bool, str]:
        """시험이 parent(enx)에만 임시로 올린 IPv6 제거 (vlan IF 삭제는 teardown)."""
        name = (ifname or "").strip()
        cidr = (cidr6 or "").strip()
        if not name or not cidr:
            return True, "skip"
        lip = cidr.split("/")[0].strip()
        s = self._guardrails_dhcp_host_sudo_line
        script = "\n".join(
            [
                "set +e",
                f"{s(f'ip -6 addr del {cidr} dev {name}')} 2>/dev/null || "
                f"{s(f'ip -6 addr del {lip}/64 dev {name}')} 2>/dev/null || true",
                "echo CAP_V6_DEL",
            ]
        )
        self._guardrails_dhcp_server_run(script, timeout=15)
        return True, f"del {lip}@{name}"

    def _guardrails_find_ip_by_ru_mac(
        self, mac_colon: str, vlan_if: str = "", family: str = "v4"
    ) -> list[str]:
        """Find RU IP via neigh / leases (pcap miss 보완). v4/v6."""
        mac = (mac_colon or "").strip().lower()
        if not mac:
            return []
        fam = "v6" if str(family).lower().startswith("v6") else "v4"
        mac_flex = mac.replace(":", "[-:]")
        if fam == "v6":
            parts = [
                f"ip -6 neigh show 2>/dev/null | grep -iE {shlex.quote(mac_flex)} || true",
            ]
            if vlan_if:
                # MAC 매칭 실패해도 vlan IF 의 global neigh 후보를 본다
                parts.append(
                    f"ip -6 neigh show dev {shlex.quote(vlan_if)} 2>/dev/null || true"
                )
                parts.append(
                    f"ip -6 neigh show dev {shlex.quote(vlan_if)} 2>/dev/null "
                    f"| grep -iE {shlex.quote(mac_flex)} || true"
                )
            parts.append(
                "for f in /var/lib/dhcp/dhcpd6.leases /var/lib/dhcp/dhcpd.leases "
                "/var/lib/dhcpd/dhcpd6.leases; do "
                f"test -r \"$f\" && grep -iE {shlex.quote(mac_flex)} -A8 \"$f\" 2>/dev/null; "
                "done || true"
            )
            cmd = "; ".join(parts)
            _ok, text = self._guardrails_dhcp_server_run(cmd, timeout=15)
            found: list[str] = []
            # MAC 매칭 줄 우선
            prefer: list[str] = []
            for line in (text or "").splitlines():
                ips_on = re.findall(
                    r"\b([0-9a-fA-F]{0,4}(?::[0-9a-fA-F]{0,4}){2,})\b", line
                )
                hit_mac = bool(re.search(mac_flex, line, re.I))
                for ip in ips_on:
                    if not self._guardrails_is_plausible_ip(ip, "v6"):
                        continue
                    # dhcp_host 자기주소(::252/::253) 제외
                    if re.search(r":(0*252|0*253)$", ip, re.I):
                        continue
                    if hit_mac:
                        if ip not in prefer:
                            prefer.append(ip)
                    elif ip not in found:
                        found.append(ip)
            return prefer + [ip for ip in found if ip not in prefer]
        parts = [
            f"ip -4 neigh show 2>/dev/null | grep -iE {shlex.quote(mac_flex)} || true",
        ]
        if vlan_if:
            parts.append(
                f"ip -4 neigh show dev {shlex.quote(vlan_if)} 2>/dev/null || true"
            )
            parts.append(
                f"ip -4 addr show dev {shlex.quote(vlan_if)} 2>/dev/null | head -5 || true"
            )
        parts.append(
            "for f in /var/lib/dhcp/dhcpd.leases /var/lib/dhcp/dhcpd6.leases "
            "/var/lib/kea/kea-leases4.csv; do "
            f"test -r \"$f\" && grep -iE {shlex.quote(mac_flex)} -A20 \"$f\" 2>/dev/null; "
            "done | head -c 12000 || true"
        )
        ok, text = self._guardrails_dhcp_server_run("; ".join(parts), timeout=25)
        if not (text or "").strip():
            return []
        found: list[str] = []
        for m in re.finditer(r"\b(10\.\d{1,3}\.\d{1,3}\.\d{1,3})\b", text or ""):
            ip = m.group(1)
            if ip.endswith((".0", ".255", ".252", ".253")):
                continue  # skip net/bcast/likely server
            if ip not in found:
                found.append(ip)
        for m in re.finditer(
            r"lease\s+(\d{1,3}(?:\.\d{1,3}){3})\s*\{", text or "", re.I
        ):
            ip = m.group(1)
            if ip not in found:
                found.append(ip)
        return found

    def _guardrails_peek_pcap_assigned_ips(
        self,
        remote_pcap: str,
        family: str = "v4",
        timeout: int = 12,
        mac_colon: str = "",
    ) -> list[str]:
        """Best-effort read of growing pcap; leased IPs (MAC 있으면 해당 프레임만)."""
        fam = "v6" if str(family).lower().startswith("v6") else "v4"
        mac = (mac_colon or "").strip().lower()
        # offline filter: RU MAC 프레임만 decode → 다른 클라 IAADDR 혼입 방지
        ether = ""
        if mac and re.fullmatch(r"(?i)[0-9a-f]{2}(?::[0-9a-f]{2}){5}", mac):
            ether = f" ether host {mac}"
        if fam == "v6":
            gre = (
                "IAADDR|iaaddr|IA_NA|Advertise|Reply|Request|"
                "2001:|240:|fc00:|fd|"
                + (mac.replace(":", "|") if mac else "ether")
            )
        else:
            gre = (
                "Your-IP|yiaddr|Client-IP|Requested-IP|IAADDR|iaaddr|ACK|OFFER|"
                + (mac.replace(":", "|") if mac else "ether")
            )
        cmd = (
            f"tcpdump -nn -vv -e -r {shlex.quote(remote_pcap)}{ether} 2>/dev/null "
            f"| grep -iE {shlex.quote(gre)} "
            f"| tail -n 240 || true"
        )
        ok, text = self._guardrails_dhcp_server_run(cmd, timeout=max(5, int(timeout)))
        if not ok and not (text or "").strip():
            return []
        return self._guardrails_parse_dhcp_assigned_ips(
            text or "", family, mac_colon=mac
        )

    def _guardrails_peek_pcap_assigned_ip(
        self, remote_pcap: str, family: str = "v4", mac_colon: str = ""
    ) -> str:
        """Newest leased IP in pcap (MAC 필터 권장)."""
        ips = self._guardrails_peek_pcap_assigned_ips(
            remote_pcap, family, mac_colon=mac_colon
        )
        return ips[-1] if ips else ""

    @staticmethod
    def _guardrails_same_v4_slash24(a: str, b: str) -> bool:
        try:
            pa = [int(x) for x in a.split(".")]
            pb = [int(x) for x in b.split(".")]
            return len(pa) == 4 and len(pb) == 4 and pa[:3] == pb[:3]
        except Exception:
            return False

    def _guardrails_is_pre_discovery_ip(self, ip: str, family: str = "v4") -> bool:
        """True if IP is still on untag plane — not Discovery(tag) lease.

        Lab mapping:
          v4 untag 10.0.60.x  → tag 10.0.61.x
          v6 untag 2001:1200:1100:1000::/64 → tag 2001:1300:1100:1000::/64
        """
        ip = self._guardrails_strip_ip_cidr(ip)
        if not ip:
            return False
        fam = "v6" if str(family).lower().startswith("v6") else "v4"
        if fam == "v4":
            probe = ""
            try:
                probe = self._guardrails_strip_ip_cidr(
                    (self.fields.get("ALLOWED_IP").get() or "").strip()  # type: ignore[union-attr]
                )
            except Exception:
                probe = ""
            if not probe:
                probe = self._guardrails_strip_ip_cidr(self._guardrails_gf("probe_v4"))
            if probe and self._guardrails_same_v4_slash24(ip, probe):
                return True
            # lab default untag plane (.60); tag 는 .61
            if ip.startswith("10.0.60."):
                return True
            return False
        # lab default untag /64 (1200); tag 는 1300
        if ip.lower().startswith("2001:1200:1100:1000:"):
            return True
        probe6 = ""
        try:
            probe6 = self._guardrails_strip_ip_cidr(
                (self.fields.get("ALLOWED_IP_V6").get() or "").strip()  # type: ignore[union-attr]
            )
        except Exception:
            probe6 = ""
        if not probe6:
            probe6 = self._guardrails_strip_ip_cidr(self._guardrails_gf("probe_v6"))
        if not (probe6 and ":" in ip and ":" in probe6):
            return False
        try:
            import ipaddress

            a = ipaddress.IPv6Address(ip)
            b = ipaddress.IPv6Address(probe6)
            # same /64 as Settings/probe_v6 → still untag/pre-Discovery
            return (int(a) >> 64) == (int(b) >> 64)
        except Exception:
            return False

    def _guardrails_pick_discovery_ip(self, ips: list[str], family: str = "v4") -> str:
        """Newest leased IP that is NOT the pre-Discovery subnet."""
        for ip in reversed(ips or []):
            if ip and not self._guardrails_is_pre_discovery_ip(ip, family):
                return ip
        return ""

    def _guardrails_ru_force_dhcp_renew(
        self,
        host: str,
        family: str = "v4",
        command: str | None = None,
        *,
        via_dhcp_host: bool = False,
        vlan_bind: bool | None = None,
    ) -> tuple[bool, str]:
        """Run renew CLI on RU over SSH. Empty command → 'dhcp renew force'.

        via_dhcp_host=True: SSH from 9.249 (dhcp_host).
        vlan_bind=True: -b 10.0.61.252 (Discovery .61).
        vlan_bind=False: 수동 untag 와 동일 — enx(.60) 기본경로로 admin@10.0.60.x.
        vlan_bind=None: 대상이 pre-discovery(.60)면 False, 아니면 True.
        """
        host = self._guardrails_strip_ip_cidr(host)
        if not host:
            return False, "renew: no host"
        fam = (family or "v4").lower()
        primary = (command if command is not None else self._guardrails_gf("dhcp_renew_cmd") or "").strip()
        if not primary:
            primary = "dhcp renew force"
        primary = primary.lower()
        # VLAN Discovery 원복: vlan-discovery renew 가 거절되면 일반 renew 도 시도
        cmd_candidates: list[str] = []
        for c in (
            primary,
            "dhcp vlan-discovery renew force",
            "dhcp renew force",
        ):
            c = (c or "").strip().lower()
            if c and c not in cmd_candidates:
                cmd_candidates.append(c)
        if via_dhcp_host:
            user = self._guardrails_gf("oru_cli_id")
            pw = self._guardrails_gf("oru_cli_pw")
            if not user:
                return False, "renew: RU SSH ID 필요"
            if not (pw or "").strip():
                return False, "renew: RU SSH PW(oru_cli_pw) 비어 있음 — 로그인 실패 가능"
            if vlan_bind is None:
                vlan_bind = not self._guardrails_is_pre_discovery_ip(host, fam)
            # fe80::…%if → link-local renew (GUA 타이밍 실패 폴백)
            host_bare = host.split("%", 1)[0].strip()
            scope_if = host.split("%", 1)[1].strip() if "%" in host else ""
            is_ll = host_bare.lower().startswith("fe80:")
            bind_ip = ""
            ifname = "dhcp_host"
            if is_ll and scope_if:
                ifname = scope_if
                vlan_bind = bool(vlan_bind) if vlan_bind is not None else ("vlan" in scope_if.lower())
            elif vlan_bind:
                # Discovery VLAN IF (v4=.61 / v6=같은 /64 ::252)
                st_if, det_if, _owned, ifname = self._guardrails_solid_vlan_if_prepare()
                if st_if != "PASS":
                    return False, f"renew: dhcp_host IF 준비 실패 ({det_if})"
                _p, _n, cidr = self._guardrails_capture_host_vlan_if_names()
                bind_ip = (cidr.split("/")[0] or "").strip()
                if fam == "v6":
                    local6 = self._guardrails_v6_solid_cidr(host_bare, plane="tag")
                    if local6:
                        ok6, det6, _added = self._guardrails_dhcp_host_ensure_v6_addr(
                            ifname, local6
                        )
                        bind_ip = local6.split("/")[0]
                        if not ok6:
                            return False, f"renew: vlan IF IPv6 부여 실패 ({det6})"
            else:
                # 수동 캡처와 동일: ssh admin@10.0.60.167 (vlan bind 없음)
                ifname = (self._guardrails_gf("dhcp_if") or "enx").strip() or "enx"
                if fam == "v6":
                    local6 = self._guardrails_v6_solid_cidr(host_bare, plane="untag")
                    if local6:
                        ok6, det6, added = self._guardrails_dhcp_host_ensure_v6_addr(
                            ifname, local6
                        )
                        bind_ip = local6.split("/")[0] if ok6 else ""
                        if not ok6:
                            return False, f"renew: parent IPv6 부여 실패 ({det6})"
                        if added:
                            self._guardrails_arm_parent_v6_cleanup(ifname, local6)
            # 죽은 tag IP 에 40s SSH 낭비 방지 (untag=enx / tag=vlan IF)
            # LL 은 GUA ping 대신 L2 ND 만 — ping 사전검사 생략(타이밍 목적)
            if not is_ll:
                try:
                    up_p, det_p = self._guardrails_probe_from_dhcp_host(
                        host_bare, via_if=ifname, family=fam
                    )
                    if not up_p:
                        return (
                            False,
                            f"renew-fail unreachable {host} "
                            f"({'tag' if vlan_bind else 'untag'} ping FAIL: {det_p[:120]})",
                        )
                except Exception as exc:
                    self._guardrails_log(f"dhcp renew: ping 사전검사 스킵 ({exc})")
            else:
                self._guardrails_log(
                    f"dhcp renew: link-local 경로 — ping 생략, ssh {user}@{host}"
                )
            ssh_flag = "-6" if (is_ll or fam != "v4") else "-4"
            bind_opt = ""
            if is_ll:
                # scope 는 user@fe80::x%if 에 포함 — -B 중복 불필요
                bind_opt = ""
            elif fam == "v6" and bind_ip:
                bind_opt = f"-b {shlex.quote(bind_ip)} "
            elif vlan_bind and bind_ip and fam == "v4":
                bind_opt = f"-b {shlex.quote(bind_ip)} "
            elif vlan_bind and fam == "v6" and ifname:
                bind_opt = f"-B {shlex.quote(ifname)} "
            how = f"dhcp_host/{ifname}→{host}"
            ssh_dest = host  # fe80::…%vlan100 유지
            last_fail = ""
            for cmd in cmd_candidates:
                # 수동과 동일: 로그인 후 프롬프트에 명령 입력.
                # 금지: sshpass + ssh -tt + '원격argv명령' — stdin 충돌로 로그인/명령 둘 다 실패.
                self._guardrails_log(
                    f"dhcp renew SSH 시도 중… {user}@{ssh_dest} "
                    f"({'LL' if is_ll else ('vlan_bind '+bind_ip if vlan_bind else 'untag '+ifname)}) "
                    f"cmd=`{cmd}` — 대화형 CLI 주입 (최대 ~40s)"
                )
                feeder = (
                    "sleep 2.5; "
                    f"printf '%s\\r\\n' {shlex.quote(cmd)}; "
                    "sleep 2.5; "
                    "printf 'exit\\r\\n'; "
                    "sleep 0.5"
                )
                remote = (
                    "export SSHPASS=" + shlex.quote(pw) + "; "
                    "command -v sshpass >/dev/null 2>&1 || { echo RENEW_NO_SSHPASS; exit 41; }; "
                    f"echo RENEW_TRY user={shlex.quote(user)} cmd={shlex.quote(cmd)} "
                    f"bind={shlex.quote(bind_ip or '-')} to {shlex.quote(ssh_dest)}; "
                    "set +e; "
                    f"sout=$( ( {feeder} ) | sshpass -e ssh {ssh_flag} {bind_opt}-tt "
                    "-o StrictHostKeyChecking=no -o UserKnownHostsFile=/dev/null -o LogLevel=ERROR "
                    "-o PreferredAuthentications=password -o PubkeyAuthentication=no "
                    "-o ConnectTimeout=12 -o ConnectionAttempts=1 "
                    + shlex.quote(f"{user}@{ssh_dest}")
                    + " 2>&1 ); rc=$?; "
                    'echo "RENEW_SSH_OUT=$(echo \"$sout\" | tr \'\\n\' \'|\' | head -c 500)"; '
                    "echo RENEW_SSH_RC=$rc; "
                    "if echo \"$sout\" | grep -Eqi "
                    "'Connection refused|No route to host|Connection timed out|"
                    "Could not resolve|Name or service not known|Network is unreachable'; then "
                    "echo RENEW_FAIL_CONNECT; exit 43; fi; "
                    "if echo \"$sout\" | grep -Eqi "
                    "'Permission denied|Authentication failed|access denied|"
                    "sshpass:.*denied'; then "
                    "echo RENEW_FAIL_AUTH; exit 42; fi; "
                    "login_ok=0; "
                    "if echo \"$sout\" | grep -Eqi "
                    "'O-RAN\\.MP\\.DAEMON|mplane#|Developer mode|oru1-mplane|"
                    "Welcome|login successful'; then "
                    "echo RENEW_LOGIN_OK; login_ok=1; fi; "
                    "if echo \"$sout\" | grep -Eqi "
                    "'Connection reset|closed by remote|Broken pipe|Connection to .* closed'; then "
                    "echo RENEW_OK_DROPPED; exit 0; fi; "
                    "if [ \"$login_ok\" -eq 1 ]; then echo RENEW_OK; exit 0; fi; "
                    "if [ \"$rc\" -eq 0 ]; then echo RENEW_OK; exit 0; fi; "
                    "echo RENEW_FAIL; exit \"$rc\""
                )
                _ok, text = self._guardrails_dhcp_server_run(remote, timeout=45)
                out = text or ""
                snippet = out.replace("\n", " | ")[:400]
                self._guardrails_log(f"dhcp renew SSH 결과: {snippet[:320]}")
                if "RENEW_NO_SSHPASS" in out:
                    return False, f"renew-fail `{cmd}` via {how}: sshpass 없음 on dhcp_host"
                if "RENEW_FAIL_CONNECT" in out:
                    # IP/경로 문제 — 다른 CLI 명령 재시도 무의미
                    return (
                        False,
                        f"renew-fail `{cmd}` via {how}: SSH 연결 실패 "
                        f"(대상 IP/경로): {snippet}",
                    )
                if "RENEW_FAIL_AUTH" in out:
                    return (
                        False,
                        f"renew-fail `{cmd}` via {how}: SSH 로그인 실패 "
                        f"(CLI-ID/PW 확인): {snippet}",
                    )
                if (
                    "RENEW_OK" in out
                    or "RENEW_OK_DROPPED" in out
                    or "RENEW_LOGIN_OK" in out
                    or re.search(
                        r"(?i)O-RAN\.MP\.DAEMON|mplane#|oru1-mplane|Welcome", out
                    )
                ):
                    tag = "dropped-ok" if "RENEW_OK_DROPPED" in out else "ok"
                    return True, f"renew `{cmd}` via {how} ({tag}): {snippet}"
                last_fail = f"renew-fail `{cmd}` via {how}: {snippet}"
                # 로그인/연결은 됐는데 판정 실패 → 다음 후보 명령 시도
                if len(cmd_candidates) > 1:
                    self._guardrails_log(
                        f"dhcp renew: `{cmd}` 미확인 — 다음 명령 후보 시도"
                    )
            return False, last_fail or f"renew-fail via {how}"
        cmd = primary
        ok, text, how = self._guardrails_ru_ssh_run(cmd, family=fam, timeout=60, host_override=host)
        snippet = (text or "").replace("\n", " | ")[:220]
        if ok:
            return True, f"renew `{cmd}` via {how}: {snippet}"
        # renew 직후 세션 끊김도 성공 후보
        low = (text or "").lower()
        if any(
            x in low
            for x in (
                "connection reset",
                "connection closed",
                "broken pipe",
                "closed by remote",
            )
        ):
            return True, f"renew `{cmd}` via {how} (dropped-ok): {snippet}"
        return False, f"renew-fail `{cmd}` via {how}: {snippet}"

    def _guardrails_match_output(self, text: str, expect_regex: str, required_ips: list[str] | None = None) -> tuple[bool, str]:
        notes: list[str] = []
        if expect_regex:
            try:
                if not re.search(expect_regex, text or "", re.I | re.M):
                    return False, f"expect_regex 미일치: {expect_regex[:60]}"
                notes.append("regex OK")
            except re.error as exc:
                return False, f"regex 오류: {exc}"
        for ip in required_ips or []:
            if ip and ip not in (text or ""):
                return False, f"IP 미검출: {ip}"
            if ip:
                notes.append(f"ip {ip}")
        return True, ", ".join(notes) if notes else "OK"

    def _guardrails_extract_dhcp_option_lines(
        self, decode: str, family: str = "v4", limit: int = 16
    ) -> list[str]:
        """pcap decode 에서 옵션/벤더 관련 줄을 중복 없이 정리."""
        text = decode or ""
        fam = "v6" if str(family).lower().startswith("v6") else "v4"
        if fam == "v4":
            pats = (
                r"(?i)Option\s*60[^\n]{0,160}",
                r"(?i)Vendor-Class[^\n]{0,160}",
                r"(?i)Option\s*43[^\n]{0,160}",
                r"(?i)Vendor-Specific[^\n]{0,120}",
                r"(?i)Your-IP\s+\d{1,3}(?:\.\d{1,3}){3}",
            )
        else:
            pats = (
                r"(?i)Option\s*16[^\n]{0,160}",
                r"(?i)Vendor Class[^\n]{0,160}",
                r"(?i)Option\s*17[^\n]{0,160}",
                r"(?i)vendor[- ]?opts[^\n]{0,140}",
                r"(?i)IA[_-]?NA[^\n]{0,100}",
                r"(?i)IAADDR\s+[0-9a-fA-F:]+",
            )
        out: list[str] = []
        seen: set[str] = set()
        for pat in pats:
            for m in re.finditer(pat, text):
                s = " ".join(m.group(0).split())
                if len(s) > 180:
                    s = s[:177] + "..."
                key = s.lower()
                if key in seen:
                    continue
                seen.add(key)
                out.append(s)
                if len(out) >= limit:
                    return out
        return out

    def _guardrails_format_dhcp_options_detail(
        self,
        *,
        item_id: str,
        status: str,
        fam: str,
        boot_ok: bool,
        opt_ok: bool,
        disc_ok: bool,
        ia_na_ok: bool,
        learned_ip: str,
        decode: str,
        opt_re: str,
        notes: list[str],
        sec_reset_to_down: float | None,
        sec_reset_to_ip: float | None,
        sec_reset_to_boot: float | None,
        sec_down_to_boot: float | None,
    ) -> str:
        """상세 결과용 줄별 요약 (옵션 / 수신 IP / reset 후 시간)."""
        lines: list[str] = [
            f"{status}  {item_id}",
            "",
            "[판정]",
            f"  boot(ping)     : {'OK' if boot_ok else 'FAIL'}",
            f"  option(pcap)   : {'OK' if opt_ok else 'FAIL'}",
            f"  discovery      : {'OK' if disc_ok else 'FAIL'}",
        ]
        if fam == "v6":
            lines.append(f"  ia_na          : {'OK' if ia_na_ok else 'FAIL'}")
        lines.extend(["", "[타이밍]  (reset 기준)"])
        def _sec(v: float | None) -> str:
            if v is None:
                return "-"
            return f"{v:.0f}s"

        lines.append(f"  reset → SSH down     : {_sec(sec_reset_to_down)}")
        lines.append(f"  reset → IP 수신      : {_sec(sec_reset_to_ip)}")
        lines.append(f"  reset → ping 복구    : {_sec(sec_reset_to_boot)}")
        if sec_down_to_boot is not None:
            lines.append(f"  SSH down → ping 복구 : {_sec(sec_down_to_boot)}")

        lines.extend(["", "[수신 IP]"])
        lip = (learned_ip or "").strip()
        if lip:
            lines.append(f"  Discovery/lease : {lip}")
        else:
            lines.append("  Discovery/lease : (미검출)")
        # 상세: MAC 필터 없이 전체 lease 도 보여 주되, 사용 IP 는 별도 표기
        leased = self._guardrails_parse_dhcp_assigned_ips(decode or "", fam)
        uniq: list[str] = []
        for ip in leased:
            if ip and ip not in uniq:
                uniq.append(ip)
        if uniq:
            label = "Your-IP" if fam == "v4" else "IAADDR"
            lines.append(f"  pcap {label} (전체):")
            for ip in uniq[-8:]:
                mark = "  ← 사용" if lip and ip == lip else ""
                lines.append(f"    - {ip}{mark}")

        lines.extend(["", "[옵션]"])
        opt_lines = self._guardrails_extract_dhcp_option_lines(decode or "", fam)
        if opt_re:
            lines.append(f"  expect_regex : {opt_re[:100]}")
        if opt_lines:
            for ol in opt_lines:
                lines.append(f"  · {ol}")
        else:
            lines.append("  (pcap에서 옵션 문자열 미추출)")

        # notes 중 핵심만 (한 줄짜리 잡음 축소)
        key_notes: list[str] = []
        for n in notes or []:
            ns = str(n).strip()
            if not ns:
                continue
            if ns.startswith(("family=", "ssh=", "if=", "bpf=", "dhcp_host=")):
                continue
            if ns.startswith(("pcap_ip:", "boot:", "option:", "discovery:", "ia_na")):
                continue
            key_notes.append(ns)
        if key_notes:
            lines.extend(["", "[기타]"])
            for n in key_notes[-20:]:
                lines.append(f"  · {n}")

        lines.extend(
            [
                "",
                "[환경]",
                f"  family={fam}",
            ]
        )
        for n in notes or []:
            ns = str(n).strip()
            if ns.startswith(("if=", "bpf=", "dhcp_host=", "ssh=")):
                lines.append(f"  {ns}")
        return "\n".join(lines)

    def _guardrails_check_dhcpv6_ia_na(self, decode: str, regex: str) -> tuple[bool, str]:
        """Require IA_NA (Option 3) in DHCPv6 Solicit — field issue: missing → no IPv6 lease."""
        text = decode or ""
        if not text.strip():
            return False, "ia_na: decode 없음"
        pat = (regex or "").strip() or r"(?i)(IA[_-]?NA|Identity Association for Non-temporary)"
        try:
            re.compile(pat)
        except re.error as exc:
            return False, f"ia_na regex 오류: {exc}"

        # Prefer Solicit-scoped: at least one Solicit must include IA_NA
        solicit_iter = list(re.finditer(r"(?i)\bdhcp6\s+solicit\b|\bsolicit\s*\(", text))
        msg_pat = re.compile(
            r"(?i)\bdhcp6\s+(solicit|advertise|request|reply|confirm|renew|rebind|"
            r"release|decline|information-request)\b"
        )
        if solicit_iter:
            ok_any = False
            missing_n = 0
            for m in solicit_iter:
                nxt = msg_pat.search(text, m.end())
                end = nxt.start() if nxt else min(len(text), m.start() + 3000)
                chunk = text[m.start() : end]
                if re.search(pat, chunk, re.I | re.M):
                    ok_any = True
                else:
                    missing_n += 1
            if ok_any:
                note = "ia_na OK (Solicit)"
                if missing_n:
                    note += f"; other_solicit_missing={missing_n}"
                return True, note
            return False, (
                "ia_na FAIL: Solicit에 IA_NA(Option 3) 없음 "
                "(현장 IPv6 미할당과 동일 패턴)"
            )

        if re.search(pat, text, re.I | re.M):
            return True, "ia_na OK (decode, no Solicit label)"
        return False, "ia_na FAIL: IA_NA/Option3 미검출"

    def _guardrails_log(self, msg: str) -> None:
        """Thread-safe GUI log for M-Plane Test.

        Important: write to log_buffer on the calling thread (worker OK).
        Only schedule UI open/paint on the Tk main thread — after() from a
        worker is unreliable on Windows and was dropping all M-Plane lines.
        """
        line = msg if msg.endswith("\n") else msg + "\n"
        if not line.startswith("[M-Plane Test]"):
            line = f"[M-Plane Test] {line}"
        # 1) Always enqueue immediately (works from worker threads).
        try:
            with self.log_lock:
                self.log_buffer.append(line)
                self._recent_log_for_session += line
                if len(self._recent_log_for_session) > 128_000:
                    self._recent_log_for_session = self._recent_log_for_session[-128_000:]
        except Exception:
            pass

        # 2) Ask main thread to show Logs + paint.
        def _ui() -> None:
            try:
                if self.log is None or self.log_window is None or not self.log_window.winfo_exists():
                    self.open_log_window()
            except Exception:
                try:
                    self.open_log_window()
                except Exception:
                    return
            try:
                if self.log is None or not self.log.winfo_exists():
                    return
                chunk = ""
                with self.log_lock:
                    if self.hidden_log_chunks:
                        chunk = "".join(self.hidden_log_chunks)
                        self.hidden_log_chunks.clear()
                    if self.log_buffer:
                        chunk += "".join(self.log_buffer)
                        self.log_buffer.clear()
                if not chunk:
                    return
                self.log.configure(state="normal")
                self.log.insert("end", chunk)
                self.log.see("end")
                self.log.configure(state="disabled")
            except Exception:
                pass

        try:
            self.after(0, _ui)
        except Exception:
            pass

    def _guardrails_trigger_oru_reboot_v6(self) -> tuple[bool, str]:
        """IPv6 CallHome reset via copied helper (conformance_oru_reboot_v6.sh). v4 helper untouched."""
        helper = "conformance_oru_reboot_v6.sh"
        # reset 시점 RU 는 아직 untag — CallHome listen = untag controller
        try:
            self._apply_lab_controller_listen_ips("untag")
        except Exception:
            pass
        local_v6 = ""
        allowed_v6 = ""
        try:
            local_v6 = self._guardrails_strip_ip_cidr(
                (self.fields.get("LOCAL_IP_V6").get() or "").strip()  # type: ignore[union-attr]
            )
            allowed_v6 = self._guardrails_strip_ip_cidr(
                (self.fields.get("ALLOWED_IP_V6").get() or "").strip()  # type: ignore[union-attr]
            )
        except Exception:
            pass
        if not local_v6:
            local_v6 = self._lab_controller_listen_ip("v6", "untag")
        # Settings ALLOWED_IP_V6 가 SSOT (⚙ probe 는 Settings 미러만)
        if not local_v6 or not allowed_v6:
            self._guardrails_log(
                "ORU reset(v6) 실패: Settings LOCAL_IP_V6 / ALLOWED_IP_V6 필요"
            )
            return False, "no-v6-settings"
        if ":" not in local_v6 or ":" not in allowed_v6:
            self._guardrails_log("ORU reset(v6) 실패: LOCAL_IP_V6/ALLOWED_IP_V6 가 IPv6 형식이 아님")
            return False, "bad-v6-settings"

        ssh_user = self.remote_user.get().strip()
        ssh_host = self.remote_host.get().strip()
        ssh_port = self.remote_port.get().strip() or "22"
        ssh_password = self.remote_password.get()
        key_path = (self.remote_key_path.get() or "").strip()
        if not ssh_user or not ssh_host:
            self._guardrails_log("ORU reset(v6) 실패: Settings SSH_USER/SSH_HOST 필요")
            return False, "no-ssh"

        try:
            import paramiko  # type: ignore
        except Exception as exc:
            self._guardrails_log(f"ORU reset(v6) 실패: paramiko 필요 ({exc})")
            return False, "no-paramiko"

        opts = self._conformance_default_run_options()
        remote_dir = opts.remote_dir.rstrip("/")
        cfg_remote = f"{remote_dir}/{_conf_manifest.CONFORMANCE_REMOTE_GUI_CONFIG_NAME}"
        lp = self._conformance_script_local_path(helper)
        if lp is None:
            cand = self._conformance_local_dir() / helper
            if cand.is_file():
                lp = cand
        if lp is None:
            self._guardrails_log(f"ORU reset(v6) 실패: 로컬 헬퍼 없음 {helper}")
            return False, "no-helper"

        def log_line(msg: str) -> None:
            self._guardrails_log(msg)

        client: Any = None
        try:
            try:
                self._conformance_cancel_event.clear()
            except Exception:
                pass
            client = paramiko.SSHClient()
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            client.connect(
                hostname=ssh_host,
                port=int(ssh_port),
                username=ssh_user,
                password=ssh_password if ssh_password else None,
                key_filename=key_path if key_path else None,
                timeout=20,
                auth_timeout=20,
                banner_timeout=20,
                look_for_keys=not bool(ssh_password),
                allow_agent=True,
            )
            _stdin, _stdout, _stderr = client.exec_command(f"mkdir -p {shlex.quote(remote_dir)}")
            _stdout.channel.recv_exit_status()
            sftp = client.open_sftp()
            rp = f"{remote_dir}/{helper}"
            try:
                # Windows checkout may be CRLF — normalize to LF before upload
                raw = lp.read_bytes().replace(b"\r\n", b"\n").replace(b"\r", b"\n")
                sftp.putfo(io.BytesIO(raw), rp, len(raw))
                try:
                    sftp.chmod(rp, 0o755)
                except OSError:
                    pass
                log_line(f"uploaded {helper} (LF normalized)")
            except Exception as exc:
                log_line(f"[ERROR] {helper} 업로드 실패: {exc}")
                return False, "upload-fail"
            try:
                cfg_payload = self._conformance_effective_config_json_text()
                cfg_bytes = cfg_payload.encode("utf-8")
                sftp.putfo(io.BytesIO(cfg_bytes), cfg_remote, len(cfg_bytes))
            except Exception as exc:
                log_line(f"[WARN] 재부팅용 config 갱신 실패(기존 config 사용): {exc}")

            envp = self._conformance_bash_env_exports(opts, None)
            host_log = self._conformance_host_run_log_path(helper)
            dir_q = shlex.quote(str(PurePosixPath(host_log).parent))
            log_q = shlex.quote(host_log)
            rp_q = shlex.quote(rp)
            cfg_q = shlex.quote(cfg_remote)
            # listen 기본 300s → M-Plane 은 90s (중지/타임아웃 반응)
            listen_to = 90
            try:
                listen_to = max(30, int(self._guardrails_gf("reset_listen_sec", "90") or "90"))
            except Exception:
                listen_to = 90
            runner = (
                f"{envp}"
                f"export CONFORMANCE_SCRIPT_BASENAME={shlex.quote(helper)} ; "
                f"export CALLHOME_LISTEN_TIMEOUT={int(listen_to)} ; "
                f"chmod +x {rp_q} 2>/dev/null ; bash {rp_q} --config {cfg_q}"
            )
            wrapped = (
                f"set -o pipefail; "
                f"mkdir -p {dir_q} && : > {log_q} && chmod 0644 {log_q} || exit 1; "
                f"( {runner} ) 2>&1 | tee -a {log_q}; "
                "_cf_rc=${PIPESTATUS[0]}; "
                'exit "${_cf_rc:-0}"'
            )
            cmd_remote = "bash -lc " + shlex.quote(wrapped)
            log_line(f"---- START {helper} (ORU reset IPv6) ----")
            log_line(f"remote host log file: {host_log}")
            log_line(
                f"[INFO] IPv6 CallHome listen LOCAL_IP_V6={local_v6} "
                f"ALLOWED_IP_V6={allowed_v6} listen≤{listen_to}s "
                f"(v4 헬퍼/스크립트 미사용)"
            )
            log_line(
                "[INFO] 순서: CallHome 로그인 성공 → "
                "<reset xmlns=\"urn:o-ran:operations:1.0\"/> 전송. "
                "이 단계에서는 dhcp renew force 를 보내지 않음 "
                "(renew 는 VLAN 원복/정리 때만)"
            )
            _stdin, stdout, stderr = client.exec_command(cmd_remote, get_pty=True)
            ch = stdout.channel
            with self._conformance_run_transport_lock:
                self._conformance_run_script_channel = ch
            t_listen0 = time.monotonic()
            # GUI 측 하드캡 (스크립트 listen + 여유)
            hard_cap = float(listen_to) + 45.0
            while not ch.exit_status_ready():
                if (
                    self._guardrails_cancel.is_set()
                    or self._conformance_cancel_event.is_set()
                ):
                    try:
                        ch.close()
                    except Exception:
                        pass
                    try:
                        # 원격 listen/netopeer 잔여 정리 (best-effort)
                        client.exec_command(
                            f"pkill -f {shlex.quote(helper)} 2>/dev/null; "
                            f"fuser -k 4334/tcp 2>/dev/null || true"
                        )
                    except Exception:
                        pass
                    log_line("재부팅 헬퍼(v6) 실행 중 사용자 중지")
                    with self._conformance_run_transport_lock:
                        self._conformance_run_script_channel = None
                    return False, "cancelled"
                if time.monotonic() - t_listen0 > hard_cap:
                    try:
                        ch.close()
                    except Exception:
                        pass
                    log_line(
                        f"재부팅 헬퍼(v6) 타임아웃 ({hard_cap:.0f}s) — "
                        "CallHome 미수신. ALLOWED_IP_V6/LOCAL_IP_V6·ip6tables 확인"
                    )
                    with self._conformance_run_transport_lock:
                        self._conformance_run_script_channel = None
                    return False, "listen-timeout"
                if ch.recv_ready():
                    chunk = ch.recv(4096).decode(errors="ignore")
                    if chunk:
                        for line in chunk.splitlines():
                            log_line(line)
                else:
                    time.sleep(0.1)
            try:
                rem = stdout.read().decode(errors="ignore")
                for line in rem.splitlines():
                    log_line(line)
            except Exception:
                pass
            rc = ch.recv_exit_status()
            log_line(f"---- END {helper} exit={rc} ----")
            with self._conformance_run_transport_lock:
                self._conformance_run_script_channel = None
            try:
                sftp.close()
            except Exception:
                pass
            if rc == 0:
                log_line("ORU reset RPC 전송 완료 (IPv6 Call Home helper)")
                return True, "conformance_oru_reboot_v6"
            log_line(f"[WARN] 재부팅 헬퍼(v6) 실패 exit={rc}")
            return False, "conformance_oru_reboot_v6-fail"
        except Exception as exc:
            self._guardrails_log(f"ORU reset(v6) 예외: {exc}")
            return False, f"exception:{exc}"
        finally:
            if client is not None:
                try:
                    client.close()
                except Exception:
                    pass

    def _guardrails_trigger_ru_reset(self, fam: str) -> tuple[bool, str]:
        """RU reset: v4 → existing conformance_oru_reboot.sh; v6 → copied v6 helper."""
        mode = (self._guardrails_gf("reset_mode", "auto") or "auto").strip().lower()
        if mode in ("0", "manual", "prompt", "none", "off"):
            self._guardrails_log("Reset mode=manual — RU를 직접 재부팅하세요")
            return True, "manual"

        if (fam or "").lower().startswith("v6"):
            return self._guardrails_trigger_oru_reboot_v6()

        # reset 시점 RU 는 아직 untag — CallHome listen = untag controller
        try:
            self._apply_lab_controller_listen_ips("untag")
        except Exception:
            pass

        ssh_user = self.remote_user.get().strip()
        ssh_host = self.remote_host.get().strip()
        ssh_port = self.remote_port.get().strip() or "22"
        ssh_password = self.remote_password.get()
        key_path = (self.remote_key_path.get() or "").strip()
        if not ssh_user or not ssh_host:
            self._guardrails_log("ORU reset 실패: Settings SSH_USER/SSH_HOST 필요")
            return False, "no-ssh"

        try:
            import paramiko  # type: ignore
        except Exception as exc:
            self._guardrails_log(f"ORU reset 실패: paramiko 필요 ({exc})")
            return False, "no-paramiko"

        opts = self._conformance_default_run_options()
        remote_dir = opts.remote_dir.rstrip("/")
        cfg_remote = f"{remote_dir}/{_conf_manifest.CONFORMANCE_REMOTE_GUI_CONFIG_NAME}"

        def log_line(msg: str) -> None:
            self._guardrails_log(msg)

        client: Any = None
        try:
            try:
                self._conformance_cancel_event.clear()
            except Exception:
                pass
            client = paramiko.SSHClient()
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            client.connect(
                hostname=ssh_host,
                port=int(ssh_port),
                username=ssh_user,
                password=ssh_password if ssh_password else None,
                key_filename=key_path if key_path else None,
                timeout=20,
                auth_timeout=20,
                banner_timeout=20,
                look_for_keys=not bool(ssh_password),
                allow_agent=True,
            )
            _stdin, _stdout, _stderr = client.exec_command(f"mkdir -p {shlex.quote(remote_dir)}")
            _stdout.channel.recv_exit_status()
            ok_lip, det_lip = self._guardrails_verify_local_ip_on_remote(client)
            self._guardrails_log(f"ORU reset(v4) LOCAL_IP 검사 → {'OK' if ok_lip else 'FAIL'}: {det_lip}")
            if not ok_lip:
                return False, f"local-ip-missing:{det_lip}"
            sftp = client.open_sftp()
            self._guardrails_log(
                "ORU reset: Conformance 헬퍼 호출 "
                "(_conformance_trigger_oru_reboot → conformance_oru_reboot.sh)"
            )
            ok = bool(
                self._conformance_trigger_oru_reboot(
                    client, sftp, opts, remote_dir, cfg_remote, log_line
                )
            )
            try:
                sftp.close()
            except Exception:
                pass
            if ok:
                return True, "conformance_oru_reboot"
            return False, "conformance_oru_reboot-fail"
        except Exception as exc:
            self._guardrails_log(f"ORU reset 예외: {exc}")
            return False, f"exception:{exc}"
        finally:
            if client is not None:
                try:
                    client.close()
                except Exception:
                    pass

    def _guardrails_run_dhcp_options_family(self, item_id: str) -> tuple[str, str]:
        """Boot(L2SW없음) 또는 VLAN Discovery(L2SW/ACL) + tcpdump option + pcap."""
        self._guardrails_settings_item_id = item_id
        self._guardrails_fill_defaults_from_context(item_id)
        self._guardrails_sync_probe_from_settings(item_id)
        self._guardrails_sync_oru_cli_from_settings(item_id)
        fam = self._guardrails_resolve_ssh_family(item_id)
        enable_vlan_disc = self._guardrails_item_mode(item_id) == "vlan"
        mac_src = (
            self._guardrails_strip_ip_cidr(self._guardrails_gf("probe_v6"))
            if fam == "v6"
            else self._guardrails_strip_ip_cidr(self._guardrails_gf("probe_v4"))
        )
        self._guardrails_log(
            f"{item_id}: mode={'vlan' if enable_vlan_disc else 'boot'} — Settings/probe "
            f"v4={self._guardrails_gf('probe_v4') or '-'} "
            f"v6={self._guardrails_gf('probe_v6') or '-'} "
            f"l2sw={self._guardrails_gf('l2sw_ip') or '-'} "
            f"| MAC조회IP[{fam}]={mac_src or '-'}"
        )

        # Boot/VLAN 공통: BPF ether host 용 RU MAC 은 매번 neigh 로 갱신
        # (Boot 에서 생략하면 예전 00:90:… 같은 stale MAC 으로 pcap 0패킷 발생)
        st_mac, det_mac = self._guardrails_ensure_ru_mac(item_id, force=True)
        if st_mac == "FAIL":
            return "FAIL", f"RU MAC 자동 조회 실패: {det_mac}"
        if st_mac == "WARN":
            self._guardrails_log(f"{item_id}: RU MAC WARN — {det_mac}")
        else:
            self._guardrails_log(f"{item_id}: RU MAC OK — {det_mac}")

        if enable_vlan_disc:
            if not self._guardrails_gf("l2sw_ip") or not self._guardrails_gf("l2sw_id"):
                return "FAIL", "⚙ L2SW IP/ID 필요 (VLAN Discovery)"

            # L2SW IF (enable → show mac → 해당 MAC 포트)
            st_if, det_if = self._guardrails_ensure_l2sw_if_from_mac(item_id, force=True)
            if st_if == "FAIL":
                return "FAIL", f"L2SW IF 자동 조회 실패: {det_if}"
            if st_if == "WARN":
                self._guardrails_log(f"{item_id}: L2SW IF 조회 WARN — {det_if}")
            if not self._guardrails_l2sw_normalize_if():
                return "FAIL", "L2SW IF 없음 (RU MAC show mac 조회 실패)"
            ll0, how_ll0 = self._guardrails_ru_linklocal_scoped(tag=False)
            if ll0:
                self._guardrails_log(f"{item_id}: 사전 LL(계산) — {how_ll0}")
        else:
            # Boot도 단일 스택: 반대 계열 DHCP ACL 필요 → L2SW IF 확보
            if not self._guardrails_gf("l2sw_ip") or not self._guardrails_gf("l2sw_id"):
                return "FAIL", "⚙ L2SW IP/ID 필요 (Boot 단일스택 ACL)"
            st_if, det_if = self._guardrails_ensure_l2sw_if_from_mac(item_id, force=True)
            if st_if == "FAIL":
                return "FAIL", f"L2SW IF 자동 조회 실패 (Boot ACL): {det_if}"
            if st_if == "WARN":
                self._guardrails_log(f"{item_id}: L2SW IF 조회 WARN — {det_if}")
            if not self._guardrails_l2sw_normalize_if():
                return "FAIL", "L2SW IF 없음 — Boot ACL 적용 불가"
            self._guardrails_log(
                f"{item_id}: Boot 단일스택 — L2SW IF OK ({det_if}); "
                + (
                    "v4 DHCP 차단 → v6만"
                    if fam == "v6"
                    else "v6 DHCP 차단 → v4만"
                )
            )

        # Capture IF (dhcp_host ifconfig → 10.0.60.x)
        st_cap, det_cap = self._guardrails_ensure_capture_if(item_id, prefer_tag=False)
        if st_cap == "FAIL":
            return "FAIL", f"Capture IF 자동 조회 실패: {det_cap}"
        if st_cap == "WARN":
            self._guardrails_log(f"{item_id}: Capture IF WARN — {det_cap}")
        iface = (self._guardrails_gf("dhcp_if") or "").strip()
        if not iface:
            return "FAIL", "Capture IF 없음 (dhcp_host ifconfig 10.0.60.x 확인)"
        # v6 untag: parent enx 에 2001:1200:…::252 임시 부여 → 시험 종료 시 삭제
        if fam == "v6":
            tgt6 = self._guardrails_strip_ip_cidr(self._guardrails_gf("probe_v6"))
            if not tgt6:
                try:
                    tgt6 = self._guardrails_strip_ip_cidr(
                        (self.fields.get("ALLOWED_IP_V6").get() or "").strip()  # type: ignore[union-attr]
                    )
                except Exception:
                    tgt6 = ""
            cidr6_untag = self._guardrails_v6_solid_cidr(tgt6, plane="untag") or (
                "2001:1200:1100:1000::252/64"
            )
            ok6, det6, _added = self._guardrails_dhcp_host_ensure_v6_addr(iface, cidr6_untag)
            if not ok6:
                return "FAIL", f"Capture IF IPv6 임시 부여 실패: {det6}"
            self._guardrails_arm_parent_v6_cleanup(iface, cidr6_untag)
            self._guardrails_log(
                f"{item_id}: v6 untag Capture IF → {iface} {cidr6_untag} "
                "(netplan 고정 유지 · 시험 후 삭제 안 함)"
            )
        host, how = self._guardrails_ru_ssh_target(fam)
        if not host:
            need = "Settings ALLOWED_IP" if fam == "v4" else "Settings ALLOWED_IP_V6 또는 RU MAC+M-Plane NIC"
            return "FAIL", f"SSH 대상 불가: {how} ({need})"
        if not self._guardrails_gf("oru_cli_id"):
            return "FAIL", "RU SSH ID 필요 (Settings ★ RU SSH ID)"

        use_sudo = (self._guardrails_gf("use_sudo", "1") or "1").strip().lower() not in ("0", "false", "no", "n")
        sudo = "sudo -n " if use_sudo else ""
        # Re-apply family regex fix right before match (saved ⚙ may still hold v4 Option 60)
        self._guardrails_fill_defaults_from_context(item_id)
        opt_re = self._guardrails_gf("option_regex")
        if fam == "v6" and (
            not opt_re
            or re.search(r"Option\s*60", opt_re, re.I)
            or not re.search(r"Option\s*16", opt_re, re.I)
        ):
            opt_re = r"(?i)(Option\s*16|vendor.class|Vendor Class|VENDOR_CLASS)"
            try:
                sk = self._guardrails_store_key(item_id)
                blob = dict(self._guardrails_per_test_settings.get(sk) or {})
                blob["option_regex"] = opt_re
                self._guardrails_per_test_settings[sk] = blob
            except Exception:
                pass
            self._guardrails_log(f"{item_id}: option_regex → Option 16 강제 (v4 Option60 잔존 방지)")
        enable_disc = (self._guardrails_gf("enable_discovery", "0") or "0").strip().lower() not in (
            "0", "false", "no", "n", "",
        )
        enable_ia_na = fam == "v6"  # IA_NA(Opt3) 필수 — 설정 UI 숨김, 항상 검사
        ia_na_re = self._guardrails_gf(
            "ia_na_regex",
            r"(?i)(IA[_-]?NA|Identity Association for Non-temporary)",
        )
        # enable_vlan_disc: item mode == vlan (체크박스 제거)
        vlan_vid = (self._guardrails_gf("vlan_discovery_vid", "61") or "61").strip() or "61"
        vlan_name = (self._guardrails_gf("vlan_discovery_name") or "").strip()
        renew_cmd_cfg = (self._guardrails_gf("dhcp_renew_cmd") or "").strip()
        renew_cmd = renew_cmd_cfg.lower() if renew_cmd_cfg else "dhcp renew force"
        learned_ip = ""  # from pcap when VLAN Discovery moves RU to new subnet (e.g. .61)
        # prepare PASS뿐 아니라 FAIL(부분 적용: vlan만 생성 등)에도 종료 시 remove 시도
        vlan_needs_cleanup = False
        solid_vlan_owned = False  # GUI가 solid에 임시 만든 IF만 종료 시 삭제
        solid_vlan_if = ""
        reset_completed = False  # CallHome reset RPC/헬퍼 성공 후 True — 중지 시 LL renew 분기
        acl_applied = False  # Boot/VLAN 공통 단일스택 ACL — 종료 시 원복
        pass_sec = self._guardrails_int("pass_sec", 240)
        timeout_sec = self._guardrails_int("timeout_sec", 540)
        poll_sec = self._guardrails_int("poll_sec", 5)
        stable_sec = self._guardrails_int("stable_sec", 10)
        down_detect_sec = self._guardrails_int("down_detect_sec", 180)

        def _maybe_force_dhcp_renew() -> None:
            """Renew via SSH. ⚙ 빈칸 → 'dhcp renew force', 설정 시 그 명령(소문자).

            VLAN Discovery 중에는 기존 .60 probe 로 renew 하지 않음(경로 끊김).
            learned_ip 또는 pcap의 Discovery(.61) IP 만 시도.
            """
            nonlocal learned_ip
            if not enable_vlan_disc and not renew_cmd_cfg:
                notes.append("dhcp_renew:SKIP")
                return
            targets: list[str] = []
            if enable_vlan_disc:
                if learned_ip and not self._guardrails_is_pre_discovery_ip(learned_ip, fam):
                    targets.append(learned_ip.strip())
                try:
                    for ip in _collect_discovery_candidate_ips():
                        if ip not in targets:
                            targets.append(ip)
                except Exception:
                    pass
                if not targets:
                    notes.append("dhcp_renew:SKIP (Discovery IP 없음, .60 probe 생략)")
                    self._guardrails_log(
                        f"{item_id}: DHCP renew 생략 — 시험 VLAN IP 없음 "
                        "(기존 probe .60 은 VLAN Discovery 중 사용 안 함)"
                    )
                    return
            else:
                for cand in (
                    learned_ip,
                    self._guardrails_strip_ip_cidr(
                        self._guardrails_gf("probe_v4" if fam == "v4" else "probe_v6")
                    ),
                ):
                    c = (cand or "").strip()
                    if c and c not in targets:
                        targets.append(c)
                if not targets:
                    notes.append("dhcp_renew:SKIP no ip")
                    self._guardrails_log(f"{item_id}: DHCP renew — 대상 IP 없음, 생략")
                    return
            for target in targets:
                self._guardrails_log(
                    f"{item_id}: DHCP renew `{renew_cmd}` → {target}"
                    + (" (via dhcp_host)" if enable_vlan_disc else "")
                )
                ok_r, det_r = self._guardrails_ru_force_dhcp_renew(
                    target, fam, command=renew_cmd, via_dhcp_host=enable_vlan_disc
                )
                notes.append(f"dhcp_renew:{'OK' if ok_r else 'WARN'}@{target} {det_r}")
                self._guardrails_log(
                    f"{item_id}: dhcp renew → {'OK' if ok_r else 'WARN'}: {det_r[:200]}"
                )
                if ok_r:
                    break

        def _teardown_solid_vlan_if() -> None:
            """시험 종료 시 dhcp_host 임시 vlan IF 삭제 (reuse 포함)."""
            nonlocal solid_vlan_owned, solid_vlan_if
            if not (solid_vlan_if or "").strip():
                solid_vlan_owned = False
                return
            st_s, det_s = self._guardrails_solid_vlan_if_teardown(solid_vlan_if)
            notes.append(f"solid_if:{det_s}")
            self._guardrails_log(f"{item_id}: solid 임시 IF 삭제 → {st_s}: {det_s}")
            solid_vlan_owned = False
            solid_vlan_if = ""

        def _defer_solid_vlan_to_pending(pending: dict[str, Any]) -> None:
            """버튼 원복까지 solid IF 유지 — ownership을 pending으로 이전."""
            nonlocal solid_vlan_owned, solid_vlan_if
            pending["solid_vlan_owned"] = True  # 원복 시 반드시 삭제
            pending["solid_vlan_if"] = solid_vlan_if
            solid_vlan_owned = False
            solid_vlan_if = ""

        def _renew_via_targets(targets: list[str], *, tag: bool) -> bool:
            """tag=True: pcap IP + vlan100. tag=False: probe(.60) untag 경로. 성공 시 True."""
            label = "tag/pcap" if tag else "untag/probe"
            any_target = False
            for target in targets:
                c = (target or "").strip()
                if not c:
                    continue
                any_target = True
                self._guardrails_log(
                    f"{item_id}: DHCP renew `{renew_cmd}` → {c} ({label}, "
                    f"ssh {self._guardrails_gf('oru_cli_id') or 'admin'}@{c})"
                )
                ok_r, det_r = self._guardrails_ru_force_dhcp_renew(
                    c,
                    fam,
                    command=renew_cmd,
                    via_dhcp_host=True,
                    vlan_bind=tag,
                )
                notes.append(f"dhcp_renew:{'OK' if ok_r else 'WARN'}@{c}/{label} {det_r}")
                self._guardrails_log(
                    f"{item_id}: dhcp renew → {'OK' if ok_r else 'WARN'}: {det_r[:200]}"
                )
                if ok_r:
                    return True
            if not any_target:
                notes.append(f"dhcp_renew:SKIP no {label} ip")
            return False

        def _cleanup_vlan_discovery() -> None:
            """정리: add base(1) → renew → 즉시 remove 시험VLAN(100).

            remove 가 늦으면 RU 가 다시 tag(.61) 로 IP 를 받음.
            renew: tag → untag probe → neigh(현재 IP) 순 폴백
            (reset 후 RU 가 이미 untag 새 IP(::141)면 옛 tag SSH 실패함).
            """
            nonlocal vlan_needs_cleanup
            if not vlan_needs_cleanup:
                _teardown_solid_vlan_if()
                return
            probe = self._guardrails_strip_ip_cidr(
                self._guardrails_gf("probe_v4" if fam == "v4" else "probe_v6")
            )
            base = self._guardrails_vlan_discovery_base_vid()
            tag_ip = ""
            if learned_ip and not self._guardrails_is_pre_discovery_ip(learned_ip, fam):
                tag_ip = learned_ip.strip()
            if not tag_ip:
                try:
                    for ip in _collect_discovery_candidate_ips():
                        if ip and not self._guardrails_is_pre_discovery_ip(ip, fam):
                            tag_ip = ip
                            break
                except Exception:
                    pass

            self._guardrails_log(
                f"{item_id}: VLAN Discovery 정리 — "
                f"① add {base} → ② renew → ③ 즉시 remove {vlan_vid}"
            )
            # 1) base add (시험 VLAN 100 은 아직 trunk에 유지 → tag SSH 가능)
            st_a, det_a = self._guardrails_l2sw_vlan_discovery_add_base(base)
            notes.append(f"vlan_base_add:{det_a}")
            self._guardrails_log(f"{item_id}: L2 base add → {st_a}: {det_a}")
            time.sleep(1.0)
            # 2) renew: tag → untag probe → neigh 현재 IP
            renewed = False
            if tag_ip:
                renewed = _renew_via_targets([tag_ip], tag=True)
            if (not renewed) and probe and probe != tag_ip:
                if tag_ip:
                    self._guardrails_log(
                        f"{item_id}: tag renew 실패 — untag probe 로 재시도 ({probe})"
                    )
                renewed = _renew_via_targets([probe], tag=False)
            if (not renewed) and mac_colon:
                try:
                    live = self._guardrails_find_ip_by_ru_mac(
                        mac_colon, solid_vlan_if or "", fam
                    )
                except Exception:
                    live = []
                seen = {x for x in (tag_ip, probe) if x}
                tag_live = [
                    ip
                    for ip in live
                    if ip
                    and ip not in seen
                    and not self._guardrails_is_pre_discovery_ip(ip, fam)
                ]
                untag_live = [
                    ip
                    for ip in live
                    if ip
                    and ip not in seen
                    and self._guardrails_is_pre_discovery_ip(ip, fam)
                ]
                if tag_live or untag_live:
                    self._guardrails_log(
                        f"{item_id}: renew 폴백 — neigh 현재 IP "
                        f"tag={tag_live or '-'} untag={untag_live or '-'}"
                    )
                if (not renewed) and tag_live:
                    renewed = _renew_via_targets(tag_live, tag=True)
                if (not renewed) and untag_live:
                    renewed = _renew_via_targets(untag_live, tag=False)
            # GUA 타이밍 실패 시: MAC→EUI-64 link-local (사전 확정) 폴백
            if not renewed:
                ll_tag, how_ll = self._guardrails_ru_linklocal_scoped(tag=True)
                if ll_tag:
                    self._guardrails_log(
                        f"{item_id}: renew 폴백 — link-local(tag) {how_ll}"
                    )
                    renewed = _renew_via_targets([ll_tag], tag=True)
                if not renewed:
                    ll_u, how_u = self._guardrails_ru_linklocal_scoped(tag=False)
                    if ll_u:
                        self._guardrails_log(
                            f"{item_id}: renew 폴백 — link-local(untag) {how_u}"
                        )
                        renewed = _renew_via_targets([ll_u], tag=False)
            if not renewed:
                notes.append("dhcp_renew:WARN all-targets-failed")
                self._guardrails_log(
                    f"{item_id}: ⚠ renew 전부 실패 — L2 vlan remove 는 계속 "
                    f"(tag={tag_ip or '-'} probe={probe or '-'})"
                )
            # 3) 즉시 remove 100 (대기 없음 — 다시 tag lease 받는 것 방지)
            self._guardrails_log(
                f"{item_id}: renew 직후 즉시 trunk remove vlan {vlan_vid}"
            )
            st_r, det_r = self._guardrails_l2sw_vlan_discovery_remove_test(vlan_vid)
            notes.append(f"vlan_remove:{det_r}")
            self._guardrails_log(f"{item_id}: L2 vlan remove → {st_r}: {det_r}")
            vlan_needs_cleanup = False
            _teardown_solid_vlan_if()

        ru_mac = self._guardrails_gf("ru_mac")
        mac_norm = self._guardrails_normalize_mac(ru_mac)
        mac_colon = ":".join(mac_norm[i : i + 2] for i in range(0, 12, 2)) if mac_norm else ""
        if fam == "v4":
            dhcp_ports = "udp port 67 or udp port 68"
        else:
            dhcp_ports = "udp port 546 or udp port 547"
        # 802.1Q: BPF 'udp port 67' alone misses VLAN-tagged DHCP on parent USB.
        # Manual: tcpdump -i enx… -e 'vlan or port 67 or port 68' 는 잡히는데
        # GUI 예전 필터는 untag만 잡아 .61 Your-IP 분석 실패했음.
        if enable_vlan_disc:
            bpf = (
                f"({dhcp_ports}) or "
                f"(vlan {vlan_vid} and ({dhcp_ports})) or "
                f"(vlan and ({dhcp_ports}))"
            )
        else:
            bpf = dhcp_ports
        if mac_colon:
            bpf = f"ether host {mac_colon} and ({bpf})"

        stamp = time.strftime("%Y%m%d_%H%M%S")
        base = f"mplane_{item_id}_{stamp}"
        remote_pcap = f"/tmp/{base}.pcap"
        remote_pid = f"/tmp/{base}.pid"
        remote_log = f"/tmp/{base}.td.log"
        try:
            log_path = (self.fields.get("LOG_PATH").get() or "").strip()  # type: ignore[union-attr]
        except Exception:
            log_path = ""
        if not log_path:
            log_path = "/var/tmp/log"
        log_path = log_path.rstrip("/")
        dest_pcap = f"{log_path}/{base}.pcap"
        dest_txt = f"{log_path}/{base}_decode.txt"

        notes: list[str] = [f"family={fam}", f"ssh={how}", f"if={iface}", f"bpf={bpf}"]
        notes.append(f"dhcp_host={self._guardrails_gf('dhcp_host') or 'solid'}")

        def _collect_discovery_candidate_ips(*, with_lease: bool = False) -> list[str]:
            """USB pcap(RU MAC 프레임의 lease) + optional neigh. Discovery 대역만."""
            if self._guardrails_cancel.is_set():
                return []
            merged: list[str] = []
            try:
                for ip in self._guardrails_peek_pcap_assigned_ips(
                    remote_pcap, fam, timeout=10, mac_colon=mac_colon
                ):
                    if ip not in merged:
                        merged.append(ip)
            except Exception:
                pass
            if with_lease and mac_colon and not self._guardrails_cancel.is_set():
                try:
                    vif = solid_vlan_if or ""
                    for ip in self._guardrails_find_ip_by_ru_mac(mac_colon, vif, fam):
                        if ip not in merged:
                            merged.append(ip)
                except Exception:
                    pass
            return [
                ip
                for ip in merged
                if ip
                and self._guardrails_is_plausible_ip(ip, fam)
                and not self._guardrails_is_pre_discovery_ip(ip, fam)
            ]

        def _sleep_poll(sec: float) -> None:
            """poll sleep that wakes quickly on 중지."""
            end = time.monotonic() + max(0.0, float(sec))
            while time.monotonic() < end:
                if self._guardrails_cancel.is_set():
                    return
                time.sleep(min(0.2, end - time.monotonic()))

        def _abort_fast(msg: str) -> tuple[str, str]:
            """사용자 중지 정리.

            - reset 이전: renew/LL 생략 (빠른 정리만)
            - reset 완료 후: link-local 로 renew 명령 후 VLAN은 버튼 원복
            """
            nonlocal vlan_needs_cleanup, learned_ip, reset_completed, acl_applied
            do_ll_renew = bool(reset_completed) and bool(enable_vlan_disc)
            if do_ll_renew:
                self._guardrails_log(
                    f"{item_id}: {msg} — reset 이후 중지 → link-local renew 시도"
                )
                renewed = False
                try:
                    for tag_ll in (True, False):
                        ll, how_ll = self._guardrails_ru_linklocal_scoped(tag=tag_ll)
                        if not ll:
                            continue
                        label = "LL/tag" if tag_ll else "LL/untag"
                        self._guardrails_log(
                            f"{item_id}: 중지 시 renew `{renew_cmd}` → {ll} ({label}) — {how_ll}"
                        )
                        ok_r, det_r = self._guardrails_ru_force_dhcp_renew(
                            ll,
                            fam,
                            command=renew_cmd,
                            via_dhcp_host=True,
                            vlan_bind=tag_ll,
                        )
                        notes.append(
                            f"abort_ll_renew:{'OK' if ok_r else 'WARN'}@{ll}/{label}"
                        )
                        self._guardrails_log(
                            f"{item_id}: 중지 LL renew → {'OK' if ok_r else 'WARN'}: "
                            f"{(det_r or '')[:200]}"
                        )
                        if ok_r:
                            renewed = True
                            break
                except Exception as exc:
                    notes.append(f"abort_ll_renew:ERR:{exc}")
                    self._guardrails_log(f"{item_id}: 중지 LL renew 예외: {exc}")
                if not renewed:
                    notes.append("abort_ll_renew:SKIP/FAIL")
                    self._guardrails_log(
                        f"{item_id}: 중지 LL renew 미성공 — VLAN 원복은 버튼으로"
                    )
            else:
                self._guardrails_log(
                    f"{item_id}: {msg} — 빠른 정리 "
                    f"({'reset 이전' if not reset_completed else 'Boot'} · renew/LL 생략)"
                )
            try:
                _stop_tcpdump()
            except Exception:
                pass
            if enable_vlan_disc and vlan_needs_cleanup:
                # 버튼 원복 때 renew 하려면 Discovery IP 가 필요 — 중지 직전 한 번 더 수집
                if not (learned_ip or "").strip() or self._guardrails_is_pre_discovery_ip(
                    learned_ip, fam
                ):
                    try:
                        for ip in _collect_discovery_candidate_ips(with_lease=True):
                            if ip and not self._guardrails_is_pre_discovery_ip(ip, fam):
                                learned_ip = ip
                                self._guardrails_log(
                                    f"{item_id}: 중지 직전 Discovery IP 확보 → {learned_ip}"
                                )
                                break
                    except Exception:
                        pass
                    if not (learned_ip or "").strip():
                        try:
                            raw = self._guardrails_peek_pcap_assigned_ips(
                                remote_pcap,
                                fam,
                                timeout=10,
                                mac_colon=mac_colon,
                            )
                            pick = self._guardrails_pick_discovery_ip(raw, fam)
                            if pick:
                                learned_ip = pick
                                self._guardrails_log(
                                    f"{item_id}: 중지 직전 pcap Discovery IP → {learned_ip}"
                                )
                        except Exception:
                            pass
                pending = {
                    "item_id": item_id,
                    "fam": fam,
                    "vid": vlan_vid,
                    "base_vid": self._guardrails_vlan_discovery_base_vid(),
                    "learned_ip": (learned_ip or "").strip(),
                    "probe": self._guardrails_strip_ip_cidr(
                        self._guardrails_gf("probe_v4" if fam == "v4" else "probe_v6")
                    ),
                    "renew_cmd": renew_cmd,
                    "ru_mac": mac_colon,
                    "abort_ll_renew_done": do_ll_renew,
                }
                _defer_solid_vlan_to_pending(pending)  # Discovery SSH용 IF는 버튼 원복까지 유지
                self.after(0, lambda p=pending: self._guardrails_arm_vlan_restore(p))
                notes.append("vlan_cleanup:DEFERRED(중지—버튼원복)")
                vlan_needs_cleanup = False
                lip = (learned_ip or "").strip() or "-"
                self._guardrails_log(
                    f"{item_id}: VLAN/solid IF 원복은 「VLAN untag 원복」 버튼으로 "
                    f"(저장 IP={lip}"
                    + (", LL renew 시도함" if do_ll_renew else ", reset 전·renew 생략")
                    + ")"
                )
                if lip == "-" and not do_ll_renew:
                    self._guardrails_log(
                        f"{item_id}: ⚠ Discovery IP 미저장 — 원복 시 tag renew 실패 가능 "
                        "(RU show dhcp IP를 ⚙에 없으므로 neigh/pcap 재시도는 버튼에서)"
                    )
            else:
                _teardown_solid_vlan_if()
            if acl_applied:
                try:
                    rm_st, rm_detail = self._guardrails_remove_acl(timeout=40)
                    notes.append(f"acl_abort:{rm_st}:{rm_detail[:80]}")
                    acl_applied = False
                except Exception as exc:
                    notes.append(f"acl_abort:ERR:{exc}")
            else:
                notes.append("acl_abort:SKIP")
            return "INFO", f"{msg} | " + " | ".join(notes)

        def _stop_tcpdump() -> tuple[bool, str]:
            stop_cmd = (
                f"pid=$(cat {shlex.quote(remote_pid)} 2>/dev/null); "
                f"test -n \"$pid\" && kill \"$pid\" 2>/dev/null; "
                f"pkill -f {shlex.quote(base + '.pcap')} 2>/dev/null; sleep 0.3; "
                f"ls -la {shlex.quote(remote_pcap)} 2>/dev/null || echo NO_PCAP"
            )
            return self._guardrails_dhcp_server_run(stop_cmd, timeout=12)

        start_cmd = (
            f"{sudo}tcpdump -n -i {shlex.quote(iface)} -U -w {shlex.quote(remote_pcap)} "
            f"{shlex.quote(bpf)} >{shlex.quote(remote_log)} 2>&1 & "
            f"echo $! >{shlex.quote(remote_pid)}; sleep 0.5; "
            f"test -s {shlex.quote(remote_pid)} && echo START_OK || "
            f"(echo START_FAIL; cat {shlex.quote(remote_log)} 2>/dev/null | tail -20)"
        )
        ok, text = self._guardrails_dhcp_server_run(start_cmd, timeout=30)
        if not ok or "START_OK" not in (text or ""):
            self._guardrails_log(f"{item_id}: tcpdump 시작 실패: {(text or '')[:200]}")
            return "FAIL", f"tcpdump 시작 실패: {(text or '')[:240]} | " + " | ".join(notes)
        notes.append("tcpdump started")
        self._guardrails_log(f"{item_id}: tcpdump 시작 OK (if={iface}, bpf={bpf})")

        # Boot/VLAN 공통: 단일 스택 — 시험 계열만 DHCP 통과 (반대 계열 ACL)
        # v4 시험: v6-dhcp-block / v6 시험: acl 110(67/68 drop)
        self._guardrails_log(
            f"{item_id}: L2SW ACL 적용 중 ({fam}"
            + (", Boot 단일스택" if not enable_vlan_disc else ", VLAN Discovery")
            + ")…"
        )
        st, detail = self._guardrails_apply_acl(fam)
        if st != "PASS":
            _stop_tcpdump()
            self._guardrails_log(f"{item_id}: ACL 실패 — {detail}")
            return st, detail + " | " + " | ".join(notes)
        notes.append(detail)
        acl_applied = True

        # VLAN Discovery 초반: vlan 생성 + trunk add 만 (base remove는 reset 직후)
        if enable_vlan_disc:
            self._guardrails_log(
                f"{item_id}: VLAN Discovery — vlan {vlan_vid}"
                + (f" ({vlan_name})" if vlan_name else "")
                + " trunk add (base remove는 reset 직후)"
            )
            st_v, det_v = self._guardrails_l2sw_vlan_discovery_prepare(vlan_vid, vlan_name)
            notes.append(f"vlan_prepare:{det_v}")
            self._guardrails_log(f"{item_id}: VLAN prepare → {st_v}: {det_v}")
            vlan_needs_cleanup = True  # PASS/FAIL 무관 — 종료/원복 시 vlan100 삭제
            if st_v != "PASS":
                notes.append("vlan_prepare:WARN (시험은 계속, 종료 시 remove 시도)")
            # 9.249(dhcp_host) Capture IF에 임시 .VID IF (v4 .61 + v6 ::252)
            st_sf, det_sf, solid_vlan_owned, solid_vlan_if = (
                self._guardrails_solid_vlan_if_prepare(vlan_vid, primary_if=iface)
            )
            notes.append(f"cap_if:{det_sf}")
            self._guardrails_log(f"{item_id}: dhcp_host 임시 IF → {st_sf}: {det_sf}")
            if st_sf != "PASS":
                _stop_tcpdump()
                _cleanup_vlan_discovery()
                self._guardrails_remove_acl()
                return (
                    "FAIL",
                    f"dhcp_host vlan IF 주소 부여 실패 — ⚙ dhcp_pw(sudo) / "
                    f"vlan_discovery_solid_cidr_v6 확인. {det_sf} | "
                    + " | ".join(notes),
                )
            if fam == "v6":
                self._guardrails_log(
                    f"{item_id}: v6 tag 경로용 IF={solid_vlan_if} "
                    f"cidr6={self._guardrails_v6_solid_cidr(plane='tag') or '-'} "
                    "(종료/원복 시 IF 삭제)"
                )

        expect = "IPv4 inet" if fam == "v4" else "global inet6"
        if self._guardrails_cancel.is_set():
            return _abort_fast("사용자 중지 (reset 이전)")
        rst_ok, rst_how = self._guardrails_trigger_ru_reset(fam)
        t_reset = time.monotonic()
        notes.append(f"reset:{rst_how}")
        mode = (self._guardrails_gf("reset_mode", "auto") or "auto").strip().lower()
        if self._guardrails_cancel.is_set() and not rst_ok:
            # 헬퍼 도중 중지·실패 — reset 미완료로 간주 (LL renew 안 함)
            return _abort_fast("사용자 중지 (reset 이전/미완료)")
        if rst_ok:
            reset_completed = True
            self._guardrails_log(
                f"{item_id}: reset 완료 표시 ({rst_how}) — 이후 중지 시 LL renew"
            )
        if not rst_ok and mode in ("auto", "mplane"):
            if mode == "mplane":
                _stop_tcpdump()
                _cleanup_vlan_discovery()
                if acl_applied:
                    self._guardrails_remove_acl()
                    acl_applied = False
                return "FAIL", f"ORU reset 실패 ({rst_how}) | " + " | ".join(notes)
            self._guardrails_log(
                f"자동 reset 실패({rst_how}) — 수동 재부팅 후 복구 감시 계속"
            )
            rst_how = "manual-fallback"
            t_reset = time.monotonic()
            # 수동 폴백은 reset RPC 미전송 — LL renew 조건에서 제외 유지
            reset_completed = False

        # reset 직후 바로 base(vlan1) remove — Discovery 강제
        if enable_vlan_disc and vlan_needs_cleanup:
            base_vid = self._guardrails_vlan_discovery_base_vid()
            self._guardrails_log(
                f"{item_id}: reset 직후 — trunk vlan {base_vid} remove (Discovery 강제)"
            )
            st_rb, det_rb = self._guardrails_l2sw_vlan_discovery_remove_base(base_vid)
            notes.append(f"vlan_base_remove:{det_rb}")
            self._guardrails_log(f"{item_id}: VLAN base remove → {st_rb}: {det_rb}")
            if st_rb != "PASS":
                notes.append("vlan_base_remove:WARN")

        if (rst_how or "").startswith("manual"):
            self._guardrails_log(
                f"{item_id}: ACL({fam})+tcpdump 준비됨. 지금 RU를 재부팅하세요 → "
                f"{expect} 복구 감시. pcap→{dest_pcap}"
            )
        else:
            self._guardrails_log(
                f"{item_id}: reset 전송({rst_how}). SSH down→{expect} 복구 감시 "
                f"(down≤{down_detect_sec}s / up≤{timeout_sec}s). pcap→{dest_pcap}"
            )

        t_acl = time.monotonic()
        saw_down = False
        last_tick = -30
        while time.monotonic() - t_acl < down_detect_sec:
            if self._guardrails_cancel.is_set():
                return _abort_fast("사용자 중지")
            elapsed_w = int(time.monotonic() - t_acl)
            if elapsed_w - last_tick >= 30:
                last_tick = elapsed_w
                self._guardrails_log(f"{item_id}: 재부팅(SSH down) 대기 중… {elapsed_w}s/{down_detect_sec}s")
            up, pdetail = self._guardrails_probe_once(fam)
            if not up:
                saw_down = True
                notes.append(f"down:{pdetail}")
                self._guardrails_log(f"{item_id}: SSH down 감지 ({elapsed_w}s) — 복구 대기 시작")
                break
            _sleep_poll(poll_sec)
        if not saw_down:
            if self._guardrails_cancel.is_set():
                return _abort_fast("사용자 중지")
            _stop_tcpdump()
            _cleanup_vlan_discovery()
            if acl_applied:
                self._guardrails_remove_acl()
                acl_applied = False
            self._guardrails_log(f"{item_id}: FAIL — 재부팅(SSH down) 미감지")
            return "FAIL", "재부팅(SSH unhealthy) 미감지. ACL/tcpdump 정리. | " + " | ".join(notes)

        t_down = time.monotonic()
        t0 = t_down
        t_ip_learned: float | None = None
        t_boot_ok: float | None = None
        recovered = False
        last = ""
        last_tick = -30
        ignore_old_ip_logged = ""
        lease_tick = -30
        healthy_since: float | None = None
        if enable_vlan_disc:
            if fam == "v6":
                self._guardrails_log(
                    f"{item_id}: 복구 감시 = DHCPv6 IAADDR / ping6 "
                    f"(vlan {vlan_vid}, IF={solid_vlan_if or '-'}) — v4/.61 대기 안 함"
                )
            else:
                self._guardrails_log(
                    f"{item_id}: 복구 감시 = DHCPv4 Your-IP / ping "
                    f"(vlan {vlan_vid}, IF={solid_vlan_if or '-'})"
                )
        self._guardrails_log(
            f"{item_id}: 복구 PASS 조건 = SSH healthy {stable_sec}s 연속 유지 "
            f"(poll={poll_sec}s, 한 번 성공만으로 PASS 안 함)"
        )
        while time.monotonic() - t0 < timeout_sec:
            if self._guardrails_cancel.is_set():
                return _abort_fast("사용자 중지")
            elapsed_w = int(time.monotonic() - t0)
            if elapsed_w - last_tick >= 30:
                last_tick = elapsed_w
                hold = ""
                if healthy_since is not None:
                    hold = f" hold={time.monotonic() - healthy_since:.0f}/{stable_sec}s"
                self._guardrails_log(
                    f"{item_id}: 주소 복구 대기 중… {elapsed_w}s/{timeout_sec}s "
                    f"pcap_ip={learned_ip or '-'}{hold} last={last[:60]}"
                )
            # VLAN Discovery: USB pcap(vlan BPF)
            # v6: neigh(RU MAC) 우선 — pcap 마지막 IAADDR(::456) 고아 방지
            if enable_vlan_disc:
                lease_every = 10 if fam == "v6" else 30
                use_lease = elapsed_w - lease_tick >= lease_every
                if use_lease:
                    lease_tick = elapsed_w
                cand = _collect_discovery_candidate_ips(with_lease=use_lease)
                if self._guardrails_cancel.is_set():
                    return _abort_fast("사용자 중지")
                neigh_ips: list[str] = []
                # v6: 매 poll neigh (use_lease 가 아니어도) — ping fail 고착 방지
                if fam == "v6" and solid_vlan_if:
                    try:
                        neigh_ips = [
                            ip
                            for ip in self._guardrails_find_ip_by_ru_mac(
                                mac_colon, solid_vlan_if, "v6"
                            )
                            if ip
                            and self._guardrails_is_plausible_ip(ip, "v6")
                            and not self._guardrails_is_pre_discovery_ip(ip, "v6")
                        ]
                    except Exception:
                        neigh_ips = []
                # 시도 순서: neigh → pcap(최신부터). ::456 만 붙잡고 재시도하지 않음
                try_ips: list[str] = []
                for ip in neigh_ips + list(reversed(cand or [])):
                    if (
                        ip
                        and self._guardrails_is_plausible_ip(ip, fam)
                        and not self._guardrails_is_pre_discovery_ip(ip, fam)
                        and ip not in try_ips
                    ):
                        try_ips.append(ip)
                if not try_ips and not (cand or neigh_ips):
                    raw = self._guardrails_peek_pcap_assigned_ips(
                        remote_pcap, fam, timeout=8, mac_colon=mac_colon
                    )
                    if raw and raw[-1] != ignore_old_ip_logged:
                        ignore_old_ip_logged = raw[-1]
                        probe_now = self._guardrails_strip_ip_cidr(
                            self._guardrails_gf("probe_v4" if fam == "v4" else "probe_v6")
                        )
                        hint = ""
                        last_raw = raw[-1]
                        if last_raw and self._guardrails_is_pre_discovery_ip(
                            last_raw, fam
                        ):
                            if probe_now and last_raw != probe_now:
                                hint = (
                                    f" — untag {last_raw}≠probe {probe_now}: "
                                    "동일 RU MAC이면 DHCP 재할당 가능 "
                                    "(다른 장비 아님일 수 있음). Discovery(.61/tag) 대기"
                                )
                            else:
                                hint = (
                                    " — untag(.60/1200) IP라 Discovery 미인정 "
                                    "(.61/1300 tag lease 대기)"
                                )
                        self._guardrails_log(
                            f"{item_id}: Discovery IP 미검출 — raw_pcap={','.join(raw[-5:])}"
                            f"{hint} (재시도 중)"
                        )
                    for ip in reversed(raw or []):
                        if (
                            ip
                            and self._guardrails_is_plausible_ip(ip, fam)
                            and not self._guardrails_is_pre_discovery_ip(ip, fam)
                            and ip not in try_ips
                        ):
                            try_ips.append(ip)
                up = False
                if not try_ips:
                    if fam == "v6":
                        last = "Discovery IPv6(IAADDR/neigh) 대기 — v4(.61) 미사용"
                    else:
                        last = "Discovery IP(서브넷 10.0.61.x) 대기 중"
                else:
                    # learned 가 try 목록에 있으면 먼저, 실패 시 나머지 전부 ping
                    order = []
                    if learned_ip and learned_ip in try_ips:
                        order.append(learned_ip)
                    for ip in try_ips:
                        if ip not in order:
                            order.append(ip)
                    for tip in order:
                        if fam == "v6" and solid_vlan_if:
                            c6 = self._guardrails_v6_local_cidr_for_target(tip)
                            if c6:
                                self._guardrails_dhcp_host_ensure_v6_addr(
                                    solid_vlan_if, c6
                                )
                        up, last = self._guardrails_probe_from_dhcp_host(
                            tip, via_if=solid_vlan_if or None, family=fam
                        )
                        if tip != learned_ip:
                            src = "neigh" if tip in neigh_ips else "pcap"
                            self._guardrails_log(
                                f"{item_id}: Discovery IP 시도 → {tip} ({src}, "
                                f"try={','.join(order[:6])}, ping={'OK' if up else 'FAIL'})"
                            )
                        if up:
                            if tip != learned_ip:
                                learned_ip = tip
                                if t_ip_learned is None:
                                    t_ip_learned = time.monotonic()
                                notes.append(f"pcap_ip:{learned_ip}")
                                self._guardrails_log(
                                    f"{item_id}: Discovery IP 확정 → {learned_ip} "
                                    f"(reset+{t_ip_learned - t_reset:.0f}s)"
                                )
                            break
                    if not up:
                        # 표시용: neigh 우선, 없으면 첫 후보 (고아 ::456 단독 고착 방지)
                        prefer = (neigh_ips[0] if neigh_ips else order[0])
                        if prefer != learned_ip:
                            self._guardrails_log(
                                f"{item_id}: ping 미복구 — 후보 갱신 "
                                f"{learned_ip or '-'} → {prefer} (try={','.join(order[:6])})"
                            )
                            learned_ip = prefer
                            if t_ip_learned is None:
                                t_ip_learned = time.monotonic()
                            notes.append(f"pcap_ip:{learned_ip}")
            else:
                up, last = self._guardrails_probe_once(fam)
            if up:
                now_h = time.monotonic()
                if healthy_since is None:
                    healthy_since = now_h
                    self._guardrails_log(
                        f"{item_id}: 주소 복구 1차 OK — {stable_sec}s 유지 확인 중… "
                        f"(down+{elapsed_w}s) {last[:100]}"
                    )
                held = now_h - healthy_since
                if held >= float(stable_sec):
                    recovered = True
                    t_boot_ok = now_h
                    if t_ip_learned is None and (learned_ip or "").strip():
                        t_ip_learned = t_boot_ok
                    self._guardrails_log(
                        f"{item_id}: 주소 복구 OK (유지 {held:.0f}s≥{stable_sec}s, "
                        f"down+{elapsed_w}s, reset+{t_boot_ok - t_reset:.0f}s) {last[:120]}"
                    )
                    notes.append(f"stable:{stable_sec}s")
                    break
            else:
                if healthy_since is not None:
                    dropped = time.monotonic() - healthy_since
                    self._guardrails_log(
                        f"{item_id}: 복구 유지 실패 ({dropped:.0f}s/{stable_sec}s) — 재대기 "
                        f"last={last[:80]}"
                    )
                healthy_since = None
            # 유지 확인 중에는 poll 을 조금 더 촘촘히 (최대 3s)
            wait_s = min(poll_sec, 3) if healthy_since is not None else poll_sec
            _sleep_poll(wait_s)
        elapsed = time.monotonic() - t0
        boot_ok = recovered and elapsed <= pass_sec
        if not recovered:
            notes.append(f"boot:FAIL timeout {elapsed:.0f}s last={last}")
            self._guardrails_log(f"{item_id}: boot FAIL — 복구 타임아웃 {elapsed:.0f}s")
        elif not boot_ok:
            notes.append(f"boot:FAIL slow {elapsed:.0f}s>PASS≤{pass_sec} {last}")
            self._guardrails_log(f"{item_id}: boot FAIL — 복구 느림 {elapsed:.0f}s > PASS≤{pass_sec}")
        else:
            notes.append(f"boot:OK {elapsed:.0f}s {last}")

        self._guardrails_log(f"{item_id}: tcpdump 중지·pcap 디코드 중…")
        ok_stop, stop_text = _stop_tcpdump()
        notes.append(f"tcpdump_stop:{'OK' if ok_stop else 'WARN'}")
        decode = ""
        opt_ok = False
        disc_ok = True
        ia_na_ok = True
        if "NO_PCAP" in (stop_text or ""):
            notes.append("option:FAIL no pcap")
            if enable_ia_na:
                ia_na_ok = False
                notes.append("ia_na:FAIL no pcap")
            self._guardrails_log(f"{item_id}: pcap 파일 없음")
        else:
            decode_cmd = (
                f"tcpdump -nn -vv -e -r {shlex.quote(remote_pcap)} 2>/dev/null | head -c 400000; "
                f"echo; echo DECODE_BYTES=$(wc -c <{shlex.quote(remote_pcap)} 2>/dev/null || echo 0)"
            )
            _ok_dec, decode = self._guardrails_dhcp_server_run(decode_cmd, timeout=90)
            m_bytes = re.search(r"DECODE_BYTES\s*=\s*(\d+)", decode or "")
            pcap_bytes = int(m_bytes.group(1)) if m_bytes else -1
            if pcap_bytes >= 0 and pcap_bytes <= 24:
                self._guardrails_log(
                    f"{item_id}: pcap 비어 있음 (DECODE_BYTES={pcap_bytes}) — "
                    "DHCP 패킷 미캡처. IF/BPF/재부팅 타이밍 확인"
                )
                notes.append(f"pcap:EMPTY bytes={pcap_bytes}")
            else:
                self._guardrails_log(
                    f"{item_id}: pcap decode OK size≈{pcap_bytes if pcap_bytes >= 0 else '?'}B"
                )
            if enable_vlan_disc and not learned_ip:
                cand = _collect_discovery_candidate_ips()
                if not cand:
                    ips_found = self._guardrails_parse_dhcp_assigned_ips(
                        decode or "", fam, mac_colon=mac_colon
                    )
                    cand = [
                        ip
                        for ip in ips_found
                        if self._guardrails_is_plausible_ip(ip, fam)
                        and not self._guardrails_is_pre_discovery_ip(ip, fam)
                    ]
                if cand:
                    learned_ip = cand[-1]
                    if t_ip_learned is None:
                        t_ip_learned = time.monotonic()
                    notes.append(f"pcap_ip:{learned_ip}")
                    self._guardrails_log(f"{item_id}: 최종 Discovery IP → {learned_ip}")
                else:
                    self._guardrails_log(
                        f"{item_id}: Discovery IP 분석 실패 — USB pcap(vlan BPF)·neigh·lease 확인"
                    )
            persist = self._guardrails_dhcp_persist_pcap(
                remote_pcap=remote_pcap,
                decode_text=decode or "",
                dest_pcap=dest_pcap,
                dest_txt=dest_txt,
                log_path=log_path,
            )
            notes.append(persist)
            local_path = self._guardrails_dhcp_download_pcap(dest_pcap, f"{base}.pcap")
            if local_path:
                notes.append(f"local={local_path}")
                self._guardrails_log(f"pcap saved: remote={dest_pcap} local={local_path}")
                try:
                    if not hasattr(self, "_guardrails_detail_by_id") or self._guardrails_detail_by_id is None:
                        self._guardrails_detail_by_id = {}
                    # stash latest local pcap path for detail view
                    self._guardrails_detail_by_id[f"{item_id}__pcap"] = local_path
                except Exception:
                    pass
            else:
                self._guardrails_log(f"pcap on remote LOG: {dest_pcap}")

            # Empty pcap cannot match option
            if pcap_bytes >= 0 and pcap_bytes <= 24:
                opt_ok, opt_detail = False, f"expect_regex 미검사(pcap empty): {opt_re}"
            else:
                opt_ok, opt_detail = self._guardrails_match_output(decode or "", opt_re, None)
            notes.append(f"option:{'OK' if opt_ok else opt_detail}")
            self._guardrails_log(f"{item_id}: option → {'OK' if opt_ok else opt_detail}")
            if enable_ia_na:
                if pcap_bytes >= 0 and pcap_bytes <= 24:
                    ia_na_ok, ia_na_detail = False, "ia_na FAIL: pcap empty"
                else:
                    ia_na_ok, ia_na_detail = self._guardrails_check_dhcpv6_ia_na(decode or "", ia_na_re)
                notes.append(ia_na_detail)
                self._guardrails_log(f"{item_id}: {ia_na_detail}")
            else:
                notes.append("ia_na:SKIP")
            if enable_disc:
                disc_re = self._guardrails_gf("discovery_regex")
                ctrls = self._guardrails_gf("expected_controllers")
                if not ctrls:
                    # VLAN Discovery 시험 중이면 tag controller, 아니면 untag
                    plane = "tag" if enable_vlan_disc else "untag"
                    try:
                        self._apply_lab_controller_listen_ips(plane)
                    except Exception:
                        pass
                    try:
                        if fam == "v6":
                            ctrls = self._lab_controller_listen_ip("v6", plane)
                        else:
                            ctrls = self._lab_controller_listen_ip("v4", plane)
                    except Exception:
                        ctrls = ""
                ips = [x.strip() for x in re.split(r"[,;\s]+", ctrls) if x.strip()]
                if pcap_bytes >= 0 and pcap_bytes <= 24:
                    disc_ok, disc_detail = False, "pcap empty"
                else:
                    disc_ok, disc_detail = self._guardrails_match_output(decode or "", disc_re, ips)
                notes.append(f"discovery:{'OK' if disc_ok else disc_detail}")
                self._guardrails_log(f"{item_id}: discovery → {'OK' if disc_ok else disc_detail}")
            else:
                notes.append("discovery:SKIP")

        # ping 재확인은 원복(renew/vlan remove) 전에 — 원복 후 .112 ping FAIL 은 정상
        if enable_vlan_disc and boot_ok and (learned_ip or "").strip():
            up2, last2 = self._guardrails_probe_from_dhcp_host(
                learned_ip, via_if=solid_vlan_if or None, family=fam
            )
            self._guardrails_log(
                f"{item_id}: 원복 전 ping 재확인 → {'OK' if up2 else 'FAIL'}: {last2[:160]}"
            )
            if not up2:
                boot_ok = False
                notes.append(f"boot:FAIL recheck {last2}")
                self._guardrails_log(
                    f"{item_id}: ping 재확인 실패 — Discovery 구간 L3 미달로 FAIL"
                )

        # 시험 VLAN 서브넷 IP를 받았을 때만 원복 팝업 (기존 .60 은 제외)
        if (
            enable_vlan_disc
            and vlan_needs_cleanup
            and (learned_ip or "").strip()
            and not self._guardrails_is_pre_discovery_ip(learned_ip, fam)
        ):
            self._guardrails_log(
                f"{item_id}: VLAN Discovery IP 수신({learned_ip}) — untag 원복 여부 확인"
            )
            restore_now = self._guardrails_ask_vlan_untag_restore(
                vid=vlan_vid,
                learned_ip=learned_ip,
                item_id=item_id,
            )
            if restore_now:
                _cleanup_vlan_discovery()
                self._guardrails_log(
                    f"{item_id}: 원복 후 {learned_ip} ping 실패는 정상 "
                    "(renew/trunk 원복으로 Discovery 서브넷 경로 종료)"
                )
            else:
                pending = {
                    "item_id": item_id,
                    "fam": fam,
                    "vid": vlan_vid,
                    "base_vid": self._guardrails_vlan_discovery_base_vid(),
                    "learned_ip": (learned_ip or "").strip(),
                    "probe": self._guardrails_strip_ip_cidr(
                        self._guardrails_gf("probe_v4" if fam == "v4" else "probe_v6")
                    ),
                    "renew_cmd": renew_cmd,
                    "ru_mac": mac_colon,
                }
                _defer_solid_vlan_to_pending(pending)
                self.after(0, lambda p=pending: self._guardrails_arm_vlan_restore(p))
                notes.append("vlan_cleanup:DEFERRED(유지—버튼으로 원복)")
                self._guardrails_log(
                    f"{item_id}: VLAN 유지 선택 — renew/L2SW 원복 생략 "
                    "(「VLAN untag 원복」 누를 때 renew 전송)"
                )
                vlan_needs_cleanup = False
        else:
            _cleanup_vlan_discovery()
        if acl_applied:
            rm_st, rm_detail = self._guardrails_remove_acl()
            # ACL 잔존은 시험 결과에 반영 (v4 leasefail 등 현장 장애의 대표 원인)
            if rm_st == "PASS":
                notes.append(rm_detail)
            else:
                notes.append(f"ACL원복FAIL:{rm_detail}")
            self._guardrails_log(f"{item_id}: ACL 원복 → {rm_st}: {rm_detail[:160]}")
            acl_applied = False
        else:
            rm_st, rm_detail = "PASS", "acl:SKIP"
            notes.append(rm_detail)
            self._guardrails_log(f"{item_id}: ACL 원복 생략 (미적용)")

        self._guardrails_log(
            f"{item_id}: 최종판정 boot(ping)={'OK' if boot_ok else 'FAIL'} "
            f"option(pcap)={'OK' if opt_ok else 'FAIL'} "
            f"discovery={'OK' if disc_ok else 'FAIL'} ia_na={'OK' if ia_na_ok else 'FAIL'} "
            f"acl_restore={'OK' if rm_st == 'PASS' else 'FAIL'}"
        )
        st_final = (
            "PASS"
            if (boot_ok and opt_ok and disc_ok and ia_na_ok and rm_st == "PASS")
            else "FAIL"
        )
        sec_reset_to_down = (t_down - t_reset) if t_down and t_reset else None
        sec_reset_to_ip = (
            (t_ip_learned - t_reset) if t_ip_learned is not None else None
        )
        sec_reset_to_boot = (t_boot_ok - t_reset) if t_boot_ok is not None else None
        sec_down_to_boot = (t_boot_ok - t_down) if t_boot_ok is not None else None
        detail_fmt = self._guardrails_format_dhcp_options_detail(
            item_id=item_id,
            status=st_final,
            fam=fam,
            boot_ok=boot_ok,
            opt_ok=opt_ok,
            disc_ok=disc_ok,
            ia_na_ok=ia_na_ok,
            learned_ip=learned_ip or "",
            decode=decode or "",
            opt_re=opt_re or "",
            notes=notes,
            sec_reset_to_down=sec_reset_to_down,
            sec_reset_to_ip=sec_reset_to_ip,
            sec_reset_to_boot=sec_reset_to_boot,
            sec_down_to_boot=sec_down_to_boot,
        )
        return st_final, detail_fmt

    def _guardrails_dhcp_server_run(self, remote_body: str, timeout: int = 60) -> tuple[bool, str]:
        """Run bash on DHCP server (direct on solid if dhcp_host empty).

        solid→sshpass→dhcp_host 로 중첩 quote 되면 `sudo -S` / `ip addr add` 가 깨짐.
        본문은 base64 로 실어 보내 원격에서만 decode|bash.
        """
        import base64

        b64 = base64.b64encode((remote_body or "").encode("utf-8")).decode("ascii")
        # wrapper 는 안전 charset 만. 파일로 풀어 stdin 파이프가 ssh/sshpass 를 먹지 않게 함
        wrapped = (
            f"f=/tmp/ds_run_$$.sh; echo {b64} | base64 -d >\"$f\" && bash \"$f\"; "
            f"ec=$?; rm -f \"$f\"; exit $ec"
        )
        host = self._guardrails_gf("dhcp_host")
        if not host:
            return self._guardrails_ssh_exec(f"bash -lc {shlex.quote(wrapped)}", timeout=timeout)
        user = self._guardrails_gf("dhcp_id")
        pw = self._guardrails_gf("dhcp_pw")
        port = "22"
        if not user:
            return False, "DHCP SSH ID 필요 (dhcp_host 지정 시)"
        inner = f"bash -lc {shlex.quote(wrapped)}"
        jump = (
            "export SSHPASS=" + shlex.quote(pw) + "; "
            "sshpass -e ssh -n "
            f"-p {shlex.quote(port)} "
            "-o StrictHostKeyChecking=no -o UserKnownHostsFile=/dev/null -o LogLevel=ERROR "
            "-o ConnectTimeout=10 -o BatchMode=no "
            + shlex.quote(f"{user}@{host}")
            + " "
            + shlex.quote(inner)
        )
        return self._guardrails_ssh_exec(f"bash -lc {shlex.quote(jump)}", timeout=timeout)

    def _guardrails_dhcp_persist_pcap(
        self,
        *,
        remote_pcap: str,
        decode_text: str,
        dest_pcap: str,
        dest_txt: str,
        log_path: str,
    ) -> str:
        """Copy pcap into solid LOG_PATH and write decode txt beside it."""
        host = self._guardrails_gf("dhcp_host")
        b64 = ""
        try:
            import base64

            b64 = base64.b64encode((decode_text or "")[:350000].encode("utf-8", errors="replace")).decode("ascii")
        except Exception:
            b64 = ""
        if b64:
            write_txt = (
                f"mkdir -p {shlex.quote(log_path)}; "
                "python3 -c "
                + shlex.quote(
                    "import base64,pathlib; pathlib.Path(%r).write_bytes(base64.b64decode(%r)); print('TXT_OK')"
                    % (dest_txt, b64)
                )
            )
        else:
            write_txt = f"mkdir -p {shlex.quote(log_path)}; echo NO_DECODE >{shlex.quote(dest_txt)}"

        if not host:
            cmd = (
                f"mkdir -p {shlex.quote(log_path)}; "
                f"cp -f {shlex.quote(remote_pcap)} {shlex.quote(dest_pcap)}; "
                f"{write_txt}; "
                f"ls -la {shlex.quote(dest_pcap)}"
            )
            ok, text = self._guardrails_ssh_exec(f"bash -lc {shlex.quote(cmd)}", timeout=60)
            return f"persist solid:{'OK' if ok else text[:80]}"

        user = self._guardrails_gf("dhcp_id")
        pw = self._guardrails_gf("dhcp_pw")
        port = "22"
        cmd = (
            f"mkdir -p {shlex.quote(log_path)}; "
            "export SSHPASS=" + shlex.quote(pw) + "; "
            f"sshpass -e scp -P {shlex.quote(port)} "
            "-o StrictHostKeyChecking=no -o UserKnownHostsFile=/dev/null -o LogLevel=ERROR "
            + shlex.quote(f"{user}@{host}:{remote_pcap}")
            + " "
            + shlex.quote(dest_pcap)
            + "; "
            f"{write_txt}; "
            f"ls -la {shlex.quote(dest_pcap)}"
        )
        ok, text = self._guardrails_ssh_exec(f"bash -lc {shlex.quote(cmd)}", timeout=90)
        return f"persist scp:{'OK' if ok else (text or '')[:80]}"

    def _guardrails_dhcp_download_pcap(self, remote_pcap: str, local_name: str) -> str:
        """SFTP pcap from solid to local AppData O-RAN-Netconf/pcaps/."""
        ssh_user = self.remote_user.get().strip()
        ssh_host = self.remote_host.get().strip()
        ssh_port = self.remote_port.get().strip() or "22"
        ssh_password = self.remote_password.get()
        if not ssh_user or not ssh_host:
            return ""
        try:
            local_dir = self.config_path.parent / "pcaps"
            local_dir.mkdir(parents=True, exist_ok=True)
            local_path = local_dir / local_name
            import paramiko  # type: ignore

            client = paramiko.SSHClient()
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            client.connect(
                hostname=ssh_host,
                port=int(ssh_port),
                username=ssh_user,
                password=ssh_password or None,
                timeout=12,
                allow_agent=True,
                look_for_keys=True,
            )
            try:
                sftp = client.open_sftp()
                try:
                    sftp.get(remote_pcap, str(local_path))
                finally:
                    sftp.close()
            finally:
                client.close()
            return str(local_path)
        except Exception as exc:
            self.after(0, self.append_log, f"[M-Plane Test] local pcap download 실패: {exc}\n")
            return ""

    def _guardrails_run_dhcp_boot(self) -> tuple[str, str]:
        """Legacy combined entry — prefer selecting ACL v4/v6 items separately."""
        return "SKIP", "ACL은 MP-DHCP-ACLv4 / MP-DHCP-ACLv6로 단독 실행하세요"

    def _guardrails_run_vlan_discovery(self) -> tuple[str, str]:
        """Standalone item retired — use dhcp_v4_vlan / dhcp_v6_vlan."""
        return (
            "SKIP",
            "MP-VLAN-1은 MP-DHCPv4-VLAN / MP-DHCPv6-VLAN 항목으로 분리됨 "
            "(Boot=L2SW없음 재시작, VLAN=L2SW trunk/ACL Discovery)",
        )

    @staticmethod
    def _guardrails_normalize_mac(mac: str) -> str:
        raw = re.sub(r"[^0-9A-Fa-f]", "", mac or "")
        if len(raw) != 12:
            return ""
        return raw.lower()

    @classmethod
    def _guardrails_mac_to_linklocal(cls, mac: str) -> str:
        """EUI-64 link-local (fe80::...) from MAC."""
        raw = cls._guardrails_normalize_mac(mac)
        if not raw:
            return ""
        b = [int(raw[i : i + 2], 16) for i in range(0, 12, 2)]
        b[0] ^= 0x02
        eui = b[0:3] + [0xFF, 0xFE] + b[3:6]
        parts = [
            f"{eui[0]:02x}{eui[1]:02x}",
            f"{eui[2]:02x}{eui[3]:02x}",
            f"{eui[4]:02x}{eui[5]:02x}",
            f"{eui[6]:02x}{eui[7]:02x}",
        ]
        return "fe80::" + ":".join(parts)

    def _guardrails_ru_linklocal_scoped(self, *, tag: bool) -> tuple[str, str]:
        """사전 계산 가능한 RU link-local + dhcp_host scope IF.

        Returns (fe80::…%ifname, detail). MAC 없으면 ("", reason).
        tag=True → vlan<VID> (Discovery), tag=False → Capture IF(enx).
        """
        mac = (self._guardrails_gf("ru_mac") or "").strip()
        ll = self._guardrails_mac_to_linklocal(mac)
        if not ll:
            return "", "RU MAC 없음 — LL 계산 불가"
        parent, vlan_if, _cidr = self._guardrails_capture_host_vlan_if_names()
        if tag:
            ifname = (vlan_if or "").strip()
            if not ifname:
                return "", "tag scope IF(vlan) 없음"
        else:
            ifname = (self._guardrails_gf("dhcp_if") or parent or "").strip()
            if not ifname:
                return "", "untag scope IF(enx) 없음"
        return f"{ll}%{ifname}", f"EUI-64 MAC={mac} → {ll}%{ifname}"

    @staticmethod
    def _guardrails_strip_ip_cidr(addr: str) -> str:
        """Strip trailing /prefix from IPv4/IPv6 (SSH/CallHome host must be bare address)."""
        a = (addr or "").strip()
        if not a:
            return ""
        if "/" in a and not a.startswith("/"):
            left, right = a.rsplit("/", 1)
            if right.isdigit() and 0 <= int(right) <= 128:
                return left.strip()
        return a

    def _guardrails_ru_ssh_target(self, family: str = "v6") -> tuple[str, str]:
        """Return (ssh_host, detail). Settings ALLOWED_IP* SSOT → probe 미러 → LL."""
        fam = (family or "v6").lower()
        if fam == "v4":
            v4 = ""
            try:
                v4 = self._guardrails_strip_ip_cidr(
                    (self.fields.get("ALLOWED_IP").get() or "").strip()  # type: ignore[union-attr]
                )
            except Exception:
                v4 = ""
            if not v4:
                v4 = self._guardrails_strip_ip_cidr(self._guardrails_gf("probe_v4"))
            if not v4:
                return "", "RU IPv4 없음 (Settings ALLOWED_IP)"
            return v4, f"v4 {v4}"
        # IPv6: Settings ALLOWED_IP_V6 → probe 미러 → LL last resort
        override = ""
        try:
            override = self._guardrails_strip_ip_cidr(
                (self.fields.get("ALLOWED_IP_V6").get() or "").strip()  # type: ignore[union-attr]
            )
        except Exception:
            override = ""
        if not override:
            override = self._guardrails_strip_ip_cidr(self._guardrails_gf("probe_v6"))
        nic = self._guardrails_gf("mplane_if")
        if override:
            if override.lower().startswith("fe80:") and "%" not in override and nic:
                return f"{override}%{nic}", f"override-ll {override}%{nic}"
            return override, f"v6-global {override}"
        mac = self._guardrails_gf("ru_mac")
        ll = self._guardrails_mac_to_linklocal(mac)
        if not ll:
            return "", "RU IPv6 없음 (Settings ALLOWED_IP_V6 / RU MAC)"
        if not nic:
            return "", "M-Plane NIC 필요 (LL SSH scope) — global이면 ALLOWED_IP_V6 사용"
        return f"{ll}%{nic}", f"ll {ll}%{nic}"

    def _guardrails_ru_check_cmd(self, family: str = "v6") -> str:
        """Command run on RU via SSH — PASS when healthy regex matches stdout."""
        ru_if = self._guardrails_gf("ru_if_name")
        fam = (family or "v6").lower()
        if fam == "v4":
            if ru_if:
                return (
                    f"ip -4 -o addr show dev {shlex.quote(ru_if)} 2>/dev/null; "
                    f"ip -o link show dev {shlex.quote(ru_if)} 2>/dev/null | head -1"
                )
            return "ip -4 -o addr show 2>/dev/null | grep -v '127\\.0\\.0\\.1'"
        if ru_if:
            return (
                f"ip -6 -o addr show dev {shlex.quote(ru_if)} scope global 2>/dev/null; "
                f"ip -o link show dev {shlex.quote(ru_if)} 2>/dev/null | head -1"
            )
        return "ip -6 -o addr show scope global 2>/dev/null"

    def _guardrails_probe_once(
        self, family: str | None = None, host_override: str | None = None
    ) -> tuple[bool, str]:
        """SSH into RU and require family address — not LL/v4 ping alone."""
        iid = getattr(self, "_guardrails_settings_item_id", None) or "dhcp_boot"
        fam = (family or self._guardrails_item_family(iid)).lower()
        if host_override:
            host = self._guardrails_strip_ip_cidr(host_override)
            how = f"pcap-ip {host}"
        else:
            host, how = self._guardrails_ru_ssh_target(fam)
        if not host:
            return False, how
        user = self._guardrails_gf("oru_cli_id")
        pw = self._guardrails_gf("oru_cli_pw")
        if not user:
            return False, "RU SSH ID 필요 (Settings ★ RU SSH ID)"
        check = self._guardrails_ru_check_cmd(fam)
        ssh_flag = "-4" if fam == "v4" else "-6"
        remote = (
            "export SSHPASS=" + shlex.quote(pw) + "; "
            f"sshpass -e ssh {ssh_flag} -n "
            "-o StrictHostKeyChecking=no -o UserKnownHostsFile=/dev/null -o LogLevel=ERROR "
            "-o ConnectTimeout=5 -o BatchMode=no "
            + shlex.quote(f"{user}@{host}")
            + " "
            + shlex.quote(check)
        )
        ok, text = self._guardrails_ssh_exec(f"bash -lc {shlex.quote(remote)}", timeout=20)
        # OpenSSH host-key 안내 문구는 실패 원인 아님 — 로그 노이즈 제거
        clean = re.sub(
            r"(?im)^\s*Warning:\s*Permanently added[^\n]*\n?",
            "",
            text or "",
        ).strip()
        if not ok:
            snip = re.sub(r"\s+", " ", clean)[:180]
            return False, f"ssh-fail via {how}: {snip or 'connect/auth fail'}"
        text = clean or text
        if fam == "v4":
            # Prefer inet/CIDR from RU output when present; otherwise SSH-to-IPv4 alone = healthy.
            pat = (
                self._guardrails_gf("healthy_regex_v4", r"inet\s+\d+\.\d+\.\d+\.\d+/")
                or r"inet\s+\d+\.\d+\.\d+\.\d+/"
            )
            try:
                matched = re.search(pat, text or "", re.I | re.M) is not None
            except re.error as exc:
                return False, f"healthy_regex_v4 오류: {exc}"
            found = [
                m.group(1)
                for m in re.finditer(r"inet\s+(\d+\.\d+\.\d+\.\d+)/", text or "")
                if not m.group(1).startswith("127.")
            ]
            if matched and found:
                return True, f"healthy {found[0]} via {how}"
            if matched:
                return True, f"healthy(regex) via {how}"
            return True, f"healthy ssh-ok via {how} (no inet/CIDR in RU output)"

        pat = self._guardrails_gf("healthy_regex", r"inet6\s+[0-9a-fA-F:]+/") or r"inet6\s+[0-9a-fA-F:]+/"
        try:
            matched = re.search(pat, text or "", re.I | re.M) is not None
        except re.error as exc:
            return False, f"healthy_regex 오류: {exc}"
        if matched:
            globals_found = [
                m.group(1)
                for m in re.finditer(r"inet6\s+([0-9a-fA-F:]+)/", text or "", re.I)
                if not m.group(1).lower().startswith("fe80:")
            ]
            if globals_found:
                return True, f"healthy {globals_found[0]} via {how}"
            if matched:
                return True, f"healthy(regex) via {how}"
        # SSH succeeded on v6 path (LL or global) — treat as recovered even without inet6 dump
        return True, f"healthy ssh-ok via {how} (no inet6/CIDR in RU output)"

    def _guardrails_l2sw_normalize_if(self, raw: str | None = None) -> str:
        """Dasan: return 'ethernet 0/22' only (no leading 'interface')."""
        ifc = (raw if raw is not None else self._guardrails_gf("l2sw_if", "")) or ""
        ifc = ifc.strip()
        if not ifc:
            return ""
        low = ifc.lower()
        # ⚙에 'interface ethernet 0/22'로 넣으면 'interface interface …' 중복 방지
        if low.startswith("interface "):
            ifc = ifc.split(None, 1)[1].strip()
            low = ifc.lower()
        if low.startswith("ethernet"):
            return "ethernet" + ifc[len("ethernet") :]
        if low.startswith("eth ") or low == "eth":
            return "ethernet" + ifc[3:]
        if re.match(r"^\d+/\d+", ifc):
            return f"ethernet {ifc}"
        return ifc

    @staticmethod
    def _guardrails_mac_display_forms(mac_norm: str) -> tuple[str, str, str, str]:
        """Return (colon, hyphen, dotted, bare) from 12-hex normalized MAC."""
        n = (mac_norm or "").lower()
        if len(n) != 12:
            return "", "", "", ""
        parts = [n[i : i + 2] for i in range(0, 12, 2)]
        colon = ":".join(parts)
        hyphen = "-".join(parts)
        dotted = f"{n[0:4]}.{n[4:8]}.{n[8:12]}"
        return colon, hyphen, dotted, n

    def _guardrails_parse_mac_table_port(self, text: str, mac_norm: str) -> str:
        """Parse Dasan/Cisco-like MAC table output → 'ethernet X/Y' for matching MAC."""
        n = self._guardrails_normalize_mac(mac_norm) or self._guardrails_normalize_mac(
            "".join(ch for ch in (mac_norm or "") if ch.isalnum())
        )
        if not text or len(n) != 12:
            return ""
        colon, hyphen, dotted, bare = self._guardrails_mac_display_forms(n)
        mac_pat = re.compile(
            rf"(?:{re.escape(colon)}|{re.escape(hyphen)}|{re.escape(dotted)}|{re.escape(bare)})",
            re.I,
        )
        if_pat = re.compile(
            r"(?:interface\s+)?(?:ethernet|eth)\s*(\d+\s*/\s*\d+)"
            r"|(?:\bGi|\bTe|\bFa)\s*(\d+/\d+)"
            r"|(?<![\d.])(\d+/\d+)(?![\d.])",
            re.I,
        )
        hits: list[str] = []
        for line in (text or "").splitlines():
            if not mac_pat.search(line):
                continue
            if re.search(r"\bCPU\b", line, re.I):
                continue
            m = if_pat.search(line)
            if not m:
                continue
            port = (m.group(1) or m.group(2) or m.group(3) or "").replace(" ", "")
            if not port or not re.fullmatch(r"\d+/\d+", port):
                continue
            ifc = self._guardrails_l2sw_normalize_if(f"ethernet {port}")
            if ifc and ifc not in hits:
                hits.append(ifc)
        if not hits:
            return ""
        # Prefer first physical ethernet hit (usually the RU access port)
        return hits[0]

    def _guardrails_l2sw_discover_if_by_ru_mac(
        self, mac: str | None = None
    ) -> tuple[bool, str, str]:
        """enable → show mac* → parse port for RU MAC. Returns (ok, ifc, detail)."""
        mac_raw = (mac if mac is not None else self._guardrails_gf("ru_mac") or "").strip()
        mac_norm = self._guardrails_normalize_mac(mac_raw)
        if not mac_norm:
            return False, "", "RU MAC 없음"
        colon, _hy, dotted, _bare = self._guardrails_mac_display_forms(mac_norm)
        # Dasan M3500: enable 후 show mac / show mac address-table …
        cmds = [
            "enable",
            "terminal length 0",
            f"show mac address-table address {colon}",
            f"show mac address-table address {dotted}",
            "show mac address-table",
            "show mac",
        ]
        ok, out = self._guardrails_l2sw_run_cmds(
            cmds,
            settle_s=1.2,
            timeout=180,
            require_config=False,
            tolerate_invalid=True,
        )
        ifc = self._guardrails_parse_mac_table_port(out or "", mac_norm)
        if ifc:
            return True, ifc, f"show mac → {ifc} (MAC {colon})"
        snippet = re.sub(r"\s+", " ", (out or "")[-280:]).strip()
        if not ok and not snippet:
            return False, "", "L2SW show mac 실패 (출력 없음)"
        return False, "", f"MAC {colon} 포트 미발견 (show mac): {snippet or '(empty)'}"

    def _guardrails_ensure_l2sw_if_from_mac(
        self, item_id: str | None = None, *, force: bool = False
    ) -> tuple[str, str]:
        """Lookup RU MAC on L2SW and auto-set l2sw_if. PASS/WARN/FAIL."""
        iid = item_id or getattr(self, "_guardrails_settings_item_id", None) or "dhcp_v4"
        self._guardrails_fill_defaults_from_context(iid)
        mac = (self._guardrails_gf("ru_mac") or "").strip()
        mac_norm = self._guardrails_normalize_mac(mac)
        if not mac_norm:
            cur = self._guardrails_l2sw_normalize_if()
            if cur:
                return "WARN", f"RU MAC 없음 — 기존 L2SW IF 유지 ({cur})"
            return "FAIL", "RU MAC 필요 (L2SW IF 자동 조회)"
        if not self._guardrails_gf("l2sw_ip") or not self._guardrails_gf("l2sw_id"):
            return "FAIL", "L2SW IP/ID 필요 (MAC→IF 조회)"
        cache = getattr(self, "_guardrails_l2sw_if_cache", None)
        now = time.monotonic()
        if (
            not force
            and isinstance(cache, dict)
            and cache.get("mac") == mac_norm
            and (now - float(cache.get("ts") or 0)) < 90.0
            and cache.get("ifc")
        ):
            ifc_c = self._guardrails_l2sw_normalize_if(str(cache.get("ifc") or ""))
            if ifc_c:
                old = self._guardrails_l2sw_normalize_if()
                if old != ifc_c:
                    sk = self._guardrails_store_key(iid)
                    cur_blob = dict(
                        (getattr(self, "_guardrails_per_test_settings", {}) or {}).get(sk) or {}
                    )
                    cur_blob["l2sw_if"] = ifc_c
                    self._guardrails_set_vals(iid, cur_blob)
                return "PASS", f"cached show mac → {ifc_c}"
        self._guardrails_log(f"{iid}: L2SW IF 자동 조회 (enable → show mac, RU MAC={mac})…")
        ok, ifc, detail = self._guardrails_l2sw_discover_if_by_ru_mac(mac)
        if not ok or not ifc:
            cur = self._guardrails_l2sw_normalize_if()
            if cur and not force:
                self._guardrails_log(f"{iid}: L2SW IF 조회 실패 — 기존 IF 유지 ({cur}): {detail}")
                return "WARN", f"{detail} · 기존 IF={cur}"
            return "FAIL", detail
        self._guardrails_l2sw_if_cache = {"mac": mac_norm, "ifc": ifc, "ts": now}
        old = self._guardrails_l2sw_normalize_if()
        if old == ifc:
            self._guardrails_log(f"{iid}: L2SW IF 확인 OK — {ifc}")
            return "PASS", detail
        sk = self._guardrails_store_key(iid)
        cur_blob = dict((getattr(self, "_guardrails_per_test_settings", {}) or {}).get(sk) or {})
        cur_blob["l2sw_if"] = ifc
        self._guardrails_set_vals(iid, cur_blob)
        self._guardrails_log(f"{iid}: L2SW IF 자동 설정 {old or '-'} → {ifc}")
        return "PASS", f"{detail} · 설정 {old or '-'}→{ifc}"

    def _guardrails_parse_lab_capture_if(
        self, text: str, *, prefer_tag: bool = False
    ) -> tuple[str, str]:
        """Parse ifconfig / ip -o addr → (ifname, inet) on 10.0.60.x or 10.0.61.x."""
        raw = text or ""
        hits60: list[tuple[str, str]] = []
        hits61: list[tuple[str, str]] = []
        # ip -o -4 addr: "2: enx00e04c681c64    inet 10.0.60.99/8 ..."
        for m in re.finditer(
            r"(?m)^\d+:\s+(\S+?)(?:@\S+)?\s+inet\s+(10\.0\.(60|61)\.\d+)(?:/\d+)?",
            raw,
        ):
            name, ip, plane = m.group(1), m.group(2), m.group(3)
            name = name.split("@", 1)[0].strip().rstrip(":")
            if plane == "61":
                hits61.append((name, ip))
            else:
                hits60.append((name, ip))
        # ifconfig blocks
        cur_if = ""
        for line in raw.splitlines():
            m_if = re.match(r"^(\S+?):\s+flags=", line)
            if m_if:
                cur_if = m_if.group(1).strip()
                continue
            m_inet = re.search(r"\binet\s+(10\.0\.(60|61)\.\d+)\b", line)
            if m_inet and cur_if:
                ip, plane = m_inet.group(1), m_inet.group(2)
                pair = (cur_if, ip)
                if plane == "61":
                    if pair not in hits61:
                        hits61.append(pair)
                else:
                    if pair not in hits60:
                        hits60.append(pair)
        # Prefer parent USB NIC (enx…) over vlan subifs like enx….100
        def _rank(name: str) -> tuple[int, str]:
            n = name.lower()
            if "." in n or "@" in n:
                return (2, n)
            if n.startswith("enx"):
                return (0, n)
            return (1, n)

        if prefer_tag and hits61:
            hits61.sort(key=lambda t: _rank(t[0]))
            return hits61[0]
        if hits60:
            hits60.sort(key=lambda t: _rank(t[0]))
            return hits60[0]
        if hits61:
            hits61.sort(key=lambda t: _rank(t[0]))
            return hits61[0]
        return "", ""

    def _guardrails_ensure_capture_if(
        self, item_id: str | None = None, *, prefer_tag: bool = False, force: bool = False
    ) -> tuple[str, str]:
        """dhcp_host ifconfig/ip addr → Capture IF (10.0.60/61) 자동 설정."""
        iid = item_id or getattr(self, "_guardrails_settings_item_id", None) or "dhcp_v4"
        self._guardrails_fill_defaults_from_context(iid)
        cur = (self._guardrails_gf("dhcp_if") or "").strip()
        cache = getattr(self, "_guardrails_capture_if_cache", None)
        now = time.monotonic()
        if (
            not force
            and isinstance(cache, dict)
            and (now - float(cache.get("ts") or 0)) < 90.0
            and cache.get("ifc")
            and bool(cache.get("tag")) == bool(prefer_tag)
        ):
            ifc_c = str(cache.get("ifc") or "").strip()
            if ifc_c:
                if cur != ifc_c:
                    sk = self._guardrails_store_key(iid)
                    blob = dict(
                        (getattr(self, "_guardrails_per_test_settings", {}) or {}).get(sk) or {}
                    )
                    blob["dhcp_if"] = ifc_c
                    self._guardrails_set_vals(iid, blob)
                return "PASS", f"cached capture IF={ifc_c}"
        where = self._guardrails_gf("dhcp_host") or "solid"
        self._guardrails_log(
            f"{iid}: Capture IF 자동 조회 (dhcp_host={where} ifconfig/ip addr, "
            f"prefer={'tag.61' if prefer_tag else 'untag.60'})…"
        )
        cmd = (
            "echo '=== ip -o -4 addr ==='; "
            "ip -o -4 addr show 2>/dev/null || true; "
            "echo '=== ifconfig -a ==='; "
            "ifconfig -a 2>/dev/null || true"
        )
        ok, out = self._guardrails_dhcp_server_run(cmd, timeout=30)
        ifc, inet = self._guardrails_parse_lab_capture_if(out or "", prefer_tag=prefer_tag)
        if not ifc:
            if cur and not force:
                self._guardrails_log(
                    f"{iid}: Capture IF 미발견 — 기존 유지 ({cur})"
                    + (f" ok={ok}" if not ok else "")
                )
                return "WARN", f"10.0.60/61 IF 미발견 · 기존={cur}"
            snip = re.sub(r"\s+", " ", (out or "")[:220]).strip()
            return "FAIL", f"Capture IF 미발견 (10.0.60.x/61.x): {snip or '(empty)'}"
        self._guardrails_capture_if_cache = {
            "ifc": ifc,
            "inet": inet,
            "tag": bool(prefer_tag),
            "ts": now,
        }
        if cur == ifc:
            self._guardrails_log(f"{iid}: Capture IF 확인 OK — {ifc} ({inet})")
            return "PASS", f"{ifc} ({inet})"
        sk = self._guardrails_store_key(iid)
        blob = dict((getattr(self, "_guardrails_per_test_settings", {}) or {}).get(sk) or {})
        blob["dhcp_if"] = ifc
        self._guardrails_set_vals(iid, blob)
        self._guardrails_log(f"{iid}: Capture IF 자동 설정 {cur or '-'} → {ifc} ({inet})")
        return "PASS", f"{cur or '-'}→{ifc} ({inet})"

    def _guardrails_ensure_ru_mac(
        self, item_id: str | None = None, *, force: bool = True
    ) -> tuple[str, str]:
        """RU MAC 자동: 시험 family IP → dhcp_host ping → neigh (UI 숨김).

        v4 시험: Settings ALLOWED_IP / probe_v4 + ping/arp/ip -4 neigh
        v6 시험: Settings ALLOWED_IP_V6 / probe_v6 + ping6/ip -6 neigh
        """
        iid = item_id or getattr(self, "_guardrails_settings_item_id", None) or "dhcp_v4"
        self._guardrails_fill_defaults_from_context(iid)
        self._guardrails_sync_probe_from_settings(iid)
        cur = self._guardrails_normalize_mac(self._guardrails_gf("ru_mac"))
        fam = self._guardrails_resolve_ssh_family(iid)

        ru_ip = ""
        if fam == "v6":
            try:
                ru_ip = self._guardrails_strip_ip_cidr(
                    (self.fields.get("ALLOWED_IP_V6").get() or "").strip()  # type: ignore[union-attr]
                )
            except Exception:
                ru_ip = ""
            if not ru_ip:
                ru_ip = self._guardrails_strip_ip_cidr(self._guardrails_gf("probe_v6"))
            # strip scope if present for neigh lookup
            ru_ip = (ru_ip or "").split("%", 1)[0].strip()
            if not ru_ip or ":" not in ru_ip:
                if cur and not force:
                    colon = ":".join(cur[i : i + 2] for i in range(0, 12, 2))
                    return "WARN", f"ALLOWED_IP_V6 없음 · 기존 MAC={colon}"
                return "FAIL", "RU MAC 조회용 Settings ★ RU IPv6(ALLOWED_IP_V6) 필요"
        else:
            try:
                ru_ip = self._guardrails_strip_ip_cidr(
                    (self.fields.get("ALLOWED_IP").get() or "").strip()  # type: ignore[union-attr]
                )
            except Exception:
                ru_ip = ""
            if not ru_ip:
                ru_ip = self._guardrails_strip_ip_cidr(self._guardrails_gf("probe_v4"))
            if not ru_ip or ":" in ru_ip:
                if cur and not force:
                    colon = ":".join(cur[i : i + 2] for i in range(0, 12, 2))
                    return "WARN", f"ALLOWED_IP 없음 · 기존 MAC={colon}"
                return "FAIL", "RU MAC 조회용 Settings ★ RU IPv4(ALLOWED_IP) 필요"

        cache = getattr(self, "_guardrails_ru_mac_cache", None)
        now = time.monotonic()
        if (
            not force
            and isinstance(cache, dict)
            and cache.get("ip") == ru_ip
            and cache.get("fam") == fam
            and cache.get("mac")
            and (now - float(cache.get("ts") or 0)) < 90.0
        ):
            mac_c = self._guardrails_normalize_mac(str(cache.get("mac") or ""))
            if mac_c:
                colon = ":".join(mac_c[i : i + 2] for i in range(0, 12, 2))
                if cur != mac_c:
                    sk = self._guardrails_store_key(iid)
                    blob = dict(
                        (getattr(self, "_guardrails_per_test_settings", {}) or {}).get(sk) or {}
                    )
                    blob["ru_mac"] = colon
                    self._guardrails_set_vals(iid, blob)
                return "PASS", f"cached neigh [{fam}] {ru_ip}→{colon}"

        self._guardrails_log(
            f"{iid}: RU MAC 자동 조회 [{fam}] (ping → neigh {ru_ip})…"
        )
        if fam == "v6":
            cmd = (
                f"ping -6 -c 2 -W 1 {shlex.quote(ru_ip)} >/dev/null 2>&1 || "
                f"ping6 -c 2 -W 1 {shlex.quote(ru_ip)} >/dev/null 2>&1 || "
                f"ping -6 -c 1 -w 2 {shlex.quote(ru_ip)} >/dev/null 2>&1 || true; "
                f"ip -6 neigh show {shlex.quote(ru_ip)} 2>/dev/null; "
                f"ip neigh show {shlex.quote(ru_ip)} 2>/dev/null || true"
            )
        else:
            cmd = (
                f"ping -4 -c 2 -W 1 {shlex.quote(ru_ip)} >/dev/null 2>&1 || "
                f"ping -c 2 -W 1 {shlex.quote(ru_ip)} >/dev/null 2>&1 || "
                f"ping -c 1 -w 2 {shlex.quote(ru_ip)} >/dev/null 2>&1 || true; "
                f"ip -4 neigh show {shlex.quote(ru_ip)} 2>/dev/null; "
                f"ip neigh show {shlex.quote(ru_ip)} 2>/dev/null; "
                f"arp -n {shlex.quote(ru_ip)} 2>/dev/null || true"
            )
        _ok, out = self._guardrails_dhcp_server_run(cmd, timeout=25)
        mac_n = ""
        for m in re.finditer(
            r"([0-9a-fA-F]{2}(?::[0-9a-fA-F]{2}){5})", out or "", re.I
        ):
            cand = self._guardrails_normalize_mac(m.group(1))
            if cand and cand != "000000000000" and not cand.startswith("ffff"):
                mac_n = cand
                break
        if not mac_n:
            m2 = re.search(
                rf"{re.escape(ru_ip)}\s+\S+\s+([0-9a-fA-F:]{{17}})", out or "", re.I
            )
            if m2:
                mac_n = self._guardrails_normalize_mac(m2.group(1))
        if not mac_n:
            if cur and not force:
                colon = ":".join(cur[i : i + 2] for i in range(0, 12, 2))
                return "WARN", f"neigh MAC 미발견 · 기존={colon}"
            snip = re.sub(r"\s+", " ", (out or "")[:180]).strip()
            need = "ALLOWED_IP_V6" if fam == "v6" else "ALLOWED_IP"
            return (
                "FAIL",
                f"RU MAC neigh 미발견 [{fam}] ({ru_ip}) — dhcp_host에서 {need} reachability 확인"
                + (f": {snip}" if snip else ""),
            )

        colon = ":".join(mac_n[i : i + 2] for i in range(0, 12, 2))
        self._guardrails_ru_mac_cache = {
            "ip": ru_ip,
            "mac": mac_n,
            "fam": fam,
            "ts": now,
        }
        sk = self._guardrails_store_key(iid)
        blob = dict((getattr(self, "_guardrails_per_test_settings", {}) or {}).get(sk) or {})
        blob["ru_mac"] = colon
        self._guardrails_set_vals(iid, blob)
        self._guardrails_log(f"{iid}: RU MAC ← [{fam}] ping/neigh {ru_ip} ({colon})")
        return "PASS", f"neigh [{fam}] {ru_ip}→{colon}"

    def _guardrails_acl_apply_cmds(self, family: str = "v6") -> list[str]:
        """Dasan ACL — 포트에 access-group 만 적용 (ACL 본문은 스위치에 유지).

        v4 시험: ipv6 access-group v6-dhcp-block (546/547 drop)
        v6 시험: ip access-group <acl_num> (67/68 drop)
        ACL list 생성/삭제 안 함 — 사전 등록된 list 가정.
        """
        acl = self._guardrails_gf("acl_num", "110") or "110"
        ifc = self._guardrails_l2sw_normalize_if()
        if not ifc:
            ifc = "ethernet 0/0"  # placeholder — caller must ensure real IF first
        fam = (family or "v6").lower()
        v6_acl = "v6-dhcp-block"
        cmds: list[str] = [
            "enable",
            "configure terminal",
            f"interface {ifc}",
            # 포트에 반대 계열/잔여 group 만 정리 (ACL 정의는 삭제하지 않음)
            f"no ip access-group {acl} in",
            f"no ipv6 access-group {v6_acl} in",
        ]
        if fam == "v4":
            cmds.extend(
                [
                    f"ipv6 access-group {v6_acl} in",
                    "exit",
                    "exit",
                ]
            )
        else:
            cmds.extend(
                [
                    f"ip access-group {acl} in",
                    "exit",
                    "exit",
                ]
            )
        return cmds

    def _guardrails_acl_remove_cmds(self) -> list[str]:
        """포트에서 access-group 만 제거 — ACL list(no access-list) 는 유지."""
        acl = self._guardrails_gf("acl_num", "110") or "110"
        ifc = self._guardrails_l2sw_normalize_if()
        if not ifc:
            ifc = "ethernet 0/0"
        v6_acl = "v6-dhcp-block"
        return [
            "enable",
            "configure terminal",
            f"interface {ifc}",
            f"no ip access-group {acl} in",
            f"no ipv6 access-group {v6_acl} in",
            "exit",
            f"interface {ifc}",
            f"no ip access-group {acl} in",
            f"no ipv6 access-group {v6_acl} in",
            "exit",
            "exit",
        ]

    def _guardrails_apply_acl(self, family: str | None = None) -> tuple[str, str]:
        iid = getattr(self, "_guardrails_settings_item_id", None) or "dhcp_boot"
        fam = (family or self._guardrails_item_family(iid)).lower()
        self._guardrails_fill_defaults_from_context(iid)
        # MAC 먼저 → L2SW show mac 으로 IF
        st_mac, det_mac = self._guardrails_ensure_ru_mac(iid, force=True)
        if st_mac == "FAIL":
            return "FAIL", f"RU MAC 조회 실패: {det_mac}"
        st_if, det_if = self._guardrails_ensure_l2sw_if_from_mac(iid, force=True)
        if st_if == "FAIL":
            return "FAIL", f"L2SW IF 자동 조회 실패: {det_if}"
        ifc = self._guardrails_l2sw_normalize_if()
        if not ifc:
            return "FAIL", "L2SW IF 없음 (RU MAC show mac 조회 실패)"
        ok, detail = self._guardrails_l2sw_run_cmds(self._guardrails_acl_apply_cmds(fam), settle_s=0.8)
        if not ok:
            return "FAIL", f"ACL 포트 적용 실패: {detail}"
        if fam == "v4":
            return (
                "PASS",
                f"port {ifc}: ipv6 access-group v6-dhcp-block in "
                f"(ACL 본문 유지, 포트만 적용; MAC={det_mac}; IF={det_if})",
            )
        acl = self._guardrails_gf("acl_num", "110")
        return (
            "PASS",
            f"port {ifc}: ip access-group {acl} in "
            f"(ACL 본문 유지, 포트만 적용; MAC={det_mac}; IF={det_if})",
        )

    def _guardrails_acl_still_on_if(self, show_out: str, acl: str, v6_acl: str) -> list[str]:
        """show running-config interface 출력에서 남은 access-group 목록."""
        text = show_out or ""
        left: list[str] = []
        if re.search(rf"(?im)^\s*ip\s+access-group\s+{re.escape(acl)}\s+in\b", text):
            left.append(f"ip access-group {acl} in")
        if re.search(
            rf"(?im)^\s*ipv6\s+access-group\s+{re.escape(v6_acl)}\s+in\b", text
        ):
            left.append(f"ipv6 access-group {v6_acl} in")
        return left

    def _guardrails_remove_acl(self, timeout: int | None = None) -> tuple[str, str]:
        """포트에서 access-group 만 detach (ACL list 는 스위치에 유지). show 로 확인."""
        self._guardrails_fill_defaults_from_context()
        acl = self._guardrails_gf("acl_num", "110") or "110"
        v6_acl = "v6-dhcp-block"
        ifc = self._guardrails_l2sw_normalize_if()
        if not ifc:
            st_if, det_if = self._guardrails_ensure_l2sw_if_from_mac()
            ifc = self._guardrails_l2sw_normalize_if()
            if not ifc:
                return "FAIL", f"L2SW IF 없음 (ACL 해제 대상 포트 불명): {det_if}"
            self._guardrails_log(f"ACL 해제 전 IF 조회: {st_if} {det_if}")
        to = int(timeout) if timeout is not None else None

        def _detach_port() -> tuple[bool, str]:
            return self._guardrails_l2sw_run_cmds(
                self._guardrails_acl_remove_cmds(), settle_s=1.0, timeout=to
            )

        def _show_if() -> str:
            _ok_s, out_s = self._guardrails_l2sw_run_cmds(
                [
                    "enable",
                    f"show running-config interface {ifc}",
                    "exit",
                ],
                settle_s=0.6,
                timeout=max(60, to or 60),
                require_config=False,
            )
            return out_s or ""

        ok, detail = _detach_port()
        if not ok:
            self._guardrails_log(f"ACL 포트 해제 CLI 실패: {detail[:200]}")
            return "FAIL", f"ACL 포트 해제 실패: {detail}"

        show1 = _show_if()
        left = self._guardrails_acl_still_on_if(show1, acl, v6_acl)
        if left:
            self._guardrails_log(
                f"ACL 포트 잔존 감지 → 재시도: {', '.join(left)} on {ifc}"
            )
            ok2, detail2 = _detach_port()
            show2 = _show_if()
            left2 = self._guardrails_acl_still_on_if(show2, acl, v6_acl)
            if left2:
                self._guardrails_log(
                    f"ACL 포트 원복 FAIL — 여전히 적용 중: {', '.join(left2)} "
                    f"(수동: interface {ifc} → no … access-group … in)"
                )
                return (
                    "FAIL",
                    f"ACL 포트 잔존: {', '.join(left2)} on {ifc} "
                    f"(CLI={'OK' if ok2 else 'FAIL'})",
                )
            self._guardrails_log(f"ACL 포트 재시도 후 해제 확인 OK on {ifc}")
            return "PASS", f"access-group removed from {ifc} (retry; ACL 본문 유지)"

        self._guardrails_log(
            f"ACL 포트 해제 확인 OK — {ifc} 에 access-group 없음 (ACL 본문 유지)"
        )
        return "PASS", f"access-group removed from {ifc} (ACL 본문 유지)"

    def _guardrails_l2sw_trunk_vlan(self, action: str, vid: str | None = None) -> tuple[str, str]:
        """Dasan: switchport trunk allowed vlan remove|add <vid> on configured L2SW IF."""
        act = (action or "").strip().lower()
        if act not in ("remove", "add"):
            return "FAIL", f"vlan action 불가: {action!r} (remove|add)"
        v = (vid if vid is not None else self._guardrails_gf("vlan_discovery_vid", "61")) or "61"
        v = str(v).strip()
        if not re.fullmatch(r"\d{1,4}", v) or not (1 <= int(v) <= 4094):
            return "FAIL", f"VLAN ID 오류: {v!r}"
        ifc = self._guardrails_l2sw_normalize_if()
        cmds = [
            "enable",
            "configure terminal",
            f"interface {ifc}",
            f"switchport trunk allowed vlan {act} {v}",
            "exit",
            "exit",
        ]
        ok, detail = self._guardrails_l2sw_run_cmds(cmds, settle_s=0.8)
        if not ok:
            return "FAIL", f"trunk vlan {act} {v} 실패: {detail}"
        return "PASS", f"trunk allowed vlan {act} {v} on {ifc}"

    def _guardrails_sync_vlan_restore_btn(self) -> None:
        btn = getattr(self, "guardrails_vlan_restore_btn", None)
        if btn is None:
            return
        pending = getattr(self, "_guardrails_vlan_restore_pending", None)
        try:
            if pending:
                vid = pending.get("vid") or "?"
                btn.configure(state="normal", text=f"VLAN untag 원복 (vlan {vid})")
            else:
                btn.configure(state="disabled", text="VLAN untag 원복")
        except tk.TclError:
            pass

    def _guardrails_arm_vlan_restore(self, pending: dict[str, Any]) -> None:
        """Keep VLAN state; enable toolbar restore for later."""
        self._guardrails_vlan_restore_pending = dict(pending or {})
        self._guardrails_sync_vlan_restore_btn()
        self._guardrails_log(
            f"VLAN untag 원복 대기: vlan {pending.get('vid')} ip={pending.get('learned_ip') or '-'} "
            f"({pending.get('item_id')})"
        )

    def _guardrails_disarm_vlan_restore(self) -> None:
        self._guardrails_vlan_restore_pending = None
        self._guardrails_sync_vlan_restore_btn()

    def _guardrails_ask_vlan_untag_restore(
        self, *, vid: str, learned_ip: str, item_id: str
    ) -> bool:
        """Worker-thread safe modal: True=지금 원복, False=유지(추후 버튼)."""
        box: dict[str, bool | None] = {"restore": None}
        done = threading.Event()

        def _show() -> None:
            win = tk.Toplevel(self)
            win.title("VLAN Discovery — untag 원복")
            try:
                win.transient(self)
            except tk.TclError:
                pass
            win.grab_set()
            frm = ttk.Frame(win, padding=14)
            frm.pack(fill="both", expand=True)
            ttk.Label(
                frm,
                text=(
                    f"[{item_id}] VLAN {vid} 에서 IP {learned_ip} 수신 — 시험을 종료합니다.\n\n"
                    f"스위치에서 시험 vlan {vid} 을 삭제하고 untag(base)로 원복할까요?\n\n"
                    f"· 원복 실행: add base → renew → 즉시 remove vlan {vid}\n"
                    "· 현재 유지: renew/원복 안 함 — 나중에 「VLAN untag 원복」 버튼"
                ),
                justify="left",
                wraplength=480,
            ).pack(anchor="w", pady=(0, 12))
            bf = ttk.Frame(frm)
            bf.pack(fill="x")

            def _choose(restore: bool) -> None:
                box["restore"] = restore
                try:
                    win.grab_release()
                except tk.TclError:
                    pass
                win.destroy()
                done.set()

            ttk.Button(bf, text="원복 실행", command=lambda: _choose(True)).pack(
                side="left", padx=(0, 8)
            )
            ttk.Button(bf, text="현재 유지", command=lambda: _choose(False)).pack(side="left")
            win.protocol("WM_DELETE_WINDOW", lambda: _choose(False))
            try:
                win.focus_force()
            except tk.TclError:
                pass

        self.after(0, _show)
        if not done.wait(timeout=3600):
            self._guardrails_log(f"{item_id}: VLAN 원복 확인 타임아웃 — 유지로 처리")
            return False
        return bool(box["restore"])

    def _guardrails_run_vlan_untag_restore(self, pending: dict[str, Any]) -> tuple[str, str]:
        """add base → renew → 즉시 remove 시험VLAN."""
        item_id = str(pending.get("item_id") or "dhcp_v4")
        fam = "v4" if str(pending.get("fam") or "v4").lower().startswith("v4") else "v6"
        vid = str(pending.get("vid") or "").strip()
        learned = str(pending.get("learned_ip") or "").strip()
        probe = str(pending.get("probe") or "").strip()
        renew_cmd = str(pending.get("renew_cmd") or "dhcp renew force").strip() or "dhcp renew force"
        sif = str(pending.get("solid_vlan_if") or "").strip()
        ru_mac = str(pending.get("ru_mac") or "").strip()
        if not ru_mac:
            mac_norm = self._guardrails_normalize_mac(self._guardrails_gf("ru_mac"))
            ru_mac = (
                ":".join(mac_norm[i : i + 2] for i in range(0, 12, 2)) if mac_norm else ""
            )
        prev = getattr(self, "_guardrails_settings_item_id", None)
        self._guardrails_settings_item_id = item_id
        notes: list[str] = []
        try:
            user = self._guardrails_gf("oru_cli_id") or "admin"
            base = str(pending.get("base_vid") or self._guardrails_vlan_discovery_base_vid())
            # 중지 시 learned 비어 있던 경우: neigh/lease 로 tag IP 재확보 (원복 전)
            if (not learned) or self._guardrails_is_pre_discovery_ip(learned, fam):
                try:
                    for ip in self._guardrails_find_ip_by_ru_mac(ru_mac, sif, fam):
                        if ip and not self._guardrails_is_pre_discovery_ip(ip, fam):
                            learned = ip
                            self._guardrails_log(
                                f"{item_id}: 원복 전 Discovery IP 재확보 → {learned}"
                            )
                            break
                except Exception as exc:
                    self._guardrails_log(f"{item_id}: Discovery IP 재확보 실패: {exc}")
            # 1) add base
            st_a, det_a = self._guardrails_l2sw_vlan_discovery_add_base(base)
            notes.append(det_a)
            self._guardrails_log(f"{item_id}: L2 base add → {st_a}: {det_a}")
            time.sleep(1.0)
            # 2) renew — tag → untag probe → neigh(현재 IP). reset 후 IP 변경(::141) 대응
            renewed = False
            if learned and not self._guardrails_is_pre_discovery_ip(learned, fam):
                self._guardrails_log(
                    f"{item_id}: renew `{renew_cmd}` → ssh {user}@{learned} (tag/{fam})"
                )
                ok_r, det_r = self._guardrails_ru_force_dhcp_renew(
                    learned, fam, command=renew_cmd, via_dhcp_host=True, vlan_bind=True
                )
                notes.append(f"dhcp_renew:{'OK' if ok_r else 'WARN'}@{learned}/tag")
                self._guardrails_log(
                    f"{item_id}: dhcp renew → {'OK' if ok_r else 'WARN'}: {det_r[:200]}"
                )
                renewed = bool(ok_r)
            if (not renewed) and probe and probe != learned:
                self._guardrails_log(
                    f"{item_id}: renew `{renew_cmd}` → ssh {user}@{probe} (untag/{fam})"
                    + (" — tag 실패 후 재시도" if learned else "")
                )
                ok_r, det_r = self._guardrails_ru_force_dhcp_renew(
                    probe, fam, command=renew_cmd, via_dhcp_host=True, vlan_bind=False
                )
                notes.append(f"dhcp_renew:{'OK' if ok_r else 'WARN'}@{probe}/untag")
                self._guardrails_log(
                    f"{item_id}: dhcp renew → {'OK' if ok_r else 'WARN'}: {det_r[:200]}"
                )
                renewed = bool(ok_r)
            if (not renewed) and ru_mac:
                try:
                    live = self._guardrails_find_ip_by_ru_mac(ru_mac, sif, fam)
                except Exception:
                    live = []
                seen = {x for x in (learned, probe) if x}
                for ip in live:
                    if not ip or ip in seen:
                        continue
                    is_untag = self._guardrails_is_pre_discovery_ip(ip, fam)
                    label = "untag/neigh" if is_untag else "tag/neigh"
                    self._guardrails_log(
                        f"{item_id}: renew `{renew_cmd}` → ssh {user}@{ip} ({label}) "
                        "— tag/probe 실패 후 neigh 폴백"
                    )
                    ok_r, det_r = self._guardrails_ru_force_dhcp_renew(
                        ip,
                        fam,
                        command=renew_cmd,
                        via_dhcp_host=True,
                        vlan_bind=not is_untag,
                    )
                    notes.append(
                        f"dhcp_renew:{'OK' if ok_r else 'WARN'}@{ip}/{label}"
                    )
                    self._guardrails_log(
                        f"{item_id}: dhcp renew → {'OK' if ok_r else 'WARN'}: "
                        f"{det_r[:200]}"
                    )
                    if ok_r:
                        renewed = True
                        break
            if not renewed:
                for tag_ll in (True, False):
                    ll, how_ll = self._guardrails_ru_linklocal_scoped(tag=tag_ll)
                    if not ll:
                        continue
                    label = "LL/tag" if tag_ll else "LL/untag"
                    self._guardrails_log(
                        f"{item_id}: renew `{renew_cmd}` → ssh {user}@{ll} ({label}) "
                        f"— {how_ll}"
                    )
                    ok_r, det_r = self._guardrails_ru_force_dhcp_renew(
                        ll,
                        fam,
                        command=renew_cmd,
                        via_dhcp_host=True,
                        vlan_bind=tag_ll,
                    )
                    notes.append(f"dhcp_renew:{'OK' if ok_r else 'WARN'}@{ll}/{label}")
                    self._guardrails_log(
                        f"{item_id}: dhcp renew → {'OK' if ok_r else 'WARN'}: "
                        f"{det_r[:200]}"
                    )
                    if ok_r:
                        renewed = True
                        break
            if not renewed and not learned and not probe:
                notes.append("dhcp_renew:SKIP no tag/probe ip")
                self._guardrails_log(
                    f"{item_id}: ⚠ renew 생략 — Discovery IP·probe 없음 "
                    f"(RU는 vlan {vid} IP만 있을 수 있음). L2 원복만 진행"
                )
            elif not renewed:
                notes.append("dhcp_renew:WARN all-targets-failed")
                self._guardrails_log(
                    f"{item_id}: ⚠ renew 전부 실패 — L2 remove 계속 "
                    f"(tag={learned or '-'} probe={probe or '-'})"
                )
            # 3) 즉시 remove 시험 VLAN
            self._guardrails_log(f"{item_id}: renew 직후 즉시 remove vlan {vid}")
            st_r, det_r = self._guardrails_l2sw_vlan_discovery_remove_test(vid)
            notes.append(det_r)
            self._guardrails_log(f"{item_id}: L2 vlan remove → {st_r}: {det_r}")
            # tag 시험용 vlan IF 는 원복 끝에서 항상 삭제 (IPv6 ::252 포함)
            if sif:
                st_s, det_s = self._guardrails_solid_vlan_if_teardown(sif)
                notes.append(f"solid_if:{det_s}")
                self._guardrails_log(f"{item_id}: solid 임시 IF 삭제 → {st_s}: {det_s}")
            st_c = "PASS" if st_a == "PASS" and st_r == "PASS" else "FAIL"
            return st_c, " | ".join(notes)
        finally:
            if prev is not None:
                self._guardrails_settings_item_id = prev

    def _guardrails_btn_vlan_untag_restore(self) -> None:
        pending = getattr(self, "_guardrails_vlan_restore_pending", None)
        if not pending:
            messagebox.showinfo("M-Plane Test", "원복 대기 중인 VLAN이 없습니다.")
            return
        if self.guardrails_busy:
            messagebox.showwarning("M-Plane Test", "검증 중에는 VLAN 원복을 실행할 수 없습니다.")
            return
        vid = pending.get("vid") or "?"
        lip = pending.get("learned_ip") or "-"
        fam_p = pending.get("fam") or "?"
        if not messagebox.askyesno(
            "VLAN untag 원복",
            f"순서: add base → renew(ssh) → remove vlan {vid}\n"
            f"base vlan {pending.get('base_vid') or 1}, family={fam_p}, "
            f"tag IP={lip}\n\n"
            f"tag IP 가 '-' 이면 renew 가 안 나갈 수 있습니다.",
        ):
            return

        snap = dict(pending)

        def _work() -> None:
            try:
                st, detail = self._guardrails_run_vlan_untag_restore(snap)
                self.after(0, self._guardrails_disarm_vlan_restore)
                self.after(
                    0,
                    lambda: messagebox.showinfo(
                        "VLAN untag 원복",
                        f"{st}\n{detail}",
                    ),
                )
            except Exception as exc:
                self.after(
                    0,
                    lambda: messagebox.showerror("VLAN untag 원복", str(exc)),
                )

        threading.Thread(target=_work, daemon=True).start()

    def _guardrails_vlan_discovery_base_vid(self) -> str:
        """Untagged/base VLAN removed during Discovery and restored on untag 원복 (default 1)."""
        b = (self._guardrails_gf("vlan_discovery_base_vid", "1") or "1").strip() or "1"
        if not re.fullmatch(r"\d{1,4}", b) or not (1 <= int(b) <= 4094):
            return "1"
        return b

    def _guardrails_capture_host_vlan_if_names(
        self, vid: str | None = None, primary_if: str | None = None
    ) -> tuple[str, str, str]:
        """Return (parent, ifname, cidr) for temporary IF on dhcp_host (9.249).

        Linux IFNAMSIZ=16 → 이름 최대 15자.
        enx00e04c681c64.100 은 19자라 생성 실패 → vlan<VID> 짧은 이름 사용.
        """
        v = str(vid or self._guardrails_gf("vlan_discovery_vid", "100") or "100").strip()
        parent = (self._guardrails_gf("vlan_discovery_solid_parent") or "").strip()
        if not parent:
            parent = (primary_if or self._guardrails_gf("dhcp_if") or "").strip()
        if not parent:
            try:
                parent = (self.fields.get("LOCAL_IF").get() or "").strip()  # type: ignore[union-attr]
            except Exception:
                parent = ""
        if not parent:
            parent = "enx00e04c681c64"
        m = re.fullmatch(r"(.+)\.(\d+)", parent)
        if m:
            parent = m.group(1)
        cidr = (self._guardrails_gf("vlan_discovery_solid_cidr", "10.0.61.252/24") or "").strip()
        if not cidr:
            cidr = "10.0.61.252/24"
        # "10.0.61.252" 만 넣으면 ip addr add 가 /32 가 됨 → .61 라우트 안 생김 → ping GW로 나감
        if "/" not in cidr:
            cidr = f"{cidr}/24"
        else:
            try:
                import ipaddress

                iface = ipaddress.ip_interface(cidr)
                # /32 는 Discovery 서브넷 라우트가 안 생김 → /24 로 강제
                if iface.network.prefixlen >= 32:
                    cidr = f"{iface.ip}/24"
            except Exception:
                if cidr.endswith("/32"):
                    cidr = cidr.rsplit("/", 1)[0] + "/24"
        # prefer parent.vid when ≤15 chars; else vlan<vid>
        long_name = f"{parent}.{v}"
        if len(long_name) <= 15:
            ifname = long_name
        else:
            ifname = f"vlan{v}"
            if len(ifname) > 15:
                ifname = f"v{v}"[:15]
        return parent, ifname, cidr

    def _guardrails_dhcp_host_sudo_line(self, cmd: str) -> str:
        """One remote line = 수동 `sudo <cmd>` 와 동일 (비밀번호는 dhcp_pw).

        tcpdump 는 NOPASSWD 인 경우가 많아 sudo -n 으로도 되는데,
        `ip addr add` 는 보통 암호 필요 → dhcp_pw 비면 여기서 실패함.
        """
        use_sudo = (self._guardrails_gf("use_sudo", "1") or "1").strip().lower() not in (
            "0", "false", "no", "n",
        )
        if not use_sudo:
            return cmd
        pw = (self._guardrails_gf("dhcp_pw") or "").strip()
        if pw:
            # 수동: sudo ip addr add …  /  자동: printf pw | sudo -S -k -- ip addr add …
            return f"printf '%s\\n' {shlex.quote(pw)} | sudo -S -k -- {cmd}"
        return f"sudo -n -- {cmd}"

    def _guardrails_dhcp_host_sudo_prefix(self) -> str:
        """Deprecated alias — heal 등 구형 호출용."""
        use_sudo = (self._guardrails_gf("use_sudo", "1") or "1").strip().lower() not in (
            "0", "false", "no", "n",
        )
        if not use_sudo:
            return ""
        pw = (self._guardrails_gf("dhcp_pw") or "").strip()
        if pw:
            return f"printf '%s\\n' {shlex.quote(pw)} | sudo -S -k -- "
        return "sudo -n -- "

    def _guardrails_dhcp_vlan_l3_fix_lines(
        self, *, ifname: str, parent: str, cidr: str, net_cidr: str, ip_only: str
    ) -> list[str]:
        """수동으로 성공한 순서와 동일:

        - carrier=0 이어도 ignore_routes_with_linkdown=0
        - /32 주소면 /24 로 교체
        - ip route add 10.0.61.0/24 dev vlan100 scope link src … metric 0
        """
        s = self._guardrails_dhcp_host_sudo_line
        return [
            f"{s(f'ip link set {parent} up')} 2>/dev/null || true",
            f"{s(f'ip link set {ifname} up')} 2>/dev/null || true",
            f"{s(f'sysctl -w net.ipv4.conf.{ifname}.ignore_routes_with_linkdown=0')} 2>/dev/null || true",
            f"{s('sysctl -w net.ipv4.conf.all.ignore_routes_with_linkdown=0')} 2>/dev/null || true",
            # /32 남아 있으면 connected /24 안 생김 → 제거 후 /24 재부여
            f"if ip -4 -o addr show dev {shlex.quote(ifname)} 2>/dev/null | grep -Eq '{ip_only}/32'; then",
            f"  echo CAP_VLAN_FIX_32",
            f"  {s(f'ip addr del {ip_only}/32 dev {ifname}')} 2>/dev/null || true",
            f"  {s(f'ip addr add {cidr} dev {ifname}')} 2>&1 || true",
            "fi",
            f"if ! ip -4 -o addr show dev {shlex.quote(ifname)} 2>/dev/null | grep -Fq {shlex.quote(ip_only)}; then",
            f"  {s(f'ip addr add {cidr} dev {ifname}')} 2>&1 || true",
            "fi",
            f"{s(f'ip route del {net_cidr}')} 2>/dev/null || true",
            f"{s(f'ip route del {net_cidr} via 192.168.1.1')} 2>/dev/null || true",
            f"{s(f'ip route del {net_cidr} via 192.168.9.254')} 2>/dev/null || true",
            # 사용자가 수동으로 성공한 명령과 동일
            f"{s(f'ip route add {net_cidr} dev {ifname} scope link src {ip_only} metric 0')} 2>&1 "
            f"|| {s(f'ip route replace {net_cidr} dev {ifname} src {ip_only} metric 0')} 2>&1 "
            f"|| true",
            "echo CAP_VLAN_ROUTE_DONE",
            f"echo LINK=$(ip -br link show {shlex.quote(ifname)} 2>&1 | tr ' ' '_')",
            f"echo CARRIER=$(cat /sys/class/net/{ifname}/carrier 2>/dev/null || echo NA)",
            f"echo OPER=$(cat /sys/class/net/{ifname}/operstate 2>/dev/null || echo NA)",
            f"echo ADDR=$(ip -4 -o addr show dev {shlex.quote(ifname)} 2>&1 | tr ' ' '_' | head -c 160)",
            f"echo ROUTES=$(ip -4 route show {shlex.quote(net_cidr)} 2>&1 | tr '\\n' ';' | head -c 200)",
            f"echo ROUTES_DEV=$(ip -4 route show dev {shlex.quote(ifname)} 2>&1 | tr '\\n' ';' | head -c 200)",
        ]

    def _guardrails_solid_vlan_if_prepare(
        self, vid: str | None = None, primary_if: str | None = None
    ) -> tuple[str, str, bool, str]:
        """Create temporary vlan IF on dhcp_host for ping/SSH (v4 + v6).

        수동과 동일:
          ip link add vlan100 … / ip addr add 10.0.61.252/24
          / ip -6 addr add 2001:…::252/64  (dhcp_v6 / ⚙ vlan_discovery_solid_cidr_v6)
        """
        parent, ifname, cidr = self._guardrails_capture_host_vlan_if_names(vid, primary_if)
        # parent enx 에 .61 이 남아 있으면 먼저 제거 (이전 renew/ping 버그 잔존)
        self._guardrails_dhcp_host_strip_tag_addr_from_parent()
        v = str(vid or self._guardrails_gf("vlan_discovery_vid", "100") or "100").strip()
        if not re.fullmatch(r"\d{1,4}", v):
            return "FAIL", f"VLAN ID 오류: {v!r}", False, ifname
        if not re.fullmatch(r"[\w.-]+", parent) or not re.fullmatch(r"[\w.-]+", ifname):
            return "FAIL", f"IF name 오류: {parent!r}/{ifname!r}", False, ifname
        ip_only = cidr.split("/")[0].strip()
        try:
            import ipaddress

            net_cidr = str(ipaddress.ip_interface(cidr).network)
        except Exception:
            net_cidr = "10.0.61.0/24"
        cidr6 = self._guardrails_v6_solid_cidr(plane="tag")
        fam_now = "v6"
        try:
            fam_now = self._guardrails_resolve_ssh_family(
                getattr(self, "_guardrails_settings_item_id", None)
            )
        except Exception:
            pass
        pw = (self._guardrails_gf("dhcp_pw") or "").strip()
        if not pw:
            self._guardrails_log(
                "dhcp_host: ⚠ dhcp_pw 비어 있음 — sudo -n 으로 ip addr add 시도 "
                "(수동 sudo 는 되고 GUI 만 실패하는 대표 원인)"
            )
        s = self._guardrails_dhcp_host_sudo_line
        v6_lines: list[str] = [
            s(f"sysctl -w net.ipv6.conf.{ifname}.disable_ipv6=0") + " 2>/dev/null || true",
            s("sysctl -w net.ipv6.conf.all.disable_ipv6=0") + " 2>/dev/null || true",
        ]
        if cidr6:
            lip6 = cidr6.split("/")[0].strip()
            try:
                import ipaddress

                net6 = str(ipaddress.IPv6Interface(cidr6 if "/" in cidr6 else f"{cidr6}/64").network)
            except Exception:
                net6 = ""
            v6_lines.extend(
                [
                    f"echo CAP_VLAN6_CIDR={shlex.quote(cidr6)}",
                    f"{s(f'ip -6 addr add {cidr6} dev {ifname} nodad')} 2>&1 || "
                    f"{s(f'ip -6 addr add {cidr6} dev {ifname}')} 2>&1 || true",
                    (
                        f"{s(f'ip -6 route replace {net6} dev {ifname} metric 0')} 2>&1 || true"
                        if net6
                        else "true"
                    ),
                    f"echo ADDR6=$(ip -6 -o addr show dev {shlex.quote(ifname)} 2>&1 | tr ' ' '_' | head -c 200)",
                    f"if ip -6 -o addr show dev {shlex.quote(ifname)} 2>/dev/null "
                    f"| grep -Fq {shlex.quote(lip6)}; then echo CAP_VLAN6_ADDR_OK; "
                    "else echo CAP_VLAN6_ADDR_FAIL; fi",
                ]
            )
        else:
            v6_lines.append("echo CAP_VLAN6_SKIP")
        script = "\n".join(
            [
                "set +e",
                f"echo CAP_VLAN_BEGIN parent={parent} if={ifname} cidr={cidr} "
                f"cidr6={cidr6 or '-'} pw={'set' if pw else 'EMPTY'}",
                s("modprobe 8021q") + " >/dev/null 2>&1 || true",
                s(f"ip link set {parent} up") + " 2>&1 | tail -n 2 || true",
                f"if ip -o link show {ifname} >/dev/null 2>&1; then",
                "  echo CAP_VLAN_LINK_EXISTS",
                "else",
                "  "
                + s(
                    f"ip link add name {ifname} link {parent} "
                    f"type vlan protocol 802.1Q id {v}"
                )
                + " 2>&1",
                "  echo CAP_VLAN_LINK_RC=$?",
                "  echo CAP_VLAN_LINK_NEW",
                "fi",
                s(f"ip link set {ifname} up") + " 2>&1 | tail -n 2 || true",
                # 주소만 빠지는 케이스: 기존 IF reuse 후 addr 없음 → 수동 add 와 동일
                s(f"ip addr flush dev {ifname}") + " 2>/dev/null || true",
                s(f"ip addr add {cidr} dev {ifname}") + " 2>&1",
                "echo CAP_VLAN_ADDR_RC=$?",
                *self._guardrails_dhcp_vlan_l3_fix_lines(
                    ifname=ifname,
                    parent=parent,
                    cidr=cidr,
                    net_cidr=net_cidr,
                    ip_only=ip_only,
                ),
                f"rg0=$(ip -4 route get {ip_only} 2>&1 || true)",
                'echo "ROUTE_SELF=$rg0"',
                f"if ip -4 -o addr show dev {ifname} 2>/dev/null | grep -Fq {shlex.quote(ip_only)}; then",
                "  echo CAP_VLAN_ADDR_OK",
                "else",
                "  echo CAP_VLAN_ADDR_FAIL",
                "fi",
                *v6_lines,
            ]
        )
        _ok, text = self._guardrails_dhcp_server_run(script, timeout=40)
        out = text or ""
        owned = "CAP_VLAN_LINK_NEW" in out
        if "CAP_VLAN_ADDR_OK" not in out:
            return (
                "FAIL",
                f"dhcp_host `sudo ip addr add {cidr} dev {ifname}` 실패 "
                f"(dhcp_pw={'있음' if pw else '없음/EMPTY'}). 출력: {out[:300]}",
                owned,
                ifname,
            )
        how = "created" if owned else "reuse"
        detail = f"{how} {ifname} {cidr} on dhcp_host"
        if cidr6:
            if "CAP_VLAN6_ADDR_OK" in out:
                detail += f" + {cidr6}"
            else:
                detail += f" +v6FAIL({cidr6})"
                if fam_now == "v6":
                    return (
                        "FAIL",
                        f"dhcp_host vlan IF IPv6 부여 실패 ({cidr6}). "
                        f"출력: {out[:280]}",
                        owned,
                        ifname,
                    )
                self._guardrails_log(
                    f"dhcp_host: vlan IF IPv6 부여 실패(비치명) {cidr6} — {out[:120]}"
                )
        return "PASS", detail, owned, ifname

    def _guardrails_solid_vlan_if_teardown(self, ifname: str | None) -> tuple[str, str]:
        """Delete temporary vlan IF on dhcp_host (only when owned by this test)."""
        self._guardrails_dhcp_host_strip_tag_addr_from_parent()
        name = (ifname or "").strip()
        if not name or not re.fullmatch(r"[\w.-]+", name):
            return "SKIP", "no ifname"
        s = self._guardrails_dhcp_host_sudo_line
        cmd = "\n".join(
            [
                "set +e",
                f"if ip link show {name} >/dev/null 2>&1; then",
                f"  {s(f'ip link delete {name}')} 2>&1",
                "  echo CAP_VLAN_DELETED",
                "else",
                "  echo CAP_VLAN_GONE",
                "fi",
            ]
        )
        _ok, text = self._guardrails_dhcp_server_run(cmd, timeout=20)
        out = text or ""
        if "CAP_VLAN_DELETED" in out or "CAP_VLAN_GONE" in out:
            return "PASS", f"deleted {name} on dhcp_host"
        return "FAIL", f"dhcp_host {name} 삭제 실패: {out[:200]}"

    def _guardrails_v4_local_cidr_for_target(self, target: str) -> str:
        """Unused — untag 는 기존 IF IP(.99 등)만 사용, .252 를 만들지 않음."""
        return ""

    def _guardrails_dhcp_host_existing_untag_v4(
        self, ifname: str, peer: str
    ) -> tuple[str, str]:
        """parent enx 에 이미 있는 untag IPv4 (같은 /24). .252 를 만들지 않음.

        예: 10.0.60.99/24 유지. 여러 개면 .252/.253 보다 기존 주소 우선.
        Returns (ip, cidr) or ("", "").
        """
        name = (ifname or "").strip()
        peer_ip = self._guardrails_strip_ip_cidr(peer)
        if not name or not peer_ip or ":" in peer_ip:
            return "", ""
        try:
            import ipaddress

            peer_net = ipaddress.IPv4Interface(f"{peer_ip}/24").network
        except Exception:
            return "", ""
        _ok, text = self._guardrails_dhcp_server_run(
            f"ip -4 -o addr show dev {shlex.quote(name)} 2>/dev/null || true",
            timeout=12,
        )
        found: list[tuple[str, str]] = []
        for m in re.finditer(
            r"\binet\s+(\d{1,3}(?:\.\d{1,3}){3})(?:/(\d{1,2}))?", text or ""
        ):
            ip = m.group(1)
            pfx = m.group(2) or "24"
            try:
                import ipaddress

                if ipaddress.IPv4Address(ip) not in peer_net:
                    continue
            except Exception:
                continue
            if ip == peer_ip:
                continue
            found.append((ip, f"{ip}/{pfx}"))
        if not found:
            return "", ""
        preferred = [x for x in found if not x[0].endswith((".252", ".253"))]
        pick = (preferred or found)[0]
        return pick[0], pick[1]

    def _guardrails_dhcp_host_strip_invented_untag_252(self, parent: str) -> None:
        """예전에 ping heal 이 올린 임시 10.0.60.252 만 제거.

        Settings LOCAL_IP (CallHome listen, 예: .252/.253) 는 절대 삭제하지 않음.
        """
        name = (parent or "").strip()
        if not name:
            return
        try:
            local_ip = self._guardrails_strip_ip_cidr(
                (self.fields.get("LOCAL_IP").get() or "").strip()  # type: ignore[union-attr]
            )
        except Exception:
            local_ip = ""
        # LOCAL_IP 가 .252 이면 heal 정리 대상이 아님
        if local_ip == "10.0.60.252":
            return
        s = self._guardrails_dhcp_host_sudo_line
        script = "\n".join(
            [
                "set +e",
                f"has_ops=$(ip -4 -o addr show dev {shlex.quote(name)} 2>/dev/null "
                f"| grep -E 'inet 10\\.0\\.60\\.[0-9]+' | grep -vE '\\.252/' || true)",
                f"has252=$(ip -4 -o addr show dev {shlex.quote(name)} 2>/dev/null "
                f"| grep -E 'inet 10\\.0\\.60\\.252/' || true)",
                'if [ -n "$has_ops" ] && [ -n "$has252" ]; then',
                f"  {s('ip addr del 10.0.60.252/24 dev ' + name)} 2>/dev/null || true",
                "  echo CAP_STRIP_UNTAG_252",
                "fi",
            ]
        )
        try:
            _ok, text = self._guardrails_dhcp_server_run(script, timeout=12)
            if text and "CAP_STRIP_UNTAG_252" in text:
                self._guardrails_log(
                    f"dhcp_host: parent {name} 에서 heal용 10.0.60.252 제거 "
                    f"(LOCAL_IP={local_ip or '-'} 유지)"
                )
        except Exception:
            pass

    def _guardrails_verify_local_ip_on_remote(self, client: Any) -> tuple[bool, str]:
        """CallHome LOCAL_IP 가 헬퍼 실행 호스트에 있는지 확인."""
        try:
            local_ip = self._guardrails_strip_ip_cidr(
                (self.fields.get("LOCAL_IP").get() or "").strip()  # type: ignore[union-attr]
            )
        except Exception:
            local_ip = ""
        if not local_ip or ":" in local_ip:
            return False, "Settings LOCAL_IP 비어 있음"
        cmd = (
            f"ip -4 -o addr show 2>/dev/null | grep -F {shlex.quote(local_ip + '/')} "
            f"|| ip -4 -o addr show 2>/dev/null | grep -F {shlex.quote('inet ' + local_ip)} "
            f"|| true"
        )
        try:
            _stdin, stdout, _stderr = client.exec_command(cmd)
            out = (stdout.read() or b"").decode(errors="ignore").strip()
        except Exception as exc:
            return False, f"LOCAL_IP 검사 실패: {exc}"
        if local_ip in out:
            return True, f"LOCAL_IP {local_ip} present: {out.splitlines()[0][:120]}"
        return (
            False,
            f"LOCAL_IP {local_ip} 가 헬퍼 호스트에 없음 — "
            f"listen 불가 → LOGIN=NOK. solid에 {local_ip} (untag eno2 / tag vlan IF) 확인",
        )

    def _guardrails_dhcp_host_strip_tag_addr_from_parent(self) -> None:
        """Safety: .61(tag) must not sit on parent enx — only on vlan<VID>."""
        try:
            parent, vlan_if, tag_cidr = self._guardrails_capture_host_vlan_if_names()
        except Exception:
            return
        tag_ip = (tag_cidr.split("/")[0] or "").strip()
        if not parent or not tag_ip or parent == vlan_if:
            return
        self._guardrails_dhcp_host_strip_invented_untag_252(parent)
        s = self._guardrails_dhcp_host_sudo_line
        script = "\n".join(
            [
                "set +e",
                f"if ip -4 -o addr show dev {shlex.quote(parent)} 2>/dev/null "
                f"| grep -Fq {shlex.quote(tag_ip)}; then",
                f"  {s(f'ip addr del {tag_cidr} dev {parent}')} 2>/dev/null || "
                f"  {s(f'ip addr del {tag_ip}/24 dev {parent}')} 2>/dev/null || "
                f"  {s(f'ip addr del {tag_ip} dev {parent}')} 2>/dev/null || true",
                f"  echo CAP_STRIP_TAG_FROM_PARENT {parent} {tag_ip}",
                "fi",
            ]
        )
        try:
            _ok, text = self._guardrails_dhcp_server_run(script, timeout=15)
            if text and "CAP_STRIP_TAG_FROM_PARENT" in text:
                self._guardrails_log(
                    f"dhcp_host: parent {parent} 에서 tag IP {tag_ip} 제거 "
                    f"(tag는 {vlan_if} 전용)"
                )
        except Exception:
            pass

    def _guardrails_probe_from_dhcp_host(
        self,
        ip: str,
        via_if: str | None = None,
        family: str = "v4",
    ) -> tuple[bool, str]:
        """Ping RU from dhcp_host (v4=ping / v6=ping6).

        untag(.60/1200): parent enx 만 — .61/1300 주소를 parent 에 올리지 않음.
        tag(.61/1300): vlan<VID> 만.
        """
        host = self._guardrails_strip_ip_cidr(ip)
        if not host:
            return False, "ping: no ip"
        fam = "v6" if (
            str(family).lower().startswith("v6") or ":" in host
        ) else "v4"
        parent, vlan_if, tag_cidr = self._guardrails_capture_host_vlan_if_names()
        via = (via_if or "").strip()
        s = self._guardrails_dhcp_host_sudo_line
        # 잘못 parent 에 남은 .61 정리
        self._guardrails_dhcp_host_strip_tag_addr_from_parent()

        if fam == "v6":
            is_untag = self._guardrails_is_pre_discovery_ip(host, "v6")
            if is_untag:
                ifname = parent if (not via or via == vlan_if) else via
                if ifname == vlan_if:
                    ifname = parent
            else:
                ifname = via if (via and via != parent) else vlan_if
                if ifname == parent:
                    ifname = vlan_if
            local6 = self._guardrails_v6_local_cidr_for_target(host)
            if not local6:
                return False, f"ping6: local cidr 계산 실패 for {host}"
            lip = local6.split("/")[0]
            if host == lip:
                return False, f"ping6: target is local {ifname} ({host})"
            ok6, det6 = self._guardrails_dhcp_host_ensure_v6_addr(ifname, local6)[:2]
            if not ok6:
                return False, f"ping6: IF IPv6 부여 실패 ({det6})"
            if self._guardrails_is_pre_discovery_ip(host, "v6"):
                self._guardrails_arm_parent_v6_cleanup(ifname, local6)
            heal = "\n".join(
                [
                    "set +e",
                    f"{s(f'ip link set {parent} up')} 2>/dev/null || true",
                    f"rg=$(ip -6 route get {shlex.quote(host)} 2>&1)",
                    'echo "ROUTE_GET6=$rg"',
                    f'if ! echo "$rg" | grep -Eq {shlex.quote(f"dev {ifname}")}; then',
                    "  echo PING_FAIL bad_route6_dev",
                    "  exit 0",
                    "fi",
                    "echo ROUTE6_OK",
                    "okn=0",
                    "for i in 1 2; do",
                    f"  po=$(ping -6 -c 1 -W 2 -I {shlex.quote(lip)} "
                    f"{shlex.quote(host)} 2>&1) || true",
                    '  echo "PING6_OUT_$i: $(echo "$po" | tr \'\\n\' \' \' | head -c 220)"',
                    "  echo \"$po\" | grep -Eqi 'bytes from' || continue",
                    "  echo \"$po\" | grep -Eqi "
                    "'Destination.*Unreachable|100% packet loss' && continue",
                    "  echo \"$po\" | grep -Eqi "
                    "'1 received|1 packets received|0% packet loss' || continue",
                    "  okn=$((okn+1))",
                    "done",
                    'if [ "$okn" -ge 2 ]; then echo PING_OK okn=$okn; else echo PING_FAIL okn=$okn; fi',
                ]
            )
            _ok, text = self._guardrails_dhcp_server_run(heal, timeout=30)
            out = text or ""
            last_ok = out.rfind("PING_OK")
            last_fail = out.rfind("PING_FAIL")
            if last_ok >= 0 and last_ok > last_fail:
                return True, f"ping6-ok route:{ifname}/{lip}→{host}"
            return False, f"ping6-fail via {ifname}→{host}: {out.replace(chr(10), ' | ')[:240]}"

        # ---- IPv4 ----
        is_untag = self._guardrails_is_pre_discovery_ip(host, "v4")
        if is_untag:
            # parent enx + 기존 IP 유지(예: 10.0.60.99). .252 추가/변경 금지
            ifname = parent if (not via or via == vlan_if) else via
            if ifname == vlan_if:
                ifname = parent
            self._guardrails_dhcp_host_strip_invented_untag_252(ifname)
            ip_only, cidr = self._guardrails_dhcp_host_existing_untag_v4(ifname, host)
            if not ip_only:
                return (
                    False,
                    f"ping: {ifname} 에 untag IPv4(같은 /24) 없음 — "
                    f"기존 IP(예: 10.0.60.99/24)를 유지하세요 (.252 자동추가 안 함)",
                )
            try:
                import ipaddress

                net_cidr = str(ipaddress.ip_interface(cidr).network)
            except Exception:
                net_cidr = "10.0.60.0/24"
            if host == ip_only:
                return False, f"ping: target is local {ifname} addr ({host})"
            # 주소 추가/flush 없음 — link up + ping 만
            heal = "\n".join(
                [
                    "set +e",
                    f"{s(f'ip link set {ifname} up')} 2>/dev/null || true",
                    f"echo CAP_UNTAG_USE existing={ip_only}",
                    f"rg=$(ip -4 route get {shlex.quote(host)} 2>&1)",
                    'echo "ROUTE_GET=$rg"',
                    'if echo "$rg" | grep -Eq "\\bvia\\b"; then',
                    "  echo PING_FAIL bad_route_via_gw",
                    f"  ip -4 route | head -n 20 | tr '\\n' ';' || true",
                    "  exit 0",
                    "fi",
                    f'if ! echo "$rg" | grep -Eq {shlex.quote(f"dev {ifname}")}; then',
                    "  echo PING_FAIL bad_route_dev",
                    "  exit 0",
                    "fi",
                    "echo ROUTE_OK",
                    "okn=0",
                    "for i in 1 2; do",
                    f"  po=$(ping -c 1 -W 2 -I {shlex.quote(ip_only)} "
                    f"{shlex.quote(host)} 2>&1) || true",
                    '  echo "PING_OUT_$i: $(echo "$po" | tr \'\\n\' \' \' | head -c 220)"',
                    f"  echo \"$po\" | grep -Eqi {shlex.quote(f'bytes from {host}')} || continue",
                    "  echo \"$po\" | grep -Eqi "
                    "'Destination Host Unreachable|Network Unreachable|100% packet loss' "
                    "&& continue",
                    "  echo \"$po\" | grep -Eqi "
                    "'1 received|1 packets received|0% packet loss' || continue",
                    "  okn=$((okn+1))",
                    "done",
                    'if [ "$okn" -ge 2 ]; then echo PING_OK okn=$okn; '
                    "else echo PING_FAIL okn=$okn; fi",
                ]
            )
            _ok, text = self._guardrails_dhcp_server_run(heal, timeout=25)
            out = text or ""
            last_ok = out.rfind("PING_OK")
            last_fail = out.rfind("PING_FAIL")
            if last_ok >= 0 and last_ok > last_fail:
                return True, f"ping-ok route:{ifname}/{ip_only}→{host}"
            return False, f"ping-fail via {ifname}/{ip_only}→{host}: {out.replace(chr(10), ' | ')[:240]}"

        # tag: vlan IF only — .61.252 부여/정리 허용
        ifname = via if (via and via != parent) else vlan_if
        if ifname == parent:
            ifname = vlan_if
        cidr = tag_cidr
        ip_only = cidr.split("/")[0].strip()
        try:
            import ipaddress

            net_cidr = str(ipaddress.ip_interface(cidr).network)
        except Exception:
            net_cidr = "10.0.61.0/24"
        if host == ip_only:
            return False, f"ping: target is local {ifname} addr ({host})"
        heal = "\n".join(
            [
                "set +e",
                f"if ! ip -4 -o addr show dev {shlex.quote(ifname)} 2>/dev/null "
                f"| grep -Fq {shlex.quote(ip_only)}; then",
                f"  {s(f'ip addr flush dev {ifname}')} 2>/dev/null || true",
                f"  {s(f'ip addr add {cidr} dev {ifname}')} 2>&1 || true",
                "  echo CAP_VLAN_HEAL",
                "fi",
                *self._guardrails_dhcp_vlan_l3_fix_lines(
                    ifname=ifname,
                    parent=parent,
                    cidr=cidr,
                    net_cidr=net_cidr,
                    ip_only=ip_only,
                ),
                f"rg=$(ip -4 route get {shlex.quote(host)} 2>&1)",
                'echo "ROUTE_GET=$rg"',
                'if echo "$rg" | grep -Eq "\\bvia\\b"; then',
                "  echo PING_FAIL bad_route_via_gw",
                f"  ip -4 route | head -n 20 | tr '\\n' ';' || true",
                "  exit 0",
                "fi",
                f'if ! echo "$rg" | grep -Eq {shlex.quote(f"dev {ifname}")}; then',
                "  echo PING_FAIL bad_route_dev",
                "  exit 0",
                "fi",
                "echo ROUTE_OK",
                "okn=0",
                "for i in 1 2; do",
                f"  po=$(ping -c 1 -W 2 {shlex.quote(host)} 2>&1) || true",
                '  echo "PING_OUT_$i: $(echo "$po" | tr \'\\n\' \' \' | head -c 220)"',
                f"  echo \"$po\" | grep -Eqi {shlex.quote(f'bytes from {host}')} || continue",
                "  echo \"$po\" | grep -Eqi "
                "'Destination Host Unreachable|Network Unreachable|100% packet loss' "
                "&& continue",
                "  echo \"$po\" | grep -Eqi '1 received|1 packets received|0% packet loss' || continue",
                "  okn=$((okn+1))",
                "done",
                'if [ "$okn" -ge 2 ]; then echo PING_OK okn=$okn; else echo PING_FAIL okn=$okn; fi',
            ]
        )
        _ok, text = self._guardrails_dhcp_server_run(heal, timeout=25)
        out = text or ""
        last_ok = out.rfind("PING_OK")
        last_fail = out.rfind("PING_FAIL")
        if last_ok >= 0 and last_ok > last_fail:
            return True, f"ping-ok route:{ifname}/{ip_only}→{host}"
        return False, f"ping-fail via {ifname}→{host}: {out.replace(chr(10), ' | ')[:240]}"

    def _guardrails_l2sw_vlan_discovery_prepare(
        self, vid: str | None = None, name: str | None = None
    ) -> tuple[str, str]:
        """Early step: create test VLAN + trunk add only (base remove is post-reset)."""
        v = (vid if vid is not None else self._guardrails_gf("vlan_discovery_vid", "61")) or "61"
        v = str(v).strip()
        if not re.fullmatch(r"\d{1,4}", v) or not (1 <= int(v) <= 4094):
            return "FAIL", f"VLAN ID 오류: {v!r}"
        base = self._guardrails_vlan_discovery_base_vid()
        if base == v:
            return "FAIL", f"시험 VLAN과 base VLAN이 같음: {v}"
        vname = (
            name
            if name is not None
            else (self._guardrails_gf("vlan_discovery_name") or "").strip()
        )
        ifc = self._guardrails_l2sw_normalize_if()
        # Dasan: vlan N 후 exit 금지. base(vlan1) remove는 reset 직후 별도 호출.
        cmds: list[str] = [
            "enable",
            "configure terminal",
            f"vlan {v}",
        ]
        if vname:
            cmds.append(f"name {vname}")
        cmds.extend(
            [
                f"interface {ifc}",
                f"switchport trunk allowed vlan add {v}",
                "exit",
                "exit",
            ]
        )
        ok, detail = self._guardrails_l2sw_run_cmds(cmds, settle_s=0.8)
        if not ok:
            return "FAIL", f"vlan {v} prepare 실패: {detail}"
        nm = f" name={vname}" if vname else ""
        return "PASS", f"trunk add {v}{nm} on {ifc}"

    def _guardrails_l2sw_vlan_discovery_remove_base(
        self, base_vid: str | None = None
    ) -> tuple[str, str]:
        """Right after ORU reset: remove base(vlan1) from trunk so Discovery is forced."""
        base = (
            str(base_vid).strip()
            if base_vid is not None
            else self._guardrails_vlan_discovery_base_vid()
        )
        if not re.fullmatch(r"\d{1,4}", base) or not (1 <= int(base) <= 4094):
            return "FAIL", f"base VLAN ID 오류: {base!r}"
        ifc = self._guardrails_l2sw_normalize_if()
        cmds = [
            "enable",
            "configure terminal",
            f"interface {ifc}",
            f"switchport trunk allowed vlan remove {base}",
            "exit",
            "exit",
        ]
        ok, detail = self._guardrails_l2sw_run_cmds(cmds, settle_s=0.8)
        if not ok:
            return "FAIL", f"vlan {base} remove 실패: {detail}"
        return "PASS", f"trunk remove base {base} on {ifc}"

    def _guardrails_l2sw_vlan_discovery_add_base(
        self, base_vid: str | None = None
    ) -> tuple[str, str]:
        """원복 1단계: trunk에 base(vlan1) 만 add (시험 VLAN은 아직 유지)."""
        base = (
            str(base_vid).strip()
            if base_vid is not None
            else self._guardrails_vlan_discovery_base_vid()
        )
        if not re.fullmatch(r"\d{1,4}", base) or not (1 <= int(base) <= 4094):
            base = "1"
        ifc = self._guardrails_l2sw_normalize_if()
        cmds = [
            "enable",
            "configure terminal",
            f"interface {ifc}",
            f"switchport trunk allowed vlan add {base}",
            "exit",
            "exit",
        ]
        ok, detail = self._guardrails_l2sw_run_cmds(cmds, settle_s=0.5)
        if not ok:
            return "FAIL", f"base {base} add 실패: {detail}"
        return "PASS", f"trunk add base {base} on {ifc}"

    def _guardrails_l2sw_vlan_discovery_remove_test(
        self, vid: str | None = None
    ) -> tuple[str, str]:
        """원복 2단계: 시험 VLAN trunk remove + no vlan (renew 직후 즉시)."""
        v = (vid if vid is not None else self._guardrails_gf("vlan_discovery_vid", "100")) or "100"
        v = str(v).strip()
        if not re.fullmatch(r"\d{1,4}", v) or not (1 <= int(v) <= 4094):
            return "FAIL", f"VLAN ID 오류: {v!r}"
        ifc = self._guardrails_l2sw_normalize_if()
        cmds = [
            "enable",
            "configure terminal",
            f"interface {ifc}",
            f"switchport trunk allowed vlan remove {v}",
            "exit",
            f"no vlan {v}",
            "exit",
        ]
        ok, detail = self._guardrails_l2sw_run_cmds(cmds, settle_s=0.4)
        if not ok:
            return "FAIL", f"vlan {v} remove 실패: {detail}"
        return "PASS", f"vlan {v} 삭제(trunk remove+no vlan) on {ifc}"

    def _guardrails_l2sw_vlan_discovery_cleanup(
        self, vid: str | None = None, base_vid: str | None = None
    ) -> tuple[str, str]:
        """원복 일괄: base add → 시험 VLAN remove (한 세션)."""
        st1, d1 = self._guardrails_l2sw_vlan_discovery_add_base(base_vid)
        if st1 != "PASS":
            return st1, d1
        st2, d2 = self._guardrails_l2sw_vlan_discovery_remove_test(vid)
        if st2 != "PASS":
            return st2, f"{d1} | {d2}"
        return "PASS", f"{d1} → {d2}"

    def _guardrails_run_dhcp_family(self, fam: str, item_id: str | None = None) -> tuple[str, str]:
        """ACL → reboot → SSH check address recovery → cleanup."""
        fam = "v4" if str(fam).lower().startswith("v4") else "v6"
        iid = item_id or ("dhcp_v4_only_boot" if fam == "v4" else "dhcp_v6_only_boot")
        self._guardrails_settings_item_id = iid
        self._guardrails_fill_defaults_from_context(iid)
        pass_sec = self._guardrails_int("pass_sec", 240)
        timeout_sec = self._guardrails_int("timeout_sec", 540)
        poll_sec = self._guardrails_int("poll_sec", 5)
        stable_sec = self._guardrails_int("stable_sec", 10)
        down_detect_sec = self._guardrails_int("down_detect_sec", 180)

        host, how = self._guardrails_ru_ssh_target(fam)
        if not host:
            need = "Settings ALLOWED_IP" if fam == "v4" else "Settings ALLOWED_IP_V6 또는 RU MAC+M-Plane NIC"
            return "FAIL", f"SSH 대상 불가: {how} ({need})"
        if not self._guardrails_gf("oru_cli_id"):
            return "FAIL", "RU SSH ID 필요 (Settings ★ RU SSH ID)"

        notes: list[str] = [f"family={fam}", f"ssh-target {how}"]
        st, detail = self._guardrails_apply_acl(fam)
        if st != "PASS":
            return st, detail
        notes.append(detail)
        expect = "IPv4 inet" if fam == "v4" else "global inet6"
        self.after(
            0,
            self.append_log,
            f"[M-Plane Test] ACL 적용({fam}). RU 재부팅 후 SSH로 {expect} 복구를 감시합니다.\n",
        )

        t_acl = time.monotonic()
        saw_down = False
        while time.monotonic() - t_acl < down_detect_sec:
            if self._guardrails_cancel.is_set():
                self._guardrails_remove_acl()
                return "INFO", "사용자 중지 (ACL 원복 시도)"
            up, pdetail = self._guardrails_probe_once(fam)
            if not up:
                saw_down = True
                notes.append(f"unhealthy/down: {pdetail}")
                break
            time.sleep(poll_sec)

        if not saw_down:
            self._guardrails_remove_acl()
            return (
                "FAIL",
                "재부팅(SSH unhealthy) 미감지. RU 재부팅·설정 확인. ACL 원복. "
                + " | ".join(notes),
            )

        t0 = time.monotonic()
        recovered = False
        last = ""
        healthy_since: float | None = None
        self._guardrails_log(
            f"{iid}: 복구 PASS 조건 = SSH healthy {stable_sec}s 연속 유지"
        )
        while time.monotonic() - t0 < timeout_sec:
            if self._guardrails_cancel.is_set():
                self._guardrails_remove_acl()
                return "INFO", f"사용자 중지 (down 이후 {time.monotonic() - t0:.0f}s, ACL 원복)"
            up, last = self._guardrails_probe_once(fam)
            if up:
                now_h = time.monotonic()
                if healthy_since is None:
                    healthy_since = now_h
                    self._guardrails_log(
                        f"{iid}: 주소 복구 1차 OK — {stable_sec}s 유지 확인 중… {last[:100]}"
                    )
                held = now_h - healthy_since
                if held >= float(stable_sec):
                    recovered = True
                    self._guardrails_log(
                        f"{iid}: 주소 복구 OK (유지 {held:.0f}s≥{stable_sec}s) {last[:120]}"
                    )
                    notes.append(f"stable:{stable_sec}s")
                    break
            else:
                if healthy_since is not None:
                    self._guardrails_log(
                        f"{iid}: 복구 유지 실패 — 재대기 last={last[:80]}"
                    )
                healthy_since = None
            time.sleep(min(poll_sec, 3) if healthy_since is not None else poll_sec)

        elapsed = time.monotonic() - t0
        rm_st, rm_detail = self._guardrails_remove_acl()
        notes.append(rm_detail if rm_st == "PASS" else f"원복주의:{rm_detail}")

        if not recovered:
            return "FAIL", f"복구 타임아웃 {elapsed:.0f}s (한도 {timeout_sec}s) last={last} | " + " | ".join(notes)

        verdict = "PASS" if elapsed <= pass_sec else "FAIL"
        why = f"[{fam}] 복구 {elapsed:.0f}s (PASS≤{pass_sec}s, stable≥{stable_sec}s) {last}"
        if verdict == "FAIL":
            why += " — renew(~300s) 대기 의심"
        return verdict, why + " | " + " | ".join(notes)

    def _guardrails_run_dhcp_v6_only(self) -> tuple[str, str]:
        return self._guardrails_run_dhcp_family("v6", item_id="dhcp_v6_only_boot")

    def _guardrails_run_dhcp_v4_only(self) -> tuple[str, str]:
        return self._guardrails_run_dhcp_family("v4", item_id="dhcp_v4_only_boot")

    def _guardrails_open_settings(self, item_id: str) -> None:
        schema = self._GUARDRAILS_PER_TEST_SCHEMA.get(item_id)
        if not schema:
            messagebox.showinfo("M-Plane Test", f"{item_id}: 설정 스키마 없음")
            return
        self._guardrails_fill_defaults_from_context(item_id)
        win = tk.Toplevel(self)
        win.title(f"설정 — {schema['title']}")
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        fr = ttk.Frame(win, padding=12)
        fr.pack(fill="both", expand=True)
        share_note = ""
        if item_id in self._GUARDRAILS_DHCP_ITEM_IDS:
            share_note = "  (캡처호스트·RU SSH 공유 · Option/IA_NA regex는 family 공유 · L2SW/VLAN은 VLAN 항목)"
        ttk.Label(fr, text=schema["title"] + share_note, font=("", 11, "bold")).grid(
            row=0, column=0, columnspan=3, sticky="w", pady=(0, 4)
        )
        if item_id in ("dhcp_v4_vlan", "dhcp_v6_vlan"):
            ttk.Label(
                fr,
                text="★ = 시험 VLAN ID  ·  RU MAC=ALLOWED_IP ping→neigh 자동  ·  "
                "RU ID·PW·IPv4/IPv6 = Settings 공용  ·  Capture/L2SW IF 자동",
                foreground="#b45309",
                font=("", 9, "bold"),
            ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(0, 8))
            row_i = 2
        elif item_id in ("dhcp_v4", "dhcp_v6"):
            ttk.Label(
                fr,
                text="Boot: 재시작 후 IP 재수신 + Option tcpdump  ·  "
                "RU ID·PW·IP = Settings 공용  ·  Capture IF 자동",
                foreground="#64748b",
                font=("", 9),
            ).grid(row=1, column=0, columnspan=3, sticky="w", pady=(0, 8))
            row_i = 2
        else:
            row_i = 1
        entries: dict[str, tk.StringVar] = {}
        sk = self._guardrails_store_key(item_id)
        # Merge stored blob + sibling shared values for display
        cur = dict((getattr(self, "_guardrails_per_test_settings", {}) or {}).get(sk) or {})
        if not cur:
            cur = dict((getattr(self, "_guardrails_per_test_settings", {}) or {}).get(item_id) or {})
        for field in schema["fields"]:
            key = str(field.get("key") or "")
            if not key:
                continue
            got = self._guardrails_get_val(item_id, key, str(field.get("default") or ""))
            if got and not (cur.get(key) or "").strip():
                cur[key] = got
        for field in schema["fields"]:
            if field.get("hidden"):
                continue
            emph = bool(field.get("emphasize"))
            lbl_kw: dict[str, Any] = {"text": field["label"]}
            if emph:
                lbl_kw["foreground"] = "#b45309"
                lbl_kw["font"] = ("", 9, "bold")
            ttk.Label(fr, **lbl_kw).grid(row=row_i, column=0, sticky="w", padx=(0, 8), pady=2)
            raw_val = cur.get(field["key"])
            if raw_val is None or str(raw_val).strip() == "":
                raw_val = self._guardrails_get_val(
                    item_id, str(field["key"]), str(field.get("default") or "")
                )
            if raw_val is None or str(raw_val).strip() == "":
                raw_val = field.get("default") or ""
            sv = tk.StringVar(value=str(raw_val))
            if field.get("widget") == "checkbox":
                on = str(raw_val).strip().lower() not in ("0", "false", "no", "n", "")
                sv.set("1" if on else "0")
                ttk.Checkbutton(fr, variable=sv, onvalue="1", offvalue="0").grid(
                    row=row_i, column=1, sticky="w", pady=2
                )
            else:
                w = 40 if field.get("wide") else 14
                show_pw = "*" if field.get("password") else ""
                ent = ttk.Entry(fr, textvariable=sv, width=w, show=show_pw)
                ent.grid(row=row_i, column=1, sticky="we", pady=2)
            entries[field["key"]] = sv
            hint = field.get("hint") or ""
            if hint:
                ttk.Label(
                    fr,
                    text=hint,
                    foreground="#b45309" if emph else "#64748b",
                    font=("", 8, "bold") if emph else ("", 8),
                ).grid(row=row_i, column=2, sticky="w", padx=(6, 0), pady=2)
            row_i += 1

        def _apply() -> None:
            vals: dict[str, str] = {}
            for field in schema["fields"]:
                key = field["key"]
                if field.get("hidden"):
                    # keep previously saved / default — not shown in UI
                    continue
                sv = entries.get(key)
                if sv is None:
                    continue
                v = (sv.get() or "").strip()
                if field.get("widget") == "checkbox":
                    vals[key] = "1" if v in ("1", "true", "yes", "on") else "0"
                elif v:
                    vals[key] = v
            # merge with shared store so v4-only regex + v6-only regex both keep
            merged = dict(cur)
            merged.update(vals)
            # drop empty
            merged = {k: v for k, v in merged.items() if str(v).strip()}
            self._guardrails_set_vals(item_id, merged)
            sib = self._guardrails_dhcp_sibling_key(item_id)
            try:
                a4 = self._guardrails_strip_ip_cidr(
                    (self.fields.get("ALLOWED_IP").get() or "").strip()  # type: ignore[union-attr]
                )
            except Exception:
                a4 = ""
            try:
                a6 = self._guardrails_strip_ip_cidr(
                    (self.fields.get("ALLOWED_IP_V6").get() or "").strip()  # type: ignore[union-attr]
                )
            except Exception:
                a6 = ""
            self.append_log(
                f"[GUI] M-Plane Test 항목 설정 저장: {item_id}/{sk} ({len(merged)}개 필드)"
                + (f" · 공통항목→{sib} 동기화" if sib else "")
                + (f" · RU IP=Settings({a4 or '-'}" + (f"/{a6}" if a6 and ':' in a6 else "") + ")" if (a4 or a6) else "")
                + "\n"
            )
            for other in self._GUARDRAILS_PER_TEST_SCHEMA:
                self._guardrails_sync_tree_row(other)
            win.destroy()

        def _reset() -> None:
            for field in schema["fields"]:
                if field.get("hidden"):
                    continue
                sv = entries.get(field["key"])
                if sv is not None:
                    sv.set(str(field.get("default") or ""))

        def _mac_if_lookup_now() -> None:
            vals = {k: (sv.get() or "").strip() for k, sv in entries.items() if (sv.get() or "").strip()}
            merged = dict(cur)
            merged.update(vals)
            self._guardrails_set_vals(item_id, merged)
            self._guardrails_settings_item_id = item_id

            def work() -> None:
                st_mac, det_mac = self._guardrails_ensure_ru_mac(item_id, force=True)
                if st_mac == "FAIL":
                    self._guardrails_log(f"⚙ MAC→IF ({item_id}) → FAIL: {det_mac}")
                    self.after(
                        0,
                        lambda: messagebox.showerror(
                            "M-Plane Test", f"RU MAC 조회 실패\n{det_mac}", parent=win
                        ),
                    )
                    return
                st, detail = self._guardrails_ensure_l2sw_if_from_mac(item_id, force=True)
                ifc = self._guardrails_l2sw_normalize_if()
                mac = (self._guardrails_gf("ru_mac") or "").strip()

                def _ui() -> None:
                    if ifc and "l2sw_if" in entries:
                        entries["l2sw_if"].set(ifc)
                    messagebox.showinfo(
                        "M-Plane Test",
                        f"MAC: {st_mac}\n{det_mac}\n\nIF: {st}\n{detail}",
                        parent=win,
                    )

                self._guardrails_log(
                    f"⚙ MAC→IF ({item_id}) → MAC={mac or '-'} / {st}: {detail}"
                )
                self.after(0, _ui)

            threading.Thread(target=work, daemon=True).start()

        def _acl_apply_now() -> None:
            # persist current form values first
            vals = {k: (sv.get() or "").strip() for k, sv in entries.items() if (sv.get() or "").strip()}
            merged = dict(cur)
            merged.update(vals)
            self._guardrails_set_vals(item_id, merged)
            self._guardrails_settings_item_id = item_id
            fam = self._guardrails_item_family(item_id)

            def work() -> None:
                st, detail = self._guardrails_apply_acl(fam)
                ifc = self._guardrails_l2sw_normalize_if()

                def _ui() -> None:
                    if ifc and "l2sw_if" in entries:
                        entries["l2sw_if"].set(ifc)
                    messagebox.showinfo("M-Plane Test", f"{st}\n{detail}", parent=win)

                self._guardrails_log(f"⚙ ACL 적용({fam}/{item_id}) → {st}: {detail}")
                self.after(0, _ui)

            threading.Thread(target=work, daemon=True).start()

        def _acl_remove_now() -> None:
            vals = {k: (sv.get() or "").strip() for k, sv in entries.items() if (sv.get() or "").strip()}
            merged = dict(cur)
            merged.update(vals)
            self._guardrails_set_vals(item_id, merged)
            self._guardrails_settings_item_id = item_id

            def work() -> None:
                st, detail = self._guardrails_remove_acl()
                self._guardrails_log(f"⚙ ACL 제거({item_id}) → {st}: {detail}")
                self.after(0, lambda: messagebox.showinfo("M-Plane Test", f"{st}\n{detail}", parent=win))

            threading.Thread(target=work, daemon=True).start()

        btn_fr = ttk.Frame(fr)
        btn_fr.grid(row=row_i + 1, column=0, columnspan=3, pady=(10, 0), sticky="w")
        ttk.Button(btn_fr, text="초기화", command=_reset, width=8).pack(side="left", padx=(0, 8))
        ttk.Button(btn_fr, text="적용", command=_apply, width=8).pack(side="left", padx=(0, 8))
        ttk.Button(btn_fr, text="취소", command=win.destroy, width=8).pack(side="left", padx=(0, 16))
        if item_id in ("dhcp_v4_vlan", "dhcp_v6_vlan", "dhcp_v4_only_boot", "dhcp_v6_only_boot"):
            ttk.Button(btn_fr, text="MAC→IF 조회", command=_mac_if_lookup_now, width=12).pack(
                side="left", padx=(0, 6)
            )
            ttk.Button(btn_fr, text="ACL만 적용(테스트)", command=_acl_apply_now, width=16).pack(
                side="left", padx=(0, 6)
            )
            ttk.Button(btn_fr, text="ACL만 제거", command=_acl_remove_now, width=12).pack(side="left")
            ttk.Label(
                fr,
                text="※ VLAN 검증 순서: ① ALLOWED_IP ping→ip neigh(MAC) ② L2SW show mac(IF) "
                "③ Capture IF(ifconfig). 「MAC→IF 조회」도 동일 순서.",
                foreground="#64748b",
                wraplength=520,
            ).grid(row=len(schema["fields"]) + 2, column=0, columnspan=3, sticky="w", pady=(8, 0))
        win.update_idletasks()
        pw, ph = win.winfo_width(), win.winfo_height()
        sx = self.winfo_rootx() + (self.winfo_width() - pw) // 2
        sy = self.winfo_rooty() + (self.winfo_height() - ph) // 3
        win.geometry(f"+{max(0, sx)}+{max(0, sy)}")

    def _guardrails_l2sw_run_cmds(
        self,
        cmds: list[str],
        settle_s: float = 0.5,
        timeout: int | None = None,
        *,
        require_config: bool = True,
        tolerate_invalid: bool = False,
    ) -> tuple[bool, str]:
        """Send CLI lines to L2SW via solid→sshpass (same path as Conformance 3151).

        require_config=False: show-only 세션 (enable→show) — (config 미진입 OK.
        tolerate_invalid=True: % Invalid input 있어도 출력 반환 (show mac 변형 시도용).
        """
        host = self._guardrails_gf("l2sw_ip")
        user = self._guardrails_gf("l2sw_id")
        pw = self._guardrails_gf("l2sw_pw")
        ifc = self._guardrails_l2sw_normalize_if() or "-"
        if not host or not user:
            return False, "L2SW IP/ID 필요 (⚙ 설정)"
        if " " in user or user.lower().startswith("ethernet") or user.lower().startswith("eth"):
            return False, (
                f"L2SW ID가 이상함: {user!r} (포트명이 ID에 들어간 듯). "
                "⚙ L2SW ID=admin, L2SW IF=ethernet 0/22 로 다시 저장하세요."
            )
        # Build feeder on solid: sshpass → L2SW interactive SSH
        feeder_parts: list[str] = ["sleep 2.5"]
        for line in cmds:
            feeder_parts.append(f"printf '%s\\r\\n' {shlex.quote(line)}")
            pause = settle_s
            low = line.strip().lower()
            if low in ("enable", "configure terminal", "configure", "conf t"):
                pause = max(settle_s, 1.2)
            elif low.startswith("show mac"):
                pause = max(settle_s, 1.8)
            feeder_parts.append(f"sleep {pause}")
        # Dasan: privileged '#' 에서 'end' 는 Invalid — exit 만 사용
        feeder_parts.append("printf 'exit\\r\\n'")
        feeder_parts.append("sleep 0.5")
        feeder = "; ".join(feeder_parts)
        # Worst-case: ~2.5 + N*(settle+1.2) + SSH handshake; keep SSH timeout generous
        est_s = int(2.5 + len(cmds) * (max(settle_s, 0.8) + 0.3) + 25)
        ssh_to = int(timeout) if timeout is not None else max(120, est_s + 15)
        remote = (
            "export SSHPASS=" + shlex.quote(pw) + "; "
            f"( {feeder} ) | sshpass -e ssh -tt "
            "-o StrictHostKeyChecking=no -o UserKnownHostsFile=/dev/null -o LogLevel=ERROR "
            "-o ConnectTimeout=12 "
            + shlex.quote(f"{user}@{host}")
            + " 2>&1; "
            "echo L2SW_SSH_DONE"
        )
        self._guardrails_log(
            f"L2SW CLI via solid → {user}@{host} if={ifc} ({len(cmds)} cmds, 최대 ~{ssh_to}s 대기)"
        )
        ok, text = self._guardrails_ssh_exec(f"bash -lc {shlex.quote(remote)}", timeout=ssh_to)
        out = text or ""
        # ssh -tt often returns non-zero on exit; judge by CLI transcript
        if "Connection timed out" in out or "Connection refused" in out:
            return False, out[:300]
        if "Permission denied" in out:
            return False, out[:300]
        if "no matching host key type" in out.lower() or "host key type found" in out.lower():
            self._guardrails_log(f"[L2SW] FAIL host key: {out[:240].replace(chr(10), ' | ')}")
            return False, (
                "SSH host key 협상 실패(ssh-rsa/dss). "
                "옵션 재시도 후에도 실패면 solid OpenSSH 확인: " + out[:200]
            )
        if "L2SW_SSH_DONE" not in out and not ok:
            if "timed out" in out.lower() or "timeout" in out.lower():
                return False, f"timed out (solid→{host}): {out[:200]}"
            return False, out[:300] or "exit fail"
        # Real Dasan failure: bad syntax / wrong mode
        # Soft messages while ACL still attached / already present — not hard FAIL
        soft_markers = (
            "Access-group already configured",
            "This object already exist",
            "object already exist",
            "Modification of ACL is not permitted",
            "This object is in use",
            "object is in use",
            "Port is already added to the Vlan",
            "already added to the Vlan",
            "is not a member",
            "not a member of the VLAN",
        )
        if any(b in out for b in soft_markers):
            self._guardrails_log("[L2SW] note: ACL/VLAN already/in-use 메시지 (무시 가능)")
        if "% Invalid input" in out or "Invalid input detected" in out:
            if tolerate_invalid:
                self._guardrails_log(
                    f"[L2SW] Invalid input (tolerated), out={out[-400:].replace(chr(10), ' | ')}"
                )
            else:
                # Ignore Invalid only if it's clearly from a soft conflict line nearby — still FAIL otherwise
                self._guardrails_log(f"[L2SW] FAIL Invalid input, out={out[-500:].replace(chr(10), ' | ')}")
                return False, "Invalid input (enable/configure/IF 문법 확인): " + out[-400:]
        if require_config and "(config" not in out.lower():
            self._guardrails_log(f"[L2SW] FAIL no config mode, out={out[-400:].replace(chr(10), ' | ')}")
            return False, "config mode 미진입 (enable→configure terminal): " + out[-300:]
        self._guardrails_log(f"[L2SW] ok, out={out[-400:].replace(chr(10), ' | ')}")
        return True, out or "(ok)"

    @staticmethod
    def _guardrails_parse_oran_alarm_table(query: str) -> dict[str, dict[str, str]]:
        """Parse `show alarm information oran` rows → {alarm_id: fields}."""
        out: dict[str, dict[str, str]] = {}
        for raw in (query or "").splitlines():
            line = raw.strip()
            if not line or line.startswith("#") or line.startswith("---"):
                continue
            # skip banners / headers
            low = line.lower()
            if "alarm id" in low and "fault" in low:
                continue
            if low.startswith(("alarm oper", "alarm detection", "alarm notification", "total count")):
                continue
            # Typical: AlarmId FaultId Severity Name... Source Config Interval Related
            m = re.match(
                r"^(\d+)\s+(\d+)\s+(CRITICAL|MAJOR|MINOR|WARNING|NORMAL|INDETERMINATE)\s+(.+)$",
                line,
                re.I,
            )
            if not m:
                # looser: leading alarm-id then rest
                m2 = re.match(r"^(\d+)\s+(\S+)\s+(\S+)\s+(.+)$", line)
                if not m2:
                    continue
                aid, fid, sev, rest = m2.group(1), m2.group(2), m2.group(3), m2.group(4)
            else:
                aid, fid, sev, rest = m.group(1), m.group(2), m.group(3), m.group(4)
            # peel source/config from the rightish tokens when possible
            toks = rest.split()
            source = ""
            config = ""
            name = rest
            # source keywords often: module, ecpri, ant-line-tx, ...
            for i, t in enumerate(toks):
                if t.lower() in (
                    "module",
                    "ecpri",
                    "ant-line-tx",
                    "ant-line-rx",
                    "carrier-tx",
                    "carrier-rx",
                ) or t.lower().startswith("ant-") or t.lower().startswith("carrier-"):
                    source = t
                    name = " ".join(toks[:i]).strip()
                    rem = toks[i + 1 :]
                    if rem:
                        config = rem[0]
                    break
            out[aid] = {
                "alarm_id": aid,
                "fault_id": fid,
                "severity": sev,
                "name": name or rest[:60],
                "source": source,
                "config": config,
                "raw": line,
            }
        return out

    @staticmethod
    def _guardrails_format_fault_alarm_detail(raw: str) -> str:
        """Parse mplane_fault_alarm.sh markers → detail (noti 중심, CLI 표 전문 없음)."""
        text = raw or ""
        lines_out: list[str] = []

        def _between(a: str, b: str) -> str:
            i = text.find(a)
            if i < 0:
                return ""
            i += len(a)
            j = text.find(b, i)
            if j < 0:
                return text[i:].strip()
            return text[i:j].strip()

        query = _between("===ALARM_QUERY_BEGIN===", "===ALARM_QUERY_END===")
        active = _between("===ACTIVE_ALARMS_BEGIN===", "===ACTIVE_ALARMS_END===")
        # Prefer compact alarm_row: from new script; fall back to full CLI parse
        table: dict[str, dict[str, str]] = {}
        for ln in (query or "").splitlines():
            s = ln.strip()
            if not s.startswith("alarm_row:"):
                continue
            parts = s[len("alarm_row:") :].split("|")
            if len(parts) < 3:
                continue
            aid = parts[0].strip()
            table[aid] = {
                "alarm_id": aid,
                "fault_id": parts[1].strip() if len(parts) > 1 else "",
                "severity": parts[2].strip() if len(parts) > 2 else "",
                "name": parts[3].strip() if len(parts) > 3 else "",
                "source": parts[4].strip() if len(parts) > 4 else "",
                "config": parts[5].strip() if len(parts) > 5 else "",
            }
        if not table:
            table = CallhomeGUI._guardrails_parse_oran_alarm_table(query)

        active_present = "—"
        active_ids = ""
        active_count = ""
        m_ap = re.search(r"(?m)^active_present:\s*(\S+)", active)
        if m_ap:
            active_present = m_ap.group(1).strip().upper()
        m_ai = re.search(r"(?m)^active_ids:\s*(.*)$", active)
        if m_ai:
            active_ids = (m_ai.group(1) or "").strip()
        m_ac = re.search(r"(?m)^active_count:\s*(\S+)", active)
        if m_ac:
            active_count = m_ac.group(1).strip()

        lines_out.extend(
            [
                "[0] 시험 전 Active (요약만 — CLI 표는 로그 생략)",
                "------------------------------------------------------------",
                f"  active_present: {active_present}"
                + (f"  count={active_count}" if active_count else "")
                + (f"  ids={active_ids}" if active_ids else ""),
            ]
        )
        if active_present == "YES":
            lines_out.append(
                "  ※ 이미 active 인 Alarm Id 는 시험에서 제외 (아래 [1.5])"
            )

        skipped = _between("===SKIPPED_ACTIVE_BEGIN===", "===SKIPPED_ACTIVE_END===")
        skip_ids = ""
        skip_count = ""
        m_sc = re.search(r"(?m)^skipped_count:\s*(\S+)", skipped)
        if m_sc:
            skip_count = m_sc.group(1).strip()
        m_si = re.search(r"(?m)^skipped_ids:\s*(.*)$", skipped)
        if m_si:
            skip_ids = (m_si.group(1) or "").strip()

        alarm_count = ""
        alarm_ids = ""
        m_n = re.search(r"(?m)^alarm_count:\s*(\S+)", query)
        if m_n:
            alarm_count = m_n.group(1).strip()
        m_ids = re.search(r"(?m)^alarm_ids:\s*(.*)$", query)
        if m_ids:
            alarm_ids = (m_ids.group(1) or "").strip()

        lines_out.extend(
            [
                "",
                "[1] Catalog (information → Alarm Id 목록)",
                "------------------------------------------------------------",
                f"  alarm_count: {alarm_count or len(table)}",
                f"  alarm_ids: {alarm_ids or ','.join(table.keys()) or '(없음)'}",
            ]
        )

        lines_out.extend(
            [
                "",
                "[1.5] 시험 제외 (이미 active)",
                "------------------------------------------------------------",
                f"  skipped_count: {skip_count or ('0' if not skip_ids else len([x for x in skip_ids.split(',') if x]))}",
                f"  skipped_ids: {skip_ids or '(없음)'}",
            ]
        )
        skip_rows = [
            ln.strip()
            for ln in (skipped or "").splitlines()
            if ln.strip().startswith("skip_row:")
        ]
        if skip_rows:
            for ln in skip_rows:
                parts = ln[len("skip_row:") :].split("|")
                aid = parts[0] if parts else "?"
                fid = parts[1] if len(parts) > 1 else ""
                sev = parts[2] if len(parts) > 2 else ""
                name = parts[3] if len(parts) > 3 else ""
                reason = parts[4] if len(parts) > 4 else "already_active"
                extra = f" fault-id={fid}" if fid else ""
                extra2 = f"  {sev}" if sev else ""
                extra3 = f"  {name}" if name else ""
                lines_out.append(f"  SKIP alarm-id {aid}:{extra}{extra2}{extra3}  ({reason})")
        elif active_ids:
            for aid in [x.strip() for x in active_ids.split(",") if x.strip()]:
                info = table.get(aid) or {}
                name = info.get("name") or ""
                fid = info.get("fault_id") or ""
                sev = info.get("severity") or ""
                lines_out.append(
                    f"  SKIP alarm-id {aid}:"
                    + (f" fault-id={fid}" if fid else "")
                    + (f"  {sev}" if sev else "")
                    + (f"  {name}" if name else "")
                    + "  (already_active)"
                )

        lines_out.extend(
            [
                "",
                "[2] 시험 결과 (결과 | alarm | fault | severity | name | 발생(s) | 해지(s))",
                "------------------------------------------------------------",
                "  발생(s)=test alarm CLI → raise noti 수신까지",
                "  해지(s)=no test alarm CLI → clear noti 수신까지",
            ]
        )

        def _parse_wall(s: str):
            s = (s or "").strip()
            if not s or s == "—":
                return None
            s = s.replace("T", " ").replace("Z", "")
            s = re.sub(r"\.\d+$", "", s)
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
                try:
                    return datetime.strptime(s[:19], fmt)
                except Exception:
                    continue
            return None

        def _delta_sec(t0: str, t1: str, noti_ok: str) -> str:
            if (noti_ok or "").strip().upper() != "OK":
                return "TIMEOUT"
            a = _parse_wall(t0)
            b = _parse_wall(t1)
            if a is None or b is None:
                return "—"
            sec = (b - a).total_seconds()
            if sec < 0:
                sec = 0.0
            if sec == int(sec):
                return str(int(sec))
            return f"{sec:.1f}"

        parts = re.split(r"===ALARM alarm_id=|===FAULT fault_id=", text)
        result_rows: list[tuple[str, str, str, str, str, str, str]] = []
        for part in parts[1:]:
            m_id = re.match(r"([^\s=]+)===\s*", part)
            if not m_id:
                continue
            aid = m_id.group(1).strip()
            body = part[m_id.end() :]
            end = body.find("===FAULT_END===")
            if end < 0:
                end = body.find("===ALARM_END===")
            if end >= 0:
                body = body[:end]

            def _kv(key: str, _body: str = body) -> str:
                mm = re.search(rf"(?m)^{re.escape(key)}:\s*(.*)$", _body)
                return (mm.group(1).strip() if mm else "") or "—"

            info = table.get(aid) or {}
            name = (info.get("name") or "").strip()
            fid = (info.get("fault_id") or _kv("yang_fault_id") or "").strip()
            sev = (info.get("severity") or "").strip()
            fr = (_kv("fault_result") or "—").strip().upper()
            if fr not in ("PASS", "FAIL"):
                fr = "FAIL" if "FAIL" in fr else (fr or "—")
            raise_sec = _delta_sec(_kv("raise_sent"), _kv("raise_noti_wall"), _kv("raise_noti"))
            clear_sec = _delta_sec(_kv("clear_sent"), _kv("clear_noti_wall"), _kv("clear_noti"))
            result_rows.append((fr, aid, fid, sev, name, raise_sec, clear_sec))

        lines_out.append("===RESULT_TABLE_BEGIN===")
        lines_out.append("# result|alarm_id|fault_id|severity|name|raise_sec|clear_sec")
        if not result_rows:
            lines_out.append("  (시험 결과 행 없음)")
        else:
            lines_out.append(
                f"  {'결과':<6} {'alarm-id':<12} {'fault-id':<12} {'severity':<10} "
                f"{'name':<34} {'발생(s)':>8} {'해지(s)':>8}"
            )
            for fr, aid, fid, sev, name, rs, cs in result_rows:
                nm = (name or "")[:34]
                lines_out.append(
                    f"  {fr:<6} {aid:<12} {fid:<12} {sev:<10} "
                    f"{nm:<34} {rs:>8} {cs:>8}"
                )
                safe_name = (name or "").replace("|", "/")
                lines_out.append(f"RESULT_ROW|{fr}|{aid}|{fid}|{sev}|{safe_name}|{rs}|{cs}")
        lines_out.append("===RESULT_TABLE_END===")

        summ = _between("===SUMMARY===", "===SUMMARY_END===")
        lines_out.extend(["", "[요약]", f"  {(summ or '(없음)').strip()}"])

        if "===ALARM alarm_id=" not in text and "===FAULT fault_id=" not in text and text.strip():
            lines_out.extend(["", "[원본 출력 일부]", text[-2500:]])

        return "\n".join(lines_out).rstrip() + "\n"

    def _guardrails_run_fault_alarm(self) -> tuple[str, str]:
        """CallHome session held; batch ORU test-alarm raise/clear + noti; structured detail."""
        self._guardrails_fill_defaults_from_context("fault_alarm")
        helper = "mplane_fault_alarm.sh"
        # Alarm IDs 설정 없음 — 항상 show alarm information oran 전체
        try:
            sk = self._guardrails_store_key("fault_alarm")
            blob = dict(
                (getattr(self, "_guardrails_per_test_settings", {}) or {}).get(sk) or {}
            )
            if "fault_ids" in blob:
                blob.pop("fault_ids", None)
                self._guardrails_set_vals("fault_alarm", blob)
                self._guardrails_log("[fault_alarm] 저장값 fault_ids 제거 → information 전체 시험")
        except Exception:
            pass

        show_cmd = (self._guardrails_gf("show_cmd") or "show alarm information oran").strip()
        active_show_cmd = (
            self._guardrails_gf("active_show_cmd") or "show alarm active-alarms"
        ).strip()
        _canon_raise = "test alarm alarm-id {alarm_id} source-id {source_id} start-alarm"
        _canon_clear = "no test alarm alarm-id {alarm_id} source-id {source_id}"
        raise_tmpl = (self._guardrails_gf("raise_tmpl") or _canon_raise).strip()
        clear_tmpl = (self._guardrails_gf("clear_tmpl") or _canon_clear).strip()
        source_id = (self._guardrails_gf("source_id") or "0").strip() or "0"

        def _sanitize_alarm_tmpl(s: str, canon: str, *, expect_start: bool) -> str:
            """Reject broken templates (duplicated source-id / trailing })."""
            t = re.sub(r"\}+\s*$", "", (s or "").strip()).strip()
            bad = (
                not t
                or t.count("source-id") != 1
                or (t.count("{alarm_id}") + t.count("{fault_id}")) < 1
                or t.count("{source_id}") != 1
                or "start-alarm source-id" in t
                or t.endswith("}")
            )
            if expect_start and "start-alarm" not in t:
                bad = True
            if (not expect_start) and not t.lower().startswith("no "):
                # clear must be `no test alarm ...`
                if "cancel-alarm" in t or "stop-alarm" in t:
                    bad = True
            if bad:
                return canon
            return t

        raise_tmpl = _sanitize_alarm_tmpl(raise_tmpl, _canon_raise, expect_start=True)
        clear_tmpl = _sanitize_alarm_tmpl(clear_tmpl, _canon_clear, expect_start=False)
        # persist cleaned templates so ⚙ 재실행 때도 깨진 값 안 씀
        try:
            sk = self._guardrails_store_key("fault_alarm")
            blob = dict(
                (getattr(self, "_guardrails_per_test_settings", {}) or {}).get(sk) or {}
            )
            if blob.get("raise_tmpl") != raise_tmpl or blob.get("clear_tmpl") != clear_tmpl:
                blob["raise_tmpl"] = raise_tmpl
                blob["clear_tmpl"] = clear_tmpl
                blob["source_id"] = source_id
                blob.pop("fault_ids", None)
                self._guardrails_set_vals("fault_alarm", blob)
                self._guardrails_log(
                    f"[fault_alarm] 템플릿 정리 저장 RAISE={raise_tmpl!r} CLEAR={clear_tmpl!r}"
                )
        except Exception:
            pass
        self._guardrails_log(
            f"[fault_alarm] RAISE_TMPL={raise_tmpl!r} CLEAR_TMPL={clear_tmpl!r} "
            f"SOURCE_ID={source_id!r} (ids=show information oran 전체)"
        )
        try:
            alarm_to = max(10, int(self._guardrails_gf("alarm_timeout_sec", "60") or "60"))
        except Exception:
            alarm_to = 60
        try:
            listen_to = max(30, int(self._guardrails_gf("listen_timeout_sec", "180") or "180"))
        except Exception:
            listen_to = 180
        require_noti = "1" if (self._guardrails_gf("require_noti", "1") or "1").strip() not in (
            "0",
            "false",
            "no",
            "n",
        ) else "0"
        skip_normal = "1" if (self._guardrails_gf("skip_normal", "1") or "1").strip() not in (
            "0",
            "false",
            "no",
            "n",
        ) else "0"

        fam = self._guardrails_resolve_ssh_family("fault_alarm")
        oru_host, how = self._guardrails_ru_ssh_target(fam)
        if not oru_host:
            return "FAIL", f"RU SSH 대상 없음: {how}"
        oru_id = self._guardrails_gf("oru_cli_id")
        oru_pw = self._guardrails_gf("oru_cli_pw")
        if not oru_id:
            return "FAIL", "RU SSH ID 필요 (Settings)"

        # CallHome listen plane: untag controller
        try:
            self._apply_lab_controller_listen_ips("untag")
        except Exception:
            pass

        local_ip = ""
        allowed_ip = ""
        try:
            if fam == "v6":
                local_ip = self._guardrails_strip_ip_cidr(
                    (self.fields.get("LOCAL_IP_V6").get() or "").strip()  # type: ignore[union-attr]
                )
                allowed_ip = self._guardrails_strip_ip_cidr(
                    (self.fields.get("ALLOWED_IP_V6").get() or "").strip()  # type: ignore[union-attr]
                )
            else:
                local_ip = self._guardrails_strip_ip_cidr(
                    (self.fields.get("LOCAL_IP").get() or "").strip()  # type: ignore[union-attr]
                )
                allowed_ip = self._guardrails_strip_ip_cidr(
                    (self.fields.get("ALLOWED_IP").get() or "").strip()  # type: ignore[union-attr]
                )
        except Exception:
            pass
        if not local_ip:
            local_ip = self._lab_controller_listen_ip(fam, "untag")
        if not allowed_ip:
            allowed_ip = oru_host
        if not local_ip or not allowed_ip:
            return "FAIL", "LOCAL_IP / ALLOWED_IP(또는 RU IP) 필요"

        ssh_user = self.remote_user.get().strip()
        ssh_host = self.remote_host.get().strip()
        ssh_port = self.remote_port.get().strip() or "22"
        ssh_password = self.remote_password.get()
        key_path = (self.remote_key_path.get() or "").strip()
        if not ssh_user or not ssh_host:
            return "FAIL", "Settings SSH_USER/SSH_HOST 필요"

        try:
            import paramiko  # type: ignore
        except Exception as exc:
            return "FAIL", f"paramiko 필요: {exc}"

        opts = self._conformance_default_run_options()
        remote_dir = opts.remote_dir.rstrip("/")
        cfg_remote = f"{remote_dir}/{_conf_manifest.CONFORMANCE_REMOTE_GUI_CONFIG_NAME}"
        lp = self._conformance_script_local_path(helper)
        if lp is None:
            cand = self._conformance_local_dir() / helper
            if cand.is_file():
                lp = cand
        if lp is None:
            return "FAIL", f"로컬 헬퍼 없음: {helper}"

        n_faults = 40  # information 표 기준 상한 (실제 개수는 remote catalog)
        hard_cap = float(listen_to) + float(alarm_to) * max(1, n_faults) * 2.5 + 120.0

        self._guardrails_log(
            f"[fault_alarm] session batch ALL from information oran fam={fam} "
            f"oru={oru_host} listen={local_ip}:{4334} ≤{listen_to}s hard={hard_cap:.0f}s "
            f"(PASS=raise/clear NETCONF noti)"
        )
        if getattr(self, "is_running", False):
            self._guardrails_log(
                "[fault_alarm] 경고: GUI Start(CallHome)가 켜져 있으면 포트 충돌 가능 — 중지 권장"
            )

        client: Any = None
        captured: list[str] = []
        try:
            try:
                self._conformance_cancel_event.clear()
            except Exception:
                pass
            client = paramiko.SSHClient()
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            client.connect(
                hostname=ssh_host,
                port=int(ssh_port),
                username=ssh_user,
                password=ssh_password if ssh_password else None,
                key_filename=key_path if key_path else None,
                timeout=20,
                auth_timeout=20,
                banner_timeout=20,
                look_for_keys=not bool(ssh_password),
                allow_agent=True,
            )
            _stdin, _stdout, _stderr = client.exec_command(f"mkdir -p {shlex.quote(remote_dir)}")
            _stdout.channel.recv_exit_status()
            sftp = client.open_sftp()
            rp = f"{remote_dir}/{helper}"
            try:
                raw = lp.read_bytes().replace(b"\r\n", b"\n").replace(b"\r", b"\n")
                sftp.putfo(io.BytesIO(raw), rp, len(raw))
                try:
                    sftp.chmod(rp, 0o755)
                except OSError:
                    pass
                self._guardrails_log(f"uploaded {helper}")
            except Exception as exc:
                return "FAIL", f"업로드 실패: {exc}"
            try:
                cfg_payload = self._conformance_effective_config_json_text()
                sftp.putfo(io.BytesIO(cfg_payload.encode("utf-8")), cfg_remote, len(cfg_payload.encode("utf-8")))
            except Exception as exc:
                self._guardrails_log(f"[WARN] config 갱신 실패: {exc}")
            # 템플릿 {source_id} 는 bash 인용에서 깨짐 → base64 로 env 파일 전달
            import base64 as _b64

            env_remote = f"{remote_dir}/mplane_fault_alarm.env.sh"
            raise_b64 = _b64.b64encode(raise_tmpl.encode("utf-8")).decode("ascii")
            clear_b64 = _b64.b64encode(clear_tmpl.encode("utf-8")).decode("ascii")
            env_lines = [
                "#!/usr/bin/env bash",
                f"export LOCAL_IP={shlex.quote(local_ip)}",
                f"export ALLOWED_IP={shlex.quote(allowed_ip)}",
                f"export FAULT_IDS=all",
                f"export SHOW_CMD={shlex.quote(show_cmd)}",
                f"export ACTIVE_SHOW_CMD={shlex.quote(active_show_cmd)}",
                f"export RAISE_TMPL_B64={shlex.quote(raise_b64)}",
                f"export CLEAR_TMPL_B64={shlex.quote(clear_b64)}",
                f"export SOURCE_ID={shlex.quote(source_id)}",
                f"export ALARM_TIMEOUT_SEC={int(alarm_to)}",
                f"export CALLHOME_LISTEN_TIMEOUT={int(listen_to)}",
                f"export ORU_CLI_ID={shlex.quote(oru_id)}",
                f"export ORU_CLI_PW={shlex.quote(oru_pw)}",
                f"export ORU_SSH_IP={shlex.quote(oru_host)}",
                f"export SSH_FAMILY={shlex.quote(fam)}",
                f"export REQUIRE_NOTI={shlex.quote(require_noti)}",
                f"export SKIP_NORMAL={shlex.quote(skip_normal)}",
                "",
            ]
            env_body = "\n".join(env_lines).replace("\r\n", "\n").replace("\r", "\n")
            env_bytes = env_body.encode("utf-8")
            try:
                sftp.putfo(io.BytesIO(env_bytes), env_remote, len(env_bytes))
                try:
                    sftp.chmod(env_remote, 0o644)
                except OSError:
                    pass
                self._guardrails_log(f"uploaded mplane_fault_alarm.env.sh (tmpl via base64)")
            except Exception as exc:
                return "FAIL", f"env 업로드 실패: {exc}"
            try:
                sftp.close()
            except Exception:
                pass

            envp = self._conformance_bash_env_exports(opts, None)
            host_log = self._conformance_host_run_log_path(helper)
            dir_q = shlex.quote(str(PurePosixPath(host_log).parent))
            log_q = shlex.quote(host_log)
            rp_q = shlex.quote(rp)
            cfg_q = shlex.quote(cfg_remote)
            env_q = shlex.quote(env_remote)
            runner = (
                f"{envp}"
                f"set -a ; . {env_q} ; set +a ; "
                f"export CONFORMANCE_SCRIPT_BASENAME={shlex.quote(helper)} ; "
                f"chmod +x {rp_q} 2>/dev/null ; bash {rp_q} --config {cfg_q}"
            )
            wrapped = (
                f"set -o pipefail; "
                f"mkdir -p {dir_q} && : > {log_q} && chmod 0644 {log_q} || exit 1; "
                f"( {runner} ) 2>&1 | tee -a {log_q}; "
                "_cf_rc=${PIPESTATUS[0]}; "
                'exit "${_cf_rc:-0}"'
            )
            cmd_remote = "bash -lc " + shlex.quote(wrapped)
            self._guardrails_log(f"---- START {helper} ----")
            self._guardrails_log(f"remote host log: {host_log}")
            self._guardrails_log(
                f"[fault_alarm] env RAISE={raise_tmpl!r} CLEAR={clear_tmpl!r} "
                f"ids=information-oran-all SKIP_NORMAL={skip_normal} (noti-only pass)"
            )
            _stdin, stdout, stderr = client.exec_command(cmd_remote, get_pty=True)
            ch = stdout.channel
            with self._conformance_run_transport_lock:
                self._conformance_run_script_channel = ch
            t0 = time.monotonic()
            while not ch.exit_status_ready():
                if self._guardrails_cancel.is_set() or self._conformance_cancel_event.is_set():
                    try:
                        ch.close()
                    except Exception:
                        pass
                    try:
                        client.exec_command(
                            f"pkill -f {shlex.quote(helper)} 2>/dev/null; "
                            f"fuser -k 4334/tcp 2>/dev/null || true"
                        )
                    except Exception:
                        pass
                    with self._conformance_run_transport_lock:
                        self._conformance_run_script_channel = None
                    return "FAIL", "사용자 중지"
                if time.monotonic() - t0 > hard_cap:
                    try:
                        ch.close()
                    except Exception:
                        pass
                    with self._conformance_run_transport_lock:
                        self._conformance_run_script_channel = None
                    detail = self._guardrails_format_fault_alarm_detail("\n".join(captured))
                    return "FAIL", f"타임아웃 ({hard_cap:.0f}s)\n\n{detail}"
                if ch.recv_ready():
                    chunk = ch.recv(4096).decode(errors="ignore")
                    if chunk:
                        for line in chunk.splitlines():
                            captured.append(line)
                            self._guardrails_log(line)
                else:
                    time.sleep(0.15)
            # drain
            try:
                rest = stdout.read().decode(errors="ignore")
                if rest:
                    for line in rest.splitlines():
                        captured.append(line)
                        self._guardrails_log(line)
            except Exception:
                pass
            rc = ch.recv_exit_status()
            with self._conformance_run_transport_lock:
                self._conformance_run_script_channel = None
            raw_out = "\n".join(captured)
            detail = self._guardrails_format_fault_alarm_detail(raw_out)
            status = "PASS" if rc == 0 and "PASS" in raw_out else "FAIL"
            # prefer SUMMARY line
            msum = re.search(r"===SUMMARY===\s*\n([^\n]+)", raw_out)
            if msum:
                head = msum.group(1).strip()
                if head.upper().startswith("PASS"):
                    status = "PASS"
                elif head.upper().startswith("FAIL"):
                    status = "FAIL"
                detail = f"{head}\n\n{detail}"
            else:
                detail = f"rc={rc}\n\n{detail}"
            return status, detail
        except Exception as exc:
            return "FAIL", f"fault_alarm 실행 오류: {exc}\n\n" + self._guardrails_format_fault_alarm_detail(
                "\n".join(captured)
            )
        finally:
            try:
                if client is not None:
                    client.close()
            except Exception:
                pass

    def _guardrails_run_one(self, item_id: str) -> tuple[str, str]:
        """Return (PASS|FAIL|INFO|SKIP, detail)."""
        self._guardrails_settings_item_id = item_id
        if item_id in ("dhcp_v4", "dhcp_v4_vlan", "dhcp_v4_only_boot", "dhcp_options"):
            try:
                return self._guardrails_run_dhcp_options_family(
                    "dhcp_v4_vlan" if item_id == "dhcp_v4_vlan" else "dhcp_v4"
                )
            finally:
                self._guardrails_cleanup_parent_v6_pending()
        if item_id in ("dhcp_v6", "dhcp_v6_vlan", "dhcp_v6_only_boot"):
            try:
                return self._guardrails_run_dhcp_options_family(
                    "dhcp_v6_vlan" if item_id == "dhcp_v6_vlan" else "dhcp_v6"
                )
            finally:
                self._guardrails_cleanup_parent_v6_pending()
        if item_id == "dhcp_boot":
            return self._guardrails_run_dhcp_boot()
        if item_id == "vlan_discovery":
            return self._guardrails_run_vlan_discovery()
        if item_id == "fault_alarm":
            return self._guardrails_run_fault_alarm()
        if item_id in (
            "netconf_capability",
            "config_states",
            "config_admin_state",
            "config_oper_state",
            "config_availability_state",
            "config_usage_state",
            "performance_mgmt",
        ):
            if item_id == "performance_mgmt":
                return "SKIP", "목록 등록만 — O-RAN Player 연동 후 구현"
            return "SKIP", "목록 등록만 — 추후 구현"
        return "SKIP", f"미정의 항목: {item_id}"

    def _guardrails_result_tag(self, text: str) -> str:
        u = (text or "").upper()
        if u.startswith("PASS"):
            return "res_pass"
        if u.startswith("FAIL"):
            return "res_fail"
        if u.startswith("INFO") or u.startswith("SKIP"):
            return "res_stop"
        if u.startswith("RUN") or "검증" in (text or ""):
            return "res_run"
        return "res_idle"

    def _guardrails_sync_tree_row(self, item_id: str) -> None:
        tree = getattr(self, "guardrails_list_tree", None)
        if tree is None:
            return
        try:
            if not tree.exists(item_id):
                return
        except tk.TclError:
            return
        bv = self.guardrails_check_vars.get(item_id)
        rv = self.guardrails_result_vars.get(item_id)
        pick = "☑" if (bv is not None and bv.get()) else "☐"
        has_cfg = item_id in self._GUARDRAILS_PER_TEST_SCHEMA
        cfg_mark = "⚙" if has_cfg else ""
        sk = self._guardrails_store_key(item_id) if has_cfg else item_id
        store = getattr(self, "_guardrails_per_test_settings", {}) or {}
        stored = store.get(sk) or store.get(item_id) or {}
        if has_cfg and stored:
            cfg_mark = "⚙✓"
        result = (rv.get() if rv is not None else "—") or "—"
        # 목록에는 짧은 판정만 (상세는 더블클릭 팝업 / Logs)
        short = result.strip()
        for prefix in ("PASS", "FAIL", "SKIP", "INFO", "RUN"):
            if short.upper().startswith(prefix):
                short = prefix if prefix != "RUN" else "RUN…"
                break
        else:
            if len(short) > 12:
                short = short[:10] + "…"
        vals = list(tree.item(item_id, "values"))
        if len(vals) >= 6:
            vals[0] = pick
            vals[4] = cfg_mark
            vals[5] = short
            tag = self._guardrails_result_tag(result)
            parity = "row_odd" if tree.index(item_id) % 2 else "row_even"
            tree.item(item_id, values=tuple(vals), tags=(parity, tag))

    def _build_guardrails_tab(self, parent: ttk.Frame) -> None:
        if not hasattr(self, "_guardrails_user_items") or self._guardrails_user_items is None:
            self._guardrails_user_items = []
        if not hasattr(self, "_guardrails_per_test_settings") or self._guardrails_per_test_settings is None:
            self._guardrails_per_test_settings = {}
        self.guardrails_check_vars.clear()
        self.guardrails_result_vars.clear()

        intro = ttk.LabelFrame(parent, text="M-Plane Test — 현장/시스템 검증 자동화 (Conformance 외)", padding=8)
        intro.pack(fill="x", padx=8, pady=(8, 4))
        ttk.Label(
            intro,
            text=(
                "엑셀 통합 ORU 검증 중 Conformance와 겹치지 않는 M-Plane 항목을 자동화합니다 "
                "(TLS CallHome·CCM·802.1X·Management 일부 제외). "
                "DHCP v4/v6는 검증 실행 시 ACL 자동 적용·원복(solid→L2SW). "
                "L2SW/ACL 값은 항목 ⚙에 두고, ACL 수동 테스트도 ⚙ 안에서 합니다."
            ),
            foreground="#475569",
            wraplength=1000,
            justify="left",
        ).pack(anchor="w")

        bar = ttk.Frame(parent)
        bar.pack(fill="x", padx=8, pady=6)
        ttk.Button(bar, text="일괄 선택", command=self._guardrails_select_all).pack(side="left", padx=(0, 4))
        ttk.Button(bar, text="일괄 해지", command=self._guardrails_clear_all).pack(side="left", padx=(0, 8))
        ttk.Button(
            bar,
            text="선택 항목 검증 실행",
            command=self._guardrails_run_checked,
            style="Big.TButton",
        ).pack(side="left", padx=(0, 8))
        self.guardrails_stop_btn = ttk.Button(
            bar, text="검증 중지", command=self._guardrails_stop, state="disabled"
        )
        self.guardrails_stop_btn.pack(side="left", padx=(0, 8))
        ttk.Label(bar, text="반복").pack(side="left", padx=(0, 2))
        ttk.Entry(bar, textvariable=self.guardrails_run_repeat_var, width=5).pack(side="left")
        ttk.Label(bar, text="(0=무한)", foreground="#64748b").pack(side="left", padx=(2, 8))
        ttk.Button(bar, text="Probe 1회", command=self._guardrails_btn_probe).pack(side="left")
        ttk.Button(
            bar, text="로컬 pcap 폴더", command=self._guardrails_open_local_pcap_folder
        ).pack(side="left", padx=(8, 0))
        ttk.Button(
            bar, text="결과 txt 폴더", command=self._guardrails_open_local_results_folder
        ).pack(side="left", padx=(8, 0))
        self.guardrails_vlan_restore_btn = ttk.Button(
            bar,
            text="VLAN untag 원복",
            command=self._guardrails_btn_vlan_untag_restore,
            state="disabled",
        )
        self.guardrails_vlan_restore_btn.pack(side="left", padx=(8, 0))
        ttk.Label(
            bar,
            text="(ACL 수동적용/제거는 항목 ⚙ / VLAN원복은 Discovery 유지 후)",
            foreground="#64748b",
        ).pack(side="left", padx=(12, 0))
        self._guardrails_sync_vlan_restore_btn()

        mid = ttk.Frame(parent)
        mid.pack(fill="both", expand=True, padx=4, pady=(0, 8))

        tree_fr = ttk.Frame(mid)
        tree_fr.pack(fill="both", expand=True)
        cols = ("pick", "ref", "scope", "summary", "config", "result")
        tree = ttk.Treeview(tree_fr, columns=cols, show="headings", selectmode="browse", takefocus=1)
        self.guardrails_list_tree = tree
        tree.heading("pick", text="선택")
        tree.column("pick", width=44, anchor="center", stretch=False)
        tree.heading("ref", text="참조")
        tree.column("ref", width=88, anchor="center", stretch=False)
        tree.heading("scope", text="범위")
        tree.column("scope", width=72, anchor="center", stretch=False)
        tree.heading("summary", text="개요")
        tree.column("summary", width=360, anchor="w", stretch=True)
        tree.heading("config", text="설정")
        tree.column("config", width=44, anchor="center", stretch=False)
        tree.heading("result", text="결과")
        tree.column("result", width=100, anchor="center", stretch=False)
        tree.tag_configure("row_even", background="#ffffff")
        tree.tag_configure("row_odd", background="#f0f4f8")
        tree.tag_configure("res_idle", foreground="#94a3b8")
        tree.tag_configure("res_run", foreground="#d97706")
        tree.tag_configure("res_pass", foreground="#15803d")
        tree.tag_configure("res_fail", foreground="#b91c1c")
        tree.tag_configure("res_stop", foreground="#64748b")

        def _on_tree_click(evt: tk.Event) -> None:
            row = tree.identify_row(evt.y)
            if not row:
                return
            col = tree.identify_column(evt.x)
            if col == "#1":
                bv = self.guardrails_check_vars.get(row)
                if bv is not None:
                    bv.set(not bv.get())
                    self._guardrails_sync_tree_row(row)
                    try:
                        self._on_any_setting_changed()
                    except Exception:
                        pass
            elif col == "#5":
                if row in self._GUARDRAILS_PER_TEST_SCHEMA:
                    self._guardrails_open_settings(row)

        def _on_tree_double(evt: tk.Event) -> None:
            row = tree.identify_row(evt.y)
            if not row:
                return
            col = tree.identify_column(evt.x)
            if col == "#5":
                return  # settings gear — single click opens
            self._guardrails_open_result_detail(row)

        tree.bind("<Button-1>", _on_tree_click)
        tree.bind("<Double-Button-1>", _on_tree_double)

        def _wheel_tree(evt: tk.Event) -> None:
            if evt.delta:
                tree.yview_scroll(int(-evt.delta / 120), "units")

        tree.bind("<MouseWheel>", _wheel_tree)
        ys = ttk.Scrollbar(tree_fr, orient="vertical", command=tree.yview)
        xs = ttk.Scrollbar(tree_fr, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=ys.set, xscrollcommand=xs.set)
        tree.grid(row=0, column=0, sticky="nsew")
        ys.grid(row=0, column=1, sticky="ns")
        xs.grid(row=1, column=0, sticky="ew")
        tree_fr.rowconfigure(0, weight=1)
        tree_fr.columnconfigure(0, weight=1)
        self._guardrails_detail_text = None
        self._guardrails_rebuild_list()

    @staticmethod
    def _guardrails_fault_alarm_rows_from_legacy_detail(full: str) -> list[tuple[str, ...]]:
        """Parse old multi-line raise/clear detail → Treeview rows."""
        rows: list[tuple[str, ...]] = []

        def _parse_wall(s: str):
            s = (s or "").strip().replace("T", " ").replace("Z", "")
            s = re.sub(r"\.\d+$", "", s)
            try:
                return datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")
            except Exception:
                return None

        def _delta(a: str, b: str, ok: str) -> str:
            if (ok or "").upper() != "OK":
                return "TIMEOUT"
            ta, tb = _parse_wall(a), _parse_wall(b)
            if ta is None or tb is None:
                return "—"
            sec = max(0.0, (tb - ta).total_seconds())
            return str(int(sec)) if sec == int(sec) else f"{sec:.1f}"

        # catalog map from "[2] ..." lines if present
        names: dict[str, tuple[str, str, str]] = {}
        for ln in (full or "").splitlines():
            m = re.search(
                r"alarm-id\s+(\d+)\s*:\s*fault-id=(\S+)\s+(\S+)\s+(.+)$",
                ln.strip(),
            )
            if m:
                names[m.group(1)] = (m.group(2), m.group(3), m.group(4).strip())

        blocks = re.split(r"(?m)^\s*alarm-id\s+(\d+)\s*(?:\([^)]*\))?\s*:\s*(PASS|FAIL)\s*$", full or "")
        # split → [pre, id, result, body, id, result, body, ...]
        i = 1
        while i + 2 <= len(blocks):
            aid = blocks[i].strip()
            fr = blocks[i + 1].strip().upper()
            body = blocks[i + 2] if i + 2 < len(blocks) else ""
            i += 3
            m_r = re.search(
                r"raise\s+noti=(\S+)\s+sent=(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\s+wall=(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})",
                body,
            )
            m_c = re.search(
                r"clear\s+noti=(\S+)\s+sent=(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})\s+wall=(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})",
                body,
            )
            rs = _delta(m_r.group(2), m_r.group(3), m_r.group(1)) if m_r else "—"
            cs = _delta(m_c.group(2), m_c.group(3), m_c.group(1)) if m_c else "—"
            fid, sev, name = names.get(aid, ("", "", ""))
            if not name:
                m_n = re.search(rf"alarm-id\s+{re.escape(aid)}\s*\(([^)]+)\)\s*:\s*{fr}", full or "")
                if m_n:
                    name = m_n.group(1).strip()
            rows.append((fr, aid, fid, sev, name, rs, cs))
        return rows

    def _guardrails_open_result_detail(self, item_id: str) -> None:
        """Popup full result (진행 로그는 Logs 창)."""
        store = getattr(self, "_guardrails_detail_by_id", None) or {}
        full = store.get(item_id) or ""
        if not full:
            rv = self.guardrails_result_vars.get(item_id)
            full = (rv.get() if rv is not None else "") or "(결과 없음)"
        pcap = store.get(f"{item_id}__pcap") or ""
        saved_txt = store.get(f"{item_id}__txt") or ""
        body = full
        if pcap:
            body = body + f"\n\nlocal pcap:\n{pcap}"
        if saved_txt:
            body = body + f"\n\nresult txt:\n{saved_txt}"

        win = tk.Toplevel(self)
        win.title(f"M-Plane Test 상세 — {item_id}")
        try:
            win.geometry(getattr(self, "guardrails_detail_win_geometry", None) or "980x560")
        except Exception:
            win.geometry("980x560")
        win.minsize(520, 300)
        fr = ttk.Frame(win, padding=8)
        fr.pack(fill="both", expand=True)

        # fault_alarm: 표(결과/발생/해지) + 아래 요약 텍스트
        result_rows: list[tuple[str, ...]] = []
        if item_id == "fault_alarm":
            for ln in (full or "").splitlines():
                if not ln.startswith("RESULT_ROW|"):
                    continue
                parts = ln.split("|")
                if len(parts) >= 8:
                    result_rows.append(
                        (
                            parts[1],
                            parts[2],
                            parts[3],
                            parts[4],
                            parts[5],
                            parts[6],
                            parts[7],
                        )
                    )
            # 구 포맷(다줄 raise/clear) 폴백
            if not result_rows:
                result_rows = self._guardrails_fault_alarm_rows_from_legacy_detail(full)

        row0 = 0
        if result_rows:
            treef = ttk.Frame(fr)
            treef.grid(row=0, column=0, columnspan=2, sticky="nsew", pady=(0, 8))
            cols = ("result", "aid", "fid", "sev", "name", "raise_s", "clear_s")
            tv = ttk.Treeview(treef, columns=cols, show="headings", height=min(18, max(6, len(result_rows))))
            tv.heading("result", text="결과")
            tv.heading("aid", text="alarm-id")
            tv.heading("fid", text="fault-id")
            tv.heading("sev", text="severity")
            tv.heading("name", text="name")
            tv.heading("raise_s", text="발생(s)")
            tv.heading("clear_s", text="해지(s)")
            tv.column("result", width=64, anchor="center", stretch=False)
            tv.column("aid", width=72, anchor="center", stretch=False)
            tv.column("fid", width=72, anchor="center", stretch=False)
            tv.column("sev", width=88, anchor="center", stretch=False)
            tv.column("name", width=280, anchor="w", stretch=True)
            tv.column("raise_s", width=72, anchor="e", stretch=False)
            tv.column("clear_s", width=72, anchor="e", stretch=False)
            tv.tag_configure("PASS", foreground="#15803d")
            tv.tag_configure("FAIL", foreground="#b91c1c")
            for r in result_rows:
                tv.insert("", "end", values=r, tags=(r[0],))
            ys_t = ttk.Scrollbar(treef, orient="vertical", command=tv.yview)
            tv.configure(yscrollcommand=ys_t.set)
            tv.grid(row=0, column=0, sticky="nsew")
            ys_t.grid(row=0, column=1, sticky="ns")
            treef.rowconfigure(0, weight=1)
            treef.columnconfigure(0, weight=1)
            row0 = 1
            # hide RESULT_ROW noise from text body
            body_lines = [
                ln
                for ln in body.splitlines()
                if not ln.startswith("RESULT_ROW|")
                and ln.strip() not in ("===RESULT_TABLE_BEGIN===", "===RESULT_TABLE_END===")
                and not ln.strip().startswith("# result|alarm_id|")
            ]
            body = "\n".join(body_lines)

        txt = tk.Text(fr, wrap="none", font=("Consolas", 10))
        ys = ttk.Scrollbar(fr, orient="vertical", command=txt.yview)
        xs = ttk.Scrollbar(fr, orient="horizontal", command=txt.xview)
        txt.configure(yscrollcommand=ys.set, xscrollcommand=xs.set)
        txt.grid(row=row0, column=0, sticky="nsew")
        ys.grid(row=row0, column=1, sticky="ns")
        xs.grid(row=row0 + 1, column=0, sticky="ew")
        fr.rowconfigure(row0, weight=1)
        fr.columnconfigure(0, weight=1)
        txt.insert("1.0", body)
        txt.configure(state="disabled")
        btn = ttk.Frame(win, padding=(8, 0, 8, 8))
        btn.pack(fill="x")
        if pcap:

            def _open_pcap_dir() -> None:
                try:
                    d = str(Path(pcap).parent)
                    if sys.platform.startswith("win"):
                        os.startfile(d)  # type: ignore[attr-defined]
                    else:
                        subprocess.Popen(["xdg-open", d])
                except Exception as exc:
                    messagebox.showerror("M-Plane Test", str(exc), parent=win)

            ttk.Button(btn, text="pcap 폴더 열기", command=_open_pcap_dir).pack(side="left")

        def _save_txt_now() -> None:
            path = self._guardrails_save_result_txt(item_id, full)
            if path:
                try:
                    if not hasattr(self, "_guardrails_detail_by_id") or self._guardrails_detail_by_id is None:
                        self._guardrails_detail_by_id = {}
                    self._guardrails_detail_by_id[f"{item_id}__txt"] = path
                except Exception:
                    pass
                messagebox.showinfo("M-Plane Test", f"저장됨\n{path}", parent=win)
            else:
                messagebox.showerror("M-Plane Test", "txt 저장 실패", parent=win)

        def _open_txt_dir() -> None:
            self._guardrails_open_local_results_folder()

        ttk.Button(btn, text="txt로 저장", command=_save_txt_now).pack(side="left", padx=(8, 0))
        ttk.Button(btn, text="결과 txt 폴더", command=_open_txt_dir).pack(side="left", padx=(8, 0))
        ttk.Button(btn, text="닫기", command=win.destroy).pack(side="right")

        def _save_geo(_evt: tk.Event | None = None) -> None:
            try:
                self.guardrails_detail_win_geometry = win.geometry()
            except Exception:
                pass

        win.bind("<Configure>", _save_geo)

    def _guardrails_selected_item_id(self) -> str:
        tree = getattr(self, "guardrails_list_tree", None)
        if tree is not None:
            sel = tree.selection()
            if sel:
                return str(sel[0])
        for i in self._guardrails_catalog():
            bv = self.guardrails_check_vars.get(i["id"])
            if bv is not None and bv.get():
                return i["id"]
        return "dhcp_v4"

    def _guardrails_btn_apply_acl(self) -> None:
        messagebox.showinfo(
            "M-Plane Test",
            "ACL 수동 적용은 항목 ⚙ 설정 창의「ACL만 적용(테스트)」를 사용하세요.\n"
            "정상 시험은「선택 항목 검증 실행」이 자동으로 ACL을 적용/원복합니다.",
        )

    def _guardrails_btn_remove_acl(self) -> None:
        messagebox.showinfo(
            "M-Plane Test",
            "ACL 수동 제거는 항목 ⚙ 설정 창의「ACL만 제거」를 사용하세요.",
        )

    def _guardrails_btn_probe(self) -> None:
        iid = self._guardrails_selected_item_id()
        self._guardrails_settings_item_id = iid
        self._guardrails_fill_defaults_from_context(iid)
        if iid in self._GUARDRAILS_DHCP_ITEM_IDS:
            fam = self._guardrails_resolve_ssh_family(iid)
        else:
            fam = self._guardrails_item_family(iid)

        def work() -> None:
            up, detail = self._guardrails_probe_once(fam)
            msg = f"[{fam}] {'HEALTHY' if up else 'UNHEALTHY'}: {detail}"
            self.after(0, self.append_log, f"[M-Plane Test] SSH check → {msg}\n")
            self.after(0, lambda: messagebox.showinfo("M-Plane Test Probe", msg))

        threading.Thread(target=work, daemon=True).start()

    def _guardrails_local_pcap_dir(self) -> Path:
        d = self.config_path.parent / "pcaps"
        d.mkdir(parents=True, exist_ok=True)
        return d

    def _guardrails_local_results_dir(self) -> Path:
        d = self.config_path.parent / "mplane_results"
        d.mkdir(parents=True, exist_ok=True)
        return d

    def _guardrails_save_result_txt(
        self,
        item_id: str,
        detail: str,
        *,
        iteration: int | None = None,
    ) -> str:
        """Save M-Plane Test detail text next to pcaps (AppData …/mplane_results/)."""
        try:
            d = self._guardrails_local_results_dir()
            stamp = time.strftime("%Y%m%d_%H%M%S")
            safe = re.sub(r"[^\w.\-]+", "_", str(item_id or "item")).strip("_") or "item"
            iter_part = f"_r{iteration}" if iteration is not None and iteration != 1 else ""
            path = d / f"mplane_{safe}{iter_part}_{stamp}.txt"
            title = ""
            try:
                for it in self._guardrails_catalog():
                    if it.get("id") == item_id:
                        title = str(it.get("title") or it.get("ref") or "")
                        break
            except Exception:
                title = ""
            header = [
                f"M-Plane Test 상세결과",
                f"item_id : {item_id}",
            ]
            if title:
                header.append(f"title   : {title}")
            header.append(f"saved   : {time.strftime('%Y-%m-%d %H:%M:%S')}")
            if iteration is not None:
                header.append(f"repeat  : {iteration}")
            header.append("=" * 60)
            body = (detail or "").rstrip() + "\n"
            path.write_text("\n".join(header) + "\n" + body, encoding="utf-8")
            return str(path)
        except Exception as exc:
            self._guardrails_log(f"결과 txt 저장 실패: {exc}")
            return ""

    def _guardrails_open_local_pcap_folder(self) -> None:
        d = self._guardrails_local_pcap_dir()
        try:
            if sys.platform.startswith("win"):
                os.startfile(str(d))  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.Popen(["open", str(d)])
            else:
                subprocess.Popen(["xdg-open", str(d)])
        except Exception as exc:
            messagebox.showerror("M-Plane Test", f"폴더 열기 실패:\n{d}\n{exc}")

    def _guardrails_open_local_results_folder(self) -> None:
        d = self._guardrails_local_results_dir()
        try:
            if sys.platform.startswith("win"):
                os.startfile(str(d))  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.Popen(["open", str(d)])
            else:
                subprocess.Popen(["xdg-open", str(d)])
        except Exception as exc:
            messagebox.showerror("M-Plane Test", f"폴더 열기 실패:\n{d}\n{exc}")

    def _guardrails_rebuild_list(self) -> None:
        tree = getattr(self, "guardrails_list_tree", None)
        if tree is None:
            return
        for child in tree.get_children():
            tree.delete(child)
        self.guardrails_check_vars.clear()
        # keep result text across rebuild if possible
        prev_results = {k: v.get() for k, v in self.guardrails_result_vars.items()}
        self.guardrails_result_vars.clear()

        for idx, item in enumerate(self._guardrails_catalog()):
            iid = item["id"]
            bv = tk.BooleanVar(value=True)
            rv = tk.StringVar(value=prev_results.get(iid) or "—")
            self.guardrails_check_vars[iid] = bv
            self.guardrails_result_vars[iid] = rv
            has_cfg = iid in self._GUARDRAILS_PER_TEST_SCHEMA
            sk = self._guardrails_store_key(iid) if has_cfg else iid
            store = getattr(self, "_guardrails_per_test_settings", {}) or {}
            stored = store.get(sk) or store.get(iid) or {}
            cfg_mark = "⚙✓" if (has_cfg and stored) else ("⚙" if has_cfg else "")
            result = rv.get() or "—"
            short = result if len(result) <= 48 else result[:45] + "..."
            parity = "row_odd" if idx % 2 else "row_even"
            tag = self._guardrails_result_tag(result)
            tree.insert(
                "",
                "end",
                iid=iid,
                values=(
                    "☑",
                    item.get("ref", iid),
                    item.get("scope", ""),
                    item.get("title", iid),
                    cfg_mark,
                    short,
                ),
                tags=(parity, tag),
            )

    def _guardrails_select_all(self) -> None:
        for iid, bv in self.guardrails_check_vars.items():
            bv.set(True)
            self._guardrails_sync_tree_row(iid)
        self._on_any_setting_changed()

    def _guardrails_clear_all(self) -> None:
        for iid, bv in self.guardrails_check_vars.items():
            bv.set(False)
            self._guardrails_sync_tree_row(iid)
        self._on_any_setting_changed()

    def _guardrails_stop(self) -> None:
        self._guardrails_cancel.set()
        # ORU reset(v6) 헬퍼는 conformance_cancel 만 보고 있었음 → 중지가 안 먹힘
        try:
            self._conformance_cancel_event.set()
        except Exception:
            pass
        try:
            with self._conformance_run_transport_lock:
                ch = getattr(self, "_conformance_run_script_channel", None)
            if ch is not None:
                try:
                    ch.close()
                except Exception:
                    pass
        except Exception:
            pass
        self._guardrails_log("검증 중지 요청 — 대기 중 작업 건너뛰고 빠른 정리로 종료합니다")

    def _guardrails_parse_repeat_count(self) -> int | None:
        """Return repeat count: 0=infinite, 1+=finite. None if invalid."""
        raw = (self.guardrails_run_repeat_var.get() or "").strip()
        if not raw:
            return 1
        try:
            n = int(raw)
        except ValueError:
            return None
        if n < 0:
            return None
        return n

    def _guardrails_run_checked(self) -> None:
        if self.guardrails_busy:
            messagebox.showwarning("M-Plane Test", "이미 검증 중입니다.")
            return
        repeat_count = self._guardrails_parse_repeat_count()
        if repeat_count is None:
            messagebox.showwarning("M-Plane Test", "반복 횟수는 0(무한) 이상의 정수로 입력하세요.")
            return
        selected = [
            i["id"]
            for i in self._guardrails_catalog()
            if self.guardrails_check_vars.get(i["id"], tk.BooleanVar(value=False)).get()
        ]
        if not selected:
            messagebox.showwarning("M-Plane Test", "검증할 항목을 하나 이상 선택하세요.")
            return
        dhcp_sel = [i for i in selected if i in self._GUARDRAILS_DHCP_ITEM_IDS]
        if dhcp_sel:
            for iid in dhcp_sel:
                self._guardrails_settings_item_id = iid
                self._guardrails_fill_defaults_from_context(iid)
                fam = self._guardrails_item_family(iid)
                mode = self._guardrails_item_mode(iid)
                if self._guardrails_gf("dhcp_host") and not self._guardrails_gf("dhcp_id"):
                    messagebox.showwarning("M-Plane Test", "DHCP SSH host가 있으면 DHCP SSH ID도 필요합니다.")
                    self._guardrails_open_settings(iid)
                    return
                if mode == "vlan":
                    if not self._guardrails_gf("l2sw_ip") or not self._guardrails_gf("l2sw_id"):
                        messagebox.showwarning("M-Plane Test", "VLAN Discovery: ⚙ L2SW IP/ID 필요.")
                        self._guardrails_open_settings(iid)
                        return
                    if not (self._guardrails_gf("vlan_discovery_vid") or "").strip():
                        messagebox.showwarning("M-Plane Test", "VLAN Discovery: ⚙ ★ 시험 VLAN ID 필요.")
                        self._guardrails_open_settings(iid)
                        return
                # Capture IF / L2SW IF / RU MAC 은 실행 시 자동 조회
                host, how = self._guardrails_ru_ssh_target(fam)
                if not host:
                    messagebox.showwarning(
                        "M-Plane Test",
                        f"[{fam}] SSH 대상 불가: {how}\n⚙ 확인.",
                    )
                    self._guardrails_open_settings(iid)
                    return
                if not self._guardrails_gf("oru_cli_id"):
                    messagebox.showwarning("M-Plane Test", "RU SSH ID가 필요합니다 (Settings ★ RU SSH ID).")
                    self._guardrails_open_settings(iid)
                    return
                if fam == "v6":
                    try:
                        self._apply_lab_controller_listen_ips("untag")
                    except Exception:
                        pass
                    lv6 = ""
                    av6 = ""
                    try:
                        lv6 = (self.fields.get("LOCAL_IP_V6").get() or "").strip()  # type: ignore[union-attr]
                        av6 = (self.fields.get("ALLOWED_IP_V6").get() or "").strip()  # type: ignore[union-attr]
                    except Exception:
                        pass
                    if not lv6:
                        lv6 = self._lab_controller_listen_ip("v6", "untag")
                    rmode = (self._guardrails_gf("reset_mode", "auto") or "auto").strip().lower()
                    if rmode not in ("0", "manual", "prompt", "none", "off") and (
                        not lv6 or not av6 or ":" not in lv6 or ":" not in av6
                    ):
                        messagebox.showwarning(
                            "M-Plane Test",
                            f"{iid} M-Plane reset(IPv6)에는 Settings의\n"
                            "★ ALLOWED_IP_V6 (global IPv6)가 필요합니다.\n"
                            "(LOCAL_IP_V6는 untag/tag 자동)\n"
                            "또는 항목 ⚙ Reset mode=manual.",
                        )
                        return
            labels = [i for i in self._guardrails_catalog() if i["id"] in dhcp_sel]
            label_txt = ", ".join(x.get("ref") or x["id"] for x in labels)
            tips = []
            boot_sel = [i for i in dhcp_sel if self._guardrails_item_mode(i) == "boot"]
            vlan_dhcp_sel = [i for i in dhcp_sel if self._guardrails_item_mode(i) == "vlan"]
            if boot_sel:
                tips.append(
                    "Boot: 재부팅 → 주소 복구 + tcpdump Option → pcap을 LOG_PATH 저장"
                )
            if vlan_dhcp_sel:
                tips.append(
                    "VLAN Discovery: ACL + vlan/trunk → Discovery → renew/원복 + Option tcpdump"
                )
            if repeat_count == 0:
                tips.append("반복: 무한(중지 버튼으로 종료)")
            elif repeat_count > 1:
                tips.append(f"반복: {repeat_count}회")
            if not messagebox.askokcancel(
                "M-Plane Test",
                f"선택: {label_txt}\n\n" + "\n".join(f"· {t}" for t in tips) + "\n\n실행할까요?",
            ):
                return
        elif repeat_count == 0 or repeat_count > 1:
            rep_txt = "무한" if repeat_count == 0 else f"{repeat_count}회"
            if not messagebox.askokcancel(
                "M-Plane Test",
                f"선택 {len(selected)}항목 · 반복 {rep_txt}\n\n실행할까요?",
            ):
                return
        self.guardrails_busy = True
        self._guardrails_cancel.clear()
        try:
            self.guardrails_stop_btn.configure(state="normal")
        except tk.TclError:
            pass
        try:
            self.open_log_window()
        except Exception:
            pass
        rep_note = "무한" if repeat_count == 0 else str(repeat_count)
        self._guardrails_log(
            f"검증 시작: {', '.join(selected)}  반복={rep_note}  "
            "(진행은 Logs의 Live Output에 표시됩니다. Log Load는 원격 *.log 전용)"
        )
        try:
            self._on_any_setting_changed()
        except Exception:
            pass

        def worker() -> None:
            iteration = 0
            while True:
                iteration += 1
                if self._guardrails_cancel.is_set():
                    self._guardrails_log("사용자 중지")
                    break
                if repeat_count == 0:
                    self._guardrails_log(f"=== M-Plane Test 반복 {iteration} (0=무한) ===")
                elif repeat_count > 1:
                    self._guardrails_log(f"=== M-Plane Test 반복 {iteration}/{repeat_count} ===")

                aborted = False
                for iid in selected:
                    if self._guardrails_cancel.is_set():
                        self._guardrails_log("사용자 중지")
                        aborted = True
                        break
                    self.after(
                        0,
                        lambda i=iid: self.guardrails_result_vars.get(i)
                        and self.guardrails_result_vars[i].set("RUN…"),
                    )
                    self.after(0, lambda i=iid: self._guardrails_sync_tree_row(i))
                    tag = (
                        f"{iid} [{iteration}]"
                        if repeat_count != 1
                        else iid
                    )
                    self._guardrails_log(f"—— 실행 중: {tag} ——")
                    st, detail = self._guardrails_run_one(iid)
                    # 상세 팝업은 줄바꿈 유지 (한 줄 | 로 합치지 않음)
                    full = detail if (detail or "").strip() else st
                    if not full.lstrip().startswith(st):
                        full = f"{st}\n{full}"
                    if repeat_count != 1:
                        suffix = f"\n(반복 {iteration}" + (
                            "" if repeat_count == 0 else f"/{repeat_count}"
                        ) + ")"
                        full = full + suffix
                    try:
                        if not hasattr(self, "_guardrails_detail_by_id") or self._guardrails_detail_by_id is None:
                            self._guardrails_detail_by_id = {}
                        self._guardrails_detail_by_id[iid] = full
                    except Exception:
                        pass
                    saved = self._guardrails_save_result_txt(
                        iid,
                        full,
                        iteration=iteration if repeat_count != 1 else None,
                    )
                    if saved:
                        try:
                            self._guardrails_detail_by_id[f"{iid}__txt"] = saved
                        except Exception:
                            pass
                        self._guardrails_log(f"{tag}: 상세결과 txt 저장 → {saved}")
                    # 목록 표시용은 판정만; 전체 문구는 detail store + Logs
                    one = st
                    if repeat_count != 1:
                        one = f"{st} ({iteration}" + ("" if repeat_count == 0 else f"/{repeat_count}") + ")"
                    rv = self.guardrails_result_vars.get(iid)
                    if rv is not None:
                        self.after(0, rv.set, one)
                    self.after(0, lambda i=iid: self._guardrails_sync_tree_row(i))
                    # Logs 는 첫 줄 요약 + 나머지는 들여쓰기
                    d_first = (detail or "").splitlines()[0] if detail else ""
                    self._guardrails_log(f"{tag} → {st}: {d_first}")
                    for dline in (detail or "").splitlines()[1:]:
                        if dline.strip():
                            self._guardrails_log(f"  {dline}")

                if aborted or self._guardrails_cancel.is_set():
                    break
                if repeat_count == 1:
                    break
                if repeat_count > 1 and iteration >= repeat_count:
                    break
                self._guardrails_log(
                    f"다음 반복 준비 (완료 {iteration}"
                    + ("" if repeat_count == 0 else f"/{repeat_count}")
                    + ")"
                )

            def _done() -> None:
                self.guardrails_busy = False
                try:
                    self.guardrails_stop_btn.configure(state="disabled")
                except tk.TclError:
                    pass
                self._save_current_config()
                self._guardrails_log("검증 종료")

            self.after(0, _done)

        threading.Thread(target=worker, daemon=True).start()

    def _load_saved_config(self) -> None:
        self.append_log(f"[GUI] Config file: {self.config_path}\n")
        if not self.config_path.exists():
            self.append_log("[GUI] No saved config yet — using defaults.\n")
            return
        self._config_hydrating = True
        try:
            saved = json.loads(self.config_path.read_text(encoding="utf-8"))
            if isinstance(saved, dict):
                self._apply_config(saved)
                self.append_log(f"[GUI] Loaded last config: {self.config_path.name}\n")
        except Exception as exc:
            self.append_log(f"[GUI] Failed to load config: {exc}\n")
        finally:
            self._config_hydrating = False

    def _save_current_config(self) -> None:
        self._config_save_job = None
        if getattr(self, "_config_hydrating", False):
            return
        try:
            payload = self._collect_current_config()
            self.config_path.parent.mkdir(parents=True, exist_ok=True)
            self.config_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
            cl = getattr(self, "_conformance_last_run_snapshot_cache", None)
            if (
                hasattr(self, "conformance_last_run_hint_var")
                and isinstance(cl, dict)
                and cl.get("by_script")
            ):
                ts = str(cl.get("saved_at") or "").strip()
                summ = cl.get("summary")
                if not isinstance(summ, dict):
                    bs = cl.get("by_script")
                    summ = (
                        self._conformance_summarize_pass_fail_counts(bs)
                        if isinstance(bs, dict)
                        else {}
                    )
                cnt = self._conformance_format_pass_fail_counts_ko(summ) if isinstance(summ, dict) else ""
                if ts and cnt:
                    self.conformance_last_run_hint_var.set(
                        f"Conformance 마지막 결과가 설정 파일에 저장되었습니다. — {cnt} (기록 시각: {ts})"
                    )
                elif ts:
                    self.conformance_last_run_hint_var.set(
                        f"Conformance 마지막 결과가 설정 파일에 저장되었습니다. (기록 시각: {ts})"
                    )
        except Exception as exc:
            self.append_log(f"[GUI] Failed to save config: {exc}\n")

    def _resolve_auth_mode(self, ssh_password: str) -> tuple[str, str | None]:
        if not ssh_password:
            return "openssh", None
        try:
            import paramiko  # type: ignore  # noqa: F401
            return "paramiko", None
        except Exception:
            pass
        if shutil.which("sshpass") is not None:
            return "sshpass", None
        if shutil.which("plink") is not None and shutil.which("pscp") is not None:
            return "putty", "sshpass not found; using PuTTY plink/pscp."
        return "openssh", (
            "SSH_PASSWORD is set but no password backend found.\n"
            "Install paramiko OR sshpass OR PuTTY(plink, pscp), or use SSH_KEY_PATH."
        )

    def fetch_latest_remote_log(self) -> None:
        self._save_current_config()
        self.append_log("[GUI] Fetching latest remote log...\n")
        threading.Thread(target=self._fetch_latest_remote_log_worker, args=(False,), daemon=True).start()

    def load_full_remote_log(self) -> None:
        self._save_current_config()
        self.append_log("[GUI] Loading full remote log...\n")
        threading.Thread(target=self._fetch_latest_remote_log_worker, args=(True,), daemon=True).start()

    def _fetch_latest_remote_log_worker(self, full_load: bool = False) -> None:
        t0 = time.perf_counter()
        ssh_user = self.remote_user.get().strip()
        ssh_host = self.remote_host.get().strip()
        ssh_port = self.remote_port.get().strip()
        ssh_password = self.remote_password.get().strip()
        key_path = self.remote_key_path.get().strip()
        log_path = self.fields.get("LOG_PATH").get().strip()
        lines = self.fetch_lines_var.get().strip() or "500"

        if not ssh_user or not ssh_host or not ssh_port or not log_path:
            self.after(0, lambda: messagebox.showerror("Error", "SSH info and LOG_PATH are required."))
            return
        auth_mode, info_or_error = self._resolve_auth_mode(ssh_password)
        if info_or_error and "Install" in info_or_error:
            self.after(0, lambda: messagebox.showerror("Error", info_or_error))
            return
        if info_or_error:
            self.after(0, self.append_log, f"[GUI] {info_or_error}\n")

        if auth_mode == "paramiko":
            ok, text = self._fetch_latest_log_paramiko(
                ssh_user, ssh_host, ssh_port, ssh_password, key_path, log_path, lines, full_load
            )
        else:
            ok, text = self._fetch_latest_log_cli(
                ssh_user, ssh_host, ssh_port, ssh_password, key_path, auth_mode, log_path, lines, full_load
            )
        if ok:
            if full_load:
                # Keep current GUI logs intact, and open fetched full log in Notepad.
                self.after(0, self._open_text_in_notepad, text)
                self.after(0, self.append_log, "[GUI] Log Load complete. Opened in Notepad.\n")
            else:
                self.after(0, self._load_log_text_direct, text)
        else:
            self.after(0, lambda: messagebox.showerror("Error", text))
        self._perf_record("fetch_remote_log", time.perf_counter() - t0)

    def _fetch_latest_log_paramiko(
        self,
        ssh_user: str,
        ssh_host: str,
        ssh_port: str,
        ssh_password: str,
        key_path: str,
        log_path: str,
        lines: str,
        full_load: bool = False,
    ) -> tuple[bool, str]:
        try:
            import paramiko  # type: ignore
            client = paramiko.SSHClient()
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            client.connect(
                hostname=ssh_host,
                port=int(ssh_port),
                username=ssh_user,
                password=ssh_password if ssh_password else None,
                key_filename=key_path if key_path else None,
                timeout=8,
                auth_timeout=8,
                banner_timeout=8,
                look_for_keys=not bool(ssh_password),
                allow_agent=True,
            )
            if full_load:
                cmd = (
                    f"LATEST=$(ls -1t {shlex.quote(log_path)}/*.log 2>/dev/null | head -n 1); "
                    f'if [ -z "$LATEST" ]; then echo "__NO_LOG__"; else echo "__LOG__:$LATEST"; cat "$LATEST"; fi'
                )
            else:
                cmd = (
                    f"LATEST=$(ls -1t {shlex.quote(log_path)}/*.log 2>/dev/null | head -n 1); "
                    f'if [ -z "$LATEST" ]; then echo "__NO_LOG__"; else echo "__LOG__:$LATEST"; tail -n {shlex.quote(lines)} "$LATEST"; fi'
                )
            _, stdout, stderr = client.exec_command(cmd)
            out = stdout.read().decode(errors="ignore")
            err = stderr.read().decode(errors="ignore").strip()
            client.close()
            if "__NO_LOG__" in out:
                return False, f"No .log file found in LOG_PATH: {log_path}"
            if "__LOG__:" in out:
                return True, out.split("\n", 1)[1] if "\n" in out else out
            return False, err or out or "Unknown error while reading remote log."
        except Exception as exc:
            return False, f"Remote log fetch failed: {exc}"

    def _fetch_latest_log_cli(
        self,
        ssh_user: str,
        ssh_host: str,
        ssh_port: str,
        ssh_password: str,
        key_path: str,
        auth_mode: str,
        log_path: str,
        lines: str,
        full_load: bool = False,
    ) -> tuple[bool, str]:
        use_password = bool(ssh_password)
        ssh_cmd = self._ssh_base_cmd(ssh_user, ssh_host, ssh_port, key_path, use_password, auth_mode, ssh_password)
        run_env = os.environ.copy()
        if use_password and auth_mode == "sshpass":
            run_env["SSHPASS"] = ssh_password
        if full_load:
            cmd = (
                f"LATEST=$(ls -1t {shlex.quote(log_path)}/*.log 2>/dev/null | head -n 1); "
                f'if [ -z "$LATEST" ]; then echo "__NO_LOG__"; else echo "__LOG__:$LATEST"; cat "$LATEST"; fi'
            )
        else:
            cmd = (
                f"LATEST=$(ls -1t {shlex.quote(log_path)}/*.log 2>/dev/null | head -n 1); "
                f'if [ -z "$LATEST" ]; then echo "__NO_LOG__"; else echo "__LOG__:$LATEST"; tail -n {shlex.quote(lines)} "$LATEST"; fi'
            )
        result = subprocess.run(ssh_cmd + [cmd], capture_output=True, text=True, env=run_env)
        out = result.stdout or ""
        err = (result.stderr or "").strip()
        if "__NO_LOG__" in out:
            return False, f"No .log file found in LOG_PATH: {log_path}"
        if "__LOG__:" in out:
            return True, out.split("\n", 1)[1] if "\n" in out else out
        return False, err or out or "Unknown error while reading remote log."

    def _set_running(self, running: bool) -> None:
        self.is_running = running
        self.start_btn.config(state="disabled" if running else "normal")
        self.stop_btn.config(state="normal" if running else "disabled")
        if running:
            self.status_var.set("Running")
            self._transport_reconnect_pending = False
            self._session_lost_logged = False
        else:
            self.session_established = False
            self.manual_send_ready = False
            self.auto_xml_send_done = False
            self._conformance_auto_sync_scheduled = False
            self._cancel_session_watch()
            if self._user_stop_requested:
                self.status_var.set("Idle")
            elif self.auto_start_var.get():
                self.status_var.set("Reconnecting...")
                self.start_btn.config(state="disabled")
                self.stop_btn.config(state="normal")
                self.after(400, self._enqueue_auto_restart_if_configured)
            else:
                self.status_var.set("Idle")
        self._sync_manual_send_widgets()
        self.after(0, self._refresh_log_target_hint_line)

    def manual_start_script(self) -> None:
        self._conformance_dbg("manual_start_script: 호출 (_user_stop_requested 해제 후 start_script)")
        self._user_stop_requested = False
        self._cancel_auto_start_retry_job()
        self.start_script()

    def start_script(self) -> None:
        self._cancel_auto_start_retry_job()
        if self.proc is not None and self.proc.poll() is None:
            messagebox.showinfo("Info", "Script is already running.")
            return
        pch = self.paramiko_channel
        if pch is not None and not getattr(pch, "closed", False):
            messagebox.showinfo("Info", "Script is already running.")
            return

        self._save_current_config()
        # Start/CallHome listen = lab untag controller
        try:
            self._apply_lab_controller_listen_ips("untag")
        except Exception:
            pass
        self.remote_cfg_cache.clear()
        env = os.environ.copy()
        field_values: dict[str, str] = {}
        for key, var in self.fields.items():
            value = var.get().strip()
            env[key] = value
            field_values[key] = value

        self._start_log_path_hint = field_values.get("LOG_PATH", "").strip()
        self.after(0, self._refresh_log_target_hint_line)

        self.stop_event.clear()
        self.session_established = False
        self.manual_send_ready = False
        self.auto_xml_send_done = False
        self._remote_start_failed_handled = False
        self._conformance_auto_sync_scheduled = False
        self._transport_reconnect_pending = False
        self._session_lost_logged = False
        self._session_lost_at_mono = 0.0
        with self.log_lock:
            self._recent_log_for_session = ""
        self._set_running(True)
        self.status_var.set("Preparing...")
        self.append_log("[GUI] BUILD_MARKER: start_script_v20260512_1725\n")
        self.append_log("\n[GUI] Preparing execution...\n")
        threading.Thread(
            target=self._start_script_worker,
            args=(env, field_values),
            daemon=True,
        ).start()
        self._schedule_session_watch()

    def _cancel_auto_start_retry_job(self) -> None:
        jid = self._auto_start_retry_job
        self._auto_start_retry_job = None
        if jid is not None:
            try:
                self.after_cancel(jid)
            except Exception:
                pass

    def _enqueue_auto_restart_if_configured(self) -> None:
        if self._conformance_run_busy or self._conformance_stop_idle_wait:
            return
        if self.is_running:
            return
        if self._user_stop_requested:
            return
        if getattr(self, "_remote_start_failed_handled", False):
            return
        if not self.auto_start_var.get():
            self.append_log("[GUI] Auto-reconnect disabled (uncheck setting or click Stop).\n")
            self.status_var.set("Idle")
            self.start_btn.config(state="normal")
            self.stop_btn.config(state="disabled")
            return
        if self.proc is not None and self.proc.poll() is None:
            return
        ch = self.paramiko_channel
        if ch is not None and not getattr(ch, "closed", False):
            return
        self._cancel_auto_start_retry_job()
        sec = max(1, self._auto_restart_delay_ms // 1000)
        self.append_log(f"[GUI] Auto-reconnect: retry Start in {sec}s (Stop to cancel).\n")
        self._auto_start_retry_job = self.after(self._auto_restart_delay_ms, self._fire_auto_start_retry_tick)

    def _fire_auto_start_retry_tick(self) -> None:
        self._auto_start_retry_job = None
        if not self.auto_start_var.get():
            return
        if self._user_stop_requested:
            return
        if self._conformance_run_busy or self._conformance_stop_idle_wait:
            return
        if self.is_running:
            return
        if self.proc is not None and self.proc.poll() is None:
            return
        pch = self.paramiko_channel
        if pch is not None and not getattr(pch, "closed", False):
            return
        self.append_log("[GUI] Auto-reconnect: retrying Start...\n")
        self.start_script()

    def _on_auto_start_toggled(self) -> None:
        self._save_current_config()
        if not self.auto_start_var.get():
            self._cancel_auto_start_retry_job()

    def _start_script_worker(self, env: dict[str, str], field_values: dict[str, str]) -> None:
        try:
            ssh_user = self.remote_user.get().strip()
            ssh_host = self.remote_host.get().strip()
            ssh_port = self.remote_port.get().strip()
            ssh_password = self.remote_password.get().strip()
            remote_script = self.remote_script_path.get().strip()
            key_path = self.remote_key_path.get().strip()

            if not ssh_user or not ssh_host or not ssh_port or not remote_script:
                self.after(
                    0,
                    lambda: messagebox.showerror(
                        "Error", "SSH_USER, SSH_HOST, SSH_PORT, REMOTE_SCRIPT_PATH are required."
                    ),
                )
                self.after(0, self._set_running, False)
                return

            auth_mode, info_or_error = self._resolve_auth_mode(ssh_password)
            if info_or_error and "Install" in info_or_error:
                self.after(0, lambda: messagebox.showerror("Error", info_or_error))
                self.after(0, self._set_running, False)
                return
            if ssh_password and auth_mode == "paramiko":
                self.after(0, self.append_log, "[GUI] Using paramiko password auth.\n")
            if info_or_error:
                self.after(0, self.append_log, f"[GUI] {info_or_error}\n")

            remote_log_path = field_values.get("LOG_PATH", "").strip()
            if not remote_log_path:
                self.after(0, lambda: messagebox.showerror("Error", "LOG_PATH is required."))
                self.after(0, self._set_running, False)
                return

            ok, info = self._ensure_remote_script(
                ssh_user,
                ssh_host,
                ssh_port,
                key_path,
                ssh_password,
                auth_mode,
                remote_script,
                remote_log_path,
            )
            self.after(0, self.append_log, f"[GUI] {info}\n")
            if not ok:
                self.after(0, lambda: messagebox.showerror("Error", info))
                self.after(0, self._set_running, False)
                return

            # 이전 miniDU/netopeer·stale lock 정리 후 Start (flock 충돌·sleep infinity 방지)
            self._cleanup_remote_daemons_blocking(
                reason="pre-start", include_start_script=True, timeout_s=12.0
            )

            exports = " ".join(
                f"{k}={shlex.quote(v)}" for k, v in field_values.items() if v != ""
            )
            remote_cmd = f"{exports} exec bash {shlex.quote(remote_script)}"
            ch_port_probe = (field_values.get("CALLHOME_PORT") or field_values.get("PORT") or "4334").strip() or "4334"
            product_probe = (field_values.get("PRODUCT") or "").strip()
            use_password = bool(ssh_password)
            if auth_mode == "paramiko":
                self.after(0, self.append_log, f"\n[GUI] Started (paramiko): {ssh_user}@{ssh_host}\n")
                self.after(0, self.append_log, "[POSTCHECK] scheduled (paramiko start).\n")
                threading.Thread(
                    target=self._start_postcheck_worker,
                    args=(ssh_user, ssh_host, ssh_port, ssh_password, key_path, remote_script, remote_log_path, ch_port_probe, product_probe),
                    daemon=True,
                ).start()
                self._start_remote_paramiko(
                    ssh_user=ssh_user,
                    ssh_host=ssh_host,
                    ssh_port=ssh_port,
                    ssh_password=ssh_password,
                    key_path=key_path,
                    remote_cmd=remote_cmd,
                )
                return
            cmd = self._ssh_base_cmd(
                ssh_user, ssh_host, ssh_port, key_path, use_password, auth_mode, ssh_password
            ) + [remote_cmd]
            if use_password and auth_mode == "sshpass":
                env = env.copy()
                env["SSHPASS"] = ssh_password

            self.proc = subprocess.Popen(
                cmd,
                stdin=subprocess.PIPE,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                env=env,
                creationflags=subprocess.CREATE_NEW_PROCESS_GROUP if os.name == "nt" else 0,
            )
            self.after(0, self.append_log, f"\n[GUI] Started: {' '.join(cmd)}\n")
            self.after(0, lambda: self.status_var.set("Running"))
            threading.Thread(target=self._read_output, daemon=True).start()
            threading.Thread(target=self._watch_process, daemon=True).start()
            self.after(0, self.append_log, "[POSTCHECK] scheduled (subprocess start).\n")
            threading.Thread(
                target=self._start_postcheck_worker,
                args=(ssh_user, ssh_host, ssh_port, ssh_password, key_path, remote_script, remote_log_path, ch_port_probe, product_probe),
                daemon=True,
            ).start()
        except Exception as exc:
            self.after(0, lambda: messagebox.showerror("Error", f"Failed to start script:\n{exc}"))
            self.after(0, self._set_running, False)

    def _start_postcheck_worker(
        self,
        ssh_user: str,
        ssh_host: str,
        ssh_port: str,
        ssh_password: str,
        key_path: str,
        remote_script: str,
        remote_log_path: str,
        ch_port: str,
        product: str,
    ) -> None:
        # Start 직후 무출력/정지처럼 보일 때 원격 상태를 즉시 보여준다.
        self.after(0, self.append_log, "[POSTCHECK] worker started.\n")
        time.sleep(2)
        try:
            import paramiko  # type: ignore
        except Exception:
            self.after(0, self.append_log, "[POSTCHECK] paramiko import failed.\n")
            return
        client: Any | None = None
        try:
            self.after(0, self.append_log, "[POSTCHECK] connecting...\n")
            client = paramiko.SSHClient()
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            client.connect(
                hostname=ssh_host,
                port=int(ssh_port),
                username=ssh_user,
                password=ssh_password if ssh_password else None,
                key_filename=key_path if key_path else None,
                timeout=8,
                auth_timeout=8,
                banner_timeout=8,
                look_for_keys=not bool(ssh_password),
                allow_agent=True,
            )
            self.after(0, self.append_log, "[POSTCHECK] connected. probing remote state...\n")
            log_dir = (remote_log_path or "/var/tmp/log").rstrip("/")
            pat = f"{product}_*.log" if product else "*.log"
            cmd = (
                f"echo '[POSTCHECK] script={shlex.quote(remote_script)}'; "
                f"echo '[POSTCHECK] process:'; pgrep -af {shlex.quote(Path(remote_script).name)} || true; "
                f"echo '[POSTCHECK] listen:{ch_port}'; (ss -tnlp | grep ':{ch_port} ' || true); "
                f"echo '[POSTCHECK] recent log:'; "
                f"(ls -1t {shlex.quote(log_dir)}/{pat} 2>/dev/null | head -n 1 | xargs -r tail -n 25) || true"
            )
            _i, o, e = client.exec_command(cmd)
            out = o.read().decode(errors="replace")
            err = e.read().decode(errors="replace")
            blob = out.strip()
            if err.strip():
                blob += ("\n" if blob else "") + f"[POSTCHECK][stderr] {err.strip()}"
            if blob:
                self.after(0, self.append_log, f"{blob}\n")
        except Exception as exc:
            self.after(0, self.append_log, f"[POSTCHECK] failed: {exc}\n")
        finally:
            try:
                if client is not None:
                    client.close()
            except Exception:
                pass

    def _start_remote_paramiko(
        self,
        ssh_user: str,
        ssh_host: str,
        ssh_port: str,
        ssh_password: str,
        key_path: str,
        remote_cmd: str,
    ) -> None:
        self.after(0, self.append_log, "[GUI][paramiko] ENTER _start_remote_paramiko v20260512_1725\n")
        try:
            import paramiko  # type: ignore
        except Exception:
            self.after(0, lambda: messagebox.showerror("Error", "paramiko is not installed. Run: py -m pip install paramiko"))
            self.after(0, self._set_running, False)
            return

        try:
            self.after(0, self.append_log, "[GUI][paramiko] connect starting...\n")
            client = paramiko.SSHClient()
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            client.connect(
                hostname=ssh_host,
                port=int(ssh_port),
                username=ssh_user,
                password=ssh_password if ssh_password else None,
                key_filename=key_path if key_path else None,
                timeout=8,
                auth_timeout=8,
                banner_timeout=8,
                look_for_keys=not bool(ssh_password),
                allow_agent=True,
            )
            self.after(0, self.append_log, "[GUI][paramiko] connect ok, opening shell...\n")
            transport = client.get_transport()
            if transport is None:
                raise RuntimeError("SSH transport is unavailable")
            transport.set_keepalive(20)

            # Use interactive shell channel for long-running script stability.
            channel = client.invoke_shell(width=200, height=50)
            self.after(0, self.append_log, "[GUI][paramiko] shell opened, sending start command...\n")
            channel.send(remote_cmd + "\n")
            self.after(0, self.append_log, "[GUI][paramiko] start command sent.\n")

            self.paramiko_client = client
            self.paramiko_channel = channel
            self.after(0, lambda: self.status_var.set("Running"))
            threading.Thread(target=self._read_paramiko_output, daemon=True).start()
            threading.Thread(target=self._watch_paramiko_channel, daemon=True).start()
        except Exception as exc:
            self.after(0, lambda: messagebox.showerror("Error", f"Failed to start remote command:\n{exc}"))
            self.after(0, self._set_running, False)

    def _read_paramiko_output(self) -> None:
        ch = self.paramiko_channel
        if ch is None:
            return
        pending: list[str] = []
        last_flush = time.monotonic()
        coalesce_sec = 0.08
        coalesce_max_chars = 262144

        def _flush_paramiko_log(*, force: bool = False) -> None:
            nonlocal pending, last_flush
            now = time.monotonic()
            blob = "".join(pending)
            if not blob:
                return
            if not force:
                elapsed = now - last_flush
                has_nl = "\n" in blob
                if not has_nl and len(blob) < coalesce_max_chars and elapsed < 0.35:
                    return
                if has_nl and elapsed < coalesce_sec and len(blob) < coalesce_max_chars:
                    return
            pending.clear()
            last_flush = now
            self.after(0, lambda b=blob: self.append_log(b))

        try:
            while not ch.exit_status_ready():
                had_data = False
                if ch.recv_ready():
                    data = ch.recv(4096).decode(errors="ignore")
                    if data:
                        had_data = True
                        self._append_session_log_tail(data)
                        self._update_session_activity_from_log(data)
                        self._detect_session_lost(data)
                        self._detect_session_established(data)
                        pending.append(data)
                        _flush_paramiko_log(force=False)
                if ch.recv_stderr_ready():
                    data = ch.recv_stderr(4096).decode(errors="ignore")
                    if data:
                        had_data = True
                        self._append_session_log_tail(data)
                        self._update_session_activity_from_log(data)
                        self._detect_session_lost(data)
                        self._detect_session_established(data)
                        pending.append(data)
                        _flush_paramiko_log(force=False)
                if self.stop_event.is_set():
                    break
                if not had_data:
                    _flush_paramiko_log(force=False)
                    time.sleep(0.05)
        except Exception as exc:
            _flush_paramiko_log(force=True)
            self.after(0, lambda e=exc: self._handle_transport_ended(f"ssh read error: {e}", -1))
            return
        _flush_paramiko_log(force=True)

    def _handle_transport_ended(self, reason: str, rc: int = 0) -> None:
        """SSH/subprocess transport lost while Start active — schedule GUI-side reconnect."""
        if self._user_stop_requested:
            if self.is_running:
                self.append_log(f"\n[GUI] Process exited with code {rc}\n")
                self._set_running(False)
            return
        if self._transport_reconnect_pending:
            return
        if not self.is_running:
            return
        self._transport_reconnect_pending = True
        self.stop_event.set()
        ch = self.paramiko_channel
        if ch is not None:
            try:
                ch.close()
            except Exception:
                pass
        if self.paramiko_client is not None:
            try:
                self.paramiko_client.close()
            except Exception:
                pass
        self.paramiko_channel = None
        self.paramiko_client = None
        proc = self.proc
        if proc is not None and proc.poll() is None:
            try:
                if os.name == "nt":
                    proc.send_signal(signal.CTRL_BREAK_EVENT)
                else:
                    proc.terminate()
            except Exception:
                pass
        self.proc = None
        self.append_log(f"\n[GUI] Connection lost: {reason} (rc={rc}).\n")
        if self.auto_start_var.get():
            self._cleanup_remote_daemons_async(
                reason=f"auto-reconnect prep ({reason})", include_start_script=True
            )
        self._set_running(False)

    def _watch_paramiko_channel(self) -> None:
        ch = self.paramiko_channel
        if ch is None:
            return
        rc = 0
        try:
            while True:
                client = self.paramiko_client
                transport = client.get_transport() if client is not None else None
                if transport is None or not transport.is_active():
                    rc = -1
                    break
                # invoke_shell channel can stay open without exit status;
                # treat explicit channel close as graceful end.
                if ch.closed:
                    rc = 0
                    break
                time.sleep(0.2)
        except Exception:
            rc = -1
        self.after(0, lambda r=rc: self._handle_transport_ended("ssh channel ended", r))

    def _read_output(self) -> None:
        if self.proc is None or self.proc.stdout is None:
            return
        for line in self.proc.stdout:
            if self.stop_event.is_set():
                break
            self._detect_session_lost(line)
            self._detect_session_established(line)
            self.after(0, self.append_log, line)

    def _watch_process(self) -> None:
        if self.proc is None:
            return
        rc = self.proc.wait()
        self.proc = None
        self.after(0, lambda r=rc: self._handle_transport_ended("local ssh process ended", r))

    def stop_script(self) -> None:
        self._user_stop_requested = True
        self._cancel_auto_start_retry_job()
        if self.paramiko_channel is not None:
            self.stop_event.set()
            try:
                self.paramiko_channel.close()
            except Exception:
                pass
            if self.paramiko_client is not None:
                try:
                    self.paramiko_client.close()
                except Exception:
                    pass
            self.paramiko_channel = None
            self.paramiko_client = None
            self.append_log("[GUI] Stop requested.\n")
            # 채널만 닫으면 원격 miniDU_callhome.sh 본체가 남아 supervision 이 계속될 수 있다.
            self._cleanup_remote_daemons_async(reason="manual stop(paramiko)", include_start_script=True)
            self._set_running(False)
            return

        if self.proc is None or self.proc.poll() is not None:
            # 로컬 프로세스가 이미 없어도 원격 잔존 데몬이 남아있을 수 있으므로 정리 시도.
            self._cleanup_remote_daemons_async(reason="manual stop(no local proc)", include_start_script=True)
            self._set_running(False)
            return

        self.stop_event.set()
        try:
            if os.name == "nt":
                self.proc.send_signal(signal.CTRL_BREAK_EVENT)
            else:
                self.proc.terminate()
        except Exception:
            pass

        self.append_log("[GUI] Stop requested.\n")
        self._cleanup_remote_daemons_async(reason="manual stop(subprocess)", include_start_script=True)

    def _build_remote_daemon_cleanup_script(self, include_start_script: bool) -> str:
        """Shell snippet run on remote via bash -lc (graceful FIFO + pkill fallbacks)."""
        fifo = shlex.quote(self.remote_control_fifo)
        # 짧은 sleep: GUI 종료·Stop 시 원격 정리 대기 시간을 줄임 (이전에는 sleep 1×4 로 최소 ~4s).
        _w = "sleep 0.35"
        parts: list[str] = [
            "set +e",
            f"if [ -p {fifo} ]; then printf '%s\\n' disconnect >{fifo} 2>/dev/null; fi",
            _w,
        ]
        if include_start_script:
            sp = (self.remote_script_path.get() or "").strip() or "/var/tmp/miniDU_callhome.sh"
            sq = shlex.quote(sp)
            parts.append(f"pkill -TERM -f {sq} 2>/dev/null || true")
            parts.append(_w)
            parts.append(f"pkill -KILL -f {sq} 2>/dev/null || true")
            parts.append("rm -f /var/tmp/netconf_tmp/miniDU_callhome.lock")
        parts.extend(
            [
                "pkill -TERM -f netopeer2-cli 2>/dev/null || true",
                _w,
                "pkill -KILL -f netopeer2-cli 2>/dev/null || true",
                "rm -f /var/tmp/netconf_tmp/miniDU_callhome.lock",
                "exit 0",
            ]
        )
        return "; ".join(parts)

    def _cleanup_remote_daemons_impl(
        self, reason: str, include_start_script: bool, *, subprocess_timeout: float | None
    ) -> None:
        pref = "[Remote-cleanup]"

        def log(msg: str) -> None:
            try:
                self.after(0, self.append_log, f"{pref} {msg}\n")
            except Exception:
                pass

        log(f"시작 ({reason})")
        ssh_user, ssh_host, ssh_port, ssh_password, key_path = self._remote_conn()
        if not ssh_user or not ssh_host or not ssh_port:
            log("건너뜀: Settings SSH 정보 없음")
            return
        auth_mode, info_or_error = self._resolve_auth_mode(ssh_password)
        if info_or_error and "Install" in info_or_error:
            log(f"건너뜀: {info_or_error}")
            return
        inner = self._build_remote_daemon_cleanup_script(include_start_script)
        remote_line = "bash -lc " + shlex.quote(inner)
        timeout = subprocess_timeout if subprocess_timeout is not None else 18.0
        conn_to = min(10.0, max(2.5, float(timeout) * 0.55))

        if auth_mode == "paramiko":
            try:
                import paramiko  # type: ignore

                client = paramiko.SSHClient()
                client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                client.connect(
                    hostname=ssh_host,
                    port=int(ssh_port),
                    username=ssh_user,
                    password=ssh_password if ssh_password else None,
                    key_filename=key_path if key_path else None,
                    timeout=conn_to,
                    auth_timeout=min(10.0, conn_to + 1.0),
                    banner_timeout=min(10.0, conn_to + 1.0),
                    look_for_keys=not bool(ssh_password),
                    allow_agent=True,
                )
                try:
                    _stdin, stdout, stderr = client.exec_command(remote_line, timeout=timeout)
                    rc = stdout.channel.recv_exit_status()
                    err = stderr.read().decode(errors="ignore").strip()
                    if rc != 0 and err:
                        log(f"종료 코드 {rc}: {err[:500]}")
                    else:
                        log("원격 정리 명령 완료")
                finally:
                    client.close()
            except Exception as exc:
                log(f"실패: {exc}")
            return

        use_password = bool(ssh_password)
        run_env = os.environ.copy()
        if use_password and auth_mode == "sshpass":
            run_env["SSHPASS"] = ssh_password
        ssh_cmd = self._ssh_base_cmd(ssh_user, ssh_host, ssh_port, key_path, use_password, auth_mode, ssh_password)
        try:
            r = subprocess.run(
                ssh_cmd + [remote_line],
                capture_output=True,
                text=True,
                env=run_env,
                timeout=timeout,
            )
            if r.returncode != 0:
                tail = (r.stderr or r.stdout or "").strip()[:500]
                log(f"종료 코드 {r.returncode}: {tail or '(no output)'}")
            else:
                log("원격 정리 명령 완료")
        except subprocess.TimeoutExpired:
            log(f"시간 초과 ({timeout:.0f}s)")
        except Exception as exc:
            log(f"실패: {exc}")

    def _cleanup_remote_daemons_async(self, *, reason: str, include_start_script: bool) -> None:
        threading.Thread(
            target=self._cleanup_remote_daemons_impl,
            args=(reason, include_start_script),
            kwargs={"subprocess_timeout": 18.0},
            daemon=True,
        ).start()

    def _cleanup_remote_daemons_blocking(
        self, *, reason: str, include_start_script: bool = True, timeout_s: float = 6.0
    ) -> None:
        def run() -> None:
            try:
                self._cleanup_remote_daemons_impl(
                    reason, include_start_script, subprocess_timeout=max(4.0, timeout_s)
                )
            except BaseException:
                pass

        t = threading.Thread(target=run, daemon=True)
        t.start()
        t.join(timeout=max(1.5, timeout_s + 0.75))
        if t.is_alive():
            try:
                self.append_log(f"[Remote-cleanup] 대기 시간 초과 ({timeout_s:.0f}s), 창 닫기 계속\n")
            except Exception:
                pass

    def _detect_remote_start_failure(self, text: str) -> None:
        if not self.is_running:
            return
        u = text.upper()
        if "MINIDU_CALLHOME.SH ALREADY RUNNING" in u.replace("-", "_"):
            self._flag_remote_start_failed("miniDU lock/flock — Stop 후 Start 하세요.")
            return
        if "ANOTHER MINIDU_CALLHOME.SH INSTANCE" in u.replace("-", "_"):
            self._flag_remote_start_failed("원격 miniDU가 이미 실행 중입니다 — Stop 후 Start 하세요.")

    def _flag_remote_start_failed(self, msg: str) -> None:
        if getattr(self, "_remote_start_failed_handled", False):
            return
        self._remote_start_failed_handled = True

        def _apply() -> None:
            self.append_log(f"[GUI] Start failed: {msg}\n")
            self._set_running(False)
            try:
                messagebox.showerror("Start failed", msg)
            except Exception:
                pass

        self.after(0, _apply)

    def _detect_session_lost(self, text: str) -> None:
        if not self.is_running or self._user_stop_requested:
            return
        if not (self.session_established or self.manual_send_ready):
            return
        recent = self._recent_session_log_chunk(text)
        upper = recent.upper()
        if not self._log_indicates_session_lost(upper, recent_only=upper):
            return
        self.session_established = False
        self.manual_send_ready = False
        self.auto_xml_send_done = False
        self.after(0, self._sync_manual_send_widgets)
        if not self._session_lost_logged:
            self._session_lost_logged = True
            self._session_lost_at_mono = time.monotonic()
            self.after(
                0,
                self.append_log,
                "[GUI] NETCONF session lost — remote script reconnecting; waiting for login...\n",
            )
            self.after(0, lambda: self.status_var.set("Session lost — waiting..."))
        self._schedule_session_watch()

    def _maybe_reconnect_start_after_conformance(self) -> None:
        """Conformance run ended while Start session was lost — retry if auto-reconnect is on."""
        if not getattr(self, "_session_lost_logged", False):
            return
        if self.session_established and self.manual_send_ready:
            return
        if not self.is_running or self._user_stop_requested:
            return
        if self._transport_reconnect_pending:
            return
        if not self.auto_start_var.get():
            return
        self.append_log("[GUI] Conformance finished — forcing Start reconnect after session loss.\n")
        self._handle_transport_ended("session lost after conformance", -1)

    def _detect_session_established(self, text: str) -> None:
        if self.session_established:
            upper = text.upper()
            if (not self.manual_send_ready) and ("SUPERVISION ACTIVE" in upper):
                self.manual_send_ready = True
                self.after(0, self._sync_manual_send_widgets)
                self.after(0, self.append_log, "[GUI] Manual send enabled (SUPERVISION ACTIVE).\n")
            return
        upper = text.upper()
        markers = [
            "ACCEPTED A CONNECTION ON",
            "AUTHENTICATION SUCCESSFUL",
            "LOGIN SUCCESSFUL",
            "<HELLO",
            "CALLHOME DETECTED",
            "CALL HOME DETECTED",
            "SUPERVISION ACTIVE",
        ]
        if any(m in upper for m in markers):
            if getattr(self, "_session_lost_logged", False):
                # Planned Conformance reboot/reconnect must keep PASS/FAIL totals.
                if not bool(getattr(self, "_conformance_run_busy", False)):
                    self._conformance_reset_session_run_stats("ORU 재부팅/재연결")
            self.session_established = True
            self._session_lost_logged = False
            self._session_lost_at_mono = 0.0
            self._last_session_activity_mono = time.monotonic()
            self.after(0, self.append_log, "[GUI] Session established detected. Netconf Client is available.\n")
            self.manual_send_ready = True
            self.after(0, lambda: self.status_var.set("Running"))
            self.after(0, self._sync_manual_send_widgets)
            self.after(0, self.append_log, "[GUI] Manual send enabled.\n")
            # Conformance auto-upload disabled while tab is viewer-only
            self.after(2000, self._schedule_conformance_auto_sync_once)
            if self.auto_xml_send_var.get() and not self.auto_xml_send_done:
                self.auto_xml_send_done = True
                self.after(0, self.append_log, "[GUI] Auto XML Send scheduled in 4 seconds.\n")
                self.after(4000, self._run_auto_xml_send_once)
        if "SUPERVISION ACTIVE" in upper and not self.manual_send_ready:
            had_session = self.session_established
            self.session_established = True
            self.manual_send_ready = True
            self.after(0, self._sync_manual_send_widgets)
            if had_session:
                self.after(0, self.append_log, "[GUI] Manual send enabled (SUPERVISION ACTIVE).\n")
            else:
                self.after(0, self.append_log, "[GUI] Session + manual send via SUPERVISION ACTIVE (no prior hello marker).\n")

    def _run_auto_xml_send_once(self) -> None:
        if not (self.is_running and self.auto_xml_send_var.get() and self.session_established):
            self.auto_xml_send_done = False
            return
        self.append_log("[GUI] Auto XML Send triggered (All Tabs Send 1x).\n")
        self.send_all_once()

    def _send_scheduler_payload(self, payload: str) -> None:
        t0 = time.perf_counter()
        payload = payload.strip()
        if not payload:
            return
        trace_id = self._trace("send_start", mode=self.send_mode_var.get(), bytes=str(len(payload)))
        payload = self._prepare_payload_for_remote_send(payload)
        # Current runtime path sends through miniDU_callhome.sh stdin bridge,
        # so commands must be line-based CLI text.
        data = payload + "\n"
        if self.exec_mode.get() == "remote":
            # Fast path: write directly to running remote shell channel stdin.
            if self.paramiko_channel is not None:
                t_direct = time.perf_counter()
                self.paramiko_channel.send(data)
                self._perf_record("send_direct_channel", time.perf_counter() - t_direct)
                self._perf_record("send_payload", time.perf_counter() - t0)
                self._trace("send_done", trace=str(trace_id), path="direct_channel")
                return
            ok, msg = self._send_remote_control_command(payload)
            if not ok:
                # FIFO channel can be missing if an old script is still running.
                # Fallback to stdin bridge path for robustness.
                if "__FIFO_MISSING__" in msg and self.proc is not None and self.proc.stdin is not None and self.proc.poll() is None:
                    self.proc.stdin.write(data)
                    self.proc.stdin.flush()
                    self.append_log("[GUI] FIFO missing, fallback to stdin bridge.\n")
                    self._perf_record("send_payload", time.perf_counter() - t0)
                    self._trace("send_done", trace=str(trace_id), path="stdin_fallback")
                    return
                raise RuntimeError(msg)
            self._perf_record("send_payload", time.perf_counter() - t0)
            self._trace("send_done", trace=str(trace_id), path="fifo_remote")
            return
        if self.paramiko_channel is not None:
            self.paramiko_channel.send(data)
            self._perf_record("send_payload", time.perf_counter() - t0)
            self._trace("send_done", trace=str(trace_id), path="paramiko_channel")
            return
        if self.proc is not None and self.proc.stdin is not None and self.proc.poll() is None:
            self.proc.stdin.write(data)
            self.proc.stdin.flush()
            self._perf_record("send_payload", time.perf_counter() - t0)
            self._trace("send_done", trace=str(trace_id), path="local_stdin")
            return
        raise RuntimeError("No active process/channel for sending")

    def _send_remote_control_command(self, command: str) -> tuple[bool, str]:
        t0 = time.perf_counter()
        ssh_user, ssh_host, ssh_port, ssh_password, key_path = self._remote_conn()
        if not ssh_user or not ssh_host or not ssh_port:
            return False, "SSH connection info is incomplete."
        auth_mode, info_or_error = self._resolve_auth_mode(ssh_password)
        if info_or_error and "Install" in info_or_error:
            return False, info_or_error
        shell_cmd = (
            f"if [ -p {shlex.quote(self.remote_control_fifo)} ]; then "
            f"printf '%s\\n' {shlex.quote(command)} > {shlex.quote(self.remote_control_fifo)}; "
            f"else echo '__FIFO_MISSING__'; exit 2; fi"
        )

        if auth_mode == "paramiko":
            try:
                import paramiko  # type: ignore
                client = self.paramiko_client
                owned = False
                if client is None:
                    client = paramiko.SSHClient()
                    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                    client.connect(
                        hostname=ssh_host,
                        port=int(ssh_port),
                        username=ssh_user,
                        password=ssh_password if ssh_password else None,
                        key_filename=key_path if key_path else None,
                        timeout=8,
                        auth_timeout=8,
                        banner_timeout=8,
                        look_for_keys=not bool(ssh_password),
                        allow_agent=True,
                    )
                    owned = True
                _, stdout, stderr = client.exec_command(shell_cmd)
                rc = stdout.channel.recv_exit_status()
                out = stdout.read().decode(errors="ignore").strip()
                err = stderr.read().decode(errors="ignore").strip()
                if owned:
                    client.close()
                if rc != 0 or "__FIFO_MISSING__" in out:
                    return False, err or out or f"Remote control send failed (rc={rc})"
                self._perf_record("send_fifo_remote", time.perf_counter() - t0)
                return True, "ok"
            except Exception as exc:
                return False, f"Remote control send failed: {exc}"

        use_password = bool(ssh_password)
        run_env = os.environ.copy()
        if use_password and auth_mode == "sshpass":
            run_env["SSHPASS"] = ssh_password
        ssh_cmd = self._ssh_base_cmd(ssh_user, ssh_host, ssh_port, key_path, use_password, auth_mode, ssh_password)
        result = subprocess.run(ssh_cmd + [shell_cmd], capture_output=True, text=True, env=run_env)
        out = (result.stdout or "").strip()
        err = (result.stderr or "").strip()
        if result.returncode != 0 or "__FIFO_MISSING__" in out:
            return False, err or out or f"Remote control send failed (rc={result.returncode})"
        self._perf_record("send_fifo_remote", time.perf_counter() - t0)
        return True, "ok"

    @staticmethod
    def _looks_like_netconf_xml_document(text: str) -> bool:
        raw = (text or "").lstrip("\ufeff").strip()
        body = re.sub(r"^\s*<\?xml[^>]*\?>\s*", "", raw, flags=re.IGNORECASE)
        lower = body.lstrip().lower()
        if lower.startswith(("<rpc", "<get", "<get-config", "<edit-config", "<config", "<filter")):
            return True
        # Be tolerant for templates that start with comments or extra wrappers.
        return any(tok in lower for tok in ("<rpc", "<get ", "<get>", "<get-config", "<edit-config", "<config", "<filter"))

    @staticmethod
    def _is_rpc_get_like(text: str) -> bool:
        raw = (text or "").lstrip("\ufeff").strip()
        body = re.sub(r"^\s*<\?xml[^>]*\?>\s*", "", raw, flags=re.IGNORECASE).lower()
        if body.startswith("<rpc"):
            return ("<get-config" in body) or bool(re.search(r"<get\b", body))
        return body.startswith("<get-config") or bool(re.search(r"^\s*<get\b", body))

    @staticmethod
    def _user_rpc_xml_from_template(xml_text: str) -> tuple[str, bool]:
        """
        netopeer2-cli user-rpc adds the <rpc> envelope itself.
        Return inner operation XML unchanged; True if <rpc> wrapper was stripped.
        """
        text = (xml_text or "").lstrip("\ufeff").strip()
        # Some templates/log copy paths can contain escaped quotes (\" / \').
        # Normalize them before XML extraction so netopeer2-cli receives valid XML.
        text = text.replace('\\"', '"').replace("\\'", "'")
        decl_m = re.match(r"^\s*(<\?xml[^>]*\?>\s*)", text, flags=re.IGNORECASE)
        decl = decl_m.group(1) if decl_m else ""
        body = text[len(decl) :] if decl else text
        # Some tabs contain appended rpc-reply/log text after the request XML.
        # Pick the first <rpc>...</rpc> request block and strip only that wrapper.
        m = re.search(r"<rpc\b([^>]*)>([\s\S]*?)</rpc>", body.strip(), flags=re.IGNORECASE)
        if not m:
            return text, False
        rpc_attrs = m.group(1) or ""
        inner = m.group(2).strip()
        # Keep NETCONF default namespace when stripping <rpc>.
        # Example: <rpc xmlns="urn:...:netconf:base:1.0"><get>...</get></rpc>
        # -> <get xmlns="urn:...:netconf:base:1.0">...</get>
        ns_m = re.search(r'\bxmlns\s*=\s*["\']([^"\']+)["\']', rpc_attrs, flags=re.IGNORECASE)
        if ns_m and inner:
            ns = ns_m.group(1).strip()
            if ns and not re.search(r"^\s*<[^>]+\bxmlns\s*=", inner, flags=re.IGNORECASE):
                inner = re.sub(
                    r"^\s*<([a-zA-Z0-9_\-:]+)",
                    rf'<\1 xmlns="{ns}"',
                    inner,
                    count=1,
                )
        return f"{decl}{inner}" if decl else inner, True

    def _upload_and_build_user_rpc_command(
        self,
        xml_text: str,
        cache_kind: str = "user_rpc",
        with_out: bool = False,
    ) -> str:
        """Upload XML for user-rpc (operation body only; netopeer2-cli wraps <rpc>)."""
        payload, stripped_rpc = self._user_rpc_xml_from_template(xml_text)
        cache_key = self._cache_key(cache_kind, payload)
        remote_path = self.remote_cfg_cache.get(cache_key, "")
        if not remote_path:
            remote_path = f"/var/tmp/netconf_tmp/gui_{cache_kind}_{int(time.time() * 1000)}.xml"
            ok, msg = self._upload_text_to_remote(remote_path, payload)
            if not ok:
                raise RuntimeError(msg)
            self.remote_cfg_cache[cache_key] = remote_path
            self.append_log(f"[GUI] Uploaded XML for user-rpc: {remote_path}\n")
        else:
            self.append_log(f"[GUI] Reused cached user-rpc XML: {remote_path}\n")
        if stripped_rpc:
            self.append_log("[GUI] user-rpc: <rpc> wrapper omitted (CLI adds envelope automatically).\n")
        cmd = f"user-rpc --content {remote_path}"
        if with_out:
            out_path = f"/var/tmp/netconf_tmp/gui_user_rpc_out_{int(time.time() * 1000)}.xml"
            cmd = f"{cmd} --out {out_path}"
            self.append_log(f"[GUI] user-rpc response out: {out_path}\n")
        return cmd

    def _prepare_payload_for_remote_send(self, payload: str) -> str:
        t0 = time.perf_counter()
        if self.send_mode_var.get() == "raw_rpc":
            stripped = payload.lstrip("\ufeff").strip()
            xml_body = re.sub(r"^\s*<\?xml[^>]*\?>\s*", "", stripped, flags=re.IGNORECASE)
            low = xml_body.lower()
            # Keep GET/GET-CONFIG RPC payloads as-is via user-rpc (no CLI conversion).
            # Preserve legacy path for edit-config so existing Apply/SET behavior is unchanged.
            if low.startswith("<rpc") and ("<edit-config" not in low) and (
                ("<get-config" in low) or re.search(r"<get\b", low)
            ):
                cmd = self._upload_and_build_user_rpc_command(
                    stripped,
                    cache_kind="raw_get_rpc",
                    with_out=True,
                )
                self._perf_record("prepare_payload", time.perf_counter() - t0)
                self.append_log(f"[GUI] RAW RPC GET(as-is) -> {cmd}\n")
                return cmd
            # Generic/custom RPC (e.g. software-activate): send as user-rpc as-is.
            if low.startswith("<rpc") and ("<edit-config" not in low):
                cmd = self._upload_and_build_user_rpc_command(
                    stripped,
                    cache_kind="raw_custom_rpc",
                    with_out=True,
                )
                self._perf_record("prepare_payload", time.perf_counter() - t0)
                self.append_log(f"[GUI] RAW RPC custom(as-is) -> {cmd}\n")
                return cmd
            t_parse = time.perf_counter()
            translated = self._translate_raw_rpc_to_cli(payload.strip())
            self._perf_record("raw_rpc_translate", time.perf_counter() - t_parse)
            self._perf_record("prepare_payload", time.perf_counter() - t0)
            return translated
        lines = [ln for ln in payload.splitlines() if ln.strip()]
        if not lines:
            self._perf_record("prepare_payload", time.perf_counter() - t0)
            return payload
        first = lines[0].strip()
        xml_body = "\n".join(lines[1:]).strip()

        # Case 1) command + inline XML in one tab:
        # edit-config ... --config=/var/tmp/netconf_tmp/x.xml
        # <config>...</config>
        if "--config=" in first and xml_body.startswith("<"):
            remote_path = self._extract_config_path(first)
            if remote_path:
                cache_key = self._cache_key("inline", xml_body, remote_path)
                cached = self.remote_cfg_cache.get(cache_key)
                if not cached:
                    self._perf_record("cfg_cache_miss", 0.0)
                    ok, msg = self._upload_text_to_remote(remote_path, xml_body)
                    if not ok:
                        raise RuntimeError(msg)
                    self.remote_cfg_cache[cache_key] = remote_path
                    self.append_log(f"[GUI] Uploaded inline XML to {remote_path}\n")
                else:
                    self._perf_record("cfg_cache_hit", 0.0)
                    self.append_log(f"[GUI] Reused cached XML: {cached}\n")
                self._perf_record("prepare_payload", time.perf_counter() - t0)
                return first

        # Case 2) local config path in command: --config=C:\...\x.xml
        if "--config=" in first:
            src_path = self._extract_config_path(first)
            if src_path and Path(src_path).exists():
                remote_path = f"/var/tmp/netconf_tmp/{Path(src_path).name}"
                ok, msg = self._upload_file_to_remote(src_path, remote_path)
                if not ok:
                    raise RuntimeError(msg)
                self.append_log(f"[GUI] Uploaded local file to {remote_path}\n")
                self._perf_record("prepare_payload", time.perf_counter() - t0)
                return re.sub(r"--config=\S+", f"--config={remote_path}", first)

        # Case 3) raw XML only in tab -> auto edit-config
        if payload.lstrip().startswith("<config"):
            cache_key = self._cache_key("auto_config", payload)
            remote_path = self.remote_cfg_cache.get(cache_key, "")
            if not remote_path:
                self._perf_record("cfg_cache_miss", 0.0)
                remote_path = f"/var/tmp/netconf_tmp/gui_config_{int(time.time())}.xml"
                ok, msg = self._upload_text_to_remote(remote_path, payload)
                if not ok:
                    raise RuntimeError(msg)
                self.remote_cfg_cache[cache_key] = remote_path
            else:
                self._perf_record("cfg_cache_hit", 0.0)
            auto_cmd = f"edit-config --target running --defop merge --config={remote_path}"
            self.append_log(f"[GUI] Auto command: {auto_cmd}\n")
            self._perf_record("prepare_payload", time.perf_counter() - t0)
            return auto_cmd

        self._perf_record("prepare_payload", time.perf_counter() - t0)
        return payload

    def _extract_rpc_xpath_filter(self, text: str) -> str | None:
        """Return XPath string when RPC uses ``<filter type="xpath" select="..."/>``."""
        if not re.search(r"<filter\b", text, re.IGNORECASE):
            return None
        if not re.search(r'\btype\s*=\s*["\']xpath["\']', text, re.IGNORECASE):
            return None
        m = re.search(
            r'<filter\b[\s\S]*?\bselect\s*=\s*"([^"]+)"',
            text,
            re.IGNORECASE,
        )
        if not m:
            m = re.search(
                r"<filter\b[\s\S]*?\bselect\s*=\s*'([^']+)'",
                text,
                re.IGNORECASE,
            )
        xpath = (m.group(1).strip() if m else "") or ""
        return xpath or None

    def _upload_subtree_filter_for_cli(self, filter_xml: str, cache_kind: str) -> str:
        cache_key = self._cache_key(cache_kind, filter_xml)
        remote_path = self.remote_cfg_cache.get(cache_key, "")
        if not remote_path:
            remote_path = f"/var/tmp/netconf_tmp/gui_{cache_kind}_{int(time.time() * 1000)}.xml"
            ok, msg = self._upload_text_to_remote(remote_path, filter_xml)
            if not ok:
                raise RuntimeError(msg)
            self.remote_cfg_cache[cache_key] = remote_path
        return remote_path

    @staticmethod
    def _subtree_filter_xml_from_node(filter_node: ET.Element) -> str:
        inner_parts = [
            ET.tostring(c, encoding="unicode")
            for c in list(filter_node)
            if isinstance(c.tag, str)
        ]
        filter_xml = "".join(inner_parts).strip()
        if not filter_xml:
            filter_xml = ET.tostring(filter_node, encoding="unicode")
            m_f = re.match(
                r"^\s*<filter\b[^>]*>([\s\S]*)</filter>\s*$",
                filter_xml,
                flags=re.IGNORECASE,
            )
            if m_f:
                filter_xml = m_f.group(1).strip()
        return filter_xml

    @staticmethod
    def _get_config_source_from_op(op: ET.Element) -> str:
        def local(tag: str) -> str:
            return tag.split("}", 1)[-1] if "}" in tag else tag

        source_node = next((c for c in list(op) if isinstance(c.tag, str) and local(c.tag) == "source"), None)
        if source_node is None:
            return "running"
        source_child = next((c for c in list(source_node) if isinstance(c.tag, str)), None)
        if source_child is None:
            return "running"
        return local(source_child.tag)

    def _translate_raw_rpc_to_cli(self, payload: str) -> str:
        text = payload.strip()
        # Support XML declaration from ATOM exports.
        text = re.sub(r"^\s*<\?xml[^>]*\?>\s*", "", text, flags=re.IGNORECASE)
        if not text.startswith("<rpc"):
            return text

        # Prefer raw extraction to preserve original namespace/prefix text.
        lower = text.lower()
        xpath = self._extract_rpc_xpath_filter(text)
        if xpath and ("<get-config" in lower or re.search(r"<get\b", lower)):
            if "<get-config" in lower:
                m_src = re.search(
                    r"<source>\s*<([a-zA-Z0-9_\-:]+)\s*/>\s*</source>",
                    text,
                    flags=re.IGNORECASE | re.DOTALL,
                )
                source = m_src.group(1).split(":")[-1] if m_src else "running"
                cmd = f"get-config --source {source} --filter-xpath {shlex.quote(xpath)}"
            else:
                cmd = f"get --filter-xpath {shlex.quote(xpath)}"
            self.append_log(f"[GUI] RAW RPC -> {cmd}\n")
            return cmd
        if "<edit-config" in lower:
            target = "running"
            defop = "merge"
            m_target = re.search(r"<target>\s*<([a-zA-Z0-9_\-:]+)\s*/>\s*</target>", text, flags=re.IGNORECASE | re.DOTALL)
            if m_target:
                target = m_target.group(1).split(":")[-1]
            m_defop = re.search(r"<default-operation>\s*([^<]+)\s*</default-operation>", text, flags=re.IGNORECASE | re.DOTALL)
            if m_defop:
                defop = m_defop.group(1).strip()

            m_cfg = re.search(r"(<config\b[\s\S]*?</config>)", text, flags=re.IGNORECASE)
            if m_cfg:
                cfg_xml = m_cfg.group(1)
                # netopeer2-cli edit-config wraps provided config once.
                # If we pass outer <config> tag too, it becomes <config><config>...</config></config>.
                inner = re.match(r"^\s*<config\b[^>]*>([\s\S]*)</config>\s*$", cfg_xml, flags=re.IGNORECASE)
                if inner:
                    cfg_xml = inner.group(1).strip()
                cache_key = self._cache_key("raw_edit", cfg_xml, target, defop)
                remote_path = self.remote_cfg_cache.get(cache_key, "")
                if not remote_path:
                    self._perf_record("cfg_cache_miss", 0.0)
                    remote_path = f"/var/tmp/netconf_tmp/gui_cfg_{int(time.time() * 1000)}.xml"
                    ok, msg = self._upload_text_to_remote(remote_path, cfg_xml)
                    if not ok:
                        raise RuntimeError(msg)
                    self.remote_cfg_cache[cache_key] = remote_path
                else:
                    self._perf_record("cfg_cache_hit", 0.0)
                cmd = f"edit-config --target {target} --defop {defop} --config={remote_path}"
                self.append_log(f"[GUI] RAW RPC -> {cmd}\n")
                self.append_log(f"[GUI] edit-config config uploaded: {remote_path}\n")
                return cmd

        try:
            root = ET.fromstring(text)
        except Exception:
            return text

        def local(tag: str) -> str:
            return tag.split("}", 1)[-1] if "}" in tag else tag

        op = None
        for child in list(root):
            if isinstance(child.tag, str):
                op = child
                break
        if op is None:
            return text

        op_name = local(op.tag)
        if op_name == "edit-config":
            target = "running"
            defop = "merge"
            target_node = next((c for c in list(op) if local(c.tag) == "target"), None)
            if target_node is not None:
                target_child = next((c for c in list(target_node) if isinstance(c.tag, str)), None)
                if target_child is not None:
                    target = local(target_child.tag)
            defop_node = next((c for c in list(op) if local(c.tag) == "default-operation"), None)
            if defop_node is not None and (defop_node.text or "").strip():
                defop = (defop_node.text or "").strip()
            config_node = next((c for c in list(op) if local(c.tag) == "config"), None)
            if config_node is None:
                return text
            cfg_xml = ET.tostring(config_node, encoding="unicode")
            remote_path = f"/var/tmp/netconf_tmp/gui_cfg_{int(time.time() * 1000)}.xml"
            ok, msg = self._upload_text_to_remote(remote_path, cfg_xml)
            if not ok:
                raise RuntimeError(msg)
            cmd = f"edit-config --target {target} --defop {defop} --config={remote_path}"
            self.append_log(f"[GUI] RAW RPC -> {cmd}\n")
            self.append_log(f"[GUI] edit-config config uploaded: {remote_path}\n")
            return cmd

        if op_name == "get-config":
            source = self._get_config_source_from_op(op)
            filter_node = next((c for c in list(op) if local(c.tag) == "filter"), None)
            if filter_node is not None:
                ftype = (filter_node.get("type") or "").strip().lower()
                select = (filter_node.get("select") or "").strip()
                if ftype == "xpath" and select:
                    cmd = f"get-config --source {source} --filter-xpath {shlex.quote(select)}"
                    self.append_log(f"[GUI] RAW RPC -> {cmd}\n")
                    return cmd
                filter_xml = self._subtree_filter_xml_from_node(filter_node)
                if filter_xml:
                    remote_path = self._upload_subtree_filter_for_cli(filter_xml, "get_config_filter")
                    cmd = f"get-config --source {source} --filter={remote_path}"
                    self.append_log(f"[GUI] RAW RPC -> {cmd}\n")
                    return cmd
            cmd = f"get-config --source {source}"
            self.append_log(f"[GUI] RAW RPC -> {cmd}\n")
            return cmd

        if op_name == "get":
            filter_node = next((c for c in list(op) if local(c.tag) == "filter"), None)
            if filter_node is not None:
                ftype = (filter_node.get("type") or "").strip().lower()
                select = (filter_node.get("select") or "").strip()
                if ftype == "xpath" and select:
                    cmd = f"get --filter-xpath {shlex.quote(select)}"
                    self.append_log(f"[GUI] RAW RPC -> {cmd}\n")
                    return cmd
                filter_xml = self._subtree_filter_xml_from_node(filter_node)
                if filter_xml:
                    remote_path = self._upload_subtree_filter_for_cli(filter_xml, "get_filter")
                    cmd = f"get --filter={remote_path}"
                    self.append_log(f"[GUI] RAW RPC -> {cmd}\n")
                    return cmd
            cmd = "get"
            self.append_log(f"[GUI] RAW RPC -> {cmd}\n")
            return cmd

        # Fallback: unsupported raw RPC operation for current CLI bridge.
        raise RuntimeError(
            "Unsupported RAW RPC operation in current mode. "
            "Use edit-config/get/get-config RPC, or CLI command mode."
        )

    @staticmethod
    def _cache_key(kind: str, *parts: str) -> str:
        h = hashlib.sha1()
        h.update(kind.encode("utf-8"))
        for p in parts:
            h.update(b"\n--\n")
            h.update((p or "").encode("utf-8", errors="ignore"))
        return h.hexdigest()

    @staticmethod
    def _extract_config_path(command: str) -> str | None:
        m = re.search(r"--config=([^\s]+)", command)
        return m.group(1).strip() if m else None

    def _remote_conn(self) -> tuple[str, str, str, str, str]:
        return (
            self.remote_user.get().strip(),
            self.remote_host.get().strip(),
            self.remote_port.get().strip(),
            self.remote_password.get().strip(),
            self.remote_key_path.get().strip(),
        )

    def _upload_text_to_remote(self, remote_path: str, content: str) -> tuple[bool, str]:
        fd, tmp = tempfile.mkstemp(prefix="gui_cfg_", suffix=".xml")
        os.close(fd)
        try:
            Path(tmp).write_text(content, encoding="utf-8")
            return self._upload_file_to_remote(tmp, remote_path)
        finally:
            try:
                Path(tmp).unlink(missing_ok=True)
            except Exception:
                pass

    def _upload_file_to_remote(self, local_path: str, remote_path: str) -> tuple[bool, str]:
        t0 = time.perf_counter()
        trace_id = self._trace("upload_start", src=local_path, dst=remote_path)
        ssh_user, ssh_host, ssh_port, ssh_password, key_path = self._remote_conn()
        if not ssh_user or not ssh_host or not ssh_port:
            return False, "SSH connection info is incomplete."
        auth_mode, info_or_error = self._resolve_auth_mode(ssh_password)
        if info_or_error and "Install" in info_or_error:
            return False, info_or_error

        remote_parent = str(Path(remote_path).parent).replace("\\", "/")

        if auth_mode == "paramiko":
            try:
                import paramiko  # type: ignore
                client = self.paramiko_client
                owned = False
                if client is None:
                    client = paramiko.SSHClient()
                    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                    client.connect(
                        hostname=ssh_host,
                        port=int(ssh_port),
                        username=ssh_user,
                        password=ssh_password if ssh_password else None,
                        key_filename=key_path if key_path else None,
                        timeout=8,
                        auth_timeout=8,
                        banner_timeout=8,
                        look_for_keys=not bool(ssh_password),
                        allow_agent=True,
                    )
                    owned = True
                _, stdout, stderr = client.exec_command(f"mkdir -p {shlex.quote(remote_parent)}")
                if stdout.channel.recv_exit_status() != 0:
                    err = stderr.read().decode(errors="ignore").strip()
                    if owned:
                        client.close()
                    return False, f"Remote mkdir failed: {err}"
                sftp = client.open_sftp()
                sftp.put(local_path, remote_path)
                sftp.close()
                if owned:
                    client.close()
                self._perf_record("upload_remote", time.perf_counter() - t0)
                self._trace("upload_done", trace=str(trace_id), mode="paramiko")
                return True, "ok"
            except Exception as exc:
                return False, f"Remote upload failed: {exc}"

        use_password = bool(ssh_password)
        run_env = os.environ.copy()
        if use_password and auth_mode == "sshpass":
            run_env["SSHPASS"] = ssh_password
        ssh_cmd = self._ssh_base_cmd(ssh_user, ssh_host, ssh_port, key_path, use_password, auth_mode, ssh_password)
        mkdir_result = subprocess.run(ssh_cmd + [f"mkdir -p {shlex.quote(remote_parent)}"], capture_output=True, text=True, env=run_env)
        if mkdir_result.returncode != 0:
            return False, (mkdir_result.stderr or mkdir_result.stdout).strip() or "Remote mkdir failed"
        scp_cmd = self._scp_base_cmd(ssh_port, key_path, use_password, auth_mode, ssh_password) + [
            local_path,
            f"{ssh_user}@{ssh_host}:{remote_path}",
        ]
        scp_result = subprocess.run(scp_cmd, capture_output=True, text=True, env=run_env)
        if scp_result.returncode != 0:
            return False, (scp_result.stderr or scp_result.stdout).strip() or "Remote upload failed"
        self._perf_record("upload_remote", time.perf_counter() - t0)
        self._trace("upload_done", trace=str(trace_id), mode=auth_mode)
        return True, "ok"

    def _collect_scheduler_payloads(self, selected_only: bool = False) -> list[tuple[str, str]]:
        if not self.message_tabs:
            return []
        if selected_only:
            idx = self.msg_notebook.index("current")
            tab = self.message_tabs[idx]
            return [(tab["title_var"].get() or f"MSG-{idx+1}", tab["text"].get("1.0", "end"))]
        # Respect current notebook visual order (left -> right).
        payloads: list[tuple[str, str]] = []
        for i in range(len(self.message_tabs)):
            frame_name = self.msg_notebook.tabs()[i]
            tab = next((t for t in self.message_tabs if str(t["frame"]) == frame_name), None)
            if tab is None:
                continue
            payloads.append((tab["title_var"].get() or f"MSG-{i+1}", tab["text"].get("1.0", "end")))
        return payloads

    def send_selected_once(self) -> None:
        if not self.session_established:
            messagebox.showwarning("Warning", "Session is not established yet.")
            return
        if not self.manual_send_ready:
            messagebox.showwarning("Warning", "Manual send is enabled after session login.")
            return
        payloads = self._collect_scheduler_payloads(selected_only=True)
        threading.Thread(target=self._send_payloads_once_worker, args=(payloads,), daemon=True).start()

    def send_all_once(self) -> None:
        if not self.session_established:
            messagebox.showwarning("Warning", "Session is not established yet.")
            return
        if not self.manual_send_ready:
            messagebox.showwarning("Warning", "Manual send is enabled after session login.")
            return
        payloads = self._collect_scheduler_payloads(selected_only=False)
        threading.Thread(target=self._send_payloads_once_worker, args=(payloads,), daemon=True).start()

    def _send_payloads_once_worker(self, payloads: list[tuple[str, str]]) -> None:
        self.after(0, lambda: self.send_selected_btn.config(state="disabled"))
        self.after(0, lambda: self.send_all_btn.config(state="disabled"))
        mb = self.mplane_apply_btn
        if mb is not None:
            self.after(0, lambda b=mb: b.config(state="disabled"))
        sent = 0
        try:
            for title, payload in payloads:
                if not payload.strip():
                    continue
                try:
                    self._send_scheduler_payload(payload)
                    self.after(0, self.append_log, f"[GUI] Sent once from tab: {title}\n")
                    sent += 1
                except Exception as exc:
                    self.after(0, self.append_log, f"[GUI] Send once failed ({title}): {exc}\n")
            if sent == 0:
                self.after(0, self.append_log, "[GUI] No payload to send.\n")
        finally:
            self.after(0, self._sync_manual_send_widgets)

    def on_close(self) -> None:
        if self._config_save_job is not None:
            try:
                self.after_cancel(self._config_save_job)
            except Exception:
                pass
            self._config_save_job = None
        self._conformance_omit_last_run_from_config_save = True
        self._conformance_last_run_snapshot_cache = None
        if hasattr(self, "conformance_last_run_hint_var"):
            self.conformance_last_run_hint_var.set("")
        self._save_current_config()
        if self._conformance_run_busy:
            self._conformance_stop_run()
        self.stop_script()
        # GUI 종료 시 관련 원격 데몬까지 정리
        self._cleanup_remote_daemons_blocking(reason="GUI close", include_start_script=True, timeout_s=4.0)
        self.destroy()


if __name__ == "__main__":
    app = CallhomeGUI()
    app.mainloop()

# EXE build (Windows):
#   py -m pip install pyinstaller paramiko openpyxl
#   py -m PyInstaller --noconfirm --onefile --windowed --collect-submodules openpyxl --name "O-RAN Netconf" callhome_gui.py
# If `py` launcher is unavailable in PowerShell, use:
#   python -m pip install pyinstaller paramiko openpyxl
#   python -m PyInstaller --noconfirm --onefile --windowed --collect-submodules openpyxl --name "O-RAN Netconf" callhome_gui.py
# Output:
#   dist\O-RAN Netconf.exe

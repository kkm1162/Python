import io
import os
import re
import shlex
import shutil
import signal
import sys
import subprocess
import threading
import tkinter as tk
import json
import base64
import hashlib
import tempfile
import time
import xml.dom.minidom
import xml.etree.ElementTree as ET
import importlib
from datetime import datetime
import tkinter.font as tkfont
from pathlib import Path, PurePosixPath
from tkinter import filedialog, messagebox, simpledialog, ttk
from typing import Any

import conformance_manifest as _conf_manifest
from conformance_mixin import ConformanceMixin


# CC table column keys (spreadsheet/grid); must match mplane_control.CC_COL_MAP round-trip.
_MPLANE_CC_TREE_COLS: tuple[str, ...] = (
    "enabled",
    "cc",
    "ant",
    "scs",
    "prb",
    "iq",
    "compression",
    "typ",
    "dl_mhz",
    "bw",
    "ul_mhz",
    "t_au",
    "n_ta",
    "t_du",
)

_MPLANE_CC_GRID_ROWS = 16
_MPLANE_CC_ENTRY_KEYS: tuple[str, ...] = _MPLANE_CC_TREE_COLS[1:]  # all carrier columns except "enabled"

_MPLANE_PHY_ROWS = 16


def _app_bundle_root() -> Path:
    """Directory containing the running script, or the folder with the .exe when frozen (PyInstaller)."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


# Default template next to the app: <bundle>/mplane/mplane.xlsx (works for exe + script).
_MPLANE_MASTER_XLSX = _app_bundle_root() / "mplane" / "mplane.xlsx"


class CallhomeGUI(tk.Tk, ConformanceMixin):
    def __init__(self) -> None:
        super().__init__()
        self.title("O-RAN Netconf")
        self.geometry("1180x820")

        self.proc: subprocess.Popen | None = None
        self.stop_event = threading.Event()
        self.is_running = False
        self.paramiko_client: Any | None = None
        self.paramiko_channel: Any | None = None
        self.session_established = False
        self.manual_send_ready = False
        self.message_tabs: list[dict[str, Any]] = []
        self.remote_cfg_cache: dict[str, str] = {}
        self.log_buffer: list[str] = []
        self.hidden_log_chunks: list[str] = []
        self.hidden_render_active = False
        self.log_lock = threading.Lock()
        self.max_log_lines = 500
        self.gui_log_max_lines_var = tk.StringVar(value="500")
        self.flush_tick = 0
        self.log_window: tk.Toplevel | None = None
        self.log: tk.Text | None = None
        self.rpc_error_list: tk.Listbox | None = None
        self.rpc_event_window: tk.Toplevel | None = None
        self.rpc_event_list: tk.Listbox | None = None
        self.rpc_error_items: list[dict[str, str]] = []
        self.rpc_error_seen: set[str] = set()
        self.rpc_event_window_geometry = "980x520"
        self.log_window_geometry = "1100x720"
        self.remote_control_fifo = "/var/tmp/netconf_tmp/netconf_control.fifo"
        self.script_path = self._resolve_local_script_path()
        self.config_path = Path(__file__).with_name("callhome_gui_config.json")
        self.exec_mode = tk.StringVar(value="remote")
        self.remote_user = tk.StringVar(value="oranuser")
        self.remote_host = tk.StringVar(value="10.0.20.128")
        self.remote_port = tk.StringVar(value="22")
        self.remote_password = tk.StringVar(value="")
        self.remote_script_path = tk.StringVar(value="/var/tmp/miniDU_callhome.sh")
        self.remote_key_path = tk.StringVar(value="")

        self.fields: dict[str, tk.StringVar] = {}
        self.status_var = tk.StringVar(value="Idle")
        self.log_target_hint_var = tk.StringVar(value="")
        self._start_log_path_hint: str = ""
        self._conformance_active_host_log: str | None = None
        self._conformance_last_host_log: str | None = None
        self.send_mode_var = tk.StringVar(value="raw_rpc")
        self.perf_debug_var = tk.BooleanVar(value=False)
        self.auto_xml_send_var = tk.BooleanVar(value=False)
        self.auto_xml_send_done = False
        self.auto_start_var = tk.BooleanVar(value=False)
        self._user_stop_requested = False
        self._auto_start_retry_job: str | None = None
        self._auto_restart_delay_ms = 10_000
        self.perf_stats: dict[str, float] = {}
        self.perf_counts: dict[str, int] = {}
        self.perf_max: dict[str, float] = {}
        self.perf_last: dict[str, float] = {}
        self.perf_last_report = time.time()
        self.perf_text_var = tk.StringVar(value="PERF: off")
        self.trace_seq = 0
        self._all_send_shortcut_pending_until = 0.0
        self._geometry_save_job: str | None = None
        self._log_geometry_save_job: str | None = None
        self._config_save_job: str | None = None
        self._last_find_text = ""
        self.find_replace_window: tk.Toplevel | None = None
        self.find_replace_target: tk.Text | None = None
        self.find_var = tk.StringVar(value="")
        self.replace_var = tk.StringVar(value="")
        self.find_ignore_case_var = tk.BooleanVar(value=True)
        self.find_preview_list: tk.Listbox | None = None
        self.find_matches: list[tuple[str, str, str]] = []
        self.mplane_xlsx_path = tk.StringVar(value=str(_MPLANE_MASTER_XLSX))
        self._mplane_loaded_xlsx_path: str = ""
        self._mplane_rpc_raw: dict[str, str] = {}
        self._mplane_baselines: dict[str, str] = {}
        self.mplane_fields: dict[str, tk.StringVar] = {}
        self.mplane_reload_btn: ttk.Button | None = None
        self.mplane_apply_btn: ttk.Button | None = None
        self.mplane_save_btn: ttk.Button | None = None
        self.mplane_warnings_var = tk.StringVar(value="")
        self.mplane_cc_cell_vars: list[list[tk.StringVar]] = []
        self.mplane_cc_enabled_vars: list[tk.BooleanVar] = []
        # Physical (per-sheet) carrier grids
        self.mplane_phy_enabled: dict[str, list[tk.BooleanVar]] = {}
        self.mplane_phy_prb: dict[str, list[tk.StringVar]] = {}
        self.mplane_phy_center_mhz: dict[str, list[tk.StringVar]] = {}
        self._recent_log_for_session: str = ""
        self._session_watch_job: str | None = None
        self._session_watch_rounds: int = 0
        self._mplane_tables: dict[str, tuple[list[str], list[list[str]]]] = {}
        self._mplane_table_widgets: dict[str, dict[str, Any]] = {}
        self._mplane_table_vars: dict[str, dict[str, Any]] = {}
        self._mplane_selection_sheet: str | None = None
        self._mplane_selection_anchor: tuple[int, int] | None = None
        self._mplane_selection_cells: set[tuple[int, int]] = set()
        self._mplane_drag_selecting: bool = False
        self._mplane_cell_index: dict[str, dict[tk.Entry, tuple[int, int]]] = {}
        self._mplane_undo_stack: dict[str, list[dict[tuple[int, int], str]]] = {}
        self._mplane_redo_stack: dict[str, list[dict[tuple[int, int], str]]] = {}
        self._mplane_sync_lock: bool = False
        self._mplane_last_valid_cell_values: dict[tuple[str, int, int], str] = {}
        self._mplane_prev_selection_sheet: str | None = None
        self._mplane_prev_selection_cells: set[tuple[int, int]] = set()
        self.mplane_find_window: tk.Toplevel | None = None
        self._mplane_find_sheet: str | None = None
        self._mplane_find_hits: list[tuple[int, int, int, int]] = []
        self._mplane_find_idx: int = -1
        self.mplane_cc_on_vars: list[tk.BooleanVar] = []
        self.mplane_debug_log_path = Path(__file__).with_name("mplane_ui_debug.log")
        self.conformance_check_vars: dict[str, tk.BooleanVar] = {}
        self.conformance_scroll_canvas: tk.Canvas | None = None
        self.conformance_list_tree: ttk.Treeview | None = None
        self._conformance_run_active_targets: set[str] | None = None
        self._conformance_run_labels: dict[str, ttk.Label] = {}
        self._conformance_run_busy: bool = False
        self._conformance_stop_idle_wait: bool = False
        self._conformance_cancel_event = threading.Event()
        self._conformance_run_transport_lock = threading.Lock()
        self._conformance_run_ssh_client: Any | None = None
        self._conformance_run_script_channel: Any | None = None
        self._conformance_progress: dict[str, dict[str, Any]] = {}
        self._conformance_script_meta_cache: dict[str, dict[str, Any]] = {}
        self._conformance_detail_win: tk.Toplevel | None = None
        self._conformance_detail_text: tk.Text | None = None
        self._conformance_detail_refresh_job: str | None = None
        self._conformance_detail_last_body: str | None = None
        self._conformance_item_detail_wins: dict[str, tk.Toplevel] = {}
        self._conformance_item_detail_texts: dict[str, tk.Text] = {}
        self._conformance_item_detail_refresh_jobs: dict[str, str | None] = {}
        self._conformance_item_detail_last_body: dict[str, str] = {}
        self._conformance_detail_lines: dict[str, list[str]] = {}
        self._conformance_detail_capture_key: str | None = None
        self._conformance_detail_lock = threading.Lock()
        self._conformance_per_test_settings: dict[str, dict[str, str]] = {}
        self._conformance_oru_boost_active: bool = False
        self._conformance_oru_boost_remote_dir: str | None = None
        self._conformance_detail_run_started_wall: dict[str, str] = {}
        self._conformance_detail_run_started_mono: dict[str, float] = {}
        self._conformance_detail_run_ended_wall: dict[str, str] = {}
        self._conformance_detail_run_ended_mono: dict[str, float] = {}
        self._conformance_auto_sync_scheduled = False
        self._conformance_tab_frame: ttk.Frame | None = None
        self.conformance_path_hint_var: tk.StringVar | None = None
        self.conformance_run_remote_dir_var = tk.StringVar(value=_conf_manifest.CONFORMANCE_REMOTE_DIR)
        self.conformance_run_sw_pkg_var = tk.StringVar(value="")
        self.conformance_run_sw_remote_dir_var = tk.StringVar(value="/var/tmp/conformance/sw_pkg")
        self.conformance_restart_start_after_run_var = tk.BooleanVar(value=True)
        self.conformance_debug_var = tk.BooleanVar(value=True)
        self._conformance_stopped_start_for_run: bool = False
        self.conformance_run_rpc_timeout_var = tk.StringVar(value="30")
        self.conformance_run_idle_timeout_var = tk.StringVar(value="120")
        self.conformance_run_supervision_interval_var = tk.StringVar(value="60")
        self.conformance_run_supervision_reset_cycles_var = tk.StringVar(value="30")
        self.conformance_run_supervision_negative_fail_cycle_var = tk.StringVar(value="3")
        self.conformance_run_conn_delay_var = tk.StringVar(value="3")
        self.conformance_post_listen_wait_var = tk.StringVar(value="0")
        self.conformance_last_run_hint_var = tk.StringVar(value="")
        self._conformance_last_run_snapshot_cache: dict[str, Any] | None = None
        self._conformance_omit_last_run_from_config_save: bool = False
        self._conformance_extra_uploads: list[tuple[str, str]] = []
        self._build_ui()
        self._setup_shortcuts()
        self._load_saved_config()
        self._setup_auto_persist()
        self._fix_mplane_path_after_load()
        self.after(250, self._conformance_refresh_status_labels)
        self.after(300, self._refresh_log_target_hint_line)

    def _fix_mplane_path_after_load(self) -> None:
        """If JSON points to a missing file, use <bundle>/mplane/mplane.xlsx when deployed next to exe."""
        cur = self.mplane_xlsx_path.get().strip()
        if cur:
            self.mplane_xlsx_path.set(self._normalize_mplane_workbook_path(cur))
            cur = self.mplane_xlsx_path.get().strip()
        try:
            exists = bool(cur) and Path(cur).expanduser().resolve().exists()
        except OSError:
            exists = False
        if not exists and _MPLANE_MASTER_XLSX.exists():
            self.mplane_xlsx_path.set(str(_MPLANE_MASTER_XLSX.resolve()))

    def _setup_xml_editor_theme(self, text_widget: tk.Text) -> None:
        self._apply_code_text_theme(text_widget)
        base_font = tkfont.Font(font=text_widget.cget("font"))
        italic_font = base_font.copy()
        italic_font.configure(slant="italic")
        text_widget.tag_configure("xml_tag", foreground="#93c5fd")
        text_widget.tag_configure("xml_attr", foreground="#f0abfc")
        text_widget.tag_configure("xml_comment", foreground="#a5b4fc", font=italic_font)

    @staticmethod
    def _apply_code_text_theme(text_widget: tk.Text) -> None:
        text_widget.configure(
            bg="#0b1220",
            fg="#f8fafc",
            insertbackground="#f8fafc",
            selectbackground="#1d4ed8",
            selectforeground="#f8fafc",
            inactiveselectbackground="#334155",
            font=("Consolas", 11),
            padx=8,
            pady=6,
            spacing1=1,
            spacing2=1,
            spacing3=1,
        )

    def _highlight_xml(self, text_widget: tk.Text) -> None:
        for tag in ("xml_tag", "xml_attr", "xml_comment"):
            text_widget.tag_remove(tag, "1.0", "end")
        content = text_widget.get("1.0", "end-1c")
        if not content:
            return
        for m in re.finditer(r"<!--[\s\S]*?-->", content):
            text_widget.tag_add("xml_comment", f"1.0+{m.start()}c", f"1.0+{m.end()}c")
        for m in re.finditer(r"</?[\w:\-\.]+(?:\s+[\w:\-\.]+(?:\s*=\s*\"[^\"]*\")?)*\s*/?>", content):
            text_widget.tag_add("xml_tag", f"1.0+{m.start()}c", f"1.0+{m.end()}c")
            seg = m.group(0)
            seg_offset = m.start()
            for am in re.finditer(r"\s([\w:\-\.]+)\s*=", seg):
                a_start = seg_offset + am.start(1)
                a_end = seg_offset + am.end(1)
                text_widget.tag_add("xml_attr", f"1.0+{a_start}c", f"1.0+{a_end}c")

    def _on_xml_editor_changed(self, text_widget: tk.Text) -> None:
        self.after_idle(lambda: self._highlight_xml(text_widget))

    def _setup_style(self) -> None:
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except Exception:
            pass
        self.configure(bg="#f1f5f9")
        style.configure(".", font=("Segoe UI", 10), background="#f1f5f9")
        style.configure("TFrame", background="#f1f5f9")
        style.configure("TLabelframe.Label", font=("Segoe UI", 10, "bold"))
        style.configure("Header.TLabel", font=("Segoe UI", 10, "bold"))
        style.configure("Status.TLabel", font=("Segoe UI", 10, "bold"), foreground="#0b5cab")
        style.configure("TButton", padding=7)
        style.configure("Big.TButton", padding=9, font=("Segoe UI", 10, "bold"))
        style.configure("TLabelframe", background="#f8fafc", borderwidth=1, relief="solid")
        style.configure("TNotebook", background="#f1f5f9", borderwidth=0)
        style.configure("TNotebook.Tab", padding=(16, 9), font=("Segoe UI", 10, "bold"))
        style.map(
            "TNotebook.Tab",
            background=[("selected", "#dbeafe"), ("!selected", "#e2e8f0")],
            foreground=[("selected", "#0f172a"), ("!selected", "#334155")],
        )

        # M-Plane combobox styling (compression column): normal/selected/disabled backgrounds.
        style.configure("MPlane.TCombobox", fieldbackground="#ffffff", background="#ffffff", foreground="#111827")
        style.configure("MPlaneSel.TCombobox", fieldbackground="#bfdbfe", background="#bfdbfe", foreground="#111827")
        style.configure("MPlaneDis.TCombobox", fieldbackground="#e5e7eb", background="#e5e7eb", foreground="#6b7280")
        style.map(
            "MPlane.TCombobox",
            fieldbackground=[("disabled", "#e5e7eb")],
            foreground=[("disabled", "#6b7280")],
        )

    def _setup_shortcuts(self) -> None:
        for seq in ("<Control-Shift-KeyPress-N>", "<Control-Shift-n>", "<Control-Shift-N>"):
            self.bind_all(seq, self._on_shortcut_connect, add="+")
        for seq in ("<Control-Shift-KeyPress-B>", "<Control-Shift-b>", "<Control-Shift-B>"):
            self.bind_all(seq, self._on_shortcut_disconnect, add="+")
        for seq in ("<Control-Shift-KeyPress-Q>", "<Control-Shift-q>", "<Control-Shift-Q>"):
            self.bind_all(seq, self._on_shortcut_open_logs, add="+")
        # Extra global logs shortcut (works consistently from M-Plane grid editors too).
        for seq in ("<Control-l>", "<Control-L>", "<Control-KeyPress-l>", "<Control-KeyPress-L>"):
            self.bind_all(seq, self._on_shortcut_open_logs, add="+")
        for seq in ("<Control-Shift-KeyPress-A>", "<Control-Shift-a>", "<Control-Shift-A>"):
            self.bind_all(seq, self._on_shortcut_all_armed, add="+")
        # Some layouts report Ctrl+Shift+Space as Ctrl+Space; bind both.
        for seq in ("<Control-Shift-KeyPress-space>", "<Control-Shift-space>", "<Control-space>"):
            self.bind_all(seq, self._on_shortcut_send, add="+")
        for seq in ("<Control-f>", "<Control-F>", "<Control-KeyPress-f>", "<Control-KeyPress-F>"):
            self.bind_all(seq, self._on_shortcut_find, add="+")
        for seq in ("<Control-h>", "<Control-H>", "<Control-KeyPress-h>", "<Control-KeyPress-H>"):
            self.bind_all(seq, self._on_shortcut_replace, add="+")

    def _setup_auto_persist(self) -> None:
        """Persist common settings automatically on change."""
        watched_vars: list[tk.Variable] = [
            self.remote_user,
            self.remote_host,
            self.remote_port,
            self.remote_password,
            self.remote_script_path,
            self.remote_key_path,
            self.gui_log_max_lines_var,
            self.auto_xml_send_var,
            self.auto_start_var,
            self.perf_debug_var,
            self.mplane_xlsx_path,
        ]
        watched_vars.extend(self.fields.values())
        for var in watched_vars:
            try:
                var.trace_add("write", self._on_any_setting_changed)
            except Exception:
                pass

    def _on_any_setting_changed(self, *_args: Any) -> None:
        if self._config_save_job is not None:
            try:
                self.after_cancel(self._config_save_job)
            except Exception:
                pass
        # Debounce frequent keypress updates.
        self._config_save_job = self.after(700, self._save_current_config)

    def _on_shortcut_connect(self, _event: tk.Event) -> str:
        self.manual_start_script()
        return "break"

    def _on_shortcut_disconnect(self, _event: tk.Event) -> str:
        self.stop_script()
        return "break"

    def _on_shortcut_open_logs(self, _event: tk.Event) -> str:
        self.open_log_window()
        return "break"

    def _on_shortcut_all_armed(self, _event: tk.Event) -> str:
        # Arm "all send" for a short window: Ctrl+Shift+A then Ctrl+Shift+Space.
        self._all_send_shortcut_pending_until = time.time() + 1.2
        self.append_log("[GUI] Shortcut armed: Ctrl+Shift+A+Space => All Tabs Send (1x)\n")
        return "break"

    def _on_shortcut_send(self, _event: tk.Event) -> str:
        if time.time() <= self._all_send_shortcut_pending_until:
            self._all_send_shortcut_pending_until = 0.0
            self.send_all_once()
        else:
            self.send_selected_once()
        return "break"

    def _on_shortcut_find(self, event: tk.Event) -> str | None:
        w = event.widget
        if isinstance(w, tk.Text) and any(tab.get("text") is w for tab in self.message_tabs):
            return self._on_xml_find(event)
        sheet = self._mplane_sheet_from_widget(w)
        if sheet is None and self._mplane_selection_sheet is not None:
            sheet = self._mplane_selection_sheet
        if sheet is None:
            text = self._selected_message_text_widget()
            if text is not None:
                self._open_find_replace_panel(text, focus_replace=False)
                return "break"
            return None
        self._open_mplane_find_panel(sheet, focus_replace=False)
        return "break"

    def _on_shortcut_replace(self, event: tk.Event) -> str | None:
        w = event.widget
        if isinstance(w, tk.Text) and any(tab.get("text") is w for tab in self.message_tabs):
            return self._on_xml_replace(event)
        sheet = self._mplane_sheet_from_widget(w)
        if sheet is None and self._mplane_selection_sheet is not None:
            sheet = self._mplane_selection_sheet
        if sheet is None:
            text = self._selected_message_text_widget()
            if text is not None:
                self._open_find_replace_panel(text, focus_replace=True)
                return "break"
            return None
        self._open_mplane_find_panel(sheet, focus_replace=True)
        return "break"

    def _mplane_sheet_from_widget(self, widget: Any) -> str | None:
        for sheet, mapping in self._mplane_cell_index.items():
            if widget in mapping:
                return sheet
        return None

    def _selected_message_text_widget(self) -> tk.Text | None:
        if not self.message_tabs:
            return None
        try:
            idx = self.msg_notebook.index("current")
        except Exception:
            return None
        if idx < 0 or idx >= len(self.message_tabs):
            return None
        w = self.message_tabs[idx].get("text")
        return w if isinstance(w, tk.Text) else None

    def _target_message_text_widget(self) -> tk.Text | None:
        # Priority: widget under mouse -> focused widget -> currently selected tab.
        try:
            under_mouse = self.winfo_containing(self.winfo_pointerx(), self.winfo_pointery())
        except Exception:
            under_mouse = None
        if isinstance(under_mouse, tk.Text) and any(tab.get("text") is under_mouse for tab in self.message_tabs):
            return under_mouse

        w = self.focus_get()
        if isinstance(w, tk.Text) and any(tab.get("text") is w for tab in self.message_tabs):
            return w

        selected = self._selected_message_text_widget()
        if selected is not None:
            return selected
        return None

    @staticmethod
    def _line_is_commented(line: str) -> bool:
        stripped = line.lstrip()
        return bool(stripped) and stripped.startswith("<!--") and stripped.endswith("-->")

    @staticmethod
    def _comment_line(line: str) -> str:
        if not line.strip():
            return line
        m = re.match(r"^(\s*)", line)
        indent = m.group(1) if m else ""
        body = line[len(indent):]
        return f"{indent}<!-- {body} -->"

    @staticmethod
    def _uncomment_line(line: str) -> str:
        m = re.match(r"^(\s*)<!--\s?(.*?)\s?-->\s*$", line)
        if not m:
            return line
        indent = m.group(1) or ""
        body = m.group(2) or ""
        return f"{indent}{body}"

    def _on_shortcut_toggle_comment(self, event: tk.Event) -> str | None:
        text = event.widget if isinstance(event.widget, tk.Text) else None
        if text is None or not any(tab.get("text") is text for tab in self.message_tabs):
            return None

        sel_ranges = text.tag_ranges("sel")
        if len(sel_ranges) >= 2:
            sel_start = text.index(sel_ranges[0])
            sel_end = text.index(sel_ranges[1])
        else:
            sel_start = text.index("insert")
            sel_end = text.index("insert")

        start_line = int(sel_start.split(".")[0])
        end_line = int(sel_end.split(".")[0])
        if sel_end.endswith(".0") and sel_end != sel_start:
            end_line = max(start_line, end_line - 1)

        start_idx = f"{start_line}.0"
        end_idx = f"{end_line}.end"
        selected_block = text.get(start_idx, end_idx)
        stripped = selected_block.strip()

        # Toggle as ONE XML comment block: <!-- ... -->
        should_uncomment = stripped.startswith("<!--") and stripped.endswith("-->")

        old_auto = bool(int(text.cget("autoseparators")))
        text.configure(autoseparators=False)
        text.edit_separator()
        text.mark_set("_orig_insert", "insert")
        try:
            if should_uncomment:
                m = re.match(r"^\s*<!--\s*([\s\S]*?)\s*-->\s*$", selected_block)
                new_block = m.group(1) if m else selected_block
            else:
                new_block = f"<!--\n{selected_block}\n-->"
            text.delete(start_idx, end_idx)
            text.insert(start_idx, new_block)
            text.mark_set("insert", "_orig_insert")
            text.tag_remove("sel", "1.0", "end")
        finally:
            text.mark_unset("_orig_insert")
            text.edit_separator()
            text.configure(autoseparators=old_auto)

        self._on_xml_editor_changed(text)
        return "break"

    def _on_xml_undo(self, event: tk.Event) -> str | None:
        text = event.widget if isinstance(event.widget, tk.Text) else None
        if text is None or not any(tab.get("text") is text for tab in self.message_tabs):
            return None
        try:
            text.edit_undo()
        except tk.TclError:
            return "break"
        self._on_xml_editor_changed(text)
        return "break"

    def _on_xml_redo(self, event: tk.Event) -> str | None:
        text = event.widget if isinstance(event.widget, tk.Text) else None
        if text is None or not any(tab.get("text") is text for tab in self.message_tabs):
            return None
        try:
            text.edit_redo()
        except tk.TclError:
            return "break"
        self._on_xml_editor_changed(text)
        return "break"

    def _on_xml_find(self, event: tk.Event) -> str | None:
        text = event.widget if isinstance(event.widget, tk.Text) else None
        if text is None or not any(tab.get("text") is text for tab in self.message_tabs):
            return None
        self._open_find_replace_panel(text, focus_replace=False)
        return "break"

    def _on_xml_replace(self, event: tk.Event) -> str | None:
        text = event.widget if isinstance(event.widget, tk.Text) else None
        if text is None or not any(tab.get("text") is text for tab in self.message_tabs):
            return None
        self._open_find_replace_panel(text, focus_replace=True)
        return "break"

    def _open_find_replace_panel(self, text: tk.Text, focus_replace: bool = False) -> None:
        self.find_replace_target = text
        if self.find_replace_window is None or not self.find_replace_window.winfo_exists():
            win = tk.Toplevel(self)
            win.title("Find / Replace")
            win.geometry("700x420")
            self.find_replace_window = win

            top = ttk.Frame(win, padding=10)
            top.pack(fill="x")
            ttk.Label(top, text="Find").grid(row=0, column=0, padx=(0, 8), pady=4, sticky="w")
            find_entry = ttk.Entry(top, textvariable=self.find_var)
            find_entry.grid(row=0, column=1, padx=(0, 8), pady=4, sticky="we")
            ttk.Label(top, text="Replace").grid(row=1, column=0, padx=(0, 8), pady=4, sticky="w")
            replace_entry = ttk.Entry(top, textvariable=self.replace_var)
            replace_entry.grid(row=1, column=1, padx=(0, 8), pady=4, sticky="we")
            ttk.Checkbutton(top, text="Ignore case", variable=self.find_ignore_case_var).grid(
                row=0, column=2, padx=6, pady=4, sticky="w"
            )
            ttk.Button(top, text="Preview", command=self._refresh_find_preview).grid(row=0, column=3, padx=4, pady=4)
            ttk.Button(top, text="Find Next", command=self._find_next_from_panel).grid(row=0, column=4, padx=4, pady=4)
            ttk.Button(top, text="Replace Current", command=self._replace_current_from_panel).grid(
                row=1, column=3, padx=4, pady=4
            )
            ttk.Button(top, text="Replace All", command=self._replace_all_from_panel).grid(row=1, column=4, padx=4, pady=4)
            top.columnconfigure(1, weight=1)

            preview_box = ttk.LabelFrame(win, text="Preview (double-click to jump)", padding=8)
            preview_box.pack(fill="both", expand=True, padx=10, pady=(0, 10))
            self.find_preview_list = tk.Listbox(preview_box, activestyle="none")
            self.find_preview_list.pack(side="left", fill="both", expand=True)
            self.find_preview_list.bind("<Double-Button-1>", self._jump_to_preview_selection)
            self.find_preview_list.bind("<<ListboxSelect>>", self._jump_to_preview_selection)
            pv_scroll = ttk.Scrollbar(preview_box, orient="vertical", command=self.find_preview_list.yview)
            pv_scroll.pack(side="right", fill="y")
            self.find_preview_list.configure(yscrollcommand=pv_scroll.set)

            def _on_close() -> None:
                if self.find_replace_target is not None and self.find_replace_target.winfo_exists():
                    self.find_replace_target.tag_remove("xml_find_hit", "1.0", "end")
                self.find_replace_window = None
                self.find_preview_list = None
                self.find_matches.clear()
                win.destroy()

            win.protocol("WM_DELETE_WINDOW", _on_close)
        else:
            self.find_replace_window.deiconify()
            self.find_replace_window.lift()

        if not self.find_var.get().strip() and self._last_find_text:
            self.find_var.set(self._last_find_text)
        self._refresh_find_preview()
        if focus_replace:
            self.find_replace_window.focus_force()

    def _open_find_panel_for_current_tab(self) -> None:
        text = self._selected_message_text_widget()
        if text is None:
            return
        self._open_find_replace_panel(text, focus_replace=False)

    def _open_replace_panel_for_current_tab(self) -> None:
        text = self._selected_message_text_widget()
        if text is None:
            return
        self._open_find_replace_panel(text, focus_replace=True)

    def _refresh_find_preview(self) -> None:
        text = self.find_replace_target
        if text is None or not text.winfo_exists():
            return
        keyword = self.find_var.get()
        self._last_find_text = keyword.strip() or self._last_find_text
        text.tag_remove("xml_find_hit", "1.0", "end")
        self.find_matches.clear()
        if self.find_preview_list is not None:
            self.find_preview_list.delete(0, "end")
        if not keyword:
            self.status_var.set("Find: enter text")
            return

        nocase = bool(self.find_ignore_case_var.get())
        pos = "1.0"
        while True:
            hit = text.search(keyword, pos, stopindex="end", nocase=nocase)
            if not hit:
                break
            end = f"{hit}+{len(keyword)}c"
            line_no = int(hit.split(".")[0])
            line = text.get(f"{line_no}.0", f"{line_no}.end").strip()
            preview = f"L{line_no}: {line[:180]}"
            self.find_matches.append((hit, end, preview))
            pos = end
            if len(self.find_matches) >= 1000:
                break

        for hit, end, _ in self.find_matches:
            text.tag_add("xml_find_hit", hit, end)
        text.tag_configure("xml_find_hit", background="#78350f", foreground="#fffbeb")
        if self.find_preview_list is not None:
            for _, _, preview in self.find_matches:
                self.find_preview_list.insert("end", preview)
        self.status_var.set(f"Find matches: {len(self.find_matches)}")

    def _jump_to_preview_selection(self, _event: tk.Event | None = None) -> None:
        text = self.find_replace_target
        if text is None or self.find_preview_list is None:
            return
        sel = self.find_preview_list.curselection()
        if not sel:
            return
        idx = int(sel[0])
        if idx < 0 or idx >= len(self.find_matches):
            return
        hit, end, _ = self.find_matches[idx]
        text.tag_remove("sel", "1.0", "end")
        text.tag_add("sel", hit, end)
        text.mark_set("insert", end)
        text.see(hit)

    def _find_next_from_panel(self) -> None:
        text = self.find_replace_target
        if text is None or not text.winfo_exists():
            return
        keyword = self.find_var.get()
        if not keyword:
            return
        nocase = bool(self.find_ignore_case_var.get())
        start_at = text.index("insert")
        hit = text.search(keyword, start_at, stopindex="end", nocase=nocase)
        if not hit:
            hit = text.search(keyword, "1.0", stopindex=start_at, nocase=nocase)
        if not hit:
            self.status_var.set("Find: no match")
            return
        end = f"{hit}+{len(keyword)}c"
        text.tag_remove("sel", "1.0", "end")
        text.tag_add("sel", hit, end)
        text.mark_set("insert", end)
        text.see(hit)
        self._refresh_find_preview()

    def _replace_current_from_panel(self) -> None:
        text = self.find_replace_target
        if text is None or not text.winfo_exists():
            return
        keyword = self.find_var.get()
        if not keyword:
            return
        replace_text = self.replace_var.get()
        sel = text.tag_ranges("sel")
        if len(sel) >= 2:
            s0, s1 = text.index(sel[0]), text.index(sel[1])
            selected = text.get(s0, s1)
            same = selected.lower() == keyword.lower() if self.find_ignore_case_var.get() else selected == keyword
            if same:
                text.delete(s0, s1)
                text.insert(s0, replace_text)
                text.mark_set("insert", f"{s0}+{len(replace_text)}c")
                self._on_xml_editor_changed(text)
                self._refresh_find_preview()
                self._find_next_from_panel()
                return
        self._find_next_from_panel()

    def _replace_all_from_panel(self) -> None:
        text = self.find_replace_target
        if text is None or not text.winfo_exists():
            return
        keyword = self.find_var.get()
        if not keyword:
            return
        replace_text = self.replace_var.get()
        nocase = bool(self.find_ignore_case_var.get())
        count = 0
        pos = "1.0"
        old_auto = bool(int(text.cget("autoseparators")))
        text.configure(autoseparators=False)
        text.edit_separator()
        try:
            while True:
                hit = text.search(keyword, pos, stopindex="end", nocase=nocase)
                if not hit:
                    break
                end = f"{hit}+{len(keyword)}c"
                text.delete(hit, end)
                text.insert(hit, replace_text)
                pos = f"{hit}+{len(replace_text)}c"
                count += 1
        finally:
            text.edit_separator()
            text.configure(autoseparators=old_auto)
        self._on_xml_editor_changed(text)
        self._refresh_find_preview()
        self.status_var.set(f"Replace all: {count} changed")

    @staticmethod
    def _ssh_options(use_password: bool) -> list[str]:
        # Fail fast when auth/network is not ready, instead of hanging UI.
        options = [
            "-o",
            "ConnectTimeout=8",
            "-o",
            "ConnectionAttempts=1",
            "-o",
            "ServerAliveInterval=5",
            "-o",
            "ServerAliveCountMax=1",
        ]
        if not use_password:
            options = ["-o", "BatchMode=yes", *options]
        return options

    def _build_ui(self) -> None:
        self._setup_style()
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=10)

        settings_tab = ttk.Frame(self.notebook)
        scheduler_tab = ttk.Frame(self.notebook)
        mplane_tab = ttk.Frame(self.notebook)
        conformance_tab = ttk.Frame(self.notebook)
        self._conformance_tab_frame = conformance_tab
        shortcuts_tab = ttk.Frame(self.notebook)
        self.notebook.add(settings_tab, text="Settings")
        self.notebook.add(mplane_tab, text="M-Plane Control")
        self.notebook.add(scheduler_tab, text="Netconf Client")
        self.notebook.add(conformance_tab, text="Conformance")
        self.notebook.add(shortcuts_tab, text="Shortcuts")

        mode_frame = ttk.LabelFrame(settings_tab, text="Execution Mode", padding=8)
        mode_frame.pack(fill="x", padx=8, pady=(8, 0))
        ttk.Label(mode_frame, text="Remote Linux (SSH)", style="Header.TLabel").grid(
            row=0, column=0, padx=8, pady=6, sticky="w"
        )
        ttk.Label(
            mode_frame,
            text="Run script on Linux via SSH.",
            foreground="#555555",
        ).grid(row=1, column=0, columnspan=4, padx=8, pady=(0, 6), sticky="w")

        remote_frame = ttk.LabelFrame(settings_tab, text="Remote SSH (used in remote mode)", padding=8)
        remote_frame.pack(fill="x", padx=8, pady=10)
        ttk.Label(remote_frame, text="SSH_USER", width=24).grid(row=0, column=0, padx=8, pady=6, sticky="w")
        ttk.Entry(remote_frame, textvariable=self.remote_user, width=36).grid(row=0, column=1, padx=8, pady=6, sticky="we")
        ttk.Label(remote_frame, text="SSH_HOST", width=24).grid(row=0, column=2, padx=8, pady=6, sticky="w")
        ttk.Entry(remote_frame, textvariable=self.remote_host, width=36).grid(row=0, column=3, padx=8, pady=6, sticky="we")

        ttk.Label(remote_frame, text="SSH_PORT", width=24).grid(row=1, column=0, padx=8, pady=6, sticky="w")
        ttk.Entry(remote_frame, textvariable=self.remote_port, width=36).grid(row=1, column=1, padx=8, pady=6, sticky="we")
        ttk.Label(remote_frame, text="REMOTE_SCRIPT_PATH", width=24).grid(row=1, column=2, padx=8, pady=6, sticky="w")
        ttk.Entry(remote_frame, textvariable=self.remote_script_path, width=36).grid(
            row=1, column=3, padx=8, pady=6, sticky="we"
        )

        ttk.Label(remote_frame, text="SSH_PASSWORD(optional)", width=24).grid(row=2, column=0, padx=8, pady=6, sticky="w")
        ttk.Entry(remote_frame, textvariable=self.remote_password, width=36, show="*").grid(
            row=2, column=1, padx=8, pady=6, sticky="we"
        )
        ttk.Label(
            remote_frame,
            text="SSH_USER/HOST는 Linux 접속 계정 정보입니다. PASSWORD 기반 접속을 사용합니다.",
            foreground="#555555",
        ).grid(row=3, column=0, columnspan=4, padx=8, pady=(0, 6), sticky="w")
        remote_frame.columnconfigure(1, weight=1)
        remote_frame.columnconfigure(3, weight=1)

        form = ttk.LabelFrame(settings_tab, text="ORU Netconf configration", padding=8)
        form.pack(fill="x", padx=8, pady=10)

        defaults = {
            "USER": "oranuser",
            "PASSWORD": "o-ran-password",
            "ALLOWED_IP": "10.0.20.128",
            "LOCAL_IP": "10.0.20.254",
            "CALLHOME_PORT": "4334",
            "NETCONF_PORT": "830",
            "PRODUCT": "nDLPU",
            "LOG_PATH": "/var/tmp/log/nDLPU",
            # O-RU 쉘 자동화 시 sshpass 등으로 CLI-ID/CLI-PW 전달에 사용할 수 있음
            "CLI-ID": "",
            "CLI-PW": "",
        }

        for i, (key, value) in enumerate(defaults.items()):
            ttk.Label(form, text=key, width=24).grid(row=i // 2, column=(i % 2) * 2, padx=8, pady=6, sticky="w")
            var = tk.StringVar(value=value)
            self.fields[key] = var
            show = "*" if key in ("PASSWORD", "CLI-PW") else ""
            ttk.Entry(form, textvariable=var, width=36, show=show).grid(
                row=i // 2,
                column=(i % 2) * 2 + 1,
                padx=8,
                pady=6,
                sticky="we",
            )

        form.columnconfigure(1, weight=1)
        form.columnconfigure(3, weight=1)
        ttk.Label(
            form,
            text=(
                "이 값들은 miniDU_callhome.sh 실행 시 환경변수로 전달됩니다 (예: LOG_PATH, NETCONF_PORT). "
                "CLI-ID/CLI-PW·ALLOWED_IP 등은 장비·자동화 스크립트에서 참조할 수 있습니다."
            ),
            foreground="#555555",
        ).grid(row=(len(defaults) + 1) // 2 + 1, column=0, columnspan=4, padx=8, pady=(0, 6), sticky="w")

        btn_frame = ttk.Frame(settings_tab)
        btn_frame.pack(fill="x", padx=8, pady=8)

        self.start_btn = ttk.Button(btn_frame, text="Start", command=self.manual_start_script, style="Big.TButton")
        self.start_btn.pack(side="left")
        self.stop_btn = ttk.Button(btn_frame, text="Stop", command=self.stop_script, state="disabled", style="Big.TButton")
        self.stop_btn.pack(side="left", padx=8)
        ttk.Button(btn_frame, text="Logs", command=self.open_log_window, style="Big.TButton").pack(side="left", padx=8)
        ttk.Checkbutton(btn_frame, text="Auto XML Send", variable=self.auto_xml_send_var).pack(side="left", padx=10)
        ttk.Checkbutton(
            btn_frame,
            text="Auto Start (after manual Start: retry every 10s if failed/disconnected)",
            variable=self.auto_start_var,
            command=self._on_auto_start_toggled,
        ).pack(side="left", padx=10)
        ttk.Label(btn_frame, text="Status:", style="Header.TLabel").pack(side="right", padx=(0, 6))
        ttk.Label(btn_frame, textvariable=self.status_var, style="Status.TLabel").pack(side="right")

        sched_ctrl = ttk.LabelFrame(scheduler_tab, text="Scheduler Control", padding=8)
        sched_ctrl.pack(fill="x", padx=8, pady=(8, 6))
        self.send_selected_btn = ttk.Button(
            sched_ctrl, text="Selected Tab Send (1x)", command=self.send_selected_once, style="Big.TButton"
        )
        self.send_selected_btn.pack(side="left", padx=(8, 6), pady=8)
        self.send_all_btn = ttk.Button(
            sched_ctrl, text="All Tabs Send (1x)", command=self.send_all_once, style="Big.TButton"
        )
        self.send_all_btn.pack(side="left", padx=6, pady=8)
        ttk.Button(sched_ctrl, text="Logs", command=self.open_log_window, style="Big.TButton").pack(
            side="left", padx=6, pady=8
        )
        ttk.Button(sched_ctrl, text="RPC Events", command=self.open_rpc_event_window, style="Big.TButton").pack(
            side="left", padx=6, pady=8
        )
        self.send_selected_btn.config(state="disabled")
        self.send_all_btn.config(state="disabled")
        ttk.Label(sched_ctrl, text="RAW RPC (ATOM) Mode", style="Header.TLabel").pack(side="left", padx=(16, 6), pady=8)
        ttk.Checkbutton(sched_ctrl, text="Perf DEBUG", variable=self.perf_debug_var).pack(
            side="left", padx=(16, 6), pady=8
        )
        ttk.Label(
            sched_ctrl,
            text="RAW RPC 모드에서는 <rpc>...</rpc> 원문을 그대로 전송합니다.",
            foreground="#555555",
        ).pack(side="right", padx=8)

        msg_tools = ttk.Frame(scheduler_tab)
        msg_tools.pack(fill="x", padx=8, pady=4)
        ttk.Button(msg_tools, text="Add Message Tab", command=self.add_message_tab).pack(side="left")
        ttk.Button(msg_tools, text="Load XML/Text File", command=self.load_file_into_selected_tab).pack(side="left", padx=6)
        ttk.Button(msg_tools, text="Load Excel Tabs", command=self.load_excel_tabs).pack(side="left", padx=6)
        ttk.Button(msg_tools, text="Find", command=self._open_find_panel_for_current_tab).pack(side="left", padx=6)
        ttk.Button(msg_tools, text="Replace", command=self._open_replace_panel_for_current_tab).pack(side="left", padx=6)
        ttk.Button(msg_tools, text="Move Left", command=self.move_selected_tab_left).pack(side="left", padx=6)
        ttk.Button(msg_tools, text="Move Right", command=self.move_selected_tab_right).pack(side="left", padx=6)
        ttk.Button(msg_tools, text="Remove Selected Tab", command=self.remove_selected_message_tab).pack(side="left", padx=6)

        self.msg_notebook = ttk.Notebook(scheduler_tab)
        self.msg_notebook.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        self.add_message_tab(initial_title="MSG-1")

        self._build_mplane_tab(mplane_tab)
        self._build_conformance_tab(conformance_tab)

        shortcuts_frame = ttk.LabelFrame(shortcuts_tab, text="Keyboard Shortcuts", padding=10)
        shortcuts_frame.pack(fill="both", expand=True, padx=8, pady=8)
        shortcut_guide = tk.Text(
            shortcuts_frame,
            wrap="word",
            height=16,
            relief="flat",
            state="normal",
        )
        self._apply_code_text_theme(shortcut_guide)
        shortcut_guide.insert(
            "1.0",
            "\n".join(
                [
                    "Connection / Control",
                    "  - Ctrl + Shift + N : Connect (Start)",
                    "  - Ctrl + Shift + B : Disconnect (Stop)",
                    "  - Settings: Auto Start — manual Start only; if checked, retry every 10s after failure/exit until Stop or unchecked.",
                    "  - Ctrl + Shift + Q : Open Logs Window",
                    "    * Same action as the Logs button in Settings/Netconf Client.",
                    "    * In Logs: set \"GUI keep lines\" + Apply to change how much live log is kept.",
                    "",
                    "XML Send",
                    "  - Ctrl + Shift + Space : Selected Tab Send (1x)",
                    "  - Ctrl + Shift + A + Space : All Tabs Send (1x)",
                    "",
                    "M-Plane Control tab",
                    "  - Reload workbook: ON column is checkboxes; edit grid, Apply when session is ready.",
                    "",
                    "XML Edit",
                    "  - Ctrl + / : Comment/Uncomment selected lines",
                    "  - Ctrl + Z : Undo in XML editor",
                    "  - Ctrl + Shift + Z : Redo in XML editor",
                    "  - Ctrl + F : Open Find/Replace panel (Find)",
                    "  - Ctrl + H : Open Find/Replace panel (Replace)",
                    "    * Panel supports Preview, Find Next, Replace Current, Replace All.",
                    "    * If no selection, toggles comment on current line.",
                    "    * Works on full line units for selected range.",
                ]
            ),
        )
        shortcut_guide.configure(state="disabled")
        shortcut_guide.pack(fill="both", expand=True)

        self.after(120, self._flush_log_buffer)
        self.bind("<Configure>", self._on_main_window_configure)
        self.protocol("WM_DELETE_WINDOW", self.on_close)

        try:
            self.notebook.bind("<<NotebookTabChanged>>", self._on_main_notebook_tab_changed)
        except tk.TclError:
            pass
        self._sync_manual_send_widgets()

    def _on_main_notebook_tab_changed(self, _event: tk.Event | None = None) -> None:
        self.after_idle(self._sync_manual_send_widgets)
        try:
            if self._conformance_tab_frame is not None and self.notebook.select() == str(self._conformance_tab_frame):
                self.after_idle(self._conformance_refresh_status_labels)
        except tk.TclError:
            pass

    @staticmethod
    def _resolve_local_script_path() -> Path:
        name = "miniDU_callhome.sh"
        candidates: list[Path] = []
        app_dir = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
        candidates.append(app_dir / name)
        candidates.append(Path(__file__).resolve().with_name(name))
        meipass = getattr(sys, "_MEIPASS", "")
        if meipass:
            candidates.append(Path(meipass) / name)
        # De-duplicate while preserving order.
        seen: set[str] = set()
        uniq: list[Path] = []
        for p in candidates:
            ps = str(p)
            if ps in seen:
                continue
            seen.add(ps)
            uniq.append(p)
        for p in uniq:
            if p.exists():
                return p
        return uniq[0]

    def _sync_manual_send_widgets(self) -> None:
        ok = bool(self.session_established and self.manual_send_ready)
        st = "normal" if ok else "disabled"
        try:
            self.send_selected_btn.configure(state=st)
            self.send_all_btn.configure(state=st)
        except tk.TclError:
            pass
        pb = self.mplane_apply_btn
        if pb is not None:
            try:
                # Apply for M-Plane just queues tabs into Netconf Client;
                # it should not depend on live connection/session state.
                pb.configure(state="normal")
                try:
                    pb.state(["!disabled"])
                except tk.TclError:
                    pass
            except tk.TclError:
                pass

    def _cancel_session_watch(self) -> None:
        jid = self._session_watch_job
        self._session_watch_job = None
        if jid is not None:
            try:
                self.after_cancel(jid)
            except Exception:
                pass

    def _schedule_session_watch(self) -> None:
        self._cancel_session_watch()
        self._session_watch_rounds = 0
        self._session_watch_job = self.after(2000, self._session_watch_tick)

    def _session_watch_tick(self) -> None:
        self._session_watch_job = None
        if not self.is_running:
            return
        if self.session_established and self.manual_send_ready:
            return
        self._session_watch_rounds += 1
        if self._session_watch_rounds > 100:
            return
        with self.log_lock:
            tail = self._recent_log_for_session
        if tail:
            self._detect_session_established(tail[-24_000:])
        if self.session_established and self.manual_send_ready:
            return
        self._session_watch_job = self.after(2500, self._session_watch_tick)

    def _build_mplane_cc_sheet(self, parent: ttk.Widget) -> None:
        self.mplane_cc_cell_vars.clear()
        self.mplane_cc_enabled_vars.clear()
        wrap = ttk.Frame(parent)
        wrap.pack(fill="both", expand=True)
        canv = tk.Canvas(wrap, highlightthickness=0, bg="#f8fafc", height=280)
        ys = ttk.Scrollbar(wrap, orient="vertical", command=canv.yview)
        xs = ttk.Scrollbar(wrap, orient="horizontal", command=canv.xview)
        canv.configure(yscrollcommand=ys.set, xscrollcommand=xs.set)
        canv.grid(row=0, column=0, sticky="nsew")
        ys.grid(row=0, column=1, sticky="ns")
        xs.grid(row=1, column=0, sticky="ew")
        wrap.rowconfigure(0, weight=1)
        wrap.columnconfigure(0, weight=1)

        inner = ttk.Frame(canv, padding=2)
        win_id = canv.create_window((0, 0), window=inner, anchor="nw")

        def _on_inner_configure(_e: tk.Event | None = None) -> None:
            canv.configure(scrollregion=canv.bbox("all"))

        inner.bind("<Configure>", _on_inner_configure)

        def _on_canvas_configure(e: tk.Event) -> None:
            # Always follow viewport width (do not lock to requested width).
            canv.itemconfigure(win_id, width=max(1, int(e.width)))

        canv.bind("<Configure>", _on_canvas_configure)
        # Ensure initial width sync without requiring manual resize.
        self.after_idle(lambda c=canv, w=win_id: c.itemconfigure(w, width=max(1, c.winfo_width())))

        cols = list(_MPLANE_CC_TREE_COLS)
        entry_cols = cols[1:]
        hdr = ("#", "ON") + tuple(c.replace("_", " ") for c in entry_cols)
        for j, h in enumerate(hdr):
            ttk.Label(inner, text=h, font=("Segoe UI", 9, "bold")).grid(row=0, column=j, padx=1, pady=1, sticky="w")

        for ri in range(_MPLANE_CC_GRID_ROWS):
            ttk.Label(inner, text=str(ri + 1), width=3).grid(row=ri + 1, column=0, padx=1, pady=1, sticky="e")
            bv = tk.BooleanVar(value=False)
            self.mplane_cc_enabled_vars.append(bv)
            ttk.Checkbutton(inner, variable=bv).grid(row=ri + 1, column=1, padx=2, pady=1)
            row_vars: list[tk.StringVar] = []
            for j, _c in enumerate(entry_cols, start=2):
                sv = tk.StringVar(value="")
                row_vars.append(sv)
                ttk.Entry(inner, textvariable=sv, width=9).grid(row=ri + 1, column=j, padx=1, pady=1, sticky="we")
            self.mplane_cc_cell_vars.append(row_vars)
        for j in range(len(hdr)):
            inner.columnconfigure(j, weight=1)

    def _build_mplane_phy_grid(self, parent: ttk.Widget, sheet: str, include_center: bool) -> None:
        """Simple 8-row grid for per-sheet physical params (PRB + optional Center MHz)."""
        self.mplane_phy_enabled[sheet] = []
        self.mplane_phy_prb[sheet] = []
        self.mplane_phy_center_mhz[sheet] = []

        wrap = ttk.Frame(parent)
        wrap.pack(fill="x", expand=False)
        hdr = ["#", "ON", "PRB"]
        if include_center:
            hdr.append("Center(MHz)")
        for j, h in enumerate(hdr):
            ttk.Label(wrap, text=h, font=("Segoe UI", 9, "bold")).grid(row=0, column=j, padx=3, pady=2, sticky="w")

        for ri in range(_MPLANE_PHY_ROWS):
            ttk.Label(wrap, text=str(ri + 1), width=3).grid(row=ri + 1, column=0, padx=3, pady=1, sticky="e")
            bv = tk.BooleanVar(value=False)
            self.mplane_phy_enabled[sheet].append(bv)
            ttk.Checkbutton(wrap, variable=bv).grid(row=ri + 1, column=1, padx=3, pady=1)
            prb = tk.StringVar(value="")
            self.mplane_phy_prb[sheet].append(prb)
            ttk.Entry(wrap, textvariable=prb, width=8).grid(row=ri + 1, column=2, padx=3, pady=1, sticky="we")
            if include_center:
                mhz = tk.StringVar(value="")
                self.mplane_phy_center_mhz[sheet].append(mhz)
                ttk.Entry(wrap, textvariable=mhz, width=12).grid(row=ri + 1, column=3, padx=3, pady=1, sticky="we")
        for j in range(len(hdr)):
            wrap.columnconfigure(j, weight=1 if j in (2, 3) else 0)

    def _mplane_fill_phy_grids_from_rpc(self) -> None:
        """Populate per-sheet PRB/Center grids from current rpc payloads."""
        try:
            import mplane_control as mp
        except Exception:
            return

        def _fill(sheet: str, include_center: bool) -> None:
            enabled = self.mplane_phy_enabled.get(sheet, [])
            prb_vars = self.mplane_phy_prb.get(sheet, [])
            mhz_vars = self.mplane_phy_center_mhz.get(sheet, [])
            for i in range(_MPLANE_PHY_ROWS):
                if i < len(enabled):
                    enabled[i].set(False)
                if i < len(prb_vars):
                    prb_vars[i].set("")
                if include_center and i < len(mhz_vars):
                    mhz_vars[i].set("")

            xml = (self._mplane_rpc_raw.get(sheet) or "").strip()
            if not xml:
                return
            prbs = mp.extract_nth_tag_values(xml, "number-of-prb", limit=_MPLANE_PHY_ROWS)
            centers = mp.extract_nth_tag_values(xml, "center-of-channel-bandwidth", limit=_MPLANE_PHY_ROWS) if include_center else []
            for i in range(_MPLANE_PHY_ROWS):
                pv = prbs[i] if i < len(prbs) else ""
                cv = centers[i] if i < len(centers) else ""
                if i < len(prb_vars):
                    prb_vars[i].set(pv)
                if include_center and i < len(mhz_vars):
                    mhz_vars[i].set(mp.hz_to_mhz_string(cv))
                if i < len(enabled):
                    enabled[i].set(bool(pv) or bool(cv))

        _fill("PDSCH", include_center=True)
        _fill("PUSCH", include_center=True)
        _fill("PRACH", include_center=False)

    def _mplane_fill_cc_sheet(self, cc_rows: list[dict[str, Any]]) -> None:
        for bv in self.mplane_cc_enabled_vars:
            bv.set(False)
        for sv_row in self.mplane_cc_cell_vars:
            for sv in sv_row:
                sv.set("")
        for ri, crow in enumerate(cc_rows[:_MPLANE_CC_GRID_ROWS]):
            self.mplane_cc_enabled_vars[ri].set(CallhomeGUI._mplane_enabled_from_cell(crow.get("enabled")))
            for j, ckey in enumerate(_MPLANE_CC_ENTRY_KEYS):
                raw = crow.get(ckey)
                self.mplane_cc_cell_vars[ri][j].set("" if raw is None else str(raw))

    @staticmethod
    def _mplane_enabled_from_cell(v: Any) -> bool:
        if v is None:
            return False
        if isinstance(v, bool):
            return bool(v)
        s = str(v).strip().upper()
        return s in ("ON", "1", "TRUE", "YES", "Y", "O", "ENABLE", "ENABLED")

    def _mplane_cc_rows_from_grid(self) -> list[dict[str, Any]]:
        rows_out: list[dict[str, Any]] = []
        if not self.mplane_cc_enabled_vars or not self.mplane_cc_cell_vars:
            return rows_out
        for ri in range(_MPLANE_CC_GRID_ROWS):
            if ri >= len(self.mplane_cc_enabled_vars) or ri >= len(self.mplane_cc_cell_vars):
                break
            d: dict[str, Any] = {}
            on = bool(self.mplane_cc_enabled_vars[ri].get())
            d["enabled"] = "ON" if on else "OFF"
            nonempty = on
            for j, key in enumerate(_MPLANE_CC_ENTRY_KEYS):
                if j >= len(self.mplane_cc_cell_vars[ri]):
                    break
                s = (self.mplane_cc_cell_vars[ri][j].get() or "").strip()
                d[key] = s if s else None
                if s:
                    nonempty = True
            if nonempty:
                rows_out.append(d)
        return rows_out

    def _mplane_apply_cc_grid_overrides_to_tables(self) -> None:
        """
        Apply Control-Sheet grid overrides into visible PDSCH/PUSCH/PRACH table vars.
        Targets:
          - PDSCH/PUSCH/PRACH compression <= compression
          - PUSCH t-au-offset <= t_au
          - PUSCH n-ta-offset <= n_ta
          - PDSCH t-da-offset (or t-du-offset header alias) <= t_du
        """
        cc_rows = self._mplane_cc_rows_from_grid()
        if not cc_rows:
            return

        def _col_idx(sheet: str, keys: tuple[str, ...]) -> int | None:
            meta = self._mplane_table_vars.get(sheet) or {}
            headers: list[str] = meta.get("headers") or []
            norm_headers = [self._mplane_tagify_header(h).replace("_", "-") for h in headers]
            for k in keys:
                nk = self._mplane_tagify_header(k).replace("_", "-")
                for i, hh in enumerate(norm_headers):
                    if hh == nk:
                        return i
            return None

        pdsch_comp = _col_idx("PDSCH", ("compression",))
        pusch_comp = _col_idx("PUSCH", ("compression",))
        prach_comp = _col_idx("PRACH", ("compression",))
        pusch_t_au = _col_idx("PUSCH", ("t-au-offset", "t_au_offset"))
        pusch_n_ta = _col_idx("PUSCH", ("n-ta-offset", "n_ta_offset"))
        pdsch_t_du = _col_idx("PDSCH", ("t-da-offset", "t-du-offset", "t_da_offset", "t_du_offset"))

        for i, row in enumerate(cc_rows[:_MPLANE_CC_GRID_ROWS]):
            comp_v = row.get("compression")
            if comp_v is not None and str(comp_v).strip() != "":
                for sheet_name, c_idx in (("PDSCH", pdsch_comp), ("PUSCH", pusch_comp), ("PRACH", prach_comp)):
                    if c_idx is None:
                        continue
                    meta = self._mplane_table_vars.get(sheet_name) or {}
                    vars_rows: list[list[tk.StringVar]] = meta.get("vars") or []
                    if i < len(vars_rows) and c_idx < len(vars_rows[i]):
                        vars_rows[i][c_idx].set(str(comp_v))
            if pusch_t_au is not None:
                v = row.get("t_au")
                if v is not None and str(v).strip() != "":
                    meta = self._mplane_table_vars.get("PUSCH") or {}
                    vars_rows: list[list[tk.StringVar]] = meta.get("vars") or []
                    if i < len(vars_rows) and pusch_t_au < len(vars_rows[i]):
                        vars_rows[i][pusch_t_au].set(str(v))
            if pusch_n_ta is not None:
                v = row.get("n_ta")
                if v is not None and str(v).strip() != "":
                    meta = self._mplane_table_vars.get("PUSCH") or {}
                    vars_rows: list[list[tk.StringVar]] = meta.get("vars") or []
                    if i < len(vars_rows) and pusch_n_ta < len(vars_rows[i]):
                        vars_rows[i][pusch_n_ta].set(str(v))
            if pdsch_t_du is not None:
                v = row.get("t_du")
                if (v is None or str(v).strip() == ""):
                    v = row.get("n_da")
                if v is not None and str(v).strip() != "":
                    meta = self._mplane_table_vars.get("PDSCH") or {}
                    vars_rows: list[list[tk.StringVar]] = meta.get("vars") or []
                    if i < len(vars_rows) and pdsch_t_du < len(vars_rows[i]):
                        vars_rows[i][pdsch_t_du].set(str(v))

    def _build_mplane_tab(self, parent: ttk.Frame) -> None:
        fld_meta: tuple[tuple[str, str], ...] = (
            ("cu_if_name", "Interface name"),
            ("cu_base_if", "base-interface"),
            ("cu_vlan", "VLAN ID"),
            ("cu_mac", "MAC (CU-plane sheet)"),
            ("pe_name", "Processing-element name"),
            ("odu_mac", "O-DU MAC"),
        )
        top = ttk.LabelFrame(parent, text="Workbook (.xlsx)", padding=8)
        top.pack(fill="x", padx=8, pady=(8, 6))
        top.columnconfigure(0, weight=1)
        ttk.Entry(top, textvariable=self.mplane_xlsx_path).grid(row=0, column=0, padx=(0, 8), pady=4, sticky="we")
        ttk.Button(top, text="Browse…", command=self._mplane_browse_workbook).grid(row=0, column=1, pady=4)
        self.mplane_reload_btn = ttk.Button(top, text="Reload from Excel", command=self.reload_mplane_from_excel)
        self.mplane_reload_btn.grid(row=0, column=2, padx=(8, 0), pady=4)
        self.mplane_save_btn = ttk.Button(top, text="Save as Excel", command=self.save_mplane_to_excel)
        self.mplane_save_btn.grid(row=0, column=3, padx=(8, 0), pady=4)

        gridf = ttk.LabelFrame(parent, text="Parameters (edit then Apply; LLTE / LLRE / RSRP_EffBW not loaded)", padding=8)
        gridf.pack(fill="x", padx=8, pady=6)
        self.mplane_fields.clear()
        for i, (key, label) in enumerate(fld_meta):
            self.mplane_fields[key] = tk.StringVar(value="")
            r, c = divmod(i, 2)
            c0 = c * 2
            ttk.Label(gridf, text=label).grid(row=r, column=c0, padx=8, pady=4, sticky="w")
            ttk.Entry(gridf, textvariable=self.mplane_fields[key], width=40).grid(
                row=r, column=c0 + 1, padx=8, pady=4, sticky="we"
            )
        gridf.columnconfigure(1, weight=1)
        gridf.columnconfigure(3, weight=1)

        cc_on_line = ttk.LabelFrame(parent, text="CC ON/OFF (global for DL/UL/PRACH)", padding=6)
        cc_on_line.pack(fill="x", padx=8, pady=(0, 6))
        self.mplane_cc_on_vars = []
        for i in range(_MPLANE_CC_GRID_ROWS):
            bv = tk.BooleanVar(value=True)
            self.mplane_cc_on_vars.append(bv)
            ttk.Checkbutton(
                cc_on_line,
                text=f"CC{i+1}",
                variable=bv,
                command=self._mplane_apply_off_row_styles,
            ).pack(side="left", padx=6)

        # One-screen scroll area (all tables stacked, like Excel view)
        scroll_wrap = ttk.Frame(parent)
        scroll_wrap.pack(fill="both", expand=True, padx=8, pady=6)
        canv = tk.Canvas(scroll_wrap, highlightthickness=0, bg="#f1f5f9")
        ys = ttk.Scrollbar(scroll_wrap, orient="vertical", command=canv.yview)
        xs = ttk.Scrollbar(scroll_wrap, orient="horizontal", command=canv.xview)
        canv.configure(yscrollcommand=ys.set, xscrollcommand=xs.set)
        canv.grid(row=0, column=0, sticky="nsew")
        ys.grid(row=0, column=1, sticky="ns")
        xs.grid(row=1, column=0, sticky="ew")
        scroll_wrap.rowconfigure(0, weight=1)
        scroll_wrap.columnconfigure(0, weight=1)

        inner = ttk.Frame(canv, padding=2)
        win_id = canv.create_window((0, 0), window=inner, anchor="nw")

        def _on_inner_configure(_e: tk.Event | None = None) -> None:
            canv.configure(scrollregion=canv.bbox("all"))

        inner.bind("<Configure>", _on_inner_configure)

        def _on_canvas_configure(e: tk.Event) -> None:
            # Keep content width locked to current viewport width.
            # Using requested width here causes runaway growth after window resize.
            canv.itemconfigure(win_id, width=max(1, int(e.width)))

        canv.bind("<Configure>", _on_canvas_configure)

        # Control-Sheet edit grid is intentionally hidden.
        # Values are written back to Control-Sheet cells during Excel Save.

        # PDSCH/PUSCH/PRACH full tables (editable grid; changes are applied into RPC)
        for sheet, title, include_center in (
            ("PDSCH", "PDSCH (DL) — 전체 표", True),
            ("PUSCH", "PUSCH (UL) — 전체 표", True),
            ("PRACH", "PRACH — 전체 표", False),
        ):
            sec = ttk.LabelFrame(inner, text=title, padding=8)
            sec.pack(fill="both", expand=True, padx=4, pady=(0, 10))
            hint = (
                "값은 Control-Sheet 3/4/5 절 표 셀만 사용 (A열 XML 직접 편집은 무시). "
                "Save → Apply/Conformance."
            )
            ttk.Label(sec, text=hint, foreground="#475569").pack(anchor="w", padx=4, pady=(0, 6))
            grid_host = ttk.Frame(sec)
            grid_host.pack(fill="both", expand=True, padx=2, pady=(8, 2))
            self._mplane_table_widgets[sheet] = {"host": grid_host, "include_center": include_center}

        ttk.Label(parent, textvariable=self.mplane_warnings_var, foreground="#92400e", wraplength=1020).pack(
            fill="x", padx=10, pady=(0, 2)
        )
        btm = ttk.Frame(parent)
        btm.pack(fill="x", padx=8, pady=(4, 8))
        self.mplane_apply_btn = ttk.Button(
            btm,
            text="Apply to device (send in workbook order)",
            command=self.apply_mplane_workbook_once,
            style="Big.TButton",
        )
        self.mplane_apply_btn.pack(side="left")
        ttk.Label(
            btm,
            text="Apply: 디스크 xlsx → M-* 탭 생성. Conformance(31101/31102)도 동일 xlsx 사용.",
            foreground="#64748b",
        ).pack(side="left", padx=14)

    def _mplane_browse_workbook(self) -> None:
        p = filedialog.askopenfilename(filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")])
        if p:
            self.mplane_xlsx_path.set(p)
            self._save_current_config()

    def _store_mplane_workbook_for_gui(self, source_path: str) -> str:
        """
        Keep a GUI-managed copy at ./mplane/mplane_gui_cache.xlsx and return that path.
        Never overwrite user-maintained ./mplane/mplane.xlsx automatically.
        If copy fails, return original source path.
        """
        try:
            src = Path(source_path).expanduser().resolve()
            if not src.exists():
                return source_path
            dst_dir = _app_bundle_root() / "mplane"
            dst_dir.mkdir(parents=True, exist_ok=True)
            user_main = dst_dir / "mplane.xlsx"
            # If user selected the main workbook itself, do not copy/overwrite.
            if src == user_main:
                return str(user_main)
            dst = dst_dir / "mplane_gui_cache.xlsx"
            # Cache copy only; never touches user_main.
            if src != dst:
                shutil.copy2(src, dst)
                self.append_log(f"[GUI] M-Plane workbook cached: {dst}\n")
            return str(dst)
        except Exception as exc:
            self.append_log(f"[GUI] M-Plane cache warning: {exc}\n")
            return source_path

    def _resolve_mplane_template_path(self) -> tuple[Path | None, list[Path]]:
        """
        Pick a template .xlsx for Save.
        Save should always start from the clean master template when available:
        <bundle>/mplane/mplane.xlsx -> <bundle>/mplane.xlsx -> path box -> last Reload path.
        Relative paths resolve against the bundle root (folder containing .exe or the script).
        """
        root = _app_bundle_root()
        ordered: list[Path] = []

        ordered.extend([root / "mplane" / "mplane.xlsx", root / "mplane.xlsx"])

        cfg = self.mplane_xlsx_path.get().strip()
        if cfg:
            p = Path(cfg).expanduser()
            ordered.append((root / p).resolve() if not p.is_absolute() else p.resolve())

        if self._mplane_loaded_xlsx_path:
            p = Path(self._mplane_loaded_xlsx_path).expanduser()
            ordered.append((root / p).resolve() if not p.is_absolute() else p.resolve())

        seen: set[str] = set()
        tried: list[Path] = []
        for p in ordered:
            key = str(p)
            if key in seen:
                continue
            seen.add(key)
            tried.append(p)
            try:
                if p.exists():
                    return p, tried
            except OSError:
                continue
        return None, tried

    def _normalize_mplane_workbook_path(self, raw: str) -> str:
        """Resolve workbook path for Reload/Save: absolute paths, then bundle-root-relative."""
        s = (raw or "").strip()
        if not s:
            return ""
        p = Path(s).expanduser()
        try:
            if p.is_absolute() and p.exists():
                return str(p.resolve())
            root = _app_bundle_root()
            rel = (root / p).resolve() if not p.is_absolute() else p.resolve()
            if rel.exists():
                return str(rel)
            if p.is_absolute():
                return str(p.resolve())
            return str(rel)
        except OSError:
            return s

    def save_mplane_to_excel(self) -> None:
        """Export current M-Plane UI state to a new .xlsx file."""
        self._mplane_sync_visible_widget_values()
        has_rpc = bool(self._mplane_rpc_raw)
        has_table_data = any(
            bool((self._mplane_table_vars.get(s) or {}).get("vars"))
            for s in ("PDSCH", "PUSCH", "PRACH")
        )
        if not has_rpc or not has_table_data:
            messagebox.showerror("M-Plane Save", 'No loaded data. Click "Reload from Excel" first.')
            return

        def _safe_part(text: str, fallback: str) -> str:
            s = (text or "").strip()
            if not s:
                return fallback
            s = re.sub(r'[\\/:*?"<>|]+', "_", s)
            s = re.sub(r"\s+", "_", s)
            s = s.strip("._")
            return s or fallback

        product_raw = ""
        if "PRODUCT" in self.fields:
            product_raw = self.fields["PRODUCT"].get()
        if not (product_raw or "").strip() and "pe_name" in self.mplane_fields:
            product_raw = self.mplane_fields["pe_name"].get()
        if not (product_raw or "").strip() and "cu_if_name" in self.mplane_fields:
            product_raw = self.mplane_fields["cu_if_name"].get()
        mac_raw = ""
        if "cu_mac" in self.mplane_fields:
            mac_raw = self.mplane_fields["cu_mac"].get()
        if not (mac_raw or "").strip() and "odu_mac" in self.mplane_fields:
            mac_raw = self.mplane_fields["odu_mac"].get()
        mac_raw = (mac_raw or "").replace(":", "-")
        base_name = f"{_safe_part(product_raw, 'product')}_{_safe_part(mac_raw, 'mac')}"

        # Save directly under "<exe_or_script_dir>/mplane" with auto-increment suffix.
        if getattr(sys, "frozen", False):
            root_dir = Path(sys.executable).resolve().parent
        else:
            root_dir = Path(__file__).resolve().parent
        target_dir = root_dir / "mplane"
        target_dir.mkdir(parents=True, exist_ok=True)

        candidate = target_dir / f"{base_name}.xlsx"
        if candidate.exists():
            idx = 1
            while True:
                probe = target_dir / f"{base_name}_{idx:03d}.xlsx"
                if not probe.exists():
                    candidate = probe
                    break
                idx += 1
        out_path = str(candidate)
        try:
            import openpyxl
        except Exception as exc:
            messagebox.showerror("M-Plane Save", f"openpyxl is required:\n{exc}")
            return

        # Template resolution for exe/script:
        # 1) Prefer <exe_or_script_dir>/mplane/mplane.xlsx when present (dist deployment).
        # 2) Then path box / last-loaded (relative paths resolve against bundle root).
        template_path, tried_paths = self._resolve_mplane_template_path()
        if template_path is None:
            tried = "\n".join(f"- {p}" for p in tried_paths[:12])
            messagebox.showerror(
                "M-Plane Save",
                "Template workbook not found.\n\n"
                "Place mplane.xlsx next to the app:\n"
                f"  {_app_bundle_root() / 'mplane' / 'mplane.xlsx'}\n\n"
                "Or set the full path in the M-Plane workbook field.\n\n"
                "Checked:\n"
                f"{tried}",
            )
            return
        try:
            wb = openpyxl.load_workbook(template_path)
        except Exception as exc:
            messagebox.showerror("M-Plane Save", f"Failed to open template workbook:\n{exc}")
            return
        self.append_log(f"[GUI] M-Plane Save template: {template_path}\n")

        try:
            import mplane_control as mp
            mp = importlib.reload(mp)
        except Exception as exc:
            messagebox.showerror("M-Plane Save", f"mplane_control load failed:\n{exc}")
            return

        # Build final RPC payloads from current UI values.
        rpc = dict(self._mplane_rpc_raw)
        for _sheet in ("PDSCH", "PUSCH", "PRACH"):
            rpc[_sheet] = mp.uncomment_endpoint_rows((rpc.get(_sheet) or ""), _sheet)
        rpc["ACTIVE"] = mp.uncomment_active_rows((rpc.get("ACTIVE") or ""))
        baselines = dict(self._mplane_baselines)
        live = self._mplane_collect_live()
        for name in mp.SEND_ORDER:
            body = (rpc.get(name) or "").strip()
            if not body:
                continue
            rpc[name] = mp.apply_global_baselines(body, baselines, live)
            if name == "CUplane-interface":
                rpc[name] = mp.ensure_cuplane_interface_fields(rpc[name], live)
            elif name == "Processing-element":
                rpc[name] = mp.ensure_processing_element_fields(rpc[name], live)
        warns = self._mplane_apply_full_tables_to_rpc(rpc)
        pusch_xml = (rpc.get("PUSCH") or "").strip()
        prach_xml = (rpc.get("PRACH") or "").strip()
        if pusch_xml and prach_xml:
            prach_xml, pr_warns = mp.omit_prach_rx_endpoints_present_in_pusch(prach_xml, pusch_xml)
            rpc["PRACH"] = prach_xml
            warns.extend(pr_warns)
        if warns:
            self.append_log(f"[GUI] M-Plane Save warnings: {' | '.join(warns)}\n")
        off_rows = [i + 1 for i, v in enumerate(self.mplane_cc_on_vars) if not v.get()]
        if off_rows:
            for _sheet in ("PDSCH", "PUSCH", "PRACH"):
                rpc[_sheet] = mp.comment_out_endpoint_rows((rpc.get(_sheet) or ""), _sheet, off_rows)
            rpc["ACTIVE"] = mp.comment_out_active_rows((rpc.get("ACTIVE") or ""), off_rows)

        # Control-Sheet upper CC table is untouched.
        # But detail sections (3.PDSCH / 4.PUSCH / 5.PRACH) are replaced from GUI values.
        ws_c = wb["Control-Sheet"] if "Control-Sheet" in wb.sheetnames else None
        if ws_c is not None:
            mp.write_control_sheet_simple_fields(ws_c, live)
            def _safe_set(row: int, col: int, value: Any) -> None:
                try:
                    cell = ws_c.cell(row=row, column=col)
                    if cell.__class__.__name__ == "MergedCell":
                        return
                    cell.value = value
                except Exception:
                    return

            def _gui_rows_for(sheet_name: str) -> tuple[list[str], list[list[str]]]:
                meta = self._mplane_table_vars.get(sheet_name) or {}
                headers: list[str] = meta.get("headers") or []
                vars_rows: list[list[tk.StringVar]] = meta.get("vars") or []
                rows: list[list[str]] = []
                for rr in vars_rows:
                    rows.append([(v.get() or "") for v in rr])
                return headers, rows

            # Header aliases: Control-Sheet label -> GUI table label.
            alias_keys = {
                "end point name": ("end point", "low level tx endpoint", "low level rx endpoint"),
                "link name": ("low level tx endpoint", "low level rx endpoint"),
                "array carrier name": ("tx array carrier", "rx array carrier"),
                "center freq": ("center freq",),
                "channel bw": ("channel bw",),
                "t da offset": ("t-da-offset", "t-du-offset"),
                "t au offset": ("t-au-offset",),
                "n ta offset": ("n-ta-offset",),
                "ul fft sampling offset": ("ul-fft-sampling-offset",),
                "ll fft sampling offset": ("ul-fft-sampling-offset",),
                "number of prb": ("number-of-prb",),
                "eaxc id": ("eaxc-id",),
                "processing element": ("processing-element",),
                "iq bitwidth": ("iq-bitwidth",),
                "gain correction": ("gain-correction",),
                "gain": ("gain",),
                "low level rx endpoint": ("low-level-rx-endpoint",),
                "low level tx endpoint": ("low-level-tx-endpoint",),
            }

            def _find_section_row(section_key: str) -> int | None:
                sk = mp.normalize_header(section_key)
                for rr in range(1, ws_c.max_row + 1):
                    vals = []
                    for cc in range(1, 10):
                        vv = ws_c.cell(row=rr, column=cc).value
                        if vv is not None:
                            vals.append(mp.normalize_header(str(vv)))
                    if sk in "|".join(vals):
                        return rr
                return None

            def _header_columns(header_row: int) -> list[tuple[int, str]]:
                """
                Build normalized header names using marker row (header_row-1) + header_row.
                Ex: "end-point" + "name" => "end-point name".
                """
                out_cols: list[tuple[int, str]] = []
                for cc in range(1, 80):
                    top = ws_c.cell(row=header_row - 1, column=cc).value if header_row - 1 >= 1 else None
                    bot = ws_c.cell(row=header_row, column=cc).value
                    t1 = mp.normalize_header(str(top)) if top is not None else ""
                    t2 = mp.normalize_header(str(bot)) if bot is not None else ""
                    if t1 and t2:
                        hn = f"{t1} {t2}".strip()
                    else:
                        hn = t1 or t2
                    if not hn:
                        continue
                    out_cols.append((cc, hn))
                return out_cols

            def _fill_control_detail_block(header_row: int, sheet_name: str) -> None:
                gui_headers, gui_rows = _gui_rows_for(sheet_name)
                if not gui_headers or not gui_rows:
                    return
                gui_norm = [mp.normalize_header(h) for h in gui_headers]
                ctrl_cols = _header_columns(header_row)
                if not ctrl_cols:
                    return

                # Fill up to GUI row count, keep template layout.
                for i in range(len(gui_rows)):
                    rr = header_row + 1 + i
                    row_vals = gui_rows[i]
                    row_map: dict[str, str] = {}
                    for j, kn in enumerate(gui_norm):
                        if j < len(row_vals):
                            row_map[kn] = row_vals[j]
                    # For 3.PDSCH table, sync CC ON/OFF to column A (e.g. A23~A30).
                    if sheet_name == "PDSCH":
                        on = i < len(self.mplane_cc_on_vars) and self.mplane_cc_on_vars[i].get()
                        _safe_set(rr, 1, "ON" if on else "OFF")
                    # Clear existing formula/text in target columns first,
                    # so template references do not remain.
                    for cc, _nk in ctrl_cols:
                        _safe_set(rr, cc, "")
                    # likely CC label column if present
                    for cc, nk in ctrl_cols:
                        if nk == "cc":
                            _safe_set(rr, cc, f"C{i}")
                            continue
                        # Explicit endpoint/array-carrier mapping for template header variants.
                        if ("rx array carrier" in nk) or ("tx array carrier" in nk):
                            _safe_set(
                                rr,
                                cc,
                                row_map.get("rx array carrier", "")
                                or row_map.get("tx array carrier", "")
                                or row_map.get("rx-array-carrier", "")
                                or row_map.get("tx-array-carrier", "")
                                or row_map.get("rx array carrier", "")
                                or row_map.get("tx array carrier", ""),
                            )
                            continue
                        if ("low level rx endpoint" in nk) or ("low level tx endpoint" in nk):
                            _safe_set(
                                rr,
                                cc,
                                row_map.get("low level rx endpoint", "")
                                or row_map.get("low level tx endpoint", "")
                                or row_map.get("low-level-rx-endpoint", "")
                                or row_map.get("low-level-tx-endpoint", "")
                                or row_map.get("low level rx endpoint", "")
                                or row_map.get("low level tx endpoint", ""),
                            )
                            continue
                        # DL tail columns explicit map: avoid template variant misses.
                        if sheet_name == "PDSCH":
                            if "gain" == nk or nk.endswith(" gain") or " gain " in f" {nk} ":
                                _safe_set(rr, cc, row_map.get("gain", ""))
                                continue
                            if ("t da offset" in nk) or ("t-du" in nk) or ("t du offset" in nk):
                                _safe_set(
                                    rr,
                                    cc,
                                    row_map.get("t da offset", "")
                                    or row_map.get("t du offset", "")
                                    or row_map.get("t-da-offset", "")
                                    or row_map.get("t-du-offset", ""),
                                )
                                continue
                        val = row_map.get(nk, "")
                        if not val and nk in alias_keys:
                            for ak in alias_keys[nk]:
                                val = row_map.get(mp.normalize_header(ak), "")
                                if val:
                                    break
                        if not val:
                            # Fallback for template/header typos and variants.
                            for rk, rv in row_map.items():
                                if rv and (rk in nk or nk in rk):
                                    val = rv
                                    break
                        _safe_set(rr, cc, val)

            # Fill by explicit section anchors to avoid fragile global pattern matching.
            sec_pd = _find_section_row("3. pdsch")
            sec_pu = _find_section_row("4. pusch")
            sec_pr = _find_section_row("5. prach")
            # In this workbook template, detail headers are located at section_row + 3.
            if sec_pd is not None:
                _fill_control_detail_block(sec_pd + 3, "PDSCH")
            if sec_pu is not None:
                _fill_control_detail_block(sec_pu + 3, "PUSCH")
            if sec_pr is not None:
                _fill_control_detail_block(sec_pr + 3, "PRACH")

        # IMPORTANT:
        # Do NOT rewrite whole RPC sheets here. Users rely on the original template layout
        # (tables/formulas/formatting). We only update Control-Sheet values in place.

        try:
            out = Path(out_path)
            wb.save(str(out))
            self.mplane_xlsx_path.set(str(out.resolve()))
            self._mplane_loaded_xlsx_path = str(out.resolve())
            self._save_current_config()
            self.append_log(f"[GUI] M-Plane saved: {out} (workbook path updated — Reload to refresh)\n")
            messagebox.showinfo(
                "M-Plane Save",
                f"Saved:\n{out}\n\nWorkbook path updated. Click \"Reload from Excel\" to load MAC/VLAN from the saved file.",
            )
        except Exception as exc:
            messagebox.showerror("M-Plane Save", f"Failed to save file:\n{exc}")
        finally:
            try:
                wb.close()
            except Exception:
                pass

    def reload_mplane_from_excel(self) -> None:
        path = self._normalize_mplane_workbook_path(self.mplane_xlsx_path.get())
        if not path:
            messagebox.showwarning("M-Plane", "Choose an .xlsx path first.")
            return
        self.mplane_xlsx_path.set(path)
        if self.mplane_reload_btn is not None:
            self.mplane_reload_btn.configure(state="disabled")
        self.mplane_warnings_var.set("Loading workbook...")
        self.append_log(f"[GUI] M-Plane: loading workbook in background: {path}\n")
        threading.Thread(target=self._reload_mplane_from_excel_worker, args=(path,), daemon=True).start()

    def _reload_mplane_from_excel_worker(self, path: str) -> None:
        try:
            import mplane_control as mp
            mp = importlib.reload(mp)
        except ImportError as exc:
            self.after(0, lambda: messagebox.showerror("M-Plane", f"Cannot load mplane_control / openpyxl:\n{exc}"))
            self.after(0, self._mplane_reload_done)
            return
        try:
            rpc, baselines, merged, cc_rows, tables, warnings = mp.load_workbook_payloads(path)
            payload = (dict(rpc), dict(baselines), dict(merged), list(cc_rows), dict(tables), list(warnings))
            self.after(0, lambda p=payload, src=path: self._apply_mplane_reload_result(src, p))
        except FileNotFoundError:
            self.after(0, lambda: messagebox.showerror("M-Plane", f"File not found:\n{path}"))
            self.after(0, self._mplane_reload_done)
        except Exception as exc:
            self.after(0, lambda: messagebox.showerror("M-Plane", f"Failed to load workbook:\n{exc}"))
            self.after(0, self._mplane_reload_done)

    def _apply_mplane_reload_result(
        self,
        path: str,
        payload: tuple[dict[str, str], dict[str, str], dict[str, Any], list[dict[str, Any]], dict[str, tuple[list[str], list[list[str]]]], list[str]],
    ) -> None:
        rpc, baselines, merged, cc_rows, tables, warnings = payload
        self._dbg_mplane(
            "reload_result "
            + ", ".join(
                f"{k}:h={len(v[0])},r={len(v[1])}" for k, v in sorted(tables.items())
            )
        )
        self._mplane_rpc_raw = rpc
        self._mplane_baselines = baselines
        self._mplane_tables = tables
        self._mplane_loaded_xlsx_path = path
        self.mplane_xlsx_path.set(path)
        for key, var in self.mplane_fields.items():
            v = merged.get(key, "")
            var.set("" if v is None else str(v))
        try:
            import mplane_control as mp

            live = {k: (self.mplane_fields[k].get() or "").strip() for k in ("cu_if_name", "cu_base_if", "cu_vlan") if k in self.mplane_fields}
            ifname = mp.resolve_l2vlan_interface_name(live)
            if ifname and "cu_if_name" in self.mplane_fields:
                self.mplane_fields["cu_if_name"].set(ifname)
        except Exception:
            pass
        self._mplane_fill_cc_sheet(cc_rows)
        # Sync Control-Sheet ON/OFF into the active CC toggle line used for comment-out.
        for i, bv in enumerate(self.mplane_cc_on_vars):
            if i < len(cc_rows):
                bv.set(self._mplane_enabled_from_cell(cc_rows[i].get("enabled")))
            else:
                bv.set(True)
        self._mplane_render_full_tables()
        wtext = " | ".join(warnings) if warnings else ""
        self.mplane_warnings_var.set(wtext[:2000] if len(wtext) > 2000 else wtext)
        tbl_info = ", ".join(
            f"{k}={len(v[1])}rows" for k, v in sorted(tables.items()) if v and v[1]
        )
        self.append_log(
            f"[GUI] M-Plane workbook loaded: {path} ({len(cc_rows)} CC row(s); Control-Sheet tables: {tbl_info})\n"
        )
        for w in warnings:
            self.append_log(f"[GUI] M-Plane: {w}\n")
        self._save_current_config()
        self.after_idle(self._sync_manual_send_widgets)
        self._mplane_reload_done()

    def _mplane_reload_done(self) -> None:
        if self.mplane_reload_btn is not None:
            self.mplane_reload_btn.configure(state="normal")

    def _dbg_mplane(self, msg: str) -> None:
        """Write debug traces for M-Plane table rendering."""
        try:
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            with self.mplane_debug_log_path.open("a", encoding="utf-8") as f:
                f.write(f"[{ts}] {msg}\n")
        except Exception:
            pass

    def _mplane_render_full_tables(self) -> None:
        """Render extracted Excel tables for PDSCH/PUSCH/PRACH in the one-screen layout (editable)."""
        self._dbg_mplane("render_start")
        for sheet, meta in self._mplane_table_widgets.items():
            host = meta.get("host")
            if host is None or not host.winfo_exists():
                continue
            for child in list(host.winfo_children()):
                try:
                    child.destroy()
                except Exception:
                    pass
            headers, rows = self._mplane_tables.get(sheet, ([], []))
            self._mplane_table_vars[sheet] = {"headers": headers, "vars": [], "entries": []}
            self._mplane_cell_index[sheet] = {}
            self._mplane_undo_stack[sheet] = []
            self._mplane_redo_stack[sheet] = []
            self._dbg_mplane(
                f"{sheet}: headers={len(headers)} rows={len(rows)} first_row_len={(len(rows[0]) if rows else 0)}"
            )
            if not headers:
                ttk.Label(host, text="(표를 찾지 못했습니다. 엑셀에서 헤더 행에 end_point / number-of-prb 가 있는지 확인)", foreground="#b45309").pack(anchor="w")
                continue
            ttk.Label(host, text=f"rows: {len(rows)}", foreground="#64748b").pack(anchor="w", padx=2, pady=(0, 2))
            # Use direct frame-grid rendering (no nested canvas) to avoid width=1 timing bugs.
            inner = ttk.Frame(host, padding=1)
            inner.pack(fill="x", expand=True)

            # Header row with row numbers
            ttk.Label(inner, text="#", font=("Segoe UI", 9, "bold")).grid(row=0, column=0, padx=1, pady=1, sticky="w")
            for j, h in enumerate(headers):
                ttk.Label(inner, text=h, font=("Segoe UI", 9, "bold")).grid(
                    row=0, column=j + 1, padx=1, pady=1, sticky="ew"
                )

            # Data rows (editable)
            for i, r in enumerate(rows):
                ttk.Label(inner, text=str(i + 1), foreground="#64748b", font=("Segoe UI", 9)).grid(
                    row=i + 1, column=0, padx=1, pady=1, sticky="e"
                )
                row_vars: list[tk.StringVar] = []
                row_entries: list[tk.Widget] = []
                c_comp = self._mplane_col_index(sheet, "compression")
                c_iq = self._mplane_col_index(sheet, "iq-bitwidth")
                for j in range(len(headers)):
                    v = r[j] if j < len(r) else ""
                    if c_comp is not None and j == c_comp:
                        iqv = ""
                        if c_iq is not None and c_iq < len(r):
                            iqv = (r[c_iq] or "").strip()
                        if iqv in {"16", "16.0"}:
                            v = "no comp"
                        else:
                            if (v or "").strip().lower() in {"", "no comp", "nocomp", "static"}:
                                v = "BFP"
                            nv = self._mplane_normalize_compression_value(v)
                            if nv is not None:
                                v = nv
                    sv = tk.StringVar(value=v)
                    sv.trace_add(
                        "write",
                        lambda *_a, s=sheet, rr=i, cc=j: self._mplane_on_cell_var_changed(s, rr, cc),
                    )
                    if c_iq is not None and j == c_iq:
                        sv.trace_add("write", lambda *_a: self.after_idle(self._mplane_apply_off_row_styles))
                    row_vars.append(sv)
                    self._mplane_last_valid_cell_values[(sheet, i, j)] = v
                    # Compression cell: editable dropdown + direct typing.
                    if c_comp is not None and j == c_comp:
                        ent: tk.Widget = ttk.Combobox(
                            inner,
                            textvariable=sv,
                            values=("BFP", "exponent"),
                            width=8,
                            state="normal",
                        )
                    else:
                        # Use compact tk.Entry so DL/UL can shrink like PRACH with many columns.
                        ent = tk.Entry(
                            inner,
                            textvariable=sv,
                            width=6,
                            font=("Consolas", 10),
                            relief="solid",
                            borderwidth=1,
                        )
                    # Keep endpoint key immutable to prevent accidental carrier row loss.
                    if j == 0:
                        try:
                            ent.configure(state="readonly", readonlybackground="#eef2ff")
                        except Exception:
                            ent.configure(state="disabled")
                    ent.grid(row=i + 1, column=j + 1, padx=1, pady=1, sticky="we")
                    ent.bind(
                        "<Button-1>",
                        lambda e, s=sheet, rr=i, cc=j: self._mplane_on_cell_press(e, s, rr, cc),
                    )
                    ent.bind(
                        "<B1-Motion>",
                        lambda e, s=sheet, rr=i, cc=j: self._mplane_on_cell_drag(e, s, rr, cc),
                    )
                    ent.bind(
                        "<Enter>",
                        lambda e, s=sheet, rr=i, cc=j: self._mplane_on_cell_enter(e, s, rr, cc),
                    )
                    ent.bind("<ButtonRelease-1>", self._mplane_on_cell_release)
                    ent.bind("<Control-v>", lambda e, s=sheet: self._mplane_on_paste(e, s))
                    ent.bind("<Control-V>", lambda e, s=sheet: self._mplane_on_paste(e, s))
                    ent.bind("<Control-c>", lambda e, s=sheet: self._mplane_on_copy(e, s))
                    ent.bind("<Control-C>", lambda e, s=sheet: self._mplane_on_copy(e, s))
                    ent.bind("<Control-z>", lambda e, s=sheet: self._mplane_on_undo(e, s))
                    ent.bind("<Control-Z>", lambda e, s=sheet: self._mplane_on_undo(e, s))
                    ent.bind("<Control-Shift-z>", lambda e, s=sheet: self._mplane_on_redo(e, s))
                    ent.bind("<Control-Shift-Z>", lambda e, s=sheet: self._mplane_on_redo(e, s))
                    ent.bind("<Control-y>", lambda e, s=sheet: self._mplane_on_redo(e, s))
                    ent.bind("<Control-Y>", lambda e, s=sheet: self._mplane_on_redo(e, s))
                    ent.bind("<Delete>", lambda e, s=sheet: self._mplane_on_multi_clear(e, s))
                    ent.bind("<BackSpace>", lambda e, s=sheet: self._mplane_on_multi_clear(e, s))
                    ent.bind("<FocusOut>", lambda _e, s=sheet, rr=i, cc=j: self._mplane_on_cell_focus_out(s, rr, cc))
                    ent.bind("<KeyRelease>", lambda _e, s=sheet, rr=i, cc=j: self._mplane_on_cell_keyrelease(s, rr, cc))
                    row_entries.append(ent)
                    self._mplane_cell_index[sheet][ent] = (i, j)
                self._mplane_table_vars[sheet]["vars"].append(row_vars)
                self._mplane_table_vars[sheet]["entries"].append(row_entries)
            # Responsive column sizing in direct-grid mode.
            # Keep a very small minimum so tables do not hard-clip on narrow windows.
            inner.columnconfigure(0, weight=0, minsize=20)
            for j in range(1, len(headers) + 1):
                inner.columnconfigure(j, weight=1, minsize=6, uniform=f"mplane_{sheet}")
            self._dbg_mplane(
                f"{sheet}: rendered var_rows={len(self._mplane_table_vars[sheet]['vars'])} "
                f"inner_w={inner.winfo_width()}"
            )
        self._mplane_apply_off_row_styles()
        self.bind_all("<B1-Motion>", self._mplane_on_global_drag, add="+")
        self.bind_all("<ButtonRelease-1>", self._mplane_on_global_release, add="+")
        self.bind_all("<Control-c>", self._mplane_on_global_copy, add="+")
        self.bind_all("<Control-C>", self._mplane_on_global_copy, add="+")
        self.bind_all("<Control-v>", self._mplane_on_global_paste, add="+")
        self.bind_all("<Control-V>", self._mplane_on_global_paste, add="+")
        self._dbg_mplane("render_done")

    def _mplane_apply_off_row_styles(self) -> None:
        """When CC is OFF, lock and shade corresponding table rows."""
        for sheet in ("PDSCH", "PUSCH", "PRACH"):
            meta = self._mplane_table_vars.get(sheet) or {}
            entries: list[list[tk.Widget]] = meta.get("entries") or []
            c_comp = self._mplane_col_index(sheet, "compression")
            for r_idx, row_entries in enumerate(entries):
                is_off = r_idx < len(self.mplane_cc_on_vars) and (not self.mplane_cc_on_vars[r_idx].get())
                iq16 = self._mplane_row_iq_is_16(sheet, r_idx)
                for c_idx, ent in enumerate(row_entries):
                    if is_off:
                        self._mplane_set_widget_state(ent, "readonly", "#e5e7eb")
                    else:
                        if c_idx == 0:
                            self._mplane_set_widget_state(ent, "readonly", "#eef2ff")
                        else:
                            if c_comp is not None and c_idx == c_comp and iq16:
                                self._mplane_set_widget_state(ent, "readonly", "#e5e7eb")
                            else:
                                self._mplane_set_widget_state(ent, "normal", "#ffffff")
        self._mplane_refresh_selection_visuals()

    def _mplane_on_cell_press(self, event: tk.Event, sheet: str, row: int, col: int) -> None:
        self._mplane_drag_selecting = True
        self._mplane_selection_sheet = sheet
        self._mplane_selection_anchor = (row, col)
        self._mplane_select_range(sheet, row, col, row, col)
        try:
            w = event.widget
            if isinstance(w, tk.Entry):
                w.focus_set()
                w.icursor("end")
        except Exception:
            pass

    def _mplane_on_cell_drag(self, _event: tk.Event, sheet: str, row: int, col: int) -> None:
        anchor = self._mplane_selection_anchor
        if not anchor or self._mplane_selection_sheet != sheet:
            return
        self._mplane_select_range(sheet, anchor[0], anchor[1], row, col)

    def _mplane_on_cell_enter(self, event: tk.Event, sheet: str, row: int, col: int) -> None:
        # While mouse button is pressed, extending selection must work across cell boundaries.
        if not self._mplane_drag_selecting:
            return
        if int(getattr(event, "state", 0)) & 0x0100 == 0:
            return
        anchor = self._mplane_selection_anchor
        if not anchor or self._mplane_selection_sheet != sheet:
            return
        self._mplane_select_range(sheet, anchor[0], anchor[1], row, col)

    def _mplane_on_cell_release(self, _event: tk.Event) -> None:
        self._mplane_drag_selecting = False

    def _mplane_on_global_drag(self, _event: tk.Event) -> None:
        """Global drag tracking so selection continues across entry boundaries."""
        if not self._mplane_drag_selecting:
            return
        try:
            ptr_x, ptr_y = self.winfo_pointerx(), self.winfo_pointery()
            target = self.winfo_containing(ptr_x, ptr_y)
        except Exception:
            # Combobox popdown can raise KeyError("popdown") in winfo_containing.
            return
        if target is None:
            return
        for sheet, mapping in self._mplane_cell_index.items():
            cell = target
            while cell is not None:
                if cell in mapping:
                    anchor = self._mplane_selection_anchor
                    if not anchor or self._mplane_selection_sheet != sheet:
                        return
                    rr, cc = mapping[cell]
                    self._mplane_select_range(sheet, anchor[0], anchor[1], rr, cc)
                    return
                parent = getattr(cell, "master", None)
                if parent is None:
                    break
                cell = parent

    def _mplane_on_global_release(self, _event: tk.Event) -> None:
        self._mplane_drag_selecting = False

    def _mplane_on_global_copy(self, event: tk.Event) -> str | None:
        w = event.widget
        if isinstance(w, tk.Text) and any(tab.get("text") is w for tab in self.message_tabs):
            return None
        sheet = self._mplane_selection_sheet
        if not sheet or not self._mplane_selection_cells:
            return
        return self._mplane_on_copy(event, sheet)

    def _mplane_on_global_paste(self, event: tk.Event) -> str | None:
        w = event.widget
        if isinstance(w, tk.Text) and any(tab.get("text") is w for tab in self.message_tabs):
            return None
        sheet = self._mplane_selection_sheet
        if not sheet:
            return
        return self._mplane_on_paste(event, sheet)

    def _open_mplane_find_panel(self, sheet: str, focus_replace: bool = False) -> None:
        self._mplane_find_sheet = sheet
        self._mplane_find_hits = []
        self._mplane_find_idx = -1
        if self.mplane_find_window is None or not self.mplane_find_window.winfo_exists():
            win = tk.Toplevel(self)
            win.title("M-Plane Find / Replace")
            win.geometry("520x180")
            self.mplane_find_window = win

            top = ttk.Frame(win, padding=10)
            top.pack(fill="both", expand=True)
            ttk.Label(top, text="Find").grid(row=0, column=0, padx=(0, 8), pady=6, sticky="w")
            find_entry = ttk.Entry(top, textvariable=self.find_var)
            find_entry.grid(row=0, column=1, padx=(0, 8), pady=6, sticky="we")
            ttk.Label(top, text="Replace").grid(row=1, column=0, padx=(0, 8), pady=6, sticky="w")
            replace_entry = ttk.Entry(top, textvariable=self.replace_var)
            replace_entry.grid(row=1, column=1, padx=(0, 8), pady=6, sticky="we")
            ttk.Checkbutton(top, text="Ignore case", variable=self.find_ignore_case_var).grid(
                row=0, column=2, padx=6, pady=6, sticky="w"
            )
            ttk.Button(top, text="Find Next", command=self._mplane_find_next).grid(row=0, column=3, padx=4, pady=6)
            ttk.Button(top, text="Replace Current", command=self._mplane_replace_current).grid(row=1, column=2, padx=4, pady=6)
            ttk.Button(top, text="Replace All", command=self._mplane_replace_all).grid(row=1, column=3, padx=4, pady=6)
            top.columnconfigure(1, weight=1)

            def _on_close() -> None:
                self.mplane_find_window = None
                self._mplane_find_hits = []
                self._mplane_find_idx = -1
                win.destroy()

            win.protocol("WM_DELETE_WINDOW", _on_close)
        else:
            self.mplane_find_window.deiconify()
            self.mplane_find_window.lift()
        if focus_replace:
            self.mplane_find_window.focus_force()

    def _mplane_collect_find_hits(self) -> list[tuple[int, int, int, int]]:
        sheet = self._mplane_find_sheet
        if not sheet:
            return []
        keyword = self.find_var.get()
        if not keyword:
            return []
        nocase = bool(self.find_ignore_case_var.get())
        key_cmp = keyword.casefold() if nocase else keyword
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        hits: list[tuple[int, int, int, int]] = []
        for rr, row in enumerate(var_rows):
            for cc, sv in enumerate(row):
                if cc == 0:
                    continue
                txt = sv.get() or ""
                src = txt.casefold() if nocase else txt
                pos = src.find(key_cmp)
                if pos >= 0:
                    hits.append((rr, cc, pos, pos + len(keyword)))
        return hits

    def _mplane_focus_hit(self, hit: tuple[int, int, int, int]) -> None:
        sheet = self._mplane_find_sheet
        if not sheet:
            return
        rr, cc, s_idx, e_idx = hit
        self._mplane_select_range(sheet, rr, cc, rr, cc)
        meta = self._mplane_table_vars.get(sheet) or {}
        entries: list[list[tk.Widget]] = meta.get("entries") or []
        if rr >= len(entries) or cc >= len(entries[rr]):
            return
        ent = entries[rr][cc]
        try:
            ent.focus_set()
            if hasattr(ent, "icursor"):
                ent.icursor(e_idx)  # type: ignore[attr-defined]
            if hasattr(ent, "selection_clear"):
                ent.selection_clear()  # type: ignore[attr-defined]
            if hasattr(ent, "selection_range"):
                ent.selection_range(s_idx, e_idx)  # type: ignore[attr-defined]
        except Exception:
            pass

    def _mplane_find_next(self) -> None:
        hits = self._mplane_collect_find_hits()
        self._mplane_find_hits = hits
        if not hits:
            self.status_var.set("M-Plane Find: no matches")
            return
        self._mplane_find_idx = (self._mplane_find_idx + 1) % len(hits)
        self._mplane_focus_hit(hits[self._mplane_find_idx])
        self.status_var.set(f"M-Plane Find: {self._mplane_find_idx + 1}/{len(hits)}")

    def _mplane_replace_current(self) -> None:
        sheet = self._mplane_find_sheet
        if not sheet:
            return
        hits = self._mplane_collect_find_hits()
        if not hits:
            self.status_var.set("M-Plane Replace: no matches")
            return
        idx = self._mplane_find_idx if 0 <= self._mplane_find_idx < len(hits) else 0
        rr, cc, _s_idx, _e_idx = hits[idx]
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        if rr >= len(var_rows) or cc >= len(var_rows[rr]):
            return
        before = self._mplane_snapshot_values(sheet)
        keyword = self.find_var.get()
        repl = self.replace_var.get()
        nocase = bool(self.find_ignore_case_var.get())
        text = var_rows[rr][cc].get()
        if nocase:
            pattern = re.compile(re.escape(keyword), re.IGNORECASE)
            new_text, n = pattern.subn(repl, text, count=1)
        else:
            new_text = text.replace(keyword, repl, 1)
            n = 1 if new_text != text else 0
        if n > 0:
            var_rows[rr][cc].set(new_text)
            self._mplane_push_undo(sheet, before)
        self._mplane_find_hits = []
        self._mplane_find_idx = -1
        self._mplane_find_next()

    def _mplane_replace_all(self) -> None:
        sheet = self._mplane_find_sheet
        if not sheet:
            return
        keyword = self.find_var.get()
        if not keyword:
            return
        repl = self.replace_var.get()
        nocase = bool(self.find_ignore_case_var.get())
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        before = self._mplane_snapshot_values(sheet)
        changed = 0
        for rr, row in enumerate(var_rows):
            for cc, sv in enumerate(row):
                if cc == 0:
                    continue
                txt = sv.get() or ""
                if nocase:
                    pattern = re.compile(re.escape(keyword), re.IGNORECASE)
                    new_txt, n = pattern.subn(repl, txt)
                else:
                    n = txt.count(keyword)
                    new_txt = txt.replace(keyword, repl)
                if n > 0:
                    sv.set(new_txt)
                    changed += n
        if changed:
            self._mplane_push_undo(sheet, before)
        self.status_var.set(f"M-Plane Replace All: {changed} replacement(s)")

    def _mplane_select_range(self, sheet: str, r1: int, c1: int, r2: int, c2: int) -> None:
        self._mplane_selection_sheet = sheet
        r_min, r_max = sorted((r1, r2))
        c_min, c_max = sorted((c1, c2))
        new_sel: set[tuple[int, int]] = set()
        for rr in range(r_min, r_max + 1):
            for cc in range(c_min, c_max + 1):
                new_sel.add((rr, cc))
        if self._mplane_selection_sheet == self._mplane_prev_selection_sheet and new_sel == self._mplane_selection_cells:
            return
        self._mplane_selection_cells = new_sel
        self._mplane_refresh_selection_visuals()

    def _mplane_refresh_selection_visuals(self) -> None:
        old_sheet = self._mplane_prev_selection_sheet
        old_cells = self._mplane_prev_selection_cells
        new_sheet = self._mplane_selection_sheet
        new_cells = set(self._mplane_selection_cells)

        # Fast path: update only changed cells when staying on same sheet.
        if old_sheet == new_sheet and new_sheet is not None:
            to_clear = old_cells - new_cells
            to_set = new_cells - old_cells
            for rr, cc in to_clear:
                self._mplane_style_cell(new_sheet, rr, cc, selected=False)
            for rr, cc in to_set:
                self._mplane_style_cell(new_sheet, rr, cc, selected=True)
        else:
            # Sheet changed (or initial render): repaint previous + current selection sets only.
            if old_sheet is not None:
                for rr, cc in old_cells:
                    self._mplane_style_cell(old_sheet, rr, cc, selected=False)
            if new_sheet is not None:
                for rr, cc in new_cells:
                    self._mplane_style_cell(new_sheet, rr, cc, selected=True)

        self._mplane_prev_selection_sheet = new_sheet
        self._mplane_prev_selection_cells = new_cells

    def _mplane_style_cell(self, sheet: str, r_idx: int, c_idx: int, selected: bool) -> None:
        meta = self._mplane_table_vars.get(sheet) or {}
        entries: list[list[tk.Widget]] = meta.get("entries") or []
        if r_idx >= len(entries) or c_idx >= len(entries[r_idx]):
            return
        ent = entries[r_idx][c_idx]
        is_off = r_idx < len(self.mplane_cc_on_vars) and (not self.mplane_cc_on_vars[r_idx].get())
        c_comp = self._mplane_col_index(sheet, "compression")
        iq16 = self._mplane_row_iq_is_16(sheet, r_idx)
        if selected:
            self._mplane_set_widget_state(ent, "normal", "#bfdbfe")
            return
        if is_off:
            self._mplane_set_widget_state(ent, "readonly", "#e5e7eb")
        elif c_idx == 0:
            self._mplane_set_widget_state(ent, "readonly", "#eef2ff")
        elif c_comp is not None and c_idx == c_comp and iq16:
            self._mplane_set_widget_state(ent, "readonly", "#e5e7eb")
        else:
            self._mplane_set_widget_state(ent, "normal", "#ffffff")

    def _mplane_set_widget_state(self, widget: tk.Widget, mode: str, bg: str) -> None:
        """Best-effort state/style setter for Entry and Combobox widgets."""
        if isinstance(widget, ttk.Combobox):
            try:
                if mode == "readonly":
                    widget.configure(state="disabled", style="MPlaneDis.TCombobox")
                else:
                    widget.configure(
                        state="normal",
                        style="MPlaneSel.TCombobox" if bg == "#bfdbfe" else "MPlane.TCombobox",
                    )
            except Exception:
                pass
            return
        try:
            if mode == "readonly":
                widget.configure(state="readonly", readonlybackground=bg, fg="#111827", disabledforeground="#111827")
            else:
                widget.configure(state="normal", bg=bg, fg="#111827", disabledforeground="#111827")
        except Exception:
            pass

    def _mplane_on_paste(self, event: tk.Event, sheet: str) -> str:
        before = self._mplane_snapshot_values(sheet)
        try:
            clip = self.clipboard_get()
        except Exception:
            return "break"
        if not clip:
            return "break"
        rows_in = [ln for ln in clip.replace("\r\n", "\n").replace("\r", "\n").split("\n") if ln != ""]
        if not rows_in:
            return "break"
        matrix = [line.split("\t") for line in rows_in]

        # Excel-like: if clipboard is single cell and a range is selected,
        # fill the entire selection.
        if (
            self._mplane_selection_sheet == sheet
            and len(self._mplane_selection_cells) > 1
            and len(matrix) == 1
            and len(matrix[0]) == 1
        ):
            meta = self._mplane_table_vars.get(sheet) or {}
            var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
            entries: list[list[tk.Widget]] = meta.get("entries") or []
            val0 = matrix[0][0]
            for rr, cc in sorted(self._mplane_selection_cells):
                if rr >= len(var_rows) or cc >= len(var_rows[rr]) or cc == 0:
                    continue
                if rr < len(self.mplane_cc_on_vars) and (not self.mplane_cc_on_vars[rr].get()):
                    continue
                ent = entries[rr][cc] if rr < len(entries) and cc < len(entries[rr]) else None
                if ent is not None and str(ent.cget("state")) in {"readonly", "disabled"}:
                    continue
                vset = val0
                if self._mplane_is_compression_col(sheet, cc):
                    if self._mplane_row_iq_is_16(sheet, rr):
                        continue
                    norm = self._mplane_normalize_compression_value(vset)
                    if norm is None:
                        continue
                    vset = norm
                var_rows[rr][cc].set(vset)
                self._mplane_last_valid_cell_values[(sheet, rr, cc)] = vset
            self._mplane_push_undo(sheet, before)
            return "break"

        start_row, start_col = 0, 0
        if self._mplane_selection_sheet == sheet and self._mplane_selection_cells:
            start_row = min(r for r, _ in self._mplane_selection_cells)
            start_col = min(c for _, c in self._mplane_selection_cells)
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        entries: list[list[tk.Widget]] = meta.get("entries") or []
        for dr, vals in enumerate(matrix):
            rr = start_row + dr
            if rr >= len(var_rows):
                break
            for dc, val in enumerate(vals):
                cc = start_col + dc
                if cc >= len(var_rows[rr]):
                    break
                if cc == 0:
                    continue
                if rr < len(self.mplane_cc_on_vars) and (not self.mplane_cc_on_vars[rr].get()):
                    continue
                ent = entries[rr][cc] if rr < len(entries) and cc < len(entries[rr]) else None
                if ent is not None and str(ent.cget("state")) == "readonly":
                    continue
                if self._mplane_is_compression_col(sheet, cc):
                    if self._mplane_row_iq_is_16(sheet, rr):
                        continue
                    norm = self._mplane_normalize_compression_value(val)
                    if norm is None:
                        # Invalid pasted compression value: ignore and keep prior value.
                        continue
                    val = norm
                var_rows[rr][cc].set(val)
                self._mplane_last_valid_cell_values[(sheet, rr, cc)] = val
        end_row = min(len(var_rows) - 1, start_row + len(matrix) - 1)
        end_col = start_col
        for vals in matrix:
            end_col = max(end_col, start_col + max(0, len(vals) - 1))
        if var_rows:
            end_col = min(len(var_rows[0]) - 1, end_col)
        self._mplane_select_range(sheet, start_row, start_col, end_row, end_col)
        self._mplane_push_undo(sheet, before)
        return "break"

    def _mplane_on_multi_clear(self, event: tk.Event, sheet: str) -> str | None:
        if self._mplane_selection_sheet != sheet or len(self._mplane_selection_cells) <= 1:
            return
        before = self._mplane_snapshot_values(sheet)
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        for rr, cc in sorted(self._mplane_selection_cells):
            if rr >= len(var_rows) or cc >= len(var_rows[rr]) or cc == 0:
                continue
            if rr < len(self.mplane_cc_on_vars) and (not self.mplane_cc_on_vars[rr].get()):
                continue
            var_rows[rr][cc].set("")
        self._mplane_push_undo(sheet, before)
        return "break"

    def _mplane_on_cell_var_changed(self, sheet: str, src_r: int, src_c: int) -> None:
        """Mirror active-cell edits to the full selected range."""
        if self._mplane_sync_lock:
            return
        meta0 = self._mplane_table_vars.get(sheet) or {}
        var_rows0: list[list[tk.StringVar]] = meta0.get("vars") or []
        if src_r < len(var_rows0) and src_c < len(var_rows0[src_r]):
            cur_val = var_rows0[src_r][src_c].get()
            if self._mplane_is_compression_col(sheet, src_c):
                if self._mplane_row_iq_is_16(sheet, src_r):
                    prev = self._mplane_last_valid_cell_values.get((sheet, src_r, src_c), "")
                    self._mplane_sync_lock = True
                    try:
                        var_rows0[src_r][src_c].set(prev)
                    finally:
                        self._mplane_sync_lock = False
                    return
                norm = self._mplane_normalize_compression_value(cur_val)
                if norm is not None:
                    if norm != cur_val:
                        self._mplane_sync_lock = True
                        try:
                            var_rows0[src_r][src_c].set(norm)
                        finally:
                            self._mplane_sync_lock = False
                    self._mplane_last_valid_cell_values[(sheet, src_r, src_c)] = norm
                # invalid typed text is reverted on FocusOut; keep editing flow.
            else:
                self._mplane_last_valid_cell_values[(sheet, src_r, src_c)] = cur_val
                headers0: list[str] = meta0.get("headers") or []
                if src_c < len(headers0) and self._mplane_tagify_header(headers0[src_c]) == "iq-bitwidth":
                    c_comp = self._mplane_col_index(sheet, "compression")
                    if cur_val.strip() == "16":
                        if c_comp is not None and src_r < len(var_rows0) and c_comp < len(var_rows0[src_r]):
                            self._mplane_sync_lock = True
                            try:
                                var_rows0[src_r][c_comp].set("no comp")
                            finally:
                                self._mplane_sync_lock = False
                            self._mplane_last_valid_cell_values[(sheet, src_r, c_comp)] = "no comp"
                    else:
                        if c_comp is not None and src_r < len(var_rows0) and c_comp < len(var_rows0[src_r]):
                            cv = (var_rows0[src_r][c_comp].get() or "").strip().lower()
                            if cv in {"", "no comp", "nocomp", "static"}:
                                self._mplane_sync_lock = True
                                try:
                                    var_rows0[src_r][c_comp].set("BFP")
                                finally:
                                    self._mplane_sync_lock = False
                                self._mplane_last_valid_cell_values[(sheet, src_r, c_comp)] = "BFP"
                    self._mplane_apply_off_row_styles()
        if self._mplane_selection_sheet != sheet or len(self._mplane_selection_cells) <= 1:
            return
        if (src_r, src_c) not in self._mplane_selection_cells:
            return
        if src_c == 0:
            return
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        if src_r >= len(var_rows) or src_c >= len(var_rows[src_r]):
            return
        value = var_rows[src_r][src_c].get()
        self._mplane_sync_lock = True
        try:
            for rr, cc in sorted(self._mplane_selection_cells):
                if (rr, cc) == (src_r, src_c):
                    continue
                if rr >= len(var_rows) or cc >= len(var_rows[rr]) or cc == 0:
                    continue
                if rr < len(self.mplane_cc_on_vars) and (not self.mplane_cc_on_vars[rr].get()):
                    continue
                if self._mplane_is_compression_col(sheet, cc):
                    norm = self._mplane_normalize_compression_value(value)
                    if norm is None:
                        continue
                    var_rows[rr][cc].set(norm)
                    self._mplane_last_valid_cell_values[(sheet, rr, cc)] = norm
                    continue
                var_rows[rr][cc].set(value)
                self._mplane_last_valid_cell_values[(sheet, rr, cc)] = value
        finally:
            self._mplane_sync_lock = False

    def _mplane_on_cell_focus_out(self, sheet: str, row: int, col: int) -> None:
        # Re-evaluate row lock state after editor commit.
        if self._mplane_is_iq_col(sheet, col):
            self.after_idle(self._mplane_apply_off_row_styles)
        if not self._mplane_is_compression_col(sheet, col):
            return
        if self._mplane_row_iq_is_16(sheet, row):
            return
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        if row >= len(var_rows) or col >= len(var_rows[row]):
            return
        val = var_rows[row][col].get()
        norm = self._mplane_normalize_compression_value(val)
        if norm is None:
            prev = self._mplane_last_valid_cell_values.get((sheet, row, col), "")
            self._mplane_sync_lock = True
            try:
                var_rows[row][col].set(prev)
            finally:
                self._mplane_sync_lock = False
            return
        if norm != val:
            self._mplane_sync_lock = True
            try:
                var_rows[row][col].set(norm)
            finally:
                self._mplane_sync_lock = False
        self._mplane_last_valid_cell_values[(sheet, row, col)] = norm

    def _mplane_on_cell_keyrelease(self, sheet: str, row: int, col: int) -> None:
        # Keep compression lock/unlock responsive while typing iq-bitwidth.
        if self._mplane_is_iq_col(sheet, col):
            self.after_idle(self._mplane_apply_off_row_styles)

    def _mplane_is_compression_col(self, sheet: str, col: int) -> bool:
        meta = self._mplane_table_vars.get(sheet) or {}
        headers: list[str] = meta.get("headers") or []
        if col < 0 or col >= len(headers):
            return False
        return self._mplane_tagify_header(headers[col]) == "compression"

    def _mplane_is_iq_col(self, sheet: str, col: int) -> bool:
        meta = self._mplane_table_vars.get(sheet) or {}
        headers: list[str] = meta.get("headers") or []
        if col < 0 or col >= len(headers):
            return False
        return self._mplane_tagify_header(headers[col]) == "iq-bitwidth"

    @staticmethod
    def _mplane_normalize_compression_value(v: str) -> str | None:
        s = (v or "").strip()
        if not s:
            return ""
        t = s.lower().replace("_", " ").replace("-", " ").strip()
        if t in {"bfp", "block floating point"}:
            return "BFP"
        if re.match(r"^exponent(\s*[:=]\s*\d+)?$", t):
            return "exponent"
        return None

    def _mplane_col_index(self, sheet: str, header_key: str) -> int | None:
        meta = self._mplane_table_vars.get(sheet) or {}
        headers: list[str] = meta.get("headers") or []
        want = self._mplane_tagify_header(header_key)
        for i, h in enumerate(headers):
            if self._mplane_tagify_header(h) == want:
                return i
        return None

    def _mplane_row_iq_is_16(self, sheet: str, row: int) -> bool:
        c_iq = self._mplane_col_index(sheet, "iq-bitwidth")
        if c_iq is None:
            return False
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        if row >= len(var_rows) or c_iq >= len(var_rows[row]):
            return False
        s = (var_rows[row][c_iq].get() or "").strip()
        if s in {"16", "16.0"}:
            return True
        try:
            return float(s) == 16.0
        except Exception:
            return False

    def _mplane_on_copy(self, _event: tk.Event, sheet: str) -> str | None:
        if self._mplane_selection_sheet != sheet or not self._mplane_selection_cells:
            return
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        if not var_rows:
            return "break"
        r0 = min(r for r, _ in self._mplane_selection_cells)
        r1 = max(r for r, _ in self._mplane_selection_cells)
        c0 = min(c for _, c in self._mplane_selection_cells)
        c1 = max(c for _, c in self._mplane_selection_cells)
        out_lines: list[str] = []
        for rr in range(r0, r1 + 1):
            vals: list[str] = []
            for cc in range(c0, c1 + 1):
                if rr < len(var_rows) and cc < len(var_rows[rr]):
                    vals.append(var_rows[rr][cc].get())
                else:
                    vals.append("")
            out_lines.append("\t".join(vals))
        text = "\n".join(out_lines)
        try:
            self.clipboard_clear()
            self.clipboard_append(text)
        except Exception:
            pass
        return "break"

    def _mplane_on_undo(self, _event: tk.Event, sheet: str) -> str:
        stack = self._mplane_undo_stack.get(sheet) or []
        if not stack:
            return "break"
        current = self._mplane_snapshot_values(sheet)
        snap = stack.pop()
        self._mplane_redo_stack.setdefault(sheet, []).append(current)
        self._mplane_restore_snapshot(sheet, snap)
        return "break"

    def _mplane_on_redo(self, _event: tk.Event, sheet: str) -> str:
        stack = self._mplane_redo_stack.get(sheet) or []
        if not stack:
            return "break"
        current = self._mplane_snapshot_values(sheet)
        snap = stack.pop()
        self._mplane_undo_stack.setdefault(sheet, []).append(current)
        self._mplane_restore_snapshot(sheet, snap)
        return "break"

    def _mplane_snapshot_values(self, sheet: str) -> dict[tuple[int, int], str]:
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        snap: dict[tuple[int, int], str] = {}
        for rr, row in enumerate(var_rows):
            for cc, sv in enumerate(row):
                snap[(rr, cc)] = sv.get()
        return snap

    def _mplane_restore_snapshot(self, sheet: str, snap: dict[tuple[int, int], str]) -> None:
        meta = self._mplane_table_vars.get(sheet) or {}
        var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
        for (rr, cc), val in snap.items():
            if rr < len(var_rows) and cc < len(var_rows[rr]):
                var_rows[rr][cc].set(val)

    def _mplane_push_undo(self, sheet: str, before: dict[tuple[int, int], str]) -> None:
        after = self._mplane_snapshot_values(sheet)
        if before == after:
            return
        stack = self._mplane_undo_stack.setdefault(sheet, [])
        stack.append(before)
        if len(stack) > 100:
            del stack[0 : len(stack) - 100]
        self._mplane_redo_stack[sheet] = []

    @staticmethod
    def _mplane_tagify_header(h: str) -> str:
        s = (h or "").strip().lower()
        s = re.sub(r"\s+", "-", s)
        s = re.sub(r"[^a-z0-9\\-_/()]+", "", s)
        return s

    def _load_mplane_workbook_from_disk(
        self, path: str, mp: Any
    ) -> tuple[
        dict[str, str],
        dict[str, str],
        dict[str, Any],
        list[dict[str, Any]],
        dict[str, tuple[list[str], list[list[str]]]],
        list[str],
        Any,
    ]:
        rpc, baselines, merged, cc_rows, tables, warnings = mp.load_workbook_payloads(path)
        return rpc, baselines, merged, cc_rows, tables, warnings, mp

    @staticmethod
    def _mplane_apply_tables_from_workbook(
        rpc: dict[str, str],
        tables: dict[str, tuple[list[str], list[list[str]]]],
        mp: Any,
    ) -> list[str]:
        """Merge Control-Sheet tables from a workbook into RPC XML (no GUI table vars)."""
        warns: list[str] = []
        for sheet in ("PDSCH", "PUSCH", "PRACH"):
            headers, rows = tables.get(sheet, ([], []))
            xml = (rpc.get(sheet) or "").strip()
            if not headers or not xml or not rows:
                continue
            if "low-level-tx-endpoints" in xml or "low-level-rx-endpoints" in xml:
                new_xml, ws = mp.apply_acorn_control_details_to_rpc(xml, sheet, headers, rows)
            else:
                new_xml, ws = mp.apply_full_table_to_rpc(xml, sheet, headers, rows)
            rpc[sheet] = new_xml
            warns.extend(ws)
        return warns

    @staticmethod
    def _mplane_cc_off_rows_from_workbook(cc_rows: list[dict[str, Any]]) -> list[int]:
        return [
            i + 1
            for i, row in enumerate(cc_rows)
            if not CallhomeGUI._mplane_enabled_from_cell(row.get("enabled"))
        ]

    def _mplane_apply_full_tables_to_rpc(self, rpc: dict[str, str]) -> list[str]:
        """Apply edited full-table values into RPC xml strings by nth-occurrence replacement."""
        warns: list[str] = []
        try:
            import mplane_control as mp
        except Exception as exc:
            return [f"mplane_control missing: {exc}"]

        for sheet in ("PDSCH", "PUSCH", "PRACH"):
            meta = self._mplane_table_vars.get(sheet) or {}
            headers: list[str] = meta.get("headers") or []
            var_rows: list[list[tk.StringVar]] = meta.get("vars") or []
            xml = (rpc.get(sheet) or "").strip()
            if not headers or not xml or not var_rows:
                headers, rows_fb = self._mplane_tables.get(sheet, ([], []))
                if headers and rows_fb and xml:
                    if "low-level-tx-endpoints" in xml or "low-level-rx-endpoints" in xml:
                        new_xml, ws = mp.apply_acorn_control_details_to_rpc(xml, sheet, headers, rows_fb)
                    else:
                        new_xml, ws = mp.apply_full_table_to_rpc(xml, sheet, headers, rows_fb)
                    rpc[sheet] = new_xml
                    warns.extend(ws)
                continue
            rows: list[list[str]] = []
            for r in var_rows:
                rows.append([(v.get() or "") for v in r])
            if "low-level-tx-endpoints" in xml or "low-level-rx-endpoints" in xml:
                new_xml, ws = mp.apply_acorn_control_details_to_rpc(xml, sheet, headers, rows)
            else:
                new_xml, ws = mp.apply_full_table_to_rpc(xml, sheet, headers, rows)
            rpc[sheet] = new_xml
            warns.extend(ws)

        return warns

    def _mplane_collect_live(self) -> dict[str, str]:
        return {k: (v.get() or "").strip() for k, v in self.mplane_fields.items()}

    def apply_mplane_workbook_once(self) -> None:
        path = self._normalize_mplane_workbook_path(self.mplane_xlsx_path.get())
        if not path:
            messagebox.showwarning("M-Plane", "Choose an .xlsx path first.")
            return
        if not os.path.isfile(path):
            messagebox.showwarning("M-Plane", f"Workbook not found:\n{path}")
            return
        threading.Thread(target=self._prepare_mplane_payload_tabs_worker, args=(path,), daemon=True).start()

    def _mplane_sync_visible_widget_values(self) -> None:
        """Force-sync current widget text into StringVars before Save/Apply."""
        for sheet in ("PDSCH", "PUSCH", "PRACH"):
            meta = self._mplane_table_vars.get(sheet) or {}
            vars_rows: list[list[tk.StringVar]] = meta.get("vars") or []
            entries_rows: list[list[tk.Widget]] = meta.get("entries") or []
            for r_idx, row_entries in enumerate(entries_rows):
                for c_idx, w in enumerate(row_entries):
                    if r_idx >= len(vars_rows) or c_idx >= len(vars_rows[r_idx]):
                        continue
                    try:
                        cur = w.get() if hasattr(w, "get") else vars_rows[r_idx][c_idx].get()
                    except Exception:
                        cur = vars_rows[r_idx][c_idx].get()
                    if cur != vars_rows[r_idx][c_idx].get():
                        vars_rows[r_idx][c_idx].set(cur)

    def _prepare_mplane_payload_tabs_worker(self, xlsx_path: str) -> None:
        try:
            import mplane_control as mp
            mp = importlib.reload(mp)
        except ImportError as exc:
            self.after(0, messagebox.showerror, "M-Plane", str(exc))
            return

        apply_btn = self.mplane_apply_btn
        if apply_btn:
            self.after(0, lambda b=apply_btn: b.config(state="disabled"))
        try:
            self.after(0, self.append_log, f"[GUI] Apply: xlsx only ({xlsx_path})\n")
            rpc_raw, baselines, merged, cc_rows, tables, load_warns, mp = self._load_mplane_workbook_from_disk(
                xlsx_path, mp
            )
            rpc = dict(rpc_raw)
            for w in load_warns:
                self.after(0, self.append_log, f"[GUI] M-Plane Apply: {w}\n")
            # Normalize endpoint comments first so ON rows never inherit stale OFF comments.
            for sheet in ("PDSCH", "PUSCH", "PRACH"):
                rpc[sheet] = mp.uncomment_endpoint_rows((rpc.get(sheet) or ""), sheet)
            rpc["ACTIVE"] = mp.uncomment_active_rows((rpc.get("ACTIVE") or ""))
            baselines = dict(baselines)
            live = {
                k: ("" if merged.get(k) is None else str(merged.get(k)).strip())
                for k in (
                    "cu_if_name",
                    "cu_base_if",
                    "cu_vlan",
                    "cu_mac",
                    "odu_mac",
                    "ru_mac_pe",
                    "pe_name",
                )
            }
            for name in mp.SEND_ORDER:
                body = (rpc.get(name) or "").strip()
                if not body:
                    self.after(0, self.append_log, f"[GUI] M-Plane skip (empty payload): {name}\n")
                    continue
                rpc[name] = mp.apply_global_baselines(body, baselines, live)
                if name == "CUplane-interface":
                    rpc[name] = mp.ensure_cuplane_interface_fields(rpc[name], live)
                elif name == "Processing-element":
                    rpc[name] = mp.ensure_processing_element_fields(rpc[name], live)
            table_warns = self._mplane_apply_tables_from_workbook(rpc, tables, mp)
            for w in table_warns:
                self.after(0, self.append_log, f"[GUI] M-Plane table: {w}\n")
            pusch_xml = (rpc.get("PUSCH") or "").strip()
            prach_xml = (rpc.get("PRACH") or "").strip()
            if pusch_xml and prach_xml:
                prach_xml, pr_warns = mp.omit_prach_rx_endpoints_present_in_pusch(prach_xml, pusch_xml)
                rpc["PRACH"] = prach_xml
                for w in pr_warns:
                    self.after(0, self.append_log, f"[GUI] M-Plane PRACH: {w}\n")
            # Global CC ON/OFF from Control-Sheet (xlsx), not GUI toggles.
            off_rows = self._mplane_cc_off_rows_from_workbook(cc_rows)
            if not off_rows:
                flags = merged.get("_detail_cc_on_flags")
                if isinstance(flags, list):
                    off_rows = [i + 1 for i, on in enumerate(flags) if not on]
            active_body, act_warns = mp.apply_active_from_control_tables(
                rpc.get("ACTIVE", ""),
                tables.get("PDSCH", ([], [])),
                tables.get("PUSCH", ([], [])),
                off_rows,
            )
            rpc["ACTIVE"] = active_body
            for w in act_warns:
                self.after(0, self.append_log, f"[GUI] M-Plane ACTIVE: {w}\n")
            if off_rows:
                for sheet in ("PDSCH", "PUSCH", "PRACH"):
                    rpc[sheet] = mp.comment_out_endpoint_rows((rpc.get(sheet) or ""), sheet, off_rows)
                self.after(0, self.append_log, f"[GUI] M-Plane OFF rows commented: {off_rows}\n")
            self.after(0, self._enqueue_mplane_payload_tabs, rpc, mp.SEND_ORDER)
        finally:
            if apply_btn:
                self.after(0, lambda b=apply_btn: b.config(state="normal"))

    def _enqueue_mplane_payload_tabs(self, rpc: dict[str, str], send_order: list[str]) -> None:
        created = 0
        for name in send_order:
            payload = (rpc.get(name) or "").strip()
            if not payload:
                continue
            payload = self._format_rpc_for_message_tab(name, payload)
            self.add_message_tab(initial_title=f"M-{name}")
            tab = self.message_tabs[-1]
            text_widget = tab["text"]
            text_widget.delete("1.0", "end")
            text_widget.insert("1.0", payload)
            self._highlight_xml(text_widget)
            created += 1
        if created == 0:
            self.append_log("[GUI] M-Plane: no payload to queue.\n")
            return
        # Focus Netconf Client tab and the last created message tab.
        try:
            self.notebook.select(2)  # Settings(0), M-Plane(1), Netconf(2), Conformance(3), Shortcuts(4)
        except Exception:
            pass
        try:
            self.msg_notebook.select(len(self.message_tabs) - 1)
        except Exception:
            pass
        self.append_log(f"[GUI] M-Plane queued {created} payload tab(s). Use Netconf Client send buttons.\n")

    def _format_rpc_for_message_tab(self, name: str, payload: str) -> str:
        """Pretty-format RPC payload for editor readability."""
        text = (payload or "").strip()
        if not text:
            return payload
        out = text
        try:
            pretty = xml.dom.minidom.parseString(text).toprettyxml(indent="  ")
            # Drop empty lines produced by minidom.
            pretty = "\n".join(line for line in pretty.splitlines() if line.strip())
            out = pretty
        except Exception:
            out = text

        # Improve carrier block readability for multi-CC payloads.
        if name in {"PDSCH", "PUSCH", "PRACH", "ACTIVE"}:
            # Add block spacing only when payload actually has multiple CC sections.
            cc_blocks = len(re.findall(r"<\s*low-level-(?:tx|rx)-endpoints\s*>", out))
            if cc_blocks > 1:
                out = re.sub(
                    r"\n(\s*<(?:low-level-(?:tx|rx)-endpoints|tx-array-carriers|rx-array-carriers|low-level-(?:tx|rx)-links)>)",
                    r"\n\n\n\1",
                    out,
                )
                out = re.sub(r"\n{5,}", "\n\n\n\n", out)
            else:
                # Single-CC payload: keep blocks compact (no extra blank gaps).
                out = re.sub(
                    r"\n{2,}(\s*<(?:tx-array-carriers|rx-array-carriers|low-level-(?:tx|rx)-links)>)",
                    r"\n\1",
                    out,
                )
            # Same CC: keep endpoints/array-carriers/links attached without blank lines.
            out = re.sub(
                r"</(low-level-(?:tx|rx)-endpoints|tx-array-carriers|rx-array-carriers)>\n\s*\n(\s*<(?:tx-array-carriers|rx-array-carriers|low-level-(?:tx|rx)-links)>)",
                r"</\1>\n\2",
                out,
            )
            # Keep indentation around comment wrappers readable.
            out = re.sub(r"(\n\s*-->)\n(?=\s*<)", r"\1\n\n", out)
            out = out.replace("<!--\n\n", "<!--\n").replace("\n\n-->", "\n-->")
        return out

    def _on_main_window_configure(self, event: tk.Event) -> None:
        if event.widget is not self:
            return
        if self._geometry_save_job is not None:
            try:
                self.after_cancel(self._geometry_save_job)
            except Exception:
                pass
        self._geometry_save_job = self.after(900, self._save_current_config)

    def _on_log_window_configure(self, event: tk.Event) -> None:
        if self.log_window is None or event.widget is not self.log_window:
            return
        self.log_window_geometry = self.log_window.geometry()
        if self._log_geometry_save_job is not None:
            try:
                self.after_cancel(self._log_geometry_save_job)
            except Exception:
                pass
        self._log_geometry_save_job = self.after(900, self._save_current_config)

    def _on_log_mousewheel(self, event: tk.Event) -> str:
        if self.log is None:
            return "break"
        delta = int(-1 * (event.delta / 120))
        self.log.yview_scroll(delta, "units")
        return "break"

    def _extract_rpc_error_items(self, text: str) -> list[dict[str, str]]:
        items: list[dict[str, str]] = []
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for block in re.findall(r"<rpc-error\b[\s\S]*?</rpc-error>", text, flags=re.IGNORECASE):
            block_upper = block.upper()
            if "OPERATION-FAILED" not in block_upper:
                continue
            m_msg = re.search(r"<error-message>\s*([\s\S]*?)\s*</error-message>", block, flags=re.IGNORECASE)
            m_path = re.search(r"<error-path>\s*([\s\S]*?)\s*</error-path>", block, flags=re.IGNORECASE)
            m_info = re.search(r"<error-info>\s*([\s\S]*?)\s*</error-info>", block, flags=re.IGNORECASE)
            msg = re.sub(r"\s+", " ", m_msg.group(1)).strip() if m_msg else ""
            path = re.sub(r"\s+", " ", m_path.group(1)).strip() if m_path else ""
            info = re.sub(r"\s+", " ", m_info.group(1)).strip() if m_info else ""
            detail = msg or path or info or "operation-failed"
            items.append(
                {
                    "ts": now,
                    "kind": "rpc-error",
                    "summary": f"operation-failed | {detail[:180]}",
                    "anchor": msg or path or "operation-failed",
                    "raw": re.sub(r"\s+", " ", block).strip(),
                }
            )
        # Generic signal keywords requested by user.
        # Capture line-level snippets for fail/error/alarm (latest-first UI later).
        for ln in text.splitlines():
            s = ln.strip()
            if not s:
                continue
            m = re.search(r"\b(fail(?:ed|ure)?|error|alarm)\b", s, flags=re.IGNORECASE)
            if not m:
                continue
            kw = m.group(1).lower()
            flat = re.sub(r"\s+", " ", s).strip()
            items.append(
                {
                    "ts": now,
                    "kind": kw,
                    "summary": f"{kw} | {flat[:180]}",
                    "anchor": flat[:180] or kw,
                    "raw": flat,
                }
            )
        return items

    def _register_rpc_error_summary(self, text: str) -> None:
        for item in self._extract_rpc_error_items(text):
            sig = hashlib.sha1(item["raw"].encode("utf-8", errors="ignore")).hexdigest()
            if sig in self.rpc_error_seen:
                continue
            self.rpc_error_seen.add(sig)
            self.rpc_error_items.append(item)
            self._refresh_rpc_event_list_widget()

    def _jump_to_selected_rpc_error(self, _event: tk.Event | None = None) -> None:
        if self.log is None or self.rpc_error_list is None:
            return
        sel = self.rpc_error_list.curselection()
        if not sel:
            return
        idx = int(sel[0])
        if idx < 0 or idx >= len(self.rpc_error_items):
            return
        item = self.rpc_error_items[idx]
        anchor = item.get("anchor", "operation-failed").strip() or "operation-failed"
        self.log.configure(state="normal")
        self.log.tag_remove("rpc_err_hit", "1.0", "end")
        pos = self.log.search(anchor, "1.0", stopindex="end", nocase=True)
        if not pos:
            pos = self.log.search("operation-failed", "1.0", stopindex="end", nocase=True)
        if pos:
            end = f"{pos}+{max(1, len(anchor))}c"
            self.log.tag_add("rpc_err_hit", pos, end)
            self.log.tag_configure("rpc_err_hit", background="#7f1d1d", foreground="#fee2e2")
            self.log.see(pos)
            self.status_var.set(f"RPC error line: {pos}")
        self.log.configure(state="disabled")

    def _refresh_rpc_error_summary_widget(self) -> None:
        if self.rpc_error_list is None or not self.rpc_error_list.winfo_exists():
            return
        self.rpc_error_list.delete(0, "end")
        for item in self.rpc_error_items[-300:]:
            self.rpc_error_list.insert("end", item["summary"])

    def _refresh_rpc_event_list_widget(self) -> None:
        if self.rpc_event_list is None or not self.rpc_event_list.winfo_exists():
            return
        self.rpc_event_list.delete(0, "end")
        # Latest first
        for item in reversed(self.rpc_error_items[-600:]):
            ts = item.get("ts", "")
            kind = item.get("kind", "").upper()
            summary = item.get("summary", "")
            self.rpc_event_list.insert("end", f"[{ts}] [{kind}] {summary}")

    def open_rpc_event_window(self) -> None:
        if self.rpc_event_window is not None and self.rpc_event_window.winfo_exists():
            self.rpc_event_window.deiconify()
            self.rpc_event_window.lift()
            self._refresh_rpc_event_list_widget()
            return

        win = tk.Toplevel(self)
        win.title("RPC Events")
        win.geometry(self.rpc_event_window_geometry or "980x520")
        self.rpc_event_window = win

        top = ttk.Frame(win, padding=8)
        top.pack(fill="x")
        ttk.Label(
            top,
            text="Keywords: rpc-error(operation-failed), fail, error, alarm | Newest first",
            foreground="#64748b",
        ).pack(side="left")
        ttk.Button(top, text="Refresh", command=self._refresh_rpc_event_list_widget).pack(side="right")

        body = ttk.Frame(win, padding=8)
        body.pack(fill="both", expand=True)
        self.rpc_event_list = tk.Listbox(body, activestyle="none")
        self.rpc_event_list.pack(side="left", fill="both", expand=True)
        self.rpc_event_list.bind("<<ListboxSelect>>", self._jump_to_selected_rpc_event)
        self.rpc_event_list.bind("<Double-Button-1>", self._jump_to_selected_rpc_event)
        ys = ttk.Scrollbar(body, orient="vertical", command=self.rpc_event_list.yview)
        ys.pack(side="right", fill="y")
        self.rpc_event_list.configure(yscrollcommand=ys.set)
        self._refresh_rpc_event_list_widget()

        def _on_close() -> None:
            if self.rpc_event_window is not None:
                try:
                    self.rpc_event_window_geometry = self.rpc_event_window.geometry()
                except Exception:
                    pass
            self.rpc_event_list = None
            self._save_current_config()
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", _on_close)

    def _jump_to_selected_rpc_event(self, _event: tk.Event | None = None) -> None:
        if self.rpc_event_list is None or self.log is None:
            self.open_log_window()
        if self.rpc_event_list is None or self.log is None:
            return
        sel = self.rpc_event_list.curselection()
        if not sel:
            return
        # rpc_event_list is latest-first view of rpc_error_items tail.
        view = self.rpc_error_items[-600:]
        rev = list(reversed(view))
        idx = int(sel[0])
        if idx < 0 or idx >= len(rev):
            return
        item = rev[idx]
        anchor = (item.get("anchor") or item.get("raw") or item.get("kind") or "error").strip()
        self.open_log_window()
        if self.log is None:
            return
        self.log.configure(state="normal")
        self.log.tag_remove("rpc_err_hit", "1.0", "end")
        # Search from bottom to jump to latest occurrence.
        pos = self.log.search(anchor, "end-1c", stopindex="1.0", nocase=True, backwards=True)
        if not pos:
            k = (item.get("kind") or "error").strip() or "error"
            pos = self.log.search(k, "end-1c", stopindex="1.0", nocase=True, backwards=True)
        if pos:
            end = f"{pos}+{max(1, min(len(anchor), 180))}c"
            self.log.tag_add("rpc_err_hit", pos, end)
            self.log.tag_configure("rpc_err_hit", background="#7f1d1d", foreground="#fee2e2")
            self.log.see(pos)
            self.status_var.set(f"RPC event line: {pos}")
        self.log.configure(state="disabled")

    def open_log_window(self) -> None:
        if self.log_window is not None and self.log_window.winfo_exists():
            self.log_window.deiconify()
            self.log_window.lift()
            self._flush_hidden_logs_to_widget()
            return

        win = tk.Toplevel(self)
        win.title("Netconf Logs")
        win.geometry(self.log_window_geometry or "1100x720")
        self.log_window = win
        win.bind("<Configure>", self._on_log_window_configure)

        tools = ttk.LabelFrame(win, text="Remote Log Tools", padding=8)
        tools.pack(fill="x", padx=8, pady=(8, 6))
        self.fetch_lines_var = tk.StringVar(value="300")
        ttk.Button(tools, text="Log Load", command=self.load_full_remote_log).pack(side="left", padx=(8, 6), pady=8)
        ttk.Button(tools, text="Clear Logs", command=self.clear_log).pack(side="left", padx=6, pady=8)
        ttk.Button(tools, text="RPC Events", command=self.open_rpc_event_window).pack(side="left", padx=6, pady=8)
        gui_cap = ttk.Frame(tools)
        gui_cap.pack(side="left", padx=(16, 4), pady=4)
        ttk.Label(gui_cap, text="GUI keep lines").pack(side="left", padx=(0, 4))
        ttk.Entry(gui_cap, textvariable=self.gui_log_max_lines_var, width=8).pack(side="left", padx=2)
        ttk.Button(gui_cap, text="Apply", command=self.apply_gui_log_line_limit).pack(side="left", padx=6)
        ttk.Label(gui_cap, foreground="#64748b", text="Oldest trimmed when exceeded.").pack(side="left", padx=(8, 0))
        ttk.Label(tools, textvariable=self.perf_text_var, foreground="#666666").pack(side="right", padx=8)

        log_frame = ttk.LabelFrame(win, text="Live Output / Remote Log", padding=8)
        log_frame.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        ttk.Label(
            log_frame,
            textvariable=self.log_target_hint_var,
            foreground="#0369a1",
            justify="left",
            wraplength=1040,
            font=("Segoe UI", 9),
        ).pack(anchor="w", fill="x", pady=(0, 6))
        self._refresh_log_target_hint_line()

        log_inner = ttk.Frame(log_frame)
        log_inner.pack(fill="both", expand=True)
        self.log = tk.Text(
            log_inner,
            wrap="none",
            relief="flat",
        )
        self._apply_code_text_theme(self.log)
        log_scroll_y = ttk.Scrollbar(log_inner, orient="vertical", command=self.log.yview)
        log_scroll_x = ttk.Scrollbar(log_inner, orient="horizontal", command=self.log.xview)
        self.log.configure(yscrollcommand=log_scroll_y.set, xscrollcommand=log_scroll_x.set)
        self.log.grid(row=0, column=0, sticky="nsew")
        log_scroll_y.grid(row=0, column=1, sticky="ns")
        log_scroll_x.grid(row=1, column=0, sticky="ew")
        log_inner.rowconfigure(0, weight=1)
        log_inner.columnconfigure(0, weight=1)
        self.log.configure(state="disabled")
        self.log.bind("<MouseWheel>", self._on_log_mousewheel)

        def _on_close() -> None:
            if self.log_window is not None:
                try:
                    self.log_window_geometry = self.log_window.geometry()
                except Exception:
                    pass
            self.log = None
            self.rpc_error_list = None
            self._save_current_config()
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", _on_close)
        self._flush_hidden_logs_to_widget()

    def add_message_tab(self, initial_title: str | None = None) -> None:
        idx = len(self.message_tabs) + 1
        title_var = tk.StringVar(value=initial_title or f"MSG-{idx}")
        frame = ttk.Frame(self.msg_notebook)
        top = ttk.Frame(frame)
        top.pack(fill="x", padx=8, pady=6)
        ttk.Label(top, text="Title").pack(side="left")
        ttk.Entry(top, textvariable=title_var, width=18).pack(side="left", padx=(6, 10))
        ttk.Label(top, text="(XML 또는 CLI text)").pack(side="left")

        text = tk.Text(frame, wrap="none")
        text.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        text.configure(undo=True, autoseparators=True, maxundo=-1)
        self._setup_xml_editor_theme(text)
        text.bind("<KeyRelease>", lambda _e, tw=text: self._on_xml_editor_changed(tw))
        text.bind("<FocusIn>", lambda _e, tw=text: self._on_xml_editor_changed(tw))
        for seq in ("<Control-slash>", "<Control-question>", "<Control-KeyPress-slash>", "<Control-KeyPress-KP_Divide>"):
            text.bind(seq, self._on_shortcut_toggle_comment, add="+")
        for seq in ("<Control-z>", "<Control-Z>", "<Control-KeyPress-z>", "<Control-KeyPress-Z>"):
            text.bind(seq, self._on_xml_undo, add="+")
        for seq in ("<Control-Shift-z>", "<Control-Shift-Z>", "<Control-y>", "<Control-Y>"):
            text.bind(seq, self._on_xml_redo, add="+")
        for seq in ("<Control-f>", "<Control-F>", "<Control-KeyPress-f>", "<Control-KeyPress-F>"):
            text.bind(seq, self._on_xml_find, add="+")
        for seq in ("<Control-h>", "<Control-H>", "<Control-KeyPress-h>", "<Control-KeyPress-H>"):
            text.bind(seq, self._on_xml_replace, add="+")
        tab_data = {"frame": frame, "title_var": title_var, "text": text}
        self.message_tabs.append(tab_data)
        self.msg_notebook.add(frame, text=title_var.get())
        title_var.trace_add("write", lambda *_: self._refresh_message_tab_titles())
        self._refresh_message_tab_titles()

    def _refresh_message_tab_titles(self) -> None:
        for i, tab in enumerate(self.message_tabs):
            self.msg_notebook.tab(i, text=tab["title_var"].get().strip() or f"MSG-{i+1}")

    def remove_selected_message_tab(self) -> None:
        if not self.message_tabs:
            return
        idx = self.msg_notebook.index("current")
        self.msg_notebook.forget(idx)
        self.message_tabs.pop(idx)
        if not self.message_tabs:
            self.add_message_tab(initial_title="MSG-1")
        self._refresh_message_tab_titles()

    def move_selected_tab_left(self) -> None:
        if len(self.message_tabs) < 2:
            return
        idx = self.msg_notebook.index("current")
        if idx <= 0:
            return
        tab = self.message_tabs.pop(idx)
        self.message_tabs.insert(idx - 1, tab)
        self.msg_notebook.insert(idx - 1, tab["frame"])
        self.msg_notebook.select(idx - 1)
        self._refresh_message_tab_titles()

    def move_selected_tab_right(self) -> None:
        if len(self.message_tabs) < 2:
            return
        idx = self.msg_notebook.index("current")
        if idx >= len(self.message_tabs) - 1:
            return
        tab = self.message_tabs.pop(idx)
        self.message_tabs.insert(idx + 1, tab)
        self.msg_notebook.insert(idx + 1, tab["frame"])
        self.msg_notebook.select(idx + 1)
        self._refresh_message_tab_titles()

    def load_file_into_selected_tab(self) -> None:
        if not self.message_tabs:
            return
        path = filedialog.askopenfilename(
            title="Select XML/Text file",
            filetypes=[("XML/Text", "*.xml *.txt *.rpc"), ("All files", "*.*")],
        )
        if not path:
            return
        idx = self.msg_notebook.index("current")
        text_widget = self.message_tabs[idx]["text"]
        try:
            content = Path(path).read_text(encoding="utf-8")
        except Exception:
            content = Path(path).read_text(errors="ignore")
        text_widget.delete("1.0", "end")
        text_widget.insert("1.0", content)
        self._highlight_xml(text_widget)
        self.append_log(f"[GUI] Loaded file into tab: {path}\n")

    def load_excel_tabs(self) -> None:
        path = filedialog.askopenfilename(
            title="Select Excel file",
            initialdir="F:/ORAN/nDIU",
            initialfile="nDIU_n78_Mplane정리.xlsx",
            filetypes=[("Excel Workbook", "*.xlsx *.xlsm"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            import openpyxl  # type: ignore
        except Exception:
            messagebox.showerror("Error", "openpyxl is required. Install with: py -m pip install openpyxl")
            return
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as exc:
            messagebox.showerror("Error", f"Failed to open Excel file:\n{exc}")
            return

        # Use last 6 sheet tabs (matches "bottom tabs 6개" request).
        target_sheets = wb.sheetnames[-6:]
        if not target_sheets:
            messagebox.showwarning("Warning", "No sheets found in workbook.")
            return

        while self.message_tabs:
            self.msg_notebook.forget(0)
            self.message_tabs.pop(0)

        for sname in target_sheets:
            ws = wb[sname]
            # Preserve worksheet row structure first, then normalize line endings.
            row_lines: list[str] = []
            for row in ws.iter_rows(values_only=True):
                vals = [v for v in row if v is not None]
                if not vals:
                    row_lines.append("")
                    continue
                # Use the first non-empty cell in each row to preserve visual XML layout.
                s = str(vals[0])
                s = s.replace("\r\n", "\n").replace("_x000D_", "")
                s = s.replace("\\r\\n", "\n").replace("\\n", "\n").replace("\\t", "\t")
                row_lines.append(s)

            content = "\n".join(row_lines)
            # Remove only trailing empty lines.
            content = re.sub(r"\n+\Z", "", content)
            # If XML is still one-line, pretty-print for readability.
            if content.startswith("<") and "\n" not in content and "><" in content:
                try:
                    pretty = xml.dom.minidom.parseString(content).toprettyxml(indent="  ")
                    content = "\n".join(
                        l for l in pretty.splitlines() if l.strip() and not l.strip().startswith("<?xml")
                    )
                except Exception:
                    pass
            self.add_message_tab(initial_title=sname)
            self.message_tabs[-1]["text"].delete("1.0", "end")
            self.message_tabs[-1]["text"].insert("1.0", content)
            self._highlight_xml(self.message_tabs[-1]["text"])

        self.append_log(f"[GUI] Loaded {len(target_sheets)} Excel tabs from: {path}\n")

    def _ssh_base_cmd(
        self,
        user: str,
        host: str,
        port: str,
        key_path: str,
        use_password: bool,
        auth_mode: str,
        ssh_password: str,
    ) -> list[str]:
        cmd: list[str] = []
        if auth_mode == "sshpass":
            cmd.extend(["sshpass", "-e"])
            cmd.extend(["ssh", "-p", port, *self._ssh_options(use_password)])
        elif auth_mode == "putty":
            cmd.extend(["plink", "-ssh", "-P", port])
            if use_password:
                cmd.extend(["-pw", ssh_password])
        else:
            cmd.extend(["ssh", "-p", port, *self._ssh_options(use_password)])
        if key_path:
            cmd.extend(["-i", key_path])
        cmd.append(f"{user}@{host}")
        return cmd

    def _scp_base_cmd(
        self, port: str, key_path: str, use_password: bool, auth_mode: str, ssh_password: str
    ) -> list[str]:
        cmd: list[str] = []
        if auth_mode == "sshpass":
            cmd.extend(["sshpass", "-e"])
            cmd.extend([
                "scp",
                "-P",
                port,
                "-o",
                "ConnectTimeout=8",
                "-o",
                "ConnectionAttempts=1",
            ])
            if not use_password:
                cmd.extend(["-o", "BatchMode=yes"])
        elif auth_mode == "putty":
            cmd.extend(["pscp", "-P", port])
            if use_password:
                cmd.extend(["-pw", ssh_password])
        else:
            cmd.extend([
                "scp",
                "-P",
                port,
                "-o",
                "ConnectTimeout=8",
                "-o",
                "ConnectionAttempts=1",
            ])
            if not use_password:
                cmd.extend(["-o", "BatchMode=yes"])
        if key_path:
            cmd.extend(["-i", key_path])
        return cmd

    def _ensure_remote_script(
        self,
        ssh_user: str,
        ssh_host: str,
        ssh_port: str,
        key_path: str,
        ssh_password: str,
        auth_mode: str,
        remote_script: str,
        remote_log_path: str,
    ) -> tuple[bool, str]:
        local_script = self._resolve_local_script_path()
        self.script_path = local_script
        local_exists = local_script.exists()

        if auth_mode == "paramiko":
            return self._ensure_remote_script_paramiko(
                ssh_user, ssh_host, ssh_port, key_path, ssh_password, remote_script, remote_log_path, local_exists
            )

        use_password = bool(ssh_password)
        ssh_cmd = self._ssh_base_cmd(
            ssh_user, ssh_host, ssh_port, key_path, use_password, auth_mode, ssh_password
        )
        run_env = os.environ.copy()
        if use_password and auth_mode == "sshpass":
            run_env["SSHPASS"] = ssh_password
        check_cmd = ssh_cmd + [f"test -f {shlex.quote(remote_script)}"]
        check_result = subprocess.run(check_cmd, capture_output=True, text=True, env=run_env)
        script_exists = check_result.returncode == 0

        remote_parent = str(Path(remote_script).parent).replace("\\", "/")
        mkdir_cmd = ssh_cmd + [f"mkdir -p {shlex.quote(remote_parent)}"]
        mkdir_result = subprocess.run(mkdir_cmd, capture_output=True, text=True, env=run_env)
        if mkdir_result.returncode != 0:
            err = (mkdir_result.stderr or mkdir_result.stdout).strip()
            return False, f"Failed to create remote directory: {err}"

        # If remote already has the script, allow run even when local file is missing.
        if script_exists and not local_exists:
            log_mkdir_cmd = ssh_cmd + [f"mkdir -p {shlex.quote(remote_log_path)}"]
            log_mkdir_result = subprocess.run(log_mkdir_cmd, capture_output=True, text=True, env=run_env)
            if log_mkdir_result.returncode != 0:
                err = (log_mkdir_result.stderr or log_mkdir_result.stdout).strip()
                return False, f"Failed to create remote LOG_PATH directory: {err}"
            return True, f"Remote script reused: {remote_script}, LOG_PATH ready: {remote_log_path}"

        if not local_exists:
            app_dir = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
            return (
                False,
                "Local script not found.\n"
                f"- tried: {local_script}\n"
                f"- expected near exe: {app_dir / 'miniDU_callhome.sh'}",
            )

        scp_cmd = self._scp_base_cmd(ssh_port, key_path, use_password, auth_mode, ssh_password) + [
            str(self.script_path),
            f"{ssh_user}@{ssh_host}:{remote_script}",
        ]
        scp_result = subprocess.run(scp_cmd, capture_output=True, text=True, env=run_env)
        if scp_result.returncode != 0:
            err = (scp_result.stderr or scp_result.stdout).strip()
            return False, f"Failed to upload script via scp: {err}"

        chmod_cmd = ssh_cmd + [f"chmod +x {shlex.quote(remote_script)}"]
        chmod_result = subprocess.run(chmod_cmd, capture_output=True, text=True, env=run_env)
        if chmod_result.returncode != 0:
            err = (chmod_result.stderr or chmod_result.stdout).strip()
            return False, f"Uploaded, but chmod failed: {err}"

        log_mkdir_cmd = ssh_cmd + [f"mkdir -p {shlex.quote(remote_log_path)}"]
        log_mkdir_result = subprocess.run(log_mkdir_cmd, capture_output=True, text=True, env=run_env)
        if log_mkdir_result.returncode != 0:
            err = (log_mkdir_result.stderr or log_mkdir_result.stdout).strip()
            return False, f"Failed to create remote LOG_PATH directory: {err}"

        if script_exists:
            return True, f"Remote script updated: {remote_script}, LOG_PATH ready: {remote_log_path}"
        return True, f"Remote script created: {remote_script}, LOG_PATH ready: {remote_log_path}"

    def _ensure_remote_script_paramiko(
        self,
        ssh_user: str,
        ssh_host: str,
        ssh_port: str,
        key_path: str,
        ssh_password: str,
        remote_script: str,
        remote_log_path: str,
        local_exists: bool,
    ) -> tuple[bool, str]:
        try:
            import paramiko  # type: ignore
        except Exception:
            return False, "paramiko is not installed. Run: py -m pip install paramiko"

        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:
            client.connect(
                hostname=ssh_host,
                port=int(ssh_port),
                username=ssh_user,
                password=ssh_password if ssh_password else None,
                key_filename=key_path if key_path else None,
                timeout=8,
                auth_timeout=8,
                banner_timeout=8,
                look_for_keys=not bool(ssh_password),
                allow_agent=True,
            )

            sftp = client.open_sftp()
            remote_parent = str(Path(remote_script).parent).replace("\\", "/")
            _, stdout, stderr = client.exec_command(f"mkdir -p {shlex.quote(remote_parent)}")
            if stdout.channel.recv_exit_status() != 0:
                return False, f"Failed to create remote directory: {stderr.read().decode(errors='ignore').strip()}"
            remote_exists = False
            try:
                sftp.stat(remote_script)
                remote_exists = True
            except Exception:
                remote_exists = False

            if local_exists:
                # Prefer fresh upload when local script exists.
                sftp.put(str(self.script_path), remote_script)
                sftp.chmod(remote_script, 0o755)
            elif not remote_exists:
                app_dir = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else Path(__file__).resolve().parent
                return (
                    False,
                    "Local script not found.\n"
                    f"- tried: {self.script_path}\n"
                    f"- expected near exe: {app_dir / 'miniDU_callhome.sh'}",
                )

            _, stdout, stderr = client.exec_command(f"mkdir -p {shlex.quote(remote_log_path)}")
            if stdout.channel.recv_exit_status() != 0:
                return False, f"Failed to create remote LOG_PATH directory: {stderr.read().decode(errors='ignore').strip()}"
            if local_exists:
                return True, f"Remote script ready: {remote_script}, LOG_PATH ready: {remote_log_path}"
            return True, f"Remote script reused: {remote_script}, LOG_PATH ready: {remote_log_path}"
        except Exception as exc:
            return False, f"Remote prep failed: {exc}"
        finally:
            try:
                client.close()
            except Exception:
                pass

    @staticmethod
    def _sanitize_log_text(text: str) -> str:
        # Remove control chars (e.g., BEL) that render as odd symbols in Tk.
        return re.sub(r"[\x00-\x08\x0b-\x1f\x7f]", "", text)

    @staticmethod
    def _pretty_xml_if_possible(line: str) -> str:
        raw = line.strip()
        if not (raw.startswith("<") and raw.endswith(">")):
            return line
        try:
            pretty = xml.dom.minidom.parseString(raw).toprettyxml(indent="  ")
            pretty = "\n".join(
                l for l in pretty.splitlines() if l.strip() and not l.strip().startswith("<?xml")
            )
            return pretty + "\n"
        except Exception:
            return line

    def _should_hide_line(self, line: str) -> bool:
        line_upper = line.upper()
        stripped_upper = line_upper.lstrip()

        # Never hide GUI orchestration / Conformance-runner lines (these are not NETCONF XML).
        if stripped_upper.startswith("[GUI]"):
            return False
        if "[CONFORMANCE" in line_upper or "[CONFORMANCE-RUN]" in line_upper:
            return False
        if stripped_upper.startswith(("[TRACE:", "[INFO]", "[WARN]", "[ERROR]", "[FAIL]", "[OK]")):
            return False
        if "CLIENT SENT" in line_upper:
            return False
        if "STEP " in line_upper and any(
            k in line_upper for k in ("CALLHOME", "LOGIN", "CRITERIA", "SUBSCRIPTION", "SUPERVISION")
        ):
            return False
        if "AUTHENTICATION SUCCESSFUL" in line_upper or "ACCEPTED A CONNECTION ON" in line_upper:
            return False
        if stripped_upper.strip() in ("OK", "NOK"):
            return False

        # Otherwise keep only NETCONF-centric lines (historic log filter behaviour).
        keep_tokens = [
            "CLIENT_SENT:",
            "<RPC",
            "</RPC",
            "<RPC-REPLY",
            "</RPC-REPLY",
            "<HELLO",
            "</HELLO",
            "<NOTIFICATION",
            "</NOTIFICATION",
            "<DATA",
            "</DATA",
            "<OK/>",
            "<RPC-ERROR",
            "</RPC-ERROR",
            "<ERROR-",
            "</ERROR-",
            "<ALARM-NOTIF",
            "</ALARM-NOTIF",
            "<FAULT-",
            "<IS-CLEARED",
            "<EVENT-TIME",
            "<AFFECTED-",
            "<FAULT-SEVERITY",
            "<FAULT-TEXT",
            "<SYNC",
            "<SYNCHRONIZATION",
            "</SYNC",
            "</SYNCHRONIZATION",
            "<SUPERVISION",
            "</SUPERVISION",
        ]
        if any(token in line_upper for token in keep_tokens):
            return False
        # Keep lines that look like XML element content (indented tags inside a block).
        s = line.strip()
        if s.startswith("<") and s.endswith(">"):
            return False
        return True

    def _refresh_log_target_hint_line(self) -> None:
        """로그 창 상단: Start( LOG_PATH ) vs Conformance( …/var/tmp/conformance/logs/… ) 안내."""
        v = self.log_target_hint_var
        if getattr(self, "_conformance_run_busy", False):
            cur = getattr(self, "_conformance_active_host_log", None)
            if cur:
                v.set(f"[Conformance] 원격 세션 로그 (tee): {cur}")
            else:
                v.set("[Conformance] 업로드/준비 중… (세션 로그 경로는 첫 스크립트 시작 시 표시)")
            return
        if getattr(self, "is_running", False):
            lp = (getattr(self, "_start_log_path_hint", None) or "").strip()
            if not lp:
                try:
                    w = self.fields.get("LOG_PATH")
                    lp = w.get().strip() if w is not None else ""
                except Exception:
                    lp = ""
            if lp:
                v.set(f"[Start / miniDU] 원격 로그 디렉터리 (LOG_PATH): {lp}")
            else:
                v.set("[Start / miniDU] 실행 중 — Settings의 LOG_PATH를 확인하세요.")
            return
        lp = ""
        try:
            w = self.fields.get("LOG_PATH")
            lp = w.get().strip() if w is not None else ""
        except Exception:
            pass
        prod = ""
        try:
            w = self.fields.get("PRODUCT")
            prod = w.get().strip() if w is not None else ""
        except Exception:
            pass
        safe = re.sub(r"[^0-9A-Za-z._-]+", "_", prod).strip("_") or "PRODUCT"
        lastc = getattr(self, "_conformance_last_host_log", None)
        last_txt = f"  |  마지막 Conf. 세션 로그: {lastc}" if lastc else ""
        if lp:
            v.set(
                f"[대기] Start → {lp}  |  Conformance tee 로그 → /var/tmp/conformance/logs/{safe}/CONF_{safe}_<yymmdd_HHMMSS>_<script>.log{last_txt}"
            )
        else:
            v.set(
                f"[대기] Start는 Settings LOG_PATH 필요  |  Conformance tee → /var/tmp/conformance/logs/{safe}/CONF_*.log (실행 시 생성){last_txt}"
            )

    # Pretty-print NETCONF XML in the log viewer; skip only on huge payloads to avoid UI freezes.
    _max_log_pretty_chars = 2 * 1024 * 1024

    def append_log(self, line: str) -> None:
        t0 = time.perf_counter()
        line = self._sanitize_log_text(line)
        if not line:
            return

        kept_parts: list[str] = []
        for part in line.splitlines(keepends=True):
            if self._should_hide_line(part):
                continue
            p = part
            # Keep NETCONF payload readable with lightweight pretty-print.
            s = p.strip()
            su = s.upper()
            if len(s) <= self._max_log_pretty_chars and s.startswith("<") and any(
                m in su
                for m in (
                    "<RPC",
                    "<RPC-REPLY",
                    "<NOTIFICATION",
                    "<DATA",
                    "<HELLO",
                )
            ):
                p = self._pretty_xml_if_possible(s)
            kept_parts.append(p)

        if not kept_parts:
            return
        merged = "".join(kept_parts)
        self._register_rpc_error_summary(merged)
        with self.log_lock:
            self.log_buffer.append(merged)
            self._recent_log_for_session += merged
            if len(self._recent_log_for_session) > 96_000:
                self._recent_log_for_session = self._recent_log_for_session[-96_000:]
        self._perf_record("append_log", time.perf_counter() - t0)

    def _flush_log_buffer(self) -> None:
        t0 = time.perf_counter()
        chunk = ""
        with self.log_lock:
            if self.log_buffer:
                chunk = "".join(self.log_buffer)
                self.log_buffer.clear()
        if chunk:
            # If log window is not visible, defer rendering to keep UI responsive.
            if self.log is None or self.log_window is None or not self.log_window.winfo_exists():
                with self.log_lock:
                    self.hidden_log_chunks.append(chunk)
                self._perf_record("flush_log_buffer", time.perf_counter() - t0)
                self.after(400, self._flush_log_buffer)
                return
            if self.hidden_log_chunks:
                with self.log_lock:
                    chunk = "".join(self.hidden_log_chunks) + chunk
                    self.hidden_log_chunks.clear()
            self.log.configure(state="normal")
            at_bottom_before = self.log.yview()[1] >= 0.97
            self.log.insert("end", chunk)
            # Cap line count periodically (not every flush) for smoother rendering.
            self.flush_tick += 1
            if self.flush_tick % 12 == 0:
                self._trim_log_widget_to_limit()
            if at_bottom_before:
                self.log.see("end")
            self.log.configure(state="disabled")
        self._perf_record("flush_log_buffer", time.perf_counter() - t0)
        self.after(200, self._flush_log_buffer)

    def _flush_hidden_logs_to_widget(self) -> None:
        if self.hidden_render_active:
            return
        self.hidden_render_active = True
        self.after(0, self._render_hidden_logs_batch)

    def _render_hidden_logs_batch(self) -> None:
        if self.log is None or self.log_window is None or not self.log_window.winfo_exists():
            self.hidden_render_active = False
            return
        batch = ""
        with self.log_lock:
            if self.hidden_log_chunks:
                # Render in smaller batches to avoid tab-switch stalls.
                while self.hidden_log_chunks and len(batch) < 16000:
                    batch += self.hidden_log_chunks.pop(0)
        if not batch:
            self.hidden_render_active = False
            return
        self.log.configure(state="normal")
        self.log.insert("end", batch)
        self._trim_log_widget_to_limit()
        self.log.see("end")
        self.log.configure(state="disabled")
        self.after(18, self._render_hidden_logs_batch)

    def _perf_record(self, key: str, elapsed_sec: float) -> None:
        if not self.perf_debug_var.get():
            return
        self.perf_stats[key] = self.perf_stats.get(key, 0.0) + elapsed_sec
        self.perf_counts[key] = self.perf_counts.get(key, 0) + 1
        self.perf_last[key] = elapsed_sec
        self.perf_max[key] = max(self.perf_max.get(key, 0.0), elapsed_sec)
        now = time.time()
        if now - self.perf_last_report < 5.0:
            return
        self.perf_last_report = now
        parts = []
        for k in sorted(self.perf_stats.keys()):
            count = max(1, self.perf_counts.get(k, 1))
            avg_ms = (self.perf_stats[k] / count) * 1000.0
            last_ms = self.perf_last.get(k, 0.0) * 1000.0
            max_ms = self.perf_max.get(k, 0.0) * 1000.0
            parts.append(f"{k}:avg={avg_ms:.2f}ms last={last_ms:.2f}ms max={max_ms:.2f}ms n={count}")
        if parts:
            text = "[PERF] " + " | ".join(parts)
            self.after(0, lambda t=text: self.perf_text_var.set(t))

    def _trace(self, event: str, **fields: str) -> int:
        if not self.perf_debug_var.get():
            return -1
        self.trace_seq += 1
        trace_id = self.trace_seq
        kv = " ".join(f"{k}={v}" for k, v in fields.items())
        self.append_log(f"[TRACE:{trace_id}] {event} {kv}\n")
        return trace_id

    def clear_log(self) -> None:
        if self.log is None:
            self.open_log_window()
        with self.log_lock:
            self.log_buffer.clear()
            self.hidden_log_chunks.clear()
        self.rpc_error_items.clear()
        self.rpc_error_seen.clear()
        self._refresh_rpc_error_summary_widget()
        self._refresh_rpc_event_list_widget()
        if self.log is not None:
            self.log.configure(state="normal")
            self.log.delete("1.0", "end")
            self.log.configure(state="disabled")

    def _trim_log_widget_to_limit(self) -> None:
        if self.log is None or not self.log.winfo_exists():
            return
        limit = max(50, min(int(self.max_log_lines), 200_000))
        self.log.configure(state="normal")
        current_lines = int(self.log.index("end-1c").split(".")[0])
        if current_lines > limit:
            delete_to = current_lines - limit
            self.log.delete("1.0", f"{delete_to}.0")
        self.log.configure(state="disabled")

    def apply_gui_log_line_limit(self) -> None:
        raw = self.gui_log_max_lines_var.get().strip().replace(",", "")
        try:
            n = int(raw)
        except ValueError:
            messagebox.showerror("Invalid value", "Enter an integer line count (e.g. 5000).")
            return
        n = max(100, min(n, 100_000))
        self.max_log_lines = n
        self.gui_log_max_lines_var.set(str(n))
        self._trim_log_widget_to_limit()
        self._save_current_config()
        self.status_var.set(f"GUI log buffer: keep last {n} lines")

    def _load_log_text_direct(self, text: str) -> None:
        # Used by "Log Load": render fetched remote log as-is (no NETCONF filter).
        self.clear_log()
        self.open_log_window()
        self._register_rpc_error_summary(text)
        if self.log is None:
            with self.log_lock:
                self.hidden_log_chunks.append(text)
            return
        self.log.configure(state="normal")
        self.log.insert("1.0", text)
        self._trim_log_widget_to_limit()
        self.log.see("end")
        self.log.configure(state="disabled")

    def _open_text_in_notepad(self, text: str) -> None:
        try:
            tmp_dir = Path(tempfile.gettempdir())
            out_path = tmp_dir / f"netconf_log_load_{int(time.time() * 1000)}.log"
            out_path.write_text(text, encoding="utf-8", errors="ignore")
            if os.name == "nt":
                subprocess.Popen(["notepad.exe", str(out_path)])
            else:
                messagebox.showinfo("Log Load", f"Saved log file:\n{out_path}")
            self.status_var.set(f"Log opened: {out_path.name}")
        except Exception as exc:
            messagebox.showerror("Error", f"Failed to open log in notepad:\n{exc}")

    def _collect_current_config(self) -> dict[str, Any]:
        message_data = []
        for tab in self.message_tabs:
            message_data.append(
                {
                    "title": tab["title_var"].get(),
                    "content": tab["text"].get("1.0", "end").strip(),
                }
            )
        omit = getattr(self, "_conformance_omit_last_run_from_config_save", False)
        if not omit:
            self._conformance_refresh_last_run_cache_from_progress()
        else:
            self._conformance_last_run_snapshot_cache = None
        payload: dict[str, Any] = {
            "window_geometry": self.geometry(),
            "log_window_geometry": self.log_window_geometry,
            "rpc_event_window_geometry": self.rpc_event_window_geometry,
            "remote_user": self.remote_user.get(),
            "remote_host": self.remote_host.get(),
            "remote_port": self.remote_port.get(),
            "remote_password": self.remote_password.get(),
            "remote_script_path": self.remote_script_path.get(),
            "remote_key_path": self.remote_key_path.get(),
            "perf_debug": bool(self.perf_debug_var.get()),
            "auto_xml_send": bool(self.auto_xml_send_var.get()),
            "auto_start": bool(self.auto_start_var.get()),
            "gui_max_log_lines": int(self.max_log_lines),
            "runtime_fields": {k: v.get() for k, v in self.fields.items()},
            "mplane_xlsx_path": self.mplane_xlsx_path.get(),
            "conformance_checked": {k: bool(v.get()) for k, v in self.conformance_check_vars.items()},
            "conformance_run_remote_dir": self.conformance_run_remote_dir_var.get(),
            "conformance_run_sw_pkg": self.conformance_run_sw_pkg_var.get(),
            "conformance_run_sw_remote_dir": self.conformance_run_sw_remote_dir_var.get(),
            "conformance_restart_start_after_run": bool(self.conformance_restart_start_after_run_var.get()),
            "conformance_debug": bool(self.conformance_debug_var.get()),
            "conformance_run_rpc_timeout": self.conformance_run_rpc_timeout_var.get(),
            "conformance_run_idle_timeout": self.conformance_run_idle_timeout_var.get(),
            "conformance_run_supervision_interval": self.conformance_run_supervision_interval_var.get(),
            "conformance_run_supervision_reset_cycles": self.conformance_run_supervision_reset_cycles_var.get(),
            "conformance_run_supervision_negative_fail_cycle": self.conformance_run_supervision_negative_fail_cycle_var.get(),
            "conformance_run_conn_delay": self.conformance_run_conn_delay_var.get(),
            "conformance_post_listen_wait": self.conformance_post_listen_wait_var.get(),
            "conformance_per_test_settings": self._conformance_per_test_settings,
            "conformance_extra_uploads": [
                {"local": a, "remote": b} for a, b in self._conformance_extra_uploads
            ],
            "message_tabs": message_data,
        }
        cl = self._conformance_last_run_snapshot_cache
        if not omit and isinstance(cl, dict) and cl.get("by_script"):
            payload["conformance_last_run"] = cl
        return payload

    def _apply_config(self, data: dict[str, Any]) -> None:
        geometry = data.get("window_geometry")
        if isinstance(geometry, str) and geometry.strip():
            try:
                self.geometry(geometry.strip())
            except Exception:
                pass
        log_geometry = data.get("log_window_geometry")
        if isinstance(log_geometry, str) and log_geometry.strip():
            self.log_window_geometry = log_geometry.strip()
        rpc_event_geometry = data.get("rpc_event_window_geometry")
        if isinstance(rpc_event_geometry, str) and rpc_event_geometry.strip():
            self.rpc_event_window_geometry = rpc_event_geometry.strip()
        self.exec_mode.set("remote")
        self.remote_user.set(str(data.get("remote_user", self.remote_user.get())))
        self.remote_host.set(str(data.get("remote_host", self.remote_host.get())))
        self.remote_port.set(str(data.get("remote_port", self.remote_port.get())))
        self.remote_password.set(str(data.get("remote_password", self.remote_password.get())))
        self.remote_script_path.set(str(data.get("remote_script_path", self.remote_script_path.get())))
        self.remote_key_path.set(str(data.get("remote_key_path", self.remote_key_path.get())))
        self.send_mode_var.set("raw_rpc")
        self.perf_debug_var.set(bool(data.get("perf_debug", self.perf_debug_var.get())))
        self.auto_xml_send_var.set(bool(data.get("auto_xml_send", self.auto_xml_send_var.get())))
        self.auto_start_var.set(bool(data.get("auto_start", self.auto_start_var.get())))
        glm = data.get("gui_max_log_lines")
        if glm is not None:
            try:
                n = int(glm)
                n = max(100, min(n, 100_000))
                self.max_log_lines = n
                self.gui_log_max_lines_var.set(str(n))
            except (TypeError, ValueError):
                pass
        if not self.perf_debug_var.get():
            self.perf_text_var.set("PERF: off")

        runtime_fields = data.get("runtime_fields", {})
        if isinstance(runtime_fields, dict):
            for key, value in runtime_fields.items():
                if key in self.fields:
                    self.fields[key].set(str(value))
        mxp = data.get("mplane_xlsx_path")
        if isinstance(mxp, str) and mxp.strip():
            self.mplane_xlsx_path.set(mxp.strip())
        conf_chk = data.get("conformance_checked")
        if isinstance(conf_chk, dict) and self.conformance_check_vars:
            for fname, val in conf_chk.items():
                bv = self.conformance_check_vars.get(str(fname))
                if bv is not None:
                    bv.set(bool(val))
        self._conformance_apply_last_run_from_config(data.get("conformance_last_run"))
        crd = data.get("conformance_run_remote_dir")
        if isinstance(crd, str) and crd.strip():
            self.conformance_run_remote_dir_var.set(crd.strip())
        csp = data.get("conformance_run_sw_pkg")
        if isinstance(csp, str):
            self.conformance_run_sw_pkg_var.set(csp)
        csr = data.get("conformance_run_sw_remote_dir")
        if isinstance(csr, str) and csr.strip():
            self.conformance_run_sw_remote_dir_var.set(csr.strip())
        crs2 = data.get("conformance_restart_start_after_run")
        if crs2 is not None:
            try:
                self.conformance_restart_start_after_run_var.set(bool(crs2))
            except (tk.TclError, ValueError):
                pass
        cd = data.get("conformance_debug")
        if cd is not None:
            try:
                self.conformance_debug_var.set(bool(cd))
            except (tk.TclError, ValueError):
                pass
        for var_key, cfg_key in (
            (self.conformance_run_rpc_timeout_var, "conformance_run_rpc_timeout"),
            (self.conformance_run_idle_timeout_var, "conformance_run_idle_timeout"),
            (self.conformance_run_supervision_interval_var, "conformance_run_supervision_interval"),
            (self.conformance_run_supervision_reset_cycles_var, "conformance_run_supervision_reset_cycles"),
            (self.conformance_run_supervision_negative_fail_cycle_var, "conformance_run_supervision_negative_fail_cycle"),
            (self.conformance_run_conn_delay_var, "conformance_run_conn_delay"),
            (self.conformance_post_listen_wait_var, "conformance_post_listen_wait"),
        ):
            v = data.get(cfg_key)
            if isinstance(v, str) and v.strip():
                var_key.set(v.strip())
        pts = data.get("conformance_per_test_settings")
        if isinstance(pts, dict):
            for fname, vals in pts.items():
                if isinstance(vals, dict):
                    self._conformance_per_test_settings[str(fname)] = {
                        str(k): str(v) for k, v in vals.items()
                    }
        ex = data.get("conformance_extra_uploads")
        if isinstance(ex, list):
            self._conformance_extra_uploads = []
            for it in ex:
                if isinstance(it, dict) and "local" in it and "remote" in it:
                    self._conformance_extra_uploads.append((str(it["local"]), str(it["remote"])))
        tabs = data.get("message_tabs", [])
        if isinstance(tabs, list) and tabs:
            while self.message_tabs:
                self.msg_notebook.forget(0)
                self.message_tabs.pop(0)
            for t in tabs:
                self.add_message_tab(initial_title=str(t.get("title", "MSG")))
                self.message_tabs[-1]["text"].delete("1.0", "end")
                self.message_tabs[-1]["text"].insert("1.0", str(t.get("content", "")))

    def _load_saved_config(self) -> None:
        if not self.config_path.exists():
            return
        try:
            saved = json.loads(self.config_path.read_text(encoding="utf-8"))
            if isinstance(saved, dict):
                self._apply_config(saved)
                self.append_log(f"[GUI] Loaded last config: {self.config_path.name}\n")
        except Exception as exc:
            self.append_log(f"[GUI] Failed to load config: {exc}\n")

    def _save_current_config(self) -> None:
        self._config_save_job = None
        try:
            payload = self._collect_current_config()
            self.config_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
            cl = getattr(self, "_conformance_last_run_snapshot_cache", None)
            if (
                hasattr(self, "conformance_last_run_hint_var")
                and isinstance(cl, dict)
                and cl.get("by_script")
            ):
                ts = str(cl.get("saved_at") or "").strip()
                summ = cl.get("summary")
                if not isinstance(summ, dict):
                    bs = cl.get("by_script")
                    summ = (
                        self._conformance_summarize_pass_fail_counts(bs)
                        if isinstance(bs, dict)
                        else {}
                    )
                cnt = self._conformance_format_pass_fail_counts_ko(summ) if isinstance(summ, dict) else ""
                if ts and cnt:
                    self.conformance_last_run_hint_var.set(
                        f"Conformance 마지막 결과가 설정 파일에 저장되었습니다. — {cnt} (기록 시각: {ts})"
                    )
                elif ts:
                    self.conformance_last_run_hint_var.set(
                        f"Conformance 마지막 결과가 설정 파일에 저장되었습니다. (기록 시각: {ts})"
                    )
        except Exception as exc:
            self.append_log(f"[GUI] Failed to save config: {exc}\n")

    def _resolve_auth_mode(self, ssh_password: str) -> tuple[str, str | None]:
        if not ssh_password:
            return "openssh", None
        try:
            import paramiko  # type: ignore  # noqa: F401
            return "paramiko", None
        except Exception:
            pass
        if shutil.which("sshpass") is not None:
            return "sshpass", None
        if shutil.which("plink") is not None and shutil.which("pscp") is not None:
            return "putty", "sshpass not found; using PuTTY plink/pscp."
        return "openssh", (
            "SSH_PASSWORD is set but no password backend found.\n"
            "Install paramiko OR sshpass OR PuTTY(plink, pscp), or use SSH_KEY_PATH."
        )

    def fetch_latest_remote_log(self) -> None:
        self._save_current_config()
        self.append_log("[GUI] Fetching latest remote log...\n")
        threading.Thread(target=self._fetch_latest_remote_log_worker, args=(False,), daemon=True).start()

    def load_full_remote_log(self) -> None:
        self._save_current_config()
        self.append_log("[GUI] Loading full remote log...\n")
        threading.Thread(target=self._fetch_latest_remote_log_worker, args=(True,), daemon=True).start()

    def _fetch_latest_remote_log_worker(self, full_load: bool = False) -> None:
        t0 = time.perf_counter()
        ssh_user = self.remote_user.get().strip()
        ssh_host = self.remote_host.get().strip()
        ssh_port = self.remote_port.get().strip()
        ssh_password = self.remote_password.get().strip()
        key_path = self.remote_key_path.get().strip()
        log_path = self.fields.get("LOG_PATH").get().strip()
        lines = self.fetch_lines_var.get().strip() or "500"

        if not ssh_user or not ssh_host or not ssh_port or not log_path:
            self.after(0, lambda: messagebox.showerror("Error", "SSH info and LOG_PATH are required."))
            return
        auth_mode, info_or_error = self._resolve_auth_mode(ssh_password)
        if info_or_error and "Install" in info_or_error:
            self.after(0, lambda: messagebox.showerror("Error", info_or_error))
            return
        if info_or_error:
            self.after(0, self.append_log, f"[GUI] {info_or_error}\n")

        if auth_mode == "paramiko":
            ok, text = self._fetch_latest_log_paramiko(
                ssh_user, ssh_host, ssh_port, ssh_password, key_path, log_path, lines, full_load
            )
        else:
            ok, text = self._fetch_latest_log_cli(
                ssh_user, ssh_host, ssh_port, ssh_password, key_path, auth_mode, log_path, lines, full_load
            )
        if ok:
            if full_load:
                # Keep current GUI logs intact, and open fetched full log in Notepad.
                self.after(0, self._open_text_in_notepad, text)
                self.after(0, self.append_log, "[GUI] Log Load complete. Opened in Notepad.\n")
            else:
                self.after(0, self._load_log_text_direct, text)
        else:
            self.after(0, lambda: messagebox.showerror("Error", text))
        self._perf_record("fetch_remote_log", time.perf_counter() - t0)

    def _fetch_latest_log_paramiko(
        self,
        ssh_user: str,
        ssh_host: str,
        ssh_port: str,
        ssh_password: str,
        key_path: str,
        log_path: str,
        lines: str,
        full_load: bool = False,
    ) -> tuple[bool, str]:
        try:
            import paramiko  # type: ignore
            client = paramiko.SSHClient()
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            client.connect(
                hostname=ssh_host,
                port=int(ssh_port),
                username=ssh_user,
                password=ssh_password if ssh_password else None,
                key_filename=key_path if key_path else None,
                timeout=8,
                auth_timeout=8,
                banner_timeout=8,
                look_for_keys=not bool(ssh_password),
                allow_agent=True,
            )
            if full_load:
                cmd = (
                    f"LATEST=$(ls -1t {shlex.quote(log_path)}/*.log 2>/dev/null | head -n 1); "
                    f'if [ -z "$LATEST" ]; then echo "__NO_LOG__"; else echo "__LOG__:$LATEST"; cat "$LATEST"; fi'
                )
            else:
                cmd = (
                    f"LATEST=$(ls -1t {shlex.quote(log_path)}/*.log 2>/dev/null | head -n 1); "
                    f'if [ -z "$LATEST" ]; then echo "__NO_LOG__"; else echo "__LOG__:$LATEST"; tail -n {shlex.quote(lines)} "$LATEST"; fi'
                )
            _, stdout, stderr = client.exec_command(cmd)
            out = stdout.read().decode(errors="ignore")
            err = stderr.read().decode(errors="ignore").strip()
            client.close()
            if "__NO_LOG__" in out:
                return False, f"No .log file found in LOG_PATH: {log_path}"
            if "__LOG__:" in out:
                return True, out.split("\n", 1)[1] if "\n" in out else out
            return False, err or out or "Unknown error while reading remote log."
        except Exception as exc:
            return False, f"Remote log fetch failed: {exc}"

    def _fetch_latest_log_cli(
        self,
        ssh_user: str,
        ssh_host: str,
        ssh_port: str,
        ssh_password: str,
        key_path: str,
        auth_mode: str,
        log_path: str,
        lines: str,
        full_load: bool = False,
    ) -> tuple[bool, str]:
        use_password = bool(ssh_password)
        ssh_cmd = self._ssh_base_cmd(ssh_user, ssh_host, ssh_port, key_path, use_password, auth_mode, ssh_password)
        run_env = os.environ.copy()
        if use_password and auth_mode == "sshpass":
            run_env["SSHPASS"] = ssh_password
        if full_load:
            cmd = (
                f"LATEST=$(ls -1t {shlex.quote(log_path)}/*.log 2>/dev/null | head -n 1); "
                f'if [ -z "$LATEST" ]; then echo "__NO_LOG__"; else echo "__LOG__:$LATEST"; cat "$LATEST"; fi'
            )
        else:
            cmd = (
                f"LATEST=$(ls -1t {shlex.quote(log_path)}/*.log 2>/dev/null | head -n 1); "
                f'if [ -z "$LATEST" ]; then echo "__NO_LOG__"; else echo "__LOG__:$LATEST"; tail -n {shlex.quote(lines)} "$LATEST"; fi'
            )
        result = subprocess.run(ssh_cmd + [cmd], capture_output=True, text=True, env=run_env)
        out = result.stdout or ""
        err = (result.stderr or "").strip()
        if "__NO_LOG__" in out:
            return False, f"No .log file found in LOG_PATH: {log_path}"
        if "__LOG__:" in out:
            return True, out.split("\n", 1)[1] if "\n" in out else out
        return False, err or out or "Unknown error while reading remote log."

    def _set_running(self, running: bool) -> None:
        self.is_running = running
        self.start_btn.config(state="disabled" if running else "normal")
        self.stop_btn.config(state="normal" if running else "disabled")
        self.status_var.set("Running" if running else "Idle")
        if not running:
            self.session_established = False
            self.manual_send_ready = False
            self.auto_xml_send_done = False
            self._conformance_auto_sync_scheduled = False
            self._cancel_session_watch()
        self._sync_manual_send_widgets()
        if not running:
            self.after(300, self._enqueue_auto_restart_if_configured)
        self.after(0, self._refresh_log_target_hint_line)

    def manual_start_script(self) -> None:
        self._conformance_dbg("manual_start_script: 호출 (_user_stop_requested 해제 후 start_script)")
        self._user_stop_requested = False
        self._cancel_auto_start_retry_job()
        self.start_script()

    def start_script(self) -> None:
        self._cancel_auto_start_retry_job()
        if self.proc is not None and self.proc.poll() is None:
            messagebox.showinfo("Info", "Script is already running.")
            return
        pch = self.paramiko_channel
        if pch is not None and not getattr(pch, "closed", False):
            messagebox.showinfo("Info", "Script is already running.")
            return

        self._save_current_config()
        self.remote_cfg_cache.clear()
        env = os.environ.copy()
        field_values: dict[str, str] = {}
        for key, var in self.fields.items():
            value = var.get().strip()
            env[key] = value
            field_values[key] = value

        self._start_log_path_hint = field_values.get("LOG_PATH", "").strip()
        self.after(0, self._refresh_log_target_hint_line)

        self.stop_event.clear()
        self.session_established = False
        self.manual_send_ready = False
        self.auto_xml_send_done = False
        self._conformance_auto_sync_scheduled = False
        with self.log_lock:
            self._recent_log_for_session = ""
        self._set_running(True)
        self.status_var.set("Preparing...")
        self.append_log("[GUI] BUILD_MARKER: start_script_v20260512_1725\n")
        self.append_log("\n[GUI] Preparing execution...\n")
        threading.Thread(
            target=self._start_script_worker,
            args=(env, field_values),
            daemon=True,
        ).start()
        self._schedule_session_watch()

    def _cancel_auto_start_retry_job(self) -> None:
        jid = self._auto_start_retry_job
        self._auto_start_retry_job = None
        if jid is not None:
            try:
                self.after_cancel(jid)
            except Exception:
                pass

    def _enqueue_auto_restart_if_configured(self) -> None:
        auto_on = bool(self.auto_start_var.get())
        if self._conformance_run_busy or self._conformance_stop_idle_wait:
            if auto_on:
                self._conformance_dbg(
                    f"auto_start: 스킵 (busy={self._conformance_run_busy} stop_idle_wait={self._conformance_stop_idle_wait})"
                )
            return
        if self.is_running:
            return
        if not auto_on:
            return
        if self._user_stop_requested:
            self._conformance_dbg("auto_start: 스킵 (_user_stop_requested — 마지막 Stop이 사용자 요청)")
            return
        if self.proc is not None and self.proc.poll() is None:
            self._conformance_dbg("auto_start: 스킵 (subprocess 아직 실행 중)")
            return
        ch = self.paramiko_channel
        if ch is not None and not getattr(ch, "closed", False):
            self._conformance_dbg("auto_start: 스킵 (paramiko 채널 아직 열림)")
            return
        self._cancel_auto_start_retry_job()
        sec = max(1, self._auto_restart_delay_ms // 1000)
        self._conformance_dbg(f"auto_start: {sec}s 후 Start 재시도 예약")
        self.append_log(f"[GUI] Auto-start: will retry Start in {sec}s (until Stop or unchecked).\n")
        self._auto_start_retry_job = self.after(self._auto_restart_delay_ms, self._fire_auto_start_retry_tick)

    def _fire_auto_start_retry_tick(self) -> None:
        self._auto_start_retry_job = None
        if not self.auto_start_var.get():
            return
        if self._user_stop_requested:
            return
        if self._conformance_run_busy or self._conformance_stop_idle_wait:
            return
        if self.is_running:
            return
        if self.proc is not None and self.proc.poll() is None:
            return
        pch = self.paramiko_channel
        if pch is not None and not getattr(pch, "closed", False):
            return
        self.append_log("[GUI] Auto-start: retrying...\n")
        self.start_script()

    def _on_auto_start_toggled(self) -> None:
        self._save_current_config()
        if not self.auto_start_var.get():
            self._cancel_auto_start_retry_job()

    def _start_script_worker(self, env: dict[str, str], field_values: dict[str, str]) -> None:
        try:
            ssh_user = self.remote_user.get().strip()
            ssh_host = self.remote_host.get().strip()
            ssh_port = self.remote_port.get().strip()
            ssh_password = self.remote_password.get().strip()
            remote_script = self.remote_script_path.get().strip()
            key_path = self.remote_key_path.get().strip()

            if not ssh_user or not ssh_host or not ssh_port or not remote_script:
                self.after(
                    0,
                    lambda: messagebox.showerror(
                        "Error", "SSH_USER, SSH_HOST, SSH_PORT, REMOTE_SCRIPT_PATH are required."
                    ),
                )
                self.after(0, self._set_running, False)
                return

            auth_mode, info_or_error = self._resolve_auth_mode(ssh_password)
            if info_or_error and "Install" in info_or_error:
                self.after(0, lambda: messagebox.showerror("Error", info_or_error))
                self.after(0, self._set_running, False)
                return
            if ssh_password and auth_mode == "paramiko":
                self.after(0, self.append_log, "[GUI] Using paramiko password auth.\n")
            if info_or_error:
                self.after(0, self.append_log, f"[GUI] {info_or_error}\n")

            remote_log_path = field_values.get("LOG_PATH", "").strip()
            if not remote_log_path:
                self.after(0, lambda: messagebox.showerror("Error", "LOG_PATH is required."))
                self.after(0, self._set_running, False)
                return

            ok, info = self._ensure_remote_script(
                ssh_user,
                ssh_host,
                ssh_port,
                key_path,
                ssh_password,
                auth_mode,
                remote_script,
                remote_log_path,
            )
            self.after(0, self.append_log, f"[GUI] {info}\n")
            if not ok:
                self.after(0, lambda: messagebox.showerror("Error", info))
                self.after(0, self._set_running, False)
                return

            exports = " ".join(
                f"{k}={shlex.quote(v)}" for k, v in field_values.items() if v != ""
            )
            remote_cmd = f"{exports} bash {shlex.quote(remote_script)}"
            ch_port_probe = (field_values.get("CALLHOME_PORT") or field_values.get("PORT") or "4334").strip() or "4334"
            product_probe = (field_values.get("PRODUCT") or "").strip()
            use_password = bool(ssh_password)
            if auth_mode == "paramiko":
                self.after(0, self.append_log, f"\n[GUI] Started (paramiko): {ssh_user}@{ssh_host}\n")
                self.after(0, self.append_log, "[POSTCHECK] scheduled (paramiko start).\n")
                threading.Thread(
                    target=self._start_postcheck_worker,
                    args=(ssh_user, ssh_host, ssh_port, ssh_password, key_path, remote_script, remote_log_path, ch_port_probe, product_probe),
                    daemon=True,
                ).start()
                self._start_remote_paramiko(
                    ssh_user=ssh_user,
                    ssh_host=ssh_host,
                    ssh_port=ssh_port,
                    ssh_password=ssh_password,
                    key_path=key_path,
                    remote_cmd=remote_cmd,
                )
                return
            cmd = self._ssh_base_cmd(
                ssh_user, ssh_host, ssh_port, key_path, use_password, auth_mode, ssh_password
            ) + [remote_cmd]
            if use_password and auth_mode == "sshpass":
                env = env.copy()
                env["SSHPASS"] = ssh_password

            self.proc = subprocess.Popen(
                cmd,
                stdin=subprocess.PIPE,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                env=env,
                creationflags=subprocess.CREATE_NEW_PROCESS_GROUP if os.name == "nt" else 0,
            )
            self.after(0, self.append_log, f"\n[GUI] Started: {' '.join(cmd)}\n")
            self.after(0, lambda: self.status_var.set("Running"))
            threading.Thread(target=self._read_output, daemon=True).start()
            threading.Thread(target=self._watch_process, daemon=True).start()
            self.after(0, self.append_log, "[POSTCHECK] scheduled (subprocess start).\n")
            threading.Thread(
                target=self._start_postcheck_worker,
                args=(ssh_user, ssh_host, ssh_port, ssh_password, key_path, remote_script, remote_log_path, ch_port_probe, product_probe),
                daemon=True,
            ).start()
        except Exception as exc:
            self.after(0, lambda: messagebox.showerror("Error", f"Failed to start script:\n{exc}"))
            self.after(0, self._set_running, False)

    def _start_postcheck_worker(
        self,
        ssh_user: str,
        ssh_host: str,
        ssh_port: str,
        ssh_password: str,
        key_path: str,
        remote_script: str,
        remote_log_path: str,
        ch_port: str,
        product: str,
    ) -> None:
        # Start 직후 무출력/정지처럼 보일 때 원격 상태를 즉시 보여준다.
        self.after(0, self.append_log, "[POSTCHECK] worker started.\n")
        time.sleep(6)
        try:
            import paramiko  # type: ignore
        except Exception:
            self.after(0, self.append_log, "[POSTCHECK] paramiko import failed.\n")
            return
        client: Any | None = None
        try:
            self.after(0, self.append_log, "[POSTCHECK] connecting...\n")
            client = paramiko.SSHClient()
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            client.connect(
                hostname=ssh_host,
                port=int(ssh_port),
                username=ssh_user,
                password=ssh_password if ssh_password else None,
                key_filename=key_path if key_path else None,
                timeout=8,
                auth_timeout=8,
                banner_timeout=8,
                look_for_keys=not bool(ssh_password),
                allow_agent=True,
            )
            self.after(0, self.append_log, "[POSTCHECK] connected. probing remote state...\n")
            log_dir = (remote_log_path or "/var/tmp/log").rstrip("/")
            pat = f"{product}_*.log" if product else "*.log"
            cmd = (
                f"echo '[POSTCHECK] script={shlex.quote(remote_script)}'; "
                f"echo '[POSTCHECK] process:'; pgrep -af {shlex.quote(Path(remote_script).name)} || true; "
                f"echo '[POSTCHECK] listen:{ch_port}'; (ss -tnlp | grep ':{ch_port} ' || true); "
                f"echo '[POSTCHECK] recent log:'; "
                f"(ls -1t {shlex.quote(log_dir)}/{pat} 2>/dev/null | head -n 1 | xargs -r tail -n 25) || true"
            )
            _i, o, e = client.exec_command(cmd)
            out = o.read().decode(errors="replace")
            err = e.read().decode(errors="replace")
            blob = out.strip()
            if err.strip():
                blob += ("\n" if blob else "") + f"[POSTCHECK][stderr] {err.strip()}"
            if blob:
                self.after(0, self.append_log, f"{blob}\n")
        except Exception as exc:
            self.after(0, self.append_log, f"[POSTCHECK] failed: {exc}\n")
        finally:
            try:
                if client is not None:
                    client.close()
            except Exception:
                pass

    def _start_remote_paramiko(
        self,
        ssh_user: str,
        ssh_host: str,
        ssh_port: str,
        ssh_password: str,
        key_path: str,
        remote_cmd: str,
    ) -> None:
        self.after(0, self.append_log, "[GUI][paramiko] ENTER _start_remote_paramiko v20260512_1725\n")
        try:
            import paramiko  # type: ignore
        except Exception:
            self.after(0, lambda: messagebox.showerror("Error", "paramiko is not installed. Run: py -m pip install paramiko"))
            self.after(0, self._set_running, False)
            return

        try:
            self.after(0, self.append_log, "[GUI][paramiko] connect starting...\n")
            client = paramiko.SSHClient()
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            client.connect(
                hostname=ssh_host,
                port=int(ssh_port),
                username=ssh_user,
                password=ssh_password if ssh_password else None,
                key_filename=key_path if key_path else None,
                timeout=8,
                auth_timeout=8,
                banner_timeout=8,
                look_for_keys=not bool(ssh_password),
                allow_agent=True,
            )
            self.after(0, self.append_log, "[GUI][paramiko] connect ok, opening shell...\n")
            transport = client.get_transport()
            if transport is None:
                raise RuntimeError("SSH transport is unavailable")
            transport.set_keepalive(20)

            # Use interactive shell channel for long-running script stability.
            channel = client.invoke_shell(width=200, height=50)
            self.after(0, self.append_log, "[GUI][paramiko] shell opened, sending start command...\n")
            channel.send(remote_cmd + "\n")
            self.after(0, self.append_log, "[GUI][paramiko] start command sent.\n")

            self.paramiko_client = client
            self.paramiko_channel = channel
            self.after(0, lambda: self.status_var.set("Running"))
            threading.Thread(target=self._read_paramiko_output, daemon=True).start()
            threading.Thread(target=self._watch_paramiko_channel, daemon=True).start()
        except Exception as exc:
            self.after(0, lambda: messagebox.showerror("Error", f"Failed to start remote command:\n{exc}"))
            self.after(0, self._set_running, False)

    def _read_paramiko_output(self) -> None:
        ch = self.paramiko_channel
        if ch is None:
            return
        pending: list[str] = []
        last_flush = time.monotonic()
        coalesce_sec = 0.1
        coalesce_chars = 6144

        def _flush_paramiko_log(*, force: bool = False) -> None:
            nonlocal pending, last_flush
            now = time.monotonic()
            blob = "".join(pending)
            if not blob:
                return
            if not force and (now - last_flush) < coalesce_sec and len(blob) < coalesce_chars:
                return
            pending.clear()
            last_flush = now
            self.after(0, lambda b=blob: self.append_log(b))

        try:
            while not ch.exit_status_ready():
                had_data = False
                if ch.recv_ready():
                    data = ch.recv(4096).decode(errors="ignore")
                    if data:
                        had_data = True
                        self._detect_session_established(data)
                        pending.append(data)
                        _flush_paramiko_log(force=False)
                if ch.recv_stderr_ready():
                    data = ch.recv_stderr(4096).decode(errors="ignore")
                    if data:
                        had_data = True
                        self._detect_session_established(data)
                        pending.append(data)
                        _flush_paramiko_log(force=False)
                if self.stop_event.is_set():
                    break
                if not had_data:
                    _flush_paramiko_log(force=False)
                    time.sleep(0.05)
        except Exception as exc:
            _flush_paramiko_log(force=True)
            self.after(0, self.append_log, f"\n[GUI] Output read error: {exc}\n")
            return
        _flush_paramiko_log(force=True)

    def _watch_paramiko_channel(self) -> None:
        ch = self.paramiko_channel
        if ch is None:
            return
        rc = 0
        try:
            while True:
                client = self.paramiko_client
                transport = client.get_transport() if client is not None else None
                if transport is None or not transport.is_active():
                    rc = -1
                    break
                # invoke_shell channel can stay open without exit status;
                # treat explicit channel close as graceful end.
                if ch.closed:
                    rc = 0
                    break
                time.sleep(0.2)
        except Exception:
            rc = -1
        self.after(0, self.append_log, f"\n[GUI] Process exited with code {rc}\n")
        try:
            ch.close()
        except Exception:
            pass
        if self.paramiko_client is not None:
            try:
                self.paramiko_client.close()
            except Exception:
                pass
        self.paramiko_channel = None
        self.paramiko_client = None
        self.after(0, self._set_running, False)

    def _read_output(self) -> None:
        if self.proc is None or self.proc.stdout is None:
            return
        for line in self.proc.stdout:
            if self.stop_event.is_set():
                break
            self._detect_session_established(line)
            self.after(0, self.append_log, line)

    def _watch_process(self) -> None:
        if self.proc is None:
            return
        rc = self.proc.wait()
        self.after(0, self.append_log, f"\n[GUI] Process exited with code {rc}\n")
        self.after(0, self._set_running, False)
        self.proc = None

    def stop_script(self) -> None:
        self._user_stop_requested = True
        self._cancel_auto_start_retry_job()
        if self.paramiko_channel is not None:
            self.stop_event.set()
            try:
                self.paramiko_channel.close()
            except Exception:
                pass
            if self.paramiko_client is not None:
                try:
                    self.paramiko_client.close()
                except Exception:
                    pass
            self.paramiko_channel = None
            self.paramiko_client = None
            self.append_log("[GUI] Stop requested.\n")
            # 채널만 닫으면 원격 miniDU_callhome.sh 본체가 남아 supervision 이 계속될 수 있다.
            self._cleanup_remote_daemons_async(reason="manual stop(paramiko)", include_start_script=True)
            self._set_running(False)
            return

        if self.proc is None or self.proc.poll() is not None:
            # 로컬 프로세스가 이미 없어도 원격 잔존 데몬이 남아있을 수 있으므로 정리 시도.
            self._cleanup_remote_daemons_async(reason="manual stop(no local proc)", include_start_script=True)
            self._set_running(False)
            return

        self.stop_event.set()
        try:
            if os.name == "nt":
                self.proc.send_signal(signal.CTRL_BREAK_EVENT)
            else:
                self.proc.terminate()
        except Exception:
            pass

        self.append_log("[GUI] Stop requested.\n")
        self._cleanup_remote_daemons_async(reason="manual stop(subprocess)", include_start_script=True)

    def _build_remote_daemon_cleanup_script(self, include_start_script: bool) -> str:
        """Shell snippet run on remote via bash -lc (graceful FIFO + pkill fallbacks)."""
        fifo = shlex.quote(self.remote_control_fifo)
        # 짧은 sleep: GUI 종료·Stop 시 원격 정리 대기 시간을 줄임 (이전에는 sleep 1×4 로 최소 ~4s).
        _w = "sleep 0.35"
        parts: list[str] = [
            "set +e",
            f"if [ -p {fifo} ]; then printf '%s\\n' disconnect >{fifo} 2>/dev/null; fi",
            _w,
        ]
        if include_start_script:
            sp = (self.remote_script_path.get() or "").strip() or "/var/tmp/miniDU_callhome.sh"
            sq = shlex.quote(sp)
            parts.append(f"pkill -TERM -f {sq} 2>/dev/null || true")
            parts.append(_w)
            parts.append(f"pkill -KILL -f {sq} 2>/dev/null || true")
        parts.extend(
            [
                "pkill -TERM -f netopeer2-cli 2>/dev/null || true",
                _w,
                "pkill -KILL -f netopeer2-cli 2>/dev/null || true",
                "exit 0",
            ]
        )
        return "; ".join(parts)

    def _cleanup_remote_daemons_impl(
        self, reason: str, include_start_script: bool, *, subprocess_timeout: float | None
    ) -> None:
        pref = "[Remote-cleanup]"

        def log(msg: str) -> None:
            try:
                self.after(0, self.append_log, f"{pref} {msg}\n")
            except Exception:
                pass

        log(f"시작 ({reason})")
        ssh_user, ssh_host, ssh_port, ssh_password, key_path = self._remote_conn()
        if not ssh_user or not ssh_host or not ssh_port:
            log("건너뜀: Settings SSH 정보 없음")
            return
        auth_mode, info_or_error = self._resolve_auth_mode(ssh_password)
        if info_or_error and "Install" in info_or_error:
            log(f"건너뜀: {info_or_error}")
            return
        inner = self._build_remote_daemon_cleanup_script(include_start_script)
        remote_line = "bash -lc " + shlex.quote(inner)
        timeout = subprocess_timeout if subprocess_timeout is not None else 18.0
        conn_to = min(10.0, max(2.5, float(timeout) * 0.55))

        if auth_mode == "paramiko":
            try:
                import paramiko  # type: ignore

                client = paramiko.SSHClient()
                client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                client.connect(
                    hostname=ssh_host,
                    port=int(ssh_port),
                    username=ssh_user,
                    password=ssh_password if ssh_password else None,
                    key_filename=key_path if key_path else None,
                    timeout=conn_to,
                    auth_timeout=min(10.0, conn_to + 1.0),
                    banner_timeout=min(10.0, conn_to + 1.0),
                    look_for_keys=not bool(ssh_password),
                    allow_agent=True,
                )
                try:
                    _stdin, stdout, stderr = client.exec_command(remote_line, timeout=timeout)
                    rc = stdout.channel.recv_exit_status()
                    err = stderr.read().decode(errors="ignore").strip()
                    if rc != 0 and err:
                        log(f"종료 코드 {rc}: {err[:500]}")
                    else:
                        log("원격 정리 명령 완료")
                finally:
                    client.close()
            except Exception as exc:
                log(f"실패: {exc}")
            return

        use_password = bool(ssh_password)
        run_env = os.environ.copy()
        if use_password and auth_mode == "sshpass":
            run_env["SSHPASS"] = ssh_password
        ssh_cmd = self._ssh_base_cmd(ssh_user, ssh_host, ssh_port, key_path, use_password, auth_mode, ssh_password)
        try:
            r = subprocess.run(
                ssh_cmd + [remote_line],
                capture_output=True,
                text=True,
                env=run_env,
                timeout=timeout,
            )
            if r.returncode != 0:
                tail = (r.stderr or r.stdout or "").strip()[:500]
                log(f"종료 코드 {r.returncode}: {tail or '(no output)'}")
            else:
                log("원격 정리 명령 완료")
        except subprocess.TimeoutExpired:
            log(f"시간 초과 ({timeout:.0f}s)")
        except Exception as exc:
            log(f"실패: {exc}")

    def _cleanup_remote_daemons_async(self, *, reason: str, include_start_script: bool) -> None:
        threading.Thread(
            target=self._cleanup_remote_daemons_impl,
            args=(reason, include_start_script),
            kwargs={"subprocess_timeout": 18.0},
            daemon=True,
        ).start()

    def _cleanup_remote_daemons_blocking(
        self, *, reason: str, include_start_script: bool = True, timeout_s: float = 6.0
    ) -> None:
        def run() -> None:
            try:
                self._cleanup_remote_daemons_impl(
                    reason, include_start_script, subprocess_timeout=max(4.0, timeout_s)
                )
            except BaseException:
                pass

        t = threading.Thread(target=run, daemon=True)
        t.start()
        t.join(timeout=max(1.5, timeout_s + 0.75))
        if t.is_alive():
            try:
                self.append_log(f"[Remote-cleanup] 대기 시간 초과 ({timeout_s:.0f}s), 창 닫기 계속\n")
            except Exception:
                pass

    def _detect_session_established(self, text: str) -> None:
        if self.session_established:
            upper = text.upper()
            if (not self.manual_send_ready) and ("SUPERVISION ACTIVE" in upper):
                self.manual_send_ready = True
                self.after(0, self._sync_manual_send_widgets)
                self.after(0, self.append_log, "[GUI] Manual send enabled (SUPERVISION ACTIVE).\n")
            return
        upper = text.upper()
        markers = [
            "ACCEPTED A CONNECTION ON",
            "AUTHENTICATION SUCCESSFUL",
            "LOGIN SUCCESSFUL",
            "<HELLO",
            "CALLHOME DETECTED",
            "CALL HOME DETECTED",
            "SUPERVISION ACTIVE",
        ]
        if any(m in upper for m in markers):
            self.session_established = True
            self.after(0, self.append_log, "[GUI] Session established detected. Netconf Client is available.\n")
            self.manual_send_ready = True
            self.after(0, self._sync_manual_send_widgets)
            self.after(0, self.append_log, "[GUI] Manual send enabled.\n")
            # Conformance auto-upload disabled while tab is viewer-only
            self.after(2000, self._schedule_conformance_auto_sync_once)
            if self.auto_xml_send_var.get() and not self.auto_xml_send_done:
                self.auto_xml_send_done = True
                self.after(0, self.append_log, "[GUI] Auto XML Send scheduled in 4 seconds.\n")
                self.after(4000, self._run_auto_xml_send_once)
        if "SUPERVISION ACTIVE" in upper and not self.manual_send_ready:
            had_session = self.session_established
            self.session_established = True
            self.manual_send_ready = True
            self.after(0, self._sync_manual_send_widgets)
            if had_session:
                self.after(0, self.append_log, "[GUI] Manual send enabled (SUPERVISION ACTIVE).\n")
            else:
                self.after(0, self.append_log, "[GUI] Session + manual send via SUPERVISION ACTIVE (no prior hello marker).\n")

    def _run_auto_xml_send_once(self) -> None:
        if not (self.is_running and self.auto_xml_send_var.get() and self.session_established):
            self.auto_xml_send_done = False
            return
        self.append_log("[GUI] Auto XML Send triggered (All Tabs Send 1x).\n")
        self.send_all_once()

    def _send_scheduler_payload(self, payload: str) -> None:
        t0 = time.perf_counter()
        payload = payload.strip()
        if not payload:
            return
        trace_id = self._trace("send_start", mode=self.send_mode_var.get(), bytes=str(len(payload)))
        payload = self._prepare_payload_for_remote_send(payload)
        # Current runtime path sends through miniDU_callhome.sh stdin bridge,
        # so commands must be line-based CLI text.
        data = payload + "\n"
        if self.exec_mode.get() == "remote":
            # Fast path: write directly to running remote shell channel stdin.
            if self.paramiko_channel is not None:
                t_direct = time.perf_counter()
                self.paramiko_channel.send(data)
                self._perf_record("send_direct_channel", time.perf_counter() - t_direct)
                self._perf_record("send_payload", time.perf_counter() - t0)
                self._trace("send_done", trace=str(trace_id), path="direct_channel")
                return
            ok, msg = self._send_remote_control_command(payload)
            if not ok:
                # FIFO channel can be missing if an old script is still running.
                # Fallback to stdin bridge path for robustness.
                if "__FIFO_MISSING__" in msg and self.proc is not None and self.proc.stdin is not None and self.proc.poll() is None:
                    self.proc.stdin.write(data)
                    self.proc.stdin.flush()
                    self.append_log("[GUI] FIFO missing, fallback to stdin bridge.\n")
                    self._perf_record("send_payload", time.perf_counter() - t0)
                    self._trace("send_done", trace=str(trace_id), path="stdin_fallback")
                    return
                raise RuntimeError(msg)
            self._perf_record("send_payload", time.perf_counter() - t0)
            self._trace("send_done", trace=str(trace_id), path="fifo_remote")
            return
        if self.paramiko_channel is not None:
            self.paramiko_channel.send(data)
            self._perf_record("send_payload", time.perf_counter() - t0)
            self._trace("send_done", trace=str(trace_id), path="paramiko_channel")
            return
        if self.proc is not None and self.proc.stdin is not None and self.proc.poll() is None:
            self.proc.stdin.write(data)
            self.proc.stdin.flush()
            self._perf_record("send_payload", time.perf_counter() - t0)
            self._trace("send_done", trace=str(trace_id), path="local_stdin")
            return
        raise RuntimeError("No active process/channel for sending")

    def _send_remote_control_command(self, command: str) -> tuple[bool, str]:
        t0 = time.perf_counter()
        ssh_user, ssh_host, ssh_port, ssh_password, key_path = self._remote_conn()
        if not ssh_user or not ssh_host or not ssh_port:
            return False, "SSH connection info is incomplete."
        auth_mode, info_or_error = self._resolve_auth_mode(ssh_password)
        if info_or_error and "Install" in info_or_error:
            return False, info_or_error
        shell_cmd = (
            f"if [ -p {shlex.quote(self.remote_control_fifo)} ]; then "
            f"printf '%s\\n' {shlex.quote(command)} > {shlex.quote(self.remote_control_fifo)}; "
            f"else echo '__FIFO_MISSING__'; exit 2; fi"
        )

        if auth_mode == "paramiko":
            try:
                import paramiko  # type: ignore
                client = self.paramiko_client
                owned = False
                if client is None:
                    client = paramiko.SSHClient()
                    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                    client.connect(
                        hostname=ssh_host,
                        port=int(ssh_port),
                        username=ssh_user,
                        password=ssh_password if ssh_password else None,
                        key_filename=key_path if key_path else None,
                        timeout=8,
                        auth_timeout=8,
                        banner_timeout=8,
                        look_for_keys=not bool(ssh_password),
                        allow_agent=True,
                    )
                    owned = True
                _, stdout, stderr = client.exec_command(shell_cmd)
                rc = stdout.channel.recv_exit_status()
                out = stdout.read().decode(errors="ignore").strip()
                err = stderr.read().decode(errors="ignore").strip()
                if owned:
                    client.close()
                if rc != 0 or "__FIFO_MISSING__" in out:
                    return False, err or out or f"Remote control send failed (rc={rc})"
                self._perf_record("send_fifo_remote", time.perf_counter() - t0)
                return True, "ok"
            except Exception as exc:
                return False, f"Remote control send failed: {exc}"

        use_password = bool(ssh_password)
        run_env = os.environ.copy()
        if use_password and auth_mode == "sshpass":
            run_env["SSHPASS"] = ssh_password
        ssh_cmd = self._ssh_base_cmd(ssh_user, ssh_host, ssh_port, key_path, use_password, auth_mode, ssh_password)
        result = subprocess.run(ssh_cmd + [shell_cmd], capture_output=True, text=True, env=run_env)
        out = (result.stdout or "").strip()
        err = (result.stderr or "").strip()
        if result.returncode != 0 or "__FIFO_MISSING__" in out:
            return False, err or out or f"Remote control send failed (rc={result.returncode})"
        self._perf_record("send_fifo_remote", time.perf_counter() - t0)
        return True, "ok"

    def _prepare_payload_for_remote_send(self, payload: str) -> str:
        t0 = time.perf_counter()
        if self.send_mode_var.get() == "raw_rpc":
            t_parse = time.perf_counter()
            translated = self._translate_raw_rpc_to_cli(payload.strip())
            self._perf_record("raw_rpc_translate", time.perf_counter() - t_parse)
            self._perf_record("prepare_payload", time.perf_counter() - t0)
            return translated
        lines = [ln for ln in payload.splitlines() if ln.strip()]
        if not lines:
            self._perf_record("prepare_payload", time.perf_counter() - t0)
            return payload
        first = lines[0].strip()
        xml_body = "\n".join(lines[1:]).strip()

        # Case 1) command + inline XML in one tab:
        # edit-config ... --config=/var/tmp/netconf_tmp/x.xml
        # <config>...</config>
        if "--config=" in first and xml_body.startswith("<"):
            remote_path = self._extract_config_path(first)
            if remote_path:
                cache_key = self._cache_key("inline", xml_body, remote_path)
                cached = self.remote_cfg_cache.get(cache_key)
                if not cached:
                    self._perf_record("cfg_cache_miss", 0.0)
                    ok, msg = self._upload_text_to_remote(remote_path, xml_body)
                    if not ok:
                        raise RuntimeError(msg)
                    self.remote_cfg_cache[cache_key] = remote_path
                    self.append_log(f"[GUI] Uploaded inline XML to {remote_path}\n")
                else:
                    self._perf_record("cfg_cache_hit", 0.0)
                    self.append_log(f"[GUI] Reused cached XML: {cached}\n")
                self._perf_record("prepare_payload", time.perf_counter() - t0)
                return first

        # Case 2) local config path in command: --config=C:\...\x.xml
        if "--config=" in first:
            src_path = self._extract_config_path(first)
            if src_path and Path(src_path).exists():
                remote_path = f"/var/tmp/netconf_tmp/{Path(src_path).name}"
                ok, msg = self._upload_file_to_remote(src_path, remote_path)
                if not ok:
                    raise RuntimeError(msg)
                self.append_log(f"[GUI] Uploaded local file to {remote_path}\n")
                self._perf_record("prepare_payload", time.perf_counter() - t0)
                return re.sub(r"--config=\S+", f"--config={remote_path}", first)

        # Case 3) raw XML only in tab -> auto edit-config
        if payload.lstrip().startswith("<config"):
            cache_key = self._cache_key("auto_config", payload)
            remote_path = self.remote_cfg_cache.get(cache_key, "")
            if not remote_path:
                self._perf_record("cfg_cache_miss", 0.0)
                remote_path = f"/var/tmp/netconf_tmp/gui_config_{int(time.time())}.xml"
                ok, msg = self._upload_text_to_remote(remote_path, payload)
                if not ok:
                    raise RuntimeError(msg)
                self.remote_cfg_cache[cache_key] = remote_path
            else:
                self._perf_record("cfg_cache_hit", 0.0)
            auto_cmd = f"edit-config --target running --defop merge --config={remote_path}"
            self.append_log(f"[GUI] Auto command: {auto_cmd}\n")
            self._perf_record("prepare_payload", time.perf_counter() - t0)
            return auto_cmd

        self._perf_record("prepare_payload", time.perf_counter() - t0)
        return payload

    def _translate_raw_rpc_to_cli(self, payload: str) -> str:
        text = payload.strip()
        # Support XML declaration from ATOM exports.
        text = re.sub(r"^\s*<\?xml[^>]*\?>\s*", "", text, flags=re.IGNORECASE)
        if not text.startswith("<rpc"):
            return text

        # Prefer raw extraction to preserve original namespace/prefix text.
        lower = text.lower()
        if "<edit-config" in lower:
            target = "running"
            defop = "merge"
            m_target = re.search(r"<target>\s*<([a-zA-Z0-9_\-:]+)\s*/>\s*</target>", text, flags=re.IGNORECASE | re.DOTALL)
            if m_target:
                target = m_target.group(1).split(":")[-1]
            m_defop = re.search(r"<default-operation>\s*([^<]+)\s*</default-operation>", text, flags=re.IGNORECASE | re.DOTALL)
            if m_defop:
                defop = m_defop.group(1).strip()

            m_cfg = re.search(r"(<config\b[\s\S]*?</config>)", text, flags=re.IGNORECASE)
            if m_cfg:
                cfg_xml = m_cfg.group(1)
                # netopeer2-cli edit-config wraps provided config once.
                # If we pass outer <config> tag too, it becomes <config><config>...</config></config>.
                inner = re.match(r"^\s*<config\b[^>]*>([\s\S]*)</config>\s*$", cfg_xml, flags=re.IGNORECASE)
                if inner:
                    cfg_xml = inner.group(1).strip()
                cache_key = self._cache_key("raw_edit", cfg_xml, target, defop)
                remote_path = self.remote_cfg_cache.get(cache_key, "")
                if not remote_path:
                    self._perf_record("cfg_cache_miss", 0.0)
                    remote_path = f"/var/tmp/netconf_tmp/gui_cfg_{int(time.time() * 1000)}.xml"
                    ok, msg = self._upload_text_to_remote(remote_path, cfg_xml)
                    if not ok:
                        raise RuntimeError(msg)
                    self.remote_cfg_cache[cache_key] = remote_path
                else:
                    self._perf_record("cfg_cache_hit", 0.0)
                cmd = f"edit-config --target {target} --defop {defop} --config={remote_path}"
                self.append_log(f"[GUI] RAW RPC -> {cmd}\n")
                return cmd

        try:
            root = ET.fromstring(text)
        except Exception:
            return text

        def local(tag: str) -> str:
            return tag.split("}", 1)[-1] if "}" in tag else tag

        op = None
        for child in list(root):
            if isinstance(child.tag, str):
                op = child
                break
        if op is None:
            return text

        op_name = local(op.tag)
        if op_name == "edit-config":
            target = "running"
            defop = "merge"
            target_node = next((c for c in list(op) if local(c.tag) == "target"), None)
            if target_node is not None:
                target_child = next((c for c in list(target_node) if isinstance(c.tag, str)), None)
                if target_child is not None:
                    target = local(target_child.tag)
            defop_node = next((c for c in list(op) if local(c.tag) == "default-operation"), None)
            if defop_node is not None and (defop_node.text or "").strip():
                defop = (defop_node.text or "").strip()
            config_node = next((c for c in list(op) if local(c.tag) == "config"), None)
            if config_node is None:
                return text
            cfg_xml = ET.tostring(config_node, encoding="unicode")
            remote_path = f"/var/tmp/netconf_tmp/gui_cfg_{int(time.time() * 1000)}.xml"
            ok, msg = self._upload_text_to_remote(remote_path, cfg_xml)
            if not ok:
                raise RuntimeError(msg)
            cmd = f"edit-config --target {target} --defop {defop} --config={remote_path}"
            self.append_log(f"[GUI] RAW RPC -> {cmd}\n")
            return cmd

        if op_name == "get-config":
            source = "running"
            source_node = next((c for c in list(op) if local(c.tag) == "source"), None)
            if source_node is not None:
                src_child = next((c for c in list(source_node) if isinstance(c.tag, str)), None)
                if src_child is not None:
                    source = local(src_child.tag)
            cmd = f"get-config --source {source}"
            self.append_log(f"[GUI] RAW RPC -> {cmd}\n")
            return cmd

        if op_name == "get":
            cmd = "get"
            self.append_log(f"[GUI] RAW RPC -> {cmd}\n")
            return cmd

        # Fallback: unsupported raw RPC operation for current CLI bridge.
        raise RuntimeError(
            "Unsupported RAW RPC operation in current mode. "
            "Use edit-config/get/get-config RPC, or CLI command mode."
        )

    @staticmethod
    def _cache_key(kind: str, *parts: str) -> str:
        h = hashlib.sha1()
        h.update(kind.encode("utf-8"))
        for p in parts:
            h.update(b"\n--\n")
            h.update((p or "").encode("utf-8", errors="ignore"))
        return h.hexdigest()

    @staticmethod
    def _extract_config_path(command: str) -> str | None:
        m = re.search(r"--config=([^\s]+)", command)
        return m.group(1).strip() if m else None

    def _remote_conn(self) -> tuple[str, str, str, str, str]:
        return (
            self.remote_user.get().strip(),
            self.remote_host.get().strip(),
            self.remote_port.get().strip(),
            self.remote_password.get().strip(),
            self.remote_key_path.get().strip(),
        )

    def _upload_text_to_remote(self, remote_path: str, content: str) -> tuple[bool, str]:
        fd, tmp = tempfile.mkstemp(prefix="gui_cfg_", suffix=".xml")
        os.close(fd)
        try:
            Path(tmp).write_text(content, encoding="utf-8")
            return self._upload_file_to_remote(tmp, remote_path)
        finally:
            try:
                Path(tmp).unlink(missing_ok=True)
            except Exception:
                pass

    def _upload_file_to_remote(self, local_path: str, remote_path: str) -> tuple[bool, str]:
        t0 = time.perf_counter()
        trace_id = self._trace("upload_start", src=local_path, dst=remote_path)
        ssh_user, ssh_host, ssh_port, ssh_password, key_path = self._remote_conn()
        if not ssh_user or not ssh_host or not ssh_port:
            return False, "SSH connection info is incomplete."
        auth_mode, info_or_error = self._resolve_auth_mode(ssh_password)
        if info_or_error and "Install" in info_or_error:
            return False, info_or_error

        remote_parent = str(Path(remote_path).parent).replace("\\", "/")

        if auth_mode == "paramiko":
            try:
                import paramiko  # type: ignore
                client = self.paramiko_client
                owned = False
                if client is None:
                    client = paramiko.SSHClient()
                    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
                    client.connect(
                        hostname=ssh_host,
                        port=int(ssh_port),
                        username=ssh_user,
                        password=ssh_password if ssh_password else None,
                        key_filename=key_path if key_path else None,
                        timeout=8,
                        auth_timeout=8,
                        banner_timeout=8,
                        look_for_keys=not bool(ssh_password),
                        allow_agent=True,
                    )
                    owned = True
                _, stdout, stderr = client.exec_command(f"mkdir -p {shlex.quote(remote_parent)}")
                if stdout.channel.recv_exit_status() != 0:
                    err = stderr.read().decode(errors="ignore").strip()
                    if owned:
                        client.close()
                    return False, f"Remote mkdir failed: {err}"
                sftp = client.open_sftp()
                sftp.put(local_path, remote_path)
                sftp.close()
                if owned:
                    client.close()
                self._perf_record("upload_remote", time.perf_counter() - t0)
                self._trace("upload_done", trace=str(trace_id), mode="paramiko")
                return True, "ok"
            except Exception as exc:
                return False, f"Remote upload failed: {exc}"

        use_password = bool(ssh_password)
        run_env = os.environ.copy()
        if use_password and auth_mode == "sshpass":
            run_env["SSHPASS"] = ssh_password
        ssh_cmd = self._ssh_base_cmd(ssh_user, ssh_host, ssh_port, key_path, use_password, auth_mode, ssh_password)
        mkdir_result = subprocess.run(ssh_cmd + [f"mkdir -p {shlex.quote(remote_parent)}"], capture_output=True, text=True, env=run_env)
        if mkdir_result.returncode != 0:
            return False, (mkdir_result.stderr or mkdir_result.stdout).strip() or "Remote mkdir failed"
        scp_cmd = self._scp_base_cmd(ssh_port, key_path, use_password, auth_mode, ssh_password) + [
            local_path,
            f"{ssh_user}@{ssh_host}:{remote_path}",
        ]
        scp_result = subprocess.run(scp_cmd, capture_output=True, text=True, env=run_env)
        if scp_result.returncode != 0:
            return False, (scp_result.stderr or scp_result.stdout).strip() or "Remote upload failed"
        self._perf_record("upload_remote", time.perf_counter() - t0)
        self._trace("upload_done", trace=str(trace_id), mode=auth_mode)
        return True, "ok"

    def _collect_scheduler_payloads(self, selected_only: bool = False) -> list[tuple[str, str]]:
        if not self.message_tabs:
            return []
        if selected_only:
            idx = self.msg_notebook.index("current")
            tab = self.message_tabs[idx]
            return [(tab["title_var"].get() or f"MSG-{idx+1}", tab["text"].get("1.0", "end"))]
        # Respect current notebook visual order (left -> right).
        payloads: list[tuple[str, str]] = []
        for i in range(len(self.message_tabs)):
            frame_name = self.msg_notebook.tabs()[i]
            tab = next((t for t in self.message_tabs if str(t["frame"]) == frame_name), None)
            if tab is None:
                continue
            payloads.append((tab["title_var"].get() or f"MSG-{i+1}", tab["text"].get("1.0", "end")))
        return payloads

    def send_selected_once(self) -> None:
        if not self.session_established:
            messagebox.showwarning("Warning", "Session is not established yet.")
            return
        if not self.manual_send_ready:
            messagebox.showwarning("Warning", "Manual send is enabled after session login.")
            return
        payloads = self._collect_scheduler_payloads(selected_only=True)
        threading.Thread(target=self._send_payloads_once_worker, args=(payloads,), daemon=True).start()

    def send_all_once(self) -> None:
        if not self.session_established:
            messagebox.showwarning("Warning", "Session is not established yet.")
            return
        if not self.manual_send_ready:
            messagebox.showwarning("Warning", "Manual send is enabled after session login.")
            return
        payloads = self._collect_scheduler_payloads(selected_only=False)
        threading.Thread(target=self._send_payloads_once_worker, args=(payloads,), daemon=True).start()

    def _send_payloads_once_worker(self, payloads: list[tuple[str, str]]) -> None:
        self.after(0, lambda: self.send_selected_btn.config(state="disabled"))
        self.after(0, lambda: self.send_all_btn.config(state="disabled"))
        mb = self.mplane_apply_btn
        if mb is not None:
            self.after(0, lambda b=mb: b.config(state="disabled"))
        sent = 0
        try:
            for title, payload in payloads:
                if not payload.strip():
                    continue
                try:
                    self._send_scheduler_payload(payload)
                    self.after(0, self.append_log, f"[GUI] Sent once from tab: {title}\n")
                    sent += 1
                except Exception as exc:
                    self.after(0, self.append_log, f"[GUI] Send once failed ({title}): {exc}\n")
            if sent == 0:
                self.after(0, self.append_log, "[GUI] No payload to send.\n")
        finally:
            self.after(0, self._sync_manual_send_widgets)

    def on_close(self) -> None:
        if self._config_save_job is not None:
            try:
                self.after_cancel(self._config_save_job)
            except Exception:
                pass
            self._config_save_job = None
        self._conformance_omit_last_run_from_config_save = True
        self._conformance_last_run_snapshot_cache = None
        if hasattr(self, "conformance_last_run_hint_var"):
            self.conformance_last_run_hint_var.set("")
        self._save_current_config()
        if self._conformance_run_busy:
            self._conformance_stop_run()
        self.stop_script()
        # GUI 종료 시 관련 원격 데몬까지 정리
        self._cleanup_remote_daemons_blocking(reason="GUI close", include_start_script=True, timeout_s=4.0)
        self.destroy()


if __name__ == "__main__":
    app = CallhomeGUI()
    app.mainloop()

# EXE build (Windows):
#   py -m pip install pyinstaller paramiko openpyxl
#   py -m PyInstaller --noconfirm --onefile --windowed --collect-submodules openpyxl --name "O-RAN Netconf" callhome_gui.py
# If `py` launcher is unavailable in PowerShell, use:
#   python -m pip install pyinstaller paramiko openpyxl
#   python -m PyInstaller --noconfirm --onefile --windowed --collect-submodules openpyxl --name "O-RAN Netconf" callhome_gui.py
# Output:
#   dist\O-RAN Netconf.exe
